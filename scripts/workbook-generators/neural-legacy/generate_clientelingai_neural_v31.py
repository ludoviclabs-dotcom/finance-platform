from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
import re

OUTPUT_FILE = "ClientelingAI_NEURAL.xlsx"

# =========================================================
# CONSTANTES
# =========================================================
COLORS = {
    "navy": "1F4E78",
    "navy_dark": "203864",
    "blue": "5B9BD5",
    "blue_light": "D9EAF7",
    "blue_soft": "DDEBF7",
    "green": "70AD47",
    "green_dark": "548235",
    "green_light": "E2F0D9",
    "orange": "F4B183",
    "orange_dark": "C55A11",
    "orange_light": "FCE4D6",
    "gold": "BF9000",
    "purple": "8064A2",
    "teal": "00B0F0",
    "gray": "A5A5A5",
    "gray_light": "F2F2F2",
    "red_light": "FFC7CE",
    "white": "FFFFFF",
}

THIN_GREY = Side(style="thin", color="D9D9D9")

YES_NO = '"Oui,Non"'


# =========================================================
# HELPERS STYLES / FORMAT
# =========================================================
def apply_border(cell):
    cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def style_header(ws, row=1, fill_color=COLORS["navy"]):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color=COLORS["white"], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            apply_border(cell)


def style_input_cell(cell):
    cell.fill = PatternFill("solid", fgColor=COLORS["blue_light"])
    apply_border(cell)


def style_formula_cell(cell):
    cell.fill = PatternFill("solid", fgColor=COLORS["gray_light"])
    apply_border(cell)


def style_note_cell(cell):
    cell.fill = PatternFill("solid", fgColor=COLORS["green_light"])
    apply_border(cell)


def set_number_format(ws, col_letter, start_row, end_row, fmt):
    for r in range(start_row, end_row + 1):
        ws[f"{col_letter}{r}"].number_format = fmt


def autofit_width(ws, min_width=12, max_width=42):
    for col_cells in ws.columns:
        try:
            col_letter = get_column_letter(col_cells[0].column)
        except Exception:
            continue
        length = 0
        for cell in col_cells:
            try:
                if cell.value is not None:
                    length = max(length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 2, max_width))


def freeze_top(ws, cell="A2"):
    ws.freeze_panes = cell


def color_tab(ws, color):
    ws.sheet_properties.tabColor = color


def add_table(ws, table_name, start_row, start_col, end_row, end_col, style_name="TableStyleMedium2"):
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def add_list_validation(ws, cell_range, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_conditional_status_colors(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"Prêt à envoyer"'],
                   fill=PatternFill("solid", fgColor="C6EFCE"))
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"À valider"'],
                   fill=PatternFill("solid", fgColor="FCE4D6"))
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"Bloqué"'],
                   fill=PatternFill("solid", fgColor="FFC7CE"))
    )


def create_sheet_with_headers(wb, name, headers, tab_color=COLORS["green"], header_color=COLORS["navy"]):
    ws = wb.create_sheet(name)
    color_tab(ws, tab_color)
    for idx, h in enumerate(headers, 1):
        ws.cell(1, idx, h)
    style_header(ws, 1, fill_color=header_color)
    freeze_top(ws)
    return ws


def mark_input_columns(ws, cols, start_row=2, end_row=300):
    for col in cols:
        for r in range(start_row, end_row + 1):
            style_input_cell(ws[f"{col}{r}"])


def mark_formula_columns(ws, cols, start_row=2, end_row=300):
    for col in cols:
        for r in range(start_row, end_row + 1):
            style_formula_cell(ws[f"{col}{r}"])


def add_named_list_ranges(wb):
    named_ranges = {
        "lstMaisons": "'PARAM_LISTES'!$A$2:$A$5000",
        "lstPays": "'PARAM_LISTES'!$E$2:$E$5000",
        "lstLangues": "'PARAM_LISTES'!$A$11:$A$5000",
        "lstSegments": "'PARAM_LISTES'!$F$11:$F$5000",
        "lstCanaux": "'PARAM_LISTES'!$A$21:$A$5000",
        "lstOccasions": "'PARAM_LISTES'!$H$21:$H$5000",
        "lstCategories": "'PARAM_LISTES'!$A$33:$A$5000",
        "lstStatutsInteraction": "'PARAM_LISTES'!$H$33:$H$5000",
        "lstTons": "'PARAM_LISTES'!$K$33:$K$5000",
        "lstValidation": "'PARAM_LISTES'!$O$33:$O$5000",
        "lstClients": "'CLIENTS'!$A$2:$A$5000",
    }

    for name, ref in named_ranges.items():
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def normalize_data_validations(wb):
    replacements = {
        "=tblMaisons[Maison]": "=lstMaisons",
        "=tblPays[Pays]": "=lstPays",
        "=tblLangues[Langue]": "=lstLangues",
        "=tblSegments[Segment_Client]": "=lstSegments",
        "=tblCanaux[Canal]": "=lstCanaux",
        "=tblCategories[Categorie]": "=lstCategories",
        "=tblStatutsInteraction[Statut_Interaction]": "=lstStatutsInteraction",
        "=tblOccasions[Occasion]": "=lstOccasions",
        "=tblTons[Ton]": "=lstTons",
        "=tblClients[Client_ID]": "=lstClients",
    }

    for ws in wb.worksheets:
        validations = getattr(ws.data_validations, "dataValidation", [])
        for dv in validations:
            if dv.type == "list" and dv.formula1 in replacements:
                dv.formula1 = replacements[dv.formula1]


def normalize_future_formulas(wb):
    patterns = [
        (r'(?<![_A-Z.])XLOOKUP\(', '_xlfn.XLOOKUP('),
        (r'(?<![_A-Z.])FILTER\(', '_xlfn._xlws.FILTER('),
        (r'(?<![_A-Z.])LET\(', '_xlfn.LET('),
        (r'(?<![_A-Z.])SORTBY\(', '_xlfn.SORTBY('),
        (r'(?<![_A-Z.])TEXTJOIN\(', '_xlfn.TEXTJOIN('),
    ]

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    updated = cell.value
                    for pattern, replacement in patterns:
                        updated = re.sub(pattern, replacement, updated)
                    if updated != cell.value:
                        cell.value = updated


def finalize_workbook_compatibility(wb):
    add_named_list_ranges(wb)
    normalize_data_validations(wb)
    normalize_future_formulas(wb)


# =========================================================
# ONGLET README
# =========================================================
def build_readme_sheet(wb):
    ws = wb.create_sheet("README_START")
    color_tab(ws, COLORS["blue"])

    rows = [
        ["ClientelingAI - NEURAL"],
        [""],
        ["Version", "1.1"],
        ["Fichier", OUTPUT_FILE],
        ["Objet", "Cockpit Excel pour clienteling digital luxe ultra-personnalisé"],
        [""],
        ["Fonctions clés"],
        ["- Priorisation des clients à contacter"],
        ["- Détection des moments relationnels"],
        ["- Suggestion de cadeaux / attentions"],
        ["- Génération de messages personnalisés"],
        ["- Contrôle conformité / pression / tonalité"],
        ["- Suivi des actions et pilotage CRM"],
        [""],
        ["Bonnes pratiques"],
        ["- Ne pas supprimer les colonnes des tables"],
        ["- Ne pas écraser les formules"],
        ["- Remplir uniquement les colonnes de saisie"],
        ["- Contrôler les consentements avant action"],
        ["- Éviter toute donnée sensible non nécessaire"],
        ["- Maintenir la veille juridique / marketing à jour"],
        [""],
        ["Code couleur"],
        ["Bleu clair = saisie utilisateur"],
        ["Gris clair = formule / calcul"],
        ["Vert = exploitable / note positive"],
        ["Orange = vigilance / validation"],
        ["Rouge = blocage"],
    ]

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["blue_light"])
    style_note_cell(ws["A7"])
    style_note_cell(ws["A15"])
    style_note_cell(ws["A23"])
    autofit_width(ws)
    return ws


# =========================================================
# ONGLET DATA_DICTIONARY
# =========================================================
def build_data_dictionary_sheet(wb):
    headers = ["Onglet", "Colonne", "Type", "Mode", "Description", "Exemple"]
    ws = create_sheet_with_headers(wb, "DATA_DICTIONARY", headers, tab_color=COLORS["gold"], header_color=COLORS["gold"])

    rows = [
        ["CLIENTS", "Client_ID", "Texte", "Saisie", "Identifiant unique client", "C0001"],
        ["CLIENTS", "Maison", "Liste", "Saisie", "Maison de rattachement", "Maison A"],
        ["CLIENTS", "Date_Anniversaire_Calc", "Date", "Calcul", "Prochaine date anniversaire", "2025-04-25"],
        ["CLIENTS", "CA_12M", "Nombre", "Calcul", "Chiffre d'affaires sur 12 mois", "4200"],
        ["CLIENTS", "Flag_Conformite", "Texte", "Calcul", "Indicateur conformité canal/opt-in", "OK"],
        ["ACHATS", "Montant_TTC", "Nombre", "Saisie", "Montant achat TTC", "4200"],
        ["INTERACTIONS", "Statut_Interaction", "Liste", "Saisie", "Statut de l'interaction", "Répondu"],
        ["MOMENTS", "Occasion_Detectee", "Texte", "Calcul", "Moment relationnel détecté", "Anniversaire"],
        ["SCORING", "Score_Total", "Nombre", "Calcul", "Score global de priorisation", "87"],
        ["MESSAGES_GEN", "Message_Final", "Texte", "Calcul", "Message personnalisé généré", "Bonjour Claire..."],
        ["CONTROLES", "Statut_Final", "Texte", "Calcul", "Résultat du contrôle final", "À valider"],
        ["QUEUE_ACTIONS", "Action_A_Faire", "Texte", "Calcul", "Instruction utilisateur", "Envoyer message"],
        ["EXPORT_CRM", "Validation_Finale", "Texte", "Calcul", "Message exportable ou non", "Prêt à envoyer"],
    ]

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    add_table(ws, "tblDataDictionary", 1, 1, len(rows) + 1, len(headers))
    autofit_width(ws)
    return ws


# =========================================================
# PARAM_LISTES
# =========================================================
def build_param_listes_sheet(wb):
    ws = wb.create_sheet("PARAM_LISTES")
    color_tab(ws, "4472C4")

    # tblMaisons
    start = 1
    headers = ["Maison", "Actif"]
    data = [["Maison A", "Oui"], ["Maison B", "Oui"], ["Maison C", "Oui"]]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblMaisons", start, 1, start + len(data), len(headers))

    # tblPays
    start, offset = 1, 5
    headers = ["Pays", "Code_Pays", "Zone", "Langue_Defaut", "Fuseau_Horaire", "Canal_Prioritaire", "Regle_Optin_Stricte", "Actif"]
    data = [
        ["France", "FR", "Europe", "FR", "CET", "Email", "Oui", "Oui"],
        ["Japon", "JP", "APAC", "JP", "JST", "LINE", "Oui", "Oui"],
        ["Chine", "CN", "APAC", "ZH", "CST", "WeChat", "Oui", "Oui"],
        ["Corée du Sud", "KR", "APAC", "KO", "KST", "KakaoTalk", "Oui", "Oui"],
        ["Italie", "IT", "Europe", "IT", "CET", "WhatsApp", "Oui", "Oui"],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblPays", start, offset, start + len(data), offset + len(headers) - 1)

    # tblLangues
    start = 10
    headers = ["Langue", "Code_Langue", "Actif"]
    data = [["FR", "fr", "Oui"], ["EN", "en", "Oui"], ["IT", "it", "Oui"], ["JP", "ja", "Oui"], ["ZH", "zh", "Oui"], ["KO", "ko", "Oui"]]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblLangues", start, 1, start + len(data), len(headers))

    # tblSegments
    start, offset = 10, 6
    headers = ["Segment_Client", "Priorite_Base", "Description", "Actif"]
    data = [
        ["Prospect premium", 1, "Prospect à fort potentiel", "Oui"],
        ["Client actif", 2, "Client ayant acheté récemment", "Oui"],
        ["VIP", 4, "Client à haute valeur", "Oui"],
        ["VIC", 5, "Very Important Client", "Oui"],
        ["Dormant", 2, "Client sans activité récente", "Oui"],
        ["À réactiver", 3, "Client à relancer", "Oui"],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblSegments", start, offset, start + len(data), offset + len(headers) - 1)

    # tblCanaux
    start = 20
    headers = ["Canal", "Type_Canal", "Nécessite_Optin", "Nb_Max_14J", "Delai_Min_Entre_Contacts_J", "Actif"]
    data = [
        ["Email", "Digital", "Oui", 2, 3, "Oui"],
        ["SMS", "Digital", "Oui", 2, 5, "Oui"],
        ["WhatsApp", "Messaging", "Oui", 2, 5, "Oui"],
        ["Téléphone", "Direct", "Oui", 2, 7, "Oui"],
        ["Courrier", "Physique", "Non", 1, 14, "Oui"],
        ["WeChat", "Messaging", "Oui", 2, 5, "Oui"],
        ["LINE", "Messaging", "Oui", 2, 5, "Oui"],
        ["KakaoTalk", "Messaging", "Oui", 2, 5, "Oui"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblCanaux", start, 1, start + len(data), len(headers))

    # tblOccasions
    start, offset = 20, 8
    headers = ["Occasion", "Priorite_Occasion", "Jours_Avant", "Jours_Apres", "Canal_Par_Defaut", "Validation_Humaine", "Actif"]
    data = [
        ["Anniversaire", 5, 15, 2, "WhatsApp", "Oui", "Oui"],
        ["Remerciement achat", 4, 0, 7, "Email", "Oui", "Oui"],
        ["Réactivation 90 jours", 3, 0, 999, "Email", "Oui", "Oui"],
        ["Réactivation 180 jours", 5, 0, 999, "Téléphone", "Oui", "Oui"],
        ["Nouvelle collection", 4, 10, 30, "WhatsApp", "Oui", "Oui"],
        ["Événement boutique", 5, 20, 1, "WhatsApp", "Oui", "Oui"],
        ["Cadeau ciblé", 4, 0, 30, "Email", "Oui", "Oui"],
        ["Anniversaire premier achat", 3, 7, 2, "Email", "Oui", "Oui"],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblOccasions", start, offset, start + len(data), offset + len(headers) - 1)

    # tblCategories
    start = 32
    headers = ["Categorie", "Actif"]
    data = [["Maroquinerie", "Oui"], ["Joaillerie", "Oui"], ["Horlogerie", "Oui"], ["Soie", "Oui"], ["Prêt-à-porter", "Oui"], ["Accessoires", "Oui"], ["Parfum", "Oui"]]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblCategories", start, 1, start + len(data), len(headers))

    # tblSousCategories
    start, offset = 32, 4
    headers = ["Categorie", "Sous_Categorie", "Actif"]
    data = [
        ["Maroquinerie", "Sac", "Oui"],
        ["Maroquinerie", "Petite maroquinerie", "Oui"],
        ["Joaillerie", "Bague", "Oui"],
        ["Joaillerie", "Collier", "Oui"],
        ["Horlogerie", "Montre", "Oui"],
        ["Soie", "Carré", "Oui"],
        ["Prêt-à-porter", "Veste", "Oui"],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblSousCategories", start, offset, start + len(data), offset + len(headers) - 1)

    # tblStatutsInteraction
    start, offset = 32, 8
    headers = ["Statut_Interaction", "Actif"]
    data = [["Envoyé", "Oui"], ["Répondu", "Oui"], ["Converti", "Oui"], ["Sans réponse", "Oui"], ["Refus", "Oui"], ["Planifié", "Oui"], ["Annulé", "Oui"]]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblStatutsInteraction", start, offset, start + len(data), offset + len(headers) - 1)

    # tblTons
    start, offset = 32, 11
    headers = ["Ton", "Description", "Actif"]
    data = [
        ["Très formel", "Formulation très institutionnelle", "Oui"],
        ["Formel premium", "Formel mais chaleureux", "Oui"],
        ["Chaleureux premium", "Personnel et haut de gamme", "Oui"],
        ["Événementiel exclusif", "Invitation / exclusivité", "Oui"],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblTons", start, offset, start + len(data), offset + len(headers) - 1)

    # tblValidation
    start, offset = 32, 15
    headers = ["Statut_Validation", "Actif"]
    data = [["Prêt à envoyer", "Oui"], ["À valider", "Oui"], ["Bloqué", "Oui"], ["Envoyé", "Oui"], ["Archivé", "Oui"]]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblValidation", start, offset, start + len(data), offset + len(headers) - 1)

    freeze_top(ws)
    autofit_width(ws)
    return ws


# =========================================================
# PARAM_REGLES
# =========================================================
def build_param_regles_sheet(wb):
    ws = wb.create_sheet("PARAM_REGLES")
    color_tab(ws, "2F75B5")

    # Brand rules
    start = 1
    headers = ["Maison", "Langue", "Ton_Autorisé", "Formule_Appel_Defaut", "Signature_Defaut", "Lexique_Autorisé", "Lexique_Interdit", "Mention_Obligatoire", "Validation_Humaine_Obligatoire"]
    data = [
        ["Maison A", "FR", "Formel premium", "Bonjour {PRENOM},", "Bien à vous, {CONSEILLER}", "élégance|savoir-faire|maison", "urgent|promo|gratuit", "Sous réserve de disponibilité.", "Oui"],
        ["Maison A", "EN", "Formal premium", "Dear {PRENOM},", "Kind regards, {CONSEILLER}", "craftsmanship|house|exclusive", "free|cheap|promo", "Subject to availability.", "Oui"],
        ["Maison B", "FR", "Chaleureux premium", "Bonjour {PRENOM},", "Bien cordialement, {CONSEILLER}", "heritage|raffinement|signature", "discount|deal|urgent", "Sélection personnalisée selon disponibilité.", "Oui"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblBrandRules", start, 1, start + len(data), len(headers))

    # Compliance rules
    start = 8
    headers = ["Regle_ID", "Type_Regle", "Pays", "Canal", "Description", "Seuil", "Action_Si_Echec", "Reference", "Date_MAJ"]
    data = [
        ["R001", "Opt-in", "France", "Email", "Opt-in obligatoire pour envoi marketing email", "Oui", "Bloquer", "RGPD/ePrivacy", "2025-01-10"],
        ["R002", "Pression", "France", "Email", "Maximum 2 contacts sur 14 jours", "2", "Alerte", "Règle interne CRM", "2025-01-10"],
        ["R003", "Opt-in", "Japon", "LINE", "Consentement requis pour contact promotionnel", "Oui", "Bloquer", "Règle locale", "2025-01-10"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblComplianceRules", start, 1, start + len(data), len(headers))

    # Score weights
    start = 14
    headers = ["Composante", "Poids"]
    data = [
        ["Valeur_Client", 0.25],
        ["Recence_Achat", 0.15],
        ["Frequence_Achat", 0.10],
        ["Engagement", 0.15],
        ["Occasion", 0.15],
        ["Conformite", 0.10],
        ["Risque_Attrition", 0.10],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblScoreWeights", start, 1, start + len(data), len(headers))

    # Seuils
    start, offset = 14, 4
    headers = ["Parametre", "Valeur"]
    data = [
        ["Jours_Dormance_1", 90],
        ["Jours_Dormance_2", 180],
        ["Jours_Anniversaire_Alert", 15],
        ["Jours_Post_Achat_Remerciement", 7],
        ["Score_P1", 85],
        ["Score_P2", 70],
        ["Score_P3", 50],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblSeuils", start, offset, start + len(data), offset + len(headers) - 1)

    # Lexique interdit
    start = 24
    headers = ["Mot_Interdit", "Langue", "Maison", "Raison"]
    data = [
        ["promo", "FR", "Maison A", "Non conforme au ton luxe"],
        ["gratuit", "FR", "Maison A", "Non conforme au ton luxe"],
        ["cheap", "EN", "Maison A", "Non conforme au ton luxe"],
        ["discount", "EN", "Maison B", "Non conforme au ton luxe"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblLexiqueInterdit", start, 1, start + len(data), len(headers))

    # Lexique premium
    start, offset = 24, 6
    headers = ["Mot", "Langue", "Maison", "Categorie"]
    data = [
        ["savoir-faire", "FR", "Maison A", "Brand"],
        ["raffinement", "FR", "Maison B", "Brand"],
        ["exclusive", "EN", "Maison A", "Brand"],
    ]
    for c, h in enumerate(headers, offset):
        ws.cell(start, c, h)
    for i, row in enumerate(data, start + 1):
        for c, val in enumerate(row, offset):
            ws.cell(i, c, val)
    style_header(ws, start)
    add_table(ws, "tblLexiquePremium", start, offset, start + len(data), offset + len(headers) - 1)

    freeze_top(ws)
    autofit_width(ws)
    return ws


# =========================================================
# CLIENTS
# =========================================================
def build_clients_sheet(wb):
    headers = [
        "Client_ID", "Maison", "Boutique", "Conseiller_Referent", "Civilite", "Prenom", "Nom",
        "Pays", "Ville", "Langue_Preferee", "Date_Naissance", "Date_Anniversaire_Calc",
        "Segment_Client", "Statut_Client", "Date_Dernier_Achat", "Date_Dernier_Contact",
        "Canal_Prefere", "Optin_Email", "Optin_SMS", "Optin_WhatsApp", "Optin_Telephone",
        "Optin_Courrier", "Optin_WeChat", "Optin_LINE", "Optin_KakaoTalk", "CA_12M",
        "Nb_Achats_12M", "Panier_Moyen_12M", "Categorie_Favorite", "SousCategorie_Favorite",
        "Couleur_Favorite", "Budget_Habituel", "Date_Premier_Achat", "Jours_Depuis_Achat",
        "Jours_Depuis_Contact", "Risque_Attrition", "Pression_14J", "Canal_Autorise_Principal",
        "Flag_Conformite", "Notes_Synthese", "Date_Consentement", "Source_Consentement",
        "Score_Clienteling", "Action_Recommandee"
    ]
    ws = create_sheet_with_headers(wb, "CLIENTS", headers, tab_color=COLORS["blue"], header_color=COLORS["navy"])

    sample_data = [
        ["C0001", "Maison A", "Paris Vendôme", "Alice Martin", "Mme", "Claire", "Durand", "France", "Paris", "FR", "1986-04-25", None, "VIP", "Actif", None, None, "WhatsApp", "Oui", "Oui", "Oui", "Oui", "Oui", "Non", "Non", "Non", None, None, None, None, None, "Bleu nuit", None, None, None, None, None, None, None, None, "Cliente fidèle maroquinerie.", "2024-06-14", "Boutique", None, None],
        ["C0002", "Maison B", "Tokyo Ginza", "Kenji Sato", "M.", "Hiro", "Tanaka", "Japon", "Tokyo", "JP", "1979-11-03", None, "VIC", "Actif", None, None, "LINE", "Non", "Non", "Non", "Oui", "Oui", "Non", "Oui", "Non", None, None, None, None, None, "Noir", None, None, None, None, None, None, None, None, "Client horlogerie et pièces exclusives.", "2024-04-22", "CRM Event", None, None],
    ]
    for r_idx, row in enumerate(sample_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    for r in range(2, 302):
        ws[f"L{r}"] = f'=IF(K{r}="","",DATE(YEAR(TODAY())+(DATE(YEAR(TODAY()),MONTH(K{r}),DAY(K{r}))<TODAY()),MONTH(K{r}),DAY(K{r})))'
        ws[f"O{r}"] = f'=IFERROR(MAX(FILTER(ACHATS!C:C,ACHATS!B:B=A{r})),"")'
        ws[f"P{r}"] = f'=IFERROR(MAX(FILTER(INTERACTIONS!C:C,INTERACTIONS!B:B=A{r})),"")'
        ws[f"Z{r}"] = f'=IFERROR(SUM(FILTER(ACHATS!M:M,(ACHATS!B:B=A{r})*(ACHATS!C:C>=EDATE(TODAY(),-12)))),0)'
        ws[f"AA{r}"] = f'=IFERROR(COUNTA(FILTER(ACHATS!A:A,(ACHATS!B:B=A{r})*(ACHATS!C:C>=EDATE(TODAY(),-12)))),0)'
        ws[f"AB{r}"] = f'=IF(AA{r}=0,0,Z{r}/AA{r})'
        ws[f"AG{r}"] = f'=IFERROR(MIN(FILTER(ACHATS!C:C,ACHATS!B:B=A{r})),"")'
        ws[f"AH{r}"] = f'=IF(O{r}="","",TODAY()-O{r})'
        ws[f"AI{r}"] = f'=IF(P{r}="","",TODAY()-P{r})'
        ws[f"AJ{r}"] = f'=MIN(100,AH{r}/2)'
        ws[f"AK{r}"] = f'=COUNTIFS(INTERACTIONS!B:B,A{r},INTERACTIONS!C:C,">="&TODAY()-14)'
        ws[f"AL{r}"] = f'=IF(AND(Q{r}="Email",R{r}="Oui"),"Email",IF(AND(Q{r}="SMS",S{r}="Oui"),"SMS",IF(AND(Q{r}="WhatsApp",T{r}="Oui"),"WhatsApp",IF(AND(Q{r}="Téléphone",U{r}="Oui"),"Téléphone",IF(AND(Q{r}="LINE",X{r}="Oui"),"LINE",IF(AND(Q{r}="WeChat",W{r}="Oui"),"WeChat",IF(AND(Q{r}="KakaoTalk",Y{r}="Oui"),"KakaoTalk","À vérifier")))))))'
        ws[f"AM{r}"] = f'=IF(AL{r}="À vérifier","Alerte","OK")'
        ws[f"AQ{r}"] = f'=IFERROR(XLOOKUP(A{r},SCORING!A:A,SCORING!I:I,""),"")'
        ws[f"AR{r}"] = f'=IFERROR(XLOOKUP(A{r},MOMENTS!A:A,MOMENTS!N:N,""),"")'

    add_table(ws, "tblClients", 1, 1, 301, len(headers))
    autofit_width(ws)

    add_list_validation(ws, "B2:B5000", "=tblMaisons[Maison]")
    add_list_validation(ws, "H2:H5000", "=tblPays[Pays]")
    add_list_validation(ws, "J2:J5000", "=tblLangues[Langue]")
    add_list_validation(ws, "M2:M5000", "=tblSegments[Segment_Client]")
    add_list_validation(ws, "Q2:Q5000", "=tblCanaux[Canal]")
    for col in ["R", "S", "T", "U", "V", "W", "X", "Y"]:
        add_list_validation(ws, f"{col}2:{col}5000", YES_NO)

    # styles visuels
    mark_input_columns(ws, ["A","B","C","D","E","F","G","H","I","J","K","M","N","Q","R","S","T","U","V","W","X","Y","AE","AN","AO"], 2, 300)
    mark_formula_columns(ws, ["L","O","P","Z","AA","AB","AG","AH","AI","AJ","AK","AL","AM","AQ","AR"], 2, 300)

    set_number_format(ws, "K", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "L", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "O", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "P", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "AG", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "Z", 2, 300, '#,##0.00')
    set_number_format(ws, "AB", 2, 300, '#,##0.00')
    return ws


# =========================================================
# ACHATS
# =========================================================
def build_achats_sheet(wb):
    headers = [
        "Achat_ID", "Client_ID", "Date_Achat", "Canal_Achat", "Boutique_Achat", "Pays_Achat",
        "SKU", "Produit", "Categorie", "Sous_Categorie", "Collection", "Saison",
        "Montant_TTC", "Quantite", "Couleur", "Taille", "Achat_Cadeau", "Conseiller",
        "Occasion_Achat", "Annee_Mois", "Montant_Unitaire"
    ]
    ws = create_sheet_with_headers(wb, "ACHATS", headers, tab_color=COLORS["orange_dark"], header_color=COLORS["orange_dark"])

    sample_data = [
        ["A0001", "C0001", "2025-02-15", "Boutique", "Paris Vendôme", "France", "SKU001", "Sac signature", "Maroquinerie", "Sac", "Croisière", "SS25", 4200, 1, "Bleu nuit", "", "Non", "Alice Martin", "Achat personnel", None, None],
        ["A0002", "C0001", "2024-10-10", "Boutique", "Paris Vendôme", "France", "SKU002", "Carré de soie", "Soie", "Carré", "Hiver", "FW24", 450, 1, "Ivoire", "", "Oui", "Alice Martin", "Cadeau", None, None],
        ["A0003", "C0002", "2025-01-20", "Boutique", "Tokyo Ginza", "Japon", "SKU003", "Montre édition limitée", "Horlogerie", "Montre", "Signature", "Permanent", 12800, 1, "Noir", "", "Non", "Kenji Sato", "Collection", None, None],
    ]
    for r_idx, row in enumerate(sample_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    for r in range(2, 302):
        ws[f"T{r}"] = f'=IF(C{r}="","",TEXT(C{r},"yyyy-mm"))'
        ws[f"U{r}"] = f'=IF(N{r}=0,0,M{r}/N{r})'

    add_table(ws, "tblAchats", 1, 1, 301, len(headers))
    autofit_width(ws)

    add_list_validation(ws, "B2:B5000", "=tblClients[Client_ID]")
    add_list_validation(ws, "F2:F5000", "=tblPays[Pays]")
    add_list_validation(ws, "I2:I5000", "=tblCategories[Categorie]")
    add_list_validation(ws, "Q2:Q5000", YES_NO)

    mark_input_columns(ws, ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S"], 2, 300)
    mark_formula_columns(ws, ["T","U"], 2, 300)

    set_number_format(ws, "C", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "M", 2, 300, '#,##0.00')
    set_number_format(ws, "U", 2, 300, '#,##0.00')
    return ws


# =========================================================
# INTERACTIONS
# =========================================================
def build_interactions_sheet(wb):
    headers = [
        "Interaction_ID", "Client_ID", "Date_Interaction", "Canal", "Type_Interaction", "Objet",
        "Message_Envoye", "Conseiller", "Statut_Interaction", "Date_Reponse", "Montant_Converti",
        "Prochaine_Action", "Date_Prochaine_Action", "Commentaire", "Sentiment"
    ]
    ws = create_sheet_with_headers(wb, "INTERACTIONS", headers, tab_color=COLORS["orange_dark"], header_color=COLORS["orange_dark"])

    sample_data = [
        ["I0001", "C0001", "2025-03-01", "WhatsApp", "Suivi collection", "Présentation nouveautés", "Bonjour Claire, nous serions ravis de vous présenter notre nouvelle sélection.", "Alice Martin", "Répondu", "2025-03-01", 0, "Planifier RDV", "2025-03-05", "Réponse positive", "Positif"],
        ["I0002", "C0002", "2025-02-01", "LINE", "Invitation", "Invitation événement privé", "Dear Hiro, we would be delighted to welcome you...", "Kenji Sato", "Converti", "2025-02-02", 12800, "Remerciement", "2025-02-03", "Présence confirmée", "Positif"],
    ]
    for r_idx, row in enumerate(sample_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    add_table(ws, "tblInteractions", 1, 1, 301, len(headers))
    autofit_width(ws)

    add_list_validation(ws, "B2:B5000", "=tblClients[Client_ID]")
    add_list_validation(ws, "D2:D5000", "=tblCanaux[Canal]")
    add_list_validation(ws, "I2:I5000", "=tblStatutsInteraction[Statut_Interaction]")

    mark_input_columns(ws, ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"], 2, 300)
    set_number_format(ws, "C", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "J", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "M", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "K", 2, 300, '#,##0.00')
    return ws


# =========================================================
# CATALOGUE_CADEAUX
# =========================================================
def build_catalogue_cadeaux_sheet(wb):
    headers = [
        "Gift_ID", "Maison", "Produit_Cadeau", "Categorie", "Sous_Categorie", "Univers",
        "Prix_Public", "Budget_Min", "Budget_Max", "Genre_Cible", "Occasion_Ideale", "Saison",
        "Pays_Autorise", "Disponible", "Niveau_Exclusivite", "VIP_Compatible", "VIC_Compatible",
        "Message_Associe", "Priorite_Mise_En_Avant", "Stock_Critique", "Actif_Cadeau"
    ]
    ws = create_sheet_with_headers(wb, "CATALOGUE_CADEAUX", headers, tab_color=COLORS["gray"], header_color=COLORS["gray"])

    sample_data = [
        ["G0001", "Maison A", "Mini sac iconique", "Maroquinerie", "Sac", "Femme", 3200, 2000, 5000, "Femme", "Anniversaire", "Permanent", "France", "Oui", 5, "Oui", "Oui", "Sélection anniversaire iconique", 10, "Non", None],
        ["G0002", "Maison A", "Carré de soie signature", "Soie", "Carré", "Mixte", 450, 300, 700, "Mixte", "Remerciement achat", "Permanent", "France", "Oui", 3, "Oui", "Oui", "Attention raffinée", 8, "Non", None],
        ["G0003", "Maison B", "Montre complication", "Horlogerie", "Montre", "Homme", 15000, 10000, 25000, "Homme", "Événement boutique", "Permanent", "Japon", "Oui", 5, "Oui", "Oui", "Pièce rare pour collectionneur", 9, "Non", None],
    ]
    for r_idx, row in enumerate(sample_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    for r in range(2, 302):
        ws[f"U{r}"] = f'=IF(AND(N{r}="Oui",T{r}<>"Oui"),"Oui","Non")'

    add_table(ws, "tblGifts", 1, 1, 301, len(headers))
    autofit_width(ws)

    add_list_validation(ws, "B2:B5000", "=tblMaisons[Maison]")
    add_list_validation(ws, "D2:D5000", "=tblCategories[Categorie]")
    add_list_validation(ws, "N2:N5000", YES_NO)
    add_list_validation(ws, "P2:Q5000", YES_NO)
    add_list_validation(ws, "T2:T5000", YES_NO)

    mark_input_columns(ws, ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"], 2, 300)
    mark_formula_columns(ws, ["U"], 2, 300)
    set_number_format(ws, "G", 2, 300, '#,##0.00')
    set_number_format(ws, "H", 2, 300, '#,##0.00')
    set_number_format(ws, "I", 2, 300, '#,##0.00')
    return ws


# =========================================================
# MOMENTS
# =========================================================
def build_moments_sheet(wb):
    headers = [
        "Client_ID", "Nom_Client", "Segment_Client", "Occasion_Detectee", "Date_Evenement",
        "Priorite_Occasion", "Canal_Recommande", "Cadeau_Suggere_1", "Cadeau_Suggere_2",
        "Cadeau_Suggere_3", "Message_Template_ID", "Validation_Humaine", "Statut_Moment",
        "Action_Suggeree", "Date_Action_Cible"
    ]
    ws = create_sheet_with_headers(wb, "MOMENTS", headers, tab_color=COLORS["green"], header_color=COLORS["green"])

    for r in range(2, 302):
        ws[f"A{r}"] = f'=IF(CLIENTS!A{r}="","",CLIENTS!A{r})'
        ws[f"B{r}"] = f'=IF(A{r}="","",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!F:F&" "&CLIENTS!G:G,""))'
        ws[f"C{r}"] = f'=IF(A{r}="","",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!M:M,""))'
        ws[f"D{r}"] = f'''=IF(A{r}="","",
IF(AND(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!L:L,"")<>"",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!L:L,"")-TODAY()>=0,XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!L:L,"")-TODAY()<=XLOOKUP("Jours_Anniversaire_Alert",PARAM_REGLES!D:D,PARAM_REGLES!E:E,15)),"Anniversaire",
IF(AND(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,"")<>"",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,"")>=XLOOKUP("Jours_Dormance_2",PARAM_REGLES!D:D,PARAM_REGLES!E:E,180)),"Réactivation 180 jours",
IF(AND(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,"")<>"",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,"")>=XLOOKUP("Jours_Dormance_1",PARAM_REGLES!D:D,PARAM_REGLES!E:E,90)),"Réactivation 90 jours",
IF(AND(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,"")<>"",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,"")<=XLOOKUP("Jours_Post_Achat_Remerciement",PARAM_REGLES!D:D,PARAM_REGLES!E:E,7)),"Remerciement achat","RAS"))))'''
        ws[f"E{r}"] = f'=IF(D{r}="Anniversaire",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!L:L,""),IF(D{r}="Remerciement achat",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!O:O,""),TODAY()))'
        ws[f"F{r}"] = f'=IFERROR(XLOOKUP(D{r},PARAM_LISTES!H:H,PARAM_LISTES!I:I,0),0)'
        ws[f"G{r}"] = f'=IF(A{r}="","",IF(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!Q:Q,"")<>"",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!Q:Q,""),XLOOKUP(D{r},PARAM_LISTES!H:H,PARAM_LISTES!L:L,"")))'
        ws[f"H{r}"] = f'=IFERROR(INDEX(FILTER(CATALOGUE_CADEAUX!C:C,(CATALOGUE_CADEAUX!D:D=XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AC:AC,""))*(CATALOGUE_CADEAUX!U:U="Oui")),1),"")'
        ws[f"I{r}"] = f'=IFERROR(INDEX(FILTER(CATALOGUE_CADEAUX!C:C,(CATALOGUE_CADEAUX!D:D=XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AC:AC,""))*(CATALOGUE_CADEAUX!U:U="Oui")),2),"")'
        ws[f"J{r}"] = f'=IFERROR(INDEX(FILTER(CATALOGUE_CADEAUX!C:C,(CATALOGUE_CADEAUX!D:D=XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AC:AC,""))*(CATALOGUE_CADEAUX!U:U="Oui")),3),"")'
        ws[f"K{r}"] = f'=IF(A{r}="","",XLOOKUP(D{r},TEMPLATES_MESSAGES!C:C,TEMPLATES_MESSAGES!A:A,""))'
        ws[f"L{r}"] = f'=IFERROR(XLOOKUP(D{r},PARAM_LISTES!H:H,PARAM_LISTES!M:M,"Oui"),"Oui")'
        ws[f"M{r}"] = f'=IF(D{r}="RAS","À surveiller","Actionnable")'
        ws[f"N{r}"] = f'=IF(D{r}="Anniversaire","Préparer message anniversaire",IF(D{r}="Réactivation 90 jours","Préparer message de réactivation",IF(D{r}="Remerciement achat","Envoyer remerciement premium","Aucune")))'
        ws[f"O{r}"] = f'=IF(E{r}="","",E{r}-2)'

    add_table(ws, "tblMoments", 1, 1, 301, len(headers))
    autofit_width(ws)
    mark_formula_columns(ws, [chr(c) for c in range(ord("A"), ord("O")+1)], 2, 300)
    set_number_format(ws, "E", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "O", 2, 300, "dd/mm/yyyy")
    return ws


# =========================================================
# SCORING
# =========================================================
def build_scoring_sheet(wb):
    headers = [
        "Client_ID", "Score_Valeur", "Score_Recence", "Score_Frequence", "Score_Engagement",
        "Score_Occasion", "Score_Conformite", "Score_Attrition", "Score_Total", "Priorite_Action", "Rang"
    ]
    ws = create_sheet_with_headers(wb, "SCORING", headers, tab_color=COLORS["green_dark"], header_color=COLORS["green_dark"])

    for r in range(2, 302):
        ws[f"A{r}"] = f'=IF(CLIENTS!A{r}="","",CLIENTS!A{r})'
        ws[f"B{r}"] = f'=MIN(100,ROUND(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!Z:Z,0)/100,0))'
        ws[f"C{r}"] = f'=LET(j,XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,999),MAX(0,100-j/3))'
        ws[f"D{r}"] = f'=MIN(100,XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AA:AA,0)*10)'
        ws[f"E{r}"] = f'=LET(id,A{r},tot,COUNTIF(INTERACTIONS!B:B,id),rep,COUNTIFS(INTERACTIONS!B:B,id,INTERACTIONS!I:I,"Répondu"),IF(tot=0,20,ROUND(rep/tot*100,0)))'
        ws[f"F{r}"] = f'=MIN(100,XLOOKUP(A{r},MOMENTS!A:A,MOMENTS!F:F,0)*20)'
        ws[f"G{r}"] = f'=IF(XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AM:AM,"Alerte")="OK",100,40)'
        ws[f"H{r}"] = f'=LET(j,XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!AH:AH,999),MIN(100,j/2))'
        ws[f"I{r}"] = f'=ROUND(B{r}*XLOOKUP("Valeur_Client",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.25)+C{r}*XLOOKUP("Recence_Achat",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.15)+D{r}*XLOOKUP("Frequence_Achat",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.10)+E{r}*XLOOKUP("Engagement",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.15)+F{r}*XLOOKUP("Occasion",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.15)+G{r}*XLOOKUP("Conformite",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.10)+H{r}*XLOOKUP("Risque_Attrition",PARAM_REGLES!A:A,PARAM_REGLES!B:B,0.10),0)'
        ws[f"J{r}"] = f'=IF(I{r}>=XLOOKUP("Score_P1",PARAM_REGLES!D:D,PARAM_REGLES!E:E,85),"P1",IF(I{r}>=XLOOKUP("Score_P2",PARAM_REGLES!D:D,PARAM_REGLES!E:E,70),"P2",IF(I{r}>=XLOOKUP("Score_P3",PARAM_REGLES!D:D,PARAM_REGLES!E:E,50),"P3","P4")))'
        ws[f"K{r}"] = f'=RANK.EQ(I{r},I$2:I$301,0)'

    add_table(ws, "tblScoring", 1, 1, 301, len(headers))
    autofit_width(ws)
    mark_formula_columns(ws, [chr(c) for c in range(ord("A"), ord("K")+1)], 2, 300)
    return ws


# =========================================================
# TEMPLATES_MESSAGES
# =========================================================
def build_templates_sheet(wb):
    headers = [
        "Template_ID", "Maison", "Occasion", "Segment_Client", "Canal", "Langue", "Ton",
        "Objet_Template", "Intro_Template", "Corps_Template", "CTA_Template", "Signature_Template",
        "Mention_Obligatoire", "Actif"
    ]
    ws = create_sheet_with_headers(wb, "TEMPLATES_MESSAGES", headers, tab_color=COLORS["purple"], header_color=COLORS["purple"])

    sample_data = [
        ["T001", "Maison A", "Anniversaire", "VIP", "WhatsApp", "FR", "Formel premium", "Une attention pensée pour vous, {PRENOM}", "Bonjour {PRENOM},", "À l’approche de votre anniversaire, nous serions ravis de vous proposer une attention particulière pensée pour vous.", "Si vous le souhaitez, je peux vous partager une sélection personnalisée.", "Bien à vous, {CONSEILLER}", "Sous réserve de disponibilité.", "Oui"],
        ["T002", "Maison A", "Remerciement achat", "VIP", "Email", "FR", "Formel premium", "Merci pour votre visite", "Bonjour {PRENOM},", "Je tenais à vous remercier chaleureusement pour votre récente visite et votre confiance.", "Je reste naturellement à votre disposition pour toute sélection complémentaire.", "Bien à vous, {CONSEILLER}", "Sous réserve de disponibilité.", "Oui"],
        ["T003", "Maison B", "Anniversaire", "VIC", "LINE", "JP", "Très formel", "A personal selection for you", "Dear {PRENOM},", "As your special day approaches, we would be delighted to prepare a personal selection for you.", "Please let me know if you would like me to share curated suggestions.", "Kind regards, {CONSEILLER}", "Subject to availability.", "Oui"],
    ]
    for r_idx, row in enumerate(sample_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    add_table(ws, "tblTemplates", 1, 1, 301, len(headers))
    autofit_width(ws)

    add_list_validation(ws, "B2:B5000", "=tblMaisons[Maison]")
    add_list_validation(ws, "C2:C5000", "=tblOccasions[Occasion]")
    add_list_validation(ws, "D2:D5000", "=tblSegments[Segment_Client]")
    add_list_validation(ws, "E2:E5000", "=tblCanaux[Canal]")
    add_list_validation(ws, "F2:F5000", "=tblLangues[Langue]")
    add_list_validation(ws, "G2:G5000", "=tblTons[Ton]")
    add_list_validation(ws, "N2:N5000", YES_NO)

    mark_input_columns(ws, [chr(c) for c in range(ord("A"), ord("N")+1)], 2, 300)
    return ws


# =========================================================
# MESSAGES_GEN
# =========================================================
def build_messages_gen_sheet(wb):
    headers = [
        "Message_ID", "Client_ID", "Template_ID", "Maison", "Canal", "Langue", "Objet_Final",
        "Message_Final", "Cadeau_Insere", "Conseiller", "Date_Generation", "Statut_Message"
    ]
    ws = create_sheet_with_headers(wb, "MESSAGES_GEN", headers, tab_color="8FAADC", header_color="8FAADC")

    for r in range(2, 302):
        ws[f"A{r}"] = f'=IF(MOMENTS!A{r}="","","MSG-"&TEXT(ROW()-1,"0000"))'
        ws[f"B{r}"] = f'=IF(MOMENTS!A{r}="","",MOMENTS!A{r})'
        ws[f"C{r}"] = f'''=IF(B{r}="","",IFERROR(INDEX(FILTER(TEMPLATES_MESSAGES!A:A,(TEMPLATES_MESSAGES!B:B=XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!B:B,""))*(TEMPLATES_MESSAGES!C:C=XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!D:D,""))*(TEMPLATES_MESSAGES!F:F=XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!J:J,""))*(TEMPLATES_MESSAGES!N:N="Oui")),1),""))'''
        ws[f"D{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!B:B,""))'
        ws[f"E{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!G:G,""))'
        ws[f"F{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!J:J,""))'
        ws[f"G{r}"] = f'=IF(C{r}="","",SUBSTITUTE(XLOOKUP(C{r},TEMPLATES_MESSAGES!A:A,TEMPLATES_MESSAGES!H:H,""),"{{PRENOM}}",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!F:F,"")))'
        ws[f"I{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!H:H,""))'
        ws[f"J{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!D:D,""))'
        ws[f"H{r}"] = f'''=IF(C{r}="","",SUBSTITUTE(SUBSTITUTE(TEXTJOIN(" ",TRUE,XLOOKUP(C{r},TEMPLATES_MESSAGES!A:A,TEMPLATES_MESSAGES!I:I,""),XLOOKUP(C{r},TEMPLATES_MESSAGES!A:A,TEMPLATES_MESSAGES!J:J,""),IF(I{r}<>"","Nous avons pensé à une sélection autour de "&I{r}&".",""),XLOOKUP(C{r},TEMPLATES_MESSAGES!A:A,TEMPLATES_MESSAGES!K:K,""),XLOOKUP(C{r},TEMPLATES_MESSAGES!A:A,TEMPLATES_MESSAGES!L:L,"")),"{{PRENOM}}",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!F:F,"")),"{{CONSEILLER}}",J{r}))'''
        ws[f"K{r}"] = '=TODAY()'
        ws[f"L{r}"] = '="Généré"'

    add_table(ws, "tblMessages", 1, 1, 301, len(headers))
    autofit_width(ws)
    mark_formula_columns(ws, [chr(c) for c in range(ord("A"), ord("L")+1)], 2, 300)
    set_number_format(ws, "K", 2, 300, "dd/mm/yyyy")
    return ws


# =========================================================
# CONTROLES
# =========================================================
def build_controles_sheet(wb):
    headers = [
        "Message_ID", "Client_ID", "Canal", "Pays", "Optin_OK", "Pression_OK", "Brand_OK",
        "Mention_OK", "Lexique_OK", "Validation_Humaine", "Statut_Final", "Commentaire_Controle"
    ]
    ws = create_sheet_with_headers(wb, "CONTROLES", headers, tab_color=COLORS["orange"], header_color=COLORS["orange"])

    for r in range(2, 302):
        ws[f"A{r}"] = f'=IF(MESSAGES_GEN!A{r}="","",MESSAGES_GEN!A{r})'
        ws[f"B{r}"] = f'=IF(A{r}="","",XLOOKUP(A{r},MESSAGES_GEN!A:A,MESSAGES_GEN!B:B,""))'
        ws[f"C{r}"] = f'=IF(A{r}="","",XLOOKUP(A{r},MESSAGES_GEN!A:A,MESSAGES_GEN!E:E,""))'
        ws[f"D{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!H:H,""))'
        ws[f"E{r}"] = f'''=IF(C{r}="","",IF(AND(C{r}="Email",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!R:R,"")="Oui"),"OK",IF(AND(C{r}="SMS",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!S:S,"")="Oui"),"OK",IF(AND(C{r}="WhatsApp",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!T:T,"")="Oui"),"OK",IF(AND(C{r}="Téléphone",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!U:U,"")="Oui"),"OK",IF(AND(C{r}="Courrier",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!V:V,"")="Oui"),"OK",IF(AND(C{r}="WeChat",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!W:W,"")="Oui"),"OK",IF(AND(C{r}="LINE",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!X:X,"")="Oui"),"OK",IF(AND(C{r}="KakaoTalk",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!Y:Y,"")="Oui"),"OK","Alerte")))))))))'''
        ws[f"F{r}"] = f'=IF(XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!AK:AK,0)<=2,"OK","Alerte")'
        ws[f"I{r}"] = f'=IF(SUMPRODUCT(--ISNUMBER(SEARCH(PARAM_REGLES!A25:A100,XLOOKUP(A{r},MESSAGES_GEN!A:A,MESSAGES_GEN!H:H,""))))>0,"A corriger","OK")'
        ws[f"G{r}"] = f'=IF(I{r}="OK","OK","A corriger")'
        ws[f"H{r}"] = f'=IF(XLOOKUP(XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!B:B,""),PARAM_REGLES!A:A,PARAM_REGLES!H:H,"")<>"","OK","À vérifier")'
        ws[f"J{r}"] = f'=XLOOKUP(XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!B:B,""),PARAM_REGLES!A:A,PARAM_REGLES!I:I,"Oui")'
        ws[f"K{r}"] = f'=IF(OR(E{r}="Alerte",F{r}="Alerte"),"Bloqué",IF(OR(G{r}="A corriger",I{r}="A corriger",J{r}="Oui"),"À valider","Prêt à envoyer"))'
        ws[f"L{r}"] = f'=TEXTJOIN(" | ",TRUE,IF(E{r}<>"OK","Vérifier consentement",""),IF(F{r}<>"OK","Pression relationnelle trop élevée",""),IF(I{r}<>"OK","Lexique interdit détecté",""),IF(J{r}="Oui","Validation manuelle requise",""))'

    add_table(ws, "tblControles", 1, 1, 301, len(headers))
    autofit_width(ws)
    mark_formula_columns(ws, [chr(c) for c in range(ord("A"), ord("L")+1)], 2, 300)
    add_conditional_status_colors(ws, "K2:K5000")
    return ws


# =========================================================
# QUEUE_ACTIONS
# =========================================================
def build_queue_actions_sheet(wb):
    headers = [
        "Action_ID", "Client_ID", "Client", "Conseiller", "Segment", "Occasion", "Score_Total",
        "Priorite_Action", "Canal", "Cadeau_Suggere", "Message_ID", "Statut_Controle",
        "Action_A_Faire", "Deadline", "Statut_Action", "Date_Envoi", "Resultat"
    ]
    ws = create_sheet_with_headers(wb, "QUEUE_ACTIONS", headers, tab_color=COLORS["green"], header_color=COLORS["green"])

    for r in range(2, 302):
        ws[f"A{r}"] = f'=IF(MOMENTS!A{r}="","","ACT-"&TEXT(ROW()-1,"0000"))'
        ws[f"B{r}"] = f'=IF(MOMENTS!A{r}="","",MOMENTS!A{r})'
        ws[f"C{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!F:F&" "&CLIENTS!G:G,""))'
        ws[f"D{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!D:D,""))'
        ws[f"E{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},CLIENTS!A:A,CLIENTS!M:M,""))'
        ws[f"F{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!D:D,""))'
        ws[f"G{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},SCORING!A:A,SCORING!I:I,0))'
        ws[f"H{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},SCORING!A:A,SCORING!J:J,""))'
        ws[f"I{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!G:G,""))'
        ws[f"J{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!H:H,""))'
        ws[f"K{r}"] = f'=IFERROR(INDEX(FILTER(MESSAGES_GEN!A:A,MESSAGES_GEN!B:B=B{r}),1),"")'
        ws[f"L{r}"] = f'=IF(K{r}="","",XLOOKUP(K{r},CONTROLES!A:A,CONTROLES!K:K,""))'
        ws[f"M{r}"] = f'=IF(L{r}="Prêt à envoyer","Envoyer message",IF(L{r}="À valider","Valider et ajuster message",IF(L{r}="Bloqué","Corriger conformité","Analyser")))'
        ws[f"N{r}"] = f'=IF(B{r}="","",XLOOKUP(B{r},MOMENTS!A:A,MOMENTS!O:O,TODAY()))'

    add_table(ws, "tblQueue", 1, 1, 301, len(headers))
    autofit_width(ws)
    add_list_validation(ws, "O2:O5000", '"À faire,En cours,Validé,Envoyé,Reporté,Annulé"')

    mark_formula_columns(ws, ["A","B","C","D","E","F","G","H","I","J","K","L","M","N"], 2, 300)
    mark_input_columns(ws, ["O","P","Q"], 2, 300)
    set_number_format(ws, "N", 2, 300, "dd/mm/yyyy")
    set_number_format(ws, "P", 2, 300, "dd/mm/yyyy")
    return ws


# =========================================================
# DASHBOARD
# =========================================================
def build_dashboard_sheet(wb):
    ws = wb.create_sheet("DASHBOARD")
    color_tab(ws, COLORS["navy_dark"])
    freeze_top(ws, "A1")

    rows = [
        ["ClientelingAI Dashboard"],
        [""],
        ["KPI", "Valeur"],
        ["Nb_Clients_Actionnables", '=COUNTIF(MOMENTS!M:M,"Actionnable")'],
        ["Nb_P1", '=COUNTIF(SCORING!J:J,"P1")'],
        ["Nb_Messages_Prets", '=COUNTIF(CONTROLES!K:K,"Prêt à envoyer")'],
        ["Nb_Messages_A_Valider", '=COUNTIF(CONTROLES!K:K,"À valider")'],
        ["Nb_Bloques", '=COUNTIF(CONTROLES!K:K,"Bloqué")'],
        ["CA_12M_Total", '=SUM(CLIENTS!Z:Z)'],
        ["Taux_Reponse", '=IF(COUNTA(INTERACTIONS!A:A)=0,0,COUNTIF(INTERACTIONS!I:I,"Répondu")/COUNTA(INTERACTIONS!A:A))'],
        ["Taux_Conversion", '=IF(COUNTA(INTERACTIONS!A:A)=0,0,COUNTIF(INTERACTIONS!I:I,"Converti")/COUNTA(INTERACTIONS!A:A))'],
        ["Montant_Converti_Total", '=SUM(INTERACTIONS!K:K)'],
        [""],
        ["Top 10 clients prioritaires"],
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    ws["A1"].font = Font(size=16, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy_dark"])
    style_header(ws, 3, fill_color=COLORS["navy"])

    ws["A15"] = "Client_ID"
    ws["B15"] = "Score_Total"
    ws["C15"] = "Priorite_Action"
    style_header(ws, 15, fill_color=COLORS["navy"])

    for r in range(16, 26):
        idx = r - 15
        ws[f"A{r}"] = f'=INDEX(SORTBY(SCORING!A2:A301,SCORING!I2:I301,-1),{idx})'
        ws[f"B{r}"] = f'=XLOOKUP(A{r},SCORING!A:A,SCORING!I:I,"")'
        ws[f"C{r}"] = f'=XLOOKUP(A{r},SCORING!A:A,SCORING!J:J,"")'

    set_number_format(ws, "B", 4, 12, '#,##0.00')
    autofit_width(ws)
    return ws


# =========================================================
# VEILLE_MAJ
# =========================================================
def build_veille_sheet(wb):
    headers = [
        "Veille_ID", "Domaine", "Pays_Zone", "Sujet", "Description_Synthetique",
        "Impact_Sur_ClientelingAI", "Action_Appliquee_Dans_Fichier", "Responsable",
        "Date_MAJ", "Source", "Statut"
    ]
    ws = create_sheet_with_headers(wb, "VEILLE_MAJ", headers, tab_color=COLORS["gold"], header_color=COLORS["gold"])

    sample_data = [
        ["V001", "RGPD", "UE", "Marketing direct", "Vérification régulière des bases légales et consentements", "Renforce contrôle opt-in", "Contrôle Optin_OK conservé", "Data Privacy Lead", "2025-01-15", "CNIL", "Actif"],
        ["V002", "Luxe / Brand", "Global", "Brand voice", "Mise à jour du lexique premium et interdit", "Ajuste Brand_OK", "Table lexique revue", "Brand Director", "2025-01-20", "Interne", "Actif"],
    ]
    for r_idx, row in enumerate(sample_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)

    add_table(ws, "tblVeille", 1, 1, 301, len(headers))
    ws["M1"] = "Fraîcheur veille"
    ws["M2"] = '=IF(TODAY()-MAX(I:I)>30,"ALERTE : veille à actualiser","Veille à jour")'
    style_note_cell(ws["M1"])
    style_note_cell(ws["M2"])

    mark_input_columns(ws, [chr(c) for c in range(ord("A"), ord("K")+1)], 2, 300)
    set_number_format(ws, "I", 2, 300, "dd/mm/yyyy")
    autofit_width(ws)
    return ws


# =========================================================
# EXPORT_CRM
# =========================================================
def build_export_crm_sheet(wb):
    headers = [
        "Client_ID", "Client", "Canal", "Objet", "Message_Final", "Conseiller",
        "Date_Envoi_Prevue", "Code_Campagne", "Tag_CRM", "Validation_Finale", "Commentaire"
    ]
    ws = create_sheet_with_headers(wb, "EXPORT_CRM", headers, tab_color=COLORS["teal"], header_color=COLORS["teal"])

    for r in range(2, 302):
        ws[f"A{r}"] = f'=IF(XLOOKUP(MESSAGES_GEN!A{r},CONTROLES!A:A,CONTROLES!K:K,"")="Prêt à envoyer",MESSAGES_GEN!B{r},"")'
        ws[f"B{r}"] = f'=IF(A{r}="","",XLOOKUP(A{r},CLIENTS!A:A,CLIENTS!F:F&" "&CLIENTS!G:G,""))'
        ws[f"C{r}"] = f'=IF(A{r}="","",MESSAGES_GEN!E{r})'
        ws[f"D{r}"] = f'=IF(A{r}="","",MESSAGES_GEN!G{r})'
        ws[f"E{r}"] = f'=IF(A{r}="","",MESSAGES_GEN!H{r})'
        ws[f"F{r}"] = f'=IF(A{r}="","",MESSAGES_GEN!J{r})'
        ws[f"G{r}"] = f'=IF(A{r}="","",TODAY())'
        ws[f"J{r}"] = f'=IF(A{r}="","",XLOOKUP(MESSAGES_GEN!A{r},CONTROLES!A:A,CONTROLES!K:K,""))'

    add_table(ws, "tblExport", 1, 1, 301, len(headers))
    autofit_width(ws)

    mark_formula_columns(ws, ["A","B","C","D","E","F","G","J"], 2, 300)
    mark_input_columns(ws, ["H","I","K"], 2, 300)
    set_number_format(ws, "G", 2, 300, "dd/mm/yyyy")
    return ws


# =========================================================
# MAIN
# =========================================================
def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme_sheet(wb)
    build_data_dictionary_sheet(wb)
    build_param_listes_sheet(wb)
    build_param_regles_sheet(wb)

    build_clients_sheet(wb)
    build_achats_sheet(wb)
    build_interactions_sheet(wb)
    build_catalogue_cadeaux_sheet(wb)

    build_moments_sheet(wb)
    build_scoring_sheet(wb)
    build_templates_sheet(wb)
    build_messages_gen_sheet(wb)
    build_controles_sheet(wb)

    build_queue_actions_sheet(wb)
    build_dashboard_sheet(wb)
    build_veille_sheet(wb)
    build_export_crm_sheet(wb)

    finalize_workbook_compatibility(wb)

    wb.save(OUTPUT_FILE)
    print(f"Fichier généré : {OUTPUT_FILE}")


if __name__ == "__main__":
    build_workbook()
