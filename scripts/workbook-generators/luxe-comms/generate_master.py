"""
NEURAL_LUXE_COMMS_MASTER.xlsx
Pilotage global : gouvernance, RACI, workflow map, data products, logs agreges.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, STATUS, VERT, VIOLET, YESNO,
    ajouter_table, attach_dv, cf_stale, convertir_dates, D, DT, dv_liste,
    ecrire_donnees, en_tetes, params_table, param_ligne, section, titre,
    today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_LUXE_COMMS_MASTER.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "NEURAL / LUXE / COMMUNICATION — MASTER", "Gouvernance, RACI, workflow, data products, pilotage global", 6)
    ws.column_dimensions["A"].width = 4
    for c in "BCDEF":
        ws.column_dimensions[c].width = 22
    ws.column_dimensions["C"].width = 80

    section(ws, 4, 6, "Objet du workbook", NOIR)
    ecrire_donnees(ws, 5, [
        ["", "Role",             "Fichier de pilotage central. Consolide les logs et KPIs des 5 agents."],
        ["", "Proprietaire",     "AI Ops / CMO."],
        ["", "Cadence",          "Refresh quotidien des imports de logs. Review hebdo des gates ouvertes."],
        ["", "Principe",         "AUCUNE logique de scoring critique ici. Uniquement agregation et dashboards."],
        ["", "Imports",          "Power Query depuis 5 fichiers NEURAL_AG00X_*.xlsx et FOUNDATIONS."],
        ["", "Output",           "Dashboard executive + risk register + release plan."],
    ], align=AL)

    section(ws, 13, 6, "Ordre de refresh (critique)", BORDEAUX)
    ecrire_donnees(ws, 14, [
        ["", "1", "FOUNDATIONS (referentiel doit etre frais)"],
        ["", "2", "AG-001 VoiceGuard (consomme FOUNDATIONS)"],
        ["", "3", "AG-005 GreenClaim (consomme FOUNDATIONS)"],
        ["", "4", "AG-002 LuxePressAgent (dependances AG-001, AG-005)"],
        ["", "5", "AG-004 HeritageComms"],
        ["", "6", "AG-003 LuxeEventComms (dependances AG-001, AG-004)"],
        ["", "7", "MASTER (import final)"],
    ], align=AL)


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES WORKBOOK", "tblParamsMaster consommee via XLOOKUP", 3)
    rows = [
        ("WORKBOOK", "MASTER", ""),
        ("MAISON_ID", "M-001", ""),
        ("MAISON_NAME", "Maison pilote", ""),
        ("LANG_PRIMARY", "FR", ""),
        ("LANG_SECONDARY", "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", ""),
        ("DATA_VERSION", "v2026.04", ""),
        ("PROMPT_VERSION", "v1.0.0", ""),
        ("LAST_REFRESH", D(today_iso()), ""),
        ("AGENTS_DEPLOYED", 5, ""),
        ("SLA_REVIEW_DEFAULT_H", 24, "Heures"),
        ("ESCALATION_CEO_THRESHOLD", "CRITICAL", ""),
        ("CRISIS_MODE_ON", "NO", "Si YES, SLA reduit global"),
    ]
    params_table(ws, "tblParamsMaster", start_row=4, rows=rows)


def onglet_scope_raci(wb):
    ws = wb.create_sheet("1_SCOPE_RACI")
    titre(ws, "SCOPE & RACI", "Perimetre v1 + matrice de responsabilites", 7)
    section(ws, 4, 7, "Scope v1", NOIR)
    en_tetes(ws, 5, ["SCOPE_ID", "ITEM", "IN_SCOPE", "MAISON_ID", "LANGUES", "CANAUX", "NOTE"],
             [12, 32, 12, 14, 16, 22, 40])
    scope = [
        ("S-001", "Maison pilote",                    "YES", "M-001", "FR, EN",    "Presse, event, social", "Unique maison active v1"),
        ("S-002", "5 agents (AG-001 a AG-005)",       "YES", "M-001", "FR, EN",    "Tous",                  "Brand, Press, Event, Heritage, Claim"),
        ("S-003", "Canal Paid media",                 "NO",  "-",     "-",         "-",                     "Reporte en v2"),
        ("S-004", "Crisis fast-track",                "YES", "M-001", "FR, EN",    "Tous",                  "Flag dans 1_MAISON_PROFILE"),
        ("S-005", "Multi-maison (N maisons)",         "NO",  "-",     "-",         "-",                     "Archi ready, pas active"),
        ("S-006", "Langues IT, DE, JA, ZH",          "NO",  "-",     "-",         "-",                     "Vocab seed requis v2"),
    ]
    last = ecrire_donnees(ws, 6, scope, align=AL)
    ajouter_table(ws, "tblScope", f"A5:G{last}")

    section(ws, last + 2, 7, "Matrice RACI", BORDEAUX)
    en_tetes(ws, last + 3, ["ROLE", "OWNER", "AG-001 Voice", "AG-002 Press", "AG-003 Event", "AG-004 Heritage", "AG-005 Claim"],
             [22, 24, 14, 14, 14, 14, 14])
    raci = [
        ("Brand Director",          "A. Dupont",    "A", "C", "C", "C", "I"),
        ("PR Director",             "B. Martin",    "C", "A", "C", "I", "C"),
        ("Compliance / ESG",        "C. Leblanc",   "I", "C", "I", "I", "A"),
        ("Heritage Director",       "D. Laurent",   "I", "I", "C", "A", "I"),
        ("Legal",                   "E. Moreau",    "C", "C", "C", "C", "C"),
        ("Atelier / Savoir-faire",  "F. Durand",    "I", "C", "R", "R", "C"),
        ("CMO",                     "G. Petit",     "A", "A", "A", "A", "A"),
        ("CEO",                     "H. Robert",    "I", "I", "I", "I", "I"),
        ("AI Ops",                  "Ludo",         "R", "R", "R", "R", "R"),
    ]
    last2 = ecrire_donnees(ws, last + 4, raci, align=AL)
    ajouter_table(ws, "tblRACI", f"A{last+3}:G{last2}")

    # legend RACI
    section(ws, last2 + 2, 7, "Legende RACI", VIOLET)
    ecrire_donnees(ws, last2 + 3, [
        ["R", "Responsible",  "Execute",               "", "", "", ""],
        ["A", "Accountable",  "Rend des comptes",      "", "", "", ""],
        ["C", "Consulted",    "Consulte",              "", "", "", ""],
        ["I", "Informed",     "Informe",               "", "", "", ""],
    ], align=AL)


def onglet_agent_registry(wb):
    ws = wb.create_sheet("2_AGENT_REGISTRY")
    titre(ws, "REGISTRE DES AGENTS", "Une ligne par agent — scope, mission, owner, dependances", 11)
    headers = ["AGENT_ID", "AGENT_NAME", "MISSION", "OWNER", "STATUS", "PRIORITY",
               "INPUT_MAIN", "OUTPUT_MAIN", "PRIMARY_GATE", "DEPENDENCIES", "DATA_REQUIRED"]
    en_tetes(ws, 4, headers, [10, 22, 46, 18, 12, 10, 28, 32, 14, 30, 30])

    agents = [
        ("AG-001", "MaisonVoiceGuard", "Scorer chaque communication sur la conformite charte de marque; refus automatique si score insuffisant",
         "Brand Director", "ACTIVE", "P0",
         "Texte + LANG + CONTEXT", "Score + Decision (APPROVE/REWORK/REJECT) + Log",
         "BRAND", "FOUNDATIONS",
         "Brand Charter + Vocab FR/EN"),
        ("AG-002", "LuxePressAgent", "Rediger communiques dans le registre du luxe, adapter presse lifestyle vs. business, relations Vogue/HB",
         "PR Director", "ACTIVE", "P0",
         "Brief presse + Media + Angle", "Communique finalise + Dossier presse",
         "BRAND", "AG-001, AG-005",
         "Media Directory + Claims"),
        ("AG-003", "LuxeEventComms", "Generer contenus pour defiles, lancements, expositions (invitations, scripts, social temps reel)",
         "PR Director", "ACTIVE", "P1",
         "Brief evenement + Format + Audience", "Pack evenementiel multi-format",
         "EVENT", "AG-001, AG-004",
         "Events Calendar + Heritage"),
        ("AG-004", "HeritageComms", "Valoriser l'heritage et l'histoire de la maison dans toutes les communications",
         "Heritage Director", "ACTIVE", "P1",
         "Query heritage + Contexte", "Narrative blocks sources + citations",
         "HERITAGE", "FOUNDATIONS",
         "Heritage Sourcebook"),
        ("AG-005", "GreenClaimChecker", "Verifier chaque affirmation RSE / developpement durable contre donnees reelles et reglements (Green Claims)",
         "ESG Lead", "ACTIVE", "P0",
         "Claim extrait + Wording type + Juridiction", "Decision (PASS/BLOCK) + Risk level + Evidence",
         "CLAIM", "FOUNDATIONS",
         "Claims Evidence Registry"),
    ]
    last = ecrire_donnees(ws, 5, agents, align=AL)

    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"ACTIVE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"PAUSED"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"P0"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"P1"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))

    ajouter_table(ws, "tblAgentRegistry", f"A4:K{last}")
    # R-06 : DV sur STATUS et PRIORITY
    attach_dv(ws, f"E5:E{last}", ["ACTIVE", "PAUSED", "DEPRECATED"])
    attach_dv(ws, f"F5:F{last}", ["P0", "P1", "P2", "P3"])
    attach_dv(ws, f"I5:I{last}", ["BRAND", "CLAIM", "HERITAGE", "LEGAL", "EVENT", "CRISIS"])


def onglet_workflow(wb):
    ws = wb.create_sheet("3_WORKFLOW_MAP")
    titre(ws, "WORKFLOW MAP", "Flux entre agents — dependances, triggers, gates, SLA", 9)
    headers = ["FLOW_ID", "SOURCE", "DESTINATION", "TRIGGER", "PAYLOAD", "AUTO", "GATE", "ESCALATION", "SLA_HOURS"]
    en_tetes(ws, 4, headers, [10, 18, 18, 26, 34, 10, 14, 22, 12])

    flows = [
        ("FL-001", "USER",   "AG-002", "Brief presse depose",              "Brief + Angle + Media cibles",    "NO",  "-",        "-",                 0),
        ("FL-002", "AG-002", "AG-001", "Draft presse cree",                "Texte draft + LANG",              "YES", "BRAND",    "Brand Director",    24),
        ("FL-003", "AG-002", "AG-005", "Draft contient terme RSE",         "Claims extraits",                 "YES", "CLAIM",    "ESG Lead + Legal",  48),
        ("FL-004", "AG-005", "AG-002", "Decision claim (PASS/BLOCK)",      "Decision + Risk level + Evidence","YES", "-",        "-",                 0),
        ("FL-005", "AG-001", "AG-002", "Decision brand (APPROVE/REWORK)",  "Score + Feedback",                "YES", "-",        "-",                 0),
        ("FL-006", "AG-002", "REVIEW", "Draft pret relecture humaine",     "Draft final",                     "NO",  "BRAND",    "CMO",               24),
        ("FL-007", "USER",   "AG-003", "Brief evenement depose",           "Brief + Type + Audience",         "NO",  "-",        "-",                 0),
        ("FL-008", "AG-003", "AG-004", "Besoin block patrimoine",          "Query heritage",                  "YES", "HERITAGE", "Heritage Dir",      72),
        ("FL-009", "AG-004", "AG-003", "Block patrimoine renvoye",         "Narrative + Sources + Citations", "YES", "-",        "-",                 0),
        ("FL-010", "AG-003", "AG-001", "Pack evenement a valider",         "Multi-format content",            "YES", "BRAND",    "Brand Director",    24),
        ("FL-011", "AG-003", "REVIEW", "Pack finalise",                    "Pack + Gate status",              "NO",  "EVENT",    "PR Director + CMO", 72),
        ("FL-012", "AG-001", "MASTER", "Log output quotidien",             "Outputs log",                     "YES", "-",        "-",                 0),
        ("FL-013", "AG-005", "MASTER", "Log decision quotidien",           "Risk log",                        "YES", "-",        "-",                 0),
        ("FL-014", "AG-004", "MASTER", "Log sourcing quotidien",           "Sourcing log",                    "YES", "-",        "-",                 0),
        ("FL-015", "CRISIS", "AG-001", "Flag crisis mode active",          "Texte + priorite",                "YES", "BRAND",    "CMO (fast-track)",  4),
    ]
    last = ecrire_donnees(ws, 5, flows, align=AL)
    ajouter_table(ws, "tblWorkflow", f"A4:I{last}")

    # SLA conditional format
    ws.conditional_formatting.add(f"I5:I{last}",
                                   FormulaRule(formula=[f"I5>48"], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_data_products(wb):
    ws = wb.create_sheet("4_DATA_PRODUCTS")
    titre(ws, "DATA PRODUCTS", "Sources de donnees critiques — criticite, refresh, acces", 10)
    headers = ["SOURCE_ID", "SOURCE_NAME", "OWNER", "CRITICALITY", "REFRESH_FREQ",
               "LAST_REVIEW", "NEXT_REVIEW", "PII", "ACCESS_LEVEL", "STATUT"]
    en_tetes(ws, 4, headers, [12, 36, 22, 14, 16, 14, 14, 10, 16, 14])

    sources = [
        ("DP-001", "Brand Charter",            "Brand Director",    "CRITICAL", "Trimestriel", D("2026-03-01"), D("2026-06-01"), "NO",  "PUBLIC_INT",    ""),
        ("DP-002", "Vocabulary FR",            "Brand Director",    "CRITICAL", "Mensuel",     D("2026-04-01"), D("2026-05-01"), "NO",  "PUBLIC_INT",    ""),
        ("DP-003", "Vocabulary EN",            "Brand Director",    "HIGH",     "Mensuel",     D("2026-03-15"), D("2026-04-15"), "NO",  "PUBLIC_INT",    ""),
        ("DP-004", "Heritage Sourcebook",      "Heritage Director", "CRITICAL", "Trimestriel", D("2026-02-01"), D("2026-05-01"), "NO",  "RESTRICTED",    ""),
        ("DP-005", "Media Directory",          "PR Director",       "HIGH",     "Mensuel",     D("2026-03-20"), D("2026-04-20"), "YES", "CONFIDENTIAL",  ""),
        ("DP-006", "Claims Evidence Registry", "ESG Lead",          "CRITICAL", "Bi-mensuel",  D("2026-04-01"), D("2026-04-15"), "NO",  "RESTRICTED",    ""),
        ("DP-007", "Events Calendar",          "PR Director",       "HIGH",     "Hebdomadaire",D("2026-04-15"), D("2026-04-22"), "NO",  "PUBLIC_INT",    ""),
        ("DP-008", "Prompts library",          "AI Ops",            "CRITICAL", "Ad-hoc",      D("2026-04-10"), D("2026-05-10"), "NO",  "RESTRICTED",    ""),
        ("DP-009", "Logs outputs agents",      "AI Ops",            "HIGH",     "Quotidien",   D(today_iso()),  None,            "YES", "RESTRICTED",    ""),
    ]
    last = ecrire_donnees(ws, 5, sources, align=AL)
    convertir_dates(ws, 5, last, date_cols=[6, 7])
    # Statut formule (STALE si NEXT_REVIEW depasse)
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=10, value=f'=IF(OR(G{r}="",NOT(ISNUMBER(G{r}))),"-",IF(G{r}<TODAY(),"STALE","OK"))')
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = AC
        c.border = BF

    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"STALE"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ajouter_table(ws, "tblDataProducts", f"A4:J{last}")


def onglet_review_gates(wb):
    ws = wb.create_sheet("5_REVIEW_GATES")
    titre(ws, "REVIEW GATES", "Regles d'approbation globales — blocking, SLA, escalation", 8)
    en_tetes(ws, 4, ["GATE_ID", "GATE_TYPE", "BLOCKING", "APPROVER_ROLE", "SLA_HOURS", "ESCALATE_TO", "CRISIS_FAST_TRACK_H", "NOTE"],
             [12, 14, 12, 24, 12, 22, 20, 50])
    gates = [
        ("G-001", "BRAND",    "YES", "Brand Director",        24, "CMO",           4,  "Sortie externe standard"),
        ("G-002", "CLAIM",    "YES", "ESG Lead + Legal",      48, "Legal + CMO",   8,  "Claim ESG / absolu / origine"),
        ("G-003", "HERITAGE", "YES", "Heritage Director",     72, "CEO",           24, "Reference patrimoine non sourcee"),
        ("G-004", "LEGAL",    "YES", "Legal",                 24, "CEO",           4,  "Contenu juridique / revendication regulee"),
        ("G-005", "EVENT",    "NO",  "PR Director",           72, "CMO",           12, "Pack evenementiel sans claim"),
        ("G-006", "CRISIS",   "YES", "CMO",                    2, "CEO",           2,  "Crise — decision < 2h imperatif"),
    ]
    last = ecrire_donnees(ws, 5, gates, align=AL)
    ajouter_table(ws, "tblReviewGates", f"A4:H{last}")

    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_prompts(wb):
    ws = wb.create_sheet("6_PROMPTS_MODELS")
    titre(ws, "PROMPTS & MODELS", "Inventaire des prompts + modele + parametres — pour versioning", 9)
    en_tetes(ws, 4, ["PROMPT_ID", "AGENT_ID", "PURPOSE", "VERSION", "MODEL_PROFILE", "TEMP", "MAX_TOKENS", "VALIDATED_BY", "LAST_UPDATE"],
             [12, 10, 36, 10, 22, 10, 12, 20, 14])
    prompts = [
        ("PR-001", "AG-001", "Score texte vs charte + vocab",      "v1.0.0", "claude-opus-4-7",   0.2, 2000, "Brand Director", today_iso()),
        ("PR-002", "AG-001", "Hard-fail detection",                "v1.0.0", "claude-haiku-4-5",  0.0, 800,  "Brand Director", today_iso()),
        ("PR-003", "AG-002", "Redaction communique lifestyle",     "v1.0.0", "claude-opus-4-7",   0.4, 4000, "PR Director",    today_iso()),
        ("PR-004", "AG-002", "Redaction communique business",      "v1.0.0", "claude-opus-4-7",   0.3, 4000, "PR Director",    today_iso()),
        ("PR-005", "AG-002", "Adaptation media matrix",            "v1.0.0", "claude-sonnet-4-6", 0.4, 3000, "PR Director",    today_iso()),
        ("PR-006", "AG-003", "Pack evenementiel multi-format",     "v1.0.0", "claude-opus-4-7",   0.5, 4000, "PR Director",    today_iso()),
        ("PR-007", "AG-003", "Script social live",                 "v1.0.0", "claude-sonnet-4-6", 0.6, 2000, "PR Director",    today_iso()),
        ("PR-008", "AG-004", "Bloc narratif patrimoine",           "v1.0.0", "claude-opus-4-7",   0.3, 3000, "Heritage Dir",   today_iso()),
        ("PR-009", "AG-004", "Verification citation",              "v1.0.0", "claude-haiku-4-5",  0.0, 1000, "Heritage Dir",   today_iso()),
        ("PR-010", "AG-005", "Extraction claims",                  "v1.0.0", "claude-haiku-4-5",  0.0, 1000, "ESG Lead",       today_iso()),
        ("PR-011", "AG-005", "Matching evidence",                  "v1.0.0", "claude-sonnet-4-6", 0.1, 2000, "ESG Lead",       today_iso()),
        ("PR-012", "AG-005", "Scoring risque juridiction",         "v1.0.0", "claude-opus-4-7",   0.1, 1500, "Legal + ESG",    today_iso()),
    ]
    last = ecrire_donnees(ws, 5, prompts, fmt_cols={6: "0.00", 7: "0"}, align=AL)
    ajouter_table(ws, "tblPromptsModels", f"A4:I{last}")


def onglet_release_plan(wb):
    ws = wb.create_sheet("7_RELEASE_PLAN")
    titre(ws, "RELEASE PLAN", "Sprints, jalons, dependances, readiness", 8)
    en_tetes(ws, 4, ["SPRINT", "PHASE", "LIVRABLE", "OWNER", "DEBUT", "FIN", "DEPENDS_ON", "STATUT"],
             [12, 28, 36, 18, 12, 12, 24, 14])
    plan = [
        ("S1", "Phase 0 Standards",               "Guide standard + enums + naming", "AI Ops",          D("2026-04-15"), D("2026-04-22"), "-",                        "DONE"),
        ("S1", "Phase 1 Foundations + Master",    "FOUNDATIONS + MASTER shipped",    "AI Ops + Brand",  D("2026-04-22"), D("2026-05-01"), "Phase 0",                  "IN_PROGRESS"),
        ("S2", "Phase 2 AG-001 VoiceGuard",       "Moteur brand scoring",            "AI Ops + Brand",  D("2026-05-01"), D("2026-05-15"), "Phase 1",                  "PLANNED"),
        ("S3", "Phase 3 AG-002 + AG-005",         "Wedge flow complet demo",         "AI Ops + PR/ESG", D("2026-05-15"), D("2026-06-05"), "Phase 2",                  "PLANNED"),
        ("S4", "Phase 4 AG-004 Heritage",         "Sourcing discipline",             "AI Ops + Hrtg",   D("2026-06-05"), D("2026-06-19"), "Phase 1",                  "PLANNED"),
        ("S5", "Phase 5 AG-003 Events",           "Pack evenementiel",               "AI Ops + PR",     D("2026-06-19"), D("2026-07-10"), "Phase 2, Phase 4",         "PLANNED"),
        ("S6", "Phase 6 Consolidation",           "Imports logs / dashboards",       "AI Ops",          D("2026-07-10"), D("2026-07-24"), "Phases 2..5",              "PLANNED"),
        ("S7", "Phase 7 Implementation site",     "Portage apps/neural/site",        "AI Ops + Dev",    D("2026-07-24"), D("2026-08-21"), "Phase 6",                  "PLANNED"),
    ]
    last = ecrire_donnees(ws, 5, plan, align=AL)
    convertir_dates(ws, 5, last, date_cols=[5, 6])
    ajouter_table(ws, "tblReleasePlan", f"A4:H{last}")

    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"DONE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"IN_PROGRESS"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"BLOCKED"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_global_log(wb):
    ws = wb.create_sheet("8_GLOBAL_LOG_IMPORT")
    titre(ws, "IMPORT GLOBAL DES LOGS", "Consolidation des Outputs_Log des 5 agents — aliment via Power Query", 14)
    headers = ["RUN_ID", "AGENT_ID", "MAISON_ID", "DOC_ID", "TIMESTAMP", "LANG",
               "FINAL_STATUS", "RISK_LEVEL", "GATE_TRIGGERED", "SLA_MET",
               "TOTAL_MINUTES", "TOKENS_IN", "TOKENS_OUT", "COST_USD"]
    en_tetes(ws, 4, headers, [14, 10, 10, 14, 18, 8, 14, 14, 14, 10, 12, 12, 12, 12])

    # Jeu de demo pour montrer le pattern
    demo = [
        ("R-20260415-001", "AG-001", "M-001", "DOC-0001", DT("2026-04-15 09:12"), "FR", "APPROVED",  "LOW",      "BRAND",   "YES", 4.2,  1800, 900,  0.022),
        ("R-20260415-020", "AG-002", "M-001", "DOC-0001", DT("2026-04-15 09:18"), "FR", "APPROVED",  "LOW",      "BRAND",   "YES", 8.1,  3200, 2100, 0.064),
        ("R-20260415-011", "AG-005", "M-001", "DOC-0002", DT("2026-04-15 10:23"), "FR", "REJECTED",  "CRITICAL", "CLAIM",   "YES", 3.8,  1200, 600,  0.018),
        ("R-20260415-002", "AG-001", "M-001", "DOC-0002", DT("2026-04-15 10:25"), "FR", "REJECTED",  "HIGH",     "BRAND",   "YES", 2.9,  1400, 400,  0.012),
        ("R-20260420-001", "AG-003", "M-001", "DOC-EV-001", DT("2026-04-20 11:05"), "FR", "APPROVED",  "LOW",      "EVENT",   "YES", 12.4, 4100, 3200, 0.098),
        ("R-20260415-040", "AG-004", "M-001", "DOC-HR-001", DT("2026-04-15 13:40"), "FR", "APPROVED",  "LOW",      "HERITAGE","YES", 6.7,  2200, 1500, 0.047),
        ("R-20260416-004", "AG-005", "M-001", "DOC-0012", DT("2026-04-16 11:17"), "EN", "REJECTED",  "CRITICAL", "CLAIM",   "NO",  15.3, 2800, 900,  0.042),
        ("R-20260415-026", "AG-002", "M-001", "DOC-0007", DT("2026-04-15 17:00"), "FR", "IN_REVIEW", "LOW",      "BRAND",   "YES", 9.0,  3100, 2300, 0.067),
    ]
    last = ecrire_donnees(ws, 5, demo, fmt_cols={11: "0.00", 12: "0", 13: "0", 14: "0.000"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5])
    ajouter_table(ws, "tblGlobalLogs", f"A4:N{last}")

    # CF
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"APPROVED"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"REWORK"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"REJECTED"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_global_kpis(wb):
    ws = wb.create_sheet("9_GLOBAL_KPIS")
    titre(ws, "KPIs GLOBAUX", "Calcules depuis tblGlobalLogs (8_GLOBAL_LOG_IMPORT)", 4)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 44

    section(ws, 4, 4, "KPIs operationnels", NOIR)
    kpis = [
        ("Runs total (unique)",                         '=IFERROR(ROWS(UNIQUE(tblGlobalLogs[RUN_ID])),0)'),
        ("Approval rate",                               '=IFERROR(COUNTIFS(tblGlobalLogs[FINAL_STATUS],"APPROVED")/COUNTA(tblGlobalLogs[RUN_ID]),0)'),
        ("Rework rate",                                 '=IFERROR(COUNTIFS(tblGlobalLogs[FINAL_STATUS],"REWORK")/COUNTA(tblGlobalLogs[RUN_ID]),0)'),
        ("Rejection rate",                              '=IFERROR(COUNTIFS(tblGlobalLogs[FINAL_STATUS],"REJECTED")/COUNTA(tblGlobalLogs[RUN_ID]),0)'),
        ("Temps moyen / run (min)",                     '=IFERROR(AVERAGE(tblGlobalLogs[TOTAL_MINUTES]),0)'),
        ("Runs risque HIGH / CRITICAL",                 '=COUNTIFS(tblGlobalLogs[RISK_LEVEL],"HIGH")+COUNTIFS(tblGlobalLogs[RISK_LEVEL],"CRITICAL")'),
        ("SLA respecte",                                '=IFERROR(COUNTIFS(tblGlobalLogs[SLA_MET],"YES")/COUNTA(tblGlobalLogs[RUN_ID]),0)'),
        ("Cout cumule USD",                             '=IFERROR(SUM(tblGlobalLogs[COST_USD]),0)'),
        ("Cout moyen / run USD",                        '=IFERROR(AVERAGE(tblGlobalLogs[COST_USD]),0)'),
        ("Tokens in cumule",                            '=IFERROR(SUM(tblGlobalLogs[TOKENS_IN]),0)'),
        ("Tokens out cumule",                           '=IFERROR(SUM(tblGlobalLogs[TOKENS_OUT]),0)'),
    ]
    formats = ["0", "0.0%", "0.0%", "0.0%", "0.00", "0", "0.0%", "0.000", "0.000", "0", "0"]
    for i, (lbl, formula) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = formats[i]

    section(ws, 18, 4, "KPIs par agent", BORDEAUX)
    en_tetes(ws, 19, ["AGENT_ID", "Runs", "Approved%", "Cout USD"], [12, 10, 14, 14])
    for i, ag in enumerate(["AG-001", "AG-002", "AG-003", "AG-004", "AG-005"]):
        r = 20 + i
        ws.cell(row=r, column=1, value=ag).border = BF
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=r, column=1).alignment = AC
        c = ws.cell(row=r, column=2, value=f'=COUNTIFS(tblGlobalLogs[AGENT_ID],"{ag}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"
        c = ws.cell(row=r, column=3, value=f'=IFERROR(COUNTIFS(tblGlobalLogs[AGENT_ID],"{ag}",tblGlobalLogs[FINAL_STATUS],"APPROVED")/COUNTIFS(tblGlobalLogs[AGENT_ID],"{ag}"),0)')
        c.border = BF; c.alignment = AC; c.number_format = "0.0%"
        c = ws.cell(row=r, column=4, value=f'=IFERROR(SUMIFS(tblGlobalLogs[COST_USD],tblGlobalLogs[AGENT_ID],"{ag}"),0)')
        c.border = BF; c.alignment = AC; c.number_format = "0.000"


def onglet_dashboard(wb):
    ws = wb.create_sheet("10_DASHBOARD")
    titre(ws, "DASHBOARD EXECUTIVE", "Vue d'ensemble — agents, gates, risques, cadence", 8)
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 18

    section(ws, 4, 8, "Synthese", NOIR)
    tiles = [
        ("Agents actifs",   '=COUNTIFS(tblAgentRegistry[STATUS],"ACTIVE")',     "0",      FOND_VERT),
        ("Runs total",      '=COUNTA(tblGlobalLogs[RUN_ID])',                    "0",      FOND_BLEU),
        ("Approved %",      '=IFERROR(COUNTIFS(tblGlobalLogs[FINAL_STATUS],"APPROVED")/COUNTA(tblGlobalLogs[RUN_ID]),0)', "0.0%", FOND_VERT),
        ("Risque HIGH+",    '=COUNTIFS(tblGlobalLogs[RISK_LEVEL],"HIGH")+COUNTIFS(tblGlobalLogs[RISK_LEVEL],"CRITICAL")',  "0",   FOND_JAUNE),
        ("SLA respecte",    '=IFERROR(COUNTIFS(tblGlobalLogs[SLA_MET],"YES")/COUNTA(tblGlobalLogs[RUN_ID]),0)',             "0.0%", FOND_VERT),
        ("Cout cumule $",   '=IFERROR(SUM(tblGlobalLogs[COST_USD]),0)',         "0.000",  FOND_VIOLET),
    ]
    for i, (lbl, formula, fmt, fill) in enumerate(tiles):
        col = 2 + i
        c = ws.cell(row=5, column=col, value=lbl)
        c.font = Font(name="Calibri", size=10, bold=True, color=NOIR)
        c.alignment = AC
        c.border = BF
        c.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")
        v = ws.cell(row=6, column=col, value=formula)
        v.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        v.alignment = AC
        v.border = BF_EPAIS
        v.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        v.number_format = fmt
        ws.row_dimensions[6].height = 44

    # Chart : runs par agent
    section(ws, 8, 8, "Runs par agent", BORDEAUX)
    ws.cell(row=9, column=2, value="Agent").font = Font(bold=True)
    ws.cell(row=9, column=3, value="Runs").font = Font(bold=True)
    for i, ag in enumerate(["AG-001", "AG-002", "AG-003", "AG-004", "AG-005"]):
        r = 10 + i
        ws.cell(row=r, column=2, value=ag).border = BF
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(tblGlobalLogs[AGENT_ID],"{ag}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch = BarChart()
    ch.type = "bar"
    ch.style = 11
    data = Reference(ws, min_col=3, min_row=9, max_row=14)
    labels = Reference(ws, min_col=2, min_row=10, max_row=14)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(labels)
    ch.title = "Runs par agent"
    ch.height = 9
    ch.width = 16
    ws.add_chart(ch, "E8")

    # Chart : statuts finaux
    section(ws, 22, 8, "Distribution des statuts finaux", BORDEAUX)
    ws.cell(row=23, column=2, value="Status").font = Font(bold=True)
    ws.cell(row=23, column=3, value="N").font = Font(bold=True)
    for i, st in enumerate(["APPROVED", "IN_REVIEW", "REWORK", "REJECTED"]):
        r = 24 + i
        ws.cell(row=r, column=2, value=st).border = BF
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(tblGlobalLogs[FINAL_STATUS],"{st}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch2 = PieChart()
    data2 = Reference(ws, min_col=3, min_row=23, max_row=27)
    labels2 = Reference(ws, min_col=2, min_row=24, max_row=27)
    ch2.add_data(data2, titles_from_data=True)
    ch2.set_categories(labels2)
    ch2.title = "Statuts finaux"
    ch2.height = 9
    ch2.width = 16
    ws.add_chart(ch2, "E22")


def onglet_risk_register(wb):
    ws = wb.create_sheet("11_RISK_REGISTER")
    titre(ws, "RISK REGISTER", "Risques ouverts, hypotheses, plans de mitigation", 10)
    en_tetes(ws, 4, ["RISK_ID", "DOMAINE", "DESCRIPTION", "LIKELIHOOD", "IMPACT", "SCORE", "OWNER", "MITIGATION", "STATUT", "NEXT_REVIEW"],
             [10, 16, 60, 12, 12, 10, 20, 46, 14, 14])
    risks = [
        ("RK-001", "Brand",      "VoiceGuard calibration trop stricte -> faux positifs et ralentissement ops", "MEDIUM", "HIGH", "", "Brand Dir",    "Testset + sensibility analysis + seuil ajustable par maison", "OPEN",    D("2026-05-15")),
        ("RK-002", "Claim",      "Expiration evidence non surveillee -> claim diffuse alors que preuve perimee","HIGH",   "CRITICAL","","ESG Lead",    "Alertes J-30 J-7 (CF pose) + statut STALE formule auto",         "MITIGATED",D("2026-04-30")),
        ("RK-003", "Heritage",   "Citation source tertiaire utilisee sans revalidation",                       "MEDIUM", "HIGH", "", "Heritage Dir", "Blocage hard si source = TERTIARY sans validation",             "OPEN",    D("2026-05-10")),
        ("RK-004", "Legal",      "Juridictions multiples (EU, US, UK, CH) -> divergence regle",                "HIGH",   "HIGH", "", "Legal",        "Extension matrice juridiction AG-005 en v2",                   "OPEN",    D("2026-06-30")),
        ("RK-005", "Ops cout",   "Derive cout API sans alerte budget",                                          "MEDIUM", "MEDIUM","", "AI Ops",      "Budget mensuel + alerte 80% + kill switch 120%",                 "OPEN",    D("2026-05-01")),
        ("RK-006", "Crisis",     "Fast-track bypass sans revue = risque reputationnel",                         "LOW",    "CRITICAL","","CMO",         "Log tous les fast-track + audit post-crise obligatoire",        "OPEN",    D("2026-05-20")),
        ("RK-007", "Data",       "Vocab FR non maintenu -> charte drift",                                        "MEDIUM", "MEDIUM","", "Brand Dir",    "Revue mensuelle + changelog obligatoire",                        "OPEN",    D("2026-05-01")),
        ("RK-008", "Heritage",   "Droits image non verifies sur visuels comms",                                 "MEDIUM", "HIGH", "", "Legal",        "Registre droits images a creer en v2",                          "OPEN",    D("2026-06-30")),
        ("RK-009", "Localisation","Pas de vocab IT/DE/JA/ZH -> international comms restent FR/EN-only",         "HIGH",   "MEDIUM","", "Brand Dir",    "Seed vocab 4 langues supp en v2",                                "OPEN",    D("2026-07-31")),
        ("RK-010", "Learnings",  "Boucle d'apprentissage partielle (AG-001 seulement)",                          "MEDIUM", "MEDIUM","", "AI Ops",      "LEARNINGS pose sur AG-002/003/004/005 dans le Sprint 3",        "MITIGATED",D("2026-06-15")),
    ]
    last = ecrire_donnees(ws, 5, risks, align=AL)
    convertir_dates(ws, 5, last, date_cols=[10])
    # Score = Likelihood*Impact mapping
    # Likelihood LOW=1 MEDIUM=2 HIGH=3
    # Impact LOW=1 MEDIUM=2 HIGH=3 CRITICAL=4
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=6, value=f'=IFS(D{r}="LOW",1,D{r}="MEDIUM",2,D{r}="HIGH",3,TRUE,0) * IFS(E{r}="LOW",1,E{r}="MEDIUM",2,E{r}="HIGH",3,E{r}="CRITICAL",4,TRUE,0)')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = "0"

    ajouter_table(ws, "tblRiskRegister", f"A4:J{last}")

    # CF sur score
    ws.conditional_formatting.add(f"F5:F{last}",
                                   FormulaRule(formula=[f"F5>=9"], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}",
                                   FormulaRule(formula=[f"AND(F5>=4,F5<9)"], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_lists(wb):
    ws = wb.create_sheet("12_LISTS")
    titre(ws, "LISTES DE VALIDATION", "Enums globaux miroir pour ce workbook", 5)
    en_tetes(ws, 4, ["STATUS", "RISK_LEVEL", "GATE_TYPE", "LANG", "DECISION"], [14, 14, 14, 10, 14])
    rows = max(len(STATUS), 4, 5, 6, 4)
    data = []
    risk = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
    gate = ["BRAND", "CLAIM", "HERITAGE", "LEGAL", "EVENT"]
    lang = ["FR", "EN", "IT", "DE", "JA", "ZH"]
    decision = ["APPROVE", "REWORK", "REJECT", "PENDING"]
    for i in range(rows):
        data.append([
            STATUS[i] if i < len(STATUS) else "",
            risk[i] if i < len(risk) else "",
            gate[i] if i < len(gate) else "",
            lang[i] if i < len(lang) else "",
            decision[i] if i < len(decision) else "",
        ])
    ecrire_donnees(ws, 5, data)


def build():
    wb = Workbook()
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_scope_raci(wb)
    onglet_agent_registry(wb)
    onglet_workflow(wb)
    onglet_data_products(wb)
    onglet_review_gates(wb)
    onglet_prompts(wb)
    onglet_release_plan(wb)
    onglet_global_log(wb)
    onglet_global_kpis(wb)
    onglet_dashboard(wb)
    onglet_risk_register(wb)
    onglet_lists(wb)

    order = ["0_README", "1_PARAMS", "1_SCOPE_RACI", "2_AGENT_REGISTRY",
             "3_WORKFLOW_MAP", "4_DATA_PRODUCTS", "5_REVIEW_GATES", "6_PROMPTS_MODELS",
             "7_RELEASE_PLAN", "8_GLOBAL_LOG_IMPORT", "9_GLOBAL_KPIS", "10_DASHBOARD",
             "11_RISK_REGISTER", "12_LISTS"]
    wb._sheets = [wb[n] for n in order]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
