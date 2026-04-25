"""
Build Assurance_Marketing_OVERVIEW_NEURAL.xlsx — synthèse cross-agents
de la branche NEURAL Assurances/Marketing.

7 onglets : README, Agents Synthèse, Pipeline Orchestration,
Réglementation Cross, KPIs Consolidés, Roadmap, Pricing & ROI.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Couleurs par agent
AGENT_COLORS = {
    "InsurSimplifier": "0F766E",      # vert teal foncé
    "DDA_MarketingGuard": "B91C1C",   # rouge foncé
    "MultiChannelInsur": "1E40AF",    # bleu marine
    "PreventionContent": "7C3AED",    # violet NEURAL
}


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=NEURAL_VIOLET, end_color=NEURAL_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def agent_fill(agent: str) -> PatternFill:
    c = AGENT_COLORS.get(agent, NEURAL_VIOLET)
    return PatternFill("solid", start_color=c, end_color=c)


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 70) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict[int, float]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ────────────────────────────────────────────────────────────────────
# Données : 4 agents
# ────────────────────────────────────────────────────────────────────
AGENTS = [
    {
        "name": "InsurSimplifier",
        "ordre": 1,
        "role": "Simplification rédactionnelle des clauses d'assurance",
        "input": "Texte brut clause CGU/CGV/IPID/notice",
        "output": "Version vulgarisée niveau B1-B2 + score Flesch + glossaire termes techniques",
        "kpi_principal": "Score Flesch FR (Kandel-Moles) — cible > 50 (lisible)",
        "kpi_secondaire": "% termes techniques avec définition glossaire — cible > 90%",
        "trigger": "Manuel ou batch sur catalogue produits",
        "duree_traitement": "30s par clause — 5 min par contrat complet",
        "confidence_v1": 92,
        "humans_required": "Validation juridique avant publication contractuelle",
        "regle_cle": "Acte délégué clarté contractuelle 2025 — UE 2024/879",
    },
    {
        "name": "DDA_MarketingGuard",
        "ordre": 2,
        "role": "Audit conformité DDA des communications marketing avant diffusion",
        "input": "Communication marketing (email, post, brochure, mailing, comparateur)",
        "output": "Verdict OK/WARN/KO sur 12 points DDA + redlines + score conformité + décision diffusion",
        "kpi_principal": "Score conformité (% OK + 0.5×WARN) / 12 — cible > 80% pour DIFFUSION OK",
        "kpi_secondaire": "Nb KO bloquants — cible 0 par communication",
        "trigger": "Pré-publication automatique via webhook CMS / DAM",
        "duree_traitement": "10s par communication — 1 min par campagne complète",
        "confidence_v1": 88,
        "humans_required": "Compliance officer pour validation finale + ACPR si campagne nationale",
        "regle_cle": "DDA UE 2016/97 + Code des assurances L.521 + ACPR Reco 2024-R-01",
    },
    {
        "name": "MultiChannelInsur",
        "ordre": 3,
        "role": "Adaptation discours assurance par canal de distribution",
        "input": "Brief commercial source + canal cible (agent / courtier / direct / comparateur)",
        "output": "Variante optimisée canal + mentions légales adaptées + score adaptation",
        "kpi_principal": "Score adaptation canal (registre + mentions) — cible > 90 par variante",
        "kpi_secondaire": "Score conformité DDA + DSA/DMA (comparateur uniquement) — cible > 85",
        "trigger": "Workflow marketing — input brief, génération 4 variantes simultanées",
        "duree_traitement": "45s par brief × 4 canaux — 3 min par campagne 3 produits",
        "confidence_v1": 90,
        "humans_required": "Validation directrice marketing + DPO si comparateur (DSA)",
        "regle_cle": "DDA + DSA art. 22-26 + DMA art. 6-7 + AI Act art. 50",
    },
    {
        "name": "PreventionContent",
        "ordre": 4,
        "role": "Audit ClaimGuard + génération contenu prévention conforme RGPD",
        "input": "Contenu prévention brut (route / santé / habitation) + métadonnées collecte données",
        "output": "Findings ClaimGuard + audit RGPD art. 9 + version optimisée + registre traitements",
        "kpi_principal": "Score ClaimGuard avant/après — gain moyen cible > +50 pts",
        "kpi_secondaire": "Score RGPD avant/après — gain moyen cible > +50 pts",
        "trigger": "Pré-diffusion via CMS prévention + revue annuelle automatique",
        "duree_traitement": "1 min par contenu — 8 min par campagne complète multi-domaines",
        "confidence_v1": 86,
        "humans_required": "DPO + DPIA pour traitements art. 9 + juriste si refonte produit",
        "regle_cle": "RGPD art. 9 + art. 22 + Code des assurances L.113-8/9 + Code conso L.121-1",
    },
]

# Pipeline orchestration : enchaînements possibles
PIPELINE_FLOWS = [
    {
        "scenario": "Lancement nouveau produit MRH",
        "etape": 1,
        "agent": "InsurSimplifier",
        "action": "Vulgarisation des CGV + IPID du nouveau produit MRH Premium",
        "input_source": "Service produit (juriste + actuaire)",
        "output_dest": "Service marketing + service formation réseau",
    },
    {
        "scenario": "Lancement nouveau produit MRH",
        "etape": 2,
        "agent": "MultiChannelInsur",
        "action": "Génération de 4 variantes communication (agent / courtier / direct / comparateur) à partir du brief produit simplifié",
        "input_source": "Sortie InsurSimplifier + brief commercial",
        "output_dest": "Service marketing par canal",
    },
    {
        "scenario": "Lancement nouveau produit MRH",
        "etape": 3,
        "agent": "DDA_MarketingGuard",
        "action": "Audit conformité DDA des 4 variantes avant diffusion",
        "input_source": "Sortie MultiChannelInsur",
        "output_dest": "Service compliance + retour aux marketeurs si KO",
    },
    {
        "scenario": "Lancement nouveau produit MRH",
        "etape": 4,
        "agent": "PreventionContent",
        "action": "Génération du guide prévention habitation associé au lancement (ClaimGuard + RGPD)",
        "input_source": "Brief prévention + métadonnées collecte données",
        "output_dest": "Newsletter prévention + espace client",
    },
    {
        "scenario": "Refonte campagne santé senior",
        "etape": 1,
        "agent": "PreventionContent",
        "action": "Audit RGPD du questionnaire santé existant + détection findings claim avoidance",
        "input_source": "Contenus en production",
        "output_dest": "DPO + équipe marketing santé",
    },
    {
        "scenario": "Refonte campagne santé senior",
        "etape": 2,
        "agent": "InsurSimplifier",
        "action": "Vulgarisation des clauses contractuelles santé liées aux dépistages",
        "input_source": "CGV produit santé senior",
        "output_dest": "Équipe marketing santé",
    },
    {
        "scenario": "Refonte campagne santé senior",
        "etape": 3,
        "agent": "MultiChannelInsur",
        "action": "Variantes par canal — agent général + direct (pas de courtier ni comparateur sur santé senior)",
        "input_source": "Sortie InsurSimplifier + brief refonte",
        "output_dest": "Marketing canal",
    },
    {
        "scenario": "Refonte campagne santé senior",
        "etape": 4,
        "agent": "DDA_MarketingGuard",
        "action": "Audit final avant diffusion campagne",
        "input_source": "Sorties MultiChannelInsur + PreventionContent",
        "output_dest": "Direction marketing — Go/No-Go diffusion",
    },
]

# Régulations cross-agents
REGULATIONS = [
    {
        "regulation": "DDA — Directive Distribution Assurance",
        "ref": "UE 2016/97 + Code des assurances L.521",
        "InsurSimplifier": "Indirect — clarté CGV/IPID",
        "DDA_MarketingGuard": "Cœur — 12 points checklist",
        "MultiChannelInsur": "Cœur — adaptation par canal",
        "PreventionContent": "Indirect — pas de claim avoidance",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "Acte délégué clarté contractuelle",
        "ref": "UE 2024/879 (applicable 2025)",
        "InsurSimplifier": "Cœur — score Flesch + vulgarisation",
        "DDA_MarketingGuard": "Indirect — vérifie clarté comm.",
        "MultiChannelInsur": "Indirect — clarté par canal",
        "PreventionContent": "Faible",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "DSA — Digital Services Act",
        "ref": "UE 2022/2065 art. 22-26",
        "InsurSimplifier": "Non applicable",
        "DDA_MarketingGuard": "Cas comparateur uniquement",
        "MultiChannelInsur": "Cœur sur canal comparateur",
        "PreventionContent": "Indirect — publicité prévention",
        "criticite": "BLOQUANT (comparateurs)",
    },
    {
        "regulation": "DMA — Digital Markets Act",
        "ref": "UE 2022/1925 art. 6-7",
        "InsurSimplifier": "Non applicable",
        "DDA_MarketingGuard": "Indirect comparateur",
        "MultiChannelInsur": "Cœur si gatekeeper",
        "PreventionContent": "Non applicable",
        "criticite": "Conditionnel",
    },
    {
        "regulation": "AI Act — Règlement IA UE",
        "ref": "UE 2024/1689 art. 50 (août 2026)",
        "InsurSimplifier": "Disclosure si IA utilisée publiquement",
        "DDA_MarketingGuard": "Disclosure scoring IA",
        "MultiChannelInsur": "Disclosure si comparateur scoring",
        "PreventionContent": "Disclosure si scoring prévention auto",
        "criticite": "BLOQUANT (août 2026)",
    },
    {
        "regulation": "RGPD",
        "ref": "UE 2016/679 art. 6, 9, 22, 30, 35",
        "InsurSimplifier": "Faible — pas de DCP traitées",
        "DDA_MarketingGuard": "Faible — analyse comm. publiques",
        "MultiChannelInsur": "Modéré — segments cible",
        "PreventionContent": "Cœur — données santé art. 9",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "ACPR — Recommandation publicité",
        "ref": "ACPR Reco 2024-R-01",
        "InsurSimplifier": "Indirect — clarté",
        "DDA_MarketingGuard": "Cœur",
        "MultiChannelInsur": "Cœur",
        "PreventionContent": "Indirect — pratiques commerciales",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "Code de la consommation",
        "ref": "L.121-1 (pratique trompeuse)",
        "InsurSimplifier": "Indirect — pas de tromperie clarté",
        "DDA_MarketingGuard": "Cœur — détection trompeur",
        "MultiChannelInsur": "Cœur — par canal",
        "PreventionContent": "Cœur — claim avoidance",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "Loi Lemoine (résiliation infra-annuelle)",
        "ref": "Loi 2022-270",
        "InsurSimplifier": "À mentionner dans CGV santé",
        "DDA_MarketingGuard": "Vérifie mention dans comm. santé",
        "MultiChannelInsur": "Mention par canal",
        "PreventionContent": "Indirect",
        "criticite": "Important (santé)",
    },
]

# KPIs consolidés
KPIS_CONSOLIDES = [
    {"agent": "InsurSimplifier", "kpi": "Score Flesch FR", "v1": 60, "cible_v2": 65, "unite": "points"},
    {"agent": "InsurSimplifier", "kpi": "% glossaire", "v1": 92, "cible_v2": 96, "unite": "%"},
    {"agent": "InsurSimplifier", "kpi": "Réduction longueur clause", "v1": 35, "cible_v2": 45, "unite": "%"},
    {"agent": "DDA_MarketingGuard", "kpi": "Score conformité moyen", "v1": 78, "cible_v2": 88, "unite": "%"},
    {"agent": "DDA_MarketingGuard", "kpi": "% comm. DIFFUSION OK 1er passage", "v1": 42, "cible_v2": 65, "unite": "%"},
    {"agent": "DDA_MarketingGuard", "kpi": "Faux positifs verdict KO", "v1": 8, "cible_v2": 4, "unite": "%"},
    {"agent": "MultiChannelInsur", "kpi": "Score adaptation canal moyen", "v1": 91, "cible_v2": 94, "unite": "points"},
    {"agent": "MultiChannelInsur", "kpi": "Score conformité DSA comparateur", "v1": 97, "cible_v2": 99, "unite": "points"},
    {"agent": "MultiChannelInsur", "kpi": "% variantes PRÊT (8/12)", "v1": 67, "cible_v2": 85, "unite": "%"},
    {"agent": "PreventionContent", "kpi": "Gain ClaimGuard moyen", "v1": 58, "cible_v2": 65, "unite": "+ pts"},
    {"agent": "PreventionContent", "kpi": "Gain RGPD moyen", "v1": 54, "cible_v2": 65, "unite": "+ pts"},
    {"agent": "PreventionContent", "kpi": "% findings KO détectés", "v1": 95, "cible_v2": 98, "unite": "%"},
]

# Roadmap déploiement
ROADMAP = [
    {"phase": "Phase 0 — POC", "duree": "Sprint 1-2 (4 semaines)", "agents": "InsurSimplifier", "livrable": "POC sur 5 produits assurance, démo prospect", "statut": "Démo Excel disponible (livrée 04/2026)"},
    {"phase": "Phase 1 — Pilote", "duree": "Sprint 3-6 (8 semaines)", "agents": "InsurSimplifier + DDA_MarketingGuard", "livrable": "Pilote chez 1 client assureur — 50 communications auditées", "statut": "Démo Excel disponible (livrée 04/2026)"},
    {"phase": "Phase 2 — MVP", "duree": "Sprint 7-12 (12 semaines)", "agents": "+ MultiChannelInsur", "livrable": "Déploiement sur 3 canaux + intégration CMS client", "statut": "Démo Excel disponible (livrée 04/2026)"},
    {"phase": "Phase 3 — V1", "duree": "Sprint 13-20 (16 semaines)", "agents": "+ PreventionContent + RGPD audit", "livrable": "Suite complète 4 agents + DPIA + registre traitements", "statut": "Démo Excel disponible (livrée 04/2026)"},
    {"phase": "Phase 4 — Scale", "duree": "Sprint 21-28 (16 semaines)", "agents": "Tous + extension multi-langues", "livrable": "Multi-tenant, FR + EN + DE, intégration ERP/CRM", "statut": "Roadmap"},
    {"phase": "Phase 5 — Self-service", "duree": "Sprint 29-36 (16 semaines)", "agents": "Tous + portail self-service", "livrable": "Plateforme SaaS — assureurs/courtiers en autonomie", "statut": "Vision"},
]

# Pricing
PRICING = [
    {"forfait": "Découverte", "agents": "InsurSimplifier seul", "volume": "100 clauses / mois", "prix_mois": 1200, "prix_an": 13200, "support": "Email — 48h"},
    {"forfait": "Marketing Compliance", "agents": "DDA_MarketingGuard + MultiChannelInsur", "volume": "300 communications / mois", "prix_mois": 3500, "prix_an": 38500, "support": "Email + chat — 24h"},
    {"forfait": "Suite complète", "agents": "Les 4 agents", "volume": "1000 actes / mois (clauses + comm + prévention)", "prix_mois": 6800, "prix_an": 74800, "support": "Email + chat + visio — 4h ouvrées"},
    {"forfait": "Enterprise", "agents": "Suite + intégration sur-mesure", "volume": "Volume illimité + SLA 99.5%", "prix_mois": 14500, "prix_an": 159500, "support": "Account manager dédié + DPO partenaire"},
    {"forfait": "On-premise", "agents": "Suite déployée chez le client", "volume": "Selon infra client", "prix_mois": 22000, "prix_an": 242000, "support": "Account manager + ingénieur dédié"},
]

# ROI estimé
ROI = [
    {"metric": "Coût annuel équivalent compliance officer (FTE)", "valeur": 95000, "unite": "€/an", "source": "INSEE 2024 — cadre conformité finance"},
    {"metric": "Coût annuel rédacteur juridique senior (FTE)", "valeur": 78000, "unite": "€/an", "source": "INSEE 2024 — juriste assurance"},
    {"metric": "Coût annuel agence comm. spécialisée assurance", "valeur": 280000, "unite": "€/an", "source": "Estimation marché 2024"},
    {"metric": "Heures économisées par mois (forfait Suite complète)", "valeur": 320, "unite": "h/mois", "source": "Estimation NEURAL — 1000 actes × 20 min en moyenne"},
    {"metric": "Coût horaire interne moyen (chargé)", "valeur": 85, "unite": "€/h", "source": "Estimation NEURAL"},
    {"metric": "Économie mensuelle estimée (forfait Suite)", "valeur": 27200, "unite": "€/mois", "source": "320h × 85€/h"},
    {"metric": "Prix forfait Suite complète", "valeur": 6800, "unite": "€/mois", "source": "Pricing NEURAL ci-dessus"},
    {"metric": "Marge brute mensuelle pour le client", "valeur": 20400, "unite": "€/mois", "source": "Économie - prix"},
    {"metric": "ROI annuel client (forfait Suite)", "valeur": 244800, "unite": "€/an", "source": "20400 × 12"},
    {"metric": "Taux d'amende ACPR évitée (publicité non conforme)", "valeur": 5, "unite": "% CA", "source": "DDA art. 33 — sanctions"},
]


# ────────────────────────────────────────────────────────────────────
# Onglet 0 — README
# ────────────────────────────────────────────────────────────────────
def build_readme(ws) -> None:
    write_title(ws, "NEURAL · Assurances/Marketing — OVERVIEW : Synthèse de la branche 4 agents")
    write_subtitle(ws, 2, "Vue consolidée des 4 agents NEURAL pour assureurs, courtiers, mutuelles et comparateurs")

    ws.row_dimensions[3].height = 10

    write_section_header(ws, 4, "VISION DE LA BRANCHE")
    vision = [
        ("Promesse", "Automatiser la conformité réglementaire et l'optimisation rédactionnelle des contenus assurance — du contrat au pitch comparateur, en passant par les guides de prévention."),
        ("Marché cible", "200+ assureurs en France (FFSA + Roam) | 22 000 courtiers (CSCA) | 8 comparateurs majeurs (LesFurets, Assurland, etc.) | 50+ mutuelles (Mutualité Française)"),
        ("Pain point #1", "Les services compliance et marketing assurance fonctionnent en silos. Les communications passent en compliance APRÈS rédaction, ce qui génère 30-40% de retours et allonge le time-to-market de 3 à 6 semaines."),
        ("Pain point #2", "L'arrivée de l'AI Act (août 2026) + Acte délégué clarté contractuelle (2025) + DSA pour comparateurs (2024) crée une charge réglementaire impossible à absorber par les équipes humaines existantes."),
        ("Différenciation", "NEURAL pré-calcule les verdicts compliance en amont (DDA, RGPD, DSA) — ce qui transforme la compliance d'un goulot d'étranglement en une garantie en temps réel."),
        ("Volumétrie cible", "À 12 mois : 5-8 clients pilotes (assureurs/courtiers) | À 24 mois : 25-35 clients | À 36 mois : 80-100 clients (incluant SaaS self-service)"),
    ]
    for i, (k, v) in enumerate(vision):
        row = 5 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font()
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 50

    ws.row_dimensions[11].height = 10

    write_section_header(ws, 12, "LES 4 AGENTS — BLOC SYNTHÈSE")
    agent_headers = ["Ordre", "Agent", "Rôle", "Réglementation cœur", "Confidence v1"]
    for col, h in enumerate(agent_headers, 1):
        cell = ws.cell(row=13, column=col, value=h)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[13].height = 28

    for i, a in enumerate(AGENTS):
        row = 14 + i
        ws.cell(row=row, column=1, value=a["ordre"])
        cell_agent = ws.cell(row=row, column=2, value=a["name"])
        cell_agent.fill = agent_fill(a["name"])
        cell_agent.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        ws.cell(row=row, column=3, value=a["role"])
        ws.cell(row=row, column=4, value=a["regle_cle"])
        ws.cell(row=row, column=5, value=f"{a['confidence_v1']}%")
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            cell.border = THIN_BORDER
            if col != 2:
                cell.font = body_font()
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5).font = body_font(bold=True, color=GREEN_OK)
        ws.row_dimensions[row].height = 40

    ws.row_dimensions[18].height = 10

    write_section_header(ws, 19, "MODE D'EMPLOI DES ONGLETS")
    tabs = [
        ("0_README", "Ce guide — vision, marché, synthèse 4 agents"),
        ("1_Agents_Synthese", "Fiche détaillée par agent : input/output/KPIs/déclencheurs/durée/validations humaines"),
        ("2_Pipeline_Orchestration", "Scénarios d'enchaînement réels : 'Lancement nouveau MRH' (4 étapes) + 'Refonte santé senior' (4 étapes)"),
        ("3_Reglementation_Cross", "Matrice 9 réglementations × 4 agents — qui est cœur, indirect, non applicable"),
        ("4_KPIs_Consolides", "12 KPIs cross-agents — valeur v1 + cible v2 + écart"),
        ("5_Roadmap", "6 phases de déploiement — POC → SaaS self-service (36 mois)"),
        ("6_Pricing_ROI", "5 forfaits + calcul ROI client : économie mensuelle vs prix mensuel"),
    ]
    for i, (tab, desc) in enumerate(tabs):
        row = 20 + i
        ws.cell(row=row, column=1, value=tab).font = body_font(bold=True, color=NEURAL_VIOLET)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row=row, column=2, value=desc).font = body_font()
        ws.row_dimensions[row].height = 22

    set_widths(ws, {1: 22, 2: 30, 3: 35, 4: 30, 5: 16, 6: 20, 7: 20, 8: 20})


# ────────────────────────────────────────────────────────────────────
# Onglet 1 — Agents Synthèse
# ────────────────────────────────────────────────────────────────────
def build_agents_synthese(ws) -> None:
    write_title(ws, "NEURAL · OVERVIEW — 1_Agents_Synthese : Fiche détaillée par agent")
    write_subtitle(ws, 2, "Input, output, KPIs, déclencheurs, validation humaine — pour chaque agent de la branche")

    headers = ["Agent", "Rôle", "Input", "Output", "KPI principal", "KPI secondaire",
               "Trigger", "Durée traitement", "Confidence v1", "Validation humaine requise"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, a in enumerate(AGENTS):
        row = 5 + i
        vals = [a["name"], a["role"], a["input"], a["output"],
                a["kpi_principal"], a["kpi_secondaire"], a["trigger"],
                a["duree_traitement"], f"{a['confidence_v1']}%",
                a["humans_required"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=120)

        # Agent badge
        agent_cell = ws.cell(row=row, column=1)
        agent_cell.fill = agent_fill(a["name"])
        agent_cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        agent_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Confidence
        conf_cell = ws.cell(row=row, column=9)
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")
        if a["confidence_v1"] >= 90:
            conf_cell.fill = green_fill()
            conf_cell.font = body_font(bold=True, color=GREEN_OK)
        elif a["confidence_v1"] >= 80:
            conf_cell.fill = amber_fill()
            conf_cell.font = body_font(bold=True, color=AMBER_WARN)

    set_widths(ws, {1: 20, 2: 30, 3: 26, 4: 32, 5: 28, 6: 28, 7: 24, 8: 22, 9: 14, 10: 30})


# ────────────────────────────────────────────────────────────────────
# Onglet 2 — Pipeline Orchestration
# ────────────────────────────────────────────────────────────────────
def build_pipeline(ws) -> None:
    write_title(ws, "NEURAL · OVERVIEW — 2_Pipeline_Orchestration : Scénarios d'enchaînement multi-agents")
    write_subtitle(ws, 2, "Comment les 4 agents s'orchestrent dans des cas réels — handoffs entre agents et services")

    headers = ["Scénario", "Étape", "Agent", "Action", "Source input", "Destinataire output"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    last_scenario = None
    for i, p in enumerate(PIPELINE_FLOWS):
        row = 5 + i
        is_new_scenario = p["scenario"] != last_scenario
        vals = [p["scenario"] if is_new_scenario else "",
                f"Étape {p['etape']}", p["agent"], p["action"],
                p["input_source"], p["output_dest"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=False, height=70)

        # Scénario header
        if is_new_scenario:
            scen_cell = ws.cell(row=row, column=1)
            scen_cell.font = body_font(bold=True, color=NEURAL_VIOLET)
            scen_cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left", indent=1)

        # Agent badge
        agent_cell = ws.cell(row=row, column=3)
        agent_cell.fill = agent_fill(p["agent"])
        agent_cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        agent_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Étape
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).font = body_font(bold=True)

        last_scenario = p["scenario"]

    ws.row_dimensions[13].height = 10

    write_section_header(ws, 14, "GAINS D'ORCHESTRATION VS USAGE INDIVIDUEL")
    gains = [
        ("Time-to-market lancement produit", "Avant orchestration", "6-9 semaines (allers-retours compliance + marketing)"),
        ("Time-to-market lancement produit", "Avec orchestration NEURAL", "2-3 semaines (pré-validation à chaque étape)"),
        ("Cohérence cross-canal", "Avant orchestration", "Inégale — chaque équipe canal travaille en silo"),
        ("Cohérence cross-canal", "Avec orchestration NEURAL", "Garantie — brief source unique, déclinaison contrôlée"),
        ("Traçabilité réglementaire", "Avant orchestration", "Manuelle — preuves dispersées dans plusieurs services"),
        ("Traçabilité réglementaire", "Avec orchestration NEURAL", "Automatisée — registre unique + journaux d'audit par agent"),
    ]
    for i, (k, mode, v) in enumerate(gains):
        row = 15 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.cell(row=row, column=2, value=mode).font = body_font(bold=True, color=AMBER_WARN if "Avant" in mode else GREEN_OK)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=3, value=v)
        cell.font = body_font()
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        if i % 2 == 1:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = alt_fill()

    set_widths(ws, {1: 30, 2: 12, 3: 24, 4: 38, 5: 30, 6: 30})


# ────────────────────────────────────────────────────────────────────
# Onglet 3 — Réglementation Cross
# ────────────────────────────────────────────────────────────────────
def build_reglementation_cross(ws) -> None:
    write_title(ws, "NEURAL · OVERVIEW — 3_Reglementation_Cross : Matrice 9 réglementations × 4 agents")
    write_subtitle(ws, 2, "Pour chaque réglementation : quel agent la traite (Cœur), la touche (Indirect), ne s'y applique pas")

    headers = ["Réglementation", "Référence légale",
               "InsurSimplifier", "DDA_MarketingGuard", "MultiChannelInsur", "PreventionContent",
               "Criticité"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    # Color agent column headers
    for col, agent in enumerate(["InsurSimplifier", "DDA_MarketingGuard", "MultiChannelInsur", "PreventionContent"], 3):
        cell = ws.cell(row=4, column=col)
        cell.fill = agent_fill(agent)

    for i, r in enumerate(REGULATIONS):
        row = 5 + i
        vals = [
            r["regulation"], r["ref"],
            r["InsurSimplifier"], r["DDA_MarketingGuard"],
            r["MultiChannelInsur"], r["PreventionContent"],
            r["criticite"],
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=50)

        # Color cells based on involvement
        for col in range(3, 7):
            cell = ws.cell(row=row, column=col)
            v = str(cell.value or "").lower()
            if "cœur" in v:
                cell.fill = PatternFill("solid", start_color="C7D2FE", end_color="C7D2FE")
                cell.font = body_font(bold=True, color="3730A3")
            elif "indirect" in v or "modéré" in v:
                cell.fill = amber_fill()
                cell.font = body_font(color=AMBER_WARN)
            elif "non applicable" in v or "faible" in v:
                cell.fill = PatternFill("solid", start_color="F3F4F6", end_color="F3F4F6")
                cell.font = body_font(color="6B7280")
            elif "disclosure" in v:
                cell.fill = PatternFill("solid", start_color="DDD6FE", end_color="DDD6FE")
                cell.font = body_font(color=NEURAL_VIOLET, bold=True)

        # Criticity
        crit_cell = ws.cell(row=row, column=7)
        if "BLOQUANT" in str(crit_cell.value or ""):
            crit_cell.fill = red_fill()
            crit_cell.font = body_font(bold=True, color=RED_NO)
        elif "Conditionnel" in str(crit_cell.value or "") or "Important" in str(crit_cell.value or ""):
            crit_cell.fill = amber_fill()
            crit_cell.font = body_font(bold=True, color=AMBER_WARN)
        crit_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[15].height = 10

    # Légende
    write_section_header(ws, 16, "LÉGENDE")
    legend = [
        ("Cœur", "L'agent est principalement conçu pour traiter cette réglementation", "C7D2FE", "3730A3"),
        ("Indirect / Modéré", "L'agent touche cette réglementation comme effet secondaire", "FEF3C7", AMBER_WARN),
        ("Disclosure", "L'agent doit ajouter une mention de disclosure (AI Act)", "DDD6FE", NEURAL_VIOLET),
        ("Faible / Non applicable", "L'agent n'est pas concerné", "F3F4F6", "6B7280"),
        ("BLOQUANT", "Toute non-conformité empêche la mise en production", "FEE2E2", RED_NO),
        ("Conditionnel / Important", "Sanction possible mais non systématique", "FEF3C7", AMBER_WARN),
    ]
    for i, (label, desc, bg, fg) in enumerate(legend):
        row = 17 + i
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.fill = PatternFill("solid", start_color=bg, end_color=bg)
        cell_label.font = Font(name=FONT_FAMILY, size=10, bold=True, color=fg)
        cell_label.alignment = Alignment(horizontal="center", vertical="center")
        cell_label.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        cell_desc = ws.cell(row=row, column=2, value=desc)
        cell_desc.font = body_font()
        cell_desc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        cell_desc.border = THIN_BORDER
        ws.row_dimensions[row].height = 24

    set_widths(ws, {1: 32, 2: 28, 3: 24, 4: 26, 5: 26, 6: 26, 7: 22})


# ────────────────────────────────────────────────────────────────────
# Onglet 4 — KPIs Consolidés
# ────────────────────────────────────────────────────────────────────
def build_kpis_consolides(ws) -> None:
    write_title(ws, "NEURAL · OVERVIEW — 4_KPIs_Consolides : 12 KPIs cross-agents v1 → cible v2")
    write_subtitle(ws, 2, "3 KPIs par agent — valeur observée v1 + cible v2 + écart à combler")

    headers = ["Agent", "KPI", "Valeur v1 (observée)", "Cible v2 (objectif)", "Écart", "Unité", "Statut"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, k in enumerate(KPIS_CONSOLIDES):
        row = 5 + i
        ecart = k["cible_v2"] - k["v1"]
        statut = "✓ ATTEINT" if ecart <= 0 else "À PROGRESSER"
        vals = [k["agent"], k["kpi"], k["v1"], k["cible_v2"],
                f"=D{row}-C{row}", k["unite"], statut]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=26)

        # Agent badge
        agent_cell = ws.cell(row=row, column=1)
        agent_cell.fill = agent_fill(k["agent"])
        agent_cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        agent_cell.alignment = Alignment(horizontal="center", vertical="center")

        # v1 + v2 alignment
        for col in [3, 4]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=col).font = body_font(bold=True)

        # Écart (formula, blue)
        ecart_cell = ws.cell(row=row, column=5)
        ecart_cell.alignment = Alignment(horizontal="center", vertical="center")
        ecart_cell.font = body_font(bold=True, color="0000FF")

        # Unité
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")

        # Statut
        statut_cell = ws.cell(row=row, column=7)
        statut_cell.alignment = Alignment(horizontal="center", vertical="center")
        if statut == "✓ ATTEINT":
            statut_cell.fill = green_fill()
            statut_cell.font = body_font(bold=True, color=GREEN_OK)
        else:
            statut_cell.fill = amber_fill()
            statut_cell.font = body_font(bold=True, color=AMBER_WARN)

    ws.row_dimensions[17].height = 10

    # Stats agrégées
    write_section_header(ws, 18, "STATISTIQUES AGRÉGÉES — Calculées par formule")
    stats = [
        ("Nb KPIs total", "=COUNTA(B5:B16)"),
        ("Nb KPIs déjà atteints", '=COUNTIF(G5:G16,"*ATTEINT*")'),
        ("Nb KPIs à progresser", '=COUNTIF(G5:G16,"*PROGRESSER*")'),
        ("Écart moyen v1→v2", "=AVERAGE(E5:E16)"),
    ]
    for i, (label, formula) in enumerate(stats):
        row = 19 + i
        ws.cell(row=row, column=1, value=label).font = body_font(bold=True)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = body_font(bold=True, color="0000FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    set_widths(ws, {1: 22, 2: 38, 3: 22, 4: 22, 5: 14, 6: 14, 7: 22})


# ────────────────────────────────────────────────────────────────────
# Onglet 5 — Roadmap
# ────────────────────────────────────────────────────────────────────
def build_roadmap(ws) -> None:
    write_title(ws, "NEURAL · OVERVIEW — 5_Roadmap : Déploiement par phases (POC → SaaS)")
    write_subtitle(ws, 2, "6 phases sur 36 mois — du POC sur 5 produits au SaaS multi-tenant self-service")

    headers = ["Phase", "Durée", "Agents intégrés", "Livrable principal", "Statut"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, p in enumerate(ROADMAP):
        row = 5 + i
        vals = [p["phase"], p["duree"], p["agents"], p["livrable"], p["statut"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=55)

        # Status color
        status_cell = ws.cell(row=row, column=5)
        if "Démo" in str(status_cell.value or ""):
            status_cell.fill = green_fill()
            status_cell.font = body_font(bold=True, color=GREEN_OK)
        elif "Roadmap" in str(status_cell.value or ""):
            status_cell.fill = amber_fill()
            status_cell.font = body_font(bold=True, color=AMBER_WARN)
        else:
            status_cell.fill = PatternFill("solid", start_color="E0E7FF", end_color="E0E7FF")
            status_cell.font = body_font(bold=True, color="3730A3")

        # Phase title bold violet
        ws.cell(row=row, column=1).font = body_font(bold=True, color=NEURAL_VIOLET)

    ws.row_dimensions[11].height = 10

    write_section_header(ws, 12, "DÉPENDANCES CRITIQUES PAR PHASE")
    deps = [
        ("Phase 0 — POC", "Aucune — démo Excel auto-portée"),
        ("Phase 1 — Pilote", "Signature 1 client pilote + intégration API CMS du client"),
        ("Phase 2 — MVP", "Validation pilote + recrutement 1 ingénieur ML + 1 juriste assurance"),
        ("Phase 3 — V1", "Recrutement DPO + 2 ingénieurs ML + DPIA validée par CNIL"),
        ("Phase 4 — Scale", "Levée Série A (~3-5 M€) + 8-10 recrutements + clients EN/DE"),
        ("Phase 5 — SaaS", "Plateforme multi-tenant + équipe support 24/7 + certifications ISO 27001"),
    ]
    for i, (phase, dep) in enumerate(deps):
        row = 13 + i
        ws.cell(row=row, column=1, value=phase).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=dep).font = body_font(color="475569")
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 28
        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = alt_fill()

    set_widths(ws, {1: 26, 2: 22, 3: 28, 4: 50, 5: 28})


# ────────────────────────────────────────────────────────────────────
# Onglet 6 — Pricing & ROI
# ────────────────────────────────────────────────────────────────────
def build_pricing_roi(ws) -> None:
    write_title(ws, "NEURAL · OVERVIEW — 6_Pricing_ROI : Forfaits + retour sur investissement client")
    write_subtitle(ws, 2, "5 forfaits du Découverte au On-premise + calcul ROI client (forfait Suite complète)")

    write_section_header(ws, 4, "FORFAITS NEURAL — BRANCHE ASSURANCES/MARKETING")
    headers = ["Forfait", "Agents inclus", "Volume mensuel", "Prix mensuel (€)", "Prix annuel (€)", "Support"]
    style_header_row(ws, 5, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)

    for i, p in enumerate(PRICING):
        row = 6 + i
        vals = [p["forfait"], p["agents"], p["volume"],
                p["prix_mois"], p["prix_an"], p["support"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=42)

        # Forfait bold
        ws.cell(row=row, column=1).font = body_font(bold=True, color=NEURAL_VIOLET)

        # Prix formatting
        for col in [4, 5]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = "#,##0 €"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = body_font(bold=True)

    ws.row_dimensions[11].height = 10

    # ROI calc
    write_section_header(ws, 12, "ROI CLIENT — FORFAIT SUITE COMPLÈTE (BUSINESS CASE)")
    roi_headers = ["Métrique", "Valeur", "Unité", "Source / hypothèse"]
    style_header_row(ws, 13, len(roi_headers))
    for col, h in enumerate(roi_headers, 1):
        ws.cell(row=13, column=col, value=h)

    for i, r in enumerate(ROI):
        row = 14 + i
        vals = [r["metric"], r["valeur"], r["unite"], r["source"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(roi_headers), alt=(i % 2 == 1), height=28)

        # Valeur format
        val_cell = ws.cell(row=row, column=2)
        val_cell.alignment = Alignment(horizontal="right", vertical="center")
        if "€" in r["unite"] or "%" in r["unite"]:
            val_cell.font = body_font(bold=True, color=NEURAL_VIOLET)
            if "€" in r["unite"]:
                val_cell.number_format = "#,##0"
        else:
            val_cell.font = body_font(bold=True)

    # Highlight key rows : Économie, Marge, ROI
    for row_label_match in ["Économie mensuelle", "Marge brute", "ROI annuel"]:
        for i, r in enumerate(ROI):
            if r["metric"].startswith(row_label_match):
                row = 14 + i
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = green_fill()
                    ws.cell(row=row, column=col).font = body_font(bold=True, color=GREEN_OK)

    ws.row_dimensions[24].height = 10

    write_section_header(ws, 25, "SYNTHÈSE PITCH PROSPECT")
    pitch = [
        ("Promesse chiffrée", "Pour 6 800 €/mois (forfait Suite), un assureur économise environ 27 200 €/mois en équivalent FTE compliance + rédaction juridique = ROI mensuel × 4."),
        ("Risque évité", "Une seule amende ACPR pour publicité non conforme peut atteindre 5% du CA — pour un assureur à 100 M€ de CA, l'exposition est de 5 M€. NEURAL prévient ce risque pour 82 k€/an."),
        ("Time-to-market", "Le délai entre brief produit et campagne lancée passe de 6-9 semaines à 2-3 semaines — avantage compétitif majeur sur un marché où les comparateurs réagissent en jours."),
        ("Qualité homogène", "Garantie d'une cohérence cross-canal (agent / courtier / direct / comparateur) impossible à atteindre avec des équipes en silo."),
    ]
    for i, (k, v) in enumerate(pitch):
        row = 26 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font()
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 50
        if i % 2 == 1:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = alt_fill()

    set_widths(ws, {1: 22, 2: 38, 3: 32, 4: 18, 5: 18, 6: 30})


# ────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────
def main() -> None:
    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = "0_README"
    build_readme(ws_readme)

    ws_synth = wb.create_sheet("1_Agents_Synthese")
    build_agents_synthese(ws_synth)

    ws_pipe = wb.create_sheet("2_Pipeline_Orchestration")
    build_pipeline(ws_pipe)

    ws_reg = wb.create_sheet("3_Reglementation_Cross")
    build_reglementation_cross(ws_reg)

    ws_kpis = wb.create_sheet("4_KPIs_Consolides")
    build_kpis_consolides(ws_kpis)

    ws_road = wb.create_sheet("5_Roadmap")
    build_roadmap(ws_road)

    ws_pricing = wb.create_sheet("6_Pricing_ROI")
    build_pricing_roi(ws_pricing)

    out = "Assurance_Marketing_OVERVIEW_NEURAL.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
