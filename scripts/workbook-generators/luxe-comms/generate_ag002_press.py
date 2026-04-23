"""
NEURAL_AG002_LuxePressAgent.xlsx
Generation et revision de communiques presse / contenus presse.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, STATUS, VERT, VIOLET, YESNO,
    ajouter_table, attach_dv, convertir_dates, D, DT, dv_liste, ecrire_donnees,
    en_tetes, params_table, param_ligne, section, titre, today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_AG002_LuxePressAgent.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "AG-002  —  LUXEPRESSAGENT", "Redaction communiques presse — registre luxe, lifestyle vs business", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 86

    section(ws, 4, 6, "Mission", NOIR)
    ecrire_donnees(ws, 5, [
        ["", "Role",        "Rediger communiques dans le registre du luxe. Adapter presse lifestyle vs business."],
        ["", "Input",       "Brief (objet, angle, audience, media cibles, date pub)."],
        ["", "Output",      "Draft communique + dossier presse (media kit, quote CEO, photos recommandees)."],
        ["", "Dependances", "AG-001 VoiceGuard (gate brand obligatoire), AG-005 GreenClaim (si claims RSE)."],
        ["", "Appelants",   "USER direct, Orchestrator, Crisis flow."],
        ["", "Referentiel", "FOUNDATIONS!6_MEDIA_DIRECTORY + 2_BRAND_CHARTER."],
    ], align=AL)

    section(ws, 13, 6, "Flow wedge", BORDEAUX)
    ecrire_donnees(ws, 14, [
        ["", "1", "Intake brief dans 2_BRIEF_INTAKE"],
        ["", "2", "Draft dans 4_DRAFT_BUILDER selon media matrix"],
        ["", "3", "Handoff AG-001 (5_BRAND_HANDOFF) — score + decision"],
        ["", "4", "Si claim RSE detecte: handoff AG-005 (6_CLAIM_HANDOFF)"],
        ["", "5", "Revision loop (7_REVISION_LOOP) jusqu'a APPROVE"],
        ["", "6", "Final output (8_FINAL_OUTPUTS) + log (9_LOGS)"],
    ], align=AL)


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES AGENT", "tblParamsAG002", 3)
    rows = [
        ("AGENT_ID",        "AG-002", ""),
        ("AGENT_NAME",      "LuxePressAgent", ""),
        ("MAISON_ID",       "M-001", ""),
        ("MAISON_NAME",     "Maison pilote", ""),
        ("LANG_PRIMARY",    "FR", ""),
        ("LANG_SECONDARY",  "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", ""),
        ("DATA_VERSION",    "v2026.04", ""),
        ("PROMPT_VERSION",  "v1.0.0", ""),
        ("CRISIS_MODE_ON",  "NO", "Bypass d'embargo en mode crise"),
        ("MAX_REVISIONS",   5,  "Nombre de loops avant escalation"),
        ("LENGTH_TARGET_WORDS", 350, "Cible nombre de mots draft"),
        ("LAST_REFRESH",    D(today_iso()), ""),
    ]
    params_table(ws, "tblParamsAG002", start_row=4, rows=rows)


def onglet_brief_intake(wb):
    ws = wb.create_sheet("2_BRIEF_INTAKE")
    titre(ws, "BRIEF INTAKE", "Briefs presse recus — objet, angle, audience, media", 11)
    en_tetes(ws, 4, ["BRIEF_ID", "MAISON_ID", "OBJET", "TYPE", "ANGLE", "AUDIENCE_CIBLE",
                      "MEDIA_CIBLES", "DATE_PUBLICATION", "LANG", "EMBARGO", "PRIORITE"],
             [12, 10, 40, 16, 30, 22, 30, 16, 8, 14, 12])
    briefs = [
        ("BR-001", "M-001", "Communique lancement collection joaillerie 100 ans", "Launch",      "Heritage + exclusivite", "Lifestyle + trade",   "Vogue, HB, Les Echos",  D("2026-05-15"), "FR",    "J-3",  "P1"),
        ("BR-002", "M-001", "Annonce resultats H1 2026",                            "Corporate",   "Croissance + ESG",       "Business press",      "FT, Les Echos, BoF",    D("2026-07-10"), "FR+EN", "J-1",  "P0"),
        ("BR-003", "M-001", "Rencontre artisans atelier",                          "Feature",     "Savoir-faire",           "Trade + lifestyle",   "WWD, Numero",           D("2026-04-25"), "FR",    "Aucun","P2"),
        ("BR-004", "M-001", "Pop-up Tokyo Harajuku",                               "Launch",      "Marche japonais",        "Lifestyle JP",        "Nikkei, Vogue JP",      D("2026-07-15"), "FR+JA", "J-7",  "P1"),
        ("BR-005", "M-001", "Collaboration artiste contemporain",                  "Partnership", "Art + patrimoine",       "Lifestyle + art",     "T Mag, AD France",      D("2026-10-20"), "FR+EN", "J-2",  "P1"),
        ("BR-006", "M-001", "Rapport ESG 2025",                                     "ESG",         "Transparence RSE",       "Trade + ESG",         "BoF, FT, Jing Daily",   D("2026-06-05"), "FR+EN", "J-0",  "P0"),
        ("BR-007", "M-001", "Nomination nouveau Directeur Artistique",             "Corporate",   "Vision creative",        "Business + lifestyle","Tous P1",               D("2026-05-02"), "FR+EN", "J-1",  "P0"),
    ]
    last = ecrire_donnees(ws, 5, briefs, align=AL)
    convertir_dates(ws, 5, last, date_cols=[8])
    ajouter_table(ws, "tblBriefIntake", f"A4:K{last}")
    attach_dv(ws, f"D5:D{last}", ["Launch", "Corporate", "Feature", "Partnership", "ESG", "Crisis"])
    attach_dv(ws, f"K5:K{last}", ["P0", "P1", "P2", "P3"])

    ws.conditional_formatting.add(f"K5:K{last}", CellIsRule(operator="equal", formula=['"P0"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"K5:K{last}", CellIsRule(operator="equal", formula=['"P1"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_media_matrix(wb):
    ws = wb.create_sheet("3_MEDIA_MATRIX")
    titre(ws, "MEDIA MATRIX", "Angle + format cible par type de media", 7)
    en_tetes(ws, 4, ["MEDIA_TYPE", "ANGLE", "FORMAT_TARGET", "LENGTH_WORDS", "QUOTE_CEO", "VISUALS_REQUIRED", "EMBARGO_RECOMMAND"],
             [20, 36, 30, 14, 14, 20, 18])
    matrix = [
        ("Lifestyle",    "Narration image / maison / heritage",         "Communique 400-600 mots + media kit", 500, "YES", "3-5 photos HD",  "J-3"),
        ("Business",     "Angles resultats / direction / expansion",     "Communique 300-400 mots + data",       350, "YES", "1 photo + charts","J-1"),
        ("Trade",        "Strategie / innovation / filiere",             "Communique 400-500 mots + interview",  450, "NO",  "2 photos",        "J-2"),
        ("Digital",      "Angle digital / conversion / contenu short",   "Communique 250-350 mots + visuels",    300, "NO",  "5-10 photos",     "Aucun"),
        ("Social",       "Capsule + teaser + storytelling",              "Post + carousel + short video",        150, "NO",  "10 photos",       "Aucun"),
        ("Generaliste",  "Dimension societale / culture / France",       "Communique 400-500 mots",              450, "YES", "3 photos",        "J-2"),
        ("Magazine",     "Interview + portrait + atelier",                "Dossier long 800-1500 mots",           1000,"YES", "10 photos",       "J-7"),
    ]
    last = ecrire_donnees(ws, 5, matrix, fmt_cols={4: "0"}, align=AL)
    ajouter_table(ws, "tblMediaMatrix", f"A4:G{last}")


def onglet_draft_builder(wb):
    ws = wb.create_sheet("4_DRAFT_BUILDER")
    titre(ws, "DRAFT BUILDER", "Construction drafts — consomme media matrix + brief", 10)
    en_tetes(ws, 4, ["DRAFT_ID", "BRIEF_ID", "MAISON_ID", "MEDIA_TYPE", "LANG", "ANGLE_TARGET",
                      "FORMAT_TARGET", "LENGTH_TARGET", "BODY_SUMMARY", "STATUS"],
             [12, 12, 10, 18, 8, 32, 30, 14, 50, 14])
    drafts = [
        ("DR-001", "BR-001", "M-001", "Lifestyle",   "FR", "", "", 0, "La maison devoile une collection de haute joaillerie inspiree des archives de 1920. Chaque piece est ciselee a la main par les artisans de l'atelier parisien.", "DRAFT"),
        ("DR-002", "BR-002", "M-001", "Business",    "FR", "", "", 0, "Resultats H1 2026 en croissance de X%, avec une contribution de la maroquinerie et une augmentation de l'or recycle dans la collection.",                          "DRAFT"),
        ("DR-003", "BR-003", "M-001", "Trade",       "FR", "", "", 0, "Rencontre en atelier avec les artisans de la maison, autour du geste de ciselure transmis depuis 1923.",                                                      "DRAFT"),
        ("DR-004", "BR-004", "M-001", "Lifestyle",   "JA", "", "", 0, "Pop-up a Tokyo Harajuku, presentation d'une capsule pensee pour le marche japonais.",                                                                          "DRAFT"),
        ("DR-005", "BR-005", "M-001", "Magazine",    "FR", "", "", 0, "Collaboration avec artiste contemporain — fusion art et patrimoine maison.",                                                                                   "DRAFT"),
        ("DR-006", "BR-006", "M-001", "Business",    "FR", "", "", 0, "Rapport ESG 2025 — audit LBMA, empreinte -30%, emplois artisans preserves.",                                                                                   "DRAFT"),
        ("DR-007", "BR-007", "M-001", "Generaliste", "FR", "", "", 0, "Nomination nouveau Directeur Artistique, vision creative 2026-2030.",                                                                                           "DRAFT"),
    ]
    last = ecrire_donnees(ws, 5, drafts, fmt_cols={8: "0"}, align=AL)

    # Formules XLOOKUP vers Media Matrix
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=6, value=f'=IFERROR(XLOOKUP(D{r},tblMediaMatrix[MEDIA_TYPE],tblMediaMatrix[ANGLE]),"")')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF

        c = ws.cell(row=r, column=7, value=f'=IFERROR(XLOOKUP(D{r},tblMediaMatrix[MEDIA_TYPE],tblMediaMatrix[FORMAT_TARGET]),"")')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF

        c = ws.cell(row=r, column=8, value=f'=IFERROR(XLOOKUP(D{r},tblMediaMatrix[MEDIA_TYPE],tblMediaMatrix[LENGTH_WORDS]),0)')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF
        c.number_format = "0"

    ajouter_table(ws, "tblDrafts", f"A4:J{last}")


def onglet_brand_handoff(wb):
    ws = wb.create_sheet("5_BRAND_HANDOFF")
    titre(ws, "BRAND HANDOFF → AG-001", "Drafts envoyes a AG-001 pour gate brand", 8)
    en_tetes(ws, 4, ["HANDOFF_ID", "DRAFT_ID", "SENT_TO", "SENT_AT", "BRAND_SCORE", "BRAND_DECISION",
                      "FEEDBACK", "BACK_AT"],
             [14, 12, 14, 18, 12, 16, 48, 18])
    ho = [
        ("HO-B-001", "DR-001", "AG-001", DT("2026-04-15 09:20"), 95, "APPROVE",  "Conforme charte.",                                 DT("2026-04-15 09:24")),
        ("HO-B-002", "DR-002", "AG-001", DT("2026-04-15 10:35"), 25, "REJECT",   "Superlatif + discount + concurrent.",              DT("2026-04-15 10:40")),
        ("HO-B-003", "DR-003", "AG-001", DT("2026-04-15 11:15"), 100,"APPROVE",  "Conforme, excellent vocabulaire.",                 DT("2026-04-15 11:19")),
        ("HO-B-004", "DR-004", "AG-001", DT("2026-04-15 13:45"), 50, "REJECT",   "Superlatifs EN + sustainable non prouve.",         DT("2026-04-15 13:50")),
        ("HO-B-005", "DR-005", "AG-001", DT("2026-04-15 14:30"), 80, "APPROVE",  "Conforme, bon angle heritage.",                    DT("2026-04-15 14:34")),
        ("HO-B-006", "DR-006", "AG-001", DT("2026-04-15 16:10"), 90, "APPROVE",  "Conforme, ton factuel ESG.",                       DT("2026-04-15 16:14")),
        ("HO-B-007", "DR-007", "AG-001", DT("2026-04-15 17:00"), 85, "APPROVE",  "Conforme, annonce claire.",                        DT("2026-04-15 17:04")),
    ]
    last = ecrire_donnees(ws, 5, ho, fmt_cols={5: "0"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[4, 8])
    ajouter_table(ws, "tblBrandHandoff", f"A4:H{last}")
    attach_dv(ws, f"F5:F{last}", ["APPROVE", "REWORK", "REJECT"])

    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"APPROVE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"REWORK"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"REJECT"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_claim_handoff(wb):
    ws = wb.create_sheet("6_CLAIM_HANDOFF")
    titre(ws, "CLAIM HANDOFF → AG-005", "Drafts avec claims RSE detectes — envoyes a AG-005", 8)
    en_tetes(ws, 4, ["HANDOFF_ID", "DRAFT_ID", "CLAIM_RAW", "RSE_TAG", "SENT_TO", "CLAIM_DECISION", "RISK_CLASS", "ACTION"],
             [14, 12, 40, 12, 14, 18, 14, 28])
    ho = [
        ("HO-C-001", "DR-001", "-",                                 "NO",  "-",     "-",        "-",        "NO_CLAIM_CHECK"),
        ("HO-C-002", "DR-002", "Or certifie 80% recycle",            "YES", "AG-005","PASS",     "MEDIUM",   "OK_PUBLISH"),
        ("HO-C-003", "DR-003", "-",                                 "NO",  "-",     "-",        "-",        "NO_CLAIM_CHECK"),
        ("HO-C-004", "DR-004", "fully sustainable",                  "YES", "AG-005","BLOCK",    "CRITICAL", "REWRITE_NO_CLAIM"),
        ("HO-C-005", "DR-005", "-",                                 "NO",  "-",     "-",        "-",        "NO_CLAIM_CHECK"),
        ("HO-C-006", "DR-006", "or recycle, empreinte -30%, LWG",    "YES", "AG-005","PASS",     "MEDIUM",   "OK_PUBLISH"),
        ("HO-C-007", "DR-007", "-",                                 "NO",  "-",     "-",        "-",        "NO_CLAIM_CHECK"),
    ]
    last = ecrire_donnees(ws, 5, ho, align=AL)

    # Formule action si RSE_TAG=YES
    # (On a deja le resultat seed, mais on peut derouler la logique)
    ajouter_table(ws, "tblClaimHandoff", f"A4:H{last}")

    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"BLOCK"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_revision_loop(wb):
    ws = wb.create_sheet("7_REVISION_LOOP")
    titre(ws, "REVISION LOOP", "Boucle de revision jusqu'a APPROVE ou escalation", 10)
    en_tetes(ws, 4, ["LOOP_ID", "DRAFT_ID", "ITERATION", "BRAND_DECISION", "CLAIM_DECISION", "NEXT_STEP", "ACTION_OWNER", "COMMENT", "START_AT", "END_AT"],
             [12, 12, 10, 16, 16, 26, 18, 30, 18, 18])
    loops = [
        ("LP-001", "DR-001", 1, "APPROVE", "-",       "READY_FOR_REVIEW", "PR Director",   "First pass OK",                        DT("2026-04-15 09:20"), DT("2026-04-15 09:30")),
        ("LP-002", "DR-002", 1, "REJECT",  "PASS",    "TO_REWRITE",       "AG-002",        "Superlatifs + concurrent a corriger",   DT("2026-04-15 10:35"), DT("2026-04-15 10:50")),
        ("LP-003", "DR-002", 2, "APPROVE", "PASS",    "READY_FOR_REVIEW", "PR Director",   "Rewrite OK",                            DT("2026-04-15 11:10"), DT("2026-04-15 11:15")),
        ("LP-004", "DR-004", 1, "REJECT",  "BLOCK",   "TO_REWRITE_NOCLAIM","AG-002",        "Supprimer 'fully sustainable'",        DT("2026-04-15 13:45"), DT("2026-04-15 14:00")),
        ("LP-005", "DR-004", 2, "APPROVE", "PASS",    "READY_FOR_REVIEW", "PR Director",   "Claim supprime",                        DT("2026-04-15 14:20"), DT("2026-04-15 14:25")),
        ("LP-006", "DR-005", 1, "APPROVE", "-",       "READY_FOR_REVIEW", "PR Director",   "OK first pass",                         DT("2026-04-15 14:30"), DT("2026-04-15 14:35")),
        ("LP-007", "DR-006", 1, "APPROVE", "PASS",    "READY_FOR_REVIEW", "PR Director",   "OK ESG + brand",                        DT("2026-04-15 16:10"), DT("2026-04-15 16:20")),
        ("LP-008", "DR-007", 1, "APPROVE", "-",       "READY_FOR_REVIEW", "PR Director",   "OK first pass",                         DT("2026-04-15 17:00"), DT("2026-04-15 17:10")),
    ]
    last = ecrire_donnees(ws, 5, loops, fmt_cols={3: "0"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[9, 10])
    ajouter_table(ws, "tblRevisionLoop", f"A4:J{last}")

    # CF iteration > 3 = alerte
    ws.conditional_formatting.add(f"C5:C{last}", FormulaRule(formula=[f"C5>3"], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"READY_FOR_REVIEW"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))


def onglet_final_outputs(wb):
    ws = wb.create_sheet("8_FINAL_OUTPUTS")
    titre(ws, "FINAL OUTPUTS", "Communiques finalises prets pour diffusion", 10)
    en_tetes(ws, 4, ["OUTPUT_ID", "DRAFT_ID", "BRIEF_ID", "MAISON_ID", "MEDIA_TYPE", "LANG",
                      "TITRE", "EMBARGO_UNTIL", "FIRST_PASS_OK", "REVISION_COUNT"],
             [14, 12, 12, 10, 18, 8, 44, 18, 14, 14])
    outputs = [
        ("OUT-001", "DR-001", "BR-001", "M-001", "Lifestyle",    "FR", "Une collection joaillerie, cent ans apres",              D("2026-05-12"), "YES", 1),
        ("OUT-002", "DR-002", "BR-002", "M-001", "Business",     "FR", "Resultats H1 2026 — croissance stable et discipline",    D("2026-07-09"), "NO",  2),
        ("OUT-003", "DR-003", "BR-003", "M-001", "Trade",        "FR", "Dans l'atelier, le geste transmis",                      None,            "YES", 1),
        ("OUT-004", "DR-004", "BR-004", "M-001", "Lifestyle",    "JA", "Pop-up Tokyo",                                            D("2026-07-08"), "NO",  2),
        ("OUT-005", "DR-005", "BR-005", "M-001", "Magazine",     "FR", "Dialogue avec un artiste contemporain",                   D("2026-10-18"), "YES", 1),
        ("OUT-006", "DR-006", "BR-006", "M-001", "Business",     "FR", "Rapport ESG 2025 — or recycle, eau, emplois",             D("2026-06-05"), "YES", 1),
        ("OUT-007", "DR-007", "BR-007", "M-001", "Generaliste",  "FR", "Nomination a la direction artistique",                    D("2026-05-01"), "YES", 1),
    ]
    last = ecrire_donnees(ws, 5, outputs, fmt_cols={10: "0"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[8])
    ajouter_table(ws, "tblFinalOutputs", f"A4:J{last}")
    attach_dv(ws, f"I5:I{last}", YESNO)

    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_logs(wb):
    ws = wb.create_sheet("9_LOGS")
    titre(ws, "LOGS", "Journal — export vers MASTER", 14)
    en_tetes(ws, 4, ["RUN_ID", "AGENT_ID", "MAISON_ID", "DOC_ID", "TIMESTAMP", "LANG",
                      "FINAL_STATUS", "RISK_LEVEL", "GATE_TRIGGERED", "SLA_MET",
                      "TOTAL_MINUTES", "TOKENS_IN", "TOKENS_OUT", "COST_USD"],
             [16, 10, 10, 12, 18, 8, 14, 12, 12, 10, 12, 12, 12, 12])
    rows = [
        ("R-20260415-020", "AG-002", "M-001", "DOC-0001", DT("2026-04-15 09:10"), "FR", "APPROVED", "LOW",    "BRAND", "YES", 8.1,  3200, 2100, 0.064),
        ("R-20260415-021", "AG-002", "M-001", "DOC-0002", DT("2026-04-15 10:22"), "FR", "APPROVED", "MEDIUM", "BRAND", "YES", 18.2, 5500, 3600, 0.108),
        ("R-20260415-022", "AG-002", "M-001", "DOC-0003", DT("2026-04-15 11:04"), "FR", "APPROVED", "LOW",    "BRAND", "YES", 7.9,  2900, 2000, 0.059),
        ("R-20260415-023", "AG-002", "M-001", "DOC-0004", DT("2026-04-15 13:30"), "JA", "APPROVED", "MEDIUM", "BRAND", "YES", 17.5, 5300, 3400, 0.102),
        ("R-20260415-024", "AG-002", "M-001", "DOC-0005", DT("2026-04-15 14:30"), "FR", "APPROVED", "LOW",    "BRAND", "YES", 14.2, 4200, 3100, 0.091),
        ("R-20260415-025", "AG-002", "M-001", "DOC-0006", DT("2026-04-15 16:10"), "FR", "APPROVED", "MEDIUM", "BRAND", "YES", 12.5, 4000, 2800, 0.082),
        ("R-20260415-026", "AG-002", "M-001", "DOC-0007", DT("2026-04-15 17:00"), "FR", "APPROVED", "LOW",    "BRAND", "YES", 9.0,  3100, 2300, 0.067),
    ]
    last = ecrire_donnees(ws, 5, rows, fmt_cols={11: "0.00", 12: "0", 13: "0", 14: "0.000"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5])
    ajouter_table(ws, "tblPressLogs", f"A4:N{last}")


def onglet_dashboard(wb):
    ws = wb.create_sheet("10_DASHBOARD")
    titre(ws, "DASHBOARD", "First-pass rate, revisions moyennes, cout, volume", 4)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18

    section(ws, 4, 4, "Indicateurs", NOIR)
    kpis = [
        ("First-pass validation rate",    '=IFERROR(COUNTIFS(tblFinalOutputs[FIRST_PASS_OK],"YES")/COUNTA(tblFinalOutputs[OUTPUT_ID]),0)', "0.0%"),
        ("Revisions moyennes par draft",  '=IFERROR(AVERAGE(tblFinalOutputs[REVISION_COUNT]),0)',                                          "0.0"),
        ("Drafts total",                  '=COUNTA(tblFinalOutputs[OUTPUT_ID])',                                                           "0"),
        ("Volume briefs ouverts",         '=COUNTA(tblBriefIntake[BRIEF_ID])',                                                             "0"),
        ("Cout USD cumule",               '=IFERROR(SUM(tblPressLogs[COST_USD]),0)',                                                       "0.000"),
        ("Cout moyen USD / draft",        '=IFERROR(AVERAGE(tblPressLogs[COST_USD]),0)',                                                   "0.000"),
        ("Press pickup rate (post-publi)",'=IFERROR(COUNTIFS(tblPressPickup[PICKUP],"YES")/COUNTA(tblPressPickup[PICKUP_ID]),0)',        "0.0%"),
    ]
    for i, (lbl, formula, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = fmt

    # Volume par type de media
    section(ws, 14, 4, "Volume par type de media", BORDEAUX)
    media_types = ["Lifestyle", "Business", "Trade", "Digital", "Social", "Generaliste", "Magazine"]
    for i, mt in enumerate(media_types):
        r = 15 + i
        ws.cell(row=r, column=2, value=mt).border = BF
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True)
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(tblFinalOutputs[MEDIA_TYPE],"{mt}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch = BarChart()
    ch.type = "bar"
    data = Reference(ws, min_col=3, min_row=15, max_row=14 + len(media_types))
    labels = Reference(ws, min_col=2, min_row=15, max_row=14 + len(media_types))
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(labels)
    ch.title = "Volume par type de media"
    ch.height = 9
    ch.width = 14
    ws.add_chart(ch, "E14")


def onglet_press_pickup(wb):
    """R-09 : mesure post-publication (reprise presse, reach, sentiment)."""
    ws = wb.create_sheet("11_PRESS_PICKUP")
    titre(ws, "PRESS PICKUP", "Mesure post-publication : reprise media, reach, sentiment", 10)
    en_tetes(ws, 4, ["PICKUP_ID", "OUTPUT_ID", "MEDIA_REPRIS", "PAYS", "LANG", "DATE_REPRISE",
                      "PICKUP", "REACH", "SENTIMENT", "TONALITE"],
             [12, 12, 24, 10, 8, 14, 10, 14, 14, 28])
    rows = [
        ("PK-001", "OUT-001", "Vogue FR",          "FR", "FR", D("2026-05-16"), "YES",  450000, "POSITIF", "Valorisation heritage maison"),
        ("PK-002", "OUT-001", "Harper's Bazaar",   "US", "EN", D("2026-05-18"), "YES",  280000, "POSITIF", "Tres favorable"),
        ("PK-003", "OUT-001", "Numero",            "FR", "FR", D("2026-05-20"), "YES",  80000,  "POSITIF", "Angle art / patrimoine"),
        ("PK-004", "OUT-002", "FT",                "UK", "EN", D("2026-07-11"), "YES",  520000, "NEUTRE",  "Resultats factuels"),
        ("PK-005", "OUT-002", "Les Echos",         "FR", "FR", D("2026-07-11"), "YES",  180000, "POSITIF", "Croissance saluee"),
        ("PK-006", "OUT-002", "BoF",               "UK", "EN", D("2026-07-12"), "YES",  95000,  "POSITIF", "Analyse strategie"),
        ("PK-003", "OUT-003", "WWD",               "US", "EN", D("2026-04-26"), "YES",  60000,  "POSITIF", "Savoir-faire"),
        ("PK-008", "OUT-003", "BoF",               "UK", "EN", D("2026-04-27"), "NO",   0,      "-",       "Non repris"),
        ("PK-009", "OUT-006", "Jing Daily",        "CN", "EN", D("2026-06-06"), "YES",  110000, "POSITIF", "Rapport ESG credible"),
        ("PK-010", "OUT-006", "FT",                "UK", "EN", D("2026-06-06"), "YES",  520000, "NEUTRE",  "ESG audit retenu"),
        ("PK-011", "OUT-007", "Tous P1",           "FR", "FR", D("2026-05-03"), "YES",  820000, "POSITIF", "Couverture massive"),
    ]
    last = ecrire_donnees(ws, 5, rows, fmt_cols={8: "#,##0"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[6])
    ajouter_table(ws, "tblPressPickup", f"A4:J{last}")
    attach_dv(ws, f"E5:E{last}", LANG)
    attach_dv(ws, f"G5:G{last}", YESNO)
    attach_dv(ws, f"I5:I{last}", ["POSITIF", "NEUTRE", "NEGATIF", "-"])

    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"POSITIF"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"NEGATIF"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_learnings(wb):
    """R-13 : boucle d'apprentissage AG-002."""
    ws = wb.create_sheet("12_LEARNINGS")
    titre(ws, "LEARNINGS — AG-002", "Capitalisation post-run : motifs revision, angle presse non recu", 8)
    en_tetes(ws, 4, ["LEARN_ID", "DATE", "SOURCE_RUN", "MOTIF", "FREQUENCE_MOIS", "ACTION_PROPOSEE", "OWNER", "STATUT"],
             [12, 12, 18, 40, 16, 44, 18, 14])
    lrn = [
        ("LRN-P-001", D(today_iso()), "R-20260415-023", "Draft JA necessite second pass pour vocab local", 3, "Ajouter vocab JA dans FOUNDATIONS",       "Brand Dir",  "OPEN"),
        ("LRN-P-002", D(today_iso()), "R-20260415-025", "Media matrix 'Business' trop generique pour ESG", 2, "Ajouter sous-type 'Business-ESG'",        "PR Director","OPEN"),
        ("LRN-P-003", D(today_iso()), "-",              "Taux de pickup <60% sur OUT-003",                1, "Revoir angle editorial trade",            "PR Director","OPEN"),
    ]
    last = ecrire_donnees(ws, 5, lrn, align=AL)
    convertir_dates(ws, 5, last, date_cols=[2])
    ajouter_table(ws, "tblLearningsAG002", f"A4:H{last}")
    attach_dv(ws, f"H5:H{last}", ["OPEN", "IN_REVIEW", "ACCEPTED", "REJECTED", "DONE"])


def build():
    wb = Workbook()
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_brief_intake(wb)
    onglet_media_matrix(wb)
    onglet_draft_builder(wb)
    onglet_brand_handoff(wb)
    onglet_claim_handoff(wb)
    onglet_revision_loop(wb)
    onglet_final_outputs(wb)
    onglet_logs(wb)
    onglet_dashboard(wb)
    onglet_press_pickup(wb)
    onglet_learnings(wb)

    order = ["0_README", "1_PARAMS", "2_BRIEF_INTAKE", "3_MEDIA_MATRIX",
             "4_DRAFT_BUILDER", "5_BRAND_HANDOFF", "6_CLAIM_HANDOFF", "7_REVISION_LOOP",
             "8_FINAL_OUTPUTS", "9_LOGS", "10_DASHBOARD", "11_PRESS_PICKUP", "12_LEARNINGS"]
    wb._sheets = [wb[n] for n in order]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
