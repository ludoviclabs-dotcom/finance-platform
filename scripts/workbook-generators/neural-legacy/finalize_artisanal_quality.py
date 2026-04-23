from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from generate_artisanal_quality import (
    AC,
    BF,
    BLEU,
    BLANC_CASSE,
    BORDEAUX,
    FOND_BLEU,
    FOND_JAUNE,
    FOND_ROUGE,
    FOND_VERT,
    GRIS,
    NOIR,
    OR,
    ROUGE,
    VERT,
    ecrire_donnees,
    en_tetes,
    formule_or,
    generer_artisanal_quality,
    ligne_synthese,
    mfc_statut,
    section,
    titre,
)


def _recreate_sheet(wb, name, index=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index=index)
    return ws


def finaliser_artisanal_quality():
    source_path = Path(generer_artisanal_quality())
    wb = load_workbook(source_path)

    ws = _recreate_sheet(wb, "9_ALERTES")
    ws.sheet_properties.tabColor = ROUGE
    titre(ws, "ALERTES QUALITE — AGENT ARTISANALQUALITYAI", "Alertes actives — Seuils qualite — Escalade intelligente — Avril 2026", 14)
    section(ws, 3, 14, "REGISTRE DES ALERTES QUALITE ACTIVES")
    headers = ["ID_Alerte", "Date", "Heure", "Type alerte", "Criticite", "Metier", "Produit / Lot concerne", "Description", "Impact", "Action requise", "Responsable", "Delai (j)", "Statut", "Date cloture"]
    widths = [14, 12, 8, 22, 12, 14, 22, 40, 32, 36, 20, 10, 12, 12]
    en_tetes(ws, 4, headers, widths)
    alertes = [
        ["AQ-2026-001", "2026-04-02", "10:45", "Non-conformite produit", "ELEVE", "Maroquinerie", "SAC-BI-28-CAMEL-003", "Double defaut : griffure cuir + piqure irreguliere — Piece rejetee en premiere inspection", "Retard livraison client 5 jours — Cout reprise 830 EUR — Image qualite interne", "Reprise complete — Plan formation artisan Lefevre — Audit eclairage tri", "Sophie Laurent", 5, "Resolu", "2026-04-05"],
        ["AQ-2026-002", "2026-04-05", "12:00", "Defaut matiere premiere fournisseur", "CRITIQUE", "Pret-a-Porter", "Lot soie Seteria Bianchi #SB-2026-T089", "DeltaE 2.1 + solidite frottement 2/5 — Lot entier non conforme — 3 robes impactees — Risque lot complet", "Blocage production soie — 12 pieces en attente — Retard collection — Cout >7000 EUR — Risque fournisseur", "Blocage lot — Retour fournisseur — Audit teinture Seteria Bianchi — Recherche lot remplacement", "Dir. Qualite + Dir. Achats", 10, "En cours", None],
        ["AQ-2026-003", "2026-04-03", "15:30", "Seuil qualite franchi", "MOYEN", "Joaillerie", "BRA-MAN-OR18-007", "Score qualite 82/100 — Sous le seuil atelier 90/100 — Piece validee apres reprise mais score initial faible", "Alerte monitoring — Pas de blocage — Vigilance accrue sur finition bracelet or", "Analyse cause + reorganisation poste polissage — Sensibilisation manipulation", "Alain Mercier", 7, "Resolu", "2026-04-08"],
        ["AQ-2026-004", "2026-04-10", "09:00", "Tendance degradation", "ELEVE", "Pret-a-Porter", "Atelier PAP Milan — Global", "FPY atelier Milan passe de 94% a 77% en avril — 3 non-conformites en 10 jours — Tendance negative", "Risque systemique — Qualite collection ete 2026 menacee — Retards cascade", "Audit flash atelier Milan — Revue process — Renforcement controles reception matieres", "Marco Ricci + Dir. Qualite", 5, "En cours", None],
        ["AQ-2026-005", "2026-04-10", "14:00", "Competence artisan", "MOYEN", "Pret-a-Porter", "Catherine Moreau — Artisan #ART-009", "2 defauts attribues en avril (DEF-004 raccord motif + DEF-007 boutonniere) — Besoin formation identifie", "Plan de formation individuel necessaire — Risque de repetition", "Entretien individuel + plan de formation placement patron + maintenance machine", "Resp. Atelier + DRH", 15, "En cours", None],
        ["AQ-2026-006", "2026-04-11", "10:30", "Excellence detectee", "INFO", "Horlogerie", "MON-TOUR-CHRONO-001", "Score 100/100 — Tourbillon Maitre Bonnard — Piece d'exception — Candidature prix interne", "Impact positif — Valorisation savoir-faire — Communication interne et externe", "Nomination prix excellence artisanale — Documentation pour communication Maison", "Pierre Fontaine", 30, "Ouvert", None],
    ]
    nb_al = len(alertes)
    ecrire_donnees(ws, 5, alertes)
    mfc_statut(ws, "M", 5, 4 + nb_al)
    for val, bg, fg in [("CRITIQUE", "FF0000", "FFFFFF"), ("ELEVE", "FF6600", "FFFFFF"), ("MOYEN", FOND_JAUNE, "9C6500"), ("INFO", FOND_BLEU, BLEU)]:
        ws.conditional_formatting.add(f"E5:E{4 + nb_al}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"), font=Font(color=fg, bold=True)))
    sa_r = 4 + nb_al + 2
    section(ws, sa_r, 8, "SYNTHESE ALERTES")
    synth_al = [("CRITIQUES ouvertes", f'=COUNTIFS(E5:E{4 + nb_al},"CRITIQUE",M5:M{4 + nb_al},"<>Resolu")'), ("ELEVEES ouvertes", f'=COUNTIFS(E5:E{4 + nb_al},"ELEVE",M5:M{4 + nb_al},"<>Resolu")'), ("MOYENNES ouvertes", f'=COUNTIFS(E5:E{4 + nb_al},"MOYEN",M5:M{4 + nb_al},"<>Resolu")'), ("Resolues", f'=COUNTIF(M5:M{4 + nb_al},"Resolu")'), ("Total actives", f'=COUNTIF(M5:M{4 + nb_al},"<>Resolu")')]
    for idx, (lab, form) in enumerate(synth_al):
        r = sa_r + 1 + idx
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r, column=1, value=lab)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=1).border = BF
        ws.merge_cells(f"E{r}:F{r}")
        formule_or(ws, r, 5, form, "0")
    dv_crit = DataValidation(type="list", formula1='"CRITIQUE,ELEVE,MOYEN,FAIBLE,INFO"', allow_blank=False)
    ws.add_data_validation(dv_crit)
    dv_crit.add("E5:E5000")
    dv_stat_al = DataValidation(type="list", formula1='"Ouvert,En cours,Resolu,Escalade"', allow_blank=False)
    ws.add_data_validation(dv_stat_al)
    dv_stat_al.add("M5:M5000")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{4 + nb_al}"
    print("  OK  Onglet 9_ALERTES cree — 6 alertes qualite")

    ws = _recreate_sheet(wb, "10_PARAMETRES")
    ws.sheet_properties.tabColor = GRIS
    titre(ws, "PARAMETRES & REFERENTIELS — AGENT ARTISANALQUALITYAI", "Referentiels qualite — Classifications — Seuils — Configuration systeme — Avril 2026", 8)
    section(ws, 4, 4, "REF A — CLASSIFICATION DES DEFAUTS")
    en_tetes(ws, 5, ["Code", "Type de defaut", "Description", "Severite par defaut"], [12, 24, 40, 16])
    ecrire_donnees(ws, 6, [
        ("TYP-EST", "Esthetique", "Defaut visuel n'affectant pas la fonctionnalite mais l'apparence", "Mineur a Majeur"),
        ("TYP-FON", "Fonctionnel", "Defaut affectant l'usage ou la durabilite du produit", "Majeur a Critique"),
        ("TYP-DIM", "Dimensionnel", "Ecart de mesure hors tolerance — Taille, poids, epaisseur", "Mineur a Majeur"),
        ("TYP-MAT", "Matiere premiere", "Defaut inherent a la matiere brute — Cuir, tissu, metal, pierre", "Variable"),
        ("TYP-ASS", "Assemblage", "Defaut de montage, collage, rivetage, sertissage", "Majeur a Critique"),
        ("TYP-FIN", "Finition", "Defaut de finition — Polissage, teinture, tranche, rhodiage", "Mineur"),
        ("TYP-EMB", "Emballage / Presentation", "Defaut d'emballage, etiquetage, presentation", "Mineur"),
        ("TYP-CON", "Contamination", "Presence corps etranger, tache, residu chimique", "Majeur a Critique"),
    ])
    ref_b = 6 + 8 + 2
    section(ws, ref_b, 4, "REF B — NIVEAUX DE MAITRISE ARTISANALE")
    en_tetes(ws, ref_b + 1, ["Niveau", "Annees min.", "Description", "Equivalent"], [18, 12, 40, 20])
    ecrire_donnees(ws, ref_b + 2, [
        ("Apprenti", "0-2", "En formation — Sous supervision directe permanente", "CAP / BEP"),
        ("Compagnon", "2-5", "Autonome sur taches courantes — Supervision ponctuelle", "BP / BMA"),
        ("Confirme", "5-10", "Maitrise complete metier — Peut former les apprentis", "CQP Luxe"),
        ("Expert", "10-20", "Excellence reconnue — Referent technique — Pieces complexes", "DMA / Titre RNCP"),
        ("Maitre d'art", "20+", "Savoir-faire exceptionnel — Distinction nationale — Pieces uniques", "MOF / Maitre d'Art"),
    ])
    ref_c = ref_b + 2 + 5 + 2
    section(ws, ref_c, 6, "REF C — SEUILS & PARAMETRES QUALITE")
    en_tetes(ws, ref_c + 1, ["Parametre", "Valeur", "Unite", "Min", "Max", "Description"], [32, 12, 10, 10, 10, 44])
    ecrire_donnees(ws, ref_c + 2, [
        ("Score qualite minimum — piece individuelle", 85, "/100", 80, 95, "En dessous : piece rejetee ou en reprise"),
        ("Score qualite minimum — atelier (moyenne)", 90, "/100", 85, 95, "En dessous : audit flash declenche"),
        ("First Pass Yield cible", 95, "%", 90, 99, "% pieces conformes au 1er controle"),
        ("Taux de defaut critique maximum", 0, "%", 0, 1, "Objectif zero defaut critique"),
        ("Delai max resolution defaut critique", 48, "heures", 24, 72, "Escalade Direction si depasse"),
        ("Delai max resolution defaut majeur", 120, "heures", 72, 168, "5 jours ouvres maximum"),
        ("Nb controles par artisan / mois (min)", 10, "Nb", 5, 20, "Frequence minimale pour suivi qualite"),
        ("Heures formation qualite / an (min)", 40, "heures", 24, 80, "Maintien des competences"),
        ("DeltaE max cuir — ecart colorimetrique", 1.0, "DeltaE", 0.5, 1.5, "Seuil stricte pour cuir pleine fleur"),
        ("DeltaE max textile", 1.5, "DeltaE", 1.0, 2.0, "Seuil standard textiles"),
        ("Version referentiel qualite", "v2026.04", "-", "-", "-", "Derniere MAJ Avril 2026"),
    ])
    ref_d = ref_c + 2 + 11 + 2
    section(ws, ref_d, 6, "REF D — CONNEXIONS INTER-AGENTS (Neural Enterprise)")
    en_tetes(ws, ref_d + 1, ["Agent source", "Donnees consommees", "Agent destination", "Donnees produites", "Frequence sync", "Statut"], [24, 32, 24, 32, 16, 12])
    ecrire_donnees(ws, ref_d + 2, [
        ("LuxeTraceability", "ID Lot, Matiere, Fournisseur, Certification origine", "ArtisanalQualityAI", "Score qualite lot, Statut QC, Defauts detectes", "Temps reel", "Actif"),
        ("ArtisanalQualityAI", "Score qualite par lot certifie", "LuxeTraceability", "Certification qualite artisanale — Lot valide / rejete", "Par controle", "Actif"),
        ("ArtisanalQualityAI", "Taux defaut par matiere / fournisseur", "RareMaterialSourcing", "Alerte qualite matiere -> Reevaluation sourcing", "Hebdomadaire", "Planifie"),
        ("ArtisanalQualityAI", "Score qualite lot + serialisation", "AntiCounterfeitSC", "Authenticite qualite artisanale certifiee par lot", "Par lot", "Planifie"),
    ])
    mfc_statut(ws, "F", ref_d + 2, ref_d + 5)
    ws.freeze_panes = "A5"
    print("  OK  Onglet 10_PARAMETRES cree — 4 referentiels complets")

    return wb, nb_al, source_path


def completer_dashboard_et_finaliser():
    wb, nb_al, source_path = finaliser_artisanal_quality()
    ws = wb["1_DASHBOARD"]
    roi_sheet = wb["8_KPI_QUALITE"]
    roi_tr = next(
        (row for row in range(1, roi_sheet.max_row + 1) if roi_sheet.cell(row=row, column=1).value == "TOTAL ROI"),
        roi_sheet.max_row,
    )

    cartes = [
        (1, "FIRST PASS YIELD", NOIR, '=IFERROR(COUNTIF(\'3_CONTROLES\'!K5:K5000,"Conforme")/COUNTIFS(\'3_CONTROLES\'!K5:K5000,"<>En attente"),0)', "0.0%", OR),
        (4, "SCORE QUALITE MOYEN", NOIR, '=IFERROR(AVERAGEIF(\'3_CONTROLES\'!L5:L5000,">0"),0)', "0.0", OR),
        (7, "ALERTES CRITIQUES", ROUGE, f'=COUNTIFS(\'9_ALERTES\'!E5:E{4 + nb_al},"CRITIQUE",\'9_ALERTES\'!M5:M{4 + nb_al},"<>Resolu")', "0", "FF0000"),
        (10, "COUT NON-QUALITE", NOIR, '=SUM(\'4_DEFAUTS\'!M5:M5000)', '#,##0 "EUR"', OR),
    ]
    for col_start, label, bg, formule, fmt, font_clr in cartes:
        ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_start + 2)
        c = ws.cell(row=5, column=col_start, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = AC
        ws.merge_cells(start_row=6, start_column=col_start, end_row=7, end_column=col_start + 2)
        c = ws.cell(row=6, column=col_start, value=formule)
        c.font = Font(name="Calibri", size=28, bold=True, color=font_clr)
        c.alignment = AC
        c.number_format = fmt
    ws.row_dimensions[6].height = 50
    section(ws, 9, 12, "SYNTHESE OPERATIONNELLE")
    en_tetes(ws, 10, ["Indicateur", "Valeur", "Unite"], [36, 18, 12])
    dash_metrics = [
        ("Criteres qualite references", "=COUNTA('2_CRITERES'!A5:A5000)", "Nb"),
        ("Controles realises (periode)", "=COUNTA('3_CONTROLES'!A5:A5000)", "Nb"),
        ("Controles conformes", '=COUNTIF(\'3_CONTROLES\'!K5:K5000,"Conforme")', "Nb"),
        ("Controles non conformes", '=COUNTIF(\'3_CONTROLES\'!K5:K5000,"Non conforme")', "Nb"),
        ("Defauts enregistres", "=COUNTA('4_DEFAUTS'!A5:A5000)", "Nb"),
        ("Defauts critiques ouverts", '=COUNTIFS(\'4_DEFAUTS\'!H5:H5000,"Critique",\'4_DEFAUTS\'!O5:O5000,"En cours")', "Nb"),
        ("Ateliers actifs", '=COUNTIF(\'5_ATELIERS\'!M5:M5000,"Actif")', "Nb"),
        ("Artisans references", "=COUNTA('6_ARTISANS'!A5:A5000)", "Nb"),
        ("Maitres d'Art / MOF", '=COUNTIF(\'6_ARTISANS\'!H5:H5000,"Maitre*")', "Nb"),
        ("Heures formation moyennes / artisan", "=IFERROR(AVERAGE('6_ARTISANS'!M5:M5000),0)", "h/an"),
        ("Alertes qualite actives", f'=COUNTIF(\'9_ALERTES\'!M5:M{4 + nb_al},"<>Resolu")', "Nb"),
        ("ROI annuel net estime", f"='8_KPI_QUALITE'!E{roi_tr}", "EUR"),
    ]
    for idx, (lab, form, unit) in enumerate(dash_metrics, 11):
        ws.cell(row=idx, column=1, value=lab)
        ws.cell(row=idx, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=idx, column=1).border = BF
        c = ws.cell(row=idx, column=2, value=form)
        c.font = Font(name="Calibri", size=12, bold=True, color=OR)
        c.alignment = AC
        c.border = BF
        if unit == "EUR":
            c.number_format = '#,##0 "EUR"'
        ws.cell(row=idx, column=3, value=unit)
        ws.cell(row=idx, column=3).font = Font(name="Calibri", size=10)
        ws.cell(row=idx, column=3).alignment = AC
        ws.cell(row=idx, column=3).border = BF
        if idx % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")

    fp = 11 + len(dash_metrics) + 2
    ws.merge_cells(f"A{fp}:L{fp}")
    ws.cell(row=fp, column=1, value="Agent ArtisanalQualityAI — Neural Enterprise Framework © 2026 — Connexion Agent LuxeTraceability active")
    ws.cell(row=fp, column=1).font = Font(name="Calibri", size=8, italic=True, color=GRIS)
    ws.cell(row=fp, column=1).alignment = AC
    ws.merge_cells(f"A{fp + 1}:L{fp + 1}")
    ws.cell(row=fp + 1, column=1, value="Certifie les lots trackes par LuxeTraceability — Alimente AntiCounterfeitSC & RareMaterialSourcing")
    ws.cell(row=fp + 1, column=1).font = Font(name="Calibri", size=8, italic=True, color=BORDEAUX)
    ws.cell(row=fp + 1, column=1).alignment = AC
    ws.freeze_panes = "A5"
    print("  OK  Onglet 1_DASHBOARD complete — Formules inter-onglets actives")

    ordre = ["1_DASHBOARD", "2_CRITERES", "3_CONTROLES", "4_DEFAUTS", "5_ATELIERS", "6_ARTISANS", "7_RAPPORTS_QC", "8_KPI_QUALITE", "9_ALERTES", "10_PARAMETRES"]
    wb._sheets = [wb[name] for name in ordre if name in wb.sheetnames]
    wb.active = wb.sheetnames.index("1_DASHBOARD")
    wb["10_PARAMETRES"].protection.sheet = True
    wb["10_PARAMETRES"].protection.password = "artiqual2026"
    wb["1_DASHBOARD"].protection.sheet = True
    wb["1_DASHBOARD"].protection.password = "artiqual2026"
    wb.properties.title = "Agent ArtisanalQualityAI — Neural Enterprise Framework"
    wb.properties.subject = "Qualite Artisanale Luxe — Documentation & Controle"
    wb.properties.creator = "Neural Enterprise — AI Agent Framework"
    wb.properties.description = "Agent ArtisanalQualityAI : documentation des criteres de qualite par metier, analyse des rapports de controle qualite, suivi artisans et ateliers, calcul ROI automatise. Version Avril 2026 — Connecte a Agent LuxeTraceability."
    wb.properties.keywords = "Luxe, Qualite, Artisanal, Controle, Maroquinerie, Joaillerie, Horlogerie, Pret-a-Porter, MOF, Maitre d'Art, KPI, Neural Enterprise"
    wb.properties.category = "Qualite — Artisanat — Luxe"

    output_dir = source_path.parent
    output_path = output_dir / "Agent_ArtisanalQualityAI_NeuralEnterprise_v1.0_Avril2026.xlsx"
    try:
        wb.save(output_path)
    except PermissionError:
        output_path = output_dir / f"Agent_ArtisanalQualityAI_NeuralEnterprise_v1.0_Avril2026_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(output_path)

    print("\n" + "=" * 70)
    print(f"  FICHIER GENERE : {output_path}")
    print("=" * 70)
    print(f"\n  {len(wb.sheetnames)} onglets crees :")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"     {i:2d}. {name}")
    print("\n  Formules inter-onglets : ACTIVEES")
    print("  Connexion LuxeTraceability : CONFIGUREE")
    print("  Protection : Dashboard + Parametres")
    print("  Validations donnees : Listes deroulantes actives")
    print("  Mise en forme conditionnelle : Statuts, Scores, Criticite")
    print("  Referentiel qualite : Avril 2026")
    print("=" * 70)
    return output_path


if __name__ == "__main__":
    completer_dashboard_et_finaliser()
