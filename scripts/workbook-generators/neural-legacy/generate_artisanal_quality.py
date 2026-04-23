from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NOIR = "1A1A1A"
OR = "C9A84C"
BORDEAUX = "6B1D2A"
BLANC_CASSE = "F9F6F0"
GRIS = "8C8C8C"
VERT = "2E7D32"
BLEU = "1565C0"
ROUGE = "C62828"
FOND_ROUGE = "FFC7CE"
FOND_JAUNE = "FFEB9C"
FOND_VERT = "C6EFCE"
FOND_BLEU = "D6E4F0"

BF = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)


def titre(ws, t1, t2, mc):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
    c = ws.cell(row=1, column=1, value=t1)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)
    c2 = ws.cell(row=2, column=1, value=t2)
    c2.font = Font(name="Calibri", size=10, italic=True, color=OR)
    c2.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28


def section(ws, row, mc, texte, couleur=BORDEAUX):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1, value=texte)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def en_tetes(ws, row, headers, largeurs=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
        c.alignment = AC
        c.border = BF
    ws.row_dimensions[row].height = 36
    if largeurs:
        for i, w in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def ecrire_donnees(ws, start_row, data, fmt_cols=None):
    for ri, ligne in enumerate(data, start_row):
        for ci, val in enumerate(ligne, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=9)
            c.alignment = AC
            c.border = BF
            if (ri - start_row) % 2 == 0:
                c.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")
            if fmt_cols and ci in fmt_cols:
                c.number_format = fmt_cols[ci]
    return start_row + len(data) - 1


def mfc_statut(ws, col, sr, er):
    p = f"{col}{sr}:{col}{er}"
    paires = [
        ("Conforme", FOND_VERT, "006100"), ("Valide", FOND_VERT, "006100"),
        ("Actif", FOND_VERT, "006100"), ("Excellent", FOND_VERT, "006100"),
        ("Acceptable", FOND_VERT, "006100"),
        ("En attente", FOND_JAUNE, "9C6500"), ("En cours", FOND_JAUNE, "9C6500"),
        ("Conditionnel", FOND_JAUNE, "9C6500"), ("A ameliorer", FOND_JAUNE, "9C6500"),
        ("Non conforme", FOND_ROUGE, "9C0006"), ("Rejete", FOND_ROUGE, "9C0006"),
        ("Critique", FOND_ROUGE, "9C0006"), ("Suspendu", FOND_ROUGE, "9C0006"),
        ("Insuffisant", FOND_ROUGE, "9C0006"),
    ]
    for val, bg, fg in paires:
        ws.conditional_formatting.add(
            p,
            CellIsRule(
                operator="equal",
                formula=[f'"{val}"'],
                fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
                font=Font(color=fg, bold=True),
            ),
        )


def ligne_synthese(ws, row, mc, label="SYNTHESE"):
    ws.cell(row=row, column=1, value=label)
    for ci in range(1, mc + 1):
        c = ws.cell(row=row, column=ci)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
        c.border = BF


def formule_or(ws, row, col, formule, fmt="0.0"):
    c = ws.cell(row=row, column=col, value=formule)
    c.font = Font(name="Calibri", size=11, bold=True, color=OR)
    c.number_format = fmt
    c.alignment = AC
    c.border = BF
    return c


def generer_artisanal_quality():
    wb = Workbook()
    print("\n" + "=" * 70)
    print("  AGENT ARTISANALQUALITYAI — Generation en cours...")
    print("=" * 70 + "\n")

    ws = wb.active
    ws.title = "1_DASHBOARD"
    ws.sheet_properties.tabColor = OR
    titre(
        ws,
        "AGENT ARTISANALQUALITYAI — TABLEAU DE BORD QUALITE",
        "Neural Enterprise Framework — Qualite Artisanale Luxe — Documentation & Controle — Avril 2026",
        12,
    )
    ws.merge_cells("A3:L3")
    c3 = ws.cell(
        row=3,
        column=1,
        value="Donnees temps reel | Maroquinerie • Joaillerie • Horlogerie • Pret-a-Porter | Connexion Agent LuxeTraceability",
    )
    c3.font = Font(name="Calibri", size=8, italic=True, color=GRIS)
    c3.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")
    c3.alignment = AC
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 18
    print("  OK  Onglet 1_DASHBOARD initialise")

    # Onglet 2
    ws = wb.create_sheet("2_CRITERES")
    ws.sheet_properties.tabColor = OR
    titre(
        ws,
        "CRITERES DE QUALITE PAR METIER — AGENT ARTISANALQUALITYAI",
        "Referentiel qualite artisanale — Tolerances, methodes de mesure, seuils — Par maison et metier — Avril 2026",
        14,
    )
    section(ws, 3, 14, "REFERENTIEL DES CRITERES DE QUALITE ARTISANALE")
    h = [
        "ID_Critere", "Metier", "Categorie produit", "Critere qualite",
        "Description detaillee", "Methode de mesure", "Outil / Instrument",
        "Tolerance min", "Tolerance max", "Unite", "Niveau criticite",
        "Reference norme", "Specificite Maison", "Frequence controle",
    ]
    l = [14, 16, 18, 24, 40, 24, 20, 14, 14, 10, 14, 18, 30, 16]
    en_tetes(ws, 4, h, l)
    data = [
        ["CQ-MAR-001", "Maroquinerie", "Sac a main", "Densite de piqure",
         "Nombre de points par cm sur couture sellier — Regularite absolue exigee — Pas de variation >0.5pt/cm",
         "Comptage visuel + loupe 10x", "Loupe binoculaire + regle graduee",
         5, 7, "pts/cm", "CRITIQUE",
         "Norme interne MQ-001", "Point sellier : 5 pts/cm tradition Hermes — 6 pts/cm standard maison", "100% des pieces"],
        ["CQ-MAR-002", "Maroquinerie", "Sac a main", "Symetrie des coutures",
         "Ecart maximal entre coutures symetriques gauche/droite — Mesure sur 5 points de reference",
         "Mesure au pied a coulisse digital", "Pied a coulisse Mitutoyo 0.01mm",
         0, 0.5, "mm", "CRITIQUE",
         "Norme interne MQ-002", "Tolerance 0.3mm pour collections Haute Maroquinerie", "100% des pieces"],
        ["CQ-MAR-003", "Maroquinerie", "Sac a main", "Finition des tranches",
         "Qualite du filetage de tranche — Regularite, brillance, absence de bulles, adherence",
         "Inspection visuelle + test adherence (flexion 90°)", "Loupe 10x + flexion manuelle",
         "Regulier", "Parfait", "Qualitatif", "ELEVE",
         "Norme interne MQ-003", "Tranche teintee a la main — Minimum 4 couches — Polissage buis", "100% des pieces"],
        ["CQ-MAR-004", "Maroquinerie", "Sac a main", "Pose quincaillerie",
         "Alignement et solidite des fermoirs, boucles, rivets — Test d'arrachement — Positionnement symetrique",
         "Test dynamometrique + controle visuel", "Dynamometre + gabarit positionnement",
         15, 999, "kg (arrachement)", "CRITIQUE",
         "Norme EN 17396", "Quincaillerie palladiee — Force arrachement min 15kg — Alignement ±0.3mm", "100% des pieces"],
        ["CQ-MAR-005", "Maroquinerie", "Petite maroquinerie", "Qualite du cuir",
         "Absence de defauts visuels sur parement — Griffures, veines marquees, taches, irregularites grain",
         "Inspection visuelle lumiere rasante 45°", "Table lumineuse + loupe",
         "Grade A", "Grade A+", "Grade", "CRITIQUE",
         "Norme ISO 3378", "Selection parements — Cuir pleine fleur uniquement — Taux de chute accepte 40%", "100% peaux"],
        ["CQ-JOA-001", "Joaillerie", "Bague", "Precision du sertissage",
         "Positionnement pierre dans le serti — Symetrie, hauteur, inclinaison — Pierre ne doit pas bouger sous pression moderee",
         "Inspection loupe 20x + test pression doigt", "Loupe gemmologique 20x + pince brucelles",
         0, 0.02, "mm (jeu)", "CRITIQUE",
         "Norme RJC Standards", "Serti clos : jeu max 0.01mm — Serti griffe : alignement ±0.02mm", "100% des pieces"],
        ["CQ-JOA-002", "Joaillerie", "Bracelet", "Finition de surface metal",
         "Qualite du polissage — Absence de rayures, marques d'outils, porosites — Brillance miroir (poli) ou satine uniforme",
         "Inspection visuelle + mesure rugosite", "Rugosimetre Mitutoyo SJ-310 + loupe 20x",
         "Ra 0.025", "Ra 0.05", "µm (rugosite)", "ELEVE",
         "Norme ISO 4287", "Polissage main minimum 3 passes — Finition avivage rhodie pour or blanc", "100% des pieces"],
        ["CQ-JOA-003", "Joaillerie", "Collier", "Tolerance poids metal",
         "Poids de la piece en metal precieux — Tolerance par rapport au poids theorique du dessin technique",
         "Pesee balance de precision 0.01g", "Balance Mettler Toledo XSR205",
         -2, 2, "% (ecart)", "ELEVE",
         "Norme LBMA", "Poincon de garantie verifie — Titre 750‰ confirme par touchau", "100% des pieces"],
        ["CQ-HOR-001", "Horlogerie", "Montre mecanique", "Precision du mouvement",
         "Ecart de marche journalier — Mesure en 5 positions + 2 temperatures sur 10 jours — Criteres COSC",
         "Test chronometrique 10 jours", "Witschi Chronoscope + stand 5 positions",
         -4, 6, "sec/jour", "CRITIQUE",
         "COSC ISO 3159", "Critere COSC standard — Maison exige -2/+4 sec/jour pour Haute Horlogerie", "100% mouvements"],
        ["CQ-HOR-002", "Horlogerie", "Montre", "Etancheite",
         "Resistance a la pression d'eau — Test de pression selon profondeur annoncee + marge 25%",
         "Test pression eau/air", "Appareil Witschi Proofmaster + chambre pression",
         0, 0, "bar (fuite)", "CRITIQUE",
         "ISO 22810", "Test a 125% de la pression annoncee — 3 cycles — Fuite = 0 tolere", "100% boitiers"],
        ["CQ-HOR-003", "Horlogerie", "Montre", "Finition du mouvement",
         "Qualite des decorations : cotes de Geneve, perlage, anglage, polissage — Evaluation sur 8 criteres",
         "Inspection loupe binoculaire 40x", "Loupe Leica M205 + eclairage rasant",
         "Bon", "Excellent", "Qualitatif (8 criteres)", "ELEVE",
         "Poincon de Geneve / Qualite Fleurier", "Anglage interne obligatoire — Cotes de Geneve sur ponts — Polissage speculaire vis", "100% mouvements"],
        ["CQ-PAP-001", "Pret-a-Porter", "Veste / Blazer", "Solidite des coutures",
         "Resistance a la traction des coutures — Test normalise sur eprouvette — Mesure force de rupture",
         "Test dynamometrique sur eprouvette", "Dynamometre Instron 5944 + machoires tissu",
         80, 999, "N (resistance)", "ELEVE",
         "ISO 13935-1", "Resistance min 80N couture principale — 60N couture decorative", "Par lot (5 pieces)"],
        ["CQ-PAP-002", "Pret-a-Porter", "Robe", "Alignement droit-fil",
         "Ecart du droit-fil tissu par rapport a l'axe de symetrie du vetement — Mesure sur patron et piece finie",
         "Mesure au rapporteur + laser", "Laser d'alignement + rapporteur gradue",
         0, 1, "° (degres)", "ELEVE",
         "Norme interne PAP-002", "Tolerance 0.5° pour soie — 1° pour laine — Tissus imprimes : raccord motif obligatoire", "100% des pieces"],
        ["CQ-PAP-003", "Pret-a-Porter", "Chemise", "Regularite des boutonnieres",
         "Taille, positionnement et finition des boutonnieres — Regularite de l'espacement — Solidite de la bride",
         "Mesure pied a coulisse + test boutonnage 50 cycles", "Pied a coulisse + mannequin d'essai",
         0, 0.5, "mm (ecart taille)", "MOYEN",
         "Norme interne PAP-003", "Boutonniere a la main pour Haute Couture — Machine pour PAP standard — Points de renfort", "100% pieces HC / lot PAP"],
        ["CQ-GEN-001", "Tous metiers", "Tous produits", "Conformite colorimetrique",
         "Coherence couleur entre pieces d'un meme lot et par rapport au standard colorimetrique — ΔE* ≤ 1.5",
         "Spectrophotometre", "X-Rite Ci7800 + cabine lumiere D65",
         0, 1.5, "ΔE* (ecart)", "ELEVE",
         "ISO 105-J03", "ΔE* ≤ 1.0 pour cuir — ΔE* ≤ 1.5 pour textile — Validation sous D65 + TL84", "Par lot + standard"],
        ["CQ-GEN-002", "Tous metiers", "Tous produits", "Test de solidite des teintures",
         "Resistance de la teinture au frottement sec et humide — a la lumiere — a la transpiration",
         "Test Crockmeter (frottement) + test lumiere xenon", "Crockmeter James Heal + Xenotest",
         4, 5, "Niveau (echelle 1-5)", "ELEVE",
         "ISO 105-X12 / ISO 105-B02", "Note min 4 frottement sec — Note min 3-4 frottement humide — 6 lumiere", "Par lot (3 eprouvettes)"],
    ]
    nb_crit = len(data)
    ecrire_donnees(ws, 5, data)
    mfc_statut(ws, "K", 5, 4 + nb_crit)
    ws.conditional_formatting.add(
        f"K5:K{4 + nb_crit}",
        CellIsRule(operator="equal", formula=['"CRITIQUE"'],
                   fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True)),
    )
    ws.conditional_formatting.add(
        f"K5:K{4 + nb_crit}",
        CellIsRule(operator="equal", formula=['"ELEVE"'],
                   fill=PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True)),
    )
    ws.conditional_formatting.add(
        f"K5:K{4 + nb_crit}",
        CellIsRule(operator="equal", formula=['"MOYEN"'],
                   fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid"),
                   font=Font(color="9C6500")),
    )
    tr = 4 + nb_crit + 1
    ligne_synthese(ws, tr, 14, "SYNTHESE CRITERES")
    ws.cell(row=tr, column=2, value="Total criteres :")
    formule_or(ws, tr, 3, f"=COUNTA(A5:A{4 + nb_crit})", "0")
    ws.cell(row=tr, column=5, value="CRITIQUES :")
    formule_or(ws, tr, 6, f'=COUNTIF(K5:K{4 + nb_crit},"CRITIQUE")', "0")
    ws.cell(row=tr, column=7, value="ELEVES :")
    formule_or(ws, tr, 8, f'=COUNTIF(K5:K{4 + nb_crit},"ELEVE")', "0")
    dv = DataValidation(type="list", formula1='"CRITIQUE,ELEVE,MOYEN,FAIBLE"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("K5:K5000")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{4 + nb_crit}"
    print("  OK  Onglet 2_CRITERES cree — 16 criteres qualite references")

    # Onglet 3
    ws = wb.create_sheet("3_CONTROLES")
    ws.sheet_properties.tabColor = VERT
    titre(
        ws,
        "CONTROLES QUALITE — AGENT ARTISANALQUALITYAI",
        "Registre des inspections et resultats — Tracabilite piece par piece — Avril 2026",
        16,
    )
    section(ws, 3, 16, "REGISTRE DES CONTROLES QUALITE")
    h = [
        "ID_Controle", "Date", "Heure", "Metier", "Categorie produit",
        "Ref. produit / Lot", "Lot tracabilite (LuxeTrace)", "Critere(s) controle(s)",
        "Artisan", "Atelier", "Resultat", "Score (/100)", "Defauts constates",
        "Action corrective", "Controleur", "Observations",
    ]
    l = [14, 12, 8, 16, 18, 20, 22, 24, 20, 20, 14, 12, 30, 28, 18, 34]
    en_tetes(ws, 4, h, l)
    data = [
        ["CTL-2026-0001", "2026-04-01", "09:15", "Maroquinerie", "Sac a main",
         "SAC-KE-35-NOIR-001", "LOT-2026-0001",
         "CQ-MAR-001 + CQ-MAR-002 + CQ-MAR-003", "Jean-Pierre Moreau", "Atelier Maroquinerie Paris",
         "Conforme", 97, "Aucun", "N/A", "Sophie Laurent",
         "Piece d'exception — Point sellier 5.5pts/cm — Symetrie parfaite"],
        ["CTL-2026-0002", "2026-04-01", "10:30", "Maroquinerie", "Sac a main",
         "SAC-KE-35-GOLD-002", "LOT-2026-0001",
         "CQ-MAR-001 + CQ-MAR-002 + CQ-MAR-004", "Marie Dubois", "Atelier Maroquinerie Paris",
         "Conforme", 94, "Obs. mineure : legere irregularite tranche coin superieur droit",
         "Reprise finition tranche — Couche supplementaire", "Sophie Laurent",
         "Piece validee apres reprise — Score releve de 91 a 94"],
        ["CTL-2026-0003", "2026-04-02", "08:45", "Maroquinerie", "Sac a main",
         "SAC-BI-28-CAMEL-003", "LOT-2026-0002",
         "CQ-MAR-001 + CQ-MAR-005", "Antoine Lefevre", "Atelier Maroquinerie Paris",
         "Non conforme", 62, "Defaut cuir : griffure 8mm visible sur parement avant — Densite piqure irreguliere (4.2 a 5.8 pts/cm)",
         "Remplacement parement + reprise couture complete", "Sophie Laurent",
         "Rejet initial — Cuir retourne au tri — Nouveau parement selectionne"],
        ["CTL-2026-0004", "2026-04-02", "14:00", "Joaillerie", "Bague",
         "BAG-SOL-OR18-DIA-001", "LOT-2026-0004",
         "CQ-JOA-001 + CQ-JOA-002 + CQ-JOA-003", "Maitre Francois Petit", "Atelier Haute Joaillerie Vendome",
         "Conforme", 99, "Aucun", "N/A", "Alain Mercier",
         "Piece Haute Joaillerie — Sertissage diamant 1.52ct D/IF impeccable — Polissage miroir parfait"],
        ["CTL-2026-0005", "2026-04-03", "09:00", "Joaillerie", "Bracelet",
         "BRA-MAN-OR18-007", "LOT-2026-0004",
         "CQ-JOA-002 + CQ-JOA-003", "Elise Durand", "Atelier Haute Joaillerie Vendome",
         "Conditionnel", 82, "Micro-rayure outil visible sur interieur maille — Poids -2.3% vs theorique",
         "Re-polissage interieur maillon — Validation poids dans tolerance", "Alain Mercier",
         "Conditionnel : re-polissage suffisant — Poids dans tolerance (-2.3% < -2% frontiere)"],
        ["CTL-2026-0006", "2026-04-03", "11:30", "Horlogerie", "Montre mecanique",
         "MON-ETER-AUTO-001", "N/A (mouvement manufacture)",
         "CQ-HOR-001 + CQ-HOR-002", "Maitre Claude Bonnard", "Atelier Horlogerie Geneve",
         "Conforme", 98, "Aucun", "N/A", "Pierre Fontaine",
         "Mouvement manufacture — Marche -1.8s/j (cible -2/+4) — Etancheite 10 bar confirmee"],
        ["CTL-2026-0007", "2026-04-04", "08:30", "Horlogerie", "Montre",
         "MON-CLAS-QTZ-015", "N/A",
         "CQ-HOR-002 + CQ-HOR-003", "Isabelle Martin", "Atelier Horlogerie Geneve",
         "Conforme", 91, "Obs. mineure : perlage pont legerement irregulier zone 3 — Dans tolerance",
         "N/A — Observation documentee", "Pierre Fontaine",
         "Perlage acceptable selon grille — Etancheite 5 bar OK — Mouvement quartz conforme"],
        ["CTL-2026-0008", "2026-04-04", "14:15", "Pret-a-Porter", "Veste / Blazer",
         "VES-SM-CASH-NAVY-001", "LOT-2026-0007",
         "CQ-PAP-001 + CQ-PAP-002 + CQ-GEN-001", "Nathalie Bernard", "Atelier PAP Milan",
         "Conforme", 95, "Aucun", "N/A", "Marco Ricci",
         "Cachemire Gobi — Coutures 120N — Droit-fil 0.3° — ΔE* 0.8 vs standard — Excellent"],
        ["CTL-2026-0009", "2026-04-05", "09:45", "Pret-a-Porter", "Robe",
         "ROB-SOI-IMP-FLOR-003", "LOT-2026-0009",
         "CQ-PAP-002 + CQ-GEN-001 + CQ-GEN-002", "Catherine Moreau", "Atelier PAP Milan",
         "Non conforme", 58, "Raccord motif imprime decale 3mm dos — ΔE* 2.1 (> seuil 1.5) — Solidite frottement humide : 2/5",
         "Recoupe panneaux dos — Reteinture lot tissu — Lot bloque", "Marco Ricci",
         "Triple non-conformite — Lot soie bloque — Alerte qualite fournisseur Seteria Bianchi"],
        ["CTL-2026-0010", "2026-04-07", "10:00", "Maroquinerie", "Petite maroquinerie",
         "PLM-ZIP-CUI-BURG-012", "LOT-2026-0001",
         "CQ-MAR-005 + CQ-MAR-003 + CQ-GEN-001", "Jean-Pierre Moreau", "Atelier Maroquinerie Paris",
         "Conforme", 93, "Obs. mineure : legere variation teinte cuir entre 2 panneaux (ΔE* 1.1 — dans tolerance)",
         "N/A — Variation naturelle cuir documentee", "Sophie Laurent",
         "Variation chromatique naturelle du cuir pleine fleur — Dans tolerance maison (ΔE* ≤ 1.0 cuir strict)"],
        ["CTL-2026-0011", "2026-04-08", "11:00", "Joaillerie", "Collier",
         "COL-RIV-OR-SAP-001", "LOT-2026-0005",
         "CQ-JOA-001 + CQ-JOA-002", "Maitre Francois Petit", "Atelier Haute Joaillerie Vendome",
         "En attente", 0, "Controle en cours — Sertissage 47 saphirs en riviere — Etape 32/47",
         "N/A — Controle non finalise", "Alain Mercier",
         "Piece Haute Joaillerie complexe — Controle pierre par pierre — Fin estimee 12/04"],
        ["CTL-2026-0012", "2026-04-09", "09:00", "Maroquinerie", "Sac a main",
         "SAC-KE-35-CROC-001", "LOT-2026-0008",
         "CQ-MAR-001 + CQ-MAR-002 + CQ-MAR-003 + CQ-MAR-004 + CQ-MAR-005",
         "Marie Dubois", "Atelier Maroquinerie Paris", "Conforme", 96, "Aucun",
         "N/A", "Sophie Laurent",
         "Piece crocodile — Controle exhaustif 5 criteres — Excellence — Ecailles symetriques validees"],
        ["CTL-2026-0013", "2026-04-10", "15:00", "Pret-a-Porter", "Chemise",
         "CHE-COT-BLA-SLIM-LOT15", "N/A (coton japonais)",
         "CQ-PAP-003 + CQ-GEN-002", "Lot de 15 pieces", "Atelier PAP Milan",
         "Conditionnel", 78, "2 pieces sur 15 : boutonniere #3 irreguliere (ecart 0.7mm > 0.5mm tolerance) — Solidite teinture OK",
         "Reprise boutonniere pieces #7 et #12 — Re-controle apres", "Marco Ricci",
         "13/15 conformes — 2 pieces en reprise — Lot conditionnel"],
        ["CTL-2026-0014", "2026-04-11", "08:00", "Horlogerie", "Montre mecanique",
         "MON-TOUR-CHRONO-001", "N/A (mouvement complication)",
         "CQ-HOR-001 + CQ-HOR-002 + CQ-HOR-003", "Maitre Claude Bonnard", "Atelier Horlogerie Geneve",
         "Conforme", 100, "Aucun — Piece exceptionnelle", "N/A", "Pierre Fontaine",
         "Score 100/100 — Tourbillon — Marche +0.2s/j — Finition Poincon de Geneve — Chef-d'oeuvre"],
    ]
    nb_ctl = len(data)
    ecrire_donnees(ws, 5, data)
    mfc_statut(ws, "K", 5, 4 + nb_ctl)
    ws.conditional_formatting.add(
        f"L5:L{4 + nb_ctl}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color=VERT),
    )
    tr = 4 + nb_ctl + 1
    ligne_synthese(ws, tr, 16, "SYNTHESE CONTROLES")
    ws.cell(row=tr, column=2, value="Total :")
    formule_or(ws, tr, 3, f"=COUNTA(A5:A{4 + nb_ctl})", "0")
    ws.cell(row=tr, column=5, value="Conformes :")
    formule_or(ws, tr, 6, f'=COUNTIF(K5:K{4 + nb_ctl},"Conforme")', "0")
    ws.cell(row=tr, column=7, value="Non conformes :")
    c_nc = ws.cell(row=tr, column=8, value=f'=COUNTIF(K5:K{4 + nb_ctl},"Non conforme")')
    c_nc.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    c_nc.border = BF
    ws.cell(row=tr, column=9, value="Score moyen :")
    formule_or(ws, tr, 10, f'=IFERROR(AVERAGEIF(L5:L{4 + nb_ctl},">0"),0)', "0.0")
    ws.cell(row=tr + 1, column=2, value="First Pass Yield :")
    formule_or(ws, tr + 1, 3, f'=IFERROR(COUNTIF(K5:K{4 + nb_ctl},"Conforme")/COUNTIFS(K5:K{4 + nb_ctl},"<>En attente"),0)', "0.0%")
    ws.cell(row=tr + 1, column=5, value="Taux rejet :")
    formule_or(ws, tr + 1, 6, f'=IFERROR(COUNTIF(K5:K{4 + nb_ctl},"Non conforme")/COUNTIFS(K5:K{4 + nb_ctl},"<>En attente"),0)', "0.0%")
    dv_res = DataValidation(type="list", formula1='"Conforme,Non conforme,Conditionnel,En attente"', allow_blank=False)
    ws.add_data_validation(dv_res)
    dv_res.add("K5:K5000")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:P{4 + nb_ctl}"
    print("  OK  Onglet 3_CONTROLES cree — 14 controles qualite")

    # Onglet 4
    ws = wb.create_sheet("4_DEFAUTS")
    ws.sheet_properties.tabColor = ROUGE
    titre(
        ws,
        "REGISTRE DES DEFAUTS — AGENT ARTISANALQUALITYAI",
        "Classification, analyse causale, actions correctives — Amelioration continue — Avril 2026",
        16,
    )
    section(ws, 3, 16, "REGISTRE DEFAUTS QUALITE ARTISANALE")
    h = [
        "ID_Defaut", "Date detection", "Controle source", "Metier",
        "Produit / Lot", "Type de defaut", "Classification", "Severite",
        "Description detaillee", "Cause racine identifiee", "Action corrective",
        "Action preventive", "Cout impact (€)", "Responsable", "Statut", "Date cloture",
    ]
    l = [14, 12, 16, 14, 22, 22, 16, 12, 40, 34, 34, 34, 14, 18, 12, 12]
    en_tetes(ws, 4, h, l)
    data = [
        ["DEF-2026-001", "2026-04-02", "CTL-2026-0003", "Maroquinerie",
         "SAC-BI-28-CAMEL-003", "Defaut matiere — Griffure cuir", "Esthetique",
         "Majeur", "Griffure lineaire 8mm sur parement avant cuir veau — Visible a l'oeil nu — Origine : selection cuir insuffisante",
         "Peau presentant un defaut en zone B non detecte lors du tri initial — Eclairage poste de tri insuffisant",
         "Remplacement parement complet — Selection nouvelle peau Grade A+",
         "Installation lampe LED lumiere rasante au poste de tri cuir — Formation operateur tri",
         450, "Resp. Atelier Maroquinerie", "Resolu", "2026-04-04"],
        ["DEF-2026-002", "2026-04-02", "CTL-2026-0003", "Maroquinerie",
         "SAC-BI-28-CAMEL-003", "Defaut piqure — Densite irreguliere", "Fonctionnel",
         "Majeur", "Variation densite piqure de 4.2 a 5.8 pts/cm sur couture principale — Norme 5-7 pts/cm mais variation >0.5pt/cm",
         "Tension fil machine mal reglee — Usure guide-fil detectee — Artisan fatigue (heure tardive)",
         "Decousage + reprise couture integrale — Reglage machine — Verification guide-fil",
         "Procedure de verification tension fil toutes les 2h — Rotation artisans en fin de journee",
         380, "Jean-Pierre Moreau (artisan)", "Resolu", "2026-04-05"],
        ["DEF-2026-003", "2026-04-03", "CTL-2026-0005", "Joaillerie",
         "BRA-MAN-OR18-007", "Defaut finition — Micro-rayure", "Esthetique",
         "Mineur", "Micro-rayure outil 2mm sur face interieure maillon bracelet or 18K — Non visible porte — Detectable loupe 20x",
         "Contact accidentel outil gravure lors de manipulation piece voisine — Poste de travail encombre",
         "Re-polissage localise — Avivage rhodie zone impactee",
         "Reorganisation poste travail — Separation physique pieces en cours — Protocole manipulation",
         120, "Elise Durand (artisan)", "Resolu", "2026-04-04"],
        ["DEF-2026-004", "2026-04-05", "CTL-2026-0009", "Pret-a-Porter",
         "ROB-SOI-IMP-FLOR-003", "Defaut matiere — Raccord motif", "Esthetique",
         "Majeur", "Decalage raccord motif imprime floral 3mm sur couture dos — Visible a 1m — Inacceptable pour piece de collection",
         "Erreur placement patron sur tissu imprime — Operateur coupe n'a pas aligne reperes motif",
         "Recoupe panneaux dos avec raccord correct — Piece recoupee dans nouveau metrage",
         "Formation placement patron sur imprimes — Gabarit de raccord obligatoire — Double validation avant coupe",
         580, "Resp. Atelier PAP Milan", "Resolu", "2026-04-08"],
        ["DEF-2026-005", "2026-04-05", "CTL-2026-0009", "Pret-a-Porter",
         "ROB-SOI-IMP-FLOR-003 (lot tissu)", "Defaut teinture — ΔE* hors tolerance", "Matiere premiere",
         "Critique", "ΔE* de 2.1 entre tissu livre et standard colorimetrique — Seuil 1.5 — Lot soie entier potentiellement impacte",
         "Derive bain de teinture chez fournisseur Seteria Bianchi — Lot de teinture #SB-2026-T089",
         "Lot soie bloque — Retour fournisseur pour reteinture — Alerte qualite emise vers fournisseur",
         "Audit teinture chez Seteria Bianchi — Exigence certificat colorimerique par lot — Controle reception systematique",
         2800, "Dir. Qualite + Seteria Bianchi", "En cours", None],
        ["DEF-2026-006", "2026-04-05", "CTL-2026-0009", "Pret-a-Porter",
         "ROB-SOI-IMP-FLOR-003 (lot tissu)", "Defaut teinture — Solidite insuffisante", "Fonctionnel",
         "Critique", "Solidite frottement humide 2/5 (min 3-4 requis) — Risque decoloration au porte — Lot impropre a la commercialisation",
         "Meme cause que DEF-005 — Fixation colorant insuffisante lot teinture #SB-2026-T089",
         "Lot bloque — Impossibilite de reprise — Remplacement lot complet par Seteria Bianchi",
         "Renforcement cahier des charges teinture — Clause penale fournisseur — Controle solidite reception",
         4200, "Dir. Qualite + Dir. Achats", "En cours", None],
        ["DEF-2026-007", "2026-04-10", "CTL-2026-0013", "Pret-a-Porter",
         "CHE-COT-BLA-SLIM #7 et #12", "Defaut confection — Boutonniere irreguliere", "Esthetique",
         "Mineur", "Boutonniere #3 sur 2 pieces (sur 15) : ecart taille 0.7mm vs 0.5mm tolerance — Legere irregularite visuelle",
         "Reglage machine boutonniere legerement decale — Maintenance preventive retardee de 3 jours",
         "Reprise manuelle boutonnieres pieces #7 et #12 — Reglage machine effectue",
         "Maintenance preventive machine boutonniere strictement toutes les 200 pieces — Check-list pre-production",
         60, "Catherine Moreau (operateur)", "Resolu", "2026-04-11"],
        ["DEF-2026-008", "2026-04-01", "CTL-2026-0002", "Maroquinerie",
         "SAC-KE-35-GOLD-002", "Defaut finition — Tranche irreguliere", "Esthetique",
         "Mineur", "Legere irregularite finition tranche coin superieur droit — Visible loupe 10x uniquement — Reprise simple",
         "Application de peinture de tranche legerement inegale — 3eme couche trop rapide (sechage incomplet)",
         "Poncage leger + application couche supplementaire — Polissage buis final",
         "Rappel temps de sechage inter-couches minimum 45min — Fiche de suivi sechage au poste",
         40, "Marie Dubois (artisan)", "Resolu", "2026-04-01"],
    ]
    nb_def = len(data)
    ecrire_donnees(ws, 5, data, {13: '#,##0 €'})
    mfc_statut(ws, "O", 5, 4 + nb_def)
    ws.conditional_formatting.add(
        f"H5:H{4 + nb_def}",
        CellIsRule(operator="equal", formula=['"Critique"'],
                   fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True)),
    )
    ws.conditional_formatting.add(
        f"H5:H{4 + nb_def}",
        CellIsRule(operator="equal", formula=['"Majeur"'],
                   fill=PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True)),
    )
    ws.conditional_formatting.add(
        f"H5:H{4 + nb_def}",
        CellIsRule(operator="equal", formula=['"Mineur"'],
                   fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid"),
                   font=Font(color="9C6500")),
    )
    tr = 4 + nb_def + 1
    ligne_synthese(ws, tr, 16, "SYNTHESE DEFAUTS")
    ws.cell(row=tr, column=2, value="Total defauts :")
    formule_or(ws, tr, 3, f"=COUNTA(A5:A{4 + nb_def})", "0")
    ws.cell(row=tr, column=5, value="Critiques :")
    c_cr = ws.cell(row=tr, column=6, value=f'=COUNTIF(H5:H{4 + nb_def},"Critique")')
    c_cr.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    c_cr.border = BF
    ws.cell(row=tr, column=7, value="Ouverts :")
    formule_or(ws, tr, 8, f'=COUNTIF(O5:O{4 + nb_def},"En cours")', "0")
    ws.cell(row=tr, column=12, value="Cout total :")
    formule_or(ws, tr, 13, f"=SUM(M5:M{4 + nb_def})", '#,##0 €')
    dv_sev = DataValidation(type="list", formula1='"Critique,Majeur,Mineur,Observation"', allow_blank=False)
    ws.add_data_validation(dv_sev)
    dv_sev.add("H5:H5000")
    dv_stat_d = DataValidation(type="list", formula1='"En cours,Resolu,Escalade,Accepte en l''etat"', allow_blank=False)
    ws.add_data_validation(dv_stat_d)
    dv_stat_d.add("O5:O5000")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:P{4 + nb_def}"
    print("  OK  Onglet 4_DEFAUTS cree — 8 defauts documentes")

    # Onglet 5
    ws = wb.create_sheet("5_ATELIERS")
    ws.sheet_properties.tabColor = OR
    titre(
        ws,
        "REGISTRE DES ATELIERS — AGENT ARTISANALQUALITYAI",
        "Ateliers de fabrication — Capacites, certifications, scores qualite — Avril 2026",
        14,
    )
    section(ws, 3, 14, "BASE DES ATELIERS DE FABRICATION")
    h = [
        "ID_Atelier", "Nom de l'atelier", "Ville", "Pays",
        "Metier principal", "Specialite", "Nb artisans",
        "Score qualite moyen", "Taux First Pass Yield",
        "Certifications atelier", "Date dernier audit qualite",
        "Resultat audit", "Statut", "Notes",
    ]
    l = [14, 30, 16, 12, 18, 24, 12, 16, 16, 28, 16, 14, 12, 36]
    en_tetes(ws, 4, h, l)
    data = [
        ["ATL-001", "Atelier Maroquinerie Paris", "Paris 3e", "France", "Maroquinerie", "Sacs a main — Point sellier — Cuirs exotiques", 12, None, None, "ISO 9001 + Label EPV (Entreprise du Patrimoine Vivant)", "2026-03-15", "Conforme", "Actif", "Atelier historique — Excellence artisanale — Formation compagnonnage — Maitre d'art resident"],
        ["ATL-002", "Atelier Haute Joaillerie Vendome", "Paris 1er", "France", "Joaillerie", "Haute Joaillerie — Sertissage — Polissage — Gravure", 8, None, None, "RJC CoP + Label EPV + Poincon de maitre", "2026-02-28", "Conforme", "Actif", "Place Vendome — Maitres joailliers — Pieces uniques et haute joaillerie — Coffre-fort classe 6"],
        ["ATL-003", "Atelier Horlogerie Geneve", "Geneve", "Suisse", "Horlogerie", "Mouvements manufacture — Complications — Habillage", 6, None, None, "Poincon de Geneve + COSC + ISO 9001", "2026-01-20", "Conforme", "Actif", "Manufacture horlogere — Poincon de Geneve — Tourbillons et complications"],
        ["ATL-004", "Atelier PAP Milan", "Milan", "Italie", "Pret-a-Porter", "Confection Femme — Tailleur — Couture flou", 18, None, None, "ISO 9001 + Camera Nazionale della Moda", "2026-03-01", "Conforme", "Actif", "Atelier italien — Couture et tailleur — Expertise textiles nobles — 2 modelistes seniors"],
        ["ATL-005", "Atelier Haute Couture Paris", "Paris 8e", "France", "Haute Couture", "Flou — Tailleur — Broderie — Plissage", 22, None, None, "Label Haute Couture (Federation) + EPV", "2026-02-15", "Conforme", "Actif", "Maison membre Chambre Syndicale Haute Couture — Ateliers flou et tailleur separes"],
        ["ATL-006", "Atelier Ganterie Millau", "Millau", "France", "Maroquinerie", "Gants de luxe — Coupe, couture, finition", 5, None, None, "Label EPV + IGP Gant de Millau", "2025-11-15", "Conforme", "Actif", "Savoir-faire centenaire — Coupe a la main — Derniere ganterie de tradition en France"],
        ["ATL-007", "Atelier Souliers Florence", "Florence", "Italie", "Maroquinerie", "Chaussures homme/femme — Cousu Blake — Cousu Goodyear", 10, None, None, "ISO 9001 + Consorzio Vera Pelle Italiana", "2026-01-10", "Conforme", "Actif", "Bottier florentin — Blake et Goodyear — Formes sur mesure disponibles"],
    ]
    nb_atl = len(data)
    ecrire_donnees(ws, 5, data)
    mfc_statut(ws, "L", 5, 4 + nb_atl)
    mfc_statut(ws, "M", 5, 4 + nb_atl)
    for r in range(5, 5 + nb_atl):
        ws.cell(row=r, column=8).value = f'=IFERROR(AVERAGEIFS(\'3_CONTROLES\'!L$5:L$5000,\'3_CONTROLES\'!J$5:J$5000,B{r},\'3_CONTROLES\'!L$5:L$5000,\">0\"),\"-\")'
        ws.cell(row=r, column=8).font = Font(name="Calibri", size=10, bold=True, color=OR)
        ws.cell(row=r, column=8).number_format = "0.0"
        ws.cell(row=r, column=9).value = f'=IFERROR(COUNTIFS(\'3_CONTROLES\'!J$5:J$5000,B{r},\'3_CONTROLES\'!K$5:K$5000,\"Conforme\")/COUNTIFS(\'3_CONTROLES\'!J$5:J$5000,B{r},\'3_CONTROLES\'!K$5:K$5000,\"<>En attente\"),\"-\")'
        ws.cell(row=r, column=9).font = Font(name="Calibri", size=10, bold=True, color=OR)
        ws.cell(row=r, column=9).number_format = "0.0%"
    ws.conditional_formatting.add(f"H5:H{4 + nb_atl}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color=VERT))
    tr = 4 + nb_atl + 1
    ligne_synthese(ws, tr, 14, "SYNTHESE ATELIERS")
    ws.cell(row=tr, column=2, value="Total ateliers :")
    formule_or(ws, tr, 3, f"=COUNTA(A5:A{4 + nb_atl})", "0")
    ws.cell(row=tr, column=6, value="Total artisans :")
    formule_or(ws, tr, 7, f"=SUM(G5:G{4 + nb_atl})", "0")
    ws.cell(row=tr, column=8, value="Score global :")
    formule_or(ws, tr, 9, f"=IFERROR(AVERAGE(H5:H{4 + nb_atl}),0)", "0.0")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{4 + nb_atl}"
    print("  OK  Onglet 5_ATELIERS cree — 7 ateliers references")

    # Onglet 6
    ws = wb.create_sheet("6_ARTISANS")
    ws.sheet_properties.tabColor = OR
    titre(ws, "REGISTRE DES ARTISANS — AGENT ARTISANALQUALITYAI", "Competences, certifications, evaluations qualite, formation — Capital humain artisanal — Avril 2026", 16)
    section(ws, 3, 16, "BASE DES ARTISANS ET COMPETENCES")
    h = ["ID_Artisan", "Nom", "Prenom", "Atelier", "Metier", "Specialite principale", "Annees experience", "Niveau maitrise", "Score qualite moyen", "Taux defaut (%)", "Certifications / Distinctions", "Derniere evaluation", "Heures formation (12 mois)", "Besoins formation identifies", "Statut", "Notes"]
    l = [14, 16, 14, 28, 16, 26, 12, 14, 14, 12, 30, 14, 14, 30, 12, 34]
    en_tetes(ws, 4, h, l)
    data = [
        ["ART-001", "Moreau", "Jean-Pierre", "Atelier Maroquinerie Paris", "Maroquinerie", "Point sellier — Sacs structures — Cuirs exotiques", 28, "Maitre d'art", None, None, "Maitre d'Art (2019) + MOF Maroquinier (2015) + EPV", "2026-03-15", 24, "Formation nouvelles colles ecologiques — Cuirs alternatifs", "Actif", "Maitre d'art — Formateur compagnons — 28 ans de maison — Referent cuirs exotiques"],
        ["ART-002", "Dubois", "Marie", "Atelier Maroquinerie Paris", "Maroquinerie", "Couture sellier — Petite maroquinerie — Finition tranches", 15, "Expert", None, None, "MOF Maroquinier candidat 2027 + CQP Maroquinerie Luxe", "2026-03-15", 32, "Perfectionnement cuirs exotiques — Preparation MOF 2027", "Actif", "Candidate MOF 2027 — Excellente — En progression rapide sur exotiques"],
        ["ART-003", "Lefevre", "Antoine", "Atelier Maroquinerie Paris", "Maroquinerie", "Coupe — Patronage — Sacs souples", 8, "Confirme", None, None, "CQP Maroquinerie Luxe + Formation interne avancee", "2026-03-15", 48, "Renforcement controle visuel cuir — Regularite piqure", "Actif", "Defauts DEF-001/002 — Plan amelioration en cours — Formation renforcee tri cuir"],
        ["ART-004", "Petit", "Francois", "Atelier Haute Joaillerie Vendome", "Joaillerie", "Sertissage — Haute Joaillerie — Pierres de centre", 32, "Maitre joaillier", None, None, "Maitre d'Art (2017) + MOF Joaillier (2011) + Poincon de maitre", "2026-02-28", 16, "Nouvelles techniques CAO 3D — Impression cire perdue avancee", "Actif", "Plus haut niveau expertise — Pieces de Haute Joaillerie exclusivement — Mentor 2 apprentis"],
        ["ART-005", "Durand", "Elise", "Atelier Haute Joaillerie Vendome", "Joaillerie", "Polissage — Finition — Rhodiage", 12, "Expert", None, None, "CQP Joaillerie Luxe + Formation Cartier Academy", "2026-02-28", 28, "Perfectionnement finitions pierres de couleur — Techniques laser", "Actif", "Excellente polisseuse — Obs. mineure DEF-003 — Plan attention renforce manipulation"],
        ["ART-006", "Bonnard", "Claude", "Atelier Horlogerie Geneve", "Horlogerie", "Assemblage mouvements — Complications — Tourbillons", 35, "Maitre horloger", None, None, "Maitre Horloger WOSTEP + Poincon de Geneve + Brevet federal CH", "2026-01-20", 12, "Nouvelles certifications silicium — Composants haute frequence", "Actif", "Maitre horloger referent — Score 100/100 sur tourbillon — 35 ans d'excellence"],
        ["ART-007", "Martin", "Isabelle", "Atelier Horlogerie Geneve", "Horlogerie", "Habillage — Decoration — Finition mouvements", 10, "Confirme", None, None, "CFC Horlogere + Formation Geneve Watchmaking School", "2026-01-20", 40, "Anglage avance — Perlage haute precision — Preparation Poincon de Geneve", "Actif", "En progression — Observation perlage (CTL-007) — Formation anglage en cours"],
        ["ART-008", "Bernard", "Nathalie", "Atelier PAP Milan", "Pret-a-Porter", "Tailleur femme — Cachemire — Couture structuree", 20, "Expert", None, None, "Diploma Accademia del Lusso + Maestro Sartoriale", "2026-03-01", 20, "Nouveaux textiles techniques — Couture collee — Hybrides", "Actif", "Experte tailleur italien — Specialiste cachemire — Resultats excellents"],
        ["ART-009", "Moreau", "Catherine", "Atelier PAP Milan", "Pret-a-Porter", "Couture flou — Soie — Imprimes — Boutonnieres", 7, "Confirme", None, None, "CAP Couture + Formation interne avancee soie", "2026-03-01", 52, "Perfectionnement placement patron imprimes — Maintenance machine", "Actif", "Defauts DEF-004 + DEF-007 — Plan formation renforce — Progression attendue"],
        ["ART-010", "Ricci", "Marco", "Atelier PAP Milan", "Pret-a-Porter", "Controle qualite — Chef d'atelier adjoint", 18, "Expert", None, None, "CQP Industrie Mode Luxe + Formation AFNOR Qualite", "2026-03-01", 36, "Formation auditeur interne ISO 9001 — Lean Manufacturing luxe", "Actif", "Controleur qualite principal Milan — Fiable et rigoureux — Candidat chef d'atelier"],
    ]
    nb_art = len(data)
    ecrire_donnees(ws, 5, data)
    mfc_statut(ws, "O", 5, 4 + nb_art)
    for r in range(5, 5 + nb_art):
        ws.cell(row=r, column=9).value = f'=IFERROR(AVERAGEIFS(\'3_CONTROLES\'!L$5:L$5000,\'3_CONTROLES\'!I$5:I$5000,\"*\"&C{r}&\"*\",\'3_CONTROLES\'!L$5:L$5000,\">0\"),\"-\")'
        ws.cell(row=r, column=9).font = Font(name="Calibri", size=10, bold=True, color=OR)
        ws.cell(row=r, column=9).number_format = "0.0"
        ws.cell(row=r, column=10).value = f'=IFERROR(COUNTIFS(\'4_DEFAUTS\'!N$5:N$5000,\"*\"&C{r}&\"*\")/COUNTIFS(\'3_CONTROLES\'!I$5:I$5000,\"*\"&C{r}&\"*\",\'3_CONTROLES\'!K$5:K$5000,\"<>En attente\")*100,0)'
        ws.cell(row=r, column=10).number_format = "0.0"
    ws.conditional_formatting.add(f"I5:I{4 + nb_art}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color=VERT))
    for val, fill_color, font_color in [("Maitre d'art", FOND_VERT, "006100"), ("Maitre joaillier", FOND_VERT, "006100"), ("Maitre horloger", FOND_VERT, "006100"), ("Expert", FOND_BLEU, BLEU), ("Confirme", FOND_JAUNE, "9C6500")]:
        ws.conditional_formatting.add(f"H5:H{4 + nb_art}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"), font=Font(color=font_color, bold=True)))
    tr = 4 + nb_art + 1
    ligne_synthese(ws, tr, 16, "SYNTHESE ARTISANS")
    ws.cell(row=tr, column=2, value="Total artisans :")
    formule_or(ws, tr, 3, f"=COUNTA(A5:A{4 + nb_art})", "0")
    ws.cell(row=tr, column=6, value="Maitres d'art :")
    formule_or(ws, tr, 7, f'=COUNTIF(H5:H{4 + nb_art},"Maitre*")', "0")
    ws.cell(row=tr, column=8, value="Exp. moyenne :")
    formule_or(ws, tr, 9, f"=IFERROR(AVERAGE(G5:G{4 + nb_art}),0)", '0.0" ans"')
    ws.cell(row=tr, column=12, value="Formation moy. :")
    formule_or(ws, tr, 13, f"=IFERROR(AVERAGE(M5:M{4 + nb_art}),0)", '0"h"')
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:P{4 + nb_art}"
    print("  OK  Onglet 6_ARTISANS cree — 10 artisans references")

    # Onglet 7
    ws = wb.create_sheet("7_RAPPORTS_QC")
    ws.sheet_properties.tabColor = BLEU
    titre(ws, "RAPPORTS QUALITE AGREGES — AGENT ARTISANALQUALITYAI", "Synthese mensuelle par atelier et metier — Tendances et analyse — Avril 2026", 14)
    section(ws, 3, 14, "RAPPORTS QUALITE MENSUELS — PERIODE : Q1-Q2 2026")
    h = ["ID_Rapport", "Periode", "Atelier", "Metier", "Nb controles", "Nb conformes", "Nb non conformes", "First Pass Yield (%)", "Score moyen (/100)", "Top 3 defauts", "Tendance vs M-1", "Actions prioritaires", "Responsable", "Statut rapport"]
    l = [14, 14, 28, 16, 12, 12, 14, 16, 14, 36, 14, 36, 20, 14]
    en_tetes(ws, 4, h, l)
    data = [
        ["RPT-2026-Q1-001", "Janvier 2026", "Atelier Maroquinerie Paris", "Maroquinerie", None, None, None, None, None, "1. Finition tranche (3) 2. Symetrie couture (2) 3. Qualite cuir (1)", "Amelioration", "Renforcement controle tri cuir — Formation finition tranche", "Sophie Laurent", "Valide"],
        ["RPT-2026-Q1-002", "Fevrier 2026", "Atelier Maroquinerie Paris", "Maroquinerie", None, None, None, None, None, "1. Qualite cuir (2) 2. Finition tranche (2) 3. Quincaillerie (1)", "Stable", "Installation eclairage LED poste tri — Suivi plan tranche", "Sophie Laurent", "Valide"],
        ["RPT-2026-Q1-003", "Mars 2026", "Atelier Maroquinerie Paris", "Maroquinerie", None, None, None, None, None, "1. Piqure irreguliere (1) 2. Griffure cuir (1)", "Amelioration", "Plan d'action eclairage deploye — Resultats positifs observes", "Sophie Laurent", "Valide"],
        ["RPT-2026-Q1-004", "Janvier 2026", "Atelier Haute Joaillerie Vendome", "Joaillerie", None, None, None, None, None, "1. Micro-rayure polissage (1)", "Stable", "Reorganisation postes de travail — Separation pieces en cours", "Alain Mercier", "Valide"],
        ["RPT-2026-Q1-005", "Fevrier 2026", "Atelier Haute Joaillerie Vendome", "Joaillerie", None, None, None, None, None, "Aucun defaut", "Amelioration", "RAS — Excellence maintenue", "Alain Mercier", "Valide"],
        ["RPT-2026-Q1-006", "Mars 2026", "Atelier Haute Joaillerie Vendome", "Joaillerie", None, None, None, None, None, "Aucun defaut", "Amelioration", "RAS — Preparation piece Haute Joaillerie complexe (47 saphirs)", "Alain Mercier", "Valide"],
        ["RPT-2026-Q1-007", "Q1 2026", "Atelier Horlogerie Geneve", "Horlogerie", None, None, None, None, None, "1. Perlage irregulier (1 obs.)", "Stable", "Formation perlage Isabelle Martin — Preparation Poincon de Geneve", "Pierre Fontaine", "Valide"],
        ["RPT-2026-Q2-008", "Avril 2026 (partiel)", "Atelier PAP Milan", "Pret-a-Porter", None, None, None, None, None, "1. Solidite teinture soie (2 CRIT) 2. Raccord motif (1) 3. Boutonniere (1)", "Degradation", "ALERTE — Probleme fournisseur Seteria Bianchi — Audit urgent teinture", "Marco Ricci", "En cours"],
    ]
    nb_rpt = len(data)
    ecrire_donnees(ws, 5, data)
    for r in range(5, 5 + nb_rpt):
        atelier_cell = f"C{r}"
        ws.cell(row=r, column=5).value = f'=IFERROR(COUNTIFS(\'3_CONTROLES\'!J$5:J$5000,{atelier_cell}),0)'
        ws.cell(row=r, column=5).number_format = "0"
        ws.cell(row=r, column=6).value = f'=IFERROR(COUNTIFS(\'3_CONTROLES\'!J$5:J$5000,{atelier_cell},\'3_CONTROLES\'!K$5:K$5000,"Conforme"),0)'
        ws.cell(row=r, column=6).number_format = "0"
        ws.cell(row=r, column=6).font = Font(name="Calibri", size=9, color="006100")
        ws.cell(row=r, column=7).value = f'=IFERROR(COUNTIFS(\'3_CONTROLES\'!J$5:J$5000,{atelier_cell},\'3_CONTROLES\'!K$5:K$5000,"Non conforme"),0)'
        ws.cell(row=r, column=7).number_format = "0"
        ws.cell(row=r, column=7).font = Font(name="Calibri", size=9, color="9C0006")
        ws.cell(row=r, column=8).value = f'=IFERROR(F{r}/COUNTIFS(\'3_CONTROLES\'!J$5:J$5000,{atelier_cell},\'3_CONTROLES\'!K$5:K$5000,"<>En attente"),0)'
        ws.cell(row=r, column=8).number_format = "0.0%"
        ws.cell(row=r, column=8).font = Font(name="Calibri", size=10, bold=True, color=OR)
        ws.cell(row=r, column=9).value = f'=IFERROR(AVERAGEIFS(\'3_CONTROLES\'!L$5:L$5000,\'3_CONTROLES\'!J$5:J$5000,{atelier_cell},\'3_CONTROLES\'!L$5:L$5000,">0"),0)'
        ws.cell(row=r, column=9).number_format = "0.0"
        ws.cell(row=r, column=9).font = Font(name="Calibri", size=10, bold=True, color=OR)
    mfc_statut(ws, "N", 5, 4 + nb_rpt)
    ws.conditional_formatting.add(f"K5:K{4 + nb_rpt}", FormulaRule(formula=[f'NOT(ISERROR(SEARCH("Amelioration",K5)))'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid"), font=Font(color="006100")))
    ws.conditional_formatting.add(f"K5:K{4 + nb_rpt}", FormulaRule(formula=[f'NOT(ISERROR(SEARCH("Degradation",K5)))'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid"), font=Font(color="9C0006")))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{4 + nb_rpt}"
    print("  OK  Onglet 7_RAPPORTS_QC cree — 8 rapports mensuels")

    # Onglet 8
    ws = wb.create_sheet("8_KPI_QUALITE")
    ws.sheet_properties.tabColor = OR
    titre(ws, "KPI QUALITE & ROI — AGENT ARTISANALQUALITYAI", "Indicateurs de performance qualite artisanale — Calcul ROI automatise — Avril 2026", 12)
    section(ws, 4, 12, "SECTION A — KPIs QUALITE ARTISANALE")
    h_kpi = ["ID_KPI", "Indicateur", "Categorie", "Unite", "Cible", "Valeur actuelle", "% Atteinte", "Tendance", "Source donnees", "Frequence", "Responsable", "Commentaire"]
    l_kpi = [10, 36, 16, 10, 12, 16, 12, 12, 22, 14, 22, 36]
    en_tetes(ws, 5, h_kpi, l_kpi)
    kpis = [
        ["KQ-001", "First Pass Yield global", "Performance", "%", 95, None, None, None, "3_CONTROLES", "Mensuel", "Dir. Qualite", "% pieces conformes des le 1er controle"],
        ["KQ-002", "Score qualite moyen", "Performance", "/100", 92, None, None, None, "3_CONTROLES", "Mensuel", "Dir. Qualite", "Moyenne ponderee tous ateliers"],
        ["KQ-003", "Taux de defauts critiques", "Defauts", "%", 0, None, None, None, "4_DEFAUTS", "Hebdomadaire", "Dir. Qualite", "Objectif zero defaut critique"],
        ["KQ-004", "Cout total non-qualite", "Financier", "EUR/mois", 5000, None, None, None, "4_DEFAUTS", "Mensuel", "Dir. Finance", "Couts reprises + rebuts + retours"],
        ["KQ-005", "Taux de retour client", "Client", "%", 0.5, None, None, None, "Donnees SAV (ext.)", "Mensuel", "Dir. Commercial", "Retours lies a un defaut qualite"],
        ["KQ-006", "Nb artisans Maitre d'Art / MOF", "Capital humain", "Nb", None, None, None, None, "6_ARTISANS", "Annuel", "Dir. RH", "Indicateur excellence savoir-faire"],
        ["KQ-007", "Heures formation qualite / artisan / an", "Formation", "h", 40, None, None, None, "6_ARTISANS", "Trimestriel", "Dir. RH + Qualite", "Minimum 40h/an pour maintien excellence"],
        ["KQ-008", "Taux de certification ateliers", "Certification", "%", 100, None, None, None, "5_ATELIERS", "Annuel", "Dir. Qualite", "Tous ateliers ISO 9001 ou equivalent"],
        ["KQ-009", "Delai moyen resolution defaut", "Reactivite", "jours", 3, None, None, None, "4_DEFAUTS", "Mensuel", "Dir. Qualite", "De detection a resolution complete"],
        ["KQ-010", "Taux de reprise vs rejet", "Economie", "%", 80, None, None, None, "4_DEFAUTS", "Mensuel", "Dir. Production", "% defauts resolus par reprise (vs rejet piece)"],
    ]
    nb_kpi = len(kpis)
    ecrire_donnees(ws, 6, kpis)
    for row in range(6, 6 + nb_kpi):
        ws.cell(row=row, column=6).font = Font(name="Calibri", size=10, bold=True, color=OR)
    ws.cell(row=6, column=6).value = '=IFERROR(COUNTIF(\'3_CONTROLES\'!K$5:K$5000,"Conforme")/COUNTIFS(\'3_CONTROLES\'!K$5:K$5000,"<>En attente")*100,0)'
    ws.cell(row=7, column=6).value = '=IFERROR(AVERAGEIF(\'3_CONTROLES\'!L$5:L$5000,">0"),0)'
    ws.cell(row=8, column=6).value = '=IFERROR(COUNTIF(\'4_DEFAUTS\'!H$5:H$5000,"Critique")/COUNTA(\'4_DEFAUTS\'!A$5:A$5000)*100,0)'
    ws.cell(row=9, column=6).value = '=IFERROR(SUM(\'4_DEFAUTS\'!M$5:M$5000),0)'
    ws.cell(row=9, column=6).number_format = '#,##0 "EUR"'
    ws.cell(row=11, column=5).value = '=COUNTA(\'6_ARTISANS\'!A$5:A$5000)'
    ws.cell(row=11, column=6).value = '=COUNTIF(\'6_ARTISANS\'!H$5:H$5000,"Maitre*")'
    ws.cell(row=12, column=6).value = '=IFERROR(AVERAGE(\'6_ARTISANS\'!M$5:M$5000),0)'
    ws.cell(row=13, column=6).value = '=IFERROR(COUNTIF(\'5_ATELIERS\'!M$5:M$5000,"Actif")/COUNTA(\'5_ATELIERS\'!A$5:A$5000)*100,0)'
    ws.cell(row=14, column=6).value = 2.5
    ws.cell(row=15, column=6).value = '=IFERROR(COUNTIF(\'4_DEFAUTS\'!O$5:O$5000,"Resolu")/COUNTA(\'4_DEFAUTS\'!A$5:A$5000)*100,0)'
    for r in range(6, 6 + nb_kpi):
        ws.cell(row=r, column=7).value = f'=IFERROR(IF(D{r}="%",IF(E{r}>F{r},F{r}/E{r},(100-F{r})/(100-E{r})),F{r}/E{r}),"-")'
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=7).font = Font(name="Calibri", size=10, bold=True)
    ws.conditional_formatting.add(f"G6:G{5 + nb_kpi}", CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(f"G6:G{5 + nb_kpi}", CellIsRule(operator="between", formula=["0.7", "0.899"], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid"), font=Font(color="9C6500")))
    ws.conditional_formatting.add(f"G6:G{5 + nb_kpi}", CellIsRule(operator="lessThan", formula=["0.7"], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid"), font=Font(color="9C0006")))
    roi_start = 6 + nb_kpi + 3
    section(ws, roi_start, 12, "SECTION B — CALCUL ROI AGENT ARTISANALQUALITYAI")
    roi_h = ["Poste", "Description", "Avant Agent (EUR/an)", "Apres Agent (EUR/an)", "Economie (EUR/an)", "% Reduction", "Temps avant (h/mois)", "Temps apres (h/mois)", "Heures gagnees", "ETP liberes", "Source", "Notes"]
    en_tetes(ws, roi_start + 1, roi_h, l_kpi)
    roi_data = [
        ["Inspection qualite manuelle", "Temps inspecteurs sur controles visuels et mesures", 220000, None, None, None, 360, None, None, None, "Benchmark KPMG 2026", "Automatisation partielle par vision IA"],
        ["Documentation qualite", "Redaction rapports, fiches controle, archivage", 85000, None, None, None, 140, None, None, None, "Interne", "Generation automatique rapports QC"],
        ["Gestion des non-conformites", "Detection, investigation, traitement, suivi", 130000, None, None, None, 180, None, None, None, "Interne", "Workflow automatise + alertes proactives"],
        ["Formation et evaluation", "Suivi competences, plans formation, evaluations", 60000, None, None, None, 80, None, None, None, "DRH", "Tableaux de bord competences automatises"],
        ["Couts de non-qualite evites", "Reprises, rebuts, retours clients reduits", 350000, None, None, None, 0, 0, 0, 0, "Qualite", "Reduction estimee 45% des couts non-qualite"],
        ["Cout Agent IA", "Licence, maintenance, formation utilisateurs", 0, 75000, None, None, 0, 30, None, None, "Devis Neural Enterprise", "Cout annuel tout compris"],
    ]
    nb_roi = len(roi_data)
    rhr = roi_start + 2
    ecrire_donnees(ws, rhr, roi_data, {3: '#,##0 "EUR"', 4: '#,##0 "EUR"', 5: '#,##0 "EUR"', 6: "0.0%"})
    for r in range(rhr, rhr + nb_roi):
        if ws.cell(row=r, column=4).value is None:
            ws.cell(row=r, column=4).value = f"=IFERROR(C{r}*0.30,0)"
            ws.cell(row=r, column=4).number_format = '#,##0 "EUR"'
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        ws.cell(row=r, column=5).number_format = '#,##0 "EUR"'
        ws.cell(row=r, column=5).font = Font(name="Calibri", size=9, bold=True, color="006100")
        ws.cell(row=r, column=6).value = f"=IFERROR(E{r}/C{r},0)"
        ws.cell(row=r, column=6).number_format = "0.0%"
        if ws.cell(row=r, column=8).value is None:
            ws.cell(row=r, column=8).value = f"=IFERROR(G{r}*0.30,0)"
        ws.cell(row=r, column=9).value = f"=G{r}-H{r}"
        ws.cell(row=r, column=9).font = Font(name="Calibri", size=9, bold=True, color="006100")
        ws.cell(row=r, column=10).value = f"=IFERROR(I{r}/151.67,0)"
        ws.cell(row=r, column=10).number_format = "0.00"
    roi_tr = rhr + nb_roi
    ligne_synthese(ws, roi_tr, 12, "TOTAL ROI")
    formule_or(ws, roi_tr, 3, f"=SUM(C{rhr}:C{roi_tr - 1})", '#,##0 "EUR"')
    formule_or(ws, roi_tr, 4, f"=SUM(D{rhr}:D{roi_tr - 1})", '#,##0 "EUR"')
    formule_or(ws, roi_tr, 5, f"=SUM(E{rhr}:E{roi_tr - 1})", '#,##0 "EUR"')
    ws.cell(row=roi_tr, column=6, value=f"=IFERROR(E{roi_tr}/C{roi_tr},0)")
    ws.cell(row=roi_tr, column=6).number_format = "0.0%"
    ws.cell(row=roi_tr, column=6).font = Font(name="Calibri", size=11, bold=True, color=OR)
    formule_or(ws, roi_tr, 7, f"=SUM(G{rhr}:G{roi_tr - 1})", "0")
    formule_or(ws, roi_tr, 8, f"=SUM(H{rhr}:H{roi_tr - 1})", "0")
    formule_or(ws, roi_tr, 9, f"=SUM(I{rhr}:I{roi_tr - 1})", "0")
    formule_or(ws, roi_tr, 10, f"=SUM(J{rhr}:J{roi_tr - 1})", "0.00")
    synth_r = roi_tr + 3
    section(ws, synth_r, 6, "SYNTHESE ROI", VERT)
    synth_items = [("ROI annuel net", f"=E{roi_tr}", '#,##0 "EUR"'), ("ROI en %", f"=IFERROR(E{roi_tr}/D{roi_tr},0)", "0.0%"), ("Heures economisees / mois", f"=I{roi_tr}", "0"), ("ETP liberes", f"=J{roi_tr}", "0.00"), ("Payback (mois)", f"=IFERROR(D{roi_tr}/(E{roi_tr}/12),0)", "0.0")]
    for idx, (lab, form, fmt) in enumerate(synth_items):
        r = synth_r + 1 + idx
        ws.merge_cells(f"A{r}:C{r}")
        ws.cell(row=r, column=1, value=lab)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=1).border = BF
        ws.merge_cells(f"D{r}:F{r}")
        c = ws.cell(row=r, column=4, value=form)
        c.font = Font(name="Calibri", size=14, bold=True, color=OR)
        c.number_format = fmt
        c.alignment = AC
        c.border = BF
    ws.freeze_panes = "A6"
    print("  OK  Onglet 8_KPI_QUALITE cree — 10 KPIs + ROI complet")

    output_dir = Path(__file__).resolve().parent / "data"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "NEURAL_ArtisanalQualityAI.xlsx"
    try:
        wb.save(output_path)
    except PermissionError:
        fallback_name = f"NEURAL_ArtisanalQualityAI_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = output_dir / fallback_name
        wb.save(output_path)

    print("\n" + "=" * 70)
    print(f"  Fichier genere : {output_path}")
    print(f"  Date generation : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70 + "\n")
    return output_path


if __name__ == "__main__":
    generer_artisanal_quality()
