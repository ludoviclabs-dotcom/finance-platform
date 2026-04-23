from datetime import datetime
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_DIR = Path(os.environ.get("CCM_OUTPUT_DIR", r"C:\Users\Ludo\Desktop\Corporate finance"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def next_available_path(directory: Path, filename: str) -> Path:
    candidate = directory / filename
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    index = 1
    while True:
        versioned = directory / f"{stem}_{index:02d}{suffix}"
        if not versioned.exists():
            return versioned
        index += 1


OUT_FILE = next_available_path(OUTPUT_DIR, "ccm_dashboard.xlsx")

wb = Workbook()
wb.active.title = "00_Home"

sheet_order = [
    "00_Home",
    "01_Parametres",
    "02_Processus",
    "03_Risques",
    "04_Control_Matrix",
    "05_Tests",
    "06_Exceptions",
    "07_KRI",
    "08_Evidences",
    "09_Dashboard",
    "10_CMM_Scoring",
    "11_Import_Workpapers",
]
for name in sheet_order[1:]:
    wb.create_sheet(title=name)

colors = {
    "navy": "0F172A",
    "slate": "334155",
    "blue": "1D4ED8",
    "green": "15803D",
    "soft": "F8FAFC",
    "line": "CBD5E1",
    "white": "FFFFFF",
}

thin = Side(style="thin", color=colors["line"])
header_fill = PatternFill("solid", fgColor=colors["navy"])
section_fill = PatternFill("solid", fgColor=colors["soft"])
subheader_fill = PatternFill("solid", fgColor=colors["slate"])
accent_fill = PatternFill("solid", fgColor=colors["blue"])
red_fill = PatternFill("solid", fgColor="FEE2E2")
orange_fill = PatternFill("solid", fgColor="FFEDD5")
green_fill = PatternFill("solid", fgColor="DCFCE7")

header_font = Font(name="Calibri", size=11, bold=True, color=colors["white"])
title_font = Font(name="Calibri", size=16, bold=True, color=colors["navy"])
subtitle_font = Font(name="Calibri", size=10, italic=True, color=colors["slate"])
text_font = Font(name="Calibri", size=11, color="000000")
link_font = Font(name="Calibri", size=11, color="0000FF", underline="single")

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
right = Alignment(horizontal="right", vertical="center")
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)


def setup_sheet(ws, title, subtitle=None):
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8
    ws.merge_cells("B2:J2")
    ws["B2"] = title
    ws["B2"].font = title_font
    ws["B2"].alignment = left
    if subtitle:
        ws.merge_cells("B3:J3")
        ws["B3"] = subtitle
        ws["B3"].font = subtitle_font
        ws["B3"].alignment = left
    for col in range(2, 25):
        ws.cell(row=2, column=col).fill = section_fill
        if subtitle:
            ws.cell(row=3, column=col).fill = section_fill


def add_table(ws, start_row, start_col, headers, rows, table_name, style="TableStyleMedium2"):
    for col_idx, header in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for row_idx, row_values in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_values, start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = text_font
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = right if isinstance(value, (int, float)) else left

    end_col = start_col + len(headers) - 1
    end_row = start_row + len(rows)
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = f"B{start_row + 1}"
    return {
        "header_row": start_row,
        "first_data_row": start_row + 1,
        "last_data_row": end_row,
    }


def apply_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_border_range(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def apply_equals_cf(ws, range_ref, anchor_cell, value, fill):
    ws.conditional_formatting.add(
        range_ref,
        FormulaRule(formula=[f'EXACT({anchor_cell},"{value}")'], stopIfTrue=False, fill=fill),
    )


setup_sheet(wb["01_Parametres"], "Paramètres CCM", "Listes, seuils, couleurs et règles de fréquence")
param_headers = ["Liste", "Valeur"]
param_rows = [
    ["Frequence", "Mensuelle"],
    ["Frequence", "Trimestrielle"],
    ["Frequence", "Semestrielle"],
    ["Frequence", "Annuelle"],
    ["Criticite", "Critique"],
    ["Criticite", "Haute"],
    ["Criticite", "Moyenne"],
    ["Criticite", "Faible"],
    ["TestResult", "Pass"],
    ["TestResult", "Fail"],
    ["TestResult", "Not Tested"],
    ["ExceptionStatus", "Open"],
    ["ExceptionStatus", "Mitigated"],
    ["ExceptionStatus", "Closed"],
    ["KRIStatus", "Vert"],
    ["KRIStatus", "Orange"],
    ["KRIStatus", "Rouge"],
]
add_table(wb["01_Parametres"], 5, 2, param_headers, param_rows, "tblParams")
apply_widths(wb["01_Parametres"], {"B": 22, "C": 22})

setup_sheet(wb["02_Processus"], "Référentiel Processus", "Base des processus métiers couverts par le CCM")
process_headers = ["process_id", "process_name", "owner", "entity", "family", "criticality"]
process_rows = [
    ["P001", "Order to Cash", "Finance Ops", "HQ", "Finance", "Critique"],
    ["P002", "Procure to Pay", "Procurement", "HQ", "Finance", "Haute"],
    ["P003", "Hire to Retire", "HR Ops", "FR", "HR", "Moyenne"],
    ["P004", "Record to Report", "Controlling", "HQ", "Finance", "Critique"],
]
add_table(wb["02_Processus"], 5, 2, process_headers, process_rows, "tblProcess")
apply_widths(wb["02_Processus"], {"B": 14, "C": 24, "D": 18, "E": 14, "F": 16, "G": 14})

setup_sheet(wb["03_Risques"], "Référentiel Risques", "Cartographie des risques rattachés aux processus")
risk_headers = ["risk_id", "process_id", "risk_name", "category", "impact", "likelihood", "inherent_score"]
risk_rows = [
    ["R001", "P001", "Revenue recognition error", "Financial", 5, 3, "=[@impact]*[@likelihood]"],
    ["R002", "P002", "Unauthorized supplier payment", "Fraud", 5, 2, "=[@impact]*[@likelihood]"],
    ["R003", "P003", "Payroll master data inaccuracy", "Operational", 4, 3, "=[@impact]*[@likelihood]"],
    ["R004", "P004", "Late close adjustments", "Financial", 4, 4, "=[@impact]*[@likelihood]"],
]
risk_meta = add_table(wb["03_Risques"], 5, 2, risk_headers, risk_rows, "tblRisk")
apply_widths(wb["03_Risques"], {"B": 12, "C": 12, "D": 30, "E": 16, "F": 10, "G": 12, "H": 14})
for row in range(risk_meta["first_data_row"], risk_meta["last_data_row"] + 1):
    wb["03_Risques"][f"H{row}"].number_format = "#,##0"

setup_sheet(wb["04_Control_Matrix"], "Control Matrix", "Lien processus -> risque -> contrôle -> fréquence -> KRI")
cm_headers = [
    "control_id",
    "process_id",
    "process_name",
    "risk_id",
    "risk_name",
    "control_name",
    "control_owner",
    "frequency",
    "control_type",
    "kri_id",
    "evidence_expected",
    "last_test_date",
    "next_test_due",
    "latest_result",
    "control_status",
]
cm_rows = [
    [
        "C001",
        "P001",
        '=XLOOKUP([@process_id],tblProcess[process_id],tblProcess[process_name],"")',
        "R001",
        '=XLOOKUP([@risk_id],tblRisk[risk_id],tblRisk[risk_name],"")',
        "Revenue cut-off review",
        "Finance Ops",
        "Mensuelle",
        "Detective",
        "K001",
        "Close checklist",
        '=IFERROR(MAXIFS(tblTests[test_date],tblTests[control_id],[@control_id]),"")',
        '=IF([@last_test_date]="","",EDATE([@last_test_date],1))',
        '=IFERROR(XLOOKUP([@control_id],tblTests[control_id],tblTests[result],"Not Tested",0,-1),"Not Tested")',
        '=IF([@latest_result]="Fail","Failed",IF([@next_test_due]="","To Schedule",IF([@next_test_due]<TODAY(),"Overdue","On Track")))',
    ],
    [
        "C002",
        "P002",
        '=XLOOKUP([@process_id],tblProcess[process_id],tblProcess[process_name],"")',
        "R002",
        '=XLOOKUP([@risk_id],tblRisk[risk_id],tblRisk[risk_name],"")',
        "3-way match review",
        "Procurement",
        "Mensuelle",
        "Preventive",
        "K002",
        "Invoice sample",
        '=IFERROR(MAXIFS(tblTests[test_date],tblTests[control_id],[@control_id]),"")',
        '=IF([@last_test_date]="","",EDATE([@last_test_date],1))',
        '=IFERROR(XLOOKUP([@control_id],tblTests[control_id],tblTests[result],"Not Tested",0,-1),"Not Tested")',
        '=IF([@latest_result]="Fail","Failed",IF([@next_test_due]="","To Schedule",IF([@next_test_due]<TODAY(),"Overdue","On Track")))',
    ],
    [
        "C003",
        "P003",
        '=XLOOKUP([@process_id],tblProcess[process_id],tblProcess[process_name],"")',
        "R003",
        '=XLOOKUP([@risk_id],tblRisk[risk_id],tblRisk[risk_name],"")',
        "Payroll change approval",
        "HR Ops",
        "Mensuelle",
        "Preventive",
        "K003",
        "HR approval form",
        '=IFERROR(MAXIFS(tblTests[test_date],tblTests[control_id],[@control_id]),"")',
        '=IF([@last_test_date]="","",EDATE([@last_test_date],1))',
        '=IFERROR(XLOOKUP([@control_id],tblTests[control_id],tblTests[result],"Not Tested",0,-1),"Not Tested")',
        '=IF([@latest_result]="Fail","Failed",IF([@next_test_due]="","To Schedule",IF([@next_test_due]<TODAY(),"Overdue","On Track")))',
    ],
    [
        "C004",
        "P004",
        '=XLOOKUP([@process_id],tblProcess[process_id],tblProcess[process_name],"")',
        "R004",
        '=XLOOKUP([@risk_id],tblRisk[risk_id],tblRisk[risk_name],"")',
        "Journal entry review",
        "Controlling",
        "Trimestrielle",
        "Detective",
        "K004",
        "JE evidence pack",
        '=IFERROR(MAXIFS(tblTests[test_date],tblTests[control_id],[@control_id]),"")',
        '=IF([@last_test_date]="","",EDATE([@last_test_date],3))',
        '=IFERROR(XLOOKUP([@control_id],tblTests[control_id],tblTests[result],"Not Tested",0,-1),"Not Tested")',
        '=IF([@latest_result]="Fail","Failed",IF([@next_test_due]="","To Schedule",IF([@next_test_due]<TODAY(),"Overdue","On Track")))',
    ],
]
cm_meta = add_table(wb["04_Control_Matrix"], 5, 2, cm_headers, cm_rows, "tblControl")
apply_widths(
    wb["04_Control_Matrix"],
    {"B": 12, "C": 12, "D": 22, "E": 10, "F": 28, "G": 28, "H": 18, "I": 14, "J": 14, "K": 10, "L": 22, "M": 14, "N": 14, "O": 14, "P": 14},
)
for row in range(cm_meta["first_data_row"], cm_meta["last_data_row"] + 1):
    wb["04_Control_Matrix"][f"M{row}"].number_format = "dd/mm/yyyy"
    wb["04_Control_Matrix"][f"N{row}"].number_format = "dd/mm/yyyy"

setup_sheet(wb["05_Tests"], "Résultats de tests", "Import des workpapers et score pass/fail")
test_headers = ["test_id", "control_id", "test_date", "tester", "method", "result", "score", "evidence_id", "finding_ref", "note"]
test_rows = [
    ["T001", "C001", datetime(2026, 1, 31), "Auditeur A", "Sample", "Pass", '=--([@result]="Pass")', "E001", "", "Control operating effectively"],
    ["T002", "C002", datetime(2026, 2, 28), "Auditeur B", "Sample", "Fail", '=--([@result]="Pass")', "E002", "F001", "Late approval detected"],
    ["T003", "C003", datetime(2026, 2, 20), "Auditeur C", "Inquiry", "Pass", '=--([@result]="Pass")', "E003", "", "Approval trace verified"],
    ["T004", "C004", datetime(2026, 3, 31), "Auditeur A", "Sample", "Pass", '=--([@result]="Pass")', "E004", "", "No exception on sample"],
]
test_meta = add_table(wb["05_Tests"], 5, 2, test_headers, test_rows, "tblTests")
apply_widths(wb["05_Tests"], {"B": 12, "C": 12, "D": 14, "E": 14, "F": 14, "G": 12, "H": 8, "I": 12, "J": 12, "K": 30})
for row in range(test_meta["first_data_row"], test_meta["last_data_row"] + 1):
    wb["05_Tests"][f"D{row}"].number_format = "dd/mm/yyyy"
    wb["05_Tests"][f"H{row}"].number_format = "#,##0"

setup_sheet(wb["06_Exceptions"], "Dérogations & exceptions", "Historique et suivi des écarts ouverts")
exc_headers = ["exception_id", "control_id", "open_date", "severity", "root_cause", "action_plan", "due_date", "status", "days_open", "overdue_flag"]
exc_rows = [
    [
        "X001",
        "C002",
        datetime(2026, 2, 28),
        "Haute",
        "Approval missing",
        "Reinforce maker-checker",
        datetime(2026, 4, 15),
        "Open",
        "=TODAY()-[@open_date]",
        '=--AND([@status]<>"Closed",[@due_date]<TODAY())',
    ],
    [
        "X002",
        "C004",
        datetime(2026, 3, 31),
        "Moyenne",
        "Late evidence storage",
        "Automate archive upload",
        datetime(2026, 5, 15),
        "Mitigated",
        "=TODAY()-[@open_date]",
        '=--AND([@status]<>"Closed",[@due_date]<TODAY())',
    ],
]
exc_meta = add_table(wb["06_Exceptions"], 5, 2, exc_headers, exc_rows, "tblExceptions")
apply_widths(wb["06_Exceptions"], {"B": 14, "C": 12, "D": 14, "E": 12, "F": 22, "G": 28, "H": 14, "I": 14, "J": 12, "K": 12})
for row in range(exc_meta["first_data_row"], exc_meta["last_data_row"] + 1):
    wb["06_Exceptions"][f"D{row}"].number_format = "dd/mm/yyyy"
    wb["06_Exceptions"][f"H{row}"].number_format = "dd/mm/yyyy"
    wb["06_Exceptions"][f"J{row}"].number_format = "#,##0"
    wb["06_Exceptions"][f"K{row}"].number_format = "#,##0"

setup_sheet(wb["07_KRI"], "KRI Monitoring", "Seuils, tendances et alertes sur indicateurs de risque")
kri_headers = ["kri_id", "process_id", "kri_name", "metric_date", "value", "green_max", "amber_max", "status", "trend_vs_prev"]
kri_rows = [
    ["K001", "P001", "Late close entries", datetime(2026, 3, 31), 2, 2, 4, '=IF([@value]>[@amber_max],"Rouge",IF([@value]>[@green_max],"Orange","Vert"))', 0.00],
    ["K002", "P002", "Blocked supplier payments", datetime(2026, 3, 31), 5, 3, 5, '=IF([@value]>[@amber_max],"Rouge",IF([@value]>[@green_max],"Orange","Vert"))', 0.25],
    ["K003", "P003", "Payroll change backlog", datetime(2026, 3, 31), 1, 1, 2, '=IF([@value]>[@amber_max],"Rouge",IF([@value]>[@green_max],"Orange","Vert"))', -0.20],
    ["K004", "P004", "Manual JE above threshold", datetime(2026, 3, 31), 7, 4, 6, '=IF([@value]>[@amber_max],"Rouge",IF([@value]>[@green_max],"Orange","Vert"))', 0.40],
]
kri_meta = add_table(wb["07_KRI"], 5, 2, kri_headers, kri_rows, "tblKRI")
apply_widths(wb["07_KRI"], {"B": 10, "C": 12, "D": 28, "E": 14, "F": 10, "G": 10, "H": 10, "I": 10, "J": 14})
for row in range(kri_meta["first_data_row"], kri_meta["last_data_row"] + 1):
    wb["07_KRI"][f"E{row}"].number_format = "dd/mm/yyyy"
    wb["07_KRI"][f"J{row}"].number_format = "0.0%"

setup_sheet(wb["08_Evidences"], "Index des preuves", "Référentiel des preuves liées aux tests et findings")
ev_headers = ["evidence_id", "test_id", "control_id", "evidence_type", "file_link", "date_received", "validity_status"]
ev_rows = [
    ["E001", "T001", "C001", "Checklist", "https://sharepoint.local/e001", datetime(2026, 1, 31), "Valid"],
    ["E002", "T002", "C002", "Invoice pack", "https://sharepoint.local/e002", datetime(2026, 2, 28), "Valid"],
    ["E003", "T003", "C003", "Approval form", "https://sharepoint.local/e003", datetime(2026, 2, 20), "Valid"],
    ["E004", "T004", "C004", "JE review file", "https://sharepoint.local/e004", datetime(2026, 3, 31), "Valid"],
]
ev_meta = add_table(wb["08_Evidences"], 5, 2, ev_headers, ev_rows, "tblEvidence")
apply_widths(wb["08_Evidences"], {"B": 12, "C": 10, "D": 12, "E": 18, "F": 34, "G": 14, "H": 14})
for row in range(ev_meta["first_data_row"], ev_meta["last_data_row"] + 1):
    wb["08_Evidences"][f"G{row}"].number_format = "dd/mm/yyyy"
    wb["08_Evidences"][f"F{row}"].font = link_font
    wb["08_Evidences"][f"F{row}"].hyperlink = wb["08_Evidences"][f"F{row}"].value

setup_sheet(wb["10_CMM_Scoring"], "Maturité CMM", "Calcul des scores de maturité 1 à 5")
cmm_headers = ["domain", "criterion", "weight", "score_input", "weighted_score", "cmm_level"]
cmm_rows = [
    ["Governance", "Control ownership formalized", 20, 4, "=[@weight]*[@score_input]/100", "=[@score_input]"],
    ["Testing", "Test execution discipline", 20, 3, "=[@weight]*[@score_input]/100", "=[@score_input]"],
    ["Monitoring", "KRI escalation", 20, 4, "=[@weight]*[@score_input]/100", "=[@score_input]"],
    ["Evidence", "Audit trail completeness", 20, 3, "=[@weight]*[@score_input]/100", "=[@score_input]"],
    ["Remediation", "Exception closure discipline", 20, 2, "=[@weight]*[@score_input]/100", "=[@score_input]"],
]
cmm_meta = add_table(wb["10_CMM_Scoring"], 5, 2, cmm_headers, cmm_rows, "tblCMM")
apply_widths(wb["10_CMM_Scoring"], {"B": 18, "C": 30, "D": 10, "E": 12, "F": 14, "G": 12})
for row in range(cmm_meta["first_data_row"], cmm_meta["last_data_row"] + 1):
    wb["10_CMM_Scoring"][f"D{row}"].number_format = "#,##0"
    wb["10_CMM_Scoring"][f"E{row}"].number_format = "#,##0"
    wb["10_CMM_Scoring"][f"F{row}"].number_format = "0.0"
    wb["10_CMM_Scoring"][f"G{row}"].number_format = "#,##0"

setup_sheet(wb["11_Import_Workpapers"], "Import workpapers", "Zone de collage/import brut avant normalisation")
imp_headers = ["source_file", "workpaper_ref", "control_id", "test_date", "result", "tester", "finding_ref", "evidence_id", "comment"]
imp_rows = [["WP_Q1.xlsx", "WP-01", "C001", datetime(2026, 1, 31), "Pass", "Auditeur A", "", "E001", "Imported sample row"]]
imp_meta = add_table(wb["11_Import_Workpapers"], 5, 2, imp_headers, imp_rows, "tblImport")
apply_widths(wb["11_Import_Workpapers"], {"B": 18, "C": 16, "D": 12, "E": 14, "F": 12, "G": 14, "H": 14, "I": 12, "J": 24})
for row in range(imp_meta["first_data_row"], imp_meta["last_data_row"] + 1):
    wb["11_Import_Workpapers"][f"E{row}"].number_format = "dd/mm/yyyy"

setup_sheet(wb["09_Dashboard"], "Tableau de bord CCM", "Synthèse automatisée par formules Excel")
ds = wb["09_Dashboard"]
for rng in ["B5:C8", "E5:F8", "H5:I8", "K5:L8"]:
    for row in ds[rng]:
        for cell in row:
            cell.fill = accent_fill
            cell.font = Font(name="Calibri", size=11, bold=True, color=colors["white"])
            cell.alignment = center
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

cards = {
    "B5": "Contrôles",
    "B6": '=ROWS(tblControl[control_id])',
    "E5": "Tests passés",
    "E6": '=COUNTIF(tblTests[result],"Pass")',
    "H5": "Exceptions ouvertes",
    "H6": '=COUNTIF(tblExceptions[status],"Open")',
    "K5": "KRI rouges",
    "K6": '=COUNTIF(tblKRI[status],"Rouge")',
}
for cell_ref, value in cards.items():
    ds[cell_ref] = value
    ds[cell_ref].alignment = center
    ds[cell_ref].font = Font(name="Calibri", size=12 if cell_ref.endswith("5") else 18, bold=True, color=colors["white"])
for cell_ref in ["C5", "C6", "F5", "F6", "I5", "I6", "L5", "L6"]:
    ds[cell_ref].fill = accent_fill
    ds[cell_ref].border = Border(top=thin, bottom=thin, left=thin, right=thin)

row = 11
ds.merge_cells(f"B{row}:G{row}")
ds[f"B{row}"] = "Heatmap contrôles par processus"
ds[f"B{row}"].font = title_font
row += 1
for idx, header in enumerate(["Processus", "On Track", "Overdue", "Failed", "To Schedule", "Taux conformité"], 2):
    cell = ds.cell(row=row, column=idx, value=header)
    cell.fill = subheader_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
for idx, excel_row in enumerate(range(row + 1, row + 5), start=1):
    ds[f"B{excel_row}"] = f"=INDEX(tblProcess[process_name],{idx})"
    ds[f"C{excel_row}"] = f'=COUNTIFS(tblControl[process_name],B{excel_row},tblControl[control_status],"On Track")'
    ds[f"D{excel_row}"] = f'=COUNTIFS(tblControl[process_name],B{excel_row},tblControl[control_status],"Overdue")'
    ds[f"E{excel_row}"] = f'=COUNTIFS(tblControl[process_name],B{excel_row},tblControl[control_status],"Failed")'
    ds[f"F{excel_row}"] = f'=COUNTIFS(tblControl[process_name],B{excel_row},tblControl[control_status],"To Schedule")'
    ds[f"G{excel_row}"] = f'=IFERROR(C{excel_row}/SUM(C{excel_row}:F{excel_row}),0)'
    ds[f"G{excel_row}"].number_format = "0.0%"
    for col in range(2, 8):
        ds.cell(row=excel_row, column=col).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ds.cell(row=excel_row, column=col).alignment = center
        ds.cell(row=excel_row, column=col).font = text_font

row = 19
ds.merge_cells(f"B{row}:G{row}")
ds[f"B{row}"] = "Suivi KRI"
ds[f"B{row}"].font = title_font
row += 1
for idx, header in enumerate(["KRI", "Processus", "Valeur", "Statut", "Trend", "Alerte"], 2):
    cell = ds.cell(row=row, column=idx, value=header)
    cell.fill = subheader_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
for idx, excel_row in enumerate(range(row + 1, row + 5), start=1):
    ds[f"B{excel_row}"] = f"=INDEX(tblKRI[kri_name],{idx})"
    ds[f"C{excel_row}"] = (
        f'=IFERROR(XLOOKUP(XLOOKUP(B{excel_row},tblKRI[kri_name],tblKRI[process_id],""),'
        f'tblProcess[process_id],tblProcess[process_name],""),"")'
    )
    ds[f"D{excel_row}"] = f'=XLOOKUP(B{excel_row},tblKRI[kri_name],tblKRI[value],0)'
    ds[f"E{excel_row}"] = f'=XLOOKUP(B{excel_row},tblKRI[kri_name],tblKRI[status],"")'
    ds[f"F{excel_row}"] = f'=XLOOKUP(B{excel_row},tblKRI[kri_name],tblKRI[trend_vs_prev],0)'
    ds[f"G{excel_row}"] = f'=IF(E{excel_row}="Rouge","Escalate",IF(E{excel_row}="Orange","Monitor","OK"))'
    ds[f"F{excel_row}"].number_format = "0.0%"
    for col in range(2, 8):
        ds.cell(row=excel_row, column=col).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ds.cell(row=excel_row, column=col).alignment = center
        ds.cell(row=excel_row, column=col).font = text_font

row = 27
ds.merge_cells(f"B{row}:F{row}")
ds[f"B{row}"] = "Maturité CMM"
ds[f"B{row}"].font = title_font
row += 1
for idx, header in enumerate(["Domaine", "Score", "Niveau", "Commentaire"], 2):
    cell = ds.cell(row=row, column=idx, value=header)
    cell.fill = subheader_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
for idx, excel_row in enumerate(range(row + 1, row + 6), start=1):
    ds[f"B{excel_row}"] = f"=INDEX(tblCMM[domain],{idx})"
    ds[f"C{excel_row}"] = f"=INDEX(tblCMM[score_input],{idx})"
    ds[f"D{excel_row}"] = f"=INDEX(tblCMM[cmm_level],{idx})"
    ds[f"E{excel_row}"] = f'=IF(D{excel_row}>=4,"Mature",IF(D{excel_row}=3,"Defined",IF(D{excel_row}=2,"Developing","Initial")))'
    ds[f"C{excel_row}"].number_format = "0.0"
    for col in range(2, 6):
        ds.cell(row=excel_row, column=col).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ds.cell(row=excel_row, column=col).alignment = center
        ds.cell(row=excel_row, column=col).font = text_font

ds["H27"] = "Score global"
ds["H27"].font = title_font
ds["H28"] = "=SUM(tblCMM[weighted_score])"
ds["H28"].number_format = "0.0"
ds["I27"] = "Niveau global"
ds["I27"].font = title_font
ds["I28"] = '=IF(H28>=4.5,5,IF(H28>=3.5,4,IF(H28>=2.5,3,IF(H28>=1.5,2,1))))'
for cell_ref in ["H27", "H28", "I27", "I28"]:
    ds[cell_ref].border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ds[cell_ref].alignment = center

apply_widths(ds, {"B": 26, "C": 18, "D": 14, "E": 14, "F": 14, "G": 16, "H": 16, "I": 16, "J": 12, "K": 16, "L": 12})

home = wb["00_Home"]
setup_sheet(home, "CCM - Continuous Controls Monitoring", "Classeur Excel généré automatiquement via Python")
home["B5"] = "Objectif"
home["B5"].font = title_font
home["B6"] = "Centraliser la matrice des contrôles, les tests, les exceptions, les KRI et la maturité CMM dans un classeur pilotable."
home["B6"].alignment = wrap
home["B6"].font = text_font
home.merge_cells("B6:J6")
home["B8"] = "Date de génération"
home["C8"] = datetime.now()
home["C8"].number_format = "dd/mm/yyyy hh:mm"
home["B9"] = "Fichier cible"
home["C9"] = str(OUT_FILE)
home["B11"] = "Navigation"
home["B11"].font = title_font
nav_items = [
    ("Processus", "02_Processus"),
    ("Risques", "03_Risques"),
    ("Control Matrix", "04_Control_Matrix"),
    ("Tests", "05_Tests"),
    ("Exceptions", "06_Exceptions"),
    ("KRI", "07_KRI"),
    ("Dashboard", "09_Dashboard"),
    ("CMM", "10_CMM_Scoring"),
]
for idx, (label, target) in enumerate(nav_items, start=12):
    home[f"B{idx}"] = label
    home[f"B{idx}"].font = link_font
    home[f"B{idx}"].hyperlink = f"#{target}!A1"
    home[f"C{idx}"] = f"Accéder à {target}"
apply_widths(home, {"B": 24, "C": 48, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14})

val_freq = DataValidation(type="list", formula1='"Mensuelle,Trimestrielle,Semestrielle,Annuelle"', allow_blank=True)
val_res = DataValidation(type="list", formula1='"Pass,Fail,Not Tested"', allow_blank=True)
val_exc = DataValidation(type="list", formula1='"Open,Mitigated,Closed"', allow_blank=True)
wb["04_Control_Matrix"].add_data_validation(val_freq)
val_freq.add("I6:I200")
wb["05_Tests"].add_data_validation(val_res)
val_res.add("G6:G500")
wb["06_Exceptions"].add_data_validation(val_exc)
val_exc.add("I6:I500")

apply_equals_cf(wb["04_Control_Matrix"], "P6:P500", "P6", "Failed", red_fill)
apply_equals_cf(wb["04_Control_Matrix"], "P6:P500", "P6", "Overdue", orange_fill)
apply_equals_cf(wb["04_Control_Matrix"], "P6:P500", "P6", "On Track", green_fill)
apply_equals_cf(wb["07_KRI"], "I6:I500", "I6", "Rouge", red_fill)
apply_equals_cf(wb["07_KRI"], "I6:I500", "I6", "Orange", orange_fill)
apply_equals_cf(wb["07_KRI"], "I6:I500", "I6", "Vert", green_fill)

for rng in ["C13:F16", "E21:E24", "G21:G24", "D29:D33", "H27:I28"]:
    set_border_range(ds, rng)

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

try:
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except AttributeError:
    pass

wb.save(OUT_FILE)
print(OUT_FILE)
