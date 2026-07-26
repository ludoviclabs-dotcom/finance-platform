"""scripts/water_intelligence/eea_artifact_inspector.py — cadre la conversion
opérateur d'une release EEA WEI+ à partir d'un artefact officiel local (X2A).

## Le trou que ce module comble, et celui qu'il NE comble PAS

X1 a vérifié l'identité de la release (code de jeu, titre, licence — fiche
ISO 19115 en ligne) et le connecteur (`eea_wei_plus.py`) sait valider un CSV
canonique. Ce qui manquait entre les deux : un moyen de lire le classeur
Excel que l'EEA publie réellement, sans deviner ses colonnes.

Ce module fournit CE moyen — inspection réelle d'un classeur (feuilles,
en-têtes, indicateurs de macro) et conversion vers le CSV canonique — mais ne
prétend PAS connaître la correspondance feuille/colonnes d'une release WEI+
réelle : aucun artefact officiel n'a été obtenu pendant X1 ou X2A (le lien de
téléchargement officiel conduit à une interface Nextcloud, cf.
`docs/carbonco/water-intelligence/activation/X1_LIVE_VALIDATION_HANDOFF.md`
§3.1). `MAPPING_PROFILES` est donc VIDE par construction : la conversion
refuse tant qu'un opérateur n'a pas inspecté un artefact réel avec
`inspect_workbook()` et déclaré un `ColumnMappingProfile` signé.

## Pourquoi `openpyxl` et pas une nouvelle dépendance

`openpyxl==3.1.5` est déjà dans `apps/api/requirements.txt`, utilisé ailleurs
dans le dépôt (`utils/excel_reader.py`, `services/strategic_mapping_excel.py`,
`services/vsme_export.py`). Aucune dépendance n'est ajoutée par ce module.

## Ce qui n'est jamais deviné

Ni le nom de la feuille, ni la colonne d'unité spatiale, ni celle du
trimestre, ni celle du territoire — même invariant que le connecteur
(`eea_wei_plus.py` : « aucune colonne n'est devinée »). `inspect_workbook()`
ne fait QUE constater ce qu'un classeur contient réellement ; la
correspondance vers les champs canoniques WEI+ est un jugement humain,
capturé dans un `ColumnMappingProfile` VERSIONNÉ par release, jamais déduit.

## Ce que ce module ne fait jamais

Aucun réseau (fichier LOCAL uniquement — le téléchargement reste le rôle de
`validate_eea.py`/`fetcher.py`). Aucune écriture du classeur source. Aucun
profil de correspondance pré-rempli avec des noms plausibles.
"""

from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass

try:
    import openpyxl
except ImportError:  # pragma: no cover - déjà dans requirements.txt
    openpyxl = None

#: Statut MVP tant qu'aucun `ColumnMappingProfile` VÉRIFIÉ n'existe pour la
#: release demandée. Cité par
#: `docs/carbonco/water-intelligence/activation/X1_CONNECTOR_READINESS_MATRIX.md`.
MAPPING_PROFILE_STATUS = "manual_artifact_converter_required"

#: Signature d'un projet VBA dans un conteneur zip/OOXML (xlsx/xlsm).
_VBA_PROJECT_ENTRY = "xl/vbaProject.bin"


class ArtifactError(Exception):
    """Refus explicite — jamais une conversion approximative."""


@dataclass(frozen=True)
class ColumnMappingProfile:
    """Correspondance VERSIONNÉE feuille/colonnes → CSV canonique WEI+, pour
    UNE release précise. Créée et signée par un opérateur humain après avoir
    inspecté un artefact réel avec `inspect_workbook()` — jamais déduite
    automatiquement, jamais devinée à partir d'un nom plausible.
    """

    release_key: str
    sheet_name: str
    identifier_column: str
    year_column: str
    quarter_column: str
    value_column: str
    #: `None` si le classeur ne porte aucune colonne d'unité — le connecteur
    #: applique alors `EXPECTED_UNIT` ("%") tel que documenté par la fiche
    #: officielle, jamais une unité devinée depuis le classeur.
    unit_column: str | None
    verified_by: str
    verified_on: str
    notes: str = ""

    def __post_init__(self) -> None:
        if not self.release_key.strip():
            raise ArtifactError("release_key obligatoire.")
        if not self.verified_by.strip() or not self.verified_on.strip():
            raise ArtifactError(
                "un profil de correspondance non attribué à un opérateur et une date "
                "n'est pas vérifié — il ne peut pas exister."
            )


#: VIDE par construction (cf. docstring de module). Une entrée n'est ajoutée
#: ici qu'après inspection RÉELLE d'un artefact officiel.
MAPPING_PROFILES: dict[str, ColumnMappingProfile] = {}


@dataclass(frozen=True)
class WorkbookInspection:
    """Constat FACTUEL sur un classeur — jamais une interprétation. Feuilles
    et en-têtes sont recopiés verbatim ; ce module ne décide pas lesquels
    correspondent aux champs canoniques WEI+."""

    sheet_names: tuple[str, ...]
    headers_by_sheet: dict[str, tuple[str, ...]]
    has_macro_indicators: bool


def has_macro_indicators(raw: bytes) -> bool:
    """Vrai si le conteneur zip/OOXML porte un projet VBA.

    Vérifiable SEULEMENT pour un conteneur zip (xlsx/xlsm) — un classeur
    `.xls` (OLE2, signature `\\xd0\\xcf\\x11\\xe0`) ou tout autre format n'est
    pas inspecté ici. Appeler cette fonction sur un conteneur non-zip lève
    `ArtifactError` plutôt que de rendre `False` par défaut, qui se lirait à
    tort comme « vérifié, aucune macro »."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            return _VBA_PROJECT_ENTRY in archive.namelist()
    except zipfile.BadZipFile as exc:
        raise ArtifactError(
            f"conteneur non-zip : la présence de macro n'est pas vérifiable ({exc})."
        ) from exc


def inspect_workbook(raw: bytes) -> WorkbookInspection:
    """Ouvre un classeur xlsx EN LECTURE SEULE et constate feuilles/en-têtes
    réels. N'écrit rien sur le classeur, ne devine aucune correspondance —
    c'est le SEUL geste que ce module effectue sans profil signé."""
    if openpyxl is None:  # pragma: no cover - dépendance déclarée, absence anormale
        raise ArtifactError(
            "openpyxl indisponible — dépendance pourtant déclarée dans "
            "apps/api/requirements.txt."
        )
    macro = has_macro_indicators(raw)
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        headers: dict[str, tuple[str, ...]] = {}
        for name in workbook.sheetnames:
            sheet = workbook[name]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers[name] = tuple(
                str(cell) if cell is not None else "" for cell in first_row
            )
        return WorkbookInspection(
            sheet_names=tuple(workbook.sheetnames),
            headers_by_sheet=headers,
            has_macro_indicators=macro,
        )
    finally:
        workbook.close()


def convert_to_canonical_csv(raw: bytes, *, release_key: str) -> str:
    """Convertit un classeur vers le CSV canonique WEI+
    (`spatialUnitIdentifier,year,quarter,wei_plus_pct,unit`) — REFUSE tant
    qu'aucun `ColumnMappingProfile` VÉRIFIÉ n'existe pour `release_key`.

    Ne devine JAMAIS une feuille ni une colonne : sans profil signé par un
    opérateur ayant inspecté un artefact réel, aucune valeur n'est produite.
    """
    profile = MAPPING_PROFILES.get(release_key)
    if profile is None:
        raise ArtifactError(
            f"{MAPPING_PROFILE_STATUS} : aucun profil de correspondance vérifié pour "
            f"{release_key!r}. Inspectez l'artefact avec inspect_workbook(), puis "
            "déclarez un ColumnMappingProfile signé avant de convertir — jamais une "
            "feuille ou une colonne devinée."
        )
    if openpyxl is None:  # pragma: no cover
        raise ArtifactError("openpyxl indisponible.")

    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if profile.sheet_name not in workbook.sheetnames:
            raise ArtifactError(
                f"feuille {profile.sheet_name!r} (profil {release_key!r}) absente du "
                f"classeur — feuilles réelles : {list(workbook.sheetnames)}."
            )
        sheet = workbook[profile.sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ArtifactError(f"feuille {profile.sheet_name!r} vide — aucun en-tête.")
        header = [str(cell) if cell is not None else "" for cell in header_row]

        required = [profile.identifier_column, profile.year_column,
                    profile.quarter_column, profile.value_column]
        if profile.unit_column:
            required.append(profile.unit_column)
        missing = [c for c in required if c not in header]
        if missing:
            raise ArtifactError(
                f"colonne(s) {missing} absente(s) de la feuille {profile.sheet_name!r} "
                f"(profil {release_key!r}) — en-têtes réels : {header}."
            )
        index = {name: position for position, name in enumerate(header) if name}

        lines = ["spatialUnitIdentifier,year,quarter,wei_plus_pct,unit"]
        for row in rows:
            if row is None or all(cell is None for cell in row):
                continue
            identifier = row[index[profile.identifier_column]]
            year = row[index[profile.year_column]]
            quarter = row[index[profile.quarter_column]]
            value = row[index[profile.value_column]]
            unit = row[index[profile.unit_column]] if profile.unit_column else "%"
            lines.append(f"{identifier},{year},{quarter},{value},{unit}")
        return "\n".join(lines) + "\n"
    finally:
        workbook.close()
