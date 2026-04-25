"""Verify NEURAL Banque Marketing workbooks."""

from __future__ import annotations

import sys

from openpyxl import load_workbook

from _styles import FILES, MVP_GATES, OUT_DIR, VERDICTS

FOUNDATION_SHEETS = [
    "0_README",
    "1_PARAMS",
    "2_SOURCEBOOK",
    "3_REGULATORY_RULES",
    "4_DISCLOSURE_LIBRARY",
    "5_CHANNEL_MATRIX",
    "6_CONSENT_SEGMENTS",
    "7_RESTRICTED_WORDING",
    "8_EVIDENCE_POLICY",
]

MASTER_SHEETS = [
    "0_README",
    "1_PARAMS",
    "2_AGENT_REGISTRY",
    "3_WORKFLOW_MAP",
    "4_REVIEW_GATES",
    "5_RISK_REGISTER",
    "6_KPI_DASHBOARD",
    "7_ROADMAP",
    "8_CHANGELOG",
]

AGENT_SHEETS = [
    "0_README",
    "1_PARAMS",
    "2_INPUT_SCENARIOS",
    "3_POLICY_GATES",
    "4_RESTRICTED_SIGNALS",
    "5_SCORING_ENGINE",
    "6_OUTPUT_PACK",
    "7_DASHBOARD",
    "8_TESTSET",
    "9_LEARNINGS",
]


def count_tables(wb) -> int:
    return sum(len(wb[sheet].tables) for sheet in wb.sheetnames)


def count_validations(wb) -> int:
    return sum(len(wb[sheet].data_validations.dataValidation) for sheet in wb.sheetnames)


def count_formulas(wb) -> int:
    total = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total += 1
    return total


def require_sheets(wb, expected: list[str], filename: str, errors: list[str]) -> None:
    missing = [sheet for sheet in expected if sheet not in wb.sheetnames]
    if missing:
        errors.append(f"{filename}: missing sheets {missing}")


def scenario_verdicts(wb) -> set[str]:
    ws = wb["2_INPUT_SCENARIOS"]
    verdicts = set()
    for row in range(5, ws.max_row + 1):
        value = ws.cell(row, 17).value
        if value:
            verdicts.add(str(value))
    return verdicts


def foundation_gates(wb) -> set[str]:
    ws = wb["3_REGULATORY_RULES"]
    gates = set()
    for row in range(5, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if value:
            gates.add(str(value))
    return gates


def verify() -> int:
    errors: list[str] = []
    print("=" * 74)
    print("  NEURAL / BANQUE / MARKETING - workbook verification")
    print("=" * 74)
    print(f"  Folder: {OUT_DIR}")

    if not OUT_DIR.exists():
        errors.append(f"missing output folder: {OUT_DIR}")
        print("  [FAIL] output folder missing")
        return 1

    for key, filename in FILES.items():
        path = OUT_DIR / filename
        if not path.exists():
            errors.append(f"missing workbook: {filename}")
            continue

        wb = load_workbook(path, data_only=False)
        if key == "foundations":
            require_sheets(wb, FOUNDATION_SHEETS, filename, errors)
        elif key == "master":
            require_sheets(wb, MASTER_SHEETS, filename, errors)
        else:
            require_sheets(wb, AGENT_SHEETS, filename, errors)

        table_count = count_tables(wb)
        validation_count = count_validations(wb)
        formula_count = count_formulas(wb)
        print()
        print(f"  {filename}")
        print(f"    sheets={len(wb.sheetnames)} tables={table_count} dvs={validation_count} formulas={formula_count}")

        if table_count < 4:
            errors.append(f"{filename}: expected at least 4 Excel tables")
        if validation_count < 1:
            errors.append(f"{filename}: expected data validations")
        if key != "foundations" and formula_count < 5:
            errors.append(f"{filename}: expected key formulas")

        if key not in {"foundations", "master"}:
            observed = scenario_verdicts(wb)
            print(f"    scenario verdicts={sorted(observed)}")
            missing = set(VERDICTS) - observed
            if missing:
                errors.append(f"{filename}: missing scenario verdicts {sorted(missing)}")

    foundations_path = OUT_DIR / FILES["foundations"]
    if foundations_path.exists():
        wb = load_workbook(foundations_path, data_only=False)
        gates = foundation_gates(wb)
        print()
        print("  MVP gate coverage")
        print(f"    gates={sorted(gates)}")
        missing = set(MVP_GATES) - gates
        if missing:
            errors.append(f"foundations: missing MVP gates {sorted(missing)}")

    print()
    print("  Isolation")
    print("    verifier does not write or inspect bank-comms, luxe-comms, aero-comms, insurance-supply-chain or public site outputs")

    print()
    print("=" * 74)
    if errors:
        print("  RESULT: FAIL")
        for err in errors:
            print(f"  - {err}")
        print("=" * 74)
        return 1

    print("  RESULT: OK")
    print("=" * 74)
    return 0


if __name__ == "__main__":
    sys.exit(verify())
