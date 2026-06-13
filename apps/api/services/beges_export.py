"""
beges_export.py — T4.2 : export BEGES réglementaire v5 (France).

Table de passage GHG Protocol → nomenclature BEGES v5 (6 catégories / 22 postes,
arrêté du 25/01/2022). La ventilation est une PARTITION exhaustive : total BEGES
= total GHG (réconciliation automatique testée). Export PDF + Excel (hash par
ligne) dans un ZIP auditable enregistré dans export_packages (domaine 'beges').
Réutilise les patterns de vsme_export. Dépôt manuel sur bilans-ges.ademe.fr.
"""

from __future__ import annotations

import io
import json
import logging
import os
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any

from db.database import db_available, get_db
from services import scope3_service
from services.vsme_export import _FIXED_META_DATE, _p, _sha
from utils.excel_sanitize import sanitize_cell

logger = logging.getLogger(__name__)

_DATA_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "beges_v5.json")


class BegesError(Exception):
    """Erreur de génération de l'export BEGES."""


def _data() -> dict[str, Any]:
    with open(_DATA_PATH, encoding="utf-8") as f:
        return json.load(f)


def eligibility(fte: int | None, country: str = "FR") -> dict[str, Any]:
    """Statut d'éligibilité BEGES (pur)."""
    if fte and fte > 500 and (country or "FR").upper() == "FR":
        return {"status": "obligatoire", "label": "Obligatoire tous les 4 ans (> 500 salariés en métropole)"}
    if fte and fte > 250:
        return {"status": "obligatoire_om", "label": "Obligatoire tous les 4 ans si > 250 salariés en outre-mer"}
    return {"status": "volontaire", "label": "Démarche volontaire (sous les seuils réglementaires)"}


def ventilate(scope_totals: dict[str, Any]) -> dict[str, Any]:
    """Ventile les totaux GHG (S1, S2, S3 par catégorie) dans les 22 postes BEGES.

    Fonction PURE. scope_totals = {"S1": float, "S2": float,
    "S3": {1: float, ..., "uncategorized": float}}. Retourne les 6 catégories
    avec leurs postes + total. Partition exhaustive (réconciliation garantie).
    """
    data = _data()
    mapping = data["ghg_mapping"]
    poste_values: dict[str, float] = defaultdict(float)

    if scope_totals.get("S1"):
        poste_values[mapping["S1"]] += float(scope_totals["S1"])
    if scope_totals.get("S2"):
        poste_values[mapping["S2"]] += float(scope_totals["S2"])
    for cat, val in (scope_totals.get("S3") or {}).items():
        if not val:
            continue
        key = "S3.uncategorized" if cat in ("uncategorized", "unc") else f"S3.{cat}"
        poste_values[mapping.get(key, "6.1")] += float(val)

    categories = []
    grand_total = 0.0
    for cat in data["categories"]:
        postes = [{"code": p["code"], "label": p["label"], "value": round(poste_values.get(p["code"], 0.0), 6)}
                  for p in cat["postes"]]
        cat_total = round(sum(p["value"] for p in postes), 6)
        categories.append({"code": cat["code"], "label": cat["label"], "total": cat_total, "postes": postes})
        grand_total += cat_total
    return {"standard": data["standard"], "categories": categories, "total": round(grand_total, 6)}


def read_scope_totals(company_id: int) -> dict[str, Any]:
    """Lit les totaux GHG de la company depuis facts_current."""
    totals: dict[str, Any] = {"S1": 0.0, "S2": 0.0, "S3": {}}
    if not db_available():
        return totals
    with get_db(company_id=company_id) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT code, value FROM facts_current WHERE company_id = %s AND code LIKE 'CC.GES.%%'",
                (company_id,),
            )
            for r in cur.fetchall():
                code = r["code"]
                val = float(r["value"]) if r["value"] is not None else 0.0
                if code == "CC.GES.SCOPE1":
                    totals["S1"] = val
                elif code == "CC.GES.SCOPE2_LB":
                    totals["S2"] = val
                elif code == "CC.GES.SCOPE2_MB" and not totals["S2"]:
                    totals["S2"] = val
                elif code == "CC.GES.SCOPE3":
                    totals["S3"]["uncategorized"] = val
                else:
                    n = scope3_service.category_of(code)
                    if n is not None:
                        totals["S3"][n] = val
    return totals


def build_beges_pdf(*, company_name: str, breakdown: dict[str, Any], elig: dict[str, Any], generated_at: str) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError as exc:  # pragma: no cover
        raise BegesError("fpdf2 indisponible.") from exc

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.creation_date = _FIXED_META_DATE
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 12, _p("Bilan GES réglementaire (BEGES v5)"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, _p(company_name), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, _p(f"{breakdown['standard']} — Généré le {generated_at}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, _p(f"Éligibilité : {elig['label']}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, _p(f"Total : {breakdown['total']} tCO2e"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)
    for cat in breakdown["categories"]:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(226, 232, 240)
        pdf.cell(0, 8, _p(f"Catégorie {cat['code']} — {cat['label']} : {cat['total']} tCO2e"),
                 new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.set_font("Helvetica", "", 9)
        for p in cat["postes"]:
            pdf.cell(0, 5, _p(f"   {p['code']}  {p['label']} : {p['value']} tCO2e"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)
    return bytes(pdf.output())


def build_beges_xlsx(*, company_name: str, breakdown: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise BegesError("openpyxl indisponible.") from exc

    wb = Workbook()
    naive = _FIXED_META_DATE.replace(tzinfo=None)
    wb.properties.created = naive
    wb.properties.modified = naive
    ws = wb.active
    ws.title = "BEGES v5"
    ws.append([sanitize_cell(v) for v in ["Entreprise", company_name]])
    ws.append(["Total (tCO2e)", breakdown["total"]])
    ws.append([])
    ws.append(["Catégorie", "Poste", "Libellé", "tCO2e", "Hash ligne (SHA-256)"])
    for cat in breakdown["categories"]:
        for p in cat["postes"]:
            row_hash = _sha(f"{p['code']}|{p['value']}".encode("utf-8"))
            ws.append([cat["code"], sanitize_cell(p["code"]), sanitize_cell(p["label"]), p["value"], row_hash])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_beges_report(*, company_id: int, company_name: str, fte: int | None = None,
                       country: str = "FR", scope_totals: dict[str, Any] | None = None) -> dict[str, Any]:
    if scope_totals is None:
        scope_totals = read_scope_totals(company_id)
    breakdown = ventilate(scope_totals)
    elig = eligibility(fte, country)
    now = datetime.now(tz=timezone.utc)
    pdf_bytes = build_beges_pdf(company_name=company_name, breakdown=breakdown, elig=elig,
                               generated_at=now.strftime("%d/%m/%Y"))
    xlsx_bytes = build_beges_xlsx(company_name=company_name, breakdown=breakdown)

    manifest = {
        "manifest_version": "v1", "report": "BEGES", "standard": breakdown["standard"],
        "company_name": company_name, "total_tco2e": breakdown["total"],
        "files": {
            "beges.pdf": {"sha256": _sha(pdf_bytes), "size": len(pdf_bytes)},
            "beges.xlsx": {"sha256": _sha(xlsx_bytes), "size": len(xlsx_bytes)},
        },
    }
    manifest_bytes = json.dumps(manifest, sort_keys=True, indent=2, ensure_ascii=False).encode("utf-8")
    manifest_hash = _sha(manifest_bytes)
    readme = (
        "Bilan GES reglementaire (BEGES v5) - CarbonCo\n=============================================\n\n"
        f"Entreprise : {company_name}\nTotal : {breakdown['total']} tCO2e\n"
        f"Hash manifest : {manifest_hash}\n\n"
        "Verification : sha256sum -c CHECKSUMS.sha256\n"
        "Depot officiel : https://bilans-ges.ademe.fr\n"
        f"Enregistrement : https://carbon-snowy-nine.vercel.app/verify/{manifest_hash}\n"
    ).encode("utf-8")

    embedded = {"manifest.json": manifest_bytes, "beges.pdf": pdf_bytes, "beges.xlsx": xlsx_bytes, "README.txt": readme}
    checksums = ("\n".join(f"{_sha(d)}  {n}" for n, d in sorted(embedded.items())) + "\n").encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, d in embedded.items():
            zf.writestr(name, d)
        zf.writestr("CHECKSUMS.sha256", checksums)
    zip_bytes = buf.getvalue()
    package_hash = _sha(zip_bytes)
    filename = f"beges-{company_id}-{now.strftime('%Y%m%d-%H%M%S')}-{package_hash[:12]}.zip"

    if db_available():
        try:
            with get_db(company_id=company_id) as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO export_packages
                            (company_id, package_hash, manifest_hash, domain, filename,
                             size_bytes, event_count, frozen_count, meta)
                        VALUES (%s, %s, %s, 'beges', %s, %s, 0, 0, %s)
                        ON CONFLICT (package_hash) DO NOTHING
                        """,
                        (company_id, package_hash, manifest_hash, filename, len(zip_bytes),
                         json.dumps({"company_name": company_name})),
                    )
        except Exception as exc:  # pragma: no cover
            logger.warning("Persist export_packages (beges) échoué: %s", exc)

    return {"zip_bytes": zip_bytes, "filename": filename, "package_hash": package_hash,
            "manifest_hash": manifest_hash, "breakdown": breakdown, "eligibility": elig}
