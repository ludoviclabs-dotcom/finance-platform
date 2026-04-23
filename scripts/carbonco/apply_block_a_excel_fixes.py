from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName


WORKBOOK_ROOT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\Carbon and Co")

CARBON_FILE = WORKBOOK_ROOT / "CarbonCo_Calcul_Carbone_v2.xlsx"
ESG_FILE = WORKBOOK_ROOT / "CarbonCo_ESG_Social.xlsx"
FINANCE_FILE = WORKBOOK_ROOT / "CarbonCo_Finance_DPP_v1_3.xlsx"


CC_NAMED_RANGES = {
    "CC_Raison_Sociale": "'Paramètres'!$B$4",
    "CC_Annee_Reporting": "'Paramètres'!$B$7",
    "CC_CA_Net": "'Paramètres'!$B$9",
    "CC_ETP": "'Paramètres'!$B$11",
    "CC_Secteur_Activite": "'Paramètres'!$B$6",
    "CC_Secteur_NAF": "'Paramètres'!$C$6",
    "CC_Surface_Totale": "'Paramètres'!$B$12",
    "CC_CapEx_Total": "'Paramètres'!$B$14",
    "CC_OpEx_Eligible_Taxo": "'Paramètres'!$B$15",
    "CC_GES_Scope1": "'Synthese_GES'!$C$10",
    "CC_GES_Scope2_LB": "'Synthese_GES'!$C$15",
    "CC_GES_Scope2_MB": "'Synthese_GES'!$C$17",
    "CC_GES_Scope3": "'Synthese_GES'!$C$35",
    "CC_GES_Total_S123": "'Synthese_GES'!$C$47",
    "CC_Intensite_CA": "'Synthese_GES'!$C$50",
    "CC_Intensite_ETP": "'Synthese_GES'!$C$51",
    "CC_Part_Scope1": "'Synthese_GES'!$C$53",
    "CC_Part_Scope2": "'Synthese_GES'!$C$54",
    "CC_Part_Scope3": "'Synthese_GES'!$C$55",
    "CC_Conso_Energie_MWh": "'Energie'!$E$19",
    "CC_Part_ENR": "'Energie'!$E$20",
    "CC_Taxo_CA_Aligne": "'Taxonomie'!$E$27",
    "CC_Taxo_CapEx_Aligne": "'Taxonomie'!$E$28",
    "CC_Taxo_OpEx_Aligne": "'Taxonomie'!$E$29",
    "CC_CBAM_Cout_Estime": "'CBAM'!$M$24",
    "CC_SBTI_Annee_Baseline": "'Trajectoire_SBTi'!$B$4",
    "CC_SBTI_Baseline_S12": "'Trajectoire_SBTi'!$B$5",
    "CC_SBTI_Baseline_S3": "'Trajectoire_SBTi'!$B$6",
    "CC_SBTI_Taux_S12": "'Trajectoire_SBTi'!$B$8",
    "CC_SBTI_Taux_S3": "'Trajectoire_SBTi'!$B$9",
}


def add_or_replace_defined_name(wb, name: str, target: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=target))


def ensure_claude_log(wb) -> None:
    if "Claude Log" in wb.sheetnames:
        return

    ws = wb.create_sheet("Claude Log")
    headers = ["Turn #", "Date", "User Request", "Action Taken", "Details", "Outcome"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    ws["A1"].font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor="F0E6CC")


def append_claude_log(wb, request: str, action: str, details: str, outcome: str) -> None:
    ensure_claude_log(wb)
    ws = wb["Claude Log"]
    next_row = ws.max_row + 1
    turn_value = 1
    if next_row > 2:
        prev = ws.cell(next_row - 1, 1).value
        if isinstance(prev, int):
            turn_value = prev + 1
        else:
            try:
                turn_value = int(prev) + 1
            except Exception:
                turn_value = next_row - 1
    ws.cell(next_row, 1, turn_value)
    ws.cell(next_row, 2, date.today().isoformat())
    ws.cell(next_row, 3, request)
    ws.cell(next_row, 4, action)
    ws.cell(next_row, 5, details)
    ws.cell(next_row, 6, outcome)


def fix_carbon_workbook() -> None:
    wb = load_workbook(CARBON_FILE)
    try:
        for key, target in CC_NAMED_RANGES.items():
            add_or_replace_defined_name(wb, key, target)

        append_claude_log(
            wb,
            "Block A data contract alignment",
            "Added canonical CC_* named ranges for shared enterprise, carbon, energy, taxonomie, CBAM and SBTi fields.",
            "Named ranges now expose the stable integration contract consumed by Phase 0 documentation and future backend ingestion.",
            "Carbon master workbook now contains the Phase 0 source contract.",
        )
        wb.save(CARBON_FILE)
    finally:
        wb.close()


def fix_esg_workbook() -> None:
    wb = load_workbook(ESG_FILE)
    try:
        ws = wb["Liaison_Donnees"]
        ws["E16"] = "Synthese_GES!C15"
        ws["E17"] = "Synthese_GES!C17"
        ws["E18"] = "Synthese_GES!C35"
        ws["E22"] = "Energie!E20"
        ws["E23"] = "Energie!E19"
        ws["E25"] = "Taxonomie!E27"
        ws["E26"] = "CBAM!M24"

        append_claude_log(
            wb,
            "Block A liaison hardening",
            "Corrected misleading source references in Liaison_Donnees for Scope 2 LB, Scope 2 MB, Scope 3, energy KPIs, taxonomie and CBAM.",
            "This change keeps the manual import workflow intact while aligning the workbook with the real source cells in CarbonCo_Calcul_Carbone_v2.xlsx.",
            "ESG Social liaison sheet is now structurally aligned with the Carbon source workbook.",
        )
        wb.save(ESG_FILE)
    finally:
        wb.close()


def fix_finance_workbook() -> None:
    wb = load_workbook(FINANCE_FILE)
    try:
        ws = wb["Liaison_Donnees"]
        ws["E16"] = "Synthese_GES!C15"
        ws["E17"] = "Synthese_GES!C17"
        ws["E18"] = "Synthese_GES!C35"
        ws["B19"] = "Total S1+S2(LB)+S3 (tCO2e)"
        ws["E19"] = "Calcul auto (LB)"
        ws["E23"] = "Energie!E20"
        ws["E24"] = "Taxonomie!E27"
        ws["E25"] = "Taxonomie!E28"
        ws["E26"] = "CBAM!M24"
        ws["E30"] = "Trajectoire_SBTi!B4"

        append_claude_log(
            wb,
            "Block A liaison hardening",
            "Corrected misleading source references in Liaison_Donnees and clarified that the aggregated total uses Scope 2 location-based values.",
            "Updated Scope, energy, taxonomie, CBAM and baseline-year source notes to match the real source workbook.",
            "Finance DPP liaison sheet is now structurally aligned with the Carbon source workbook.",
        )
        wb.save(FINANCE_FILE)
    finally:
        wb.close()


def main() -> None:
    fix_carbon_workbook()
    fix_esg_workbook()
    fix_finance_workbook()
    print("BLOCK_A_EXCEL_FIXES_APPLIED=1")


if __name__ == "__main__":
    main()
