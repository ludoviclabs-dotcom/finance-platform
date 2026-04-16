"""Routes for Excel file upload, preview, and validation."""

from __future__ import annotations

import io
import logging
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook as _openpyxl_load
from pydantic import BaseModel

from routers.auth import require_analyst
from services.auth_service import AuthUser
from services.carbon_service import (
    CANONICAL_CC_RANGES,
    REQUIRED_SHEETS,
    CarbonServiceError,
    build_carbon_snapshot_from_bytes,
    get_workbook_paths,
)
from services.snapshot_cache import write_snapshot
from utils.excel_reader import (
    CorruptFileError,
    ExcelReader,
    InvalidRangeError,
    SheetNotFoundError,
)

logger = logging.getLogger(__name__)

router = APIRouter()

# Named ranges strictement requis pour /excel/ingest-uploaded (domaine carbon).
# Aligné sur SNAPSHOT_FIELD_TO_KEY filtré aux REQUIRED_SNAPSHOT_FIELDS du
# carbon_service, plus le Scope 2 MB (nécessaire pour le bandeau dashboard).
_REQUIRED_CC_RANGES_CARBON: list[str] = [
    "CC_Raison_Sociale",
    "CC_Annee_Reporting",
    "CC_CA_Net",
    "CC_GES_Scope1",
    "CC_GES_Scope2_LB",
    "CC_GES_Scope3",
    "CC_GES_Total_S123",
    "CC_Conso_Energie_MWh",
    "CC_Part_ENR",
    "CC_Taxo_CA_Aligne",
]

_EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}

# Named ranges expected per domain workbook
_EXPECTED_NAMED_RANGES: dict[str, list[str]] = {
    "carbon": ["company_name", "reporting_year", "scope1_tco2e", "scope2_lb_tco2e",
               "scope3_tco2e", "energie_mwh", "enr_pct"],
    "esg": ["raison_sociale", "score_esg_global", "enjeux_evalues"],
    "finance": ["prix_ets", "capex_decarb_s12", "score_esg_investisseur"],
}

# Sheets expected per domain
_EXPECTED_SHEETS: dict[str, list[str]] = {
    "carbon": ["Bilan GES", "Energie", "Taxonomie"],
    "esg": ["VSME", "Materialite", "Scores ESG"],
    "finance": ["Finance Climat", "SFDR PAI", "Benchmark"],
}


def _check_mime(file: UploadFile) -> None:
    if file.content_type and file.content_type not in _EXCEL_MIMES:
        if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
            raise HTTPException(
                status_code=400,
                detail=f"Type de fichier invalide '{file.content_type}'. Veuillez uploader un fichier Excel (.xlsx).",
            )


# ---------------------------------------------------------------------------
# POST /upload — upload simple (inchangé)
# ---------------------------------------------------------------------------

@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(..., description="Excel file (.xlsx)"),
) -> dict[str, Any]:
    _check_mime(file)
    try:
        reader = await ExcelReader.from_upload(file)
    except CorruptFileError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    sheets = reader.sheet_names
    reader.close()
    return {"filename": file.filename, "sheets": sheets, "sheet_count": len(sheets)}


# ---------------------------------------------------------------------------
# POST /preview — lire les métadonnées + premières lignes de chaque sheet
# ---------------------------------------------------------------------------

class SheetPreview(BaseModel):
    name: str
    row_count: int | None
    col_count: int | None
    headers: list[str]           # première ligne non-vide
    sample_rows: list[list[Any]] # 5 premières lignes de données


class PreviewResponse(BaseModel):
    filename: str
    domain: str | None
    sheet_count: int
    sheets: list[SheetPreview]
    named_ranges: list[str]
    detected_domain: str | None  # "carbon" | "esg" | "finance" | None


@router.post("/preview", response_model=PreviewResponse)
async def preview_excel(
    file: UploadFile = File(...),
    domain: str | None = None,
) -> PreviewResponse:
    """
    Read workbook metadata + first rows of each sheet for preview.
    domain hint (carbon|esg|finance) is optional — used to highlight expected sheets.
    """
    _check_mime(file)
    try:
        # Need read_only=False to access named_ranges + dimensions
        import io

        from openpyxl import load_workbook as _load

        contents = await file.read()
        wb = _load(io.BytesIO(contents), read_only=False, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Fichier illisible : {exc}") from exc

    # Named ranges
    named_ranges = [nr.name for nr in wb.defined_names.definedName] if hasattr(wb, "defined_names") else []

    # Sheet previews
    sheet_previews: list[SheetPreview] = []
    for sheet_name in wb.sheetnames[:8]:  # limiter à 8 feuilles
        ws = wb[sheet_name]
        headers: list[str] = []
        sample: list[list[Any]] = []
        row_idx = 0
        for row in ws.iter_rows(max_row=7, values_only=True):
            if not any(c is not None for c in row):
                continue
            row_vals = [c for c in row if c is not None or True]
            if not headers:
                headers = [str(c) if c is not None else "" for c in row_vals]
            else:
                sample.append([c for c in row_vals])
                if len(sample) >= 5:
                    break
            row_idx += 1

        # Dimensions
        try:
            max_row = ws.max_row
            max_col = ws.max_column
        except Exception:
            max_row = None
            max_col = None

        sheet_previews.append(SheetPreview(
            name=sheet_name,
            row_count=max_row,
            col_count=max_col,
            headers=headers[:20],
            sample_rows=[r[:20] for r in sample],
        ))

    wb.close()

    # Auto-detect domain from sheet names
    detected: str | None = None
    if not domain:
        sheet_set = set(wb.sheetnames) if hasattr(wb, "sheetnames") else set()
        for d, expected in _EXPECTED_SHEETS.items():
            if any(s in sheet_set for s in expected):
                detected = d
                break
    else:
        detected = domain if domain in _EXPECTED_SHEETS else None

    return PreviewResponse(
        filename=file.filename or "",
        domain=domain,
        sheet_count=len(wb.sheetnames) if hasattr(wb, "sheetnames") else len(sheet_previews),
        sheets=sheet_previews,
        named_ranges=named_ranges[:50],
        detected_domain=detected,
    )


# ---------------------------------------------------------------------------
# POST /validate — contrôles structurels avant ingest
# ---------------------------------------------------------------------------

class ValidationIssue(BaseModel):
    level: str        # "error" | "warning" | "info"
    message: str
    sheet: str | None = None
    field: str | None = None


class ValidateResponse(BaseModel):
    filename: str
    domain: str | None
    status: str       # "ok" | "warning" | "error"
    issues: list[ValidationIssue]
    named_ranges_found: list[str]
    named_ranges_missing: list[str]
    sheets_found: list[str]
    sheets_missing: list[str]


@router.post("/validate", response_model=ValidateResponse)
async def validate_excel(
    file: UploadFile = File(...),
    domain: str | None = None,
) -> ValidateResponse:
    """
    Validate workbook structure before ingest:
    - Check expected sheets are present
    - Check expected named ranges exist
    - Check non-empty file
    - Warn on unusually small files
    """
    _check_mime(file)
    issues: list[ValidationIssue] = []

    try:
        import io

        from openpyxl import load_workbook as _load

        contents = await file.read()
        if len(contents) < 512:
            issues.append(ValidationIssue(level="error", message="Fichier trop petit — probablement vide ou corrompu."))
            return ValidateResponse(
                filename=file.filename or "",
                domain=domain,
                status="error",
                issues=issues,
                named_ranges_found=[], named_ranges_missing=[],
                sheets_found=[], sheets_missing=[],
            )

        if len(contents) < 5000:
            issues.append(ValidationIssue(level="warning", message="Fichier inhabituellement petit — vérifiez qu'il contient bien des données."))

        wb = _load(io.BytesIO(contents), read_only=False, data_only=True)
    except Exception as exc:
        issues.append(ValidationIssue(level="error", message=f"Fichier illisible : {exc}"))
        return ValidateResponse(
            filename=file.filename or "", domain=domain, status="error",
            issues=issues,
            named_ranges_found=[], named_ranges_missing=[],
            sheets_found=[], sheets_missing=[],
        )

    # Auto-detect domain
    effective_domain = domain
    if not effective_domain:
        sheet_set = set(wb.sheetnames)
        for d, expected_sheets in _EXPECTED_SHEETS.items():
            if any(s in sheet_set for s in expected_sheets):
                effective_domain = d
                break

    # Named ranges
    named_ranges_found: list[str] = [nr.name for nr in wb.defined_names.definedName] if hasattr(wb, "defined_names") else []
    expected_nr = _EXPECTED_NAMED_RANGES.get(effective_domain or "", [])
    named_ranges_missing = [n for n in expected_nr if n not in named_ranges_found]
    for missing in named_ranges_missing:
        issues.append(ValidationIssue(
            level="warning",
            message=f"Plage nommée manquante : '{missing}'",
            field=missing,
        ))

    # Sheets
    sheets_found = list(wb.sheetnames)
    expected_sh = _EXPECTED_SHEETS.get(effective_domain or "", [])
    sheets_missing = [s for s in expected_sh if s not in sheets_found]
    for missing in sheets_missing:
        issues.append(ValidationIssue(
            level="warning",
            message=f"Feuille attendue manquante : '{missing}'",
            sheet=missing,
        ))

    # Empty sheets check
    for sheet_name in wb.sheetnames[:6]:
        ws = wb[sheet_name]
        try:
            if ws.max_row is None or ws.max_row < 2:
                issues.append(ValidationIssue(
                    level="warning",
                    message=f"La feuille '{sheet_name}' semble vide.",
                    sheet=sheet_name,
                ))
        except Exception:
            pass

    wb.close()

    errors = [i for i in issues if i.level == "error"]
    warnings = [i for i in issues if i.level == "warning"]
    status = "error" if errors else "warning" if warnings else "ok"

    return ValidateResponse(
        filename=file.filename or "",
        domain=effective_domain,
        status=status,
        issues=issues,
        named_ranges_found=named_ranges_found[:50],
        named_ranges_missing=named_ranges_missing,
        sheets_found=sheets_found,
        sheets_missing=sheets_missing,
    )


# ---------------------------------------------------------------------------
# Endpoints lecture bas niveau (inchangés)
# ---------------------------------------------------------------------------

@router.post("/read-cell")
async def read_cell(
    file: UploadFile = File(...),
    sheet: str = "Sheet1",
    cell: str = "A1",
) -> dict[str, Any]:
    try:
        reader = await ExcelReader.from_upload(file)
        value = reader.get_cell(sheet, cell)
    except CorruptFileError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except SheetNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except InvalidRangeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        try:
            reader.close()
        except Exception:
            pass
    return {"sheet": sheet, "cell": cell, "value": value}


@router.post("/read-range")
async def read_range(
    file: UploadFile = File(...),
    sheet: str = "Sheet1",
    range_str: str = "A1:D10",
) -> dict[str, Any]:
    try:
        reader = await ExcelReader.from_upload(file)
        data = reader.get_range_as_dicts(sheet, range_str)
    except CorruptFileError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except SheetNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except InvalidRangeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        try:
            reader.close()
        except Exception:
            pass
    return {"sheet": sheet, "range": range_str, "row_count": len(data), "data": data}


# ---------------------------------------------------------------------------
# POST /ingest-uploaded — parse + snapshot depuis le fichier utilisateur
# ---------------------------------------------------------------------------


class IngestUploadedKpis(BaseModel):
    scope1Tco2e: float | None = None
    scope2LbTco2e: float | None = None
    scope2MbTco2e: float | None = None
    scope3Tco2e: float | None = None
    totalS123Tco2e: float | None = None


class IngestUploadedResponse(BaseModel):
    snapshotId: int | None
    version: int | None
    generatedAt: str
    domain: str
    source: Literal["user_upload"]
    kpis: IngestUploadedKpis
    validation: dict[str, Any]


def _strict_validate_carbon_workbook(contents: bytes) -> tuple[list[str], list[str]]:
    """Return (named_ranges_missing, sheets_missing) for the uploaded carbon file."""
    try:
        wb = _openpyxl_load(io.BytesIO(contents), read_only=False, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "invalid_workbook",
                "message": f"Fichier illisible : {exc}",
            },
        ) from exc

    defined_names = set(wb.defined_names) if hasattr(wb, "defined_names") else set()
    sheet_names = set(wb.sheetnames)
    wb.close()

    named_ranges_missing = [n for n in _REQUIRED_CC_RANGES_CARBON if n not in defined_names]
    # REQUIRED_SHEETS["carbon"] est un set — on garde un ordre stable pour la réponse
    required_sheets = sorted(REQUIRED_SHEETS["carbon"])
    sheets_missing = [s for s in required_sheets if s not in sheet_names]
    return named_ranges_missing, sheets_missing


@router.post("/ingest-uploaded", response_model=IngestUploadedResponse)
async def ingest_uploaded(
    file: UploadFile = File(..., description="Classeur Excel utilisateur (.xlsx)"),
    domain: str = Form("carbon"),
    user: AuthUser = Depends(require_analyst),
) -> IngestUploadedResponse:
    company_id = user.company_id
    """Ingest strict : l'utilisateur uploade son classeur, on calcule le
    snapshot puis on l'écrit avec source='user_upload'.

    Rejette (422) toute structure non conforme au template maître (named
    ranges CC_* et sheets canoniques requis).
    """
    if domain != "carbon":
        raise HTTPException(
            status_code=400,
            detail=f"Domaine '{domain}' non supporté en Phase 1.A. Utilisez 'carbon'.",
        )
    _check_mime(file)

    contents = await file.read()
    if len(contents) < 512:
        raise HTTPException(
            status_code=422,
            detail={"error": "empty_workbook", "message": "Fichier trop petit ou vide."},
        )

    named_ranges_missing, sheets_missing = _strict_validate_carbon_workbook(contents)
    if named_ranges_missing or sheets_missing:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "invalid_workbook_structure",
                "message": "Structure non conforme au template officiel.",
                "named_ranges_missing": named_ranges_missing,
                "sheets_missing": sheets_missing,
                "hint": "Téléchargez le template officiel via GET /excel/template?domain=carbon et reportez-y vos données.",
            },
        )

    try:
        snapshot = build_carbon_snapshot_from_bytes(
            contents, source_filename=file.filename
        )
    except CarbonServiceError as exc:
        logger.error("Carbon snapshot build failed for user %s: %s", user.email, exc)
        raise HTTPException(status_code=500, detail=f"Échec du calcul snapshot : {exc}") from exc

    if snapshot["validation"]["status"] == "failed":
        raise HTTPException(
            status_code=422,
            detail={
                "error": "snapshot_validation_failed",
                "message": "Les données extraites ne satisfont pas les règles minimales.",
                "failures": snapshot["validation"]["failures"],
                "warnings": snapshot["validation"]["warnings"],
            },
        )

    write_result = write_snapshot(
        "carbon",
        snapshot,
        company_id=company_id,
        source="user_upload",
    )

    carbon_kpis = snapshot.get("carbon", {}) or {}
    return IngestUploadedResponse(
        snapshotId=write_result["id"] if write_result else None,
        version=write_result["version"] if write_result else None,
        generatedAt=snapshot["generatedAt"],
        domain="carbon",
        source="user_upload",
        kpis=IngestUploadedKpis(
            scope1Tco2e=carbon_kpis.get("scope1Tco2e"),
            scope2LbTco2e=carbon_kpis.get("scope2LbTco2e"),
            scope2MbTco2e=carbon_kpis.get("scope2MbTco2e"),
            scope3Tco2e=carbon_kpis.get("scope3Tco2e"),
            totalS123Tco2e=carbon_kpis.get("totalS123Tco2e"),
        ),
        validation=snapshot["validation"],
    )


# ---------------------------------------------------------------------------
# GET /template — télécharger le template officiel pour un domaine
# ---------------------------------------------------------------------------


@router.get("/template")
async def download_template(domain: str = "carbon") -> FileResponse:
    """Sert le classeur maître comme template de départ pour l'utilisateur."""
    if domain != "carbon":
        raise HTTPException(
            status_code=400,
            detail=f"Template '{domain}' non disponible (seul 'carbon' en Phase 1.A).",
        )
    paths = get_workbook_paths()
    template_path = paths[domain]
    if not template_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Template introuvable côté serveur : {template_path.name}",
        )
    return FileResponse(
        path=str(template_path),
        filename=f"CarbonCo_Template_{domain}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
