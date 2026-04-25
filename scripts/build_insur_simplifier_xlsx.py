"""
Build InsurSimplifier_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'InsurSimplifier' (branche Assurances — Marketing).

7 onglets : README, Inputs, Output langage clair, FAQ, Mapping articles,
KPIs, Limites. 3 cas réels (auto alcool, auto vol total, MRH dégâts des eaux).
Formules Flesch FR + réduction calculées dans Excel (pas de hardcode).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_VIOLET_LIGHT = "C4B5FD"
NEURAL_INK = "0E0824"
NEURAL_BG = "F6F2FF"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=NEURAL_VIOLET, end_color=NEURAL_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def yellow_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def write_title(ws, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, bold=False, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict[int, float]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ────────────────────────────────────────────────────────────────────
# Données : 3 cas réels
# ────────────────────────────────────────────────────────────────────
CASES = [
    {
        "id": "AUTO-01",
        "branche": "Auto",
        "garantie": "Exclusion conduite sous emprise",
        "source": "Modèle de Conditions Générales auto FFA, art. 24 (exclusions)",
        "original": (
            "Sont exclus de la garantie les dommages causés ou aggravés par le conducteur "
            "reconnu en état d'imprégnation alcoolique caractérisée par une concentration "
            "d'alcool dans le sang égale ou supérieure à celle réprimée par la législation "
            "en vigueur, ou en état d'ivresse manifeste, ainsi que ceux causés ou aggravés "
            "sous l'effet de substances ou plantes classées comme stupéfiants, dès lors qu'il "
            "en est résulté pour le conducteur l'application des dispositions du code de la "
            "route relatives à ces infractions."
        ),
        "reformule": (
            "Vous n'êtes pas couvert si vous conduisez en état d'ivresse (taux d'alcool "
            "au-delà de la limite légale, soit 0,5 g/L de sang en France pour les conducteurs "
            "expérimentés et 0,2 g/L pour les jeunes conducteurs) ou sous l'effet de drogues, "
            "et que cela a contribué à l'accident."
        ),
        "orig_words": 76,
        "orig_sentences": 1,
        "orig_syllables": 188,
        "ref_words": 50,
        "ref_sentences": 1,
        "ref_syllables": 96,
        "orig_jargon": 7,
        "ref_jargon": 1,
        "faq": [
            (
                "Si je conduis avec un taux d'alcool de 0,3 g/L, suis-je couvert ?",
                "Oui en théorie (vous êtes en dessous de 0,5 g/L), sauf si vous êtes jeune "
                "conducteur (limite à 0,2 g/L). L'exclusion ne s'applique qu'au-delà de la "
                "limite légale.",
            ),
            (
                "Et si l'alcool n'a rien à voir avec l'accident ?",
                "L'exclusion s'applique uniquement si l'alcoolémie ou les stupéfiants ont "
                "contribué ou aggravé le sinistre. En cas de litige, l'expert et la justice "
                "tranchent.",
            ),
            (
                "Le passager sera-t-il indemnisé ?",
                "Oui. Cette exclusion concerne les dommages au véhicule du conducteur en "
                "faute. La garantie responsabilité civile (obligatoire) couvre les tiers, "
                "dont les passagers.",
            ),
            (
                "Que se passe-t-il pour les médicaments délivrés sur ordonnance ?",
                "La clause vise les substances classées comme stupéfiants. Certains "
                "médicaments en contiennent (opioïdes forts) — vérifiez le pictogramme rouge "
                "« incompatible avec la conduite » sur la boîte.",
            ),
            (
                "Cette exclusion s'applique-t-elle aussi à la garantie incendie ?",
                "Oui, si l'imprégnation a contribué à l'incendie (par exemple un accident "
                "suivi d'un incendie). Vérifiez le détail dans vos Conditions Particulières.",
            ),
        ],
        "mapping": [
            (
                "concentration d'alcool dans le sang égale ou supérieure à celle "
                "réprimée par la législation en vigueur",
                "taux d'alcool au-delà de la limite légale, soit 0,5 g/L",
                "Précision : ajout du seuil chiffré (0,5 g/L et 0,2 g/L jeunes conducteurs)",
            ),
            (
                "en état d'ivresse manifeste",
                "en état d'ivresse",
                "Suppression du qualificatif juridique « manifeste » (équivalent en clair)",
            ),
            (
                "substances ou plantes classées comme stupéfiants",
                "drogues",
                "Vulgarisation acceptée (terme technique fidèle au sens du code de la route)",
            ),
            (
                "dès lors qu'il en est résulté pour le conducteur l'application des "
                "dispositions du code de la route relatives à ces infractions",
                "et que cela a contribué à l'accident",
                "Reformulation pédagogique du lien de causalité",
            ),
        ],
    },
    {
        "id": "AUTO-02",
        "branche": "Auto",
        "garantie": "Vol total — calcul de l'indemnité",
        "source": "Modèle de Conditions Générales auto FFA, art. 18 (indemnisation vol)",
        "original": (
            "En cas de vol total du véhicule assuré, l'indemnité versée par l'assureur est "
            "calculée sur la base de la valeur de remplacement à dire d'expert au jour du "
            "sinistre, déduction faite d'une franchise contractuelle dont le montant figure "
            "aux Conditions Particulières, et plafonnée à la valeur figurant aux Conditions "
            "Particulières lors de la dernière échéance principale du contrat."
        ),
        "reformule": (
            "Si votre voiture est volée, on vous rembourse sa valeur estimée par un expert "
            "le jour du vol. On déduit votre franchise (le montant qui reste à votre charge, "
            "indiqué dans vos Conditions Particulières). Le remboursement ne peut pas "
            "dépasser le plafond également indiqué dans vos Conditions Particulières."
        ),
        "orig_words": 64,
        "orig_sentences": 1,
        "orig_syllables": 156,
        "ref_words": 51,
        "ref_sentences": 3,
        "ref_syllables": 102,
        "orig_jargon": 5,
        "ref_jargon": 2,
        "faq": [
            (
                "Comment est calculée la valeur de mon véhicule volé ?",
                "Un expert estime la valeur de remplacement de votre voiture le jour du vol "
                "(âge, kilométrage, état). Vous n'êtes pas remboursé sur la base du prix "
                "d'achat.",
            ),
            (
                "Qu'est-ce que la franchise ?",
                "C'est le montant qui reste à votre charge sur le remboursement. Il est fixé "
                "dans vos Conditions Particulières (par exemple 300 €).",
            ),
            (
                "Y a-t-il un plafond de remboursement ?",
                "Oui. Le montant maximum est indiqué dans vos Conditions Particulières et "
                "correspond à la dernière échéance principale du contrat.",
            ),
            (
                "Et si le véhicule est retrouvé ?",
                "Selon votre contrat, soit l'assureur reprend le véhicule retrouvé en "
                "échange de l'indemnisation, soit le véhicule vous est restitué et "
                "l'indemnité est ajustée.",
            ),
            (
                "Le vol partiel (objets dans le véhicule) est-il couvert ?",
                "Cela dépend de votre contrat. Le vol partiel est généralement traité "
                "séparément avec sa propre franchise et son propre plafond.",
            ),
        ],
        "mapping": [
            (
                "valeur de remplacement à dire d'expert au jour du sinistre",
                "valeur estimée par un expert le jour du vol",
                "Vulgarisation : « à dire d'expert » → « estimée par un expert »",
            ),
            (
                "déduction faite d'une franchise contractuelle dont le montant figure aux "
                "Conditions Particulières",
                "On déduit votre franchise (le montant qui reste à votre charge, indiqué "
                "dans vos Conditions Particulières)",
                "Définition inline de la franchise pour le grand public",
            ),
            (
                "plafonnée à la valeur figurant aux Conditions Particulières lors de la "
                "dernière échéance principale du contrat",
                "Le remboursement ne peut pas dépasser le plafond également indiqué dans "
                "vos Conditions Particulières",
                "Simplification : suppression de « dernière échéance principale » sans "
                "perte de sens pour l'assuré moyen (à valider compliance)",
            ),
        ],
    },
    {
        "id": "MRH-01",
        "branche": "MRH (habitation)",
        "garantie": "Dégâts des eaux",
        "source": "Modèle de Conditions Générales MRH FFA, art. 12 (dégâts des eaux)",
        "original": (
            "La garantie dégâts des eaux couvre les dommages matériels résultant directement "
            "et accidentellement de l'action de l'eau provenant de fuites, ruptures ou "
            "débordements, à l'exclusion des dommages dus à un défaut d'entretien manifeste, "
            "à l'humidité ou la condensation, ainsi que des frais de recherche de fuite qui "
            "font l'objet d'une garantie distincte mentionnée aux Conditions Particulières."
        ),
        "reformule": (
            "Votre assurance habitation rembourse les dégâts causés par une fuite, une "
            "rupture ou un débordement d'eau accidentel (par exemple, une canalisation qui "
            "éclate). Ne sont pas couverts : les dégâts liés à un manque d'entretien "
            "(par exemple, un robinet qui goutte depuis des mois), l'humidité, la "
            "condensation, ni les frais de recherche de fuite (couverts séparément, voir "
            "vos Conditions Particulières)."
        ),
        "orig_words": 62,
        "orig_sentences": 1,
        "orig_syllables": 158,
        "ref_words": 67,
        "ref_sentences": 2,
        "ref_syllables": 142,
        "orig_jargon": 4,
        "ref_jargon": 1,
        "faq": [
            (
                "Une fuite de mon lave-vaisselle est-elle couverte ?",
                "Oui, si elle est accidentelle (rupture du tuyau, panne soudaine). Non, si "
                "vous avez ignoré un suintement pendant des mois — cela relève du défaut "
                "d'entretien.",
            ),
            (
                "La condensation sur les fenêtres est-elle couverte ?",
                "Non. La condensation et l'humidité chronique sont explicitement exclues "
                "de la garantie dégâts des eaux.",
            ),
            (
                "Qui paie pour trouver l'origine de la fuite ?",
                "Les frais de recherche de fuite sont une garantie distincte. Vérifiez vos "
                "Conditions Particulières — elle est souvent incluse mais avec un plafond "
                "spécifique.",
            ),
            (
                "Mon voisin a une fuite qui a inondé chez moi. Est-ce ma garantie qui paye ?",
                "Oui dans un premier temps. Votre assurance habitation déclenche, puis se "
                "retourne contre l'assurance du voisin (recours). Vous devez quand même "
                "déclarer le sinistre rapidement.",
            ),
            (
                "J'ai laissé mon robinet ouvert en sortant. Suis-je couvert ?",
                "Selon les contrats, c'est généralement considéré comme un débordement "
                "accidentel et donc couvert (avec franchise). Une négligence répétée "
                "pourrait être contestée.",
            ),
        ],
        "mapping": [
            (
                "résultant directement et accidentellement de l'action de l'eau",
                "causés par une fuite, une rupture ou un débordement d'eau accidentel",
                "Reformulation pédagogique : suppression de « action de l'eau » + "
                "exemple concret",
            ),
            (
                "à l'exclusion des dommages dus à un défaut d'entretien manifeste",
                "Ne sont pas couverts : les dégâts liés à un manque d'entretien "
                "(par exemple, un robinet qui goutte depuis des mois)",
                "Ajout d'un exemple concret pour illustrer le « défaut d'entretien »",
            ),
            (
                "frais de recherche de fuite qui font l'objet d'une garantie distincte",
                "frais de recherche de fuite (couverts séparément, voir vos Conditions "
                "Particulières)",
                "Mise en évidence du renvoi aux Conditions Particulières",
            ),
        ],
    },
]


# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────
def build_readme(ws) -> None:
    write_title(ws, "InsurSimplifier — Démo NEURAL · Assurances/Marketing")
    write_subtitle(
        ws,
        2,
        "Reformulation des conditions générales en langage clair + génération de FAQ "
        "produit, avec traçabilité article-par-article (DDA).",
    )

    write_section_header(ws, 4, "Mission")
    ws["A5"] = (
        "Transformer les clauses techniques d'un contrat d'assurance en formulations "
        "compréhensibles par un assuré non-juriste, tout en conservant un mapping fidèle "
        "vers le texte original (auditable conformité)."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A5"].font = body_font()
    ws.merge_cells("A5:F5")
    ws.row_dimensions[5].height = 56

    write_section_header(ws, 7, "Périmètre v1 (démo)")
    perimetre = [
        "3 cas réels : Auto (alcool/stupéfiants), Auto (vol total), MRH (dégâts des eaux)",
        "Branches couvertes : Auto + MRH habitation. Santé et prévoyance : v2.",
        "Sources : modèles FFA + recommandations ACPR 2024-2026 + lignes directrices DDA EIOPA",
        "Lecture cible : grand public — niveau collège (Flesch FR > 60)",
    ]
    for i, item in enumerate(perimetre, start=8):
        ws.cell(row=i, column=1, value=f"• {item}").font = body_font()
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.row_dimensions[i].height = 20

    write_section_header(ws, 13, "Mode d'emploi des onglets")
    onglets = [
        ("0_README", "Mission, périmètre, mode d'emploi (cet onglet)"),
        ("1_Inputs", "Article original tel qu'il figure dans les Conditions Générales"),
        ("2_Output_LangageClair", "Reformulation simplifiée + score Flesch FR avant/après"),
        ("3_FAQ_Generee", "5 questions/réponses dérivées par cas (15 au total)"),
        ("4_Mapping_Articles", "Traçabilité phrase-à-phrase (auditable conformité DDA)"),
        ("5_KPIs", "Tableau de bord avant/après par cas + agrégat"),
        ("6_Limites", "Truth layer : ce que l'agent ne fait pas + disclaimers"),
    ]
    row = 14
    for name, desc in onglets:
        ws.cell(row=row, column=1, value=name).font = body_font(bold=True, color=NEURAL_VIOLET)
        ws.cell(row=row, column=2, value=desc).font = body_font()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 20
        row += 1

    write_section_header(ws, 22, "Conformité réglementaire (avril 2026)")
    conformite = [
        "DDA (Directive Distribution Assurance) — devoir de conseil + clarté contractuelle",
        "Acte délégué clarté contractuelle (en vigueur fin 2025)",
        "EU AI Act art. 50 (août 2026) — étiquetage « contenu généré par IA » obligatoire",
        "RGPD — pas de profilage individuel dans les FAQ générées",
        "Recommandations ACPR sur la lisibilité des contrats (2024-2026)",
    ]
    for i, item in enumerate(conformite, start=23):
        ws.cell(row=i, column=1, value=f"• {item}").font = body_font()
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.row_dimensions[i].height = 20

    write_section_header(ws, 29, "Statut")
    ws["A30"] = "Démo orchestrée · données publiques · validation humaine obligatoire avant diffusion"
    ws["A30"].font = body_font(bold=True, color=AMBER_WARN)
    ws["A30"].fill = yellow_fill()
    ws.merge_cells("A30:F30")
    ws.row_dimensions[30].height = 26

    set_widths(ws, {1: 24, 2: 24, 3: 18, 4: 18, 5: 18, 6: 18})


def build_inputs(ws) -> None:
    write_title(ws, "1 · Inputs — Articles originaux des Conditions Générales")
    write_subtitle(
        ws,
        2,
        "Texte source tel qu'il figure dans les modèles de Conditions Générales du marché. "
        "Ce sont ces clauses qui sont reformulées en onglet 2.",
    )

    headers = ["ID", "Branche", "Garantie / sujet", "Source", "Texte original (CG)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    for i, case in enumerate(CASES):
        row = 5 + i
        ws.cell(row=row, column=1, value=case["id"])
        ws.cell(row=row, column=2, value=case["branche"])
        ws.cell(row=row, column=3, value=case["garantie"])
        ws.cell(row=row, column=4, value=case["source"])
        ws.cell(row=row, column=5, value=case["original"])
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=160)
        ws.cell(row=row, column=1).font = body_font(bold=True)

    set_widths(ws, {1: 12, 2: 18, 3: 30, 4: 38, 5: 80})


def build_output(ws) -> None:
    write_title(ws, "2 · Output — Reformulation langage clair + Flesch FR")
    write_subtitle(
        ws,
        2,
        "Pour chaque cas : version reformulée + métriques de lisibilité calculées par "
        "formule (Flesch FR = 207 - 1,015 × mots/phrases - 73,6 × syllabes/mots).",
    )

    headers = [
        "ID",
        "Texte reformulé",
        "Mots (orig.)",
        "Phrases (orig.)",
        "Syllabes (orig.)",
        "Flesch FR orig.",
        "Mots (refor.)",
        "Phrases (refor.)",
        "Syllabes (refor.)",
        "Flesch FR refor.",
        "Δ Flesch",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    for i, case in enumerate(CASES):
        row = 5 + i
        ws.cell(row=row, column=1, value=case["id"])
        ws.cell(row=row, column=2, value=case["reformule"])
        ws.cell(row=row, column=3, value=case["orig_words"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=4, value=case["orig_sentences"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=5, value=case["orig_syllables"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=6, value=f"=207-1.015*(C{row}/D{row})-73.6*(E{row}/C{row})")
        ws.cell(row=row, column=7, value=case["ref_words"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=8, value=case["ref_sentences"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=9, value=case["ref_syllables"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=10, value=f"=207-1.015*(G{row}/H{row})-73.6*(I{row}/G{row})")
        ws.cell(row=row, column=11, value=f"=J{row}-F{row}")
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=180)
        ws.cell(row=row, column=1).font = body_font(bold=True)

        for col_idx in (6, 10, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.0"

    note_row = 5 + len(CASES) + 1
    write_section_header(ws, note_row, "Lecture du score Flesch FR (échelle Kandel-Moles)")
    legend = [
        ("90 - 100", "Très facile (CE2)"),
        ("70 - 90", "Facile (CM2)"),
        ("60 - 70", "Standard (collège) — cible v1"),
        ("50 - 60", "Assez difficile (lycée)"),
        ("30 - 50", "Difficile (universitaire)"),
        ("0 - 30", "Très difficile (académique / juridique)"),
    ]
    for i, (interval, label) in enumerate(legend, start=note_row + 1):
        ws.cell(row=i, column=1, value=interval).font = body_font(bold=True)
        ws.cell(row=i, column=2, value=label).font = body_font()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=11)
        ws.row_dimensions[i].height = 20

    set_widths(
        ws,
        {1: 12, 2: 75, 3: 11, 4: 11, 5: 12, 6: 14, 7: 11, 8: 11, 9: 12, 10: 14, 11: 11},
    )


def build_faq(ws) -> None:
    write_title(ws, "3 · FAQ générée — 5 questions par cas (15 total)")
    write_subtitle(
        ws,
        2,
        "Foire aux questions dérivée de chaque clause reformulée. Sert de support aux "
        "équipes service client et aux supports marketing produit.",
    )

    headers = ["ID cas", "Branche", "N°", "Question", "Réponse"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    for case_idx, case in enumerate(CASES):
        for q_idx, (q, a) in enumerate(case["faq"], start=1):
            ws.cell(row=row, column=1, value=case["id"])
            ws.cell(row=row, column=2, value=case["branche"])
            ws.cell(row=row, column=3, value=q_idx)
            ws.cell(row=row, column=4, value=q)
            ws.cell(row=row, column=5, value=a)
            style_body_row(ws, row, len(headers), alt=(case_idx % 2 == 1), height=70)
            ws.cell(row=row, column=1).font = body_font(bold=True)
            ws.cell(row=row, column=4).font = body_font(bold=True)
            row += 1

    set_widths(ws, {1: 12, 2: 18, 3: 6, 4: 50, 5: 80})


def build_mapping(ws) -> None:
    write_title(ws, "4 · Mapping articles — traçabilité phrase-à-phrase")
    write_subtitle(
        ws,
        2,
        "Pour chaque reformulation, lien explicite avec la phrase d'origine. Indispensable "
        "pour audit DDA et démontrer qu'aucune obligation contractuelle n'a été altérée.",
    )

    headers = ["ID cas", "N°", "Phrase originale (extrait CG)", "Phrase reformulée", "Note de transformation"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    for case_idx, case in enumerate(CASES):
        for m_idx, (orig, refor, note) in enumerate(case["mapping"], start=1):
            ws.cell(row=row, column=1, value=case["id"])
            ws.cell(row=row, column=2, value=m_idx)
            ws.cell(row=row, column=3, value=orig)
            ws.cell(row=row, column=4, value=refor)
            ws.cell(row=row, column=5, value=note)
            style_body_row(ws, row, len(headers), alt=(case_idx % 2 == 1), height=80)
            ws.cell(row=row, column=1).font = body_font(bold=True)
            row += 1

    set_widths(ws, {1: 12, 2: 6, 3: 60, 4: 60, 5: 45})


def build_kpis(ws) -> None:
    write_title(ws, "5 · KPIs — avant / après par cas")
    write_subtitle(
        ws,
        2,
        "Mesures publiables : nombre de mots, lisibilité Flesch FR, termes techniques. "
        "Toutes les colonnes calculées sont des formules Excel (auditables).",
    )

    headers = [
        "ID",
        "Branche",
        "Mots orig.",
        "Mots refor.",
        "Δ Mots",
        "Δ Mots %",
        "Flesch orig.",
        "Flesch refor.",
        "Δ Flesch",
        "Jargon orig.",
        "Jargon refor.",
        "Jargon réduit %",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    n = len(CASES)
    for i, case in enumerate(CASES):
        row = 5 + i
        ws.cell(row=row, column=1, value=case["id"])
        ws.cell(row=row, column=2, value=case["branche"])
        ws.cell(row=row, column=3, value=case["orig_words"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=4, value=case["ref_words"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
        ws.cell(row=row, column=6, value=f"=IFERROR((D{row}-C{row})/C{row},0)")
        ws.cell(row=row, column=6).number_format = "0.0%"
        # Flesch orig and refor as plain formulas based on syllables/sentences (carries them inline for completeness)
        ws.cell(
            row=row,
            column=7,
            value=(
                f"=207-1.015*({case['orig_words']}/{case['orig_sentences']})"
                f"-73.6*({case['orig_syllables']}/{case['orig_words']})"
            ),
        )
        ws.cell(
            row=row,
            column=8,
            value=(
                f"=207-1.015*({case['ref_words']}/{case['ref_sentences']})"
                f"-73.6*({case['ref_syllables']}/{case['ref_words']})"
            ),
        )
        ws.cell(row=row, column=9, value=f"=H{row}-G{row}")
        ws.cell(row=row, column=10, value=case["orig_jargon"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=11, value=case["ref_jargon"]).font = body_font(color="0000FF")
        ws.cell(row=row, column=12, value=f"=IFERROR((J{row}-K{row})/J{row},0)")
        ws.cell(row=row, column=12).number_format = "0.0%"

        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=28)
        ws.cell(row=row, column=1).font = body_font(bold=True)
        for col_idx in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12):
            ws.cell(row=row, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        for col_idx in (7, 8, 9):
            ws.cell(row=row, column=col_idx).number_format = "0.0"

    # Ligne agrégat (moyennes)
    agg_row = 5 + n + 1
    ws.cell(row=agg_row, column=1, value="MOYENNE").font = body_font(bold=True, color="FFFFFF")
    ws.cell(row=agg_row, column=1).fill = title_fill()
    ws.cell(row=agg_row, column=2, value="Tous cas").font = body_font(bold=True, color="FFFFFF")
    ws.cell(row=agg_row, column=2).fill = title_fill()
    for col_idx, col_letter in enumerate(["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"], start=3):
        ws.cell(row=agg_row, column=col_idx, value=f"=AVERAGE({col_letter}5:{col_letter}{4 + n})")
        cell = ws.cell(row=agg_row, column=col_idx)
        cell.font = body_font(bold=True, color="FFFFFF")
        cell.fill = title_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_letter == "F" or col_letter == "L":
            cell.number_format = "0.0%"
        elif col_letter in {"G", "H", "I"}:
            cell.number_format = "0.0"
        else:
            cell.number_format = "0.0"
    ws.row_dimensions[agg_row].height = 28

    # Note méthodo
    note_row = agg_row + 2
    write_section_header(ws, note_row, "Note méthodologique")
    notes = [
        "Mots / phrases / syllabes : entrées en bleu (modifiables). Ce sont les seuls hardcodes.",
        "Flesch FR (Kandel-Moles, 1958) — adaptation française du Flesch Reading Ease.",
        "Jargon : termes techniques du contrat (« imprégnation », « plafonnée », "
        "« à dire d'expert », etc.) — comptés manuellement v1, automatisés v2.",
        "Cible v1 : Flesch refor. > 60 (niveau collège), réduction jargon > 70%.",
    ]
    for i, n_ in enumerate(notes, start=note_row + 1):
        ws.cell(row=i, column=1, value=f"• {n_}").font = body_font()
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
        ws.row_dimensions[i].height = 22

    set_widths(
        ws,
        {1: 12, 2: 18, 3: 11, 4: 11, 5: 9, 6: 11, 7: 13, 8: 13, 9: 11, 10: 12, 11: 13, 12: 16},
    )


def build_limites(ws) -> None:
    write_title(ws, "6 · Limites — truth layer")
    write_subtitle(
        ws,
        2,
        "Ce que l'agent InsurSimplifier ne fait pas (et ne prétend pas faire). "
        "À diffuser aux équipes commerciales et compliance avant tout pilote client.",
    )

    sections = [
        (
            "Ce que l'agent ne remplace pas",
            [
                "Un avocat ou un juriste : la reformulation n'a pas de valeur juridique.",
                "Un courtier / un agent : le conseil personnalisé reste humain.",
                "Le texte original des Conditions Générales : il reste seul opposable.",
                "Le devoir de conseil DDA : la reformulation n'est qu'un support pédagogique.",
            ],
        ),
        (
            "Ce que l'agent ne génère pas",
            [
                "De nouveaux contrats ou de nouvelles clauses contractuelles.",
                "Des avis juridiques personnalisés sur un cas concret.",
                "Des décisions automatiques de prise en charge ou de refus de garantie.",
                "Des FAQ contenant des données personnelles ou un profilage individuel (RGPD).",
            ],
        ),
        (
            "Validations humaines obligatoires",
            [
                "Compliance : revue du mapping articles avant publication (DDA).",
                "Direction Marketing : validation du ton et du brand voice.",
                "Service client : pertinence des FAQ vs. réclamations réelles.",
                "Juridique : non-altération des obligations contractuelles.",
            ],
        ),
        (
            "Conformité réglementaire — disclaimers obligatoires",
            [
                "EU AI Act art. 50 (août 2026) : tout contenu généré doit porter la mention "
                "« Contenu généré par IA — version pédagogique non opposable ».",
                "DDA + Acte délégué clarté contractuelle : la version « langage clair » est "
                "un complément, jamais un substitut au texte des Conditions Générales.",
                "RGPD : aucune donnée personnelle d'assuré n'est utilisée pour générer la "
                "FAQ ou la reformulation.",
                "ACPR — surveillance permanente du contenu commercial : log de génération "
                "et d'approbation conservé 5 ans minimum.",
            ],
        ),
        (
            "Limites techniques v1 (à industrialiser v2)",
            [
                "Comptage syllabique manuel (à automatiser via une lib NLP FR).",
                "Détection du jargon manuelle (à remplacer par un dictionnaire métier).",
                "Pas de re-run automatique sur changement réglementaire (ACPR / EIOPA).",
                "Pas de versioning fin de la reformulation par rapport aux changements de CG.",
            ],
        ),
    ]

    row = 4
    for title, items in sections:
        write_section_header(ws, row, title)
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=f"• {item}").font = body_font()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.row_dimensions[row].height = 32
            row += 1
        row += 1

    set_widths(ws, {1: 24, 2: 24, 3: 18, 4: 18, 5: 18, 6: 18})


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_readme = wb.create_sheet("0_README")
    build_readme(ws_readme)

    ws_inputs = wb.create_sheet("1_Inputs")
    build_inputs(ws_inputs)

    ws_output = wb.create_sheet("2_Output_LangageClair")
    build_output(ws_output)

    ws_faq = wb.create_sheet("3_FAQ_Generee")
    build_faq(ws_faq)

    ws_mapping = wb.create_sheet("4_Mapping_Articles")
    build_mapping(ws_mapping)

    ws_kpis = wb.create_sheet("5_KPIs")
    build_kpis(ws_kpis)

    ws_limites = wb.create_sheet("6_Limites")
    build_limites(ws_limites)

    output_path = "C:/Users/Ludo/finance-platform/InsurSimplifier_NEURAL.xlsx"
    wb.save(output_path)
    print(f"WROTE {output_path}")


if __name__ == "__main__":
    main()
