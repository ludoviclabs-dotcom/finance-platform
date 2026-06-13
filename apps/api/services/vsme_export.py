"""
vsme_export.py — T3.3 : export du Rapport VSME (PDF + annexe Excel).

PDF : structure EFRAG (page de garde avec version du standard + complétude,
sections par module, tableaux quantitatifs et narratifs). Excel : un onglet par
module avec un hash par ligne (réutilise la couche T2). Les deux sont assemblés
dans un ZIP auditable (manifest + CHECKSUMS sha256sum -c), enregistré dans
export_packages (domaine 'vsme') pour /verify. Chaque valeur du rapport est
traçable vers son fact via le mapping (T3.2).

Les fonctions de rendu (build_vsme_pdf / build_vsme_xlsx) sont pures : elles
prennent un `mapping` (dict de compute_mapping) → testables sans DB.
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import re
import zipfile
from datetime import datetime, timezone
from typing import Any

from db.database import db_available, get_db
from services import vsme_catalog
from utils.excel_sanitize import sanitize_cell

logger = logging.getLogger(__name__)


# Date interne FIXE pour les métadonnées PDF/Excel → artefacts reproductibles
# (fpdf2/openpyxl embarquent sinon l'horodatage courant, qui casserait le hash).
_FIXED_META_DATE = datetime(2026, 1, 1, tzinfo=timezone.utc)


class VsmeExportError(Exception):
    """Erreur de génération du rapport VSME."""


def _sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


_CORE_DATE_RE = re.compile(
    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)"
)


def _normalize_xlsx(data: bytes) -> bytes:
    """Re-emballe un xlsx en octets REPRODUCTIBLES.

    openpyxl écrit l'heure courante (1) dans chaque entrée du zip interne et
    (2) dans docProps/core.xml (<dcterms:modified>, réécrit à la sauvegarde même
    si wb.properties est figé). On reconstruit le zip avec des entrées triées,
    une date d'entrée figée, et on neutralise les horodatages de core.xml.
    """
    src = zipfile.ZipFile(io.BytesIO(data))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in sorted(src.namelist()):
            content = src.read(name)
            if name == "docProps/core.xml":
                content = _CORE_DATE_RE.sub(rb"\g<1>2026-01-01T00:00:00Z\g<2>", content)
            info = zipfile.ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            zf.writestr(info, content)
    return out.getvalue()


def write_zip(entries: dict[str, bytes]) -> bytes:
    """Assemble un ZIP REPRODUCTIBLE (package_hash stable à contenu identique).

    Sans ZipInfo figée, zipfile.writestr inscrit l'heure courante dans chaque
    entrée → le package_hash varierait à chaque génération malgré un contenu
    figé. On trie les noms et fige la date d'entrée (même technique que
    _normalize_xlsx). Réutilisé par l'export BEGES.
    """
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in sorted(entries):
            info = zipfile.ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            zf.writestr(info, entries[name])
    return out.getvalue()


# fpdf2/Helvetica = latin-1 : translitère les caractères hors latin-1.
_PDF_MAP = str.maketrans({
    "—": "-", "–": "-", "’": "'", "‘": "'",
    "“": '"', "”": '"', "…": "...", "€": "EUR",
    "₂": "2", "₃": "3", "«": "<<", "»": ">>",
    "≥": ">=", "≤": "<=", "•": "*", " ": " ", " ": " ",
})


def _p(text: Any) -> str:
    return str(text).translate(_PDF_MAP).encode("latin-1", errors="replace").decode("latin-1")


def _fmt_value(row: dict[str, Any]) -> str:
    if row["status"] == "na":
        return "Non applicable"
    if row["value"] in (None, ""):
        return "Non renseigné"
    unit = f" {row['unit']}" if row.get("unit") else ""
    return f"{row['value']}{unit}"


def build_vsme_pdf(*, company_name: str, mapping: dict[str, Any], generated_at: str) -> bytes:
    """Génère le PDF du rapport VSME. Best-effort : lève VsmeExportError si fpdf2 absent."""
    try:
        from fpdf import FPDF
    except ImportError as exc:  # pragma: no cover
        raise VsmeExportError("fpdf2 non disponible pour la génération PDF.") from exc

    comp = mapping["completeness"]
    rows = mapping["datapoints"]
    by_module: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_module.setdefault(r["module"], []).append(r)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.creation_date = _FIXED_META_DATE  # reproductibilité (pas de now() dans le PDF)
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    # Page de garde
    pdf.set_font("Helvetica", "B", 22)
    pdf.cell(0, 14, _p("Rapport VSME"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, _p(company_name), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, _p(f"Standard : {vsme_catalog.standard_label()}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, _p(f"Référentiel : {mapping['version']} - Généré le {generated_at}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0, 6,
        _p(f"Complétude : {comp['overall_pct']}% ({comp['mandatory_filled']}/{comp['mandatory_total']} datapoints obligatoires)"),
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # Sections par module (ordre EFRAG)
    order = {m: i for i, m in enumerate(vsme_catalog.ALL_MODULES)}
    for module in sorted(by_module, key=lambda m: order.get(m, 99)):
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_fill_color(226, 232, 240)
        pdf.cell(0, 9, _p(f"Module {module}"), new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.ln(1)
        for r in by_module[module]:
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 5, _p(f"{r['code']} - {r['label']}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(60, 60, 60)
            src = f"  (source : {r['source']})" if r.get("source") and r["status"] == "auto" else ""
            pdf.multi_cell(0, 5, _p(f"   {_fmt_value(r)}{src}"), new_x="LMARGIN", new_y="NEXT")
            if r["status"] == "na" and r.get("na_justification"):
                pdf.set_text_color(120, 120, 120)
                pdf.multi_cell(0, 5, _p(f"   Justification : {r['na_justification']}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    return bytes(pdf.output())


def _row_hash(r: dict[str, Any]) -> str:
    canon = f"{r['code']}|{r['value']}|{r.get('unit') or ''}|{r['status']}"
    return _sha(canon.encode("utf-8"))


def build_vsme_xlsx(*, company_name: str, mapping: dict[str, Any]) -> bytes:
    """Génère l'annexe Excel : un onglet par module, hash par ligne."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise VsmeExportError("openpyxl non disponible pour la génération Excel.") from exc

    rows = mapping["datapoints"]
    by_module: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_module.setdefault(r["module"], []).append(r)

    wb = Workbook()
    # Propriétés figées → classeur reproductible (sinon openpyxl met now()).
    naive = _FIXED_META_DATE.replace(tzinfo=None)
    wb.properties.created = naive
    wb.properties.modified = naive
    synth = wb.active
    synth.title = "Synthèse"
    synth.append([sanitize_cell(v) for v in ["Entreprise", company_name]])
    synth.append([sanitize_cell(v) for v in ["Référentiel", mapping["version"]]])
    synth.append(["Complétude (%)", mapping["completeness"]["overall_pct"]])
    synth.append([])
    synth.append(["Module", "Obligatoires renseignés", "Total obligatoires", "%"])
    for m in mapping["completeness"]["modules"]:
        synth.append([m["module"], m["filled"], m["total"], m["pct"]])

    order = {m: i for i, m in enumerate(vsme_catalog.ALL_MODULES)}
    header = ["Code", "Datapoint", "Type", "Unité", "Valeur", "Statut", "Source", "Hash ligne (SHA-256)"]
    for module in sorted(by_module, key=lambda m: order.get(m, 99)):
        ws = wb.create_sheet(title=module[:31])
        ws.append(header)
        for r in by_module[module]:
            ws.append([
                sanitize_cell(r["code"]), sanitize_cell(r["label"]), sanitize_cell(r["type"]),
                sanitize_cell(r.get("unit") or ""),
                sanitize_cell(_fmt_value(r)), sanitize_cell(r["status"]),
                sanitize_cell(r.get("source") or ""), _row_hash(r),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return _normalize_xlsx(buf.getvalue())


def build_vsme_report(*, company_id: int, company_name: str, mapping: dict[str, Any] | None = None,
                      generated_at: datetime | None = None) -> dict[str, Any]:
    """Assemble le ZIP du rapport VSME (PDF + Excel + manifest + CHECKSUMS + README).

    `generated_at` permet de figer l'horodatage (reproductibilité du package_hash
    indépendamment de l'horloge ; le manifest_hash, lui, n'inclut jamais la date).
    """
    if mapping is None:
        from services import vsme_mapping_service
        mapping = vsme_mapping_service.compute_mapping(company_id)

    now = generated_at or datetime.now(tz=timezone.utc)
    pdf_bytes = build_vsme_pdf(company_name=company_name, mapping=mapping, generated_at=now.strftime("%d/%m/%Y"))
    xlsx_bytes = build_vsme_xlsx(company_name=company_name, mapping=mapping)

    # Manifest canonique (sans horodatage → reproductible).
    manifest = {
        "manifest_version": "v1",
        "report": "VSME",
        "standard_version": mapping["version"],
        "company_name": company_name,
        "completeness_pct": mapping["completeness"]["overall_pct"],
        "files": {
            "rapport-vsme.pdf": {"sha256": _sha(pdf_bytes), "size": len(pdf_bytes)},
            "annexe-vsme.xlsx": {"sha256": _sha(xlsx_bytes), "size": len(xlsx_bytes)},
        },
    }
    manifest_bytes = json.dumps(manifest, sort_keys=True, indent=2, ensure_ascii=False).encode("utf-8")
    manifest_hash = _sha(manifest_bytes)

    readme = (
        "Rapport VSME - CarbonCo\n=======================\n\n"
        f"Entreprise : {company_name}\nGenere le  : {now.isoformat()}\n"
        f"Hash manifest : {manifest_hash}\n\n"
        "Verifiez l'integrite : sha256sum -c CHECKSUMS.sha256\n"
        f"Enregistrement officiel : https://carbon-snowy-nine.vercel.app/verify/{manifest_hash}\n"
    ).encode("utf-8")

    embedded = {
        "manifest.json": manifest_bytes,
        "rapport-vsme.pdf": pdf_bytes,
        "annexe-vsme.xlsx": xlsx_bytes,
        "README.txt": readme,
    }
    checksums = ("\n".join(f"{_sha(d)}  {n}" for n, d in sorted(embedded.items())) + "\n").encode("utf-8")

    zip_bytes = write_zip({**embedded, "CHECKSUMS.sha256": checksums})
    package_hash = _sha(zip_bytes)
    filename = f"rapport-vsme-{company_id}-{now.strftime('%Y%m%d-%H%M%S')}-{package_hash[:12]}.zip"

    # Enregistrement dans export_packages (domaine 'vsme') pour /verify.
    if db_available():
        try:
            with get_db(company_id=company_id) as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO export_packages
                            (company_id, package_hash, manifest_hash, domain, filename,
                             size_bytes, event_count, frozen_count, meta)
                        VALUES (%s, %s, %s, 'vsme', %s, %s, 0, 0, %s)
                        ON CONFLICT (package_hash) DO NOTHING
                        """,
                        (company_id, package_hash, manifest_hash, filename, len(zip_bytes),
                         json.dumps({"company_name": company_name})),
                    )
        except Exception as exc:  # pragma: no cover
            logger.warning("Persist export_packages (vsme) échoué: %s", exc)

    return {
        "zip_bytes": zip_bytes,
        "filename": filename,
        "package_hash": package_hash,
        "manifest_hash": manifest_hash,
    }
