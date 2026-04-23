"""
NEURAL_AG003_LuxeEventComms.xlsx
Pack de contenus evenementiels multi-format — gate brand + heritage.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, STATUS, VERT, VIOLET, YESNO,
    ajouter_table, attach_dv, convertir_dates, D, DT, dv_liste, ecrire_donnees,
    en_tetes, params_table, param_ligne, section, titre, today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_AG003_LuxeEventComms.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "AG-003  —  LUXEEVENTCOMMS", "Pack evenementiel multi-format — defile, lancement, exposition", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 86

    section(ws, 4, 6, "Mission", NOIR)
    ecrire_donnees(ws, 5, [
        ["", "Role",        "Generer des packs evenementiels (invitation, script, social live, caption, teaser) coherents avec la charte."],
        ["", "Input",       "Brief evenement (type, date, audience, angle)."],
        ["", "Output",      "Pack multi-format + gate brand + requete heritage si necessaire."],
        ["", "Gate",        "EVENT (non blocking par defaut) + BRAND (blocking) + HERITAGE (si angle patrimoine)."],
        ["", "Appelants",   "USER direct, PR team."],
        ["", "Referentiel", "FOUNDATIONS!8_EVENTS_CALENDAR."],
    ], align=AL)

    section(ws, 13, 6, "Regles de completeness", BORDEAUX)
    ecrire_donnees(ws, 14, [
        ["", "Defile",       "Invitation VIP + caption social live + teaser + dossier presse OBLIGATOIRES."],
        ["", "Exhibition",   "Catalogue + wall text + parcours + invitation presse."],
        ["", "Launch",       "Teaser + post reveal + email client VIP + caption social."],
        ["", "Cocktail",     "Invitation + caption social + carte placement."],
        ["", "Client event", "Invitation personnalisee + plan + gift card."],
    ], align=AL)


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES AGENT", "tblParamsAG003", 3)
    rows = [
        ("AGENT_ID",        "AG-003", ""),
        ("AGENT_NAME",      "LuxeEventComms", ""),
        ("MAISON_ID",       "M-001", ""),
        ("MAISON_NAME",     "Maison pilote", ""),
        ("LANG_PRIMARY",    "FR", ""),
        ("LANG_SECONDARY",  "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", ""),
        ("DATA_VERSION",    "v2026.04", ""),
        ("PROMPT_VERSION",  "v1.0.0", ""),
        ("SLA_PACK_H",      72, "Delai livraison pack standard"),
        ("SLA_SOCIAL_LIVE_H", 2, "Delai social live temps reel"),
        ("CRISIS_MODE_ON",  "NO", ""),
        ("LAST_REFRESH",    D(today_iso()), ""),
    ]
    params_table(ws, "tblParamsAG003", start_row=4, rows=rows)


def onglet_event_briefs(wb):
    ws = wb.create_sheet("2_EVENT_BRIEFS")
    titre(ws, "EVENT BRIEFS", "Briefs evenements en cours", 13)
    en_tetes(ws, 4, ["EVENT_ID", "MAISON_ID", "NOM", "EVENT_TYPE", "DATE_DEBUT", "LIEU", "AUDIENCE",
                      "VIP_LEVEL", "HERITAGE_INTENT", "CLAIMS_PRESENT", "HERITAGE_REQUIRED", "CLAIMS_CHECK_REQUIRED", "PRIORITE"],
             [10, 10, 36, 16, 12, 22, 22, 10, 16, 16, 18, 20, 12])
    briefs = [
        ("EV-001", "M-001", "Defile haute couture printemps",         "Defile",       D("2026-07-05"), "Paris, Palais de Tokyo",   "VIP / Presse / Clients","HIGH",   "YES", "NO",  "", "", "P1"),
        ("EV-002", "M-001", "Exposition retrospective 100 ans",       "Exhibition",   D("2026-09-15"), "Musee des Arts Deco",      "Public / Presse",       "HIGH",   "YES", "NO",  "", "", "P1"),
        ("EV-003", "M-001", "Lancement parfum capsule automne",       "Launch",       D("2026-10-01"), "Paris, flagship",          "Presse / VIP",          "MEDIUM", "NO",  "YES", "", "", "P1"),
        ("EV-004", "M-001", "Cocktail Cannes 2026",                    "Cocktail",    D("2026-05-15"), "Cannes, hotel partenaire", "Celebrites / Presse",   "HIGH",   "NO",  "NO",  "", "", "P2"),
        ("EV-005", "M-001", "Salon joaillerie Biennale",               "Salon",       D("2026-09-22"), "Paris, Grand Palais",      "Collectionneurs",       "HIGH",   "YES", "YES", "", "", "P1"),
        ("EV-006", "M-001", "Collaboration artiste contemporain",      "Launch",      D("2026-11-10"), "Paris, galerie",           "Art + lifestyle",       "HIGH",   "YES", "YES", "", "", "P1"),
        ("EV-007", "M-001", "Client day Hong Kong",                    "Client event",D("2026-10-28"), "Hong Kong",                "Top clients Asie",      "HIGH",   "NO",  "NO",  "", "", "P1"),
        ("EV-008", "M-001", "Pop-up Tokyo Harajuku",                    "Pop-up",      D("2026-08-01"), "Tokyo",                    "Lifestyle jeune",       "MEDIUM", "YES", "YES", "", "", "P2"),
        ("EV-009", "M-001", "Week Chinese New Year activation",        "Activation",  D("2026-02-08"), "Shanghai, Pekin",          "Marche chinois",        "MEDIUM", "YES", "YES", "", "", "P2"),
    ]
    last = ecrire_donnees(ws, 5, briefs, align=AL)
    convertir_dates(ws, 5, last, date_cols=[5])

    # Heritage & claims required formulas
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=11, value=f'=IF(D{r}="Exhibition","YES",IF(I{r}="YES","YES","NO"))')
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = AC
        c.border = BF

        c = ws.cell(row=r, column=12, value=f'=IF(J{r}="YES","YES","NO")')
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = AC
        c.border = BF

    ajouter_table(ws, "tblEventBriefs", f"A4:M{last}")
    attach_dv(ws, f"D5:D{last}", ["Defile", "Exhibition", "Launch", "Cocktail", "Salon", "PR event", "Client event", "Pop-up", "Activation", "Conference", "Charity"])
    attach_dv(ws, f"H5:H{last}", ["HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"I5:I{last}", YESNO)
    attach_dv(ws, f"J5:J{last}", YESNO)
    attach_dv(ws, f"M5:M{last}", ["P0", "P1", "P2", "P3"])


def onglet_format_matrix(wb):
    ws = wb.create_sheet("3_FORMAT_MATRIX")
    titre(ws, "FORMAT MATRIX", "Formats attendus par type d'evenement (mandatory / optional)", 6)
    en_tetes(ws, 4, ["EVENT_TYPE", "FORMAT", "MANDATORY", "MANDATORY_FLAG", "OWNER", "NOTE"],
             [16, 26, 14, 18, 22, 36])
    fm = [
        ("Defile",       "Invitation VIP",        "YES", "", "PR Director",   "Envoi J-21"),
        ("Defile",       "Caption social live",    "YES", "", "Social Lead",   "Temps reel"),
        ("Defile",       "Teaser 30s video",       "YES", "", "Content Lead",  "J-3"),
        ("Defile",       "Dossier presse",         "YES", "", "PR Director",   "J-7"),
        ("Defile",       "Email client VIP",       "NO",  "", "CRM Lead",      "Optionnel"),
        ("Exhibition",   "Catalogue",              "YES", "", "Heritage Dir",  "Coedition musee"),
        ("Exhibition",   "Wall text",              "YES", "", "Heritage Dir",  "Multi-langue"),
        ("Exhibition",   "Parcours visite",        "YES", "", "Heritage Dir",  "FR + EN"),
        ("Exhibition",   "Invitation presse",      "YES", "", "PR Director",   "J-14"),
        ("Exhibition",   "Audio guide",            "NO",  "", "Heritage Dir",  "Optionnel"),
        ("Launch",       "Teaser",                 "YES", "", "Content Lead",  "J-7"),
        ("Launch",       "Post reveal",            "YES", "", "Social Lead",   "J-0"),
        ("Launch",       "Email client VIP",       "YES", "", "CRM Lead",      "J-3"),
        ("Launch",       "Caption social",          "YES", "", "Social Lead",   "J-0 a J+3"),
        ("Cocktail",     "Invitation",              "YES", "", "PR Director",   "J-10"),
        ("Cocktail",     "Caption social",          "YES", "", "Social Lead",   "Live"),
        ("Cocktail",     "Carte placement",         "YES", "", "PR Coord",      ""),
        ("Client event", "Invitation personnalisee","YES", "", "CRM Lead",      ""),
        ("Client event", "Plan / carte",            "YES", "", "CRM Lead",      ""),
        ("Client event", "Gift card",               "NO",  "", "CRM Lead",      "Optionnel"),
        ("Pop-up",       "Teaser",                  "YES", "", "Social Lead",   ""),
        ("Pop-up",       "Post reveal",             "YES", "", "Social Lead",   ""),
        ("Pop-up",       "Caption social",          "YES", "", "Social Lead",   ""),
        ("Salon",        "Catalogue",              "YES", "", "Heritage Dir",  ""),
        ("Salon",        "Invitation VIP",         "YES", "", "PR Director",   ""),
        ("Activation",   "Post social reveal",     "YES", "", "Social Lead",   "Marche-specific"),
        ("Activation",   "KOL script",              "YES", "", "Social Lead",   "Marche CN"),
    ]
    last = ecrire_donnees(ws, 5, fm, align=AL)
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=4, value=f'=IF(C{r}="YES",1,0)')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF
        c.number_format = "0"
    ajouter_table(ws, "tblFormatMatrix", f"A4:F{last}")


def onglet_content_packs(wb):
    ws = wb.create_sheet("4_CONTENT_PACKS")
    titre(ws, "CONTENT PACKS", "Etat de completion par evenement — formats livres / mandatory", 14)
    en_tetes(ws, 4, ["PACK_ID", "EVENT_ID", "EVENT_TYPE", "MAISON_ID", "LANG",
                      "INVITATION", "SOCIAL_LIVE", "TEASER", "DOSSIER_PRESSE", "EMAIL_VIP", "CATALOGUE",
                      "WALL_TEXT", "MANDATORY_TARGET", "MANDATORY_DELIVERED"],
             [12, 10, 16, 10, 8, 14, 14, 12, 16, 14, 12, 14, 18, 20])
    packs = [
        ("PCK-001", "EV-001", "Defile",       "M-001", "FR", "YES", "YES", "YES", "YES", "YES", "-",  "-",  "", ""),
        ("PCK-002", "EV-002", "Exhibition",   "M-001", "FR", "YES", "YES", "NO",  "YES", "NO",  "YES","YES","", ""),
        ("PCK-003", "EV-003", "Launch",       "M-001", "FR", "YES", "YES", "YES", "YES", "YES", "-",  "-",  "", ""),
        ("PCK-004", "EV-004", "Cocktail",     "M-001", "FR", "YES", "YES", "-",   "NO",  "-",   "-",  "-",  "", ""),
        ("PCK-005", "EV-005", "Salon",        "M-001", "FR", "YES", "YES", "-",   "YES", "YES", "YES","-",  "", ""),
        ("PCK-006", "EV-006", "Launch",       "M-001", "FR", "YES", "YES", "YES", "YES", "YES", "-",  "-",  "", ""),
        ("PCK-007", "EV-007", "Client event", "M-001", "EN", "YES", "NO",  "-",   "-",   "YES", "-",  "-",  "", ""),
        ("PCK-008", "EV-008", "Pop-up",       "M-001", "JA", "NO",  "YES", "YES", "NO",  "NO",  "-",  "-",  "", ""),
        ("PCK-009", "EV-009", "Activation",   "M-001", "ZH", "NO",  "YES", "YES", "NO",  "NO",  "-",  "-",  "", ""),
    ]
    last = ecrire_donnees(ws, 5, packs, align=AL)

    # MANDATORY_TARGET = SUMIFS(MANDATORY_FLAG WHERE EVENT_TYPE matches)
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=13, value=f'=SUMIFS(tblFormatMatrix[MANDATORY_FLAG],tblFormatMatrix[EVENT_TYPE],C{r})')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = "0"

        c = ws.cell(row=r, column=14, value=f'=COUNTIFS(F{r}:L{r},"YES")')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = "0"

    ajouter_table(ws, "tblContentPacks", f"A4:N{last}")


def onglet_heritage_requests(wb):
    ws = wb.create_sheet("5_HERITAGE_REQUESTS")
    titre(ws, "HERITAGE REQUESTS → AG-004", "Demandes blocs patrimoniaux", 8)
    en_tetes(ws, 4, ["HREQ_ID", "EVENT_ID", "QUERY", "SENT_AT", "AG004_STATUS", "BLOCK_ID_RETOUR", "SOURCE_STATUS", "COMMENT"],
             [12, 10, 40, 18, 16, 16, 16, 30])
    hr = [
        ("HR-001", "EV-001", "Motif XVI-line origine",             DT("2026-04-20 09:00"), "DONE",        "BL-005", "ACTIVE", "Bloc renvoye"),
        ("HR-002", "EV-002", "Histoire maison 1920-2025 complete", DT("2026-04-22 10:30"), "IN_PROGRESS", "",       "",       "Demande complexe"),
        ("HR-003", "EV-005", "Piece portee Cannes 1987",           DT("2026-04-25 14:00"), "DONE",        "BL-006", "ACTIVE", "Bloc renvoye"),
        ("HR-004", "EV-006", "Collaboration avec artiste — histoire precedente",DT("2026-04-28 11:15"), "IN_PROGRESS", "",  "",  "Research en cours"),
        ("HR-005", "EV-008", "Pop-up Tokyo 1972 (reference historique)", DT("2026-05-02 09:00"), "PENDING",     "",       "",       ""),
        ("HR-006", "EV-009", "Activation CNY 2018 (reference)",    DT("2026-05-02 09:30"), "PENDING",     "",       "",       ""),
    ]
    last = ecrire_donnees(ws, 5, hr, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[4])
    ajouter_table(ws, "tblHeritageRequests", f"A4:H{last}")
    attach_dv(ws, f"E5:E{last}", ["PENDING", "IN_PROGRESS", "DONE", "BLOCKED"])

    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"DONE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"IN_PROGRESS"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"PENDING"'], fill=PatternFill(start_color=FOND_BLEU, end_color=FOND_BLEU, fill_type="solid")))


def onglet_brand_gate_results(wb):
    ws = wb.create_sheet("6_BRAND_GATE_RESULTS")
    titre(ws, "BRAND GATE → AG-001", "Resultats gate brand par pack", 6)
    en_tetes(ws, 4, ["GATE_ID", "PACK_ID", "EVENT_ID", "BRAND_DECISION", "SCORE", "FEEDBACK"],
             [12, 10, 10, 16, 12, 50])
    gates = [
        ("GT-001", "PCK-001", "EV-001", "APPROVE", 92, "Conforme charte."),
        ("GT-002", "PCK-002", "EV-002", "APPROVE", 90, "OK, wall text valide."),
        ("GT-003", "PCK-003", "EV-003", "REWORK",  65, "Superlatifs dans teaser."),
        ("GT-004", "PCK-004", "EV-004", "APPROVE", 88, "Cocktail OK."),
        ("GT-005", "PCK-005", "EV-005", "APPROVE", 94, "Catalogue salon OK."),
        ("GT-006", "PCK-006", "EV-006", "APPROVE", 91, "Collaboration OK."),
        ("GT-007", "PCK-007", "EV-007", "APPROVE", 89, "Client event OK."),
        ("GT-008", "PCK-008", "EV-008", "REWORK",  70, "Caption JA a revoir."),
        ("GT-009", "PCK-009", "EV-009", "APPROVE", 85, "KOL script OK."),
    ]
    last = ecrire_donnees(ws, 5, gates, fmt_cols={5: "0"}, align=AL)
    ajouter_table(ws, "tblBrandGateEvents", f"A4:F{last}")

    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"APPROVE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"REWORK"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"REJECT"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_approval_pack(wb):
    ws = wb.create_sheet("7_APPROVAL_PACK")
    titre(ws, "APPROVAL PACK", "Consolidation brand + heritage -> ready/on hold", 8)
    en_tetes(ws, 4, ["APR_ID", "PACK_ID", "EVENT_ID", "BRAND_STATUS", "HERITAGE_STATUS", "CLAIMS_STATUS", "FINAL_STATUS", "DUE"],
             [12, 10, 10, 16, 18, 16, 16, 14])
    aprs = [
        ("APR-001", "PCK-001", "EV-001", "APPROVED", "APPROVED",    "-",     "", D("2026-07-01")),
        ("APR-002", "PCK-002", "EV-002", "APPROVED", "IN_PROGRESS", "-",     "", D("2026-09-10")),
        ("APR-003", "PCK-003", "EV-003", "REWORK",   "-",           "PASS",  "", D("2026-09-27")),
        ("APR-004", "PCK-004", "EV-004", "APPROVED", "-",           "-",     "", D("2026-05-13")),
        ("APR-005", "PCK-005", "EV-005", "APPROVED", "APPROVED",    "PASS",  "", D("2026-09-19")),
        ("APR-006", "PCK-006", "EV-006", "APPROVED", "IN_PROGRESS", "PASS",  "", D("2026-11-07")),
        ("APR-007", "PCK-007", "EV-007", "APPROVED", "-",           "-",     "", D("2026-10-26")),
        ("APR-008", "PCK-008", "EV-008", "REWORK",   "PENDING",     "PASS",  "", D("2026-07-28")),
        ("APR-009", "PCK-009", "EV-009", "APPROVED", "PENDING",     "PASS",  "", D("2026-02-05")),
    ]
    last = ecrire_donnees(ws, 5, aprs, align=AL)
    convertir_dates(ws, 5, last, date_cols=[8])

    for r in range(5, last + 1):
        c = ws.cell(row=r, column=7, value=f'=IF(AND(D{r}="APPROVED",OR(E{r}="APPROVED",E{r}="-"),OR(F{r}="PASS",F{r}="-")),"READY","ON_HOLD")')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ajouter_table(ws, "tblApprovalPack", f"A4:H{last}")

    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"READY"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"ON_HOLD"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_logs(wb):
    ws = wb.create_sheet("8_LOGS")
    titre(ws, "LOGS", "Journal — export vers MASTER", 14)
    en_tetes(ws, 4, ["RUN_ID", "AGENT_ID", "MAISON_ID", "DOC_ID", "TIMESTAMP", "LANG",
                      "FINAL_STATUS", "RISK_LEVEL", "GATE_TRIGGERED", "SLA_MET",
                      "TOTAL_MINUTES", "TOKENS_IN", "TOKENS_OUT", "COST_USD"],
             [16, 10, 10, 12, 18, 8, 14, 12, 14, 10, 12, 12, 12, 12])
    rows = [
        ("R-20260420-001", "AG-003", "M-001", "EV-001", DT("2026-04-20 11:05"), "FR", "APPROVED",   "LOW",    "EVENT", "YES", 12.4, 4100, 3200, 0.098),
        ("R-20260422-001", "AG-003", "M-001", "EV-002", DT("2026-04-22 10:30"), "FR", "IN_REVIEW",  "LOW",    "EVENT", "YES", 18.1, 5200, 4000, 0.128),
        ("R-20260423-001", "AG-003", "M-001", "EV-003", DT("2026-04-23 09:00"), "FR", "REWORK",     "MEDIUM", "EVENT", "YES", 11.2, 3600, 2700, 0.087),
        ("R-20260424-001", "AG-003", "M-001", "EV-004", DT("2026-04-24 15:00"), "FR", "APPROVED",   "LOW",    "EVENT", "YES", 7.5,  2400, 1900, 0.062),
        ("R-20260425-001", "AG-003", "M-001", "EV-005", DT("2026-04-25 14:00"), "FR", "APPROVED",   "LOW",    "EVENT", "YES", 14.0, 4500, 3500, 0.110),
        ("R-20260428-001", "AG-003", "M-001", "EV-006", DT("2026-04-28 11:15"), "FR", "IN_REVIEW",  "LOW",    "EVENT", "YES", 13.5, 4300, 3300, 0.105),
        ("R-20260501-001", "AG-003", "M-001", "EV-007", DT("2026-05-01 09:00"), "EN", "APPROVED",   "LOW",    "EVENT", "YES", 9.0,  2800, 2200, 0.072),
        ("R-20260502-001", "AG-003", "M-001", "EV-008", DT("2026-05-02 09:00"), "JA", "REWORK",     "MEDIUM", "EVENT", "YES", 16.0, 4800, 3700, 0.115),
        ("R-20260502-002", "AG-003", "M-001", "EV-009", DT("2026-05-02 09:30"), "ZH", "APPROVED",   "LOW",    "EVENT", "YES", 13.2, 4100, 3200, 0.100),
    ]
    last = ecrire_donnees(ws, 5, rows, fmt_cols={11: "0.00", 12: "0", 13: "0", 14: "0.000"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5])
    ajouter_table(ws, "tblEventLogs", f"A4:N{last}")


def onglet_dashboard(wb):
    ws = wb.create_sheet("9_DASHBOARD")
    titre(ws, "DASHBOARD EVENTS", "Completion packs, brand gate, cout", 4)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18

    section(ws, 4, 4, "KPIs", NOIR)
    kpis = [
        ("Packs ouverts",                      '=COUNTA(tblContentPacks[PACK_ID])',                                     "0"),
        ("Packs READY (approval)",             '=COUNTIFS(tblApprovalPack[FINAL_STATUS],"READY")',                       "0"),
        ("Packs ON_HOLD",                      '=COUNTIFS(tblApprovalPack[FINAL_STATUS],"ON_HOLD")',                     "0"),
        ("Brand APPROVE rate",                 '=IFERROR(COUNTIFS(tblBrandGateEvents[BRAND_DECISION],"APPROVE")/COUNTA(tblBrandGateEvents[GATE_ID]),0)', "0.0%"),
        ("Heritage requests DONE",             '=COUNTIFS(tblHeritageRequests[AG004_STATUS],"DONE")',                    "0"),
        ("Heritage requests PENDING/IP",       '=COUNTIFS(tblHeritageRequests[AG004_STATUS],"PENDING")+COUNTIFS(tblHeritageRequests[AG004_STATUS],"IN_PROGRESS")', "0"),
        ("Cout USD cumule",                    '=IFERROR(SUM(tblEventLogs[COST_USD]),0)',                                "0.000"),
    ]
    for i, (lbl, formula, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = fmt

    section(ws, 14, 4, "Completion par type d'evenement", BORDEAUX)
    types = ["Defile", "Exhibition", "Launch", "Cocktail", "Salon", "Client event", "Pop-up", "Activation"]
    for i, t in enumerate(types):
        r = 15 + i
        ws.cell(row=r, column=2, value=t).border = BF
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True)
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(tblContentPacks[EVENT_TYPE],"{t}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch = BarChart()
    ch.type = "bar"
    data = Reference(ws, min_col=3, min_row=15, max_row=14 + len(types))
    labels = Reference(ws, min_col=2, min_row=15, max_row=14 + len(types))
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(labels)
    ch.title = "Packs par type d'evenement"
    ch.height = 9
    ch.width = 14
    ws.add_chart(ch, "E14")


def onglet_learnings(wb):
    """R-13 : boucle d'apprentissage AG-003."""
    ws = wb.create_sheet("10_LEARNINGS")
    titre(ws, "LEARNINGS — AG-003", "Capitalisation : packs incomplets, brand rework, SLA social live", 8)
    en_tetes(ws, 4, ["LEARN_ID", "DATE", "SOURCE_RUN", "MOTIF", "FREQUENCE_MOIS", "ACTION_PROPOSEE", "OWNER", "STATUT"],
             [12, 12, 18, 40, 16, 44, 18, 14])
    lrn = [
        ("LRN-E-001", D(today_iso()), "R-20260423-001", "Teaser Launch souvent en rework (superlatif)",   3, "Ajouter template teaser sans superlatif",   "PR Director","OPEN"),
        ("LRN-E-002", D(today_iso()), "R-20260502-001", "Caption JA systematique 2e pass",                 2, "Seed vocab JA + revue native",              "Brand Dir",  "OPEN"),
        ("LRN-E-003", D(today_iso()), "-",              "SLA social live 2h depasse 3x en 1 mois",        3, "Pre-generation J-1 au lieu de live",        "Social Lead","OPEN"),
    ]
    last = ecrire_donnees(ws, 5, lrn, align=AL)
    convertir_dates(ws, 5, last, date_cols=[2])
    ajouter_table(ws, "tblLearningsAG003", f"A4:H{last}")
    attach_dv(ws, f"H5:H{last}", ["OPEN", "IN_REVIEW", "ACCEPTED", "REJECTED", "DONE"])


def build():
    wb = Workbook()
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_event_briefs(wb)
    onglet_format_matrix(wb)
    onglet_content_packs(wb)
    onglet_heritage_requests(wb)
    onglet_brand_gate_results(wb)
    onglet_approval_pack(wb)
    onglet_logs(wb)
    onglet_dashboard(wb)
    onglet_learnings(wb)

    order = ["0_README", "1_PARAMS", "2_EVENT_BRIEFS", "3_FORMAT_MATRIX",
             "4_CONTENT_PACKS", "5_HERITAGE_REQUESTS", "6_BRAND_GATE_RESULTS",
             "7_APPROVAL_PACK", "8_LOGS", "9_DASHBOARD", "10_LEARNINGS"]
    wb._sheets = [wb[n] for n in order]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
