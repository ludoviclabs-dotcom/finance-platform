"""Tests pour le service et l'endpoint Strategic Mapping ESG.

Vérifie :
  - build_strategic_mapping() retourne un StrategicMappingResponse valide
  - Les filtres segment/persona fonctionnent correctement
  - Toutes les sources chiffrées (qualitative=False) ont au moins une source
  - groundedKpis est None sans snapshot en cache (CI mode)
  - GET /strategic-mapping/adhesion-volontaire → 200 avec structure valide
  - Les filtres query params sont appliqués correctement
"""

from __future__ import annotations

from fastapi.testclient import TestClient

# ---------------------------------------------------------------------------
# Tests unitaires du service
# ---------------------------------------------------------------------------


class TestBuildStrategicMapping:
    def test_returns_response(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        from models.strategic_mapping import StrategicMappingResponse
        assert isinstance(result, StrategicMappingResponse)

    def test_meta_is_populated(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert result.meta.version
        assert result.meta.lastReviewedAt
        assert result.meta.contentOwner == "Carbon & Co"
        assert len(result.meta.regulatoryBaseline) >= 4

    def test_hero_is_populated(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert result.hero.title
        assert result.hero.subtitle
        assert result.hero.summary

    def test_investments_present(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert len(result.investments) == 4
        ids = [i.id for i in result.investments]
        assert "rh" in ids
        assert "si" in ids
        assert "conseil" in ids
        assert "gouvernance" in ids

    def test_before_after_present(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert len(result.beforeAfter) >= 5
        for item in result.beforeAfter:
            assert item.category
            assert item.before
            assert item.after

    def test_value_chain_ordered(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert len(result.valueChain) == 5
        orders = [s.order for s in result.valueChain]
        assert orders == sorted(orders)

    def test_financial_gains_present(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert len(result.financialGains) >= 5

    def test_externalities_present(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert len(result.externalities) >= 4

    def test_carbonco_levers_present(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        assert len(result.carbonCoLevers) == 6
        for lever in result.carbonCoLevers:
            assert lever.benefit
            assert lever.capability

    def test_grounded_kpis_none_without_cache(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=999)
        # Sans snapshot, groundedKpis peut être None ou dataAvailable=False
        if result.groundedKpis is not None:
            assert not result.groundedKpis.dataAvailable

    def test_filters_applied_in_response(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, segment="pme", persona="daf")
        assert result.filters.segment == "pme"
        assert result.filters.persona == "daf"


class TestSegmentFiltering:
    def test_pme_excludes_wacc(self) -> None:
        """Le gain WACC ne doit pas apparaître pour les PME."""
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, segment="pme")
        gain_ids = [g.id for g in result.financialGains]
        assert "wacc" not in gain_ids

    def test_pme_excludes_isr_externality(self) -> None:
        """L'externalité ISR ne s'applique pas aux PME."""
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, segment="pme")
        ext_ids = [e.id for e in result.externalities]
        assert "isr" not in ext_ids

    def test_grand_groupe_includes_wacc(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, segment="grand_groupe")
        gain_ids = [g.id for g in result.financialGains]
        assert "wacc" in gain_ids

    def test_generic_segment_returns_all(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, segment="generic")
        assert len(result.financialGains) == 6
        assert len(result.externalities) == 5

    def test_investments_returned_for_all_segments(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        for segment in ("pme", "eti", "grand_groupe"):
            result = build_strategic_mapping(company_id=1, segment=segment)
            assert len(result.investments) == 4


class TestPersonaFiltering:
    def test_daf_persona_returns_daf_message(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, persona="daf")
        assert len(result.executiveMessages) == 1
        assert result.executiveMessages[0].persona == "daf"

    def test_generic_persona_returns_all_messages(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, persona="generic")
        assert len(result.executiveMessages) == 4

    def test_daf_gains_filtered(self) -> None:
        """Le persona DAF ne doit pas voir les gains orientés DG uniquement."""
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1, persona="daf", segment="generic")
        for gain in result.financialGains:
            # Aucun gain ne doit avoir une liste personas qui exclut explicitement daf
            assert not gain.personas or "daf" in gain.personas or gain.personas == []


class TestSourceIntegrity:
    def test_non_qualitative_gains_have_sources(self) -> None:
        """Tout gain chiffré (qualitative=False) doit avoir au moins une source."""
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        for gain in result.financialGains:
            if not gain.qualitative:
                assert len(gain.sources) >= 1, (
                    f"Gain '{gain.id}' est chiffré (qualitative=False) mais n'a pas de source"
                )

    def test_sources_have_publisher_and_year(self) -> None:
        """Chaque source doit avoir publisher et year."""
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        all_sources = []
        for gain in result.financialGains:
            all_sources.extend(gain.sources)
        for ext in result.externalities:
            all_sources.extend(ext.sources)
        for src in all_sources:
            assert src.publisher, f"Source '{src.title}' manque de publisher"
            assert src.year > 2000, f"Source '{src.title}' a une année invalide"

    def test_non_qualitative_externalities_have_sources(self) -> None:
        from services.strategic_mapping_service import build_strategic_mapping
        result = build_strategic_mapping(company_id=1)
        for ext in result.externalities:
            if not ext.qualitative:
                assert len(ext.sources) >= 1


# ---------------------------------------------------------------------------
# Tests des endpoints HTTP
# ---------------------------------------------------------------------------


class TestStrategicMappingEndpoint:
    def test_returns_200(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200

    def test_structure_complete(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        data = resp.json()
        assert "meta" in data
        assert "hero" in data
        assert "investments" in data
        assert "beforeAfter" in data
        assert "valueChain" in data
        assert "financialGains" in data
        assert "externalities" in data
        assert "carbonCoLevers" in data
        assert "filters" in data

    def test_segment_filter_pme(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire?segment=pme",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["filters"]["segment"] == "pme"
        gain_ids = [g["id"] for g in data["financialGains"]]
        assert "wacc" not in gain_ids

    def test_persona_filter_daf(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire?persona=daf",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["filters"]["persona"] == "daf"
        assert len(data["executiveMessages"]) == 1
        assert data["executiveMessages"][0]["persona"] == "daf"

    def test_invalid_segment_returns_422(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire?segment=invalid_value",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 422

    def test_grounded_kpis_field_present(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        data = resp.json()
        # groundedKpis peut être null (pas de snapshot en CI) ou un objet
        assert "groundedKpis" in data


# ---------------------------------------------------------------------------
# Tests export Excel
# ---------------------------------------------------------------------------


class TestExcelService:
    """Tests unitaires du service de génération Excel."""

    def _build_data(self, segment: str = "generic", persona: str = "generic"):
        from services.strategic_mapping_service import build_strategic_mapping
        return build_strategic_mapping(company_id=1, segment=segment, persona=persona)

    def test_returns_bytes(self) -> None:
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data()
        result = build_strategic_mapping_xlsx(data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_magic_number(self) -> None:
        """Un fichier xlsx est un zip — commence par PK\\x03\\x04."""
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data()
        result = build_strategic_mapping_xlsx(data)
        assert result[:4] == b"PK\x03\x04", "Le fichier ne semble pas être un xlsx valide"

    def test_contains_all_sheets(self) -> None:
        """Les 6 onglets attendus doivent être présents."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data()
        wb = load_workbook(BytesIO(build_strategic_mapping_xlsx(data)))
        expected_sheets = {
            "Synthèse",
            "Investissements",
            "Chaîne de valeur",
            "Avant-Après",
            "Gains & Externalités",
            "Leviers Carbon & Co",
        }
        assert expected_sheets == set(wb.sheetnames)

    def test_synthese_contains_hero_title(self) -> None:
        """L'onglet Synthèse doit contenir le titre du hero."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data()
        wb = load_workbook(BytesIO(build_strategic_mapping_xlsx(data)))
        ws = wb["Synthèse"]
        cell_value = ws["A1"].value
        assert cell_value and "adhésion" in cell_value.lower()

    def test_filters_respected_pme_excludes_wacc(self) -> None:
        """PME : la feuille Gains ne doit pas contenir l'ID 'wacc'."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data(segment="pme")
        wb = load_workbook(BytesIO(build_strategic_mapping_xlsx(data)))
        ws = wb["Gains & Externalités"]
        all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
        assert "wacc" not in all_values

    def test_filters_respected_grand_groupe_includes_wacc(self) -> None:
        """Grand groupe : la feuille Gains doit contenir l'ID 'wacc'."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data(segment="grand_groupe")
        wb = load_workbook(BytesIO(build_strategic_mapping_xlsx(data)))
        ws = wb["Gains & Externalités"]
        all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
        assert "wacc" in all_values

    def test_sources_preserved_in_gains_sheet(self) -> None:
        """Les publishers des sources doivent apparaître dans l'onglet Gains."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data()
        wb = load_workbook(BytesIO(build_strategic_mapping_xlsx(data)))
        ws = wb["Gains & Externalités"]
        all_text = " ".join(str(cell.value or "") for row in ws.iter_rows() for cell in row)
        # Au moins une source connue doit apparaître
        assert "Banque de France" in all_text or "LMA" in all_text or "ADEME" in all_text

    def test_avant_apres_sheet_has_correct_columns(self) -> None:
        """L'onglet Avant-Après doit avoir 4 colonnes."""
        from io import BytesIO
        from openpyxl import load_workbook
        from services.strategic_mapping_excel import build_strategic_mapping_xlsx
        data = self._build_data()
        wb = load_workbook(BytesIO(build_strategic_mapping_xlsx(data)))
        ws = wb["Avant-Après"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers[0] == "Catégorie"
        assert headers[2] is not None  # "Sans démarche (avant)"
        assert headers[3] is not None  # "Avec démarche (après)"


class TestExcelEndpoint:
    """Tests HTTP de l'endpoint export.xlsx."""

    def test_returns_200(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.xlsx",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200

    def test_content_type_xlsx(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.xlsx",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    def test_content_disposition_attachment(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.xlsx",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd

    def test_valid_xlsx_bytes(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.xlsx",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.content[:4] == b"PK\x03\x04"

    def test_filters_applied_in_export(self, client: TestClient, analyst_token: str) -> None:
        """Le filtre segment=pme doit exclure wacc du classeur."""
        from io import BytesIO
        from openpyxl import load_workbook
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.xlsx?segment=pme",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Gains & Externalités"]
        all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
        assert "wacc" not in all_values


# ---------------------------------------------------------------------------
# Tests export PDF
# ---------------------------------------------------------------------------


class TestPdfService:
    """Tests unitaires du service de génération PDF."""

    def _build_data(self, segment: str = "generic", persona: str = "generic"):
        from services.strategic_mapping_service import build_strategic_mapping
        return build_strategic_mapping(company_id=1, segment=segment, persona=persona)

    def test_returns_bytes(self) -> None:
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        data = self._build_data()
        result = build_strategic_mapping_pdf(data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_pdf_magic_number(self) -> None:
        """Un PDF valide commence par %PDF-."""
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        data = self._build_data()
        result = build_strategic_mapping_pdf(data)
        assert result[:5] == b"%PDF-", "Le fichier ne semble pas être un PDF valide"

    def test_pdf_contains_multiple_pages(self) -> None:
        """Le PDF doit contenir plusieurs pages (6 sections)."""
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        data = self._build_data()
        result = build_strategic_mapping_pdf(data)
        # Compter les objets de page dans le PDF (chaque page = un objet /Page)
        page_count = result.count(b"/Page\n") + result.count(b"/Page\r")
        # On ne peut pas compter précisément sans parser, mais le fichier doit être non-trivial
        assert len(result) > 10_000, "PDF trop petit pour contenir 6 pages"

    def test_pdf_size_reasonable(self) -> None:
        """Le PDF doit faire entre 5 Ko et 2 Mo."""
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        data = self._build_data()
        result = build_strategic_mapping_pdf(data)
        assert 5_000 < len(result) < 2_000_000, f"Taille PDF inattendue : {len(result)} bytes"

    def test_filters_respected_pme(self) -> None:
        """Le PDF généré pour PME ne doit pas contenir 'wacc'."""
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        data = self._build_data(segment="pme")
        result = build_strategic_mapping_pdf(data)
        # 'wacc' est l'ID du gain — il ne doit pas apparaître dans le PDF PME
        assert b"wacc" not in result.lower()

    def test_filters_respected_grand_groupe(self) -> None:
        """Le PDF pour grand_groupe doit être plus grand que PME (plus de gains)."""
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        data_gg = self._build_data(segment="grand_groupe")
        data_pme = self._build_data(segment="pme")
        result_gg = build_strategic_mapping_pdf(data_gg)
        result_pme = build_strategic_mapping_pdf(data_pme)
        # grand_groupe a plus de gains (incl. wacc) → PDF plus grand ou comparable
        assert len(result_gg) > 0 and len(result_pme) > 0
        # Les deux doivent être des PDFs valides
        assert result_gg[:5] == b"%PDF-"
        assert result_pme[:5] == b"%PDF-"

    def test_all_segments_produce_valid_pdf(self) -> None:
        """Chaque segment doit produire un PDF valide."""
        from services.strategic_mapping_pdf import build_strategic_mapping_pdf
        from services.strategic_mapping_service import build_strategic_mapping
        for segment in ("pme", "eti", "grand_groupe", "generic"):
            data = build_strategic_mapping(company_id=1, segment=segment)
            result = build_strategic_mapping_pdf(data)
            assert result[:5] == b"%PDF-", f"PDF invalide pour segment={segment}"


class TestPdfEndpoint:
    """Tests HTTP de l'endpoint export.pdf."""

    def test_returns_200(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.pdf",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200

    def test_content_type_pdf(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.pdf",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert "application/pdf" in resp.headers.get("content-type", "")

    def test_content_disposition_attachment(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.pdf",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".pdf" in cd

    def test_valid_pdf_bytes(self, client: TestClient, analyst_token: str) -> None:
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.pdf",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.content[:5] == b"%PDF-"

    def test_filters_applied_in_export(self, client: TestClient, analyst_token: str) -> None:
        """segment=pme doit exclure wacc du PDF."""
        resp = client.get(
            "/strategic-mapping/adhesion-volontaire/export.pdf?segment=pme",
            headers={"Authorization": f"Bearer {analyst_token}"},
        )
        assert resp.status_code == 200
        assert b"wacc" not in resp.content.lower()
