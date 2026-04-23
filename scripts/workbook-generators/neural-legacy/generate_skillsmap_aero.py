#!/usr/bin/env python3
"""
NEURAL - SkillsMapAero
Generator for a consolidated 18-sheet Excel workbook focused on
aeronautical skills mapping, gap analysis, succession and training planning.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_NAME = "NEURAL_SkillsMapAero.xlsx"

BLACK = "1A1A2E"
GOLD = "D4AF37"
GOLD_LIGHT = "F5E6B8"
NAVY = "16213E"
GRAY_DARK = "2C2C2C"
GRAY_LIGHT = "F8F9FA"
WHITE = "FFFFFF"
RED = "E74C3C"
RED_LIGHT = "FADBD8"
ORANGE = "E67E22"
ORANGE_LIGHT = "FDEBD0"
YELLOW = "F1C40F"
YELLOW_LIGHT = "FEF9E7"
GREEN = "27AE60"
GREEN_LIGHT = "D5F5E3"
BLUE = "3498DB"
BLUE_LIGHT = "D6EAF8"
PURPLE = "8E44AD"

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
GOLD_BORDER = Border(
    left=Side(style="thin", color="555555"),
    right=Side(style="thin", color="555555"),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)

SITES = [
    "Villaroche (77)",
    "Gennevilliers (92)",
    "Chatellerault (86)",
    "Bordes (64)",
    "Istres (13)",
]

BUS = [
    "Production",
    "Maintenance (MRO)",
    "Bureau d'etudes",
    "Qualite",
    "Supply Chain",
    "HSE",
]

CATEGORIES = ["Operateur", "Technicien", "Ingenieur", "Cadre", "Manager"]

LEVELS = {
    0: ("N/A", "Non evalue", GRAY_LIGHT),
    1: ("Debutant", "Sensibilise, formation initiale suivie", RED_LIGHT),
    2: ("Operationnel guide", "Execute sous supervision directe", ORANGE_LIGHT),
    3: ("Autonome", "Maitrise en situation courante", YELLOW_LIGHT),
    4: ("Expert", "Gere situations complexes, forme les autres", GREEN_LIGHT),
    5: ("Referent", "Autorite reconnue, definit les standards", BLUE_LIGHT),
}

COMPETENCES = [
    ("COMP-001", "Soudure TIG aeronautique", "Procedes speciaux", "Soudure", 5, 15, 6000, 5, "AWS D17.1 / NF EN 4677", 50000, "Arret ligne assemblage"),
    ("COMP-002", "Soudure MIG/MAG aeronautique", "Procedes speciaux", "Soudure", 4, 10, 4500, 4, "AWS D17.1", 35000, "Sous-traitance obligatoire"),
    ("COMP-003", "Soudure faisceau electrons", "Procedes speciaux", "Soudure", 5, 20, 12000, 5, "AWS C7.1", 70000, "Arret production pieces critiques"),
    ("COMP-004", "Brasage aeronautique", "Procedes speciaux", "Brasage", 4, 8, 3500, 4, "AWS C3.6 / AMS 2675", 25000, "Retard assemblage"),
    ("COMP-005", "Ressuage PT Niveau 2", "CND/NDT", "Ressuage", 4, 10, 4000, 3, "NAS 410 / EN 4179", 28000, "Liberation pieces bloquee"),
    ("COMP-006", "Magnetoscopie MT Niveau 2", "CND/NDT", "Magnetoscopie", 4, 10, 4000, 3, "NAS 410 / EN 4179", 28000, "Liberation pieces bloquee"),
    ("COMP-007", "Ultrasons UT Niveau 2", "CND/NDT", "Ultrasons", 5, 15, 5500, 4, "NAS 410 / EN 4179", 42000, "Arret controle pieces forgees"),
    ("COMP-008", "Radiographie RT Niveau 2", "CND/NDT", "Radiographie", 5, 15, 6000, 5, "NAS 410 / EN 4179", 42000, "Arret controle soudures"),
    ("COMP-009", "Courants de Foucault ET N2", "CND/NDT", "Foucault", 4, 12, 5000, 4, "NAS 410 / EN 4179", 30000, "Controle surface degrade"),
    ("COMP-010", "Usinage CN 3 axes", "Usinage & fabrication", "Usinage", 3, 10, 3000, 2, "Interne", 18000, "Reaffectation possible"),
    ("COMP-011", "Usinage CN 5 axes", "Usinage & fabrication", "Usinage", 5, 20, 8000, 5, "Interne", 60000, "Arret production pieces complexes"),
    ("COMP-012", "Tournage CN", "Usinage & fabrication", "Tournage", 3, 8, 2500, 2, "Interne", 15000, "Reaffectation possible"),
    ("COMP-013", "Rectification de precision", "Usinage & fabrication", "Rectification", 4, 12, 4000, 4, "Interne", 26000, "Sous-traitance partielle"),
    ("COMP-014", "Electroerosion EDM", "Usinage & fabrication", "EDM", 4, 10, 4500, 4, "Interne", 30000, "Retard pieces speciales"),
    ("COMP-015", "Programmation FAO CATIA Mfg", "Usinage & fabrication", "FAO", 4, 15, 6000, 4, "Dassault certif", 38000, "Retard mise en production"),
    ("COMP-016", "Traitement thermique aero", "Traitements", "Thermique", 5, 12, 5000, 5, "AMS 2750 / NADCAP", 65000, "Arret four - toute production"),
    ("COMP-017", "Traitement surface anodisation", "Traitements", "Surface", 4, 8, 3000, 3, "NADCAP / Airbus AIMS", 22000, "Retard livraison"),
    ("COMP-018", "Shot peening controle", "Traitements", "Peening", 4, 10, 4500, 4, "AMS 2430 / NADCAP", 26000, "Sous-traitance couteuse"),
    ("COMP-019", "Controle metallurgique", "Traitements", "Metallurgie", 5, 15, 6000, 5, "ASTM E112 / interne", 45000, "Validation matiere bloquee"),
    ("COMP-020", "Conception CAO CATIA V5/V6", "Bureau d'etudes", "CAO", 4, 20, 7000, 3, "Dassault certif", 35000, "Retard conception"),
    ("COMP-021", "Conception NX Siemens", "Bureau d'etudes", "CAO", 3, 15, 6000, 3, "Siemens certif", 30000, "Retard conception"),
    ("COMP-022", "Calcul structures NASTRAN", "Bureau d'etudes", "Calcul", 5, 25, 10000, 5, "MSC certif / EASA", 80000, "Certification bloquee"),
    ("COMP-023", "Simulation thermique ABAQUS", "Bureau d'etudes", "Simulation", 5, 20, 9000, 5, "Dassault Systemes", 70000, "Certification bloquee"),
    ("COMP-024", "Cotation GPS GD&T aero", "Bureau d'etudes", "Cotation", 4, 5, 2000, 3, "ISO 1101 / ASME Y14.5", 18000, "Erreurs de fabrication"),
    ("COMP-025", "Gestion de configuration", "Bureau d'etudes", "CM", 3, 5, 1800, 2, "EN 9100 8.1", 12000, "NC audit mineur"),
    ("COMP-026", "Audit interne EN 9100", "Qualite & methodes", "Audit", 4, 5, 2500, 3, "EN 9100:2018 9.2", 18000, "Suspension agrement risquee"),
    ("COMP-027", "AMDEC/FMEA aeronautique", "Qualite & methodes", "AMDEC", 4, 3, 1500, 3, "SAE J1739 / ARP4761", 14000, "Analyse risque incomplete"),
    ("COMP-028", "Lean Manufacturing / 6 Sigma", "Qualite & methodes", "Lean", 3, 10, 4000, 2, "ASQ certif", 10000, "Pas d'impact immediat"),
    ("COMP-029", "Metrologie 3D MMT", "Qualite & methodes", "Metrologie", 4, 8, 3500, 3, "ISO 10360", 22000, "Controle dimensionnel bloque"),
    ("COMP-030", "First Article Inspection FAI", "Qualite & methodes", "FAI", 4, 3, 1200, 3, "AS9102", 16000, "Premiere piece non liberable"),
]

PRENOMS_H = [
    "Jean-Marc", "Pierre", "Thomas", "Nicolas", "Julien", "Antoine", "Sebastien",
    "Francois", "Christophe", "Laurent", "Mathieu", "David", "Philippe", "Eric",
    "Frederic", "Olivier", "Alexandre", "Guillaume", "Cedric", "Benoit",
    "Maxime", "Vincent", "Stephane", "Yannick", "Romain", "Fabien",
    "Arnaud", "Jerome", "Pascal", "Damien",
]
PRENOMS_F = [
    "Sophie", "Marie", "Claire", "Nathalie", "Isabelle", "Caroline", "Aurelie",
    "Celine", "Emilie", "Sandrine", "Valerie", "Anne", "Helene", "Laure",
    "Julie", "Camille", "Marion", "Delphine", "Pauline", "Stephanie",
    "Karine", "Virginie", "Florence", "Agathe", "Lucie", "Amandine",
    "Charlotte", "Sylvie", "Catherine", "Melanie",
]
NOMS_FAMILLE = [
    "DUPONT", "MARTIN", "BERNARD", "PETIT", "ROBERT", "RICHARD", "DURAND",
    "MOREAU", "LAURENT", "SIMON", "MICHEL", "LEFEVRE", "LEROY", "ROUX",
    "DAVID", "BERTRAND", "MOREL", "FOURNIER", "GIRARD", "BONNET",
    "LAMBERT", "FONTAINE", "ROUSSEAU", "MERCIER", "BLANC", "GUERIN",
    "MULLER", "HENRY", "BARBIER", "PERRIN",
]

POSTES_TYPES = [
    ("Soudeur TIG qualifie", "Operateur", "Production"),
    ("Soudeur polyvalent", "Operateur", "Production"),
    ("Operateur CND Niveau 2", "Technicien", "Qualite"),
    ("Operateur CND Niveau 3", "Technicien", "Qualite"),
    ("Operateur usinage CN 5 axes", "Operateur", "Production"),
    ("Operateur usinage CN 3 axes", "Operateur", "Production"),
    ("Technicien traitement thermique", "Technicien", "Production"),
    ("Technicien traitement surface", "Technicien", "Production"),
    ("Ingenieur conception CATIA", "Ingenieur", "Bureau d'etudes"),
    ("Ingenieur calcul structures", "Ingenieur", "Bureau d'etudes"),
    ("Responsable qualite PART 21", "Cadre", "Qualite"),
    ("Auditeur interne EN 9100", "Cadre", "Qualite"),
    ("Chef d'atelier production", "Manager", "Production"),
    ("Technicien metrologie", "Technicien", "Qualite"),
    ("Ingenieur methodes", "Ingenieur", "Production"),
    ("Technicien maintenance", "Technicien", "Maintenance (MRO)"),
    ("Ingenieur materiaux", "Ingenieur", "Bureau d'etudes"),
    ("Technicien FAO", "Technicien", "Production"),
    ("Responsable HSE", "Cadre", "HSE"),
    ("Coordinateur Supply Chain", "Cadre", "Supply Chain"),
]

PROFILS_POSTES = {
    "Soudeur TIG qualifie": {"COMP-001": 4, "COMP-002": 3, "COMP-005": 2, "COMP-024": 2, "COMP-027": 2, "COMP-028": 1},
    "Soudeur polyvalent": {"COMP-001": 3, "COMP-002": 4, "COMP-004": 3, "COMP-005": 2, "COMP-028": 1},
    "Operateur CND Niveau 2": {"COMP-005": 4, "COMP-006": 4, "COMP-007": 3, "COMP-008": 3, "COMP-009": 3, "COMP-030": 2},
    "Operateur CND Niveau 3": {"COMP-005": 5, "COMP-006": 5, "COMP-007": 5, "COMP-008": 4, "COMP-009": 4, "COMP-026": 3, "COMP-030": 3},
    "Operateur usinage CN 5 axes": {"COMP-011": 4, "COMP-010": 3, "COMP-015": 3, "COMP-029": 2, "COMP-024": 2},
    "Operateur usinage CN 3 axes": {"COMP-010": 4, "COMP-012": 3, "COMP-015": 2, "COMP-029": 2},
    "Technicien traitement thermique": {"COMP-016": 4, "COMP-019": 3, "COMP-029": 2, "COMP-027": 2, "COMP-028": 2},
    "Technicien traitement surface": {"COMP-017": 4, "COMP-018": 3, "COMP-019": 2, "COMP-029": 2},
    "Ingenieur conception CATIA": {"COMP-020": 5, "COMP-024": 4, "COMP-025": 3, "COMP-027": 3, "COMP-022": 2},
    "Ingenieur calcul structures": {"COMP-022": 5, "COMP-023": 4, "COMP-020": 3, "COMP-024": 3, "COMP-027": 3},
    "Responsable qualite PART 21": {"COMP-026": 5, "COMP-027": 4, "COMP-030": 4, "COMP-028": 3, "COMP-025": 3, "COMP-029": 3},
    "Auditeur interne EN 9100": {"COMP-026": 5, "COMP-027": 4, "COMP-028": 3, "COMP-030": 3, "COMP-025": 3},
    "Chef d'atelier production": {"COMP-028": 4, "COMP-027": 4, "COMP-026": 3, "COMP-030": 3, "COMP-025": 3, "COMP-024": 2},
    "Technicien metrologie": {"COMP-029": 5, "COMP-024": 4, "COMP-030": 3, "COMP-028": 2},
    "Ingenieur methodes": {"COMP-015": 4, "COMP-020": 3, "COMP-028": 3, "COMP-027": 3, "COMP-024": 3, "COMP-011": 2},
    "Technicien maintenance": {"COMP-010": 3, "COMP-012": 3, "COMP-005": 2, "COMP-006": 2, "COMP-028": 2},
    "Ingenieur materiaux": {"COMP-019": 5, "COMP-016": 4, "COMP-022": 3, "COMP-027": 3, "COMP-023": 2},
    "Technicien FAO": {"COMP-015": 4, "COMP-011": 3, "COMP-010": 3, "COMP-020": 2, "COMP-024": 2},
    "Responsable HSE": {"COMP-028": 4, "COMP-026": 3, "COMP-027": 3, "COMP-025": 2},
    "Coordinateur Supply Chain": {"COMP-025": 4, "COMP-028": 3, "COMP-030": 3, "COMP-027": 2},
}


def set_widths(ws, widths: dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def title_block(ws, title: str, subtitle: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c1 = ws.cell(1, 1, title)
    c1.font = Font("Calibri", 16, bold=True, color=GOLD)
    c1.fill = PatternFill("solid", BLACK)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2, 1, subtitle)
    c2.font = Font("Calibri", 10, italic=True, color=GOLD_LIGHT)
    c2.fill = PatternFill("solid", BLACK)
    c2.alignment = Alignment(horizontal="center", vertical="center")


def section_row(ws, row: int, text: str, ncols: int, fill: str = GOLD) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font("Calibri", 11, bold=True, color=BLACK if fill == GOLD else WHITE)
    c.fill = PatternFill("solid", fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def header_row(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        c = ws.cell(row, col, header)
        c.font = Font("Calibri", 10, bold=True, color=GOLD)
        c.fill = PatternFill("solid", BLACK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = GOLD_BORDER
    ws.row_dimensions[row].height = 30


def value_cell(ws, row: int, col: int, value, alt: bool = False, fmt: str | None = None):
    c = ws.cell(row, col, value)
    c.font = Font("Calibri", 10, color="333333")
    c.border = THIN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if alt:
        c.fill = PatternFill("solid", GRAY_LIGHT)
    if fmt:
        c.number_format = fmt
    return c


def formula_cell(ws, row: int, col: int, formula: str, alt: bool = False, fmt: str | None = None):
    return value_cell(ws, row, col, formula, alt=alt, fmt=fmt)


def criticality_label(score: int) -> str:
    if score >= 5:
        return "Critique"
    if score >= 4:
        return "Eleve"
    if score >= 3:
        return "Moyen"
    return "Standard"


def generate_employees(n: int = 60) -> list[tuple]:
    random.seed(42)
    employees = []
    all_prenoms = PRENOMS_H + PRENOMS_F
    for i in range(n):
        eid = f"EMP-{i + 1:03d}"
        nom = NOMS_FAMILLE[i % len(NOMS_FAMILLE)]
        prenom = all_prenoms[i % len(all_prenoms)]
        site = SITES[i % len(SITES)]
        poste, categorie, bu = POSTES_TYPES[i % len(POSTES_TYPES)]
        annee = random.randint(2010, 2024)
        mois = random.randint(1, 12)
        jour = random.randint(1, 28)
        date_entree = date(annee, mois, jour)
        manager_id = f"EMP-{max(1, ((i // 5) * 5) + 1):03d}" if i > 0 else ""
        contrat = random.choice(["CDI"] * 8 + ["CDD"] + ["Interimaire"])
        employees.append((eid, nom, prenom, site, bu, poste, categorie, date_entree, manager_id, contrat))
    return employees


EMPLOYEES = generate_employees(60)


def generate_matrice_data() -> list[dict]:
    random.seed(42)
    rows: list[dict] = []
    line_id = 0
    comp_lookup = {c[0]: c for c in COMPETENCES}
    for emp in EMPLOYEES:
        eid, nom, prenom, site, bu, poste, categorie, date_entree, manager_id, contrat = emp
        profil = PROFILS_POSTES.get(poste, {})
        for comp in COMPETENCES:
            cid = comp[0]
            req = profil.get(cid, 0)
            if req > 0 or random.random() < 0.08:
                anciennete = (date.today() - date_entree).days / 365
                if req > 0:
                    variation = random.choice([-1, -1, 0, 0, 0, 1, 1, 2])
                    if anciennete > 10:
                        variation += 1
                    elif anciennete < 2:
                        variation -= 1
                    level = max(1, min(5, req + variation))
                else:
                    level = random.randint(1, 3)
                eval_date = date.today() - timedelta(days=random.randint(30, 365))
                gap = level - req if req else None
                status = "Bonus" if req == 0 else ("Conforme" if gap >= 0 else "Gap")
                line_id += 1
                rows.append(
                    {
                        "row_id": f"MAT-{line_id:04d}",
                        "employee_id": eid,
                        "employee_name": f"{nom} {prenom}",
                        "site": site,
                        "poste": poste,
                        "comp_id": cid,
                        "comp_name": comp_lookup[cid][1],
                        "category": comp_lookup[cid][2],
                        "criticality_num": comp_lookup[cid][4],
                        "criticality_label": criticality_label(comp_lookup[cid][4]),
                        "level": level,
                        "required": req,
                        "gap": gap,
                        "status": status,
                        "source": random.choice(["manager", "auto-eval", "test"]),
                        "eval_date": eval_date,
                        "comment": "Expertise cle" if level >= 4 else "",
                    }
                )
    return rows


MATRICE_DATA = generate_matrice_data()


def build_01_parametres(wb: Workbook):
    ws = wb.active
    ws.title = "01_PARAMETRES"
    ws.sheet_properties.tabColor = GOLD
    set_widths(ws, {"A": 34, "B": 28, "C": 58})
    title_block(ws, "NEURAL - SkillsMapAero - Parametres", "Configuration de reference de l'agent", 3)

    rows = [
        ("Nom de l'entreprise", "Safran Aircraft Engines", "Reference d'entreprise"),
        ("Code entreprise", "SAF-AE", "Identifiant interne"),
        ("Secteur", "Aeronautique - Moteurs", "Activite principale"),
        ("Agrement EASA PART 21", "FR.21G.0049", "Numero d'agrement production"),
        ("Date de reference", datetime.now().strftime("%d/%m/%Y"), "Date de calcul"),
        ("Exercice", 2026, "Annee de pilotage"),
        ("Niveau minimum requis", 3, "Seuil de conformite"),
        ("Taux de couverture cible", 0.95, "Objectif global"),
        ("Nb minimum titulaires competence critique", 3, "Seuil d'alerte"),
        ("Objectif polyvalence", 5, "Nombre de competences qualifiees vise"),
        ("Budget formation global", 320000, "Budget annuel de reference"),
    ]

    section_row(ws, 4, "REFERENTIEL D'ENTREPRISE", 3)
    header_row(ws, 5, ["Parametre", "Valeur", "Description"])
    for idx, (name, value, desc) in enumerate(rows, 6):
        alt = idx % 2 == 0
        value_cell(ws, idx, 1, name, alt)
        fmt = "0%" if isinstance(value, float) and value <= 1 else "#,##0 €" if isinstance(value, int) and value > 10000 else None
        value_cell(ws, idx, 2, value, alt, fmt)
        value_cell(ws, idx, 3, desc, alt)

    start = 19
    section_row(ws, start, "SITES", 3)
    for offset, site in enumerate(SITES, 1):
        value_cell(ws, start + offset, 1, f"Site {offset}")
        value_cell(ws, start + offset, 2, site)
    start = 26
    section_row(ws, start, "BUSINESS UNITS", 3)
    for offset, bu in enumerate(BUS, 1):
        value_cell(ws, start + offset, 1, f"BU {offset}")
        value_cell(ws, start + offset, 2, bu)
    start = 34
    section_row(ws, start, "ECHELLE DE COMPETENCE", 3)
    header_row(ws, start + 1, ["Niveau", "Libelle", "Definition"])
    for idx, level in enumerate(range(1, 6), start + 2):
        label, desc, _ = LEVELS[level]
        value_cell(ws, idx, 1, level)
        value_cell(ws, idx, 2, label)
        value_cell(ws, idx, 3, desc)

    ws.freeze_panes = "A5"
    return ws


def build_02_ref_competences(wb: Workbook):
    ws = wb.create_sheet("02_REF_COMPETENCES")
    ws.sheet_properties.tabColor = NAVY
    widths = {get_column_letter(i): w for i, w in enumerate(
        [12, 34, 22, 16, 10, 12, 12, 10, 20, 16, 15, 15, 15, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16, 14], 1
    )}
    set_widths(ws, widths)
    title_block(ws, "Referentiel des competences aeronautiques", "Base catalogue et indicateurs de couverture", 24)
    headers = [
        "ID", "Competence", "Categorie", "Sous-cat.", "Criticite", "Duree form.", "Cout form.",
        "Rarete", "Norme", "Cout rupture", "Impact", "Niveau moyen", "Nb evals", "Criticite libelle",
        "Nb titulaires niv>=4", "Nb titulaires niv>=3", "Nb postes requis", "Ecart couvert.", "Alerte",
        "Invest. formation", "Indice rarete", "Indice risque", "Statut risque", "Taux couverture"
    ]
    header_row(ws, 4, headers)
    for idx, comp in enumerate(COMPETENCES, 5):
        alt = idx % 2 == 0
        cid, name, cat, subcat, crit, duration, cost, rarity, norm, break_cost, impact = comp
        value_cell(ws, idx, 1, cid, alt)
        value_cell(ws, idx, 2, name, alt)
        value_cell(ws, idx, 3, cat, alt)
        value_cell(ws, idx, 4, subcat, alt)
        value_cell(ws, idx, 5, crit, alt)
        value_cell(ws, idx, 6, duration, alt)
        value_cell(ws, idx, 7, cost, alt, "#,##0 €")
        value_cell(ws, idx, 8, rarity, alt)
        value_cell(ws, idx, 9, norm, alt)
        value_cell(ws, idx, 10, break_cost, alt, "#,##0 €")
        value_cell(ws, idx, 11, impact, alt)
        formula_cell(ws, idx, 12, f'=IFERROR(AVERAGEIF(\'04_MATRICE_SKILLS\'!F:F,A{idx},\'04_MATRICE_SKILLS\'!I:I),0)', alt, "0.00")
        formula_cell(ws, idx, 13, f'=COUNTIF(\'04_MATRICE_SKILLS\'!F:F,A{idx})', alt)
        value_cell(ws, idx, 14, criticality_label(crit), alt)
        formula_cell(ws, idx, 15, f'=COUNTIFS(\'04_MATRICE_SKILLS\'!F:F,A{idx},\'04_MATRICE_SKILLS\'!I:I,">="&4)', alt)
        formula_cell(ws, idx, 16, f'=COUNTIFS(\'04_MATRICE_SKILLS\'!F:F,A{idx},\'04_MATRICE_SKILLS\'!I:I,">="&3)', alt)
        formula_cell(ws, idx, 17, f'=COUNTIF(\'05_GAP_ANALYSIS\'!E:E,A{idx})', alt)
        formula_cell(ws, idx, 18, f'=P{idx}-Q{idx}', alt)
        formula_cell(ws, idx, 19, f'=IF(AND(E{idx}>=4,P{idx}<\'01_PARAMETRES\'!B9),"Critique",IF(P{idx}=0,"Alerte","Sous controle"))', alt)
        formula_cell(ws, idx, 20, f'=G{idx}*MAX(1,Q{idx}-P{idx})', alt, "#,##0 €")
        formula_cell(ws, idx, 21, f'=H{idx}*20', alt)
        formula_cell(ws, idx, 22, f'=ROUND((E{idx}*8)+(H{idx}*6)+(MAX(0,Q{idx}-P{idx})*10),0)', alt)
        formula_cell(ws, idx, 23, f'=IF(V{idx}>=70,"Critique",IF(V{idx}>=45,"Eleve",IF(V{idx}>=25,"Moyen","Faible")))', alt)
        formula_cell(ws, idx, 24, f'=IFERROR(P{idx}/MAX(Q{idx},1),0)', alt, "0.0%")

    ws.conditional_formatting.add("E5:E34", ColorScaleRule(start_type="num", start_value=1, start_color=GREEN, mid_type="num", mid_value=3, mid_color=YELLOW, end_type="num", end_value=5, end_color=RED))
    ws.conditional_formatting.add("X5:X34", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=BLUE))
    ws.auto_filter.ref = "A4:X34"
    ws.freeze_panes = "B5"
    return ws


def build_03_annuaire(wb: Workbook):
    ws = wb.create_sheet("03_ANNUAIRE")
    ws.sheet_properties.tabColor = GREEN
    set_widths(ws, {
        "A": 10, "B": 14, "C": 14, "D": 20, "E": 18, "F": 26, "G": 14, "H": 12, "I": 12,
        "J": 12, "K": 12, "L": 12, "M": 12, "N": 12, "O": 16, "P": 14, "Q": 14, "R": 14,
    })
    title_block(ws, "Annuaire du personnel", "Population de reference reliee a la matrice des competences", 18)
    headers = [
        "ID Employe", "Nom", "Prenom", "Site", "Business Unit", "Poste", "Categorie", "Date entree",
        "Manager ID", "Contrat", "Statut", "Anciennete", "Nb evals", "Niveau moyen", "Statut RH",
        "Nb competences qualifiees", "Score polyvalence", "Taux couverture"
    ]
    header_row(ws, 4, headers)

    dv_site = DataValidation(type="list", formula1='"' + ",".join(SITES) + '"')
    dv_bu = DataValidation(type="list", formula1='"' + ",".join(BUS) + '"')
    dv_cat = DataValidation(type="list", formula1='"' + ",".join(CATEGORIES) + '"')
    dv_stat = DataValidation(type="list", formula1='"Actif,Inactif,Suspendu"')
    for dv in [dv_site, dv_bu, dv_cat, dv_stat]:
        ws.add_data_validation(dv)

    for idx, emp in enumerate(EMPLOYEES, 5):
        alt = idx % 2 == 0
        eid, nom, prenom, site, bu, poste, categorie, date_entree, manager_id, contrat = emp
        value_cell(ws, idx, 1, eid, alt)
        value_cell(ws, idx, 2, nom, alt)
        value_cell(ws, idx, 3, prenom, alt)
        value_cell(ws, idx, 4, site, alt)
        value_cell(ws, idx, 5, bu, alt)
        value_cell(ws, idx, 6, poste, alt)
        value_cell(ws, idx, 7, categorie, alt)
        value_cell(ws, idx, 8, date_entree, alt, "DD/MM/YYYY")
        value_cell(ws, idx, 9, manager_id, alt)
        value_cell(ws, idx, 10, contrat, alt)
        value_cell(ws, idx, 11, "Actif", alt)
        formula_cell(ws, idx, 12, f'=DATEDIF(H{idx},TODAY(),"Y")', alt)
        formula_cell(ws, idx, 13, f'=COUNTIF(\'04_MATRICE_SKILLS\'!B:B,A{idx})', alt)
        formula_cell(ws, idx, 14, f'=IFERROR(AVERAGEIF(\'04_MATRICE_SKILLS\'!B:B,A{idx},\'04_MATRICE_SKILLS\'!I:I),0)', alt, "0.00")
        formula_cell(ws, idx, 15, f'=IF(K{idx}<>"Actif","Inactif",IF(N{idx}<2,"Sous-qualifie",IF(Q{idx}<0.6,"Polyvalence faible","OK")))', alt)
        formula_cell(ws, idx, 16, f'=COUNTIFS(\'04_MATRICE_SKILLS\'!B:B,A{idx},\'04_MATRICE_SKILLS\'!I:I,">="&3)', alt)
        formula_cell(ws, idx, 17, f'=IFERROR(P{idx}/\'01_PARAMETRES\'!B10,0)', alt, "0.0%")
        formula_cell(ws, idx, 18, f'=IFERROR(COUNTIFS(\'04_MATRICE_SKILLS\'!B:B,A{idx},\'04_MATRICE_SKILLS\'!L:L,"Conforme")/MAX(M{idx},1),0)', alt, "0.0%")
        dv_site.add(ws.cell(idx, 4))
        dv_bu.add(ws.cell(idx, 5))
        dv_cat.add(ws.cell(idx, 7))
        dv_stat.add(ws.cell(idx, 11))

    ws.conditional_formatting.add("N5:N64", ColorScaleRule(start_type="num", start_value=1, start_color=RED, mid_type="num", mid_value=3, mid_color=YELLOW, end_type="num", end_value=5, end_color=GREEN))
    ws.conditional_formatting.add("Q5:Q64", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.5, color=GOLD))
    ws.auto_filter.ref = "A4:R64"
    ws.freeze_panes = "D5"
    return ws


def build_04_matrice_skills(wb: Workbook):
    ws = wb.create_sheet("04_MATRICE_SKILLS")
    ws.sheet_properties.tabColor = RED
    set_widths(ws, {"A": 12, "B": 10, "C": 22, "D": 18, "E": 25, "F": 12, "G": 18, "H": 12, "I": 10, "J": 10, "K": 8, "L": 16, "M": 12, "N": 14, "O": 24})
    title_block(ws, "Matrice des competences", "Population x competence avec niveaux reels et requis", 15)
    headers = [
        "ID Ligne", "ID Employe", "Nom employe", "Site", "Poste", "ID Competence", "Categorie",
        "Criticite", "Niveau reel", "Niveau requis", "Gap", "Statut", "Source", "Date eval.", "Commentaire"
    ]
    header_row(ws, 4, headers)
    dv_level = DataValidation(type="whole", operator="between", formula1=0, formula2=5)
    ws.add_data_validation(dv_level)
    for idx, entry in enumerate(MATRICE_DATA, 5):
        alt = idx % 2 == 0
        value_cell(ws, idx, 1, entry["row_id"], alt)
        value_cell(ws, idx, 2, entry["employee_id"], alt)
        value_cell(ws, idx, 3, entry["employee_name"], alt)
        value_cell(ws, idx, 4, entry["site"], alt)
        value_cell(ws, idx, 5, entry["poste"], alt)
        value_cell(ws, idx, 6, entry["comp_id"], alt)
        value_cell(ws, idx, 7, entry["category"], alt)
        value_cell(ws, idx, 8, entry["criticality_label"], alt)
        value_cell(ws, idx, 9, entry["level"], alt)
        value_cell(ws, idx, 10, entry["required"] if entry["required"] else "", alt)
        formula_cell(ws, idx, 11, f'=IF(J{idx}="","",I{idx}-J{idx})', alt)
        formula_cell(ws, idx, 12, f'=IF(J{idx}="","Bonus",IF(K{idx}>=0,"Conforme",IF(K{idx}=-1,"Ecart mineur","Ecart critique")))', alt)
        value_cell(ws, idx, 13, entry["source"], alt)
        value_cell(ws, idx, 14, entry["eval_date"], alt, "DD/MM/YYYY")
        value_cell(ws, idx, 15, entry["comment"], alt)
        dv_level.add(ws.cell(idx, 9))

    last = 4 + len(MATRICE_DATA)
    ws.conditional_formatting.add(f"I5:I{last}", ColorScaleRule(start_type="num", start_value=1, start_color=RED, mid_type="num", mid_value=3, mid_color=YELLOW, end_type="num", end_value=5, end_color=GREEN))
    ws.conditional_formatting.add(f"K5:K{last}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", RED_LIGHT)))
    ws.conditional_formatting.add(f"L5:L{last}", FormulaRule(formula=['L5="Ecart critique"'], fill=PatternFill("solid", RED_LIGHT)))
    ws.auto_filter.ref = f"A4:O{last}"
    ws.freeze_panes = "E5"
    return ws


def build_05_gap_analysis(wb: Workbook):
    ws = wb.create_sheet("05_GAP_ANALYSIS")
    ws.sheet_properties.tabColor = ORANGE
    set_widths(ws, {"A": 10, "B": 10, "C": 20, "D": 18, "E": 12, "F": 11, "G": 11, "H": 14, "I": 12, "J": 12, "K": 16, "L": 28})
    title_block(ws, "Gap analysis", "Analyse des ecarts de competences et priorisation des actions", 12)
    headers = ["ID", "Employe ID", "Employe", "Site", "Competence ID", "Niv actuel", "Niv requis", "Severite", "Score", "Traitement", "Plan formation", "Action recommandee"]
    header_row(ws, 4, headers)
    rows = [r for r in MATRICE_DATA if r["required"] > 0]
    for i, row in enumerate(rows, 5):
        alt = i % 2 == 0
        value_cell(ws, i, 1, f"GAP-{i-4:04d}", alt)
        value_cell(ws, i, 2, row["employee_id"], alt)
        value_cell(ws, i, 3, row["employee_name"], alt)
        value_cell(ws, i, 4, row["site"], alt)
        value_cell(ws, i, 5, row["comp_id"], alt)
        value_cell(ws, i, 6, row["level"], alt)
        value_cell(ws, i, 7, row["required"], alt)
        formula_cell(ws, i, 8, f'=IF(F{i}-G{i}>=0,"🟢 OK",IF(F{i}-G{i}=-1,"🟡 Modere","🔴 Critique"))', alt)
        formula_cell(ws, i, 9, f'=MAX(0,(G{i}-F{i})*20)+IF(H{i}="🔴 Critique",25,0)', alt)
        formula_cell(ws, i, 10, f'=IF(H{i}="🟢 OK","Traite",IF(H{i}="🔴 Critique","Non traite","A planifier"))', alt)
        formula_cell(ws, i, 11, f'=IF(H{i}="🟢 OK","","PF-"&TEXT(ROW()-4,"0000"))', alt)
        formula_cell(ws, i, 12, f'=IF(H{i}="🟢 OK","Maintien",IF(H{i}="🔴 Critique","Formation urgente","Coaching interne"))', alt)

    last = 4 + len(rows)
    ws.conditional_formatting.add(f"I5:I{last}", ColorScaleRule(start_type="num", start_value=0, start_color=GREEN, mid_type="num", mid_value=30, mid_color=YELLOW, end_type="num", end_value=70, end_color=RED))
    ws.auto_filter.ref = f"A4:L{last}"
    ws.freeze_panes = "C5"
    return ws


def build_06_succession(wb: Workbook):
    ws = wb.create_sheet("06_SUCCESSION")
    ws.sheet_properties.tabColor = RED
    set_widths(ws, {"A": 10, "B": 25, "C": 18, "D": 14, "E": 12, "F": 22, "G": 14, "H": 12, "I": 14, "J": 12, "K": 16, "L": 14, "M": 14})
    title_block(ws, "Plans de succession", "Couverture des postes critiques et preparation des successeurs", 13)
    headers = ["ID", "Poste", "Site", "Titulaire ID", "Criticite", "Titulaire", "Successeur 1", "Pret 1", "Successeur 2", "Pret 2", "Risque depart", "Maj plan", "Plan action"]
    header_row(ws, 4, headers)
    critical_roles = [emp for emp in EMPLOYEES if emp[6] in ("Ingenieur", "Cadre", "Manager")][:20]
    random.seed(123)
    for idx, emp in enumerate(critical_roles, 5):
        alt = idx % 2 == 0
        s1 = EMPLOYEES[(idx + 2) % len(EMPLOYEES)][0]
        s2 = EMPLOYEES[(idx + 6) % len(EMPLOYEES)][0]
        value_cell(ws, idx, 1, f"SUC-{idx-4:03d}", alt)
        value_cell(ws, idx, 2, emp[5], alt)
        value_cell(ws, idx, 3, emp[3], alt)
        value_cell(ws, idx, 4, emp[0], alt)
        value_cell(ws, idx, 5, "Critique" if emp[6] in ("Manager", "Cadre") else "Eleve", alt)
        value_cell(ws, idx, 6, f"{emp[1]} {emp[2]}", alt)
        value_cell(ws, idx, 7, s1, alt)
        value_cell(ws, idx, 8, random.randint(45, 90) / 100, alt, "0%")
        value_cell(ws, idx, 9, s2, alt)
        value_cell(ws, idx, 10, random.randint(20, 75) / 100, alt, "0%")
        value_cell(ws, idx, 11, random.choice(["Faible", "Moyen", "Eleve", "Critique"]), alt)
        value_cell(ws, idx, 12, date.today() - timedelta(days=random.randint(10, 260)), alt, "DD/MM/YYYY")
        formula_cell(ws, idx, 13, f'=IF(K{idx}="Critique","Transfert immediat + mentoring",IF(H{idx}<0.6,"Programme 12 mois","Suivi semestriel"))', alt)
    ws.conditional_formatting.add("H5:H24", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GREEN))
    ws.auto_filter.ref = "A4:M24"
    ws.freeze_panes = "C5"
    return ws


def build_07_plan_formation(wb: Workbook):
    ws = wb.create_sheet("07_PLAN_FORMATION")
    ws.sheet_properties.tabColor = BLUE
    set_widths(ws, {"A": 10, "B": 26, "C": 10, "D": 20, "E": 18, "F": 14, "G": 14, "H": 12, "I": 12, "J": 12, "K": 16, "L": 16, "M": 14, "N": 14, "O": 14, "P": 14, "Q": 16, "R": 14, "S": 14})
    title_block(ws, "Plan de formation", "Actions de montee en competences deduites des gaps prioritaires", 19)
    headers = ["ID Plan", "Competence", "Employe ID", "Employe", "Site", "Priorite", "Date session", "Duree j", "Niv actuel", "Niv cible", "Modalite", "Organisme", "Budget prevu", "Budget engage", "Manager", "Cout complet", "Resultat", "Rappel", "Statut"]
    header_row(ws, 4, headers)
    gap_rows = [r for r in MATRICE_DATA if r["required"] > r["level"]][:80]
    random.seed(77)
    for idx, row in enumerate(gap_rows, 5):
        alt = idx % 2 == 0
        delta = row["required"] - row["level"]
        session_date = date.today() + timedelta(days=random.randint(15, 240))
        modality = "Formation externe" if delta >= 2 else random.choice(["Compagnonnage", "Coaching", "Formation interne"])
        planned = (delta * 1800) + random.randint(300, 1500)
        status = random.choice(["A planifier"] * 3 + ["Planifie"] * 3 + ["En cours"] * 2 + ["Realise"])
        value_cell(ws, idx, 1, f"PF-{idx-4:04d}", alt)
        value_cell(ws, idx, 2, row["comp_name"], alt)
        value_cell(ws, idx, 3, row["employee_id"], alt)
        value_cell(ws, idx, 4, row["employee_name"], alt)
        value_cell(ws, idx, 5, row["site"], alt)
        value_cell(ws, idx, 6, "Critique" if delta >= 2 else "Important", alt)
        value_cell(ws, idx, 7, session_date, alt, "DD/MM/YYYY")
        value_cell(ws, idx, 8, max(2, delta * 3), alt)
        value_cell(ws, idx, 9, row["level"], alt)
        value_cell(ws, idx, 10, row["required"], alt)
        value_cell(ws, idx, 11, modality, alt)
        value_cell(ws, idx, 12, "Organisme agree" if modality == "Formation externe" else "Academie interne", alt)
        value_cell(ws, idx, 13, planned, alt, "#,##0 €")
        formula_cell(ws, idx, 14, f'=IF(S{idx}="Realise",M{idx},IF(S{idx}="En cours",M{idx}*0.6,IF(S{idx}="Planifie",M{idx}*0.3,0)))', alt, "#,##0 €")
        formula_cell(ws, idx, 15, f'=IFERROR(VLOOKUP(C{idx},\'03_ANNUAIRE\'!A:I,9,FALSE),"")', alt)
        formula_cell(ws, idx, 16, f'=M{idx}+ROUND(M{idx}*0.15,0)', alt, "#,##0 €")
        formula_cell(ws, idx, 17, f'=IF(S{idx}="Realise","A reevaluer",IF(S{idx}="En cours","Progression","En attente"))', alt)
        formula_cell(ws, idx, 18, f'=IF(G{idx}<TODAY(),"Echeance depassee","A suivre")', alt)
        value_cell(ws, idx, 19, status, alt)
    last = 4 + len(gap_rows)
    ws.conditional_formatting.add(f"S5:S{last}", FormulaRule(formula=['S5="A planifier"'], fill=PatternFill("solid", RED_LIGHT)))
    ws.conditional_formatting.add(f"N5:N{last}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=12000, color=GOLD))
    ws.auto_filter.ref = f"A4:S{last}"
    ws.freeze_panes = "C5"
    return ws


def build_08_budget(wb: Workbook):
    ws = wb.create_sheet("08_BUDGET")
    ws.sheet_properties.tabColor = PURPLE
    set_widths(ws, {"A": 26, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 16})
    title_block(ws, "Budget formation", "Pilotage budgetaire par site et categorie", 8)
    headers = ["Perimetre", "Budget alloue", "Budget prevu", "Budget engage", "Reste a engager", "Taux conso", "Nb actions", "Alerte"]
    header_row(ws, 4, headers)
    scopes = [f"{site} - Formation" for site in SITES] + ["Production", "Qualite", "Bureau d'etudes", "Maintenance (MRO)", "HSE"]
    random.seed(9)
    for idx, scope in enumerate(scopes, 5):
        alt = idx % 2 == 0
        allocated = random.randint(18000, 65000)
        value_cell(ws, idx, 1, scope, alt)
        value_cell(ws, idx, 2, allocated, alt, "#,##0 €")
        formula_cell(ws, idx, 3, f'=SUMIF(\'07_PLAN_FORMATION\'!E:E,IFERROR(LEFT(A{idx},FIND(" -",A{idx})-1),A{idx}),\'07_PLAN_FORMATION\'!M:M)', alt, "#,##0 €")
        formula_cell(ws, idx, 4, f'=SUMIF(\'07_PLAN_FORMATION\'!E:E,IFERROR(LEFT(A{idx},FIND(" -",A{idx})-1),A{idx}),\'07_PLAN_FORMATION\'!N:N)', alt, "#,##0 €")
        formula_cell(ws, idx, 5, f'=B{idx}-D{idx}', alt, "#,##0 €")
        formula_cell(ws, idx, 6, f'=IFERROR(D{idx}/B{idx},0)', alt, "0.0%")
        formula_cell(ws, idx, 7, f'=COUNTIF(\'07_PLAN_FORMATION\'!E:E,IFERROR(LEFT(A{idx},FIND(" -",A{idx})-1),A{idx}))', alt)
        formula_cell(ws, idx, 8, f'=IF(F{idx}>1,"🔴 Depassement",IF(F{idx}>0.85,"🟡 Surveillance","🟢 OK"))', alt)
    last = 4 + len(scopes)
    ws.conditional_formatting.add(f"F5:F{last}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color=BLUE))
    ws.auto_filter.ref = f"A4:H{last}"
    ws.freeze_panes = "B5"
    return ws


def build_09_historique(wb: Workbook):
    ws = wb.create_sheet("09_HISTORIQUE")
    ws.sheet_properties.tabColor = GRAY_DARK
    set_widths(ws, {"A": 10, "B": 18, "C": 16, "D": 20, "E": 22, "F": 50})
    title_block(ws, "Historique des mises a jour", "Trace simplifiee des operations et changements du classeur", 6)
    header_row(ws, 4, ["ID", "Date", "Type", "Auteur", "Perimetre", "Commentaire"])
    events = [
        ("HIS-001", date.today() - timedelta(days=45), "Import", "Agent NEURAL", "Annuaire", "Chargement initial des 60 collaborateurs"),
        ("HIS-002", date.today() - timedelta(days=32), "Evaluation", "Manager qualite", "Matrice skills", "Mise a jour des niveaux CND et FAI"),
        ("HIS-003", date.today() - timedelta(days=21), "Budget", "RH formation", "Budget", "Reallocation budget site Bordes"),
        ("HIS-004", date.today() - timedelta(days=12), "Succession", "DRH", "Succession", "Ajout de 5 plans de succession"),
        ("HIS-005", date.today() - timedelta(days=3), "Controle", "Agent NEURAL", "Controles", "Verification integrite des references"),
    ]
    for idx, event in enumerate(events, 5):
        alt = idx % 2 == 0
        for col, value in enumerate(event, 1):
            fmt = "DD/MM/YYYY" if isinstance(value, date) else None
            value_cell(ws, idx, col, value, alt, fmt)
    ws.auto_filter.ref = "A4:F9"
    ws.freeze_panes = "A5"
    return ws


def build_10_dashboard_kpi(wb: Workbook):
    ws = wb.create_sheet("10_DASHBOARD_KPI")
    ws.sheet_properties.tabColor = GOLD
    set_widths(ws, {"A": 24, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16})
    title_block(ws, "Dashboard SkillsMapAero", f"Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}", 12)

    section_row(ws, 4, "INDICATEURS CLES", 12)
    header_row(ws, 5, ["KPI", "Valeur", "Objectif", "Statut"])
    kpis = [
        ("Taux couverture global", '=IFERROR(COUNTIF(\'04_MATRICE_SKILLS\'!L:L,"Conforme")/COUNTA(\'04_MATRICE_SKILLS\'!A:A),0)', 0.85, "0.0%"),
        ("Competences critiques couvertes", '=IFERROR(COUNTIFS(\'02_REF_COMPETENCES\'!N:N,"Critique",\'02_REF_COMPETENCES\'!X:X,">=0.8")/COUNTIF(\'02_REF_COMPETENCES\'!N:N,"Critique"),0)', 1.0, "0.0%"),
        ("Nb gaps critiques non traites", '=COUNTIFS(\'05_GAP_ANALYSIS\'!H:H,"🔴*",\'05_GAP_ANALYSIS\'!J:J,"Non traite")', 0, "0"),
        ("Score moyen competences", '=IFERROR(AVERAGE(\'04_MATRICE_SKILLS\'!I:I),0)', 3.5, "0.00"),
        ("Postes critiques avec successeur", '=IFERROR(COUNTIFS(\'06_SUCCESSION\'!E:E,"Critique",\'06_SUCCESSION\'!G:G,"<>")/COUNTIF(\'06_SUCCESSION\'!E:E,"Critique"),0)', 1.0, "0.0%"),
        ("Budget formation consomme", '=IFERROR(SUM(\'08_BUDGET\'!D:D)/SUM(\'08_BUDGET\'!B:B),0)', 0.9, "0.0%"),
        ("Employes evalues <12 mois", '=COUNTIFS(\'04_MATRICE_SKILLS\'!N:N,">="&TODAY()-365)', 50, "0"),
        ("Score qualite donnees", '=IFERROR(COUNTIF(\'11_CONTROLES\'!F5:F27,"✅*")/COUNTA(\'11_CONTROLES\'!F5:F27),0)', 0.9, "0.0%"),
    ]
    for idx, (label, formula, target, fmt) in enumerate(kpis, 6):
        alt = idx % 2 == 0
        value_cell(ws, idx, 1, label, alt)
        formula_cell(ws, idx, 2, formula, alt, fmt)
        value_cell(ws, idx, 3, target, alt, "0.0%" if isinstance(target, float) and target <= 1 else None)
        formula_cell(ws, idx, 4, f'=IF(B{idx}>=C{idx},"🟢 OK",IF(B{idx}>=C{idx}*0.8,"🟡 Sous cible","🔴 Alerte"))', alt)

    site_row = 17
    section_row(ws, site_row, "REPARTITION PAR SITE", 12)
    header_row(ws, site_row + 1, ["Site", "Effectif", "Score moyen", "Taux couverture", "Gaps critiques"])
    for pos, site in enumerate(SITES, site_row + 2):
        alt = pos % 2 == 0
        value_cell(ws, pos, 1, site, alt)
        formula_cell(ws, pos, 2, f'=COUNTIF(\'03_ANNUAIRE\'!D:D,A{pos})', alt)
        formula_cell(ws, pos, 3, f'=IFERROR(AVERAGEIF(\'04_MATRICE_SKILLS\'!D:D,A{pos},\'04_MATRICE_SKILLS\'!I:I),0)', alt, "0.00")
        formula_cell(ws, pos, 4, f'=IFERROR(COUNTIFS(\'04_MATRICE_SKILLS\'!D:D,A{pos},\'04_MATRICE_SKILLS\'!L:L,"Conforme")/MAX(COUNTIF(\'04_MATRICE_SKILLS\'!D:D,A{pos}),1),0)', alt, "0.0%")
        formula_cell(ws, pos, 5, f'=COUNTIFS(\'05_GAP_ANALYSIS\'!D:D,A{pos},\'05_GAP_ANALYSIS\'!H:H,"🔴*")', alt)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Score moyen par site"
    chart.y_axis.title = "Score"
    chart.height = 9
    chart.width = 14
    data = Reference(ws, min_col=3, min_row=site_row + 1, max_row=site_row + 1 + len(SITES))
    cats = Reference(ws, min_col=1, min_row=site_row + 2, max_row=site_row + 1 + len(SITES))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "G18")

    top_row = 26
    section_row(ws, top_row, "TOP 10 COMPETENCES A RISQUE", 12)
    header_row(ws, top_row + 1, ["Competence", "Score risque", "Priorite"])
    for rank in range(1, 11):
        row = top_row + 1 + rank
        alt = row % 2 == 0
        formula_cell(ws, row, 1, f'=IFERROR(INDEX(\'14_MATRICE_RISQUES\'!B:B,MATCH(LARGE(\'14_MATRICE_RISQUES\'!L$5:L$24,{rank}),\'14_MATRICE_RISQUES\'!L$5:L$24,0)+4),"")', alt)
        formula_cell(ws, row, 2, f'=IFERROR(LARGE(\'14_MATRICE_RISQUES\'!L$5:L$24,{rank}),0)', alt)
        formula_cell(ws, row, 3, f'=IF(B{row}>=70,"🔴 Critique",IF(B{row}>=50,"🟠 Haute",IF(B{row}>=30,"🟡 Moyenne","🟢 Basse")))', alt)

    ws.freeze_panes = "A5"
    return ws


def build_11_controles(wb: Workbook):
    ws = wb.create_sheet("11_CONTROLES")
    ws.sheet_properties.tabColor = "8B0000"
    set_widths(ws, {"A": 6, "B": 16, "C": 46, "D": 16, "E": 16, "F": 22, "G": 36})
    title_block(ws, "Controles de coherence", "Vingt-trois verifications de qualite de donnees", 7)
    ws.merge_cells("A3:D3")
    value_cell(ws, 3, 1, "Score qualite donnees")
    formula_cell(ws, 3, 5, '=ROUND((COUNTIF(F5:F27,"✅*")/COUNTA(F5:F27))*100,0)&"/100"')
    formula_cell(ws, 3, 6, '=IF(VALUE(LEFT(E3,FIND("/",E3)-1))=100,"🟢 Excellent",IF(VALUE(LEFT(E3,FIND("/",E3)-1))>=80,"🟡 Bon","🔴 Insuffisant"))')
    header_row(ws, 4, ["#", "Categorie", "Controle", "Severite", "Nb anomalies", "Resultat", "Detail"])

    controls = [
        ("Integrite", "Employes de la matrice presents dans l'annuaire", "Bloquant", '=SUMPRODUCT((COUNTIF(\'03_ANNUAIRE\'!A:A,\'04_MATRICE_SKILLS\'!B5:B500)=0)*(\'04_MATRICE_SKILLS\'!B5:B500<>""))', "Employes orphelins"),
        ("Integrite", "Competences de la matrice presentes dans le referentiel", "Bloquant", '=SUMPRODUCT((COUNTIF(\'02_REF_COMPETENCES\'!A:A,\'04_MATRICE_SKILLS\'!F5:F500)=0)*(\'04_MATRICE_SKILLS\'!F5:F500<>""))', "Competences non referencees"),
        ("Integrite", "Pas de doublon ID employe", "Bloquant", '=SUMPRODUCT((COUNTIF(\'03_ANNUAIRE\'!A5:A200,\'03_ANNUAIRE\'!A5:A200)>1)*(\'03_ANNUAIRE\'!A5:A200<>""))/2', "Doublons employes"),
        ("Integrite", "Pas de doublon ID competence", "Bloquant", '=SUMPRODUCT((COUNTIF(\'02_REF_COMPETENCES\'!A5:A200,\'02_REF_COMPETENCES\'!A5:A200)>1)*(\'02_REF_COMPETENCES\'!A5:A200<>""))/2', "Doublons competences"),
        ("Integrite", "Pas de doublon employe x competence", "Bloquant", '=SUMPRODUCT((COUNTIFS(\'04_MATRICE_SKILLS\'!B5:B500,\'04_MATRICE_SKILLS\'!B5:B500,\'04_MATRICE_SKILLS\'!F5:F500,\'04_MATRICE_SKILLS\'!F5:F500)>1)*(\'04_MATRICE_SKILLS\'!B5:B500<>""))/2', "Doublons matrice"),
        ("Coherence", "Niveaux entre 0 et 5", "Bloquant", '=COUNTIF(\'04_MATRICE_SKILLS\'!I:I,">5")+COUNTIF(\'04_MATRICE_SKILLS\'!I:I,"<0")', "Niveaux hors plage"),
        ("Coherence", "Niveaux requis renseignes sur lignes requises", "Important", '=COUNTIFS(\'04_MATRICE_SKILLS\'!J:J,"",\'04_MATRICE_SKILLS\'!L:L,"<>Bonus")', "Requis manquants"),
        ("Coherence", "Employes actifs avec au moins 1 evaluation", "Important", '=COUNTIFS(\'03_ANNUAIRE\'!K:K,"Actif",\'03_ANNUAIRE\'!M:M,0)', "Employes non evalues"),
        ("Coherence", "Competences critiques avec 0 titulaire qualifie", "Bloquant", '=COUNTIFS(\'02_REF_COMPETENCES\'!N:N,"Critique",\'02_REF_COMPETENCES\'!P:P,0)', "Ruptures critiques"),
        ("Coherence", "Surqualification systemique > 50%", "Attention", '=IF(COUNTIF(\'05_GAP_ANALYSIS\'!H:H,"🟢*")/MAX(COUNTA(\'05_GAP_ANALYSIS\'!H:H),1)<0.5,1,0)', "Possible tension globale"),
        ("Couverture", "Postes critiques avec successeur identifie", "Important", '=COUNTIFS(\'06_SUCCESSION\'!E:E,"Critique",\'06_SUCCESSION\'!G:G,"")', "Successeurs manquants"),
        ("Couverture", "Couverture critiques >= 80%", "Bloquant", '=COUNTIFS(\'02_REF_COMPETENCES\'!N:N,"Critique",\'02_REF_COMPETENCES\'!X:X,"<0.8")', "Competences critiques sous-couvertes"),
        ("Couverture", "Chaque site couvre des competences critiques", "Important", '=COUNTIF(\'10_DASHBOARD_KPI\'!E19:E23,0)', "Sites sans couverture critique"),
        ("Gaps", "Aucun gap critique non traite", "Bloquant", '=COUNTIFS(\'05_GAP_ANALYSIS\'!H:H,"🔴*",\'05_GAP_ANALYSIS\'!J:J,"Non traite")', "Gaps critiques ouverts"),
        ("Gaps", "Plan de formation sur gaps prioritaires", "Important", '=COUNTIFS(\'05_GAP_ANALYSIS\'!H:H,"🔴*",\'05_GAP_ANALYSIS\'!K:K,"")', "Gaps sans plan"),
        ("Budget", "Pas de depassement budgetaire", "Attention", '=COUNTIF(\'08_BUDGET\'!H:H,"*Depassement*")', "Lignes budget depassees"),
        ("Budget", "Budget engage coherent avec le plan", "Attention", '=IF(ABS(SUM(\'07_PLAN_FORMATION\'!N:N)-SUM(\'08_BUDGET\'!D:D))>100,1,0)', "Ecart entre budget et plan"),
        ("Temporel", "Evaluations < 12 mois", "Attention", '=COUNTIFS(\'04_MATRICE_SKILLS\'!N:N,"<"&TODAY()-365,\'04_MATRICE_SKILLS\'!N:N,"<>")', "Evaluations obsoletes"),
        ("Temporel", "Sessions passees non mises a jour", "Attention", '=COUNTIFS(\'07_PLAN_FORMATION\'!G:G,"<"&TODAY(),\'07_PLAN_FORMATION\'!S:S,"Planifie")', "Sessions a requalifier"),
        ("Qualite", "Champs obligatoires matrice complets", "Important", '=COUNTBLANK(\'04_MATRICE_SKILLS\'!B5:B500)', "ID employe manquants"),
        ("Qualite", "Justification sur niveaux >=4", "Attention", '=COUNTIFS(\'04_MATRICE_SKILLS\'!I:I,">="&4,\'04_MATRICE_SKILLS\'!O:O,"")', "Commentaires manquants"),
        ("Qualite", "Source d'evaluation renseignee", "Attention", '=COUNTIFS(\'04_MATRICE_SKILLS\'!M:M,"",\'04_MATRICE_SKILLS\'!I:I,">0")', "Sources manquantes"),
        ("Qualite", "Plans de succession recents", "Important", '=COUNTIFS(\'06_SUCCESSION\'!E:E,"Critique",\'06_SUCCESSION\'!L:L,"<"&TODAY()-180)', "Plans obsoletes"),
    ]

    for idx, (cat, desc, sev, formula, detail) in enumerate(controls, 5):
        alt = idx % 2 == 0
        value_cell(ws, idx, 1, idx - 4, alt)
        value_cell(ws, idx, 2, cat, alt)
        value_cell(ws, idx, 3, desc, alt)
        value_cell(ws, idx, 4, sev, alt)
        formula_cell(ws, idx, 5, formula, alt)
        formula_cell(ws, idx, 6, f'=IF(E{idx}=0,"✅ OK","❌ ECHEC ("&E{idx}&" anomalie(s))")', alt)
        value_cell(ws, idx, 7, detail, alt)

    ws.conditional_formatting.add("F5:F27", FormulaRule(formula=['LEFT(F5,1)="✅"'], fill=PatternFill("solid", GREEN_LIGHT)))
    ws.conditional_formatting.add("F5:F27", FormulaRule(formula=['LEFT(F5,1)="❌"'], fill=PatternFill("solid", RED_LIGHT)))
    ws.freeze_panes = "A5"
    return ws


def build_12_notes_reglementaires(wb: Workbook):
    ws = wb.create_sheet("12_NOTES_REGLEMENTAIRES")
    ws.sheet_properties.tabColor = NAVY
    set_widths(ws, {"A": 18, "B": 28, "C": 22, "D": 70})
    title_block(ws, "Notes reglementaires", "References normatives et points d'attention conformite", 4)
    header_row(ws, 4, ["Reference", "Theme", "Perimetre", "Note"])
    notes = [
        ("EN 9100", "Systeme qualite", "Qualite / production", "Cadre qualite aeronautique de reference pour audit, competences et tracabilite."),
        ("AS9102", "FAI", "Industrialisation", "Impose la robustesse des donnees de premiere piece et la maitrise documentaire."),
        ("NAS 410 / EN 4179", "CND/NDT", "Controle non destructif", "Qualification indispensable pour PT, MT, UT, RT et ET."),
        ("AMS 2750", "Traitement thermique", "Fours et pyrometrie", "Competence critique liee a la conformite NADCAP."),
        ("EASA PART 21", "Production", "Entreprise", "Competences critiques a piloter sur postes clefs et liberations."),
        ("ISO 1101 / ASME Y14.5", "Cotation", "Bureau d'etudes", "Base de convergence entre conception et fabrication."),
        ("ARP4761", "Analyse de risque", "Ingenierie", "Support aux analyses AMDEC/FMEA et plans de mitigation."),
    ]
    for idx, note in enumerate(notes, 5):
        alt = idx % 2 == 0
        for col, value in enumerate(note, 1):
            value_cell(ws, idx, col, value, alt)
    ws.freeze_panes = "A5"
    return ws


def build_13_export_rapport(wb: Workbook):
    ws = wb.create_sheet("13_EXPORT_RAPPORT")
    ws.sheet_properties.tabColor = GREEN
    set_widths(ws, {get_column_letter(i): 16 for i in range(1, 9)})
    title_block(ws, "NEURAL - Rapport cartographie des competences", "Support d'export PDF et synthese executive", 8)
    section_row(ws, 4, "SYNTHESE EXECUTIVE", 8)
    metrics = [
        ("Effectif total evalue", '=COUNTA(\'03_ANNUAIRE\'!A5:A64)'),
        ("Taux de couverture global", '=\'10_DASHBOARD_KPI\'!B6'),
        ("Competences critiques couvertes", '=\'10_DASHBOARD_KPI\'!B7'),
        ("Gaps critiques non traites", '=\'10_DASHBOARD_KPI\'!B8'),
        ("Score qualite donnees", '=\'10_DASHBOARD_KPI\'!B13'),
        ("Budget formation consomme", '=\'10_DASHBOARD_KPI\'!B11'),
    ]
    for idx, (label, formula) in enumerate(metrics, 6):
        value_cell(ws, idx, 1, label)
        formula_cell(ws, idx, 4, formula)

    section_row(ws, 14, "ALERTES CRITIQUES", 8, RED)
    alerts = [
        "Competences critiques a couverture inferieure a 80%",
        "Gaps critiques non traites",
        "Plans de succession manquants ou obsoletes",
        "Depassements budgetaires potentiels",
    ]
    for i, text in enumerate(alerts, 16):
        value_cell(ws, i, 1, f"- {text}")

    section_row(ws, 22, "RECOMMANDATIONS", 8, GREEN)
    actions = [
        "0-30 jours : traiter les gaps critiques et rafraichir les evaluations obsoletes.",
        "30-90 jours : lancer les formations prioritaires et renforcer les plans de succession.",
        "90-180 jours : harmoniser les pratiques inter-sites et industrialiser la polyvalence.",
    ]
    for i, text in enumerate(actions, 24):
        value_cell(ws, i, 1, text)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    return ws


def build_14_matrice_risques(wb: Workbook):
    ws = wb.create_sheet("14_MATRICE_RISQUES")
    ws.sheet_properties.tabColor = RED
    set_widths(ws, {"A": 12, "B": 34, "C": 20, "D": 10, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12, "M": 18, "N": 38})
    title_block(ws, "Matrice des risques competences", "Scoring multicritere sur les competences les plus sensibles", 14)
    header_row(ws, 4, ["ID", "Competence", "Categorie", "Criticite", "Tx couverture", "Nb titulaires", "Nb postes", "Substituabilite", "Impact", "Score brut", "Mitigation", "Score net", "Priorite", "Action recommandee"])
    for idx in range(5, 25):
        ref_row = idx
        alt = idx % 2 == 0
        formula_cell(ws, idx, 1, f'=IF(\'02_REF_COMPETENCES\'!A{ref_row}<>"",\'02_REF_COMPETENCES\'!A{ref_row},"")', alt)
        formula_cell(ws, idx, 2, f'=IF(A{idx}="","",\'02_REF_COMPETENCES\'!B{ref_row})', alt)
        formula_cell(ws, idx, 3, f'=IF(A{idx}="","",\'02_REF_COMPETENCES\'!C{ref_row})', alt)
        formula_cell(ws, idx, 4, f'=IF(A{idx}="","",\'02_REF_COMPETENCES\'!E{ref_row})', alt)
        formula_cell(ws, idx, 5, f'=IF(A{idx}="","",\'02_REF_COMPETENCES\'!X{ref_row})', alt, "0.0%")
        formula_cell(ws, idx, 6, f'=IF(A{idx}="","",\'02_REF_COMPETENCES\'!P{ref_row})', alt)
        formula_cell(ws, idx, 7, f'=IF(A{idx}="","",\'02_REF_COMPETENCES\'!Q{ref_row})', alt)
        formula_cell(ws, idx, 8, f'=IF(F{idx}<=1,5,IF(F{idx}<=3,4,IF(F{idx}<=5,3,IF(F{idx}<=10,2,1))))', alt)
        formula_cell(ws, idx, 9, f'=IF(D{idx}>=5,5,IF(D{idx}>=4,4,IF(D{idx}>=3,3,2)))', alt)
        formula_cell(ws, idx, 10, f'=ROUND(D{idx}*(5-MIN(E{idx},1)*5+1)*H{idx}*I{idx}/25*20,0)', alt)
        formula_cell(ws, idx, 11, f'=IF(COUNTIFS(\'07_PLAN_FORMATION\'!B:B,B{idx},\'07_PLAN_FORMATION\'!S:S,"Planifie")>0,"Oui","Non")', alt)
        formula_cell(ws, idx, 12, f'=ROUND(J{idx}*IF(K{idx}="Oui",0.6,1),0)', alt)
        formula_cell(ws, idx, 13, f'=IF(L{idx}>=70,"🔴 Critique",IF(L{idx}>=50,"🟠 Haute",IF(L{idx}>=30,"🟡 Moyenne",IF(L{idx}>=15,"🔵 Basse","🟢 Maitrise"))))', alt)
        formula_cell(ws, idx, 14, f'=IF(L{idx}>=70,"Recrutement + formation express",IF(L{idx}>=50,"Planifier formation sous 30 jours",IF(L{idx}>=30,"Integrer au plan semestriel","Surveillance normale")))', alt)
    ws.conditional_formatting.add("L5:L24", ColorScaleRule(start_type="num", start_value=0, start_color=GREEN, mid_type="num", mid_value=50, mid_color=YELLOW, end_type="num", end_value=100, end_color=RED))
    ws.freeze_panes = "A5"
    return ws


def build_15_simulation_departs(wb: Workbook):
    ws = wb.create_sheet("15_SIMULATION_DEPARTS")
    ws.sheet_properties.tabColor = PURPLE
    set_widths(ws, {"A": 14, "B": 24, "C": 18, "D": 26, "E": 14, "F": 16, "G": 28, "H": 18, "I": 14, "J": 20, "K": 28})
    title_block(ws, "Simulation d'impact des departs", "Zone de simulation simple sur cinq collaborateurs", 11)
    header_row(ws, 4, ["ID employe", "Nom prenom", "Site", "Poste", "Nb comp.", "Nb comp. critiques", "Ruptures apres depart", "Cout remplacement", "Temps remplt", "Substituabilite", "Impact global"])
    for idx in range(5, 10):
        value_cell(ws, idx, 1, f"EMP-{idx-4:03d}" if idx == 5 else "", False)
        ws.cell(idx, 1).fill = PatternFill("solid", "FFFFCC")
        formula_cell(ws, idx, 2, f'=IFERROR(VLOOKUP(A{idx},\'03_ANNUAIRE\'!A:C,2,FALSE)&" "&VLOOKUP(A{idx},\'03_ANNUAIRE\'!A:C,3,FALSE),"Saisir ID")')
        formula_cell(ws, idx, 3, f'=IFERROR(VLOOKUP(A{idx},\'03_ANNUAIRE\'!A:D,4,FALSE),"")')
        formula_cell(ws, idx, 4, f'=IFERROR(VLOOKUP(A{idx},\'03_ANNUAIRE\'!A:F,6,FALSE),"")')
        formula_cell(ws, idx, 5, f'=COUNTIF(\'04_MATRICE_SKILLS\'!B:B,A{idx})')
        formula_cell(ws, idx, 6, f'=COUNTIFS(\'04_MATRICE_SKILLS\'!B:B,A{idx},\'04_MATRICE_SKILLS\'!H:H,"Critique")')
        formula_cell(ws, idx, 7, f'=IFERROR(SUMPRODUCT((\'04_MATRICE_SKILLS\'!B$5:B$500=A{idx})*(COUNTIFS(\'04_MATRICE_SKILLS\'!F$5:F$500,\'04_MATRICE_SKILLS\'!F$5:F$500,\'04_MATRICE_SKILLS\'!L$5:L$500,"Conforme")<=2)),0)')
        formula_cell(ws, idx, 8, f'=IFERROR(SUMPRODUCT((\'04_MATRICE_SKILLS\'!B$5:B$500=A{idx})*(VLOOKUP(\'04_MATRICE_SKILLS\'!F$5:F$500,\'02_REF_COMPETENCES\'!A:J,10,FALSE))),0)', fmt="#,##0 €")
        formula_cell(ws, idx, 9, f'=IF(F{idx}>3,18,IF(F{idx}>1,12,IF(F{idx}>0,6,0)))')
        formula_cell(ws, idx, 10, f'=IF(G{idx}>=3,"🔴 Irremplacable",IF(G{idx}>=1,"🟠 Difficile","🟢 Remplacable"))')
        formula_cell(ws, idx, 11, f'=IF(G{idx}>=3,"Impact majeur",IF(G{idx}>=1,"Impact significatif","Impact faible"))')
    section_row(ws, 12, "ANALYSE CONSOLIDEE", 11)
    summary = [
        ("Total competences impactees", "=SUM(E5:E9)"),
        ("Total ruptures estimees", "=SUM(G5:G9)"),
        ("Cout total remplacement", "=SUM(H5:H9)"),
        ("Temps moyen remplacement", "=AVERAGE(I5:I9)"),
        ("Nb profils irremplacables", '=COUNTIF(J5:J9,"*Irremplacable*")'),
    ]
    for i, (label, formula) in enumerate(summary, 13):
        value_cell(ws, i, 1, label)
        fmt = "#,##0 €" if "Cout" in label else "0.0" if "Temps" in label else None
        formula_cell(ws, i, 3, formula, fmt=fmt)
    ws.freeze_panes = "A5"
    return ws


def build_16_calendrier_consolide(wb: Workbook):
    ws = wb.create_sheet("16_CALENDRIER_CONSOLIDE")
    ws.sheet_properties.tabColor = BLUE
    set_widths(ws, {"A": 14, "B": 14, "C": 16, "D": 22, "E": 26, "F": 16, "G": 16, "H": 18})
    title_block(ws, "Calendrier consolide", "Vision planning des formations et jalons RH", 8)
    header_row(ws, 4, ["Date", "Type", "Site", "Employe", "Action", "Priorite", "Statut", "Source"])
    for idx in range(5, 35):
        alt = idx % 2 == 0
        formula_cell(ws, idx, 1, '=IFERROR(INDEX(\'07_PLAN_FORMATION\'!G:G,ROW()+0),"")', alt, "DD/MM/YYYY")
        formula_cell(ws, idx, 2, '"Formation"', alt)
        formula_cell(ws, idx, 3, '=IFERROR(INDEX(\'07_PLAN_FORMATION\'!E:E,ROW()+0),"")', alt)
        formula_cell(ws, idx, 4, '=IFERROR(INDEX(\'07_PLAN_FORMATION\'!D:D,ROW()+0),"")', alt)
        formula_cell(ws, idx, 5, '=IFERROR(INDEX(\'07_PLAN_FORMATION\'!B:B,ROW()+0),"")', alt)
        formula_cell(ws, idx, 6, '=IFERROR(INDEX(\'07_PLAN_FORMATION\'!F:F,ROW()+0),"")', alt)
        formula_cell(ws, idx, 7, '=IFERROR(INDEX(\'07_PLAN_FORMATION\'!S:S,ROW()+0),"")', alt)
        value_cell(ws, idx, 8, "07_PLAN_FORMATION", alt)
    ws.freeze_panes = "A5"
    return ws


def build_17_benchmark_sites(wb: Workbook):
    ws = wb.create_sheet("17_BENCHMARK_SITES")
    ws.sheet_properties.tabColor = NAVY
    set_widths(ws, {"A": 28, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
    title_block(ws, "Benchmark inter-sites", "Comparaison des indicateurs clefs entre sites", 8)
    header_row(ws, 4, ["Indicateur"] + SITES + ["Moyenne", "Ecart-type"])
    indicators = ["Effectif", "Score moyen", "Taux couverture", "Gaps critiques", "Budget engage", "Polyvalence moyenne"]
    for idx, label in enumerate(indicators, 5):
        alt = idx % 2 == 0
        value_cell(ws, idx, 1, label, alt)
        for sidx, site in enumerate(SITES, 2):
            if label == "Effectif":
                formula = f'=COUNTIF(\'03_ANNUAIRE\'!D:D,"{site}")'
                fmt = None
            elif label == "Score moyen":
                formula = f'=IFERROR(AVERAGEIF(\'04_MATRICE_SKILLS\'!D:D,"{site}",\'04_MATRICE_SKILLS\'!I:I),0)'
                fmt = "0.00"
            elif label == "Taux couverture":
                formula = f'=IFERROR(COUNTIFS(\'04_MATRICE_SKILLS\'!D:D,"{site}",\'04_MATRICE_SKILLS\'!L:L,"Conforme")/MAX(COUNTIF(\'04_MATRICE_SKILLS\'!D:D,"{site}"),1),0)'
                fmt = "0.0%"
            elif label == "Gaps critiques":
                formula = f'=COUNTIFS(\'05_GAP_ANALYSIS\'!D:D,"{site}",\'05_GAP_ANALYSIS\'!H:H,"🔴*")'
                fmt = None
            elif label == "Budget engage":
                formula = f'=SUMIF(\'08_BUDGET\'!A:A,"{site} - Formation",\'08_BUDGET\'!D:D)'
                fmt = "#,##0 €"
            else:
                formula = f'=IFERROR(AVERAGEIF(\'03_ANNUAIRE\'!D:D,"{site}",\'03_ANNUAIRE\'!Q:Q),0)'
                fmt = "0.0%"
            formula_cell(ws, idx, sidx, formula, alt, fmt)
        formula_cell(ws, idx, 7, f'=AVERAGE(B{idx}:F{idx})', alt, "0.00")
        formula_cell(ws, idx, 8, f'=STDEV(B{idx}:F{idx})', alt, "0.00")
    ws.freeze_panes = "B5"
    return ws


def build_18_generateur_notif(wb: Workbook):
    ws = wb.create_sheet("18_GENERATEUR_NOTIF")
    ws.sheet_properties.tabColor = GOLD
    set_widths(ws, {"A": 8, "B": 18, "C": 18, "D": 18, "E": 60})
    title_block(ws, "Generateur de notifications", "Messages prets a l'emploi derives des alertes du classeur", 5)
    header_row(ws, 4, ["#", "Type", "Niveau", "Destinataire", "Message"])
    rows = [
        ("Alerte couverture", "Critique", "DRH / Direction industrielle", '="Couverture critique detectee sur "&INDEX(\'14_MATRICE_RISQUES\'!B:B,MATCH("🔴 Critique",\'14_MATRICE_RISQUES\'!M:M,0))'),
        ("Alerte gap", "Critique", "Manager de site", '="Gap critique non traite pour "&INDEX(\'05_GAP_ANALYSIS\'!C:C,MATCH("🔴 Critique",\'05_GAP_ANALYSIS\'!H:H,0))'),
        ("Alerte succession", "Important", "DRH", '="Plan de succession a revoir sur "&INDEX(\'06_SUCCESSION\'!B:B,MATCH("Critique",\'06_SUCCESSION\'!E:E,0))'),
        ("Alerte budget", "Attention", "RH formation", '="Depassement potentiel budget formation : "&TEXT(SUM(\'08_BUDGET\'!D:D),"# ##0 €")'),
        ("Alerte qualite donnees", "Important", "PMO SkillsMap", '="Score qualite donnees actuel : "&\'11_CONTROLES\'!E3'),
    ]
    for idx, row in enumerate(rows, 5):
        alt = idx % 2 == 0
        value_cell(ws, idx, 1, idx - 4, alt)
        value_cell(ws, idx, 2, row[0], alt)
        value_cell(ws, idx, 3, row[1], alt)
        value_cell(ws, idx, 4, row[2], alt)
        formula_cell(ws, idx, 5, row[3], alt)
    ws.freeze_panes = "A5"
    return ws


def reorder_sheets(wb: Workbook) -> None:
    desired = [
        "01_PARAMETRES", "02_REF_COMPETENCES", "03_ANNUAIRE", "04_MATRICE_SKILLS",
        "05_GAP_ANALYSIS", "06_SUCCESSION", "07_PLAN_FORMATION", "08_BUDGET",
        "09_HISTORIQUE", "10_DASHBOARD_KPI", "11_CONTROLES", "12_NOTES_REGLEMENTAIRES",
        "13_EXPORT_RAPPORT", "14_MATRICE_RISQUES", "15_SIMULATION_DEPARTS",
        "16_CALENDRIER_CONSOLIDE", "17_BENCHMARK_SITES", "18_GENERATEUR_NOTIF",
    ]
    lookup = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [lookup[name] for name in desired if name in lookup]


def apply_workbook_properties(wb: Workbook) -> None:
    wb.properties.title = "NEURAL - SkillsMapAero - Cartographie des competences aeronautiques"
    wb.properties.subject = "Cartographie dynamique des competences, gaps et succession"
    wb.properties.creator = "NEURAL - Intelligence augmentee pour l'entreprise"
    wb.properties.description = "Classeur genere par script Python openpyxl pour le suivi des competences aeronautiques."
    wb.properties.keywords = "NEURAL, SkillsMapAero, aeronautique, competences, RH, succession, formation"
    wb.properties.category = "Aeronautique - RH"


def main() -> Path:
    print("=" * 70)
    print(" NEURAL - SkillsMapAero - generation")
    print(f" {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 70)

    wb = Workbook()
    build_01_parametres(wb)
    build_02_ref_competences(wb)
    build_03_annuaire(wb)
    build_04_matrice_skills(wb)
    build_05_gap_analysis(wb)
    build_06_succession(wb)
    build_07_plan_formation(wb)
    build_08_budget(wb)
    build_09_historique(wb)
    build_10_dashboard_kpi(wb)
    build_11_controles(wb)
    build_12_notes_reglementaires(wb)
    build_13_export_rapport(wb)
    build_14_matrice_risques(wb)
    build_15_simulation_departs(wb)
    build_16_calendrier_consolide(wb)
    build_17_benchmark_sites(wb)
    build_18_generateur_notif(wb)
    reorder_sheets(wb)
    apply_workbook_properties(wb)

    output_path = Path(__file__).resolve().parent / "data" / OUTPUT_NAME
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Fichier genere : {output_path}")
    print(f"Onglets : {len(wb.worksheets)}")
    return output_path


if __name__ == "__main__":
    main()
