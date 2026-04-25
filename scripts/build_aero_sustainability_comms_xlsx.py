"""
Build AeroSustainabilityComms_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'AeroSustainabilityComms' (branche Aéronautique — Marketing).

7 onglets : README, Référentiel, Cas claims, Verdict par cas, Redlines,
AI Act Disclosure, Limites. 5 cas démo greenwashing (SAF, ZEROe H2, eVTOL,
compensation carbone, électrique régional).

Date de référence : 25/04/2026.
Contexte 2026 : Airbus ZEROe reporté 2040, Lilium liquidée fév 2025,
Universal Hydrogen liquidée mai 2025, Volocopter en redressement.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"
AGENT_GREEN = "059669"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=AGENT_GREEN, end_color=AGENT_GREEN)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def verdict_fill(verdict: str) -> PatternFill:
    if verdict == "OK":
        return green_fill()
    if verdict == "WARN":
        return amber_fill()
    if verdict == "KO":
        return red_fill()
    return PatternFill()


def verdict_color(verdict: str) -> str:
    if verdict == "OK":
        return GREEN_OK
    if verdict == "WARN":
        return AMBER_WARN
    if verdict == "KO":
        return RED_NO
    return "0E0824"


# ────────────────────────────────────────────────────────────────────
# Données : 11 règles greenwashing aéro
# ────────────────────────────────────────────────────────────────────
RULES = [
    {
        "id": "G01",
        "label": "LCA-EVIDENCE",
        "axis": "LCA / PEF",
        "rule": "Tout claim chiffré (g CO2eq/km, % réduction, neutralité) doit être adossé à une LCA conforme PEF (Product Environmental Footprint) certifiée par tiers indépendant.",
        "base": "EU Green Claims Directive 2024 art. 3 + Recommandation 2021/2279 PEF.",
    },
    {
        "id": "G02",
        "label": "SAF-CERTIFICATION",
        "axis": "SAF",
        "rule": "Tout claim SAF doit préciser : (1) type (HEFA / ATJ / PtL / FT-SPK), (2) fournisseur certifié ISCC EU, (3) % blend exact, (4) trajectoire ReFuelEU (2% en 2025, 6% en 2030).",
        "base": "ReFuelEU Aviation Reg. UE 2023/2405 + Directive RED III + ISCC EU.",
    },
    {
        "id": "G03",
        "label": "NO-CLIMATE-NEUTRAL",
        "axis": "Neutralité carbone",
        "rule": "Claims « carbon neutral », « net zero », « climate neutral » INTERDITS sans : (1) plan de réduction sourcé, (2) compensation registre certifié (Verra VCS, Gold Standard, ART/TREES), (3) LCA scope 1+2+3.",
        "base": "EU Green Claims Directive 2024 art. 5 + EASA Decision 2024/015 + Loi Climat FR art. 12.",
    },
    {
        "id": "G04",
        "label": "NO-VAGUE-GREEN",
        "axis": "Termes vagues",
        "rule": "Termes « vert », « écologique », « propre », « durable », « éthique » sans preuve PEF interdits — Green Claims Directive sanctionne par DGCCRF.",
        "base": "EU Green Claims Directive 2024 art. 5 + Code consommation L.121-1 (pratiques trompeuses).",
    },
    {
        "id": "G05",
        "label": "H2-REALISM",
        "axis": "Hydrogène",
        "rule": "Claims aviation H2 doivent intégrer le contexte 2026 : Airbus ZEROe reporté à 2040 (annonce fév 2025), Universal Hydrogen liquidée mai 2025. Pas de promesse commercialisation H2 < 2035.",
        "base": "Airbus communiqué 2025-02 + Code consommation L.121-1.",
    },
    {
        "id": "G06",
        "label": "EVTOL-REALITY",
        "axis": "eVTOL",
        "rule": "Claims eVTOL doivent refléter état 2026 : Joby FAA Type Cert Q4 2026 espérée, Archer financé Stellantis 1B$ 2025, Volocopter en redressement déc 2024, Lilium liquidée fév 2025. Pas de promesse service commercial < 2027.",
        "base": "Communiqués FAA + presse spécialisée 2025-2026.",
    },
    {
        "id": "G07",
        "label": "CSRD-COHERENCE",
        "axis": "CSRD",
        "rule": "Les claims marketing doivent être cohérents avec le rapport CSRD ESRS E1 (climat) publié par l'entreprise. Aucune sur-déclaration marketing par rapport au reporting réglementaire.",
        "base": "Directive CSRD UE 2022/2464 + ESRS E1 (Climate change).",
    },
    {
        "id": "G08",
        "label": "LIFECYCLE-SCOPE",
        "axis": "Périmètre LCA",
        "rule": "Tout chiffre d'émissions doit préciser le périmètre : Well-to-Tank (WtT), Tank-to-Wake (TtW), Well-to-Wake (WtW), ou cycle complet incluant production et fin de vie.",
        "base": "ICAO CORSIA Methodology + EU ETS Aviation.",
    },
    {
        "id": "G09",
        "label": "OFFSET-DISCLOSE",
        "axis": "Compensation",
        "rule": "Toute compensation carbone doit préciser : (1) registre (Verra VCS, Gold Standard, ART/TREES, Plan Vivo), (2) type projet (REDD+, biochar, IFM), (3) certificat tiers, (4) date achat. Mention « offset » sans détail interdit.",
        "base": "ICVCM Core Carbon Principles + EU Green Claims Directive 2024 art. 6.",
    },
    {
        "id": "G10",
        "label": "DGCCRF-RISK",
        "axis": "Sanction",
        "rule": "Si claim ≥ 1 KO sur G01-G09 → risque sanction DGCCRF (FR) ou CMA (UK) ou EASA Decision 2024/015 — amende jusqu'à 4% du CA.",
        "base": "Code consommation FR L.132-2 + Loi Climat & Résilience art. 12.",
    },
    {
        "id": "G11",
        "label": "AI-DISCLOSE",
        "axis": "AI Act",
        "rule": "Si claim généré ou pré-validé par IA → mention art. 50 obligatoire (applicable 02/08/2026).",
        "base": "EU AI Act 2024/1689 art. 50.",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : 5 cas démo claims
# ────────────────────────────────────────────────────────────────────
CASES = [
    {
        "id": "SUS-01",
        "type": "Campagne SAF",
        "title": "Annonce SAF 50% blend long-courrier — campagne digital + presse",
        "claim": (
            "Notre nouveau service long-courrier Paris-Singapour est désormais opéré "
            "avec un carburant durable d'aviation 50% (SAF blend). Nous réduisons ainsi "
            "drastiquement notre empreinte carbone et offrons un vol plus propre. "
            "Engagement neutralité carbone d'ici 2030 sur cette ligne. "
            "Source : nos engagements ESG."
        ),
        "verdicts": {"G01": "KO", "G02": "WARN", "G03": "KO", "G04": "KO", "G05": "OK", "G06": "OK", "G07": "WARN", "G08": "KO", "G09": "WARN", "G10": "WARN", "G11": "WARN"},
    },
    {
        "id": "SUS-02",
        "type": "Communiqué hydrogène",
        "title": "Communiqué presse — démonstrateur hydrogène 50 places",
        "claim": (
            "Notre démonstrateur d'avion régional à hydrogène 50 places réalisera son "
            "premier vol commercial en 2030. Cette technologie révolutionnaire permettra "
            "d'éliminer 100% des émissions de CO2 du transport aérien régional européen. "
            "Notre ambition : 200 appareils en service d'ici 2035. "
            "Aucune compensation carbone nécessaire grâce à cette propulsion 100% propre."
        ),
        "verdicts": {"G01": "WARN", "G02": "OK", "G03": "KO", "G04": "KO", "G05": "KO", "G06": "OK", "G07": "WARN", "G08": "KO", "G09": "OK", "G10": "WARN", "G11": "OK"},
    },
    {
        "id": "SUS-03",
        "type": "Brochure eVTOL",
        "title": "Brochure eVTOL urbain — service taxi aérien Paris 2027",
        "claim": (
            "Découvrez le futur de la mobilité urbaine. Notre eVTOL 4 passagers offrira "
            "un service de taxi aérien à Paris dès 2027, 100% électrique et 100% silencieux. "
            "Trajet aéroport CDG ↔ centre-ville en 8 minutes pour 75 €. Service écologique "
            "et durable, parfait alternative aux véhicules thermiques. Réservations "
            "ouvrent prochainement. Premier vol démonstration certifié EASA Q3 2026."
        ),
        "verdicts": {"G01": "WARN", "G02": "OK", "G03": "OK", "G04": "KO", "G05": "OK", "G06": "KO", "G07": "WARN", "G08": "WARN", "G09": "OK", "G10": "WARN", "G11": "WARN"},
    },
    {
        "id": "SUS-04",
        "type": "Page web compensation",
        "title": "Page web entreprise — programme compensation carbone vols corporate",
        "claim": (
            "Notre programme « FlyGreen » compense automatiquement 100% des émissions "
            "de CO2 de vos déplacements professionnels. Pour chaque vol, nous achetons "
            "des crédits carbone certifiés et plantons des arbres. Zéro empreinte carbone "
            "garantie pour tous les déplacements de votre entreprise. Plus de 50 000 tonnes "
            "compensées en 2025. Contribution durable au climat."
        ),
        "verdicts": {"G01": "WARN", "G02": "OK", "G03": "KO", "G04": "KO", "G05": "OK", "G06": "OK", "G07": "WARN", "G08": "WARN", "G09": "KO", "G10": "WARN", "G11": "OK"},
    },
    {
        "id": "SUS-05",
        "type": "Fiche technique électrique",
        "title": "Fiche technique avion électrique régional 9 places",
        "claim": (
            "Notre avion régional 9 places 100% électrique offre une autonomie de 250 km "
            "en 38 minutes de vol. Émissions opérationnelles : 0 g CO2/passager-km. "
            "Empreinte carbone du cycle de vie estimée à 28 g CO2eq/passager-km "
            "(étude ADEME 2025 — méthodologie ICAO CORSIA v2.3, périmètre Well-to-Wake "
            "incluant production batterie). Certification EASA Type CS-23 envisagée Q1 2027 "
            "(sous réserve). Aucune compensation requise."
        ),
        "verdicts": {"G01": "OK", "G02": "OK", "G03": "WARN", "G04": "OK", "G05": "OK", "G06": "OK", "G07": "OK", "G08": "OK", "G09": "OK", "G10": "OK", "G11": "WARN"},
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Redlines
# ────────────────────────────────────────────────────────────────────
REDLINES = [
    # SUS-01 SAF
    {"case": "SUS-01", "rule": "G01", "before": "Source : nos engagements ESG", "after": "Source : LCA Well-to-Wake conforme PEF certifiée par tiers indépendant (étude X 2025, n°ABC-2025-Q4) — réduction CO2 de 38% démontrée vs Jet A1 baseline.", "why": "Claim chiffré sans LCA = greenwashing — Green Claims Directive art. 3"},
    {"case": "SUS-01", "rule": "G02", "before": "carburant durable d'aviation 50% (SAF blend)", "after": "SAF HEFA blend 50% (fournisseur Neste, certifié ISCC EU, conforme ReFuelEU Aviation Reg. UE 2023/2405) — au-delà de l'objectif 2025 de 2% blend.", "why": "Préciser type SAF, fournisseur, certification ISCC, conformité ReFuelEU"},
    {"case": "SUS-01", "rule": "G03", "before": "Engagement neutralité carbone d'ici 2030 sur cette ligne", "after": "Plan de réduction CO2 de 50% d'ici 2030 sur cette ligne (réduction effective + SAF 80% blend visé 2030), reste 50% compensé via crédits Verra VCS validés (registre publique).", "why": "« Neutralité carbone » sans plan de réduction + registre = sanction DGCCRF — interdit Green Claims Directive 2024 art. 5"},
    {"case": "SUS-01", "rule": "G04", "before": "vol plus propre... drastiquement", "after": "vol émettant 38% moins de CO2eq qu'un vol équivalent en Jet A1 conventionnel", "why": "« Plus propre » + « drastiquement » = termes vagues interdits Green Claims Directive 2024 art. 5"},
    {"case": "SUS-01", "rule": "G08", "before": "réduisons notre empreinte carbone", "after": "réduisons les émissions Tank-to-Wake (TtW) de 38% par rapport à un vol Jet A1 standard. Périmètre Well-to-Wake (WtW) en cours d'évaluation — publication prévue Q3 2026.", "why": "Préciser périmètre LCA (WtT / TtW / WtW)"},
    # SUS-02 hydrogène
    {"case": "SUS-02", "rule": "G05", "before": "premier vol commercial en 2030... 200 appareils en service d'ici 2035", "after": "Démonstrateur en vol prévu 2030, sous réserve du contexte technologique et réglementaire. Service commercial plausible dans la période 2035-2040 (cf. Airbus ZEROe reporté à 2040, communiqué fév 2025).", "why": "Calendrier 2030/2035 H2 = irréaliste cf. report Airbus ZEROe 2040 + liquidation Universal Hydrogen mai 2025"},
    {"case": "SUS-02", "rule": "G03", "before": "éliminer 100% des émissions de CO2 du transport aérien régional européen", "after": "réduire les émissions de CO2 directes (Tank-to-Wake) à zéro pour ce démonstrateur 50 places, dans le cadre d'un plan de réduction sectoriel CSRD. Émissions Well-to-Tank (production H2) restant à compenser via crédits Verra VCS.", "why": "« Éliminer 100% » = neutralité carbone interdite sans plan + compensation. Green Claims Directive 2024"},
    {"case": "SUS-02", "rule": "G04", "before": "technologie révolutionnaire... 100% propre", "after": "technologie de propulsion à pile à combustible hydrogène — 0 émission directe en vol. Bilan carbone du cycle de vie selon production H2 (vert, gris, bleu).", "why": "« Révolutionnaire » et « 100% propre » = termes vagues interdits"},
    {"case": "SUS-02", "rule": "G08", "before": "Aucune compensation carbone nécessaire grâce à cette propulsion 100% propre", "after": "Pour atteindre une trajectoire SBTi 1.5°C, des compensations partielles seront mobilisées sur les émissions WtT (production H2 non-vert).", "why": "Périmètre LCA absent + compensation niée à tort"},
    # SUS-03 eVTOL
    {"case": "SUS-03", "rule": "G06", "before": "service de taxi aérien à Paris dès 2027... certifié EASA Q3 2026", "after": "Premier vol démonstration prévu Q3 2026, sous réserve de certification EASA (CS-23 ou SC-VTOL en cours). Service commercial plausible 2028-2029 selon trajectoire actuelle des certifications eVTOL (Joby FAA Type Cert Q4 2026 espérée).", "why": "Calendrier Paris 2027 trop ambitieux cf. réalité 2026 (Volocopter redressement, Lilium liquidée) — risque DGCCRF"},
    {"case": "SUS-03", "rule": "G04", "before": "100% électrique et 100% silencieux... écologique et durable", "after": "Propulsion 100% électrique avec niveau sonore inférieur à 65 dB(A) en survol urbain. Bilan carbone du cycle de vie en cours d'évaluation (publication 2026).", "why": "« 100% silencieux » impossible techniquement + « écologique et durable » = vague"},
    {"case": "SUS-03", "rule": "G07", "before": "[Pas de mention CSRD]", "after": "Notre rapport CSRD ESRS E1 publié en avril 2026 documente notre trajectoire climat — disponible sur demande pour les investisseurs.", "why": "Cohérence claim ↔ reporting CSRD requis"},
    # SUS-04 compensation
    {"case": "SUS-04", "rule": "G09", "before": "achetons des crédits carbone certifiés et plantons des arbres", "after": "Achetons des crédits carbone Gold Standard (registre Gold Standard for the Global Goals, vérifiés par tiers indépendant) — projets reforestation REDD+ certifiés (références projets sur demande). Date achat affichée par vol.", "why": "« Crédits certifiés » sans registre + type projet = greenwashing offset"},
    {"case": "SUS-04", "rule": "G03", "before": "Zéro empreinte carbone garantie pour tous les déplacements", "after": "Compensation carbone à hauteur de 100% des émissions Tank-to-Wake (TtW). Émissions Well-to-Tank (WtT) à votre charge selon plan de réduction interne.", "why": "« Zéro empreinte garantie » = greenwashing — absurde car compensation ≠ réduction"},
    {"case": "SUS-04", "rule": "G04", "before": "Plus de 50 000 tonnes compensées en 2025. Contribution durable au climat.", "after": "50 824 t CO2eq compensées en 2025 (registre Gold Standard, vérification tiers — rapport CSRD 2025 page 47).", "why": "« Contribution durable » = vague — citer rapport CSRD"},
    # SUS-05 électrique régional (cas le plus conforme — sert de bench)
    {"case": "SUS-05", "rule": "G03", "before": "Aucune compensation requise.", "after": "Compensation supplémentaire optionnelle via Gold Standard pour atteindre net-zero scope 3 (transport batterie usagée).", "why": "Affirmation forte « aucune compensation » à nuancer pour scope 3"},
    {"case": "SUS-05", "rule": "G11", "before": "[Aucune mention IA dans la fiche]", "after": "Si fiche pré-validée par IA contre référentiel greenwashing → mention art. 50.", "why": "AI Act art. 50"},
]

# ────────────────────────────────────────────────────────────────────
# Données : AI Act Disclosure modèles
# ────────────────────────────────────────────────────────────────────
AI_DISCLOSURE = [
    {"format": "Audit pré-publication claim", "fr": "Ce claim a été pré-audité par un système d'IA contre 11 règles greenwashing aéro (Green Claims Directive + ReFuelEU + ESRS E1). Validation finale par responsable ESG signataire CSRD. (UE 2024/1689 art. 50.)", "en": "This claim has been pre-audited by an AI system against 11 aero greenwashing rules. Final validation by CSRD signatory ESG officer. (EU 2024/1689 art. 50.)"},
    {"format": "Communiqué presse SAF/H2", "fr": "Communication validée par audit IA et bureau d'études — données LCA certifiées par tiers indépendant. (UE 2024/1689 art. 50.)", "en": "Communication validated by AI audit and engineering office — LCA data certified by independent third party. (EU 2024/1689 art. 50.)"},
    {"format": "Brochure ESG corporate", "fr": "Document avec contenus partiellement assistés IA. Données ESG cohérentes avec rapport CSRD publié. (UE 2024/1689 art. 50.)", "en": "Document with partially AI-assisted content. ESG data consistent with published CSRD report. (EU 2024/1689 art. 50.)"},
    {"format": "Page web claim chiffré", "fr": "Chiffres LCA pré-vérifiés par IA contre méthodologie ICAO CORSIA + ESRS E1. Source : LCA tiers, registre ICVCM si compensation. UE 2024/1689 art. 50.", "en": "LCA figures AI pre-verified against ICAO CORSIA + ESRS E1 methodology. Source: third-party LCA, ICVCM registry if offset. EU 2024/1689 art. 50."},
    {"format": "Post LinkedIn ESG", "fr": "Post pré-audité IA contre Green Claims Directive #AIVerified", "en": "Post AI pre-audited against Green Claims Directive #AIVerified"},
]


# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("0_README")
    set_widths(ws, {1: 26, 2: 92})
    write_title(ws, "AeroSustainabilityComms — Audit anti-greenwashing aéro", span=2)
    write_subtitle(ws, 2, "NEURAL — Branche Aéronautique / Marketing — Date de réf. : 25/04/2026", span=2)

    rows = [
        ("Mission", "Auditer les claims SAF, hydrogène, électrique, eVTOL, compensation carbone vs Green Claims Directive 2024 + Loi Climat & Résilience FR + ReFuelEU Aviation + CSRD ESRS E1 + EASA Decision 2024/015. Détection greenwashing avant diffusion."),
        ("Périmètre", "Claims environnementaux marketing aviation — campagnes digital, presse, brochures, fiches techniques, pages web, social. B2C et B2B."),
        ("Référentiel", "11 règles encodées (G01-G11) : LCA evidence, SAF certification, no climate-neutral, no vague green, H2 realism (cf. ZEROe report 2040), eVTOL reality (Joby/Volocopter/Lilium), CSRD coherence, lifecycle scope, offset disclose, DGCCRF risk, AI disclose."),
        ("Échelle verdict", "OK = conforme | WARN = à revoir | KO = bloquant, redline obligatoire avant diffusion."),
        ("Décision", "0 KO ET score ≥ 80% → DIFFUSION OK. Sinon : RELECTURE ou BLOQUÉ."),
        ("Inputs cas démo", "5 cas synthétiques 2026 : SAF 50% campagne long-courrier, communiqué H2 démonstrateur, brochure eVTOL Paris 2027, page compensation carbone corporate, fiche électrique régional (cas conforme bench)."),
        ("Outputs", "(1) Verdict matrice 11 × 5, (2) score conformité auto, (3) décision auto, (4) ~17 redlines avec base légale, (5) modèles disclosure IA FR + EN."),
        ("Contexte 2026 critique", "Airbus ZEROe REPORTÉ À 2040 (annonce fév 2025). Universal Hydrogen LIQUIDÉE mai 2025. Lilium LIQUIDÉE fév 2025. Volocopter EN REDRESSEMENT déc 2024. Joby et Archer continuent. Discours marketing aéro doit refléter cette réalité."),
        ("Statut", "Démo recruteur — claims synthétiques, anonymisés. Aucune entreprise réelle citée nominativement dans les cas (références sectorielles factuelles uniquement)."),
        ("Disclaimer", "Outil indicatif. Ne remplace pas : (1) DGCCRF / CMA pour sanction pratiques commerciales trompeuses, (2) responsable ESG signataire CSRD, (3) bureau d'études LCA tiers indépendant, (4) juriste consommation. Validation humaine obligatoire."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 75 + 1))
        row += 1


def build_referentiel(wb: Workbook) -> None:
    ws = wb.create_sheet("1_Referentiel")
    set_widths(ws, {1: 8, 2: 22, 3: 18, 4: 60, 5: 50})
    write_title(ws, "Référentiel — 11 règles anti-greenwashing aéro", span=5)
    write_subtitle(ws, 2, "Green Claims Directive + ReFuelEU + Loi Climat FR + CSRD ESRS E1 + EASA Decision 2024/015.", span=5)

    headers = ["ID", "Label", "Axe", "Règle", "Base légale / réf."]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        ws.cell(row=row, column=3, value=rule["axis"])
        ws.cell(row=row, column=4, value=rule["rule"])
        ws.cell(row=row, column=5, value=rule["base"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=80)
        row += 1


def build_cas_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("2_Cas_Claims")
    set_widths(ws, {1: 12, 2: 22, 3: 38, 4: 75})
    write_title(ws, "Cas claims — 5 communications environnementales à auditer", span=4)
    write_subtitle(ws, 2, "Cas synthétiques 2026 inspirés du marché. Aucune entreprise réelle citée nominativement.", span=4)

    headers = ["ID", "Type", "Titre", "Claim (extrait)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 4)

    row = 5
    for i, c in enumerate(CASES):
        ws.cell(row=row, column=1, value=c["id"])
        ws.cell(row=row, column=2, value=c["type"])
        ws.cell(row=row, column=3, value=c["title"])
        ws.cell(row=row, column=4, value=c["claim"])
        style_body_row(ws, row, 4, alt=(i % 2 == 1), height=140)
        row += 1


def build_verdict(wb: Workbook) -> None:
    ws = wb.create_sheet("3_Verdict_Par_Cas")
    n_cases = len(CASES)
    n_rules = len(RULES)
    cols_total = 2 + n_cases

    widths = {1: 8, 2: 22}
    for i in range(n_cases):
        widths[3 + i] = 14
    set_widths(ws, widths)

    write_title(ws, f"Matrice verdicts — {n_rules} règles × {n_cases} cas", span=cols_total)
    write_subtitle(ws, 2, "OK = conforme | WARN = ajustement | KO = bloquant. Score + décision auto en bas.", span=cols_total)

    ws.cell(row=4, column=1, value="ID")
    ws.cell(row=4, column=2, value="Règle")
    for i, c in enumerate(CASES):
        ws.cell(row=4, column=3 + i, value=c["id"])
    style_header_row(ws, 4, cols_total)

    row = 5
    for r_idx, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        for c_idx, c in enumerate(CASES):
            v = c["verdicts"].get(rule["id"], "")
            cell = ws.cell(row=row, column=3 + c_idx, value=v)
            cell.font = body_font(bold=True, color=verdict_color(v))
            cell.fill = verdict_fill(v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).font = body_font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).font = body_font()
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    write_section_header(ws, row, "Synthèse par cas", span=cols_total)
    row += 1

    last_verdict_row = 4 + n_rules

    for label, count_val in [("Nb OK", "OK"), ("Nb WARN", "WARN"), ("Nb KO", "KO")]:
        ws.cell(row=row, column=2, value=label)
        for c_idx in range(n_cases):
            col_letter = get_column_letter(3 + c_idx)
            ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"{count_val}")')
        style_body_row(ws, row, cols_total, height=22)
        ws.cell(row=row, column=2).font = body_font(bold=True)
        row += 1

    ws.cell(row=row, column=2, value="Score conformité")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        formula = f'=(COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN"))/{n_rules}'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.number_format = "0%"
    style_body_row(ws, row, cols_total, height=24)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    ws.cell(row=row, column=2, value="Décision")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ko_count = f'COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"KO")'
        score = f'(COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN"))/{n_rules}'
        formula = f'=IF({ko_count}>0,"BLOQUÉ",IF({score}>=0.8,"DIFFUSION OK","RELECTURE"))'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.font = body_font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    style_body_row(ws, row, cols_total, height=26)
    ws.cell(row=row, column=2).font = body_font(bold=True)


def build_redlines(wb: Workbook) -> None:
    ws = wb.create_sheet("4_Redlines")
    set_widths(ws, {1: 12, 2: 8, 3: 50, 4: 65, 5: 45})
    write_title(ws, "Redlines — Suggestions de correction concrètes", span=5)
    write_subtitle(ws, 2, "Pour chaque KO/WARN, redline applicable avec base légale citée.", span=5)

    headers = ["Cas", "Règle", "Avant (claim)", "Après (suggestion)", "Pourquoi"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rl in enumerate(REDLINES):
        ws.cell(row=row, column=1, value=rl["case"])
        ws.cell(row=row, column=2, value=rl["rule"])
        ws.cell(row=row, column=3, value=rl["before"])
        ws.cell(row=row, column=4, value=rl["after"])
        ws.cell(row=row, column=5, value=rl["why"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=80)
        row += 1


def build_ai_disclosure(wb: Workbook) -> None:
    ws = wb.create_sheet("5_AI_Act_Disclosure")
    set_widths(ws, {1: 28, 2: 60, 3: 60})
    write_title(ws, "AI Act art. 50 — Modèles de mention IA", span=3)
    write_subtitle(ws, 2, "Applicable 2 août 2026. Disclosure obligatoire si IA utilisée pour audit greenwashing.", span=3)

    headers = ["Format", "Mention FR", "Mention EN"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 3)

    row = 5
    for i, d in enumerate(AI_DISCLOSURE):
        ws.cell(row=row, column=1, value=d["format"])
        ws.cell(row=row, column=2, value=d["fr"])
        ws.cell(row=row, column=3, value=d["en"])
        style_body_row(ws, row, 3, alt=(i % 2 == 1), height=85)
        row += 1


def build_limites(wb: Workbook) -> None:
    ws = wb.create_sheet("6_Limites")
    set_widths(ws, {1: 32, 2: 80})
    write_title(ws, "Truth layer — Ce que AeroSustainabilityComms NE remplace PAS", span=2)
    write_subtitle(ws, 2, "Disclaimer obligatoire. Sujet : greenwashing aéro = juridiquement chaud.", span=2)

    rows = [
        ("DGCCRF / CMA / EASA", "L'agent n'a pas pouvoir de sanction. La DGCCRF (FR), CMA (UK), EASA Decision 2024/015 sont les autorités compétentes pour sanctionner les pratiques commerciales trompeuses."),
        ("Bureau d'études LCA tiers", "Aucune LCA n'est calculée par l'agent. Tout chiffre LCA doit être produit par un bureau d'études tiers certifié (PEF + ISO 14040/14044)."),
        ("Responsable ESG signataire CSRD", "L'agent vérifie cohérence claim ↔ reporting CSRD ESRS E1, mais ne signe pas le rapport. Le DAF / DRSE signataire reste responsable."),
        ("Juriste consommation", "Pour toute campagne grand public à risque (B2C), validation par juriste consommation FR / UK / DE obligatoire."),
        ("Compensation registres", "L'agent vérifie présence de registre cité (Verra, Gold Standard, ART/TREES, Plan Vivo) — il ne valide PAS la qualité du projet sous-jacent. Vérification ICVCM Core Carbon Principles requise."),
        ("EASA / DGAC / FAA — claims certification", "Les claims liés aux certifications type (CS-23, SC-VTOL, FAA Part 23) doivent être validés par le service certification interne — l'agent ne signale qu'une cohérence calendaire indicative."),
        ("Calendrier H2 / eVTOL — sources évolutives", "Le contexte 2026 (ZEROe 2040, Lilium liquidée, Universal Hydrogen liquidée, Volocopter redressement) est figé au 25/04/2026. À rafraîchir avant tout claim — situation évolutive."),
        ("ReFuelEU Aviation — quotas", "L'agent applique la trajectoire 2025 (2%) → 2030 (6%) → 2050 (70%). Vérification des quotas réels par fournisseur SAF certifié ISCC EU obligatoire."),
        ("Ce que l'agent NE génère PAS", "Pas de génération de claim conforme — uniquement audit. Pour rédaction : voir AeroTechContent + bureau d'études."),
        ("Mode démo public", "Scenario-id only. Cas anonymisés. Aucune donnée client réelle ingérée."),
        ("Limite v1", "Pas d'API connectée aux registres compensation (Verra, Gold Standard) — vérification manuelle du registre cité. v2 prévue connecteur ICVCM."),
        ("Sanctions potentielles", "DGCCRF FR : amendes jusqu'à 4% CA. EU Green Claims Directive : sanctions par État membre. EASA Decision 2024/015 : retrait certificat. À utiliser comme aide à la décision, pas comme sécurité juridique absolue."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 70 + 1))
        row += 1


def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb)
    build_referentiel(wb)
    build_cas_inputs(wb)
    build_verdict(wb)
    build_redlines(wb)
    build_ai_disclosure(wb)
    build_limites(wb)

    out = "AeroSustainabilityComms_NEURAL.xlsx"
    wb.save(out)
    print(f"OK — {out} written ({len(CASES)} cas, {len(RULES)} règles, {len(REDLINES)} redlines)")


if __name__ == "__main__":
    build()
