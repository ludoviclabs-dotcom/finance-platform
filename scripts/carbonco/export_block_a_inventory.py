from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_ROOT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\Carbon and Co")
MASTERS = [
    "CarbonCo_Calcul_Carbone_v2.xlsx",
    "CarbonCo_ESG_Social.xlsx",
    "CarbonCo_Finance_DPP_v1_3.xlsx",
]


def inventory_workbook(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        named_ranges = list(wb.defined_names.keys())
    finally:
        wb.close()

    stat = path.stat()
    return {
        "filename": path.name,
        "path": str(path),
        "size_bytes": stat.st_size,
        "last_modified": stat.st_mtime,
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "named_range_count": len(named_ranges),
        "has_claude_log": "Claude Log" in sheet_names,
    }


def main() -> None:
    results = []
    for name in MASTERS:
        path = WORKBOOK_ROOT / name
        if not path.exists():
            print(f"MISSING|{path}")
            continue

        item = inventory_workbook(path)
        results.append(item)
        print(
            "FILE|{filename}|{size_bytes}|{sheet_count}|{named_range_count}|{has_claude_log}".format(
                **item
            )
        )
        print("SHEETS|" + ",".join(item["sheet_names"]))

    print("JSON|" + json.dumps(results, ensure_ascii=True))


if __name__ == "__main__":
    main()
