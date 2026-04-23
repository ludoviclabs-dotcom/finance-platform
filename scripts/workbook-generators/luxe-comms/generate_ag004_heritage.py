"""
NEURAL_AG004_HeritageComms.xlsx
Sourcing patrimonial, validation de sources, blocs narratifs sources.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, STATUS, VERT, VIOLET, YESNO,
    ajouter_table, attach_dv, cf_stale, convertir_dates, D, DT, dv_liste,
    ecrire_donnees, en_tetes, params_table, param_ligne, section, titre,
    today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_AG004_HeritageComms.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "AG-004  —  HERITAGECOMMS", "Sourcing patrimonial — sources, faits, blocs narratifs", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 86

    section(ws, 4, 6, "Mission", NOIR)
    ecrire_donnees(ws, 5, [
        ["", "Role",        "Valoriser l'heritage maison. Aucune sortie sans source active et citation correctement formatee."],
        ["", "Input",       "Query heritage (thematique, annee, produit, angle)."],
        ["", "Output",      "Blocs narratifs sources + citations exactes."],
        ["", "Gate",        "HERITAGE — blocking si source TERTIARY ou STALE ou citation absente."],
        ["", "Appelants",   "AG-002 (enrichissement presse), AG-003 (pack evenementiel)."],
        ["", "Referentiel", "FOUNDATIONS!5_HERITAGE_SOURCEBOOK."],
    ], align=AL)

    section(ws, 13, 6, "Regle d'or", BORDEAUX)
    ecrire_donnees(ws, 14, [
        ["", "Regle 1", "Jamais d'affirmation sans source cataloguee."],
        ["", "Regle 2", "Source TERTIARY = REVIEW obligatoire avant usage."],
        ["", "Regle 3", "Citation minimum : SRC_ID + date + cote archive."],
        ["", "Regle 4", "Source STALE = blocage tant que REVIEW_DATE depassee."],
        ["", "Regle 5", "Toute sortie enrichie retourne SRC_ID pour tracabilite aval."],
    ], align=AL)


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES AGENT", "tblParamsAG004", 3)
    rows = [
        ("AGENT_ID",        "AG-004", ""),
        ("AGENT_NAME",      "HeritageComms", ""),
        ("MAISON_ID",       "M-001", ""),
        ("MAISON_NAME",     "Maison pilote", ""),
        ("LANG_PRIMARY",    "FR", ""),
        ("LANG_SECONDARY",  "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", ""),
        ("DATA_VERSION",    "v2026.04", ""),
        ("PROMPT_VERSION",  "v1.0.0", ""),
        ("CITATION_FORMAT_DEFAULT", "Maison-style", ""),
        ("MIN_SOURCES_PER_BLOCK", 1, "Nombre minimum de sources par bloc narratif"),
        ("BLOCK_TERTIARY_ONLY",   "YES", "Bloquer si sources = TERTIARY uniquement"),
        ("LAST_REFRESH",    D(today_iso()), ""),
    ]
    params_table(ws, "tblParamsAG004", start_row=4, rows=rows)


def onglet_query_intake(wb):
    ws = wb.create_sheet("2_QUERY_INTAKE")
    titre(ws, "QUERY INTAKE", "Demandes patrimoniales entrantes — AG-002 / AG-003", 8)
    en_tetes(ws, 4, ["QUERY_ID", "MAISON_ID", "THEMATIQUE", "SOURCE_AGENT", "USE_CASE", "ANGLE", "DEADLINE", "STATUS"],
             [10, 10, 40, 14, 22, 30, 16, 12])
    queries = [
        ("HQ-001", "M-001", "Origine du motif iconique XVI-line",   "AG-003", "Exposition 100 ans", "Heritage design",           D("2026-05-01"), "IN_PROGRESS"),
        ("HQ-002", "M-001", "Fondateur — portrait historique",       "AG-002", "Rapport ESG",       "Valeurs + transmission",    D("2026-05-20"), "IN_PROGRESS"),
        ("HQ-003", "M-001", "Piece portee a Cannes 1987",            "AG-003", "Defile printemps",  "Lien celebrite + maison",   D("2026-04-25"), "IN_PROGRESS"),
        ("HQ-004", "M-001", "Archives atelier annees 1950",          "AG-002", "Collaboration artiste","Atelier + continuite",  D("2026-09-30"), "IN_PROGRESS"),
        ("HQ-005", "M-001", "Naissance de la marque 1920",           "AG-003", "Pop-up Tokyo",      "Narration origine",         D("2026-07-01"), "IN_PROGRESS"),
        ("HQ-006", "M-001", "Piece Musee Galliera 1935",             "AG-002", "Presse trade",      "Reconnaissance institutionnelle",D("2026-05-15"), "IN_PROGRESS"),
        ("HQ-007", "M-001", "Savoir-faire ciselure — origine geste", "AG-002", "Feature atelier",   "Technique + transmission",  D("2026-04-25"), "DONE"),
    ]
    last = ecrire_donnees(ws, 5, queries, align=AL)
    convertir_dates(ws, 5, last, date_cols=[7])
    ajouter_table(ws, "tblHeritageQueries", f"A4:H{last}")
    attach_dv(ws, f"D5:D{last}", ["AG-002", "AG-003", "USER"])
    attach_dv(ws, f"H5:H{last}", ["PENDING", "IN_PROGRESS", "DONE", "BLOCKED"])


def onglet_source_catalog(wb):
    ws = wb.create_sheet("3_SOURCE_CATALOG")
    titre(ws, "SOURCE CATALOG", "Catalogue local AG-004 — STATUT = formule (R-03), STALE_FLAG supprime", 10)
    # R-03 : MANUAL_OVERRIDE + STATUT formule. Plus de STALE_FLAG redondant.
    en_tetes(ws, 4, ["SOURCE_ID", "TITRE", "TYPE", "ANNEE", "COTE", "MANUAL_OVERRIDE", "REVIEW_DATE",
                      "USAGE_COUNT", "CITATION_FORMAT", "STATUT"],
             [12, 36, 14, 10, 18, 18, 14, 14, 20, 14])
    sources = [
        ("SRC-001", "Carnet de dessins du fondateur",       "PRIMARY",   1923, "A-1923-001", "",          D("2026-10-01"), 4, "Maison-style", ""),
        ("SRC-002", "Lettre a la direction atelier",        "PRIMARY",   1947, "A-1947-112", "",          D("2026-09-15"), 2, "Chicago",      ""),
        ("SRC-003", "Catalogue exposition retrospective",   "SECONDARY", 1987, "A-1987-032", "",          D("2026-07-20"), 7, "Chicago",      ""),
        ("SRC-004", "Livre monographie maison",             "SECONDARY", 2005, "A-2005-008", "",          D("2026-12-01"), 3, "APA",          ""),
        ("SRC-005", "Article Le Figaro edition speciale",   "TERTIARY",  2012, "A-2012-044", "",          D("2025-01-10"), 1, "Chicago",      ""),
        ("SRC-006", "Photos atelier annees 1950",           "PRIMARY",   1952, "A-1952-007", "",          D("2026-06-30"), 5, "Maison-style", ""),
        ("SRC-007", "Interview fondateur",                  "PRIMARY",   1965, "A-1965-099", "",          D("2026-11-15"), 2, "Chicago",      ""),
        ("SRC-008", "Piece Musee Galliera",                 "SECONDARY", 1935, "A-1935-401", "",          D("2026-08-01"), 1, "Chicago",      ""),
        ("SRC-009", "Document juridique marque deposee",    "PRIMARY",   1920, "A-1920-001", "",          D("2027-01-01"), 0, "Juridique",    ""),
        ("SRC-010", "Blog post fan non officiel",           "TERTIARY",  2019, "A-2019-777", "REJECTED",  D("2024-12-01"), 0, "Non utilisable",""),
    ]
    last = ecrire_donnees(ws, 5, sources, fmt_cols={4: "0", 8: "0"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[7])

    # R-03 : STATUT formule
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=10, value=f'=IF(F{r}<>"",F{r},IF(G{r}<TODAY(),"STALE","ACTIVE"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS
    # R-08 : CF J-30 sur REVIEW_DATE
    cf_stale(ws, f"G5:G{last}", days_window=30)

    ajouter_table(ws, "tblSourceCatalog", f"A4:J{last}")
    attach_dv(ws, f"C5:C{last}", ["PRIMARY", "SECONDARY", "TERTIARY"])
    attach_dv(ws, f"F5:F{last}", ["", "REJECTED"])

    # CF sur STATUT (col J = 10 apres R-03)
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"ACTIVE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"STALE"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"REJECTED"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_approved_facts(wb):
    ws = wb.create_sheet("4_APPROVED_FACTS")
    titre(ws, "APPROVED FACTS", "Faits historiques valides — reutilisables dans comms", 8)
    en_tetes(ws, 4, ["FACT_ID", "FAIT", "ANNEE", "SOURCE_1", "SOURCE_2", "SOURCE_3", "STATUT", "SOURCE_CHECK"],
             [10, 54, 10, 14, 14, 14, 14, 18])
    facts = [
        ("FC-001", "La maison est fondee en 1920 par le fondateur X.",                                     1920, "SRC-009", "SRC-001", "",         "APPROVED", ""),
        ("FC-002", "Les premiers carnets de dessins datent de 1923.",                                      1923, "SRC-001", "",        "",         "APPROVED", ""),
        ("FC-003", "En 1935, une piece de la maison entre au Musee Galliera (inv A-1935-401).",            1935, "SRC-008", "",        "",         "APPROVED", ""),
        ("FC-004", "En 1947, l'atelier adopte le geste de ciselure enseigne par le maitre A. Dupond.",     1947, "SRC-002", "",        "",         "APPROVED", ""),
        ("FC-005", "L'atelier de 1952 comptait 18 artisans.",                                               1952, "SRC-006", "",        "",         "APPROVED", ""),
        ("FC-006", "Le fondateur donne une interview a la television en 1965.",                            1965, "SRC-007", "",        "",         "APPROVED", ""),
        ("FC-007", "En 1987, une retrospective est organisee au Musee des Arts Deco.",                     1987, "SRC-003", "SRC-004", "",         "APPROVED", ""),
        ("FC-008", "Motif XVI-line inspire des dessins de 1923.",                                          2020, "SRC-001", "SRC-004", "",         "APPROVED", ""),
        ("FC-009", "Une piece portee a Cannes 1987 par [celebrite].",                                       1987, "SRC-003", "",        "",         "APPROVED", ""),
        ("FC-010", "La maison emploie 240 artisans en 2025.",                                              2025, "SRC-004", "",        "",         "APPROVED", ""),
    ]
    last = ecrire_donnees(ws, 5, facts, fmt_cols={3: "0"}, align=AL)

    # Formule SOURCE_CHECK = MISSING_SOURCE si aucune source
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=8, value=f'=IF(COUNTA(D{r}:F{r})>=1,"OK","MISSING_SOURCE")')
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = AC
        c.border = BF

    ajouter_table(ws, "tblApprovedFacts", f"A4:H{last}")

    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"MISSING_SOURCE"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_narrative_blocks(wb):
    ws = wb.create_sheet("5_NARRATIVE_BLOCKS")
    titre(ws, "NARRATIVE BLOCKS", "Blocs narratifs reutilisables — sources + usage", 8)
    en_tetes(ws, 4, ["BLOCK_ID", "THEME", "TITRE", "TEXTE", "SOURCES", "SOURCE_STATUS", "USABILITY", "USAGE_COUNT"],
             [12, 18, 40, 64, 26, 18, 18, 14])
    blocks = [
        ("BL-001", "Origine",        "1920 — La maison nait",
         "En 1920, [Fondateur] depose la marque et ouvre son premier atelier. Les premiers carnets de dessins datent de 1923 (Archives A-1923-001).",
         "SRC-001; SRC-009", "ACTIVE", "", 2),
        ("BL-002", "Atelier",         "Le geste transmis depuis 1947",
         "En 1947, le maitre A. Dupond formalise le geste de ciselure dans une lettre conservee aux archives (A-1947-112). Aujourd'hui, 240 artisans perpetuent ce meme geste.",
         "SRC-002; SRC-004",  "ACTIVE", "", 1),
        ("BL-003", "Reconnaissance",  "1935 — Une piece au Musee Galliera",
         "En 1935, une piece de la maison entre dans les collections du Musee Galliera (inv A-1935-401), marquant la reconnaissance institutionnelle du savoir-faire.",
         "SRC-008",            "ACTIVE", "", 0),
        ("BL-004", "Patrimoine",      "1987 — Retrospective Arts Deco",
         "Le Musee des Arts Decoratifs a consacre en 1987 une retrospective a la maison (Cat A-1987-032), reunissant cent pieces de 1920 a 1985.",
         "SRC-003; SRC-004",   "ACTIVE", "", 3),
        ("BL-005", "Iconographie",    "Le motif XVI-line",
         "Le motif XVI-line trouve ses origines dans les carnets de dessins de 1923 (A-1923-001), repris et reinterprete dans les collections contemporaines.",
         "SRC-001",            "ACTIVE", "", 1),
        ("BL-006", "Culture",         "1987 — Cannes",
         "Une piece de la maison a ete portee a Cannes en 1987, captee par les agences de photo (cat A-1987-032).",
         "SRC-003",            "ACTIVE", "", 0),
    ]
    last = ecrire_donnees(ws, 5, blocks, fmt_cols={8: "0"}, align=AL)

    # USABILITY : USABLE si SOURCE_STATUS = ACTIVE
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=7, value=f'=IF(F{r}="ACTIVE","USABLE","REVIEW_REQUIRED")')
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = AC
        c.border = BF

    ajouter_table(ws, "tblNarrativeBlocks", f"A4:H{last}")

    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"USABLE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"REVIEW_REQUIRED"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_citation_control(wb):
    ws = wb.create_sheet("6_CITATION_CONTROL")
    titre(ws, "CITATION CONTROL", "Controle des citations dans les sorties finales", 7)
    en_tetes(ws, 4, ["CTRL_ID", "BLOCK_ID", "SOURCE_COUNT", "PRIMARY_COUNT", "STALE_SOURCE_COUNT", "STATUS", "NOTE"],
             [12, 12, 16, 16, 20, 14, 36])
    ctrl = [
        ("CC-001", "BL-001", 2, 2, 0, "", "2 sources primaires — OK"),
        ("CC-002", "BL-002", 2, 1, 0, "", "1 primaire + 1 secondaire"),
        ("CC-003", "BL-003", 1, 0, 0, "", "1 secondaire, acceptable"),
        ("CC-004", "BL-004", 2, 0, 0, "", "2 secondaires"),
        ("CC-005", "BL-005", 1, 1, 0, "", "1 primaire"),
        ("CC-006", "BL-006", 1, 0, 0, "", "1 secondaire"),
    ]
    last = ecrire_donnees(ws, 5, ctrl, fmt_cols={3: "0", 4: "0", 5: "0"}, align=AL)

    for r in range(5, last + 1):
        c = ws.cell(row=r, column=6, value=f'=IFS(C{r}=0,"BLOCKED",E{r}>0,"REVIEW",TRUE,"OK")')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ajouter_table(ws, "tblCitationControl", f"A4:G{last}")

    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"REVIEW"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"BLOCKED"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_outputs_log(wb):
    ws = wb.create_sheet("7_OUTPUTS_LOG")
    titre(ws, "OUTPUTS LOG", "Export vers MASTER", 14)
    en_tetes(ws, 4, ["RUN_ID", "AGENT_ID", "MAISON_ID", "DOC_ID", "TIMESTAMP", "LANG",
                      "FINAL_STATUS", "RISK_LEVEL", "GATE_TRIGGERED", "SLA_MET",
                      "TOTAL_MINUTES", "TOKENS_IN", "TOKENS_OUT", "COST_USD"],
             [16, 10, 10, 12, 18, 8, 14, 12, 14, 10, 12, 12, 12, 12])
    rows = [
        ("R-20260415-040", "AG-004", "M-001", "HQ-001", DT("2026-04-15 13:40"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 6.7,  2200, 1500, 0.047),
        ("R-20260415-041", "AG-004", "M-001", "HQ-002", DT("2026-04-15 14:20"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 5.2,  1800, 1200, 0.038),
        ("R-20260415-042", "AG-004", "M-001", "HQ-003", DT("2026-04-15 15:10"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 4.9,  1700, 1100, 0.035),
        ("R-20260415-043", "AG-004", "M-001", "HQ-004", DT("2026-04-15 16:00"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 5.6,  1900, 1300, 0.041),
        ("R-20260415-044", "AG-004", "M-001", "HQ-005", DT("2026-04-15 17:20"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 6.0,  2000, 1400, 0.044),
        ("R-20260415-045", "AG-004", "M-001", "HQ-006", DT("2026-04-15 18:00"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 4.5,  1600, 1000, 0.032),
        ("R-20260415-046", "AG-004", "M-001", "HQ-007", DT("2026-04-15 18:30"), "FR", "APPROVED", "LOW", "HERITAGE", "YES", 5.8,  1950, 1250, 0.040),
    ]
    last = ecrire_donnees(ws, 5, rows, fmt_cols={11: "0.00", 12: "0", 13: "0", 14: "0.000"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5])
    ajouter_table(ws, "tblHeritageLogs", f"A4:N{last}")


def onglet_dashboard(wb):
    ws = wb.create_sheet("8_DASHBOARD")
    titre(ws, "DASHBOARD HERITAGE", "Sources actives, citations, couverture", 4)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18

    section(ws, 4, 4, "KPIs", NOIR)
    kpis = [
        ("Sources actives",               '=COUNTIFS(tblSourceCatalog[STATUT],"ACTIVE")',          "0"),
        ("Sources STALE / REJECTED",      '=COUNTIFS(tblSourceCatalog[STATUT],"STALE")+COUNTIFS(tblSourceCatalog[STATUT],"REJECTED")', "0"),
        ("Sources PRIMARY",               '=COUNTIFS(tblSourceCatalog[TYPE],"PRIMARY")',           "0"),
        ("Faits approuves",               '=COUNTIFS(tblApprovedFacts[STATUT],"APPROVED")',        "0"),
        ("Narrative blocks USABLE",       '=COUNTIFS(tblNarrativeBlocks[SOURCE_STATUS],"ACTIVE")', "0"),
        ("Citations OK",                  '=COUNTIFS(tblCitationControl[STATUS],"OK")',            "0"),
        ("Citations BLOCKED",             '=COUNTIFS(tblCitationControl[STATUS],"BLOCKED")',       "0"),
        ("Cout USD cumule",               '=IFERROR(SUM(tblHeritageLogs[COST_USD]),0)',            "0.000"),
    ]
    for i, (lbl, formula, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = fmt

    section(ws, 15, 4, "Repartition par type", BORDEAUX)
    types = ["PRIMARY", "SECONDARY", "TERTIARY"]
    for i, t in enumerate(types):
        r = 16 + i
        ws.cell(row=r, column=2, value=t).border = BF
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True)
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(tblSourceCatalog[TYPE],"{t}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch = PieChart()
    data = Reference(ws, min_col=3, min_row=16, max_row=18)
    labels = Reference(ws, min_col=2, min_row=16, max_row=18)
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(labels)
    ch.title = "Sources par type"
    ch.height = 8
    ch.width = 14
    ws.add_chart(ch, "E15")


def onglet_learnings(wb):
    """R-13 : boucle d'apprentissage AG-004."""
    ws = wb.create_sheet("9_LEARNINGS")
    titre(ws, "LEARNINGS — AG-004", "Capitalisation : sources stale, citations manquantes, demandes rejetees", 8)
    en_tetes(ws, 4, ["LEARN_ID", "DATE", "SOURCE_RUN", "MOTIF", "FREQUENCE_MOIS", "ACTION_PROPOSEE", "OWNER", "STATUT"],
             [12, 12, 18, 40, 16, 44, 18, 14])
    lrn = [
        ("LRN-H-001", D(today_iso()), "-",              "Source SRC-005 STALE depuis 2025-01-10 encore citee dans 2 blocs", 2, "Pousser revalidation urgente", "Heritage Dir", "OPEN"),
        ("LRN-H-002", D(today_iso()), "-",              "TERTIARY sur 2 demandes - taux d'acceptation trop bas",          1, "Ecartage automatique sources TERTIARY", "Heritage Dir", "OPEN"),
        ("LRN-H-003", D(today_iso()), "-",              "HQ-002 (portrait fondateur) queries complexes repetees",          3, "Creer un bloc narratif canonique", "Heritage Dir", "OPEN"),
    ]
    last = ecrire_donnees(ws, 5, lrn, align=AL)
    convertir_dates(ws, 5, last, date_cols=[2])
    ajouter_table(ws, "tblLearningsAG004", f"A4:H{last}")
    attach_dv(ws, f"H5:H{last}", ["OPEN", "IN_REVIEW", "ACCEPTED", "REJECTED", "DONE"])


def build():
    wb = Workbook()
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_query_intake(wb)
    onglet_source_catalog(wb)
    onglet_approved_facts(wb)
    onglet_narrative_blocks(wb)
    onglet_citation_control(wb)
    onglet_outputs_log(wb)
    onglet_dashboard(wb)
    onglet_learnings(wb)

    order = ["0_README", "1_PARAMS", "2_QUERY_INTAKE", "3_SOURCE_CATALOG",
             "4_APPROVED_FACTS", "5_NARRATIVE_BLOCKS", "6_CITATION_CONTROL",
             "7_OUTPUTS_LOG", "8_DASHBOARD", "9_LEARNINGS"]
    wb._sheets = [wb[n] for n in order]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
