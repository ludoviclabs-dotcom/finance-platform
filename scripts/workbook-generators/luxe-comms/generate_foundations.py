"""
NEURAL_LUXE_COMMS_FOUNDATIONS.xlsx
Referentiel partage pour les 5 agents de la branche LUXE / Communication.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BLEU_NUIT, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, SEVERITY, STATUS, TERM_TYPE, VERT, VIOLET, WORDING_TYPE, YESNO,
    ajouter_table, ajuster_hauteurs, attach_dv, cf_stale, convertir_dates,
    D, DT, dv_liste, ecrire_donnees, en_tetes, params_table, param_ligne,
    section, titre, today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_LUXE_COMMS_FOUNDATIONS.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "NEURAL / LUXE / COMMUNICATION  —  FOUNDATIONS", "Referentiel partage  |  Maison, charte, vocabulaire, sources, medias, claims", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 82

    section(ws, 4, 6, "Objet du workbook", NOIR)
    donnees = [
        ["", "Role", "Referentiel transverse. Alimente les 5 agents via Power Query ou imports controles."],
        ["", "Principe", "Aucune logique de scoring ici. Seulement des referentiels stables a changer peu."],
        ["", "Proprietaire", "Brand & Heritage Office."],
        ["", "Mise a jour", "Revue mensuelle + changelog obligatoire onglet 10_CHANGELOG."],
        ["", "Multi-maison", "Chaque ligne porte MAISON_ID. V1 = 1 maison active par fichier."],
        ["", "Langues", "FR natif, EN partiel. IT / DE / JA / ZH prevues en v2."],
    ]
    ecrire_donnees(ws, 5, donnees, align=AL)

    section(ws, 12, 6, "Carte des onglets", BORDEAUX)
    carto = [
        ["", "1_MAISON_PROFILE", "Identite maison : ton, exclusions, lignes rouges, pays."],
        ["", "2_BRAND_CHARTER", "Regles de charte editoriale (ton, structure, formulation)."],
        ["", "3_BRAND_VOCAB_FR", "Vocabulaire FR : preferred / forbidden / preferred_missing."],
        ["", "4_BRAND_VOCAB_EN", "Miroir EN (v1 partielle)."],
        ["", "5_HERITAGE_SOURCEBOOK", "Sources patrimoniales + statut (active / stale)."],
        ["", "6_MEDIA_DIRECTORY", "Medias : angle editorial, vertical, statut relation."],
        ["", "7_CLAIMS_EVIDENCE_REGISTRY", "Claims + preuves + juridiction + expiry."],
        ["", "8_EVENTS_CALENDAR", "Evenements maison : type, lieu, sensibilite, scoring."],
        ["", "9_LISTS_REFERENTIALS", "Listes statiques de validation (enums, roles, types)."],
        ["", "10_CHANGELOG", "Journal des modifications."],
    ]
    ecrire_donnees(ws, 13, carto, align=AL)

    section(ws, 25, 6, "Interface avec les agents", VIOLET)
    iface = [
        ["", "AG-001 VoiceGuard", "Consomme 2_BRAND_CHARTER, 3_BRAND_VOCAB_FR, 4_BRAND_VOCAB_EN."],
        ["", "AG-002 LuxePressAgent", "Consomme 6_MEDIA_DIRECTORY. Publie claims vers 7_CLAIMS."],
        ["", "AG-003 LuxeEventComms", "Consomme 8_EVENTS_CALENDAR. Demande blocs a AG-004."],
        ["", "AG-004 HeritageComms", "Consomme 5_HERITAGE_SOURCEBOOK. Publie narrative blocks."],
        ["", "AG-005 GreenClaimChecker", "Consomme 7_CLAIMS_EVIDENCE_REGISTRY."],
    ]
    ecrire_donnees(ws, 26, iface, align=AL)


def onglet_maison_profile(wb):
    ws = wb.create_sheet("1_MAISON_PROFILE")
    titre(ws, "MAISON — Profil de reference", "Identite, ton, lignes rouges, pays de presence", 4)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18

    section(ws, 4, 4, "Parametres maison active", NOIR)

    params = [
        ("MAISON_ID", "M-001", "Identifiant unique maison"),
        ("MAISON_NAME", "Maison pilote", "Raison sociale / enseigne"),
        ("CATEGORIES", "Haute couture, Joaillerie, Parfumerie", "Piliers produit"),
        ("SIEGE", "Paris, France", "Pays siege"),
        ("PAYS_CIBLES", "FR, IT, UK, DE, US, JP, KR, CN, UAE", "Marches adressables"),
        ("LANG_PRIMARY", "FR", "Langue principale"),
        ("LANG_SECONDARY", "EN", "Langue secondaire"),
        ("ANNEE_FONDATION", 1920, "Annee historique"),
        ("TON_MAITRE", "Raffine, intemporel, allusif, jamais ostentatoire", "Registre editorial"),
        ("DO_LIST", "Savoir-faire; Geste artisanal; Allusion patrimoniale", "Elements obligatoires"),
        ("DONT_LIST", "Superlatifs; Discount; Exclusivite revendiquee; Claims absolus", "Interdictions"),
        ("LIGNES_ROUGES", "Greenwashing; Revendications medicales; Comparatif concurrent", "Tabous"),
        ("CODE_DEONTO", "Charte 2024 v1.3 — signataire FHL & FFPAPF", "Charte / engagement"),
        ("ESG_PILIERS", "Or recycle; Cuir LWG; Eau; Emplois artisans", "Axes ESG autorises"),
        ("REVIEW_REQUIRED_DEFAULT", "YES", "Revue humaine obligatoire par defaut"),
        ("CRISIS_MODE_ON", "NO", "Fast-track actif ? (YES bascule gate d'urgence)"),
        ("DATA_VERSION", "v2026.04", "Version du referentiel"),
        ("PROMPT_VERSION", "v1.0.0", "Version des prompts agents"),
        ("LAST_REFRESH", today_iso(), "Derniere mise a jour"),
    ]
    for i, (k, v, note) in enumerate(params):
        param_ligne(ws, 5 + i, k, v, note)

    # Named range pour ref croisee
    try:
        from openpyxl.workbook.defined_name import DefinedName
        wb.defined_names["MAISON_ID_ACTIVE"] = DefinedName("MAISON_ID_ACTIVE", attr_text="'1_MAISON_PROFILE'!$B$5")
        wb.defined_names["MAISON_NAME_ACTIVE"] = DefinedName("MAISON_NAME_ACTIVE", attr_text="'1_MAISON_PROFILE'!$B$6")
    except Exception:
        pass


def onglet_brand_charter(wb):
    ws = wb.create_sheet("2_BRAND_CHARTER")
    titre(ws, "CHARTE EDITORIALE", "Regles de ton, structure, formulation — lues par AG-001 VoiceGuard", 8)
    headers = ["RULE_ID", "CATEGORIE", "REGLE", "NIVEAU", "POIDS_PENALITE", "LANG", "EXEMPLE_OK", "EXEMPLE_KO"]
    en_tetes(ws, 4, headers, [12, 18, 62, 14, 16, 10, 42, 42])

    rules = [
        ("R-001", "Tone", "Registre allusif, jamais declamatoire",                      "CRITICAL", 25, "FR", "L'atelier dessine un bijou a la main, comme en 1920.",     "Le meilleur bijou du monde."),
        ("R-002", "Tone", "Pas de superlatif absolu (meilleur, le plus, unique)",        "HIGH",     15, "FR", "Une piece rare issue d'un savoir-faire rare.",              "La piece la plus extraordinaire jamais concue."),
        ("R-003", "Structure", "Accroche < 140 caracteres, puis expose factuel",         "MEDIUM",   10, "FR", "Un bracelet ciselé au marteau. Collection printemps.",       "Bracelet extraordinaire en or exceptionnel unique au monde."),
        ("R-004", "Identite", "Eviter nom des concurrents et comparatifs",                "CRITICAL", 25, "FR", "Nos ateliers francais.",                                     "Superieur a [Concurrent]."),
        ("R-005", "Claim", "Toute affirmation RSE doit etre prouvee (voir AG-005)",      "CRITICAL", 25, "FR", "Or certifie 80% recycle, source fournisseur LBMA.",          "Entierement eco-responsable."),
        ("R-006", "Heritage", "Toute reference historique datee doit etre sourcee",      "HIGH",     15, "FR", "Archives maison 1987 (cote A-2014-032).",                   "Depuis toujours, nous..."),
        ("R-007", "Formulation", "Utiliser le present de narration, pas le futur vendeur","MEDIUM",   10, "FR", "L'artisan polit la surface pendant 36 heures.",             "Vous allez adorer cette piece."),
        ("R-008", "Formulation", "Pas d'exhortation commerciale ('achetez', 'promo')",   "HIGH",     15, "FR", "Piece numerotee, serie limitee a 48 exemplaires.",           "Achetez maintenant, stock limite !"),
        ("R-009", "Legal", "Pas de revendication sante / medicale",                       "CRITICAL", 25, "FR", "Extrait de rose centifolia.",                                "Effet anti-age prouve."),
        ("R-010", "Tone", "Allusion > revendication directe",                             "MEDIUM",   10, "FR", "Un geste qui traverse les generations.",                    "Le meilleur savoir-faire du monde."),
        ("R-011", "Identite", "Maintenir cohesion vocabulaire (voir 3_BRAND_VOCAB_FR)",  "HIGH",     15, "FR", "Atelier, savoir-faire, geste.",                             "Usine, production, fabrication de masse."),
        ("R-012", "Inclusive", "Pas de formulation genree exclusive",                     "MEDIUM",   10, "FR", "Celle ou celui qui porte.",                                 "L'homme accompli."),
        ("R-013", "Tone", "Neutral prestige tone in EN",                                 "CRITICAL", 25, "EN", "A piece crafted in our Parisian atelier.",                  "The best luxury product ever made."),
        ("R-014", "Claim", "Do not use 'sustainable' without evidence",                   "CRITICAL", 25, "EN", "Gold certified 80% recycled, LBMA audited.",                "Fully sustainable collection."),
        ("R-015", "Crisis", "En crisis mode (flag on), by-pass SLA mais review maintenue","HIGH",     15, "FR", "Communication de crise validee en < 2h.",                   "Publication sans review."),
    ]
    last = ecrire_donnees(ws, 5, rules, fmt_cols={5: "0"}, align=AL)

    # Conditional formatting sur NIVEAU
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=PatternFill(start_color=FOND_BLEU, end_color=FOND_BLEU, fill_type="solid")))

    ajouter_table(ws, "tblBrandRulesFound", f"A4:H{last}")
    # R-06 : DV sur NIVEAU et LANG
    attach_dv(ws, f"D5:D{last}", ["CRITICAL", "HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"F5:F{last}", LANG)


def onglet_vocab_fr(wb):
    ws = wb.create_sheet("3_BRAND_VOCAB_FR")
    titre(ws, "VOCABULAIRE FR", "Lexique normatif consomme par AG-001 (scoring) et AG-002 (drafting)", 8)
    headers = ["TERM_ID", "TERME", "TERM_TYPE", "CATEGORIE", "NIVEAU", "SUGGESTION_REMPLACEMENT", "CONTEXTE", "ACTION"]
    en_tetes(ws, 4, headers, [10, 28, 18, 18, 12, 32, 32, 14])

    vocab = [
        ("VOC-FR-001", "savoir-faire",       "PREFERRED", "Heritage",    "HIGH",     "-",                           "Toujours valorise",             "BONUS"),
        ("VOC-FR-002", "geste artisanal",    "PREFERRED", "Heritage",    "HIGH",     "-",                           "Toujours valorise",             "BONUS"),
        ("VOC-FR-003", "maison",             "PREFERRED", "Identite",    "MEDIUM",   "-",                           "Remplace 'marque'",             "BONUS"),
        ("VOC-FR-004", "atelier",            "PREFERRED", "Identite",    "MEDIUM",   "-",                           "Remplace 'usine'",              "BONUS"),
        ("VOC-FR-005", "ciselure",           "PREFERRED", "Technique",   "MEDIUM",   "-",                           "Vocable metier",                "BONUS"),
        ("VOC-FR-006", "intemporel",         "PREFERRED", "Positionnement","LOW",    "-",                           "Allusif",                        "BONUS"),
        ("VOC-FR-007", "patrimoine",         "PREFERRED", "Heritage",    "HIGH",     "-",                           "Valorise ancrage",              "BONUS"),
        ("VOC-FR-008", "discount",           "FORBIDDEN", "Commercial",  "CRITICAL", "piece numerotee, serie limitee","Jamais tolere",                 "BLOCK"),
        ("VOC-FR-009", "promo",              "FORBIDDEN", "Commercial",  "CRITICAL", "selection",                   "Jamais tolere",                 "BLOCK"),
        ("VOC-FR-010", "le meilleur",        "FORBIDDEN", "Superlatif",  "CRITICAL", "une piece rare",              "Superlatif absolu",             "BLOCK"),
        ("VOC-FR-011", "unique au monde",    "FORBIDDEN", "Superlatif",  "CRITICAL", "piece numerotee",             "Superlatif absolu",             "BLOCK"),
        ("VOC-FR-012", "eco-responsable",    "FORBIDDEN", "Claim RSE",   "CRITICAL", "avec preuve AG-005 uniquement","Declenche gate AG-005 si utilise","BLOCK"),
        ("VOC-FR-013", "eternel",            "FORBIDDEN", "Superlatif",  "HIGH",     "intemporel",                  "Abus",                           "BLOCK"),
        ("VOC-FR-014", "fabrication",        "FORBIDDEN", "Identite",    "HIGH",     "atelier / geste artisanal",   "Connotation industrielle",      "BLOCK"),
        ("VOC-FR-015", "usine",              "FORBIDDEN", "Identite",    "CRITICAL", "atelier",                     "Hors charte",                    "BLOCK"),
        ("VOC-FR-016", "no 1 mondial",       "FORBIDDEN", "Superlatif",  "CRITICAL", "-",                           "Comparatif prohibe",            "BLOCK"),
        ("VOC-FR-017", "anti-age",           "FORBIDDEN", "Claim sante", "CRITICAL", "-",                           "Revendication medicale",        "BLOCK"),
        ("VOC-FR-018", "natural",            "FORBIDDEN", "Claim RSE",   "HIGH",     "ingredient d'origine naturelle","Greenwashing potentiel",       "REVIEW"),
        ("VOC-FR-019", "sans parabenes",     "FORBIDDEN", "Claim RSE",   "MEDIUM",   "formulation sans parabenes",  "Claim negatif regule",          "REVIEW"),
        ("VOC-FR-020", "exclusif",           "REVIEW",    "Superlatif",  "MEDIUM",   "reserve, numerotee",          "A encadrer",                    "REVIEW"),
        ("VOC-FR-021", "artisan",            "PREFERRED", "Identite",    "HIGH",     "-",                           "Valorise metier",               "BONUS"),
        ("VOC-FR-022", "rare",               "PREFERRED", "Positionnement","LOW",    "-",                           "Allusif",                        "BONUS"),
        ("VOC-FR-023", "prestige",           "REVIEW",    "Positionnement","MEDIUM", "-",                           "Eviter en exces",               "INFO"),
        ("VOC-FR-024", "cuir exotique",      "REVIEW",    "Materiel",    "HIGH",     "cuir certifie",               "Regulation CITES",              "REVIEW"),
        ("VOC-FR-025", "fait main",          "PREFERRED", "Technique",   "MEDIUM",   "-",                           "Si vrai uniquement",            "BONUS"),
    ]
    last = ecrire_donnees(ws, 5, vocab, align=AL)

    # Conditional formatting sur TERM_TYPE
    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"FORBIDDEN"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"PREFERRED"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"REVIEW"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))

    ajouter_table(ws, "tblVocabFR", f"A4:H{last}")
    # R-06 : DV sur TERM_TYPE et ACTION
    attach_dv(ws, f"C5:C{last}", TERM_TYPE)
    attach_dv(ws, f"E5:E{last}", ["CRITICAL", "HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"H5:H{last}", ["BLOCK", "REVIEW", "BONUS", "INFO"])


def onglet_vocab_en(wb):
    ws = wb.create_sheet("4_BRAND_VOCAB_EN")
    titre(ws, "BRAND VOCABULARY — EN", "Mirror of FR vocab. v1 partial, to be completed in v2.", 8)
    headers = ["TERM_ID", "TERM", "TERM_TYPE", "CATEGORY", "SEVERITY", "REPLACEMENT", "CONTEXT", "ACTION"]
    en_tetes(ws, 4, headers, [10, 26, 16, 16, 12, 30, 30, 14])

    vocab = [
        ("VOC-EN-001", "craftsmanship",      "PREFERRED", "Heritage",    "HIGH",     "-",                              "Always valued",            "BONUS"),
        ("VOC-EN-002", "maison",             "PREFERRED", "Identity",    "MEDIUM",   "-",                              "Preferred over 'brand'",   "BONUS"),
        ("VOC-EN-003", "atelier",            "PREFERRED", "Identity",    "MEDIUM",   "-",                              "Preferred over 'factory'", "BONUS"),
        ("VOC-EN-004", "heritage",           "PREFERRED", "Heritage",    "HIGH",     "-",                              "Positive anchor",          "BONUS"),
        ("VOC-EN-005", "discount",           "FORBIDDEN", "Commercial",  "CRITICAL", "numbered piece / limited run",   "Never acceptable",          "BLOCK"),
        ("VOC-EN-006", "sale",               "FORBIDDEN", "Commercial",  "CRITICAL", "selection",                      "Never acceptable",          "BLOCK"),
        ("VOC-EN-007", "the best",           "FORBIDDEN", "Superlative", "CRITICAL", "a rare piece",                   "Absolute superlative",      "BLOCK"),
        ("VOC-EN-008", "sustainable",        "FORBIDDEN", "ESG claim",   "CRITICAL", "requires AG-005 evidence",       "Triggers AG-005 gate",      "BLOCK"),
        ("VOC-EN-009", "eco-friendly",       "FORBIDDEN", "ESG claim",   "CRITICAL", "requires AG-005 evidence",       "Triggers AG-005 gate",      "BLOCK"),
        ("VOC-EN-010", "green",              "REVIEW",    "ESG claim",   "HIGH",     "specify criteria",               "Vague unless evidenced",    "REVIEW"),
        ("VOC-EN-011", "anti-aging",         "FORBIDDEN", "Health claim","CRITICAL", "-",                              "Medical claim regulated",   "BLOCK"),
        ("VOC-EN-012", "factory",            "FORBIDDEN", "Identity",    "CRITICAL", "atelier",                        "Off-brand",                 "BLOCK"),
        ("VOC-EN-013", "unique",             "REVIEW",    "Superlative", "MEDIUM",   "limited / rare",                 "Use sparingly",             "REVIEW"),
        ("VOC-EN-014", "handcrafted",        "PREFERRED", "Technique",   "MEDIUM",   "-",                              "Only if true",              "BONUS"),
        ("VOC-EN-015", "artisan",            "PREFERRED", "Identity",    "HIGH",     "-",                              "Valued",                    "BONUS"),
    ]
    last = ecrire_donnees(ws, 5, vocab, align=AL)

    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"FORBIDDEN"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"PREFERRED"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"C5:C{last}", CellIsRule(operator="equal", formula=['"REVIEW"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))

    ajouter_table(ws, "tblVocabEN", f"A4:H{last}")
    attach_dv(ws, f"C5:C{last}", TERM_TYPE)
    attach_dv(ws, f"E5:E{last}", ["CRITICAL", "HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"H5:H{last}", ["BLOCK", "REVIEW", "BONUS", "INFO"])


def onglet_heritage(wb):
    ws = wb.create_sheet("5_HERITAGE_SOURCEBOOK")
    titre(ws, "HERITAGE SOURCEBOOK", "Sources patrimoniales maison — validees, datees, referencees AG-004", 11)
    # R-03 : STATUT devient formule ; MANUAL_OVERRIDE permet de forcer REJECTED
    headers = ["SOURCE_ID", "TITRE", "TYPE", "ANNEE", "AUTEUR", "COTE_ARCHIVE", "MANUAL_OVERRIDE",
               "REVIEW_DATE", "CITATION_FORMAT", "USAGE_COUNT", "STATUT"]
    en_tetes(ws, 4, headers, [12, 36, 14, 10, 26, 20, 18, 14, 20, 14, 14])

    # R-01 : dates typees
    sources = [
        ("SRC-001", "Carnet de dessins du fondateur",       "PRIMARY",   1923, "Fondateur maison",     "A-1923-001", "",          D("2026-10-01"), "Maison-style",    4, ""),
        ("SRC-002", "Lettre a la direction atelier",        "PRIMARY",   1947, "Directeur atelier",    "A-1947-112", "",          D("2026-09-15"), "Chicago",         2, ""),
        ("SRC-003", "Catalogue exposition retrospective",   "SECONDARY", 1987, "Musee Arts Deco",      "A-1987-032", "",          D("2026-07-20"), "Chicago",         7, ""),
        ("SRC-004", "Livre monographie maison",             "SECONDARY", 2005, "Historien partenaire", "A-2005-008", "",          D("2026-12-01"), "APA",             3, ""),
        ("SRC-005", "Article Le Figaro edition speciale",   "TERTIARY",  2012, "Journaliste presse",   "A-2012-044", "",          D("2025-01-10"), "Chicago",         1, ""),
        ("SRC-006", "Photos atelier annees 1950",           "PRIMARY",   1952, "Photographe maison",   "A-1952-007", "",          D("2026-06-30"), "Maison-style",    5, ""),
        ("SRC-007", "Interview fondateur",                  "PRIMARY",   1965, "Emission TV",          "A-1965-099", "",          D("2026-11-15"), "Chicago",         2, ""),
        ("SRC-008", "Piece Musee Galliera, inv 12-09",      "SECONDARY", 1935, "Musee Galliera",       "A-1935-401", "",          D("2026-08-01"), "Chicago",         1, ""),
        ("SRC-009", "Document juridique marque deposee",    "PRIMARY",   1920, "INPI",                 "A-1920-001", "",          D("2027-01-01"), "Juridique",       0, ""),
        ("SRC-010", "Blog post fan non officiel",           "TERTIARY",  2019, "Blogueur tiers",       "A-2019-777", "REJECTED",  D("2024-12-01"), "Non utilisable",  0, ""),
    ]
    last = ecrire_donnees(ws, 5, sources)
    convertir_dates(ws, 5, last, date_cols=[8])

    # R-03 : STATUT = IF(MANUAL_OVERRIDE<>"", MANUAL_OVERRIDE, IF(REVIEW_DATE<TODAY(),"STALE","ACTIVE"))
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=11, value=f'=IF(G{r}<>"",G{r},IF(H{r}<TODAY(),"STALE","ACTIVE"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ws.conditional_formatting.add(f"K5:K{last}", CellIsRule(operator="equal", formula=['"ACTIVE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"K5:K{last}", CellIsRule(operator="equal", formula=['"STALE"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"K5:K{last}", CellIsRule(operator="equal", formula=['"REJECTED"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    # R-08 : CF J-30 sur REVIEW_DATE
    cf_stale(ws, f"H5:H{last}", days_window=30)

    ajouter_table(ws, "tblHeritageSources", f"A4:K{last}")
    # R-06 : DV sur TYPE et MANUAL_OVERRIDE
    attach_dv(ws, f"C5:C{last}", ["PRIMARY", "SECONDARY", "TERTIARY"])
    attach_dv(ws, f"G5:G{last}", ["", "REJECTED"])


def onglet_media(wb):
    ws = wb.create_sheet("6_MEDIA_DIRECTORY")
    titre(ws, "MEDIA DIRECTORY", "Repertoire presse — angle editorial, vertical, statut relation", 10)
    headers = ["MEDIA_ID", "NOM_MEDIA", "TYPE", "VERTICAL", "PAYS", "LANG", "ANGLE_EDITORIAL", "EMBARGO_ACCEPTED", "RELATION_STATUS", "PRIORITE"]
    en_tetes(ws, 4, headers, [12, 30, 14, 16, 12, 10, 36, 18, 18, 14])

    medias = [
        ("MED-001", "Vogue",              "Magazine",  "Lifestyle", "FR", "FR", "Narration image / maison",         "YES", "STRONG",   "P1"),
        ("MED-002", "Harper's Bazaar",    "Magazine",  "Lifestyle", "US", "EN", "Narration style / collection",     "YES", "STRONG",   "P1"),
        ("MED-003", "Financial Times",    "Presse",    "Business",  "UK", "EN", "Resultats / expansion / direction", "NO",  "STRONG",   "P1"),
        ("MED-004", "Les Echos",          "Presse",    "Business",  "FR", "FR", "Resultats / investissement",        "YES", "MEDIUM",   "P2"),
        ("MED-005", "Le Monde",           "Presse",    "Generaliste","FR", "FR", "Patrimoine / culture / societe",   "YES", "STRONG",   "P1"),
        ("MED-006", "WWD",                "Presse",    "Trade",     "US", "EN", "Industrie mode / retail",          "YES", "STRONG",   "P1"),
        ("MED-007", "Business of Fashion","Digital",   "Trade",     "UK", "EN", "Strategie / innovation",            "YES", "STRONG",   "P1"),
        ("MED-008", "T Magazine NYT",     "Magazine",  "Lifestyle", "US", "EN", "Culture / art de vivre",            "YES", "MEDIUM",   "P2"),
        ("MED-009", "Madame Figaro",      "Magazine",  "Lifestyle", "FR", "FR", "Lifestyle premium",                 "YES", "STRONG",   "P1"),
        ("MED-010", "L'Officiel",         "Magazine",  "Lifestyle", "FR", "FR", "Mode / patrimoine",                 "YES", "STRONG",   "P1"),
        ("MED-011", "GQ France",          "Magazine",  "Lifestyle", "FR", "FR", "Style masculin",                    "YES", "MEDIUM",   "P2"),
        ("MED-012", "Numero",             "Magazine",  "Lifestyle", "FR", "FR", "Avant-garde / art",                 "YES", "MEDIUM",   "P2"),
        ("MED-013", "Le Point",           "Presse",    "Generaliste","FR", "FR", "Economie / culture",                "YES", "MEDIUM",   "P2"),
        ("MED-014", "Vogue Business",     "Digital",   "Trade",     "UK", "EN", "Strategie / digital",               "YES", "STRONG",   "P1"),
        ("MED-015", "Jing Daily",         "Digital",   "Trade",     "CN", "EN", "Marche chinois / luxe",             "YES", "STRONG",   "P1"),
        ("MED-016", "WEIBO KOL partner",  "Social",    "Lifestyle", "CN", "ZH", "Amplification Chine",               "NO",  "MEDIUM",   "P2"),
        ("MED-017", "Nikkei Style",       "Magazine",  "Lifestyle", "JP", "JA", "Marche japonais",                   "YES", "MEDIUM",   "P2"),
        ("MED-018", "AD France",          "Magazine",  "Lifestyle", "FR", "FR", "Design / interieur",                "YES", "MEDIUM",   "P3"),
        ("MED-019", "Vanity Fair France", "Magazine",  "Lifestyle", "FR", "FR", "Culture / people",                  "YES", "MEDIUM",   "P2"),
        ("MED-020", "Bilan",              "Presse",    "Business",  "CH", "FR", "Horlogerie / joaillerie suisse",    "YES", "MEDIUM",   "P2"),
    ]
    last = ecrire_donnees(ws, 5, medias, align=AL)

    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"STRONG"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"COLD"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))

    ajouter_table(ws, "tblMediaDirectory", f"A4:J{last}")
    attach_dv(ws, f"F5:F{last}", LANG)
    attach_dv(ws, f"I5:I{last}", ["STRONG", "MEDIUM", "COLD"])
    attach_dv(ws, f"J5:J{last}", ["P1", "P2", "P3"])
    attach_dv(ws, f"H5:H{last}", YESNO)


def onglet_claims(wb):
    ws = wb.create_sheet("7_CLAIMS_EVIDENCE_REGISTRY")
    titre(ws, "CLAIMS EVIDENCE REGISTRY", "Registre claims + preuves + juridiction + expiry — source de verite AG-005", 12)
    # R-03 : MANUAL_OVERRIDE pour forcer UNVERIFIED / MISSING ; STATUS = formule
    headers = ["CLAIM_ID", "CLAIM", "CATEGORIE", "WORDING_TYPE", "EVIDENCE_TITLE", "EVIDENCE_SOURCE",
               "JURIDICTION", "EVIDENCE_DATE", "EVIDENCE_EXPIRY", "OWNER", "MANUAL_OVERRIDE", "STATUS"]
    en_tetes(ws, 4, headers, [12, 46, 18, 16, 36, 28, 14, 14, 14, 18, 18, 14])

    claims = [
        ("CLM-001", "Or certifie 80% recycle",                         "ESG",   "QUALIFIED", "Attestation fournisseur",   "LBMA audit 2026",         "EU",  D("2026-01-15"), D("2027-01-15"), "ESG Lead",       "", ""),
        ("CLM-002", "Cuir d'origine certifiee LWG",                    "ESG",   "QUALIFIED", "Certificat LWG",             "Leather Working Group",   "EU",  D("2025-11-01"), D("2026-11-01"), "ESG Lead",       "", ""),
        ("CLM-003", "Zero dechet atelier",                              "ESG",   "ABSOLUTE",  "",                           "",                         "EU",  None,            None,            "ESG Lead",       "UNVERIFIED", ""),
        ("CLM-004", "Eau utilisee -30% vs 2019",                       "ESG",   "QUALIFIED", "Rapport RSE 2025",           "Audit Ernst & Young 2025","EU",  D("2025-06-30"), D("2026-06-30"), "ESG Lead",       "", ""),
        ("CLM-005", "Diamants sans conflit (Kimberley)",               "ESG",   "QUALIFIED", "Certificat Kimberley Process","Fournisseur certifie",    "EU",  D("2026-02-20"), D("2027-02-20"), "Supply Lead",    "", ""),
        ("CLM-006", "Produit vegan",                                    "ESG",   "ABSOLUTE",  "Certification PETA",         "PETA",                    "EU",  D("2025-09-10"), D("2026-09-10"), "Product Lead",   "", ""),
        ("CLM-007", "Fabrique en France",                               "ORIGIN","QUALIFIED", "Label IGP atelier",          "DGCCRF",                  "FR",  D("2025-12-01"), D("2027-12-01"), "Legal",          "", ""),
        ("CLM-008", "Fait main integralement",                          "TECH",  "ABSOLUTE",  "Video atelier + process",    "Interne",                 "FR",  D("2026-01-01"), D("2026-12-31"), "Atelier Dir",    "", ""),
        ("CLM-009", "Eco-responsable",                                  "ESG",   "ABSOLUTE",  "",                           "",                         "EU",  None,            None,            "ESG Lead",       "", ""),
        ("CLM-010", "Empreinte carbone neutre",                         "ESG",   "ABSOLUTE",  "",                           "",                         "EU",  None,            None,            "ESG Lead",       "", ""),
        ("CLM-011", "Cruelty-free",                                     "ESG",   "ABSOLUTE",  "Certificat Leaping Bunny",   "Cruelty Free International","EU", D("2025-07-01"), D("2026-07-01"), "Product Lead",   "", ""),
        ("CLM-012", "Serie limitee a 48 pieces",                        "SCARCITY","QUALIFIED","Production book serialise","Interne",                  "EU",  D("2026-03-01"), D("2027-03-01"), "Atelier Dir",    "", ""),
        ("CLM-013", "Utilise au Palais de l'Elysee",                    "PROVENANCE","QUALIFIED","Archives maison",         "SRC-003",                 "FR",  D("2018-05-01"), None,            "Heritage Dir",   "", ""),
        ("CLM-014", "Porte a Cannes 1987 par [celebrite]",             "PROVENANCE","QUALIFIED","Photo agence + accord",    "Archives presse",         "EU",  D("1987-05-10"), None,            "Heritage Dir",   "", ""),
        ("CLM-015", "Or du Perou ethiquement extrait",                  "ESG",   "QUALIFIED", "Certificat Fairmined",       "ARM",                     "EU",  D("2025-10-01"), D("2026-10-01"), "Supply Lead",    "", ""),
        ("CLM-016", "Sans plastique d'emballage",                       "ESG",   "ABSOLUTE",  "Packaging spec + audit",     "Audit packaging 2025",   "EU",  D("2025-08-01"), D("2026-08-01"), "Packaging Lead", "", ""),
        ("CLM-017", "Emplois preserves artisans maroquiniers",          "ESG",   "QUALIFIED", "Rapport social 2025",        "Audit social",            "FR",  D("2025-12-31"), D("2026-12-31"), "HR Dir",         "", ""),
    ]
    last = ecrire_donnees(ws, 5, claims, align=AL)
    convertir_dates(ws, 5, last, date_cols=[8, 9])

    # R-03 : STATUS = formule (MISSING si pas d'evidence, STALE si expiree, VALID sinon ; override prioritaire)
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=12,
                     value=f'=IF(K{r}<>"",K{r},IF(OR(E{r}="",E{r}="-"),"MISSING",IF(AND(ISNUMBER(I{r}),I{r}<TODAY()),"STALE","VALID")))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ws.conditional_formatting.add(f"L5:L{last}", CellIsRule(operator="equal", formula=['"VALID"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"L5:L{last}", CellIsRule(operator="equal", formula=['"STALE"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"L5:L{last}", CellIsRule(operator="equal", formula=['"UNVERIFIED"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"L5:L{last}", CellIsRule(operator="equal", formula=['"MISSING"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    # R-08 : CF J-30 sur EVIDENCE_EXPIRY
    cf_stale(ws, f"I5:I{last}", days_window=30)

    ajouter_table(ws, "tblClaimsRegistry", f"A4:L{last}")
    # R-06 : DV
    attach_dv(ws, f"D5:D{last}", WORDING_TYPE)
    attach_dv(ws, f"K5:K{last}", ["", "UNVERIFIED", "MISSING"])
    attach_dv(ws, f"G5:G{last}", ["EU", "FR", "UK", "US", "CH", "EU + UK", "EU + US"])


def onglet_events(wb):
    ws = wb.create_sheet("8_EVENTS_CALENDAR")
    titre(ws, "EVENTS CALENDAR", "Calendrier evenementiel maison — fournit AG-003", 12)
    headers = ["EVENT_ID", "NOM", "TYPE", "DATE_DEBUT", "DATE_FIN", "LIEU", "AUDIENCE", "VIP_LEVEL",
               "CLAIMS_EXPECTED", "HERITAGE_ANGLE", "SENSIBILITE_SCORE", "STATUT"]
    en_tetes(ws, 4, headers, [12, 36, 22, 12, 12, 22, 20, 12, 16, 16, 18, 14])

    events = [
        ("EV-001", "Defile haute couture printemps",          "Defile",       D("2026-07-05"), D("2026-07-05"), "Paris, Palais de Tokyo",   "VIP / Presse / Clients", "HIGH",  "YES", "YES", "", "PLANNED"),
        ("EV-002", "Exposition retrospective 100 ans",        "Exhibition",   D("2026-09-15"), D("2026-12-15"), "Musee des Arts Deco",      "Public / Presse",        "HIGH",  "NO",  "YES", "", "PLANNED"),
        ("EV-003", "Lancement parfum capsule automne",        "Launch",       D("2026-10-01"), D("2026-10-01"), "Paris, flagship",          "Presse / VIP",           "MEDIUM","YES", "NO",  "", "PLANNED"),
        ("EV-004", "Cocktail Cannes 2026",                    "Cocktail",     D("2026-05-15"), D("2026-05-15"), "Cannes, hotel partenaire", "Celebrites / Presse",    "HIGH",  "NO",  "NO",  "", "CONFIRMED"),
        ("EV-005", "Salon joaillerie Biennale",               "Salon",        D("2026-09-22"), D("2026-09-29"), "Paris, Grand Palais",      "Collectionneurs",        "HIGH",  "YES", "YES", "", "PLANNED"),
        ("EV-006", "Rencontre artisans atelier",               "PR event",    D("2026-04-18"), D("2026-04-18"), "Atelier historique",       "Journalistes trade",     "MEDIUM","NO",  "YES", "", "CONFIRMED"),
        ("EV-007", "Collaboration artiste contemporain",       "Launch",      D("2026-11-10"), D("2026-11-10"), "Paris, galerie",           "Art / lifestyle presse", "HIGH",  "YES", "YES", "", "PLANNED"),
        ("EV-008", "Client day Hong Kong",                     "Client event",D("2026-10-28"), D("2026-10-29"), "Hong Kong",                "Top clients Asie",       "HIGH",  "NO",  "NO",  "", "PLANNED"),
        ("EV-009", "Conference ESG Milan",                     "Conference",  D("2026-06-12"), D("2026-06-12"), "Milan",                    "Trade / ESG analysts",   "MEDIUM","YES", "NO",  "", "PLANNED"),
        ("EV-010", "Pop-up Tokyo Harajuku",                    "Pop-up",      D("2026-08-01"), D("2026-08-31"), "Tokyo",                    "Lifestyle jeune",        "MEDIUM","YES", "NO",  "", "PLANNED"),
        ("EV-011", "Gala charity Noel",                        "Charity",     D("2026-12-08"), D("2026-12-08"), "Paris, Opera",             "HNW / celebrites",        "HIGH",  "NO",  "NO",  "", "PLANNED"),
        ("EV-012", "Week Chinese New Year activation",         "Activation",  D("2026-02-08"), D("2026-02-18"), "Shanghai, Pekin",          "Marche chinois",         "MEDIUM","YES", "YES", "", "CONFIRMED"),
    ]
    last = ecrire_donnees(ws, 5, events, align=AL)
    convertir_dates(ws, 5, last, date_cols=[4, 5])

    # Sensibilite_score formule
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=11, value=f'=IF(H{r}="HIGH",3,1)+IF(I{r}="YES",2,0)+IF(J{r}="YES",1,0)')
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = AC
        c.border = BF

    ajouter_table(ws, "tblEventsCalendar", f"A4:L{last}")
    attach_dv(ws, f"H5:H{last}", ["HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"I5:I{last}", YESNO)
    attach_dv(ws, f"J5:J{last}", YESNO)
    attach_dv(ws, f"L5:L{last}", ["PLANNED", "CONFIRMED", "ONGOING", "DONE", "CANCELLED"])


def onglet_lists(wb):
    ws = wb.create_sheet("9_LISTS_REFERENTIALS")
    titre(ws, "LISTES DE VALIDATION", "Enums statiques — valeurs autorisees pour tous les workbooks", 8)
    en_tetes(ws, 4, ["STATUS", "RISK_LEVEL", "GATE_TYPE", "LANG", "DECISION", "TERM_TYPE", "WORDING_TYPE", "ROLE"],
             [14, 14, 14, 10, 12, 16, 22, 22])

    rows = max(len(STATUS), 4, 5, len(LANG), 4, 3, 3, 10)
    data = []
    roles = ["Brand Director", "PR Director", "Compliance", "ESG Lead", "Legal",
             "Heritage Director", "Atelier Director", "CEO", "Editorial Committee", "CMO"]
    wording = ["ABSOLUTE", "QUALIFIED", "COMPARATIVE"]
    term_types = ["PREFERRED", "FORBIDDEN", "REVIEW"]
    risk = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
    gate = ["BRAND", "CLAIM", "HERITAGE", "LEGAL", "EVENT"]
    decision = ["APPROVE", "REWORK", "REJECT", "PENDING"]
    for i in range(rows):
        data.append([
            STATUS[i] if i < len(STATUS) else "",
            risk[i] if i < len(risk) else "",
            gate[i] if i < len(gate) else "",
            LANG[i] if i < len(LANG) else "",
            decision[i] if i < len(decision) else "",
            term_types[i] if i < len(term_types) else "",
            wording[i] if i < len(wording) else "",
            roles[i] if i < len(roles) else "",
        ])
    ecrire_donnees(ws, 5, data)

    # Section gates / approval policy
    section(ws, 20, 8, "Matrice d'approbation (Gates)", BORDEAUX)
    en_tetes(ws, 21, ["GATE_ID", "GATE_TYPE", "BLOCKING", "APPROVER_ROLE", "SLA_HOURS", "ESCALATE_TO", "CRISIS_FAST_TRACK", "NOTE"],
             [12, 14, 12, 22, 12, 22, 18, 40])
    gates = [
        ("G-001", "BRAND",    "YES", "Brand Director",        24, "CMO",          "4h",  "Gate par defaut toute sortie externe"),
        ("G-002", "CLAIM",    "YES", "ESG Lead + Legal",      48, "Legal / CMO",  "8h",  "Claim RSE / absolu / origine"),
        ("G-003", "HERITAGE", "YES", "Heritage Director",     72, "CEO",          "24h", "Reference patrimoine non sourcee"),
        ("G-004", "LEGAL",    "YES", "Legal",                 24, "CEO",          "4h",  "Contenu juridique / revendication regulee"),
        ("G-005", "EVENT",    "NO",  "PR Director",           72, "CMO",          "12h", "Pack evenementiel sans claim"),
    ]
    ecrire_donnees(ws, 22, gates, align=AL)
    ajouter_table(ws, "tblGatesPolicy", f"A21:H{21 + len(gates)}")


def onglet_changelog(wb):
    ws = wb.create_sheet("10_CHANGELOG")
    titre(ws, "CHANGELOG", "Journal des modifications — obligatoire pour tracabilite", 6)
    en_tetes(ws, 4, ["VERSION", "DATE", "AUTEUR", "TYPE", "SCOPE", "DESCRIPTION"],
             [14, 14, 20, 14, 24, 80])
    rows = [
        ("v2026.04",    D(today_iso()), "Ludo",                   "CREATE", "Workbook entier",           "Creation initiale FOUNDATIONS v1"),
        ("v2026.04.01", D(today_iso()), "Brand Office",           "SEED",   "3_BRAND_VOCAB_FR",          "Vocab FR seed 25 termes"),
        ("v2026.04.02", D(today_iso()), "ESG Lead",               "SEED",   "7_CLAIMS_EVIDENCE_REGISTRY","Seed 17 claims avec evidence"),
        ("v2026.04.03", D(today_iso()), "AI Ops",                 "FIX",    "Dates, STATUT, DV",         "Sprint correctif audit (R-01 a R-08)"),
    ]
    last = ecrire_donnees(ws, 5, rows, align=AL)
    convertir_dates(ws, 5, last, date_cols=[2])
    ajouter_table(ws, "tblChangelog", f"A4:F{last}")
    attach_dv(ws, f"D5:D{last}", ["CREATE", "UPDATE", "DELETE", "SEED", "FIX", "REVIEW"])


def onglet_dashboard(wb):
    ws = wb.create_sheet("11_DASHBOARD")
    titre(ws, "FOUNDATIONS — DASHBOARD", "Indicateurs de sante du referentiel", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18

    section(ws, 4, 6, "Sante des sources", NOIR)
    kpis = [
        ("", "Sources patrimoniales actives",        '=COUNTIFS(tblHeritageSources[STATUT],"ACTIVE")'),
        ("", "Sources patrimoniales stale / rejete", '=COUNTIFS(tblHeritageSources[STATUT],"STALE")+COUNTIFS(tblHeritageSources[STATUT],"REJECTED")'),
        ("", "Claims VALID",                         '=COUNTIFS(tblClaimsRegistry[STATUS],"VALID")'),
        ("", "Claims UNVERIFIED / MISSING",          '=COUNTIFS(tblClaimsRegistry[STATUS],"UNVERIFIED")+COUNTIFS(tblClaimsRegistry[STATUS],"MISSING")'),
        ("", "Termes FR FORBIDDEN",                  '=COUNTIFS(tblVocabFR[TERM_TYPE],"FORBIDDEN")'),
        ("", "Termes FR PREFERRED",                  '=COUNTIFS(tblVocabFR[TERM_TYPE],"PREFERRED")'),
        ("", "Medias P1 (priorite haute)",           '=COUNTIFS(tblMediaDirectory[PRIORITE],"P1")'),
        ("", "Evenements HIGH sensibilite",          '=COUNTIFS(tblEventsCalendar[VIP_LEVEL],"HIGH")'),
        ("", "Regles brand CRITICAL",                '=COUNTIFS(tblBrandRulesFound[NIVEAU],"CRITICAL")'),
    ]
    for i, (_, lbl, formula) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = "0"

    # Graphique termes vocab
    section(ws, 16, 6, "Repartition vocabulaire FR", BORDEAUX)
    lbls_vocab = [("PREFERRED", 'PREFERRED'), ("FORBIDDEN", 'FORBIDDEN'), ("REVIEW", 'REVIEW')]
    for i, (lbl, val) in enumerate(lbls_vocab):
        ws.cell(row=17 + i, column=2, value=lbl).border = BF
        ws.cell(row=17 + i, column=2).font = Font(name="Calibri", size=10, bold=True)
        c = ws.cell(row=17 + i, column=3, value=f'=COUNTIFS(tblVocabFR[TERM_TYPE],"{val}")')
        c.border = BF
        c.alignment = AC
        c.number_format = "0"

    ch = PieChart()
    data = Reference(ws, min_col=3, min_row=17, max_row=19)
    labels = Reference(ws, min_col=2, min_row=17, max_row=19)
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(labels)
    ch.title = "Vocab FR — Repartition"
    ch.height = 9
    ch.width = 14
    ws.add_chart(ch, "E4")


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES WORKBOOK", "Parametres — tblParamsFound consommee via XLOOKUP par les agents", 3)
    # R-05 : wrap en Excel Table pour XLOOKUP stable
    rows = [
        ("WORKBOOK", "FOUNDATIONS", "Type de workbook"),
        ("MAISON_ID", "M-001", "Idem 1_MAISON_PROFILE"),
        ("MAISON_NAME", "Maison pilote", ""),
        ("LANG_PRIMARY", "FR", ""),
        ("LANG_SECONDARY", "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", ""),
        ("DATA_VERSION", "v2026.04", ""),
        ("PROMPT_VERSION", "v1.0.0", ""),
        ("LAST_REFRESH", D(today_iso()), "Date du dernier refresh"),
    ]
    params_table(ws, "tblParamsFound", start_row=4, rows=rows)


def build():
    wb = Workbook()
    # remove default
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_maison_profile(wb)
    onglet_brand_charter(wb)
    onglet_vocab_fr(wb)
    onglet_vocab_en(wb)
    onglet_heritage(wb)
    onglet_media(wb)
    onglet_claims(wb)
    onglet_events(wb)
    onglet_lists(wb)
    onglet_changelog(wb)
    onglet_dashboard(wb)

    # Order sheets
    order = ["0_README", "1_PARAMS", "1_MAISON_PROFILE", "2_BRAND_CHARTER",
             "3_BRAND_VOCAB_FR", "4_BRAND_VOCAB_EN", "5_HERITAGE_SOURCEBOOK",
             "6_MEDIA_DIRECTORY", "7_CLAIMS_EVIDENCE_REGISTRY", "8_EVENTS_CALENDAR",
             "9_LISTS_REFERENTIALS", "10_CHANGELOG", "11_DASHBOARD"]
    wb._sheets = [wb[n] for n in order]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
