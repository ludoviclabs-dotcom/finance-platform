"""
Build Aero_Marketing_OVERVIEW_NEURAL.xlsx — synthèse cross-agents
de la branche NEURAL Aéronautique / Marketing.

7 onglets : README, Agents Synthèse, Pipeline Orchestration,
Réglementation Cross, KPIs Consolidés, Roadmap, Pricing & ROI.

Cible recruteur : Airbus, Safran, Thales, Dassault, Naval Group,
Daher, Liebherr Aerospace, ainsi que équipementiers et MRO.

Date de référence : 25/04/2026.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Couleurs par agent (alignées avec les Excel agents)
AGENT_COLORS = {
    "AeroTechContent": "1E40AF",       # bleu marine
    "DefenseCommsGuard": "B91C1C",     # rouge
    "AeroEventAI": "7C3AED",           # violet
    "AeroSustainabilityComms": "059669",  # vert
}


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=NEURAL_VIOLET, end_color=NEURAL_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def agent_fill(agent: str) -> PatternFill:
    c = AGENT_COLORS.get(agent, NEURAL_VIOLET)
    return PatternFill("solid", start_color=c, end_color=c)


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ────────────────────────────────────────────────────────────────────
# Données : 4 agents
# ────────────────────────────────────────────────────────────────────
AGENTS = [
    {
        "name": "AeroTechContent",
        "ordre": 1,
        "role": "Audit qualité contenus aéro B2B technique avant diffusion",
        "input": "White paper, fiche technique, RFP/RFI, brochure salon",
        "output": "Verdict OK/WARN/KO sur 10 règles + redlines + score qualité + décision diffusion",
        "kpi_principal": "Score qualité (% OK + 0.5×WARN) / 10 — cible > 80% pour DIFFUSION OK",
        "kpi_secondaire": "Nb chiffres techniques sourcés vs total — cible 100%",
        "trigger": "Pré-diffusion via webhook CMS / DAM / outil RFP",
        "duree": "30s par contenu — 5 min par dossier RFP complet",
        "confidence_v1": 89,
        "humans": "Bureau d'études signataire + DirCom + DPO si données prospects",
        "regle_cle": "AI Act art. 50 + Code défense L.2335-1 + Green Claims Directive",
    },
    {
        "name": "DefenseCommsGuard",
        "ordre": 2,
        "role": "Audit export-control + sanctions des communications marketing",
        "input": "Toute communication B2B (email, presse, social, brochure, webinaire)",
        "output": "Verdict 12 règles ITAR/EAR/EU/FR/sanctions + redlines + décision diffusion",
        "kpi_principal": "Nb KO bloquants — cible 0",
        "kpi_secondaire": "Score conformité — cible > 85% (sujet sensible)",
        "trigger": "Pré-diffusion automatique + scan inbox commercial (MailGuard)",
        "duree": "10s par communication — 2 min par campagne complète",
        "confidence_v1": 92,
        "humans": "Compliance officer export-control habilité + juriste défense + DDTC/BIS pour licence",
        "regle_cle": "ITAR US 22 CFR 120-130 + EAR 15 CFR 730-774 + Code défense L.2335-1 + OFAC SDN",
    },
    {
        "name": "AeroEventAI",
        "ordre": 3,
        "role": "Génération + audit packs événementiels salons aéro/défense 2026",
        "input": "Brief commercial + salon cible (Farnborough, ILA, Eurosatory, MEBAA)",
        "output": "Pack complet : brief presse + 5 posts social + talking points VIP + plan diffusion",
        "kpi_principal": "Score adaptation pack (tonalité ASD + AI Act + multi-langue) — cible > 85%",
        "kpi_secondaire": "Nb redlines par pack — cible < 5",
        "trigger": "Workflow événementiel — brief → 4 variantes salon en parallèle",
        "duree": "2 min par pack — 8 min pour 4 salons en parallèle",
        "confidence_v1": 87,
        "humans": "DirCom + responsable événementiel + traducteur professionnel pour multi-langue",
        "regle_cle": "ASD Charter mars 2025 + AI Act art. 50 + ITAR/EAR awareness",
    },
    {
        "name": "AeroSustainabilityComms",
        "ordre": 4,
        "role": "Audit anti-greenwashing claims SAF/H2/électrique/eVTOL/compensation",
        "input": "Campagne, communiqué, page web, brochure, fiche avec claims environnementaux",
        "output": "Verdict 11 règles greenwashing + redlines + scoring risque DGCCRF/CMA/EASA",
        "kpi_principal": "Score conformité Green Claims — cible > 80%",
        "kpi_secondaire": "Nb claims sans LCA évidence — cible 0",
        "trigger": "Pré-publication CMS + revue annuelle rapport CSRD ESRS E1",
        "duree": "1 min par claim — 10 min par rapport ESG complet",
        "confidence_v1": 86,
        "humans": "Responsable ESG signataire CSRD + bureau d'études LCA tiers + juriste consommation",
        "regle_cle": "EU Green Claims Directive 2024 + ReFuelEU Aviation 2023/2405 + CSRD ESRS E1 + EASA Decision 2024/015",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Pipeline orchestration
# ────────────────────────────────────────────────────────────────────
PIPELINE_FLOWS = [
    {
        "scenario": "Lancement white paper + campagne digitale propulsion hybride",
        "etape": 1,
        "agent": "AeroTechContent",
        "action": "Audit qualité du white paper : sources actives, chiffres validés, mention DGA si applicable, AI disclosure",
        "input_source": "Bureau d'études + équipe contenu",
        "output_dest": "Marketing digital + responsable contenu",
    },
    {
        "scenario": "Lancement white paper + campagne digitale propulsion hybride",
        "etape": 2,
        "agent": "AeroSustainabilityComms",
        "action": "Audit greenwashing des claims environnementaux : SAF certification, périmètre LCA, cohérence CSRD ESRS E1",
        "input_source": "White paper validé par AeroTechContent",
        "output_dest": "Responsable ESG + DGCCRF si campagne grand public",
    },
    {
        "scenario": "Lancement white paper + campagne digitale propulsion hybride",
        "etape": 3,
        "agent": "DefenseCommsGuard",
        "action": "Vérification finale : pas d'info dual-use, pas de programme défense classifié",
        "input_source": "Sortie AeroTechContent + AeroSustainabilityComms",
        "output_dest": "Compliance officer — Go / No-Go diffusion",
    },
    {
        "scenario": "Pack salon Farnborough 2026 multi-format",
        "etape": 1,
        "agent": "AeroEventAI",
        "action": "Génération du pack événementiel : brief presse + 5 posts + talking points + plan diffusion multi-langue (EN/FR/DE)",
        "input_source": "Brief commercial + thématique salon",
        "output_dest": "Équipe événementiel + DirCom",
    },
    {
        "scenario": "Pack salon Farnborough 2026 multi-format",
        "etape": 2,
        "agent": "DefenseCommsGuard",
        "action": "Audit export-control de tous les contenus du pack avant impression brochures + envoi presse",
        "input_source": "Sortie AeroEventAI",
        "output_dest": "Compliance officer + juriste export",
    },
    {
        "scenario": "Pack salon Farnborough 2026 multi-format",
        "etape": 3,
        "agent": "AeroSustainabilityComms",
        "action": "Audit greenwashing si claims SAF/H2 dans communications salon — cohérence avec CSRD",
        "input_source": "Brief presse + brochures",
        "output_dest": "Responsable ESG + DirCom",
    },
    {
        "scenario": "Pack salon Farnborough 2026 multi-format",
        "etape": 4,
        "agent": "AeroTechContent",
        "action": "Vérification finale chiffres techniques fiches produits + RFP-ready packaging",
        "input_source": "Tous outputs précédents",
        "output_dest": "DirCom + équipe presse",
    },
    {
        "scenario": "Réponse RFP DGA / RFI NATO programme drone",
        "etape": 1,
        "agent": "DefenseCommsGuard",
        "action": "Pré-screening export-control : ITAR Cat. VIII/XV, EAR ECCN 9A610/9A619, EU dual-use Reg. 2021/821, FR LMG L.2335-1, OFAC + EU sanctions",
        "input_source": "Brief réponse RFP + spécifications matériel",
        "output_dest": "Compliance officer — Go / No-Go avant rédaction",
    },
    {
        "scenario": "Réponse RFP DGA / RFI NATO programme drone",
        "etape": 2,
        "agent": "AeroTechContent",
        "action": "Audit qualité de la réponse RFP : sources actives, chiffres validés DGA-MOA / NCIA, mentions EDF/EDIP/EDIRPA, AI disclosure",
        "input_source": "Brief Go from DefenseCommsGuard + draft technique",
        "output_dest": "Bureau d'études signataire + DirCom",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Régulations cross-agents (12 régulations majeures)
# ────────────────────────────────────────────────────────────────────
REGULATIONS = [
    {
        "regulation": "ITAR — International Traffic in Arms Regulations",
        "ref": "US 22 CFR 120-130 (revision janv 2026)",
        "AeroTechContent": "Indirect — flag export-aware",
        "DefenseCommsGuard": "Cœur — Cat. VIII (aircraft), Cat. XV (spacecraft)",
        "AeroEventAI": "Indirect — pas d'info ITAR sur salons publics",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "EAR — Export Administration Regulations",
        "ref": "US 15 CFR 730-774 (Entity List mars 2026 + FDPR drones 2025)",
        "AeroTechContent": "Indirect",
        "DefenseCommsGuard": "Cœur — ECCN 9A610/9A619, FDPR",
        "AeroEventAI": "Indirect",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "EU Dual-Use Reg. 2021/821",
        "ref": "Annexe I (radars, optronique, propulsion, cyber)",
        "AeroTechContent": "Indirect — flag dual-use",
        "DefenseCommsGuard": "Cœur — annexe I systématique",
        "AeroEventAI": "Indirect",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "Code de la défense FR L.2335-1 LMG",
        "ref": "Arrêté LMG 27/06/2012 + LPM 2024-2030 (Loi 2023-703)",
        "AeroTechContent": "Indirect",
        "DefenseCommsGuard": "Cœur — AT/IDP/AGI",
        "AeroEventAI": "Indirect — flag programmes défense FR",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "Sanctions OFAC SDN List + EU sanctions packages",
        "ref": "OFAC SDN (mars 2026) + EU 16e paquet Russie (fév 2026)",
        "AeroTechContent": "Indirect",
        "DefenseCommsGuard": "Cœur — screening systématique",
        "AeroEventAI": "Indirect — invités webinaires/salons",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "EU AI Act — UE 2024/1689 art. 50",
        "ref": "Applicable 02/08/2026",
        "AeroTechContent": "Cœur — disclosure IA contenus générés",
        "DefenseCommsGuard": "Cœur — disclosure scoring IA screening",
        "AeroEventAI": "Cœur — disclosure IA briefs/posts/talking points",
        "AeroSustainabilityComms": "Cœur — disclosure IA audit greenwashing",
        "criticite": "BLOQUANT (août 2026)",
    },
    {
        "regulation": "EU Green Claims Directive 2024",
        "ref": "Transposition 2026 + ReFuelEU 2023/2405",
        "AeroTechContent": "Indirect — claims SAF/H2 dans contenus tech",
        "DefenseCommsGuard": "Non applicable",
        "AeroEventAI": "Indirect — claims salon",
        "AeroSustainabilityComms": "Cœur — audit systématique",
        "criticite": "BLOQUANT (DGCCRF FR)",
    },
    {
        "regulation": "CSRD ESRS E1 Climate change",
        "ref": "Directive UE 2022/2464 (obligatoire 2025 grandes entreprises)",
        "AeroTechContent": "Indirect — citation CSRD si applicable",
        "DefenseCommsGuard": "Non applicable",
        "AeroEventAI": "Indirect — communication ESG salon",
        "AeroSustainabilityComms": "Cœur — cohérence claim ↔ rapport CSRD",
        "criticite": "BLOQUANT",
    },
    {
        "regulation": "ASD Europe Responsible Defence Comms Charter",
        "ref": "Charte mars 2025 (post-Ukraine)",
        "AeroTechContent": "Cœur — tonalité contenus défense",
        "DefenseCommsGuard": "Cœur — règle D12",
        "AeroEventAI": "Cœur — règle E01",
        "AeroSustainabilityComms": "Faible",
        "criticite": "Volontaire (mais critique réputationnel)",
    },
    {
        "regulation": "NIS2 + DORA (cyber + résilience)",
        "ref": "Décret FR 2024-1308 + DORA Reg. UE 2022/2554 (applicable 17/01/2025)",
        "AeroTechContent": "Cœur — règle R06 mention cyber",
        "DefenseCommsGuard": "Indirect",
        "AeroEventAI": "Faible",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "BLOQUANT systèmes critiques",
    },
    {
        "regulation": "EDIP + EDF + EDIRPA + ASAP (financement EU défense)",
        "ref": "Reg. UE 2025/588 EDIP + EU Defence Industrial Strategy",
        "AeroTechContent": "Cœur — règle R09 mention financement",
        "DefenseCommsGuard": "Indirect",
        "AeroEventAI": "Cœur — communication programmes EU",
        "AeroSustainabilityComms": "Non applicable",
        "criticite": "Conditionnel (si subv. EU)",
    },
    {
        "regulation": "EASA Decision 2024/015 (claims environnementaux aviation)",
        "ref": "Encadrement « carbon neutral flight » et claims ESG",
        "AeroTechContent": "Indirect",
        "DefenseCommsGuard": "Non applicable",
        "AeroEventAI": "Indirect",
        "AeroSustainabilityComms": "Cœur — règle G03",
        "criticite": "BLOQUANT (EASA)",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : KPIs consolidés
# ────────────────────────────────────────────────────────────────────
KPIS = [
    {"id": "K01", "kpi": "Score conformité moyen (4 agents)", "unit": "%", "v1_actuel": 78, "v1_cible": 85, "v2_cible": 92},
    {"id": "K02", "kpi": "Nb KO bloquants par campagne", "unit": "nb", "v1_actuel": 3, "v1_cible": 0, "v2_cible": 0},
    {"id": "K03", "kpi": "% communications validées 1er passage", "unit": "%", "v1_actuel": 35, "v1_cible": 65, "v2_cible": 85},
    {"id": "K04", "kpi": "Temps audit moyen / communication", "unit": "secondes", "v1_actuel": 90, "v1_cible": 30, "v2_cible": 15},
    {"id": "K05", "kpi": "Nb redlines générées / audit (moyenne)", "unit": "nb", "v1_actuel": 5.2, "v1_cible": 4, "v2_cible": 3},
    {"id": "K06", "kpi": "% disclosure IA AI Act art. 50 conforme", "unit": "%", "v1_actuel": 8, "v1_cible": 100, "v2_cible": 100},
    {"id": "K07", "kpi": "Nb sanctions screening / mois", "unit": "nb", "v1_actuel": 0, "v1_cible": 1500, "v2_cible": 5000},
    {"id": "K08", "kpi": "% claims greenwashing détectés", "unit": "%", "v1_actuel": 0, "v1_cible": 95, "v2_cible": 99},
    {"id": "K09", "kpi": "ROI client (économie heures compliance)", "unit": "K€/mois", "v1_actuel": 0, "v1_cible": 18, "v2_cible": 32},
    {"id": "K10", "kpi": "Confiance verdict (validation HITL)", "unit": "%", "v1_actuel": 88, "v1_cible": 92, "v2_cible": 96},
    {"id": "K11", "kpi": "Couverture salons annuelle (packs)", "unit": "salons", "v1_actuel": 4, "v1_cible": 8, "v2_cible": 15},
    {"id": "K12", "kpi": "Time-to-pitch (idée → diffusion validée)", "unit": "jours", "v1_actuel": 14, "v1_cible": 5, "v2_cible": 2},
]

# ────────────────────────────────────────────────────────────────────
# Données : Roadmap
# ────────────────────────────────────────────────────────────────────
ROADMAP = [
    {"phase": "Phase 1 — POC interne", "horizon": "Q1 2026 (en cours)", "scope": "5 Excel auditables + démo recruteur", "livrable": "AeroTechContent_NEURAL.xlsx + DefenseCommsGuard_NEURAL.xlsx + AeroEventAI_NEURAL.xlsx + AeroSustainabilityComms_NEURAL.xlsx + Aero_Marketing_OVERVIEW_NEURAL.xlsx", "statut": "EN COURS"},
    {"phase": "Phase 2 — Portage UI", "horizon": "Q2 2026", "scope": "Page Next.js /secteurs/aeronautique/marketing + catalogue TS + page hub /secteurs/aeronautique", "livrable": "Site NEURAL avec branche aéro/marketing déployée Vercel", "statut": "PRÉVU"},
    {"phase": "Phase 3 — Pilote client", "horizon": "Q3 2026", "scope": "1-2 clients équipementiers / défense FR — POC 3 mois sur 1 agent", "livrable": "Démonstration ROI + retours clients + ajustements v1.5", "statut": "PRÉVU"},
    {"phase": "Phase 4 — v2 production", "horizon": "Q4 2026 - Q1 2027", "scope": "API connectées (OFAC, EU sanctions, ICVCM, ISCC EU) + dashboard SaaS", "livrable": "Plateforme SaaS avec connecteurs runtime + multi-tenant", "statut": "DESIGN"},
    {"phase": "Phase 5 — Couverture salons 2027", "horizon": "Q2 2027 (Bourget)", "scope": "Pack événementiel pré-Bourget Paris 2027 (juin) + Farnborough non bisannuel UK", "livrable": "AeroEventAI v2 avec génération automatique 8 salons", "statut": "DESIGN"},
    {"phase": "Phase 6 — International", "horizon": "Q4 2027", "scope": "Extension US (DDTC OFAC), UK (OFSI), DE (BAFA), AU (AUKUS)", "livrable": "Couverture export-control multi-juridictions", "statut": "ROADMAP"},
]

# ────────────────────────────────────────────────────────────────────
# Données : Pricing & ROI (5 personas)
# ────────────────────────────────────────────────────────────────────
PRICING = [
    {"persona": "Équipementier civil ETI", "exemple_cible": "Daher, Liebherr Aerospace, Latécoère", "agents": "AeroTechContent + AeroSustainabilityComms", "forfait_mois": 4500, "scope": "Audit white papers + claims SAF/électrique. ~50 contenus/mois. 2 salons / an.", "roi_estime": "Économie 14 K€/mois (heures compliance + marketing) — ROI x 3.1"},
    {"persona": "Équipementier défense", "exemple_cible": "Sous-traitants Thales, Safran Defence, MBDA", "agents": "DefenseCommsGuard + AeroTechContent", "forfait_mois": 8500, "scope": "Audit export-control 100% communications. Réponses RFP DGA/NATO. ~200 comms/mois.", "roi_estime": "Économie 26 K€/mois (compliance officer + risque sanction évité) — ROI x 3.1"},
    {"persona": "Compagnie eVTOL / régional électrique", "exemple_cible": "Joby Europe, Archer, Volocopter (si redressement OK)", "agents": "AeroSustainabilityComms + AeroEventAI", "forfait_mois": 5200, "scope": "Audit greenwashing + packs salons (CES, ILA, MEBAA). Communications grand public sensibles DGCCRF.", "roi_estime": "Économie 16 K€/mois + évitement sanction DGCCRF (4% CA) — ROI x 3.1"},
    {"persona": "Grand groupe aéro (Top 5)", "exemple_cible": "Airbus, Safran, Thales, Dassault, Naval Group", "agents": "Suite complète 4 agents + AeroRegWatch_Marketing + AeroEvidenceGuard_Marketing", "forfait_mois": 22000, "scope": "Couverture totale comms marketing + défense + ESG + événementiel. ~2000 comms/mois. 8 salons/an.", "roi_estime": "Économie 75 K€/mois (équipes compliance/marketing/ESG) + sécurité juridique — ROI x 3.4"},
    {"persona": "Cabinet conseil aéro", "exemple_cible": "Roland Berger Aviation, Oliver Wyman, Bain Aerospace", "agents": "API access + white-label", "forfait_mois": 3000, "scope": "Accès API pour audits clients. Volume 100 audits/mois inclus. Marque blanche.", "roi_estime": "Marge nette 60% sur revente client — ROI x 5+"},
]


# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("0_README")
    set_widths(ws, {1: 26, 2: 92})
    write_title(ws, "Aero / Marketing — OVERVIEW NEURAL — Synthèse cross-agents", span=2)
    write_subtitle(ws, 2, "NEURAL — Branche Aéronautique / Marketing — Date de réf. : 25/04/2026", span=2)

    rows = [
        ("Périmètre", "Branche Aéronautique / Marketing : 4 agents publics (AeroTechContent, DefenseCommsGuard, AeroEventAI, AeroSustainabilityComms) + 2 services réservés (AeroRegWatch_Marketing, AeroEvidenceGuard_Marketing)."),
        ("Distinction", "Cette branche couvre le marketing B2B aéro (white papers, RFP, salons, communications environnementales). DISTINCTE de la branche aero-comms (Communications & Affaires publiques corporate) déjà entamée mais non publiée — à porter ultérieurement."),
        ("Pitch", "Audit déterministe + redlines + score conformité auto. Aucun verdict KO sans citation de base légale. AI Act art. 50 disclosure intégrée dans les 4 agents (applicable août 2026). Validation humaine obligatoire — aucune autopublication."),
        ("Cibles recruteur", "Airbus, Safran, Thales, Dassault, Naval Group, Daher, Liebherr Aerospace, MBDA, Latécoère, équipementiers EU défense + cabinets conseil aéro."),
        ("Onglets", "1_Agents_Synthese (4 agents) | 2_Pipeline_Orchestration (2 scénarios × 4-3 étapes) | 3_Reglementation_Cross (12 régs) | 4_KPIs_Consolides (12 KPI) | 5_Roadmap (6 phases) | 6_Pricing_ROI (5 personas)."),
        ("Workbooks branche", "AeroTechContent_NEURAL.xlsx + DefenseCommsGuard_NEURAL.xlsx + AeroEventAI_NEURAL.xlsx + AeroSustainabilityComms_NEURAL.xlsx + Aero_Marketing_OVERVIEW_NEURAL.xlsx (ce fichier)."),
        ("Statut", "Phase 1 livrée. Démo recruteur : cas synthétiques, aucun client réel, aucun programme classifié."),
        ("Disclaimer", "Outil indicatif. Ne remplace pas DDTC, BIS, DGA, DGAC, OFAC, DGCCRF, EASA, juriste défense, compliance officer export-control, responsable ESG signataire CSRD ni bureau d'études signataire. Toute diffusion finale doit être validée humainement."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 75 + 1))
        row += 1


def build_agents_synthese(wb: Workbook) -> None:
    ws = wb.create_sheet("1_Agents_Synthese")
    set_widths(ws, {1: 22, 2: 6, 3: 38, 4: 26, 5: 36, 6: 30, 7: 30, 8: 20, 9: 12, 10: 30, 11: 36})
    write_title(ws, "Synthèse 4 agents — branche Aéro/Marketing", span=11)
    write_subtitle(ws, 2, "Mission, input, output, KPIs principaux, déclencheurs, durée traitement, validation humaine, règle clé.", span=11)

    headers = ["Agent", "#", "Rôle", "Input", "Output", "KPI principal", "KPI secondaire", "Trigger", "Confiance v1 %", "Validation humaine", "Règle clé"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 11)

    row = 5
    for i, a in enumerate(AGENTS):
        cells = [
            a["name"], a["ordre"], a["role"], a["input"], a["output"],
            a["kpi_principal"], a["kpi_secondaire"], a["trigger"], a["confidence_v1"],
            a["humans"], a["regle_cle"],
        ]
        for col, v in enumerate(cells, start=1):
            ws.cell(row=row, column=col, value=v)
        # Couleur agent dans col 1
        c = ws.cell(row=row, column=1)
        c.fill = agent_fill(a["name"])
        c.font = body_font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = THIN_BORDER
        for col in range(2, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = alt_fill()
        ws.row_dimensions[row].height = 110
        row += 1


def build_pipeline(wb: Workbook) -> None:
    ws = wb.create_sheet("2_Pipeline_Orchestration")
    set_widths(ws, {1: 50, 2: 8, 3: 24, 4: 60, 5: 32, 6: 32})
    write_title(ws, "Pipeline orchestration — 3 scénarios cross-agents", span=6)
    write_subtitle(ws, 2, "Enchaînements typiques : white paper, pack salon, RFP DGA/NATO. Cohérence cross-agents.", span=6)

    headers = ["Scénario", "Étape", "Agent", "Action", "Input source", "Output destination"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 6)

    row = 5
    prev_scenario = None
    for i, p in enumerate(PIPELINE_FLOWS):
        ws.cell(row=row, column=1, value=p["scenario"] if p["scenario"] != prev_scenario else "")
        ws.cell(row=row, column=2, value=p["etape"])
        ws.cell(row=row, column=3, value=p["agent"])
        ws.cell(row=row, column=4, value=p["action"])
        ws.cell(row=row, column=5, value=p["input_source"])
        ws.cell(row=row, column=6, value=p["output_dest"])

        # Couleur agent dans col 3
        c3 = ws.cell(row=row, column=3)
        c3.fill = agent_fill(p["agent"])
        c3.font = body_font(bold=True, color="FFFFFF")
        c3.alignment = Alignment(vertical="center", wrap_text=True)
        c3.border = THIN_BORDER

        for col in [1, 2, 4, 5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.font = body_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        ws.row_dimensions[row].height = 70
        prev_scenario = p["scenario"]
        row += 1


def build_regulation(wb: Workbook) -> None:
    ws = wb.create_sheet("3_Reglementation_Cross")
    set_widths(ws, {1: 40, 2: 36, 3: 22, 4: 22, 5: 22, 6: 22, 7: 22})
    write_title(ws, "Régulations cross-agents — 12 régulations majeures", span=7)
    write_subtitle(ws, 2, "ITAR + EAR + EU dual-use + FR LMG + sanctions + AI Act + Green Claims + CSRD + ASD + NIS2 + EDIP + EASA Decision 2024/015.", span=7)

    headers = ["Régulation", "Référence", "AeroTechContent", "DefenseCommsGuard", "AeroEventAI", "AeroSustainabilityComms", "Criticité"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 7)

    row = 5
    for i, r in enumerate(REGULATIONS):
        ws.cell(row=row, column=1, value=r["regulation"])
        ws.cell(row=row, column=2, value=r["ref"])
        ws.cell(row=row, column=3, value=r["AeroTechContent"])
        ws.cell(row=row, column=4, value=r["DefenseCommsGuard"])
        ws.cell(row=row, column=5, value=r["AeroEventAI"])
        ws.cell(row=row, column=6, value=r["AeroSustainabilityComms"])
        ws.cell(row=row, column=7, value=r["criticite"])

        # Crit cell color
        crit = r["criticite"]
        c7 = ws.cell(row=row, column=7)
        if "BLOQUANT" in crit:
            c7.fill = red_fill()
            c7.font = body_font(bold=True, color=RED_NO)
        elif "Conditionnel" in crit or "Volontaire" in crit:
            c7.fill = amber_fill()
            c7.font = body_font(bold=True, color=AMBER_WARN)
        c7.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c7.border = THIN_BORDER

        for col in [1, 2, 3, 4, 5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.font = body_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = alt_fill()

            # Highlight "Cœur" cells
            if cell.value == "Cœur" or (isinstance(cell.value, str) and cell.value.startswith("Cœur")):
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)

        ws.row_dimensions[row].height = 60
        row += 1


def build_kpis(wb: Workbook) -> None:
    ws = wb.create_sheet("4_KPIs_Consolides")
    set_widths(ws, {1: 8, 2: 56, 3: 14, 4: 16, 5: 16, 6: 16, 7: 14, 8: 14})
    write_title(ws, "KPIs consolidés — 12 indicateurs cross-agents", span=8)
    write_subtitle(ws, 2, "Comparaison v1 actuel / v1 cible / v2 cible. Calcul auto des gains.", span=8)

    headers = ["ID", "KPI", "Unité", "v1 actuel", "v1 cible", "v2 cible", "Gain v1→v1c", "Gain v1c→v2"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 8)

    row = 5
    for i, k in enumerate(KPIS):
        ws.cell(row=row, column=1, value=k["id"])
        ws.cell(row=row, column=2, value=k["kpi"])
        ws.cell(row=row, column=3, value=k["unit"])
        ws.cell(row=row, column=4, value=k["v1_actuel"])
        ws.cell(row=row, column=5, value=k["v1_cible"])
        ws.cell(row=row, column=6, value=k["v2_cible"])
        ws.cell(row=row, column=7, value=f"=E{row}-D{row}")
        ws.cell(row=row, column=8, value=f"=F{row}-E{row}")
        style_body_row(ws, row, 8, alt=(i % 2 == 1), height=24)
        for col in [4, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="center")
        row += 1


def build_roadmap(wb: Workbook) -> None:
    ws = wb.create_sheet("5_Roadmap")
    set_widths(ws, {1: 32, 2: 22, 3: 50, 4: 60, 5: 14})
    write_title(ws, "Roadmap branche Aéro/Marketing — 6 phases", span=5)
    write_subtitle(ws, 2, "Phase 1 livrée Q1 2026 (5 Excel). Phase 2 portage Next.js Q2. Pilote Q3. v2 production fin 2026/2027.", span=5)

    headers = ["Phase", "Horizon", "Scope", "Livrable", "Statut"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, r in enumerate(ROADMAP):
        ws.cell(row=row, column=1, value=r["phase"])
        ws.cell(row=row, column=2, value=r["horizon"])
        ws.cell(row=row, column=3, value=r["scope"])
        ws.cell(row=row, column=4, value=r["livrable"])
        c5 = ws.cell(row=row, column=5, value=r["statut"])

        if r["statut"] == "EN COURS":
            c5.fill = green_fill()
            c5.font = body_font(bold=True, color=GREEN_OK)
        elif r["statut"] == "PRÉVU":
            c5.fill = amber_fill()
            c5.font = body_font(bold=True, color=AMBER_WARN)
        else:
            c5.font = body_font()

        c5.alignment = Alignment(horizontal="center", vertical="center")
        c5.border = THIN_BORDER

        for col in [1, 2, 3, 4]:
            cell = ws.cell(row=row, column=col)
            cell.font = body_font()
            if col == 1:
                cell.font = body_font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = alt_fill()

        ws.row_dimensions[row].height = 80
        row += 1


def build_pricing(wb: Workbook) -> None:
    ws = wb.create_sheet("6_Pricing_ROI")
    set_widths(ws, {1: 30, 2: 38, 3: 38, 4: 14, 5: 50, 6: 50})
    write_title(ws, "Pricing & ROI — 5 personas cibles", span=6)
    write_subtitle(ws, 2, "Forfaits mensuels indicatifs. ROI calculé sur économie heures compliance + risque sanction évité.", span=6)

    headers = ["Persona", "Exemples cibles (publics)", "Agents inclus", "Forfait €/mois", "Scope mensuel", "ROI estimé"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 6)

    row = 5
    for i, p in enumerate(PRICING):
        ws.cell(row=row, column=1, value=p["persona"])
        ws.cell(row=row, column=2, value=p["exemple_cible"])
        ws.cell(row=row, column=3, value=p["agents"])
        c4 = ws.cell(row=row, column=4, value=p["forfait_mois"])
        c4.number_format = "#,##0 €"
        ws.cell(row=row, column=5, value=p["scope"])
        ws.cell(row=row, column=6, value=p["roi_estime"])
        style_body_row(ws, row, 6, alt=(i % 2 == 1), height=80)
        row += 1

    # Total + moyenne
    row += 1
    write_section_header(ws, row, "Synthèse", span=6)
    row += 1
    ws.cell(row=row, column=1, value="Forfait moyen")
    ws.cell(row=row, column=4, value=f"=AVERAGE(D5:D{4 + len(PRICING)})").number_format = "#,##0 €"
    style_body_row(ws, row, 6, height=22)
    ws.cell(row=row, column=1).font = body_font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total mensuel cumulé")
    ws.cell(row=row, column=4, value=f"=SUM(D5:D{4 + len(PRICING)})").number_format = "#,##0 €"
    style_body_row(ws, row, 6, height=22)
    ws.cell(row=row, column=1).font = body_font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total annuel cumulé")
    ws.cell(row=row, column=4, value=f"=SUM(D5:D{4 + len(PRICING)})*12").number_format = "#,##0 €"
    style_body_row(ws, row, 6, height=22)
    ws.cell(row=row, column=1).font = body_font(bold=True)


# ────────────────────────────────────────────────────────────────────
# Build main
# ────────────────────────────────────────────────────────────────────


def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb)
    build_agents_synthese(wb)
    build_pipeline(wb)
    build_regulation(wb)
    build_kpis(wb)
    build_roadmap(wb)
    build_pricing(wb)

    out = "Aero_Marketing_OVERVIEW_NEURAL.xlsx"
    wb.save(out)
    print(f"OK — {out} written ({len(AGENTS)} agents, {len(PIPELINE_FLOWS)} étapes pipeline, {len(REGULATIONS)} régs, {len(KPIS)} KPIs, {len(ROADMAP)} phases, {len(PRICING)} personas)")


if __name__ == "__main__":
    build()
