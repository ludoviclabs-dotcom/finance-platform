"""Utility class for reading Excel files uploaded via FastAPI."""

from __future__ import annotations

import io
import re
from typing import Any

try:
    import pandas as pd
    _PANDAS_AVAILABLE = True
except ImportError:
    pd = None  # type: ignore[assignment]
    _PANDAS_AVAILABLE = False

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReaderError(Exception):
    """Base exception for ExcelReader errors."""


class SheetNotFoundError(ExcelReaderError):
    """Raised when the requested sheet does not exist."""


class InvalidRangeError(ExcelReaderError):
    """Raised when a cell reference or range string is malformed."""


class CorruptFileError(ExcelReaderError):
    """Raised when the uploaded file cannot be parsed as Excel."""


# ---------------------------------------------------------------------------
# Regex helpers
# ---------------------------------------------------------------------------
_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$", re.IGNORECASE)
_RANGE_RE = re.compile(
    r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", re.IGNORECASE,
)


class ExcelReader:
    """High-level wrapper around an in-memory Excel workbook.

    Usage::

        reader = await ExcelReader.from_upload(upload_file)
        sheets = reader.sheet_names
        value  = reader.get_cell("Sheet1", "B4")
        df     = reader.get_range_as_dataframe("Sheet1", "A1:D10")
    """

    def __init__(self, workbook: Workbook) -> None:
        self._wb = workbook

    # ------------------------------------------------------------------
    # Factory
    # ------------------------------------------------------------------
    @classmethod
    async def from_upload(cls, file: UploadFile) -> "ExcelReader":
        """Create an ``ExcelReader`` from a FastAPI ``UploadFile``.

        Raises:
            CorruptFileError: If the file cannot be read by openpyxl.
        """
        try:
            contents = await file.read()
            wb = load_workbook(
                filename=io.BytesIO(contents),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise CorruptFileError(
                f"Unable to read the uploaded file as Excel: {exc}"
            ) from exc
        return cls(wb)

    # ------------------------------------------------------------------
    # Properties
    # ------------------------------------------------------------------
    @property
    def sheet_names(self) -> list[str]:
        """Return the ordered list of sheet (tab) names."""
        return self._wb.sheetnames

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _get_sheet(self, name: str) -> Worksheet:
        if name not in self._wb.sheetnames:
            raise SheetNotFoundError(
                f"Sheet '{name}' not found. "
                f"Available sheets: {', '.join(self._wb.sheetnames)}"
            )
        return self._wb[name]

    @staticmethod
    def _parse_cell_ref(ref: str) -> tuple[int, int]:
        """Return ``(row, col)`` 1-indexed from a cell reference like ``B4``."""
        m = _CELL_RE.match(ref.strip())
        if not m:
            raise InvalidRangeError(f"Invalid cell reference: '{ref}'")
        col = column_index_from_string(m.group(1).upper())
        row = int(m.group(2))
        return row, col

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def get_cell(self, sheet_name: str, cell_ref: str) -> Any:
        """Return the value of a single cell (e.g. ``"B4"``) in *sheet_name*.

        Raises:
            SheetNotFoundError: If the sheet does not exist.
            InvalidRangeError: If *cell_ref* is malformed.
        """
        ws = self._get_sheet(sheet_name)
        row, col = self._parse_cell_ref(cell_ref)
        return ws.cell(row=row, column=col).value

    def get_range_as_dicts(
        self,
        sheet_name: str,
        range_str: str,
    ) -> list[dict[str, Any]]:
        """Extract a rectangular range as a list of row-dicts.

        The first row of the range is used as column headers.  Each
        subsequent row becomes a ``{header: value}`` dictionary.

        Example::

            reader.get_range_as_dicts("Sheet1", "A1:D10")

        Raises:
            SheetNotFoundError: If the sheet does not exist.
            InvalidRangeError: If *range_str* is malformed.
        """
        ws = self._get_sheet(sheet_name)
        m = _RANGE_RE.match(range_str.strip())
        if not m:
            raise InvalidRangeError(f"Invalid range: '{range_str}'")

        col_start = column_index_from_string(m.group(1).upper())
        row_start = int(m.group(2))
        col_end = column_index_from_string(m.group(3).upper())
        row_end = int(m.group(4))

        rows: list[list[Any]] = []
        for r in range(row_start, row_end + 1):
            row_data = [
                ws.cell(row=r, column=c).value
                for c in range(col_start, col_end + 1)
            ]
            rows.append(row_data)

        if not rows:
            return []

        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def get_range_as_dataframe(self, sheet_name: str, range_str: str):  # type: ignore[return]
        """Same as :meth:`get_range_as_dicts` but returns a ``DataFrame``.

        Raises:
            ImportError: If pandas is not installed in the current environment.
        """
        if not _PANDAS_AVAILABLE:
            raise ImportError(
                "pandas is required for get_range_as_dataframe() but is not installed."
            )
        return pd.DataFrame(self.get_range_as_dicts(sheet_name, range_str))

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------
    def close(self) -> None:
        """Close the underlying workbook."""
        self._wb.close()
