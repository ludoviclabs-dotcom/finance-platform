"""
Build AeroTechContent_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'AeroTechContent' (branche Aéronautique — Marketing).

7 onglets : README, Référentiel, Cas Inputs, Verdict par Cas, Redlines/Outputs,
AI Act Disclosure, Limites. 5 cas réels (white paper drone, fiche radar, RFP DGA,
RFP NATO, brochure Farnborough). Formules score qualité technique calculées
dans Excel (pas de hardcode).

Date de référence des sources : 25/04/2026.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"
AGENT_BLUE = "1E40AF"  # bleu marine spécifique AeroTechContent

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=AGENT_BLUE, end_color=AGENT_BLUE)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def verdict_fill(verdict: str) -> PatternFill:
    if verdict == "OK":
        return green_fill()
    if verdict == "WARN":
        return amber_fill()
    if verdict == "KO":
        return red_fill()
    return PatternFill()


def verdict_color(verdict: str) -> str:
    if verdict == "OK":
        return GREEN_OK
    if verdict == "WARN":
        return AMBER_WARN
    if verdict == "KO":
        return RED_NO
    return "0E0824"


# ────────────────────────────────────────────────────────────────────
# Données : 10 règles (référentiel)
# ────────────────────────────────────────────────────────────────────
RULES = [
    {
        "id": "R01",
        "label": "SOURCE-FRESHNESS",
        "axis": "Veille",
        "rule": "Toute source citée doit avoir une date de publication ≤ 24 mois (sauf normes ISO/EASA pérennes).",
        "base": "Bonne pratique B2B technique — éviter contenu obsolète.",
        "blocking": "WARN",
    },
    {
        "id": "R02",
        "label": "NUM-VALIDATED",
        "axis": "Précision",
        "rule": "Chaque chiffre technique (km/h, kg, dB, GHz, autonomie, MTBF, dollars) doit être sourcé avec référence vérifiable.",
        "base": "Code consommation L.121-1 (pratiques commerciales trompeuses) + ASD Charter mars 2025.",
        "blocking": "KO",
    },
    {
        "id": "R03",
        "label": "AI-DISCLOSE",
        "axis": "AI Act",
        "rule": "Mention IA obligatoire si tout ou partie du contenu généré par IA — applicable 02/08/2026.",
        "base": "EU AI Act UE 2024/1689 art. 50, applicable 2 août 2026.",
        "blocking": "KO",
    },
    {
        "id": "R04",
        "label": "DGA-MENTION",
        "axis": "Défense FR",
        "rule": "Tout contenu lié à un programme défense français doit mentionner DGA-MOA / numéro programme si publiable.",
        "base": "Code de la défense L.2335-1 + LPM 2024-2030 (Loi 2023-703).",
        "blocking": "WARN",
    },
    {
        "id": "R05",
        "label": "NO-GREENWASH",
        "axis": "ESG",
        "rule": "Pas de claims SAF/H2/électrique vagues (« vert », « neutre carbone », « propre ») sans LCA chiffrée + fournisseur ISCC EU.",
        "base": "EU Green Claims Directive 2024 (transposition 2026) + Loi Climat & Résilience FR art. 12 + ReFuelEU Aviation Reg. 2023/2405.",
        "blocking": "KO",
    },
    {
        "id": "R06",
        "label": "CYBER-MENTION",
        "axis": "Cyber",
        "rule": "Mention NIS2 / DORA / certification CC EAL ou SecNumCloud obligatoire pour tout système critique embarqué ou sol.",
        "base": "Décret FR 2024-1308 (NIS2) + DORA Reg. UE 2022/2554 (applicable 17/01/2025).",
        "blocking": "WARN",
    },
    {
        "id": "R07",
        "label": "NO-CLASSIFIED",
        "axis": "Sécurité",
        "rule": "Aucune information classifiée Diffusion Restreinte / Secret Défense / Top Secret ne doit apparaître dans un contenu marketing.",
        "base": "SGDSN — Protection du secret de la défense nationale (Code défense L.2311-1).",
        "blocking": "KO",
    },
    {
        "id": "R08",
        "label": "EXPORT-AWARE",
        "axis": "Export Control",
        "rule": "Toute techno potentiellement ITAR (US 22 CFR 120-130) / EAR (15 CFR 730-774) / EU dual-use Reg. 2021/821 doit être flaggée pour audit DefenseCommsGuard.",
        "base": "ITAR USML revision janv 2026 + EAR FDPR drones 2025 + EU Reg. 2021/821 dual-use.",
        "blocking": "KO",
    },
    {
        "id": "R09",
        "label": "EU-FUNDING",
        "axis": "Financement EU",
        "rule": "Si projet financé par EDF / EDIRPA / EDIP / ASAP, mention obligatoire du programme et du % de cofinancement.",
        "base": "EU Defence Industrial Strategy 2024 + EDIP Reg. UE 2025/588 (mars 2025) + ASAP étendu 2026.",
        "blocking": "WARN",
    },
    {
        "id": "R10",
        "label": "LEGAL-FOOTER",
        "axis": "Mentions légales",
        "rule": "Footer obligatoire : raison sociale, SIREN, adresse postale, directeur publication, ORIAS si distribution assurance.",
        "base": "Code commerce R.123-237 + LCEN art. 6-III.",
        "blocking": "WARN",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : 5 cas démo
# ────────────────────────────────────────────────────────────────────
CASES = [
    {
        "id": "WPR-001",
        "type": "White Paper",
        "title": "Mini-drones surveillance maritime — État de l'art 2026",
        "audience": "Décideurs maritimes (douanes, garde-côtes, autorités portuaires)",
        "format": "PDF 18 pages, diffusion lead generation salon Eurosatory + Farnborough",
        "ai_used": "OUI — rédaction par LLM, schémas générés, traduction EN",
        "extract": (
            "Notre famille de mini-drones surveillance offre une autonomie révolutionnaire de 6 heures, "
            "une charge utile de 2,5 kg, et une signature radar quasi-nulle. La technologie de propulsion "
            "électrique 100% verte garantit zéro émission de CO2 sur l'ensemble du cycle de vol. "
            "Compatible avec les chaînes ISR de l'OTAN, ils s'intègrent à tout C2 standard. "
            "Source de l'autonomie : tests internes 2023."
        ),
        "verdicts": {"R01": "WARN", "R02": "KO", "R03": "KO", "R04": "OK", "R05": "KO", "R06": "WARN", "R07": "OK", "R08": "WARN", "R09": "OK", "R10": "OK"},
    },
    {
        "id": "TDS-001",
        "type": "Fiche technique",
        "title": "Radar AESA Bande X — Datasheet",
        "audience": "Acheteurs systèmes (équipementiers, intégrateurs aéro)",
        "format": "PDF 4 pages, distribution salons Farnborough + ILA Berlin + sur demande",
        "ai_used": "PARTIEL — synthèse spécifications, traduction EN",
        "extract": (
            "Radar AESA bande X — gain antenne 32 dBi, portée 350 km en mode air-air, "
            "résolution Doppler 0,5 m/s. MTBF documenté à 4 200 heures (réf. tests 2024). "
            "Conforme MIL-STD-810H. Solution embarquée certifiée DO-178C niveau B. "
            "Disponible sous licence d'export selon réglementation EAR ECCN 6A008. "
            "Compatible OTAN STANAG 4660."
        ),
        "verdicts": {"R01": "OK", "R02": "OK", "R03": "WARN", "R04": "WARN", "R05": "OK", "R06": "OK", "R07": "OK", "R08": "OK", "R09": "OK", "R10": "OK"},
    },
    {
        "id": "RFP-DGA-001",
        "type": "Réponse RFP",
        "title": "RFP DGA — Programme drone tactique terrestre TACT-2027 (section technique)",
        "audience": "Direction Générale de l'Armement, Service du Matériel Aérien",
        "format": "Document de réponse 80 pages, soumission via plateforme PLACE",
        "ai_used": "PARTIEL — synthèse benchmarks, génération de schémas",
        "extract": (
            "Notre solution répond au cahier des charges du programme TACT-2027 lancé par la DGA-MOA Aéro. "
            "L'autonomie de 4h dépasse de 33% la spécification minimale. "
            "Le système intègre un module de cybersécurité conforme aux exigences de l'ANSSI. "
            "Co-financement EDIP envisagé à hauteur de 35% (en cours d'instruction CINEA). "
            "Le radar embarqué utilise une technologie radar à compression d'impulsions optimisée."
        ),
        "verdicts": {"R01": "OK", "R02": "WARN", "R03": "WARN", "R04": "OK", "R05": "OK", "R06": "OK", "R07": "OK", "R08": "WARN", "R09": "OK", "R10": "OK"},
    },
    {
        "id": "RFP-NAT-001",
        "type": "Réponse RFI",
        "title": "RFI NATO — Système communication tactique sécurisée 2026",
        "audience": "NATO Communications and Information Agency (NCIA)",
        "format": "RFI réponse 45 pages, anglais",
        "ai_used": "OUI — rédaction sections techniques + traduction",
        "extract": (
            "Our tactical SDR solution offers AES-256 encryption with FIPS 140-3 certification. "
            "Compatible with NATO STANAG 5066 and 4691. Operational since 2022, "
            "deployed by 4 NATO members. Throughput: 2 Mbps in degraded environment. "
            "Cyber-resilience certified Common Criteria EAL4+. Eligible for NATO Defence Innovation Accelerator. "
            "Export licence US/UK obtained under AUKUS Pillar II Open General Licence."
        ),
        "verdicts": {"R01": "OK", "R02": "OK", "R03": "KO", "R04": "OK", "R05": "OK", "R06": "OK", "R07": "OK", "R08": "OK", "R09": "WARN", "R10": "WARN"},
    },
    {
        "id": "BRO-FAR-001",
        "type": "Brochure salon",
        "title": "Propulsion hybride-électrique aviation régionale — Brochure Farnborough 2026",
        "audience": "Compagnies aériennes régionales, leasing companies, MRO",
        "format": "Brochure print 6 pages — distribution stand FAR2026 + version PDF",
        "ai_used": "NON — rédaction humaine, photos studio",
        "extract": (
            "Notre solution de propulsion hybride-électrique permet un vol 100% décarboné "
            "sur 250 km. Autonomie totale 1 200 km grâce au système de generator de bord. "
            "Réduction CO2 de 70% par rapport aux appareils thermiques. Notre technologie est verte. "
            "Disponible en option SAF blend 50%. Certifications EASA en cours. "
            "Premier appareil livraison 2027 (sous réserve)."
        ),
        "verdicts": {"R01": "OK", "R02": "WARN", "R03": "OK", "R04": "OK", "R05": "KO", "R06": "OK", "R07": "OK", "R08": "OK", "R09": "OK", "R10": "OK"},
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Redlines (suggestions de correction)
# ────────────────────────────────────────────────────────────────────
REDLINES = [
    # WPR-001
    {"case": "WPR-001", "rule": "R02", "before": "autonomie révolutionnaire de 6 heures", "after": "autonomie de 5h45 (test interne, conditions standard ISA, charge 1,5 kg) — réf. doc test n°TI-2024-Q3-117", "why": "Chiffre marketing non sourcé, formulation évaluative interdite (« révolutionnaire »)"},
    {"case": "WPR-001", "rule": "R03", "before": "[Aucune mention IA]", "after": "Mention pied de page : « Ce document a été partiellement rédigé avec l'assistance d'un système d'IA générative — relecture humaine experte. (UE 2024/1689 art. 50). »", "why": "AI Act art. 50 applicable 02/08/2026 — disclosure obligatoire"},
    {"case": "WPR-001", "rule": "R05", "before": "propulsion électrique 100% verte... zéro émission de CO2 sur l'ensemble du cycle de vol", "after": "propulsion électrique batterie : 0 émission CO2 en vol. Bilan cycle de vie (LCA conforme PEF) = 18 g CO2eq/km — étude ADEME 2024 réf. AVI-LCA-024", "why": "Greenwashing : « 100% verte » et « ensemble du cycle » contredits par l'extraction batterie. Green Claims Directive interdit ces formulations sans LCA"},
    {"case": "WPR-001", "rule": "R08", "before": "Compatible avec les chaînes ISR de l'OTAN, ils s'intègrent à tout C2 standard", "after": "[FLAG DefenseCommsGuard] Vérifier si les protocoles ISR/C2 cités sont couverts par EAR ECCN 6A008 ou EU dual-use Reg. 2021/821 annexe I avant publication.", "why": "Mention de chaîne ISR OTAN nécessite revue export control"},
    {"case": "WPR-001", "rule": "R01", "before": "Source de l'autonomie : tests internes 2023", "after": "Source : test interne TI-2024-Q3-117 (octobre 2024) — sources externes : Frost & Sullivan Mini-Drone Market Report 2025", "why": "Source datée de 2023 — > 24 mois → WARN"},
    # TDS-001
    {"case": "TDS-001", "rule": "R03", "before": "[Mention IA absente sur la fiche]", "after": "Mention discrète sous le titre : « Synthèse spécifications assistée IA — données techniques validées par bureau d'études. »", "why": "AI partiel utilisé → disclosure même réduite obligatoire"},
    {"case": "TDS-001", "rule": "R04", "before": "[Pas de mention de programme]", "after": "Si applicable : ajouter « Développement supporté par programme DGA-MOA Aéro RAPID 2024 ».", "why": "Si programme DGA, mention conseillée — sinon retirer toute connotation défense FR"},
    # RFP-DGA-001
    {"case": "RFP-DGA-001", "rule": "R02", "before": "L'autonomie de 4h dépasse de 33% la spécification minimale", "after": "L'autonomie de 4h00 (test interne TI-DGA-2025-Q1) dépasse de 33% la spécification minimale de 3h00 (cahier des charges TACT-2027 v1.2 du 15/01/2026)", "why": "Spécifier la source de la spec et du test"},
    {"case": "RFP-DGA-001", "rule": "R03", "before": "[Aucune mention IA dans la réponse RFP]", "after": "Annexe RFP : « Sections benchmarks et schémas générés avec assistance IA — relecture par bureau d'études et signature ingénieur responsable. »", "why": "AI Act art. 50 — RFP DGA inclus dans le scope de disclosure"},
    {"case": "RFP-DGA-001", "rule": "R08", "before": "Le radar embarqué utilise une technologie radar à compression d'impulsions optimisée", "after": "[FLAG DefenseCommsGuard + DGA EXPORT] Vérifier classification matériel de guerre (Code défense L.2335-1) avant publication RFP final.", "why": "Radar à compression d'impulsions = matériel potentiellement classé"},
    # RFP-NAT-001
    {"case": "RFP-NAT-001", "rule": "R03", "before": "[No AI disclosure in NATO RFI]", "after": "Footer en anglais : « This document includes content generated with AI assistance under EU AI Act art. 50 (applicable 02/08/2026). Human technical review by certified engineer. »", "why": "AI Act applicable to all communications including RFI"},
    {"case": "RFP-NAT-001", "rule": "R09", "before": "Eligible for NATO Defence Innovation Accelerator", "after": "Eligible for NATO Defence Innovation Accelerator (DIANA) — submission planned Q3 2026. Co-funding under EU EDIP under instruction (Reg. UE 2025/588).", "why": "Précision EDIP + DIANA si applicable"},
    {"case": "RFP-NAT-001", "rule": "R10", "before": "[Footer minimal]", "after": "Add full legal footer: registered name, SIREN, head office, publication director, ORIAS if applicable.", "why": "LCEN art. 6-III applicable même pour documents export"},
    # BRO-FAR-001
    {"case": "BRO-FAR-001", "rule": "R02", "before": "Réduction CO2 de 70% par rapport aux appareils thermiques", "after": "Réduction CO2 estimée de 70% en mode 100% électrique sur 250 km par rapport à un ATR 72 thermique de référence — étude ICAO CORSIA Methodology v2.3 (2024). Hors phases de production batterie.", "why": "Chiffre marketing sans périmètre LCA → trompeur"},
    {"case": "BRO-FAR-001", "rule": "R05", "before": "vol 100% décarboné sur 250 km", "after": "vol à 0 émission CO2 directe sur 250 km en mode batterie (hors LCA cycle complet)", "why": "« Décarboné » sans LCA = greenwashing au sens Green Claims Directive 2024"},
    {"case": "BRO-FAR-001", "rule": "R05", "before": "Notre technologie est verte", "after": "[SUPPRIMER cette phrase]. Remplacer par : « Empreinte carbone du cycle de vie : 28 g CO2eq/passager-km (étude ADEME 2025) ».", "why": "« Verte » est un superlatif environnemental interdit par Green Claims Directive sans preuve PEF/LCA"},
    {"case": "BRO-FAR-001", "rule": "R05", "before": "Disponible en option SAF blend 50%", "after": "Disponible en option SAF blend 50% (HEFA certifié ISCC EU, fournisseur Neste — réf. ReFuelEU Aviation Reg. 2023/2405)", "why": "Mention SAF doit préciser type (HEFA/ATJ/PtL), certification fournisseur, et conformité ReFuelEU"},
]

# ────────────────────────────────────────────────────────────────────
# Données : AI Act Disclosure modèles
# ────────────────────────────────────────────────────────────────────
AI_DISCLOSURE = [
    {"format": "White Paper", "fr": "Ce document a été partiellement rédigé avec l'assistance d'un système d'intelligence artificielle générative. Toutes les données techniques ont fait l'objet d'une relecture par des experts humains. (Mention au titre du Règlement UE 2024/1689 art. 50 — applicable au 2 août 2026.)", "en": "This document has been partially drafted with the assistance of a generative AI system. All technical data has been reviewed by human experts. (Notice pursuant to Regulation (EU) 2024/1689 art. 50 — applicable August 2, 2026.)"},
    {"format": "Fiche technique", "fr": "Synthèse des spécifications assistée par IA — données techniques validées et signées par bureau d'études. (UE 2024/1689 art. 50.)", "en": "Specifications synthesis AI-assisted — technical data validated and signed by engineering office. (EU 2024/1689 art. 50.)"},
    {"format": "Réponse RFP/RFI", "fr": "Annexe disclosure : Les sections benchmarks, traduction et schémas du présent document ont été générées avec l'assistance d'un système d'IA. Tous les engagements techniques sont signés par un ingénieur responsable. (UE 2024/1689 art. 50.)", "en": "Disclosure annex: The benchmarks, translation and diagrams sections of this document were generated with AI assistance. All technical commitments are signed by a responsible engineer. (EU 2024/1689 art. 50.)"},
    {"format": "Brochure salon", "fr": "Document rédigé sans assistance IA — toutes images et chiffres validés par bureau d'études. (Conformité préventive UE 2024/1689 art. 50.)", "en": "Document drafted without AI assistance — all images and figures validated by engineering office. (Preventive compliance EU 2024/1689 art. 50.)"},
    {"format": "Post LinkedIn / X", "fr": "Post composé avec assistance IA #AIGen #UE2024-1689", "en": "Post composed with AI assistance #AIGen #EU2024-1689"},
]


# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("0_README")
    set_widths(ws, {1: 26, 2: 90})
    write_title(ws, "AeroTechContent — Audit qualité contenus aéro B2B technique", span=2)
    write_subtitle(ws, 2, "NEURAL — Branche Aéronautique / Marketing — Date de réf. : 25/04/2026", span=2)

    rows = [
        ("Mission", "Auditer les contenus marketing techniques aéro (white papers, fiches, RFP, brochures salons) avant diffusion : sources actives, chiffres validés, disclosure IA, mentions DGA, anti-greenwashing, mention cyber, flag export."),
        ("Périmètre", "Aéronautique civile + défense + spatial. B2B uniquement (pas de communication B2C, pas de relations gouvernementales — voir agent aero-comms AGA003 GovRelations)."),
        ("Référentiel", "10 règles encodées (R01 à R10) couvrant veille, précision, AI Act, défense FR, ESG, cyber, sécurité, export control, financement EU, mentions légales."),
        ("Échelle verdict", "OK = conforme | WARN = à revoir mais non bloquant | KO = bloquant, redline obligatoire avant diffusion."),
        ("Décision", "Score conformité ≥ 80% ET 0 KO → DIFFUSION OK. Sinon : RELECTURE (WARN cumulés) ou BLOQUÉ (KO ≥ 1)."),
        ("Inputs cas démo", "5 cas synthétiques inspirés du marché 2026 : white paper drone, fiche radar AESA, RFP DGA programme drone, RFI NATO communication, brochure Farnborough propulsion hybride."),
        ("Outputs", "(1) Verdict par règle × cas, (2) score conformité auto, (3) décision auto, (4) ~17 redlines concrètes, (5) modèles disclosure IA FR + EN."),
        ("Statut", "Démo recruteur — données synthétiques, pas de programme classifié, pas de client réel cité."),
        ("Disclaimer", "Outil indicatif. Ne remplace PAS un compliance officer export-control, un juriste défense, un bureau d'études signataire ou la DGA / DGAC. Toute diffusion finale doit être validée humainement."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 75 + 1))
        row += 1


def build_referentiel(wb: Workbook) -> None:
    ws = wb.create_sheet("1_Referentiel")
    set_widths(ws, {1: 8, 2: 22, 3: 16, 4: 60, 5: 50, 6: 12})
    write_title(ws, "Référentiel — 10 règles d'audit AeroTechContent", span=6)
    write_subtitle(ws, 2, "Chaque règle a un seuil bloquant. Les KO bloquent la diffusion. Les WARN imposent une relecture.", span=6)

    headers = ["ID", "Label", "Axe", "Règle", "Base légale / réf.", "Seuil"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 6)

    row = 5
    for i, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        ws.cell(row=row, column=3, value=rule["axis"])
        ws.cell(row=row, column=4, value=rule["rule"])
        ws.cell(row=row, column=5, value=rule["base"])
        c6 = ws.cell(row=row, column=6, value=rule["blocking"])
        c6.font = body_font(bold=True, color=verdict_color(rule["blocking"]))
        style_body_row(ws, row, 6, alt=(i % 2 == 1), height=70)
        row += 1


def build_cas_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("2_Cas_Inputs")
    set_widths(ws, {1: 14, 2: 18, 3: 36, 4: 30, 5: 28, 6: 18, 7: 70})
    write_title(ws, "Cas inputs — 5 contenus marketing à auditer", span=7)
    write_subtitle(ws, 2, "Cas synthétiques 2026 inspirés du marché. Aucun client réel ni programme classifié.", span=7)

    headers = ["ID", "Type", "Titre", "Audience", "Format diffusion", "IA utilisée", "Extrait"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 7)

    row = 5
    for i, c in enumerate(CASES):
        ws.cell(row=row, column=1, value=c["id"])
        ws.cell(row=row, column=2, value=c["type"])
        ws.cell(row=row, column=3, value=c["title"])
        ws.cell(row=row, column=4, value=c["audience"])
        ws.cell(row=row, column=5, value=c["format"])
        ws.cell(row=row, column=6, value=c["ai_used"])
        ws.cell(row=row, column=7, value=c["extract"])
        style_body_row(ws, row, 7, alt=(i % 2 == 1), height=130)
        row += 1


def build_verdict(wb: Workbook) -> None:
    ws = wb.create_sheet("3_Verdict_Par_Cas")
    n_cases = len(CASES)
    n_rules = len(RULES)
    cols_total = 2 + n_cases  # ID, Label, then 1 per case

    widths = {1: 8, 2: 22}
    for i in range(n_cases):
        widths[3 + i] = 16
    set_widths(ws, widths)

    write_title(ws, "Matrice verdicts — 10 règles × 5 cas", span=cols_total)
    write_subtitle(ws, 2, "OK = vert | WARN = ambre | KO = rouge. Score conformité auto + décision auto en bas.", span=cols_total)

    # Header row
    ws.cell(row=4, column=1, value="ID")
    ws.cell(row=4, column=2, value="Règle")
    for i, c in enumerate(CASES):
        ws.cell(row=4, column=3 + i, value=c["id"])
    style_header_row(ws, 4, cols_total)

    # Verdict rows
    row = 5
    for r_idx, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        for c_idx, c in enumerate(CASES):
            v = c["verdicts"].get(rule["id"], "")
            cell = ws.cell(row=row, column=3 + c_idx, value=v)
            cell.font = body_font(bold=True, color=verdict_color(v))
            cell.fill = verdict_fill(v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).font = body_font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).font = body_font()
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 26
        row += 1

    # Aggregation rows
    row += 1
    write_section_header(ws, row, "Synthèse par cas", span=cols_total)
    row += 1

    # OK count per case
    ws.cell(row=row, column=2, value="Nb OK")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"OK")')
    style_body_row(ws, row, cols_total, height=22)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    # WARN count
    ws.cell(row=row, column=2, value="Nb WARN")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"WARN")')
    style_body_row(ws, row, cols_total, height=22)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    # KO count
    ws.cell(row=row, column=2, value="Nb KO")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"KO")')
    style_body_row(ws, row, cols_total, height=22)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    # Score conformité
    ws.cell(row=row, column=2, value="Score conformité")
    score_row = row
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        # Score = (OK + 0.5 * WARN) / n_rules
        formula = f'=(COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"WARN"))/{n_rules}'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.number_format = "0%"
    style_body_row(ws, row, cols_total, height=24)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    # Décision auto
    ws.cell(row=row, column=2, value="Décision")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ko_count = f'COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"KO")'
        score = f'(COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{4 + n_rules},"WARN"))/{n_rules}'
        formula = f'=IF({ko_count}>0,"BLOQUÉ",IF({score}>=0.8,"DIFFUSION OK","RELECTURE"))'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.font = body_font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    style_body_row(ws, row, cols_total, height=26)
    ws.cell(row=row, column=2).font = body_font(bold=True)


def build_redlines(wb: Workbook) -> None:
    ws = wb.create_sheet("4_Redlines_Outputs")
    set_widths(ws, {1: 14, 2: 8, 3: 50, 4: 65, 5: 45})
    write_title(ws, "Redlines — Suggestions de correction concrètes", span=5)
    write_subtitle(ws, 2, "Pour chaque KO/WARN, une redline applicable. Conserver l'extrait original pour traçabilité.", span=5)

    headers = ["Cas", "Règle", "Avant (extrait original)", "Après (suggestion redline)", "Pourquoi"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rl in enumerate(REDLINES):
        ws.cell(row=row, column=1, value=rl["case"])
        ws.cell(row=row, column=2, value=rl["rule"])
        ws.cell(row=row, column=3, value=rl["before"])
        ws.cell(row=row, column=4, value=rl["after"])
        ws.cell(row=row, column=5, value=rl["why"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=70)
        row += 1


def build_ai_disclosure(wb: Workbook) -> None:
    ws = wb.create_sheet("5_AI_Act_Disclosure")
    set_widths(ws, {1: 22, 2: 60, 3: 60})
    write_title(ws, "AI Act art. 50 — Modèles de mention IA", span=3)
    write_subtitle(ws, 2, "Applicable 2 août 2026. Disclosure obligatoire si tout ou partie du contenu généré par IA.", span=3)

    headers = ["Format", "Mention FR", "Mention EN"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 3)

    row = 5
    for i, d in enumerate(AI_DISCLOSURE):
        ws.cell(row=row, column=1, value=d["format"])
        ws.cell(row=row, column=2, value=d["fr"])
        ws.cell(row=row, column=3, value=d["en"])
        style_body_row(ws, row, 3, alt=(i % 2 == 1), height=80)
        row += 1


def build_limites(wb: Workbook) -> None:
    ws = wb.create_sheet("6_Limites")
    set_widths(ws, {1: 30, 2: 80})
    write_title(ws, "Truth layer — Ce que l'agent NE remplace PAS", span=2)
    write_subtitle(ws, 2, "Disclaimer obligatoire à conserver dans tous les outputs.", span=2)

    rows = [
        ("Compliance officer export", "L'audit ITAR/EAR/dual-use est indicatif. Toute classification matériels de guerre, dual-use ou ECCN doit être validée par un officier de conformité export-control habilité (DGA / DDTC / BIS)."),
        ("Juriste défense", "L'agent ne se substitue pas à une revue juridique défense (Code défense L.2335-1, LPM 2024-2030, accords AUKUS Pillar II, accords G2G)."),
        ("Bureau d'études signataire", "Aucun chiffre technique généré par l'agent ne tient lieu d'engagement contractuel. Toute donnée doit être validée et signée par un ingénieur responsable du bureau d'études."),
        ("DGA / DGAC / EASA", "L'agent ne remplace ni la DGA pour les programmes défense FR, ni la DGAC/EASA pour la certification, ni le SGDSN pour la classification (Diffusion Restreinte / Secret Défense)."),
        ("DGCCRF / CMA / EASA Greenwashing", "L'audit anti-greenwashing est préventif. Toute campagne SAF/H2/électrique doit être contre-vérifiée par juriste consommation + responsable ESG signataire CSRD."),
        ("ANSSI / cybersécurité", "Les mentions NIS2/DORA/CC EAL/SecNumCloud sont indicatives. Toute revendication de certification doit être validée par un responsable RSSI signataire."),
        ("DPO / RGPD", "Si le contenu marketing inclut des traitements de données personnelles, validation DPO obligatoire."),
        ("Date de fraîcheur sources", "Sanctions OFAC SDN, EU listes, EAR Entity List évoluent en continu. Date de référence des sources : 25/04/2026. À rafraîchir avant toute diffusion sensible."),
        ("Ce que l'agent NE génère PAS", "Ne génère pas le contenu lui-même (rôle d'audit, pas de rédaction). Voir module rédaction sous supervision dans la roadmap v2."),
        ("Mode de fonctionnement", "Scenario-id only en démo publique. Pas d'ingestion de texte libre confidentiel sans contrat NDA + environnement isolé."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 70 + 1))
        row += 1


# ────────────────────────────────────────────────────────────────────
# Build main
# ────────────────────────────────────────────────────────────────────


def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb)
    build_referentiel(wb)
    build_cas_inputs(wb)
    build_verdict(wb)
    build_redlines(wb)
    build_ai_disclosure(wb)
    build_limites(wb)

    out = "AeroTechContent_NEURAL.xlsx"
    wb.save(out)
    print(f"OK — {out} written ({len(CASES)} cas, {len(RULES)} règles, {len(REDLINES)} redlines)")


if __name__ == "__main__":
    build()
