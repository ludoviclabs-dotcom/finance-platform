"""
strategic_mapping_excel.py — Génération du classeur Excel Value Mapping ESG.

Onglets :
  1. Synthèse         — Hero + KPIs grounded + meta
  2. Investissements  — 4 piliers avec budgets par segment
  3. Chaîne de valeur — 5 étapes avec precisionNote
  4. Avant-Après      — 7 catégories (rouge/vert)
  5. Gains & Externalités — financialGains + externalities
  6. Leviers Carbon&Co — 6 capacités avec moduleRef

Source de vérité unique : StrategicMappingResponse (même que l'endpoint JSON).
"""

from __future__ import annotations

import io
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from models.strategic_mapping import StrategicMappingResponse


# ---------------------------------------------------------------------------
# Palette couleurs Carbon & Co
# ---------------------------------------------------------------------------

_C_PRIMARY    = "059669"   # emerald-600
_C_PRIMARY_L  = "D1FAE5"   # emerald-100 (fond clair)
_C_DARK       = "0F172A"   # slate-900
_C_MUTED      = "64748B"   # slate-500
_C_LIGHT      = "F1F5F9"   # slate-100
_C_BORDER     = "CBD5E1"   # slate-300
_C_WHITE      = "FFFFFF"
_C_RED_L      = "FEE2E2"   # red-100
_C_GREEN_L    = "D1FAE5"   # emerald-100
_C_HEADER_FG  = "FFFFFF"


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 10, color: str = _C_DARK,
          italic: bool = False) -> "Font":
    return Font(bold=bold, size=size, color=color, italic=italic)


def _thin_border() -> "Border":
    side = Side(border_style="thin", color=_C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_row(ws: "object", row: int, cols: list[str]) -> None:
    """Écrit une ligne d'en-tête sur fond vert sombre."""
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _font(bold=True, size=10, color=_C_HEADER_FG)
        cell.fill = _fill(_C_PRIMARY)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = _thin_border()


def _data_cell(ws: "object", row: int, col: int, value: object,
               bg: str = _C_WHITE, bold: bool = False,
               italic: bool = False, wrap: bool = True) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, italic=italic)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    cell.border = _thin_border()


def _set_col_widths(ws: "object", widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws: "object", cell: str = "A2") -> None:
    ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# Onglet 1 — Synthèse
# ---------------------------------------------------------------------------

def _sheet_synthese(wb: "Workbook", data: StrategicMappingResponse) -> None:
    ws = wb.create_sheet("Synthèse")

    # Titre principal
    ws["A1"] = data.hero.title
    ws["A1"].font = _font(bold=True, size=14, color=_C_PRIMARY)
    ws["A1"].fill = _fill(_C_LIGHT)
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 40

    # Sous-titre
    ws["A2"] = data.hero.subtitle
    ws["A2"].font = _font(italic=True, size=10, color=_C_MUTED)
    ws.merge_cells("A2:C2")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 24

    # Résumé
    ws["A3"] = data.hero.summary
    ws["A3"].font = _font(size=10)
    ws.merge_cells("A3:C3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 60

    # Séparateur
    row = 5

    # Filtres appliqués
    ws.cell(row=row, column=1, value="Filtres appliqués").font = _font(bold=True, size=10, color=_C_MUTED)
    row += 1
    for label, val in [
        ("Segment", data.filters.segment),
        ("Persona", data.filters.persona),
        ("Horizon", data.filters.horizon),
    ]:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1

    row += 1

    # KPIs grounded si disponibles
    if data.groundedKpis and data.groundedKpis.dataAvailable:
        g = data.groundedKpis
        ws.cell(row=row, column=1, value="Données entreprise").font = _font(bold=True, size=10, color=_C_MUTED)
        row += 1
        if g.companyName:
            ws.cell(row=row, column=1, value="Entreprise").font = _font(bold=True)
            ws.cell(row=row, column=2, value=g.companyName)
            row += 1
        if g.totalS123Tco2e is not None:
            ws.cell(row=row, column=1, value="Émissions S1+2+3 (tCO₂e)").font = _font(bold=True)
            ws.cell(row=row, column=2, value=round(g.totalS123Tco2e, 1))
            row += 1
        if g.esgScoreGlobal is not None:
            ws.cell(row=row, column=1, value="Score ESG global").font = _font(bold=True)
            ws.cell(row=row, column=2, value=round(g.esgScoreGlobal, 1))
            row += 1
        if g.vsmeCompletion is not None:
            ws.cell(row=row, column=1, value="Complétion VSME (%)").font = _font(bold=True)
            ws.cell(row=row, column=2, value=round(g.vsmeCompletion, 1))
            row += 1
        if g.greenCapexPct is not None:
            ws.cell(row=row, column=1, value="Green Capex (%)").font = _font(bold=True)
            ws.cell(row=row, column=2, value=round(g.greenCapexPct, 1))
            row += 1
        row += 1

    # Métadonnées
    ws.cell(row=row, column=1, value="Métadonnées").font = _font(bold=True, size=10, color=_C_MUTED)
    row += 1
    for label, val in [
        ("Version", data.meta.version),
        ("Dernière révision", data.meta.lastReviewedAt),
        ("Prochaine révision", data.meta.nextReviewScheduled),
        ("Auteur", data.meta.contentOwner),
    ]:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Note").font = _font(bold=True)
    ws.cell(row=row, column=2,
            value="Les gains sont formulés en potentiel conditionnel — voir sources dans l'onglet Gains & Externalités.")
    ws.cell(row=row, column=2).font = _font(italic=True, size=9, color=_C_MUTED)
    ws.merge_cells(f"B{row}:C{row}")

    _set_col_widths(ws, [28, 55, 25])


# ---------------------------------------------------------------------------
# Onglet 2 — Investissements
# ---------------------------------------------------------------------------

def _sheet_investissements(wb: "Workbook", data: StrategicMappingResponse) -> None:
    ws = wb.create_sheet("Investissements")
    _header_row(ws, 1, ["Pilier", "Description", "Implique", "Segment", "Budget bas (€)", "Budget haut (€)", "Note budget", "Sources"])
    _freeze(ws)

    row = 2
    alt = False
    for pilier in data.investments:
        bg = _C_LIGHT if alt else _C_WHITE
        alt = not alt
        implies_text = "\n".join(f"• {item}" for item in pilier.implies)
        sources_text = "\n".join(
            f"{s.publisher} ({s.year})" + (f" — {s.url}" if s.url else "")
            for s in pilier.sources
        ) or "—"

        if pilier.budgetRanges:
            for br in pilier.budgetRanges:
                _data_cell(ws, row, 1, pilier.label, bg, bold=True)
                _data_cell(ws, row, 2, pilier.description, bg)
                _data_cell(ws, row, 3, implies_text, bg)
                _data_cell(ws, row, 4, br.segment)
                _data_cell(ws, row, 5, br.low)
                _data_cell(ws, row, 6, br.high)
                _data_cell(ws, row, 7, br.note or "")
                _data_cell(ws, row, 8, sources_text, bg)
                ws.row_dimensions[row].height = max(40, len(implies_text) // 3)
                row += 1
        else:
            _data_cell(ws, row, 1, pilier.label, bg, bold=True)
            _data_cell(ws, row, 2, pilier.description, bg)
            _data_cell(ws, row, 3, implies_text, bg)
            _data_cell(ws, row, 4, "—")
            _data_cell(ws, row, 5, None)
            _data_cell(ws, row, 6, None)
            _data_cell(ws, row, 7, "")
            _data_cell(ws, row, 8, sources_text, bg)
            row += 1

    _set_col_widths(ws, [20, 35, 40, 12, 14, 14, 28, 45])


# ---------------------------------------------------------------------------
# Onglet 3 — Chaîne de valeur
# ---------------------------------------------------------------------------

def _sheet_chaine_valeur(wb: "Workbook", data: StrategicMappingResponse) -> None:
    ws = wb.create_sheet("Chaîne de valeur")
    _header_row(ws, 1, ["Étape", "Label", "Description", "Note de précision"])
    _freeze(ws)

    alt = False
    for i, step in enumerate(data.valueChain, start=2):
        bg = _C_LIGHT if alt else _C_WHITE
        alt = not alt
        _data_cell(ws, i, 1, step.order, bg, bold=True)
        _data_cell(ws, i, 2, step.label, bg, bold=True)
        _data_cell(ws, i, 3, step.description, bg)
        _data_cell(ws, i, 4, step.precisionNote or "—", bg, italic=bool(step.precisionNote))
        ws.row_dimensions[i].height = 36

    _set_col_widths(ws, [8, 35, 55, 55])


# ---------------------------------------------------------------------------
# Onglet 4 — Avant / Après
# ---------------------------------------------------------------------------

def _sheet_avant_apres(wb: "Workbook", data: StrategicMappingResponse) -> None:
    ws = wb.create_sheet("Avant-Après")
    _header_row(ws, 1, ["Catégorie", "Tag d'impact", "Sans démarche (avant)", "Avec démarche (après)"])
    _freeze(ws)

    for i, item in enumerate(data.beforeAfter, start=2):
        _data_cell(ws, i, 1, item.category, bold=True)
        _data_cell(ws, i, 2, item.impactTag or "—")
        _data_cell(ws, i, 3, item.before, bg=_C_RED_L)
        _data_cell(ws, i, 4, item.after, bg=_C_GREEN_L)
        ws.row_dimensions[i].height = 32

    _set_col_widths(ws, [22, 16, 50, 50])


# ---------------------------------------------------------------------------
# Onglet 5 — Gains & Externalités
# ---------------------------------------------------------------------------

def _sheet_gains(wb: "Workbook", data: StrategicMappingResponse) -> None:
    ws = wb.create_sheet("Gains & Externalités")

    # --- Gains financiers ---
    ws["A1"] = "Gains financiers"
    ws["A1"].font = _font(bold=True, size=12, color=_C_PRIMARY)
    ws.merge_cells("A1:G1")

    _header_row(ws, 2, ["ID", "Label", "Description", "Magnitude", "Qualitatif", "Segments", "Sources"])
    _freeze(ws, "A3")

    row = 3
    for gain in data.financialGains:
        bg = _C_PRIMARY_L if not gain.qualitative else _C_WHITE
        sources_text = "\n".join(
            f"{s.publisher} ({s.year})" + (f" — {s.url}" if s.url else "")
            for s in gain.sources
        ) or "—"
        _data_cell(ws, row, 1, gain.id, bg)
        _data_cell(ws, row, 2, gain.label, bg, bold=True)
        _data_cell(ws, row, 3, gain.description, bg)
        _data_cell(ws, row, 4, gain.magnitude or "—", bg)
        _data_cell(ws, row, 5, "Non" if gain.qualitative else "Oui (chiffré)", bg)
        _data_cell(ws, row, 6, ", ".join(gain.segments), bg)
        _data_cell(ws, row, 7, sources_text, bg)
        ws.row_dimensions[row].height = 48
        row += 1

    row += 1

    # --- Externalités positives ---
    ws.cell(row=row, column=1, value="Externalités positives")
    ws.cell(row=row, column=1).font = _font(bold=True, size=12, color=_C_PRIMARY)
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    _header_row(ws, row, ["ID", "Label", "Catégorie", "Description", "Qualitatif", "Segments", "Sources"])
    row += 1

    for ext in data.externalities:
        bg = _C_LIGHT
        sources_text = "\n".join(
            f"{s.publisher} ({s.year})" + (f" — {s.url}" if s.url else "")
            for s in ext.sources
        ) or "—"
        _data_cell(ws, row, 1, ext.id, bg)
        _data_cell(ws, row, 2, ext.label, bg, bold=True)
        _data_cell(ws, row, 3, ext.category, bg)
        _data_cell(ws, row, 4, ext.description, bg)
        _data_cell(ws, row, 5, "Non" if ext.qualitative else "Oui (chiffré)", bg)
        _data_cell(ws, row, 6, ", ".join(ext.segments), bg)
        _data_cell(ws, row, 7, sources_text, bg)
        ws.row_dimensions[row].height = 48
        row += 1

    _set_col_widths(ws, [14, 30, 16, 55, 16, 24, 48])


# ---------------------------------------------------------------------------
# Onglet 6 — Leviers Carbon & Co
# ---------------------------------------------------------------------------

def _sheet_leviers(wb: "Workbook", data: StrategicMappingResponse) -> None:
    ws = wb.create_sheet("Leviers Carbon & Co")
    _header_row(ws, 1, ["ID", "Bénéfice activé", "Capacité Carbon & Co", "Module (URL interne)"])
    _freeze(ws)

    alt = False
    for i, lever in enumerate(data.carbonCoLevers, start=2):
        bg = _C_LIGHT if alt else _C_WHITE
        alt = not alt
        _data_cell(ws, i, 1, lever.id, bg)
        _data_cell(ws, i, 2, lever.benefit, bg, bold=True)
        _data_cell(ws, i, 3, lever.capability, bg)
        _data_cell(ws, i, 4, lever.moduleRef or "—", bg)
        ws.row_dimensions[i].height = 36

    _set_col_widths(ws, [14, 32, 55, 22])


# ---------------------------------------------------------------------------
# Point d'entrée public
# ---------------------------------------------------------------------------

def build_strategic_mapping_xlsx(data: StrategicMappingResponse) -> bytes:
    """
    Génère le classeur Excel Value Mapping ESG et retourne les bytes .xlsx.

    Raises:
        RuntimeError: si openpyxl n'est pas installé.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl n'est pas installé — pip install openpyxl")

    wb = Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    _sheet_synthese(wb, data)
    _sheet_investissements(wb, data)
    _sheet_chaine_valeur(wb, data)
    _sheet_avant_apres(wb, data)
    _sheet_gains(wb, data)
    _sheet_leviers(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
