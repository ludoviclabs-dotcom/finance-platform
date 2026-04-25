"""
Build DefenseCommsGuard_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'DefenseCommsGuard' (branche Aéronautique — Marketing).

7 onglets : README, Référentiel ITAR/EAR/EU/FR, Cas Inputs, Verdict par Cas,
Redlines, AI Act Disclosure, Limites. 5 cas marketing sensibles à auditer
avant diffusion (drone Iran, missile UK, radar GaN, brochure DSEI Algérie,
webinaire Chine FDPR).

Date de référence des sources : 25/04/2026.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"
AGENT_RED = "B91C1C"  # rouge spécifique DefenseCommsGuard

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=AGENT_RED, end_color=AGENT_RED)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def verdict_fill(verdict: str) -> PatternFill:
    if verdict == "OK":
        return green_fill()
    if verdict == "WARN":
        return amber_fill()
    if verdict == "KO":
        return red_fill()
    return PatternFill()


def verdict_color(verdict: str) -> str:
    if verdict == "OK":
        return GREEN_OK
    if verdict == "WARN":
        return AMBER_WARN
    if verdict == "KO":
        return RED_NO
    return "0E0824"


# ────────────────────────────────────────────────────────────────────
# Données : 12 règles export-control (référentiel)
# ────────────────────────────────────────────────────────────────────
RULES = [
    {
        "id": "D01",
        "label": "ITAR-CAT-VIII",
        "axis": "ITAR US",
        "rule": "Toute mention d'aéronef militaire (drone armé, hélicoptère armé, chasseur, bombardier) ou composant USML Cat. VIII doit être identifiée avant diffusion. Licence DDTC requise pour transfert.",
        "base": "ITAR US 22 CFR 121.1 Cat. VIII (révision DDTC janv 2026, notice 87 FR 11423).",
    },
    {
        "id": "D02",
        "label": "ITAR-CAT-XV",
        "axis": "ITAR US",
        "rule": "Tout système spatial (satellite militaire, télescope spatial, technologie GNSS hardened) entre dans USML Cat. XV — diffusion publique encadrée.",
        "base": "ITAR US 22 CFR 121.1 Cat. XV.",
    },
    {
        "id": "D03",
        "label": "EAR-ECCN-9A610-9A619",
        "axis": "EAR US",
        "rule": "Aéronefs militaires (9A610) et drones armés (9A619) sous EAR — exigent licence BIS et vérification Entity List.",
        "base": "EAR 15 CFR 774 ECCN 9A610/9A619 (BIS Entity List update mars 2026).",
    },
    {
        "id": "D04",
        "label": "EU-DUAL-USE",
        "axis": "EU dual-use",
        "rule": "Tout produit listé annexe I du Règlement EU 2021/821 (radars, optronique, propulsion, cybersécurité offensive) requiert licence EU avant export.",
        "base": "Règlement (UE) 2021/821 dual-use, annexe I.",
    },
    {
        "id": "D05",
        "label": "FR-LMG-L2335",
        "axis": "France défense",
        "rule": "Tout matériel de guerre listé arrêté du 27 juin 2012 (LMG) requiert AT/IDP/AGI avant export. Mention publique encadrée.",
        "base": "Code de la défense L.2335-1 + arrêté LMG 27/06/2012 + LPM 2024-2030 (Loi 2023-703).",
    },
    {
        "id": "D06",
        "label": "OFAC-SDN",
        "axis": "Sanctions US",
        "rule": "Aucun destinataire (entreprise, individu, pays) listé OFAC SDN ne doit recevoir le contenu — vérification consolidated list obligatoire.",
        "base": "OFAC SDN List (Treasury) — ajouts mars 2026 incluant Roselectronics RU + IRGC-AF IR.",
    },
    {
        "id": "D07",
        "label": "EU-SANCTIONS",
        "axis": "Sanctions EU",
        "rule": "Sanctions Russie (16e paquet fév 2026), Iran, Belarus, Corée du Nord — aucun produit annexe XXIII Reg. 833/2014 ne doit transiter.",
        "base": "Reg. (UE) 833/2014 + 16e paquet sanctions Russie (fév 2026) + Reg. 2023/1529 Iran.",
    },
    {
        "id": "D08",
        "label": "UK-OFSI",
        "axis": "Sanctions UK",
        "rule": "OFSI (HM Treasury) — consolidated list à vérifier pour tout destinataire UK ou produit transitant via UK post-Brexit.",
        "base": "OFSI Consolidated List + UK Sanctions and Anti-Money Laundering Act 2018.",
    },
    {
        "id": "D09",
        "label": "AUKUS-OGL",
        "axis": "AUKUS",
        "rule": "Pour produits éligibles AUKUS Pillar II (US/UK/AU) : Open General License DDTC permet transfert simplifié — vérifier scope produit.",
        "base": "AUKUS Pillar II Open General License (DDTC, sept 2024, étendue 2025).",
    },
    {
        "id": "D10",
        "label": "EAR-FDPR-DRONES",
        "axis": "FDPR drones",
        "rule": "Foreign Direct Product Rule étendue aux drones (2025) — produits étrangers contenant techno US sont soumis à EAR si destinés à pays/usage contrôlé.",
        "base": "EAR 15 CFR 734.9 + BIS extension drones (2025).",
    },
    {
        "id": "D11",
        "label": "AI-ACT-ART50",
        "axis": "AI Act",
        "rule": "Si scoring IA utilisé pour classification destinataire ou screening sanctions automatisé → disclosure obligatoire art. 50 AI Act.",
        "base": "EU AI Act 2024/1689 art. 50 (applicable 02/08/2026).",
    },
    {
        "id": "D12",
        "label": "ASD-RESPONSIBLE",
        "axis": "Tonalité",
        "rule": "Charte ASD Europe « responsible defence comms » : pas de glorification, pas de visuel inutilement choquant, contexte stratégique sourcé, pas de mention spécifique opérations en cours.",
        "base": "ASD Europe Charter on Responsible Defence Communications (mars 2025).",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : 5 cas démo sensibles
# ────────────────────────────────────────────────────────────────────
CASES = [
    {
        "id": "DEF-01",
        "type": "Email B2B prospection",
        "title": "Email cold drone tactique mini — destinataire potentiel Iran",
        "channel": "Email outbound",
        "extract": (
            "Cher prospect, suite à vos échanges au salon Dubai Airshow 2025, nous proposons "
            "notre drone tactique mini (autonomie 6h, charge 2,5 kg, vitesse 120 km/h). "
            "Il est compatible OTAN STANAG 4660 et utilise une caméra EO/IR thermique haute "
            "résolution. Notre commercial Téhéran peut planifier une démo le 15 mai. "
            "Tarif unitaire à partir de 85 K€ — exempt de TVA pour export."
        ),
        "verdicts": {"D01": "WARN", "D02": "OK", "D03": "KO", "D04": "KO", "D05": "WARN", "D06": "KO", "D07": "KO", "D08": "OK", "D09": "OK", "D10": "WARN", "D11": "WARN", "D12": "WARN"},
    },
    {
        "id": "DEF-02",
        "type": "Communiqué de presse",
        "title": "Communiqué pénétration radar furtif programme conjoint UK",
        "channel": "Press release + LinkedIn",
        "extract": (
            "Notre solution radar AESA bande X démontre des capacités de pénétration "
            "des défenses adverses inégalées. En partenariat avec un industriel britannique "
            "dans le cadre du programme défense conjoint AUKUS Pillar II, nous livrerons "
            "12 unités opérationnelles d'ici Q4 2026. Le système permet de détecter "
            "et neutraliser les menaces avant qu'elles ne soient identifiées par l'adversaire. "
            "Pour plus de détails, contactez notre responsable export."
        ),
        "verdicts": {"D01": "WARN", "D02": "OK", "D03": "WARN", "D04": "OK", "D05": "OK", "D06": "OK", "D07": "OK", "D08": "OK", "D09": "OK", "D10": "OK", "D11": "OK", "D12": "KO"},
    },
    {
        "id": "DEF-03",
        "type": "Post LinkedIn",
        "title": "Post technique sur radar GaN bande X — diffusion organique",
        "channel": "LinkedIn organique",
        "extract": (
            "Nouvelle génération de radar GaN bande X — gain antenne 32 dBi, portée 350 km, "
            "détection cible 1 m² @ 200 km. Architecture ouverte permettant intégration "
            "à toute chaîne C4ISR moderne. Idéal pour applications air-air, air-sol et "
            "surveillance maritime. Notre R&D bénéficie de financement EDF (35%) et nous "
            "exposerons une démo live à Farnborough 2026. Reservez votre slot via LinkedIn."
        ),
        "verdicts": {"D01": "OK", "D02": "OK", "D03": "WARN", "D04": "KO", "D05": "WARN", "D06": "OK", "D07": "OK", "D08": "OK", "D09": "OK", "D10": "OK", "D11": "OK", "D12": "OK"},
    },
    {
        "id": "DEF-04",
        "type": "Brochure salon",
        "title": "Brochure DSEI 2025 — distribution post-salon vers acheteur algérien",
        "channel": "Print + envoi PDF par email",
        "extract": (
            "[Brochure DSEI distribuée puis envoyée 3 semaines plus tard à acheteur "
            "algérien via email] Nos systèmes de défense sol-air courte portée offrent "
            "une protection 360°. Capacité d'engagement 12 cibles simultanées. "
            "Compatible munitions OTAN. Configurable selon le besoin du client. "
            "Disponible pour livraison Q1 2027 sous réserve d'agréments export."
        ),
        "verdicts": {"D01": "OK", "D02": "OK", "D03": "OK", "D04": "WARN", "D05": "KO", "D06": "WARN", "D07": "OK", "D08": "OK", "D09": "OK", "D10": "OK", "D11": "OK", "D12": "WARN"},
    },
    {
        "id": "DEF-05",
        "type": "Webinaire technique",
        "title": "Webinaire ouvert — drone surveillance, invité confirmé société chinoise",
        "channel": "Webinaire en ligne (Zoom public)",
        "extract": (
            "Webinaire « Drone surveillance maritime — état de l'art 2026 ». Slides "
            "techniques détaillant l'architecture hardware (CPU US Texas Instruments, "
            "FPGA Xilinx US, capteur EO/IR Made in EU), software de fusion de données "
            "propriétaire. Démonstration vidéo live des capacités. Invités confirmés : "
            "12 sociétés acheteuses dont 1 société chinoise listée Entity List 2025. "
            "Inscription ouverte sur LinkedIn."
        ),
        "verdicts": {"D01": "OK", "D02": "OK", "D03": "WARN", "D04": "WARN", "D05": "OK", "D06": "KO", "D07": "OK", "D08": "OK", "D09": "OK", "D10": "KO", "D11": "WARN", "D12": "OK"},
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Redlines (suggestions de correction)
# ────────────────────────────────────────────────────────────────────
REDLINES = [
    # DEF-01 : email Iran
    {"case": "DEF-01", "rule": "D06", "before": "commercial Téhéran peut planifier une démo", "after": "[BLOQUER ENVOI] Iran sous sanctions EU (Reg. 2023/1529) + OFAC SDN — aucun envoi n'est autorisé sans licence individuelle.", "why": "Iran = pays sanctionné US + EU. Tout envoi commercial est interdit sans licence DGA + dérogation OFAC."},
    {"case": "DEF-01", "rule": "D07", "before": "exempt de TVA pour export", "after": "[BLOQUER] Mention export vers Iran — vérifier Reg. (UE) 2023/1529 art. 5 + 5bis (interdiction technologies aéronautiques).", "why": "Réglementation EU restrictive Iran 2023/1529 interdit technologies aéronautiques sensibles."},
    {"case": "DEF-01", "rule": "D03", "before": "drone tactique mini (autonomie 6h, charge 2,5 kg)", "after": "[FLAG ECCN 9A619] — Drone tactique armable potentiellement sous EAR ECCN 9A619. Vérifier classification BIS avant tout devis.", "why": "Drone tactique armable ou armable-possible = ECCN 9A619 sous contrôle EAR strict."},
    {"case": "DEF-01", "rule": "D04", "before": "caméra EO/IR thermique haute résolution", "after": "[FLAG dual-use] — Capteur EO/IR thermique potentiellement annexe I Reg. 2021/821 (catégorie 6A002/6A003). Vérifier licence EU export.", "why": "Capteurs EO/IR haute résolution = annexe I Reg. 2021/821 — licence EU requise"},
    # DEF-02 : communiqué UK AUKUS
    {"case": "DEF-02", "rule": "D12", "before": "capacités de pénétration des défenses adverses inégalées... avant qu'elles ne soient identifiées par l'adversaire", "after": "Notre solution radar AESA bande X est conçue pour la détection longue portée et la résilience face aux contre-mesures électroniques modernes (réf. STANAG 4691).", "why": "Charte ASD Europe : éviter formulation glorifiante (« inégalées », « avant qu'elles ne soient identifiées »). Préférer description technique factuelle."},
    {"case": "DEF-02", "rule": "D01", "before": "permet de détecter et neutraliser les menaces", "after": "permet de détecter et caractériser les menaces (capacité d'engagement opéré par les systèmes intégrateurs)", "why": "« Neutraliser » suggère capacité offensive — peut requalifier produit en USML Cat. VIII. Préférer « détecter et caractériser »"},
    # DEF-03 : LinkedIn radar GaN
    {"case": "DEF-03", "rule": "D04", "before": "radar GaN bande X — gain antenne 32 dBi, portée 350 km, détection cible 1 m² @ 200 km", "after": "[FLAG dual-use Reg. 2021/821 annexe I cat. 6A008] Spécifications radar GaN bande X précises = informations techniques contrôlées. Retirer chiffres ou diffusion sur canal contrôlé uniquement.", "why": "Diffusion publique de spécifications radar haute performance peut constituer transfert de technologie contrôlée"},
    {"case": "DEF-03", "rule": "D05", "before": "Idéal pour applications air-air, air-sol et surveillance maritime", "after": "Adapté aux applications de surveillance et détection multi-environnements (préciser cas d'usage civil ou défense selon programme)", "why": "Mention multi-applications militaires sans contexte programme = peut requérir AT/IDP/AGI"},
    # DEF-04 : DSEI Algérie
    {"case": "DEF-04", "rule": "D05", "before": "[Brochure DSEI distribuée puis envoyée 3 semaines plus tard à acheteur algérien via email] Nos systèmes de défense sol-air...", "after": "[BLOQUER ENVOI ALGÉRIE] Système défense sol-air = matériel de guerre catégorie 2 LMG arrêté 27/06/2012. Export Algérie requiert AT (autorisation transit) DGA + licence individuelle. Aucune diffusion email post-salon sans licence préalable.", "why": "Matériel de guerre cat. 2 — export Algérie strictement encadré par DGA"},
    {"case": "DEF-04", "rule": "D06", "before": "envoi PDF par email à acheteur algérien", "after": "Vérifier acheteur algérien sur OFAC SDN List + liste UE sanctions Algérie (sectorielle énergie 2024). Si listé : refuser, sinon procédure DGA.", "why": "Algérie = pays sensible, vérification OFAC + EU obligatoire"},
    {"case": "DEF-04", "rule": "D12", "before": "Configurable selon le besoin du client", "after": "Configurable selon les spécifications opérationnelles convenues avec les autorités d'acquisition compétentes du pays acheteur", "why": "Tonalité ASD : neutraliser le côté commercial agressif sur du défense sol-air"},
    # DEF-05 : webinaire Chine FDPR
    {"case": "DEF-05", "rule": "D06", "before": "1 société chinoise listée Entity List 2025", "after": "[BLOQUER PARTICIPATION] Toute société listée BIS Entity List ne peut accéder aux contenus techniques sans licence individuelle. Désinscription immédiate + audit qui d'autre.", "why": "Entity List BIS = interdiction transfert technologie sans licence — applicable à webinaires techniques"},
    {"case": "DEF-05", "rule": "D10", "before": "CPU US Texas Instruments, FPGA Xilinx US", "after": "[FDPR risque élevé] Composants US (TI CPU, Xilinx FPGA) = produit final soumis à FDPR EAR 15 CFR 734.9. Diffusion technique vers Chine = transfert contrôlé. Webinaire en accès restreint requis.", "why": "FDPR étendue drones 2025 — produits intégrant techno US contrôlés même hors US"},
    {"case": "DEF-05", "rule": "D04", "before": "capteur EO/IR Made in EU... fusion de données propriétaire", "after": "[FLAG dual-use cat. 6A] Capteur EO/IR + fusion de données = annexe I Reg. 2021/821 cat. 6A002/6A005. Vérifier licence EU export avant diffusion technique.", "why": "Capteur EO/IR + algorithme fusion = dual-use catégorie 6A"},
    {"case": "DEF-05", "rule": "D11", "before": "[Pas de mention IA]", "after": "Si scoring IA utilisé pour la sélection automatique des invités ou la modération du chat → disclosure art. 50 obligatoire dans la confirmation d'inscription.", "why": "AI Act art. 50 — disclosure obligatoire si IA utilisée"},
]

# ────────────────────────────────────────────────────────────────────
# Données : AI Act Disclosure modèles
# ────────────────────────────────────────────────────────────────────
AI_DISCLOSURE = [
    {"format": "Screening sanctions auto", "fr": "Le screening sanctions de ce destinataire utilise un système d'IA assisté par rapprochement consolidé (OFAC SDN, EU Sanctions, UK OFSI). Décision finale validée humainement par compliance officer. (UE 2024/1689 art. 50.)", "en": "Sanctions screening for this recipient uses an AI-assisted consolidated matching system (OFAC SDN, EU Sanctions, UK OFSI). Final decision validated by compliance officer. (EU 2024/1689 art. 50.)"},
    {"format": "Audit communication marketing", "fr": "Cette communication a été pré-auditée par un système d'IA contre 12 règles export-control. Verdict final approuvé par responsable compliance. (UE 2024/1689 art. 50.)", "en": "This communication has been pre-audited by an AI system against 12 export-control rules. Final verdict approved by compliance officer. (EU 2024/1689 art. 50.)"},
    {"format": "Réponse RFI/RFP défense", "fr": "Annexe disclosure : la pré-classification ITAR/EAR/dual-use de cette réponse a été réalisée par un système d'IA. Validation finale par responsable export-control habilité. (UE 2024/1689 art. 50.)", "en": "Disclosure annex: ITAR/EAR/dual-use pre-classification of this response performed by an AI system. Final validation by accredited export-control officer. (EU 2024/1689 art. 50.)"},
    {"format": "Webinaire technique", "fr": "Inscription au webinaire : la liste des participants est filtrée automatiquement contre BIS Entity List + OFAC SDN par un système d'IA. (UE 2024/1689 art. 50.)", "en": "Webinar registration: participant list automatically filtered against BIS Entity List + OFAC SDN by an AI system. (EU 2024/1689 art. 50.)"},
    {"format": "Brochure salon — verso footer", "fr": "Audit pré-diffusion par IA contre référentiel ITAR/EAR/EU/FR/UK. Validation finale humaine. UE 2024/1689 art. 50.", "en": "Pre-distribution AI audit against ITAR/EAR/EU/FR/UK reference. Human final validation. EU 2024/1689 art. 50."},
]


# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("0_README")
    set_widths(ws, {1: 26, 2: 92})
    write_title(ws, "DefenseCommsGuard — Audit export-control des communications marketing", span=2)
    write_subtitle(ws, 2, "NEURAL — Branche Aéronautique / Marketing — Date de réf. : 25/04/2026", span=2)

    rows = [
        ("Mission", "Auditer toute communication marketing (email, presse, social, brochure, webinaire) contre 12 règles export-control + sanctions consolidées : ITAR (US), EAR (US), EU dual-use, France L.2335-1 LMG, OFAC SDN, EU sanctions packages, UK OFSI, AUKUS OGL, FDPR drones, AI Act art. 50, charte ASD."),
        ("Périmètre", "Aéronautique défense + dual-use civil. Communications B2B et corporate. Distinct de l'agent corporate aero-comms AGA001 (qui audite les communiqués CEO et investor relations)."),
        ("Référentiel", "12 règles encodées (D01-D12) — ITAR Cat. VIII/XV, EAR ECCN 9A610/9A619, EU Reg. 2021/821, FR Code défense L.2335-1, OFAC SDN, EU 16e paquet, UK OFSI, AUKUS OGL, EAR FDPR drones, AI Act, ASD Charter responsible defence comms."),
        ("Échelle verdict", "OK = conforme | WARN = revue compliance recommandée | KO = bloquant, redline obligatoire avant diffusion."),
        ("Décision", "0 KO ET score ≥ 80% → DIFFUSION OK. Sinon : RELECTURE compliance officer (WARN cumulés) ou BLOQUÉ (KO ≥ 1)."),
        ("Inputs cas démo", "5 cas synthétiques inspirés du marché 2026 : email drone vers Iran (sanctions), communiqué AUKUS UK, post LinkedIn radar GaN, brochure DSEI envoyée Algérie, webinaire technique avec invité chinois Entity List."),
        ("Outputs", "(1) Verdict matrice 12 × 5, (2) score conformité auto, (3) décision auto, (4) ~14 redlines concrètes avec base légale citée, (5) 5 modèles disclosure IA FR + EN."),
        ("Statut", "Démo recruteur — destinataires synthétiques anonymisés. Aucune information classifiée, aucune référence à programme réel sensible."),
        ("Disclaimer", "OUTIL INDICATIF. Ne remplace EN AUCUN CAS : (1) DDTC pour licence ITAR US, (2) BIS pour licence EAR US, (3) DGA pour AT/IDP/AGI matériels de guerre français, (4) compliance officer export-control habilité, (5) juriste défense, (6) Trésor pour validation OFAC. Toute diffusion finale doit être validée humainement et conformément aux processus internes de l'entreprise."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 75 + 1))
        row += 1


def build_referentiel(wb: Workbook) -> None:
    ws = wb.create_sheet("1_Referentiel")
    set_widths(ws, {1: 8, 2: 22, 3: 16, 4: 60, 5: 55})
    write_title(ws, "Référentiel — 12 règles export-control & sanctions", span=5)
    write_subtitle(ws, 2, "ITAR + EAR + EU dual-use + FR LMG + sanctions OFAC/EU/UK + AUKUS + FDPR + AI Act + ASD Charter.", span=5)

    headers = ["ID", "Label", "Axe", "Règle", "Base légale / réf."]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        ws.cell(row=row, column=3, value=rule["axis"])
        ws.cell(row=row, column=4, value=rule["rule"])
        ws.cell(row=row, column=5, value=rule["base"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=80)
        row += 1


def build_cas_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("2_Cas_Inputs")
    set_widths(ws, {1: 12, 2: 22, 3: 38, 4: 26, 5: 70})
    write_title(ws, "Cas inputs — 5 communications marketing à auditer", span=5)
    write_subtitle(ws, 2, "Cas synthétiques 2026. Destinataires anonymisés. Aucun client réel ni programme classifié.", span=5)

    headers = ["ID", "Type", "Titre", "Canal", "Extrait"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, c in enumerate(CASES):
        ws.cell(row=row, column=1, value=c["id"])
        ws.cell(row=row, column=2, value=c["type"])
        ws.cell(row=row, column=3, value=c["title"])
        ws.cell(row=row, column=4, value=c["channel"])
        ws.cell(row=row, column=5, value=c["extract"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=160)
        row += 1


def build_verdict(wb: Workbook) -> None:
    ws = wb.create_sheet("3_Verdict_Par_Cas")
    n_cases = len(CASES)
    n_rules = len(RULES)
    cols_total = 2 + n_cases

    widths = {1: 8, 2: 22}
    for i in range(n_cases):
        widths[3 + i] = 14
    set_widths(ws, widths)

    write_title(ws, f"Matrice verdicts — {n_rules} règles × {n_cases} cas", span=cols_total)
    write_subtitle(ws, 2, "OK = conforme | WARN = revue compliance | KO = bloquant. Score + décision auto en bas.", span=cols_total)

    ws.cell(row=4, column=1, value="ID")
    ws.cell(row=4, column=2, value="Règle")
    for i, c in enumerate(CASES):
        ws.cell(row=4, column=3 + i, value=c["id"])
    style_header_row(ws, 4, cols_total)

    row = 5
    for r_idx, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        for c_idx, c in enumerate(CASES):
            v = c["verdicts"].get(rule["id"], "")
            cell = ws.cell(row=row, column=3 + c_idx, value=v)
            cell.font = body_font(bold=True, color=verdict_color(v))
            cell.fill = verdict_fill(v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).font = body_font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).font = body_font()
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    write_section_header(ws, row, "Synthèse par cas", span=cols_total)
    row += 1

    last_verdict_row = 4 + n_rules

    ws.cell(row=row, column=2, value="Nb OK")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")')
    style_body_row(ws, row, cols_total, height=22)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    ws.cell(row=row, column=2, value="Nb WARN")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN")')
    style_body_row(ws, row, cols_total, height=22)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    ws.cell(row=row, column=2, value="Nb KO")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"KO")')
    style_body_row(ws, row, cols_total, height=22)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    ws.cell(row=row, column=2, value="Score conformité")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        formula = f'=(COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN"))/{n_rules}'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.number_format = "0%"
    style_body_row(ws, row, cols_total, height=24)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    ws.cell(row=row, column=2, value="Décision")
    for c_idx in range(n_cases):
        col_letter = get_column_letter(3 + c_idx)
        ko_count = f'COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"KO")'
        score = f'(COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN"))/{n_rules}'
        formula = f'=IF({ko_count}>0,"BLOQUÉ",IF({score}>=0.8,"DIFFUSION OK","RELECTURE"))'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.font = body_font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    style_body_row(ws, row, cols_total, height=26)
    ws.cell(row=row, column=2).font = body_font(bold=True)


def build_redlines(wb: Workbook) -> None:
    ws = wb.create_sheet("4_Redlines")
    set_widths(ws, {1: 10, 2: 8, 3: 50, 4: 65, 5: 45})
    write_title(ws, "Redlines — Suggestions de correction concrètes", span=5)
    write_subtitle(ws, 2, "Pour chaque KO/WARN, redline applicable avec base légale citée. Validation compliance officer obligatoire avant diffusion.", span=5)

    headers = ["Cas", "Règle", "Avant (extrait)", "Après (suggestion redline)", "Pourquoi"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rl in enumerate(REDLINES):
        ws.cell(row=row, column=1, value=rl["case"])
        ws.cell(row=row, column=2, value=rl["rule"])
        ws.cell(row=row, column=3, value=rl["before"])
        ws.cell(row=row, column=4, value=rl["after"])
        ws.cell(row=row, column=5, value=rl["why"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=80)
        row += 1


def build_ai_disclosure(wb: Workbook) -> None:
    ws = wb.create_sheet("5_AI_Act_Disclosure")
    set_widths(ws, {1: 28, 2: 60, 3: 60})
    write_title(ws, "AI Act art. 50 — Modèles de mention IA", span=3)
    write_subtitle(ws, 2, "Applicable 2 août 2026. Disclosure obligatoire si IA utilisée pour audit ou screening sanctions.", span=3)

    headers = ["Format", "Mention FR", "Mention EN"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 3)

    row = 5
    for i, d in enumerate(AI_DISCLOSURE):
        ws.cell(row=row, column=1, value=d["format"])
        ws.cell(row=row, column=2, value=d["fr"])
        ws.cell(row=row, column=3, value=d["en"])
        style_body_row(ws, row, 3, alt=(i % 2 == 1), height=85)
        row += 1


def build_limites(wb: Workbook) -> None:
    ws = wb.create_sheet("6_Limites")
    set_widths(ws, {1: 32, 2: 80})
    write_title(ws, "Truth layer — Ce que DefenseCommsGuard NE remplace PAS", span=2)
    write_subtitle(ws, 2, "Disclaimer obligatoire à conserver dans tous les outputs. Sujet sensible : export-control + sanctions.", span=2)

    rows = [
        ("DDTC (US Department of State)", "L'agent ne se substitue PAS au Directorate of Defense Trade Controls pour les licences ITAR US. Toute classification USML doit être validée par DDTC avant export."),
        ("BIS (US Department of Commerce)", "L'agent ne se substitue PAS au Bureau of Industry and Security pour les licences EAR US ni pour la vérification Entity List et FDPR."),
        ("DGA (Direction Générale de l'Armement)", "L'agent ne se substitue PAS à la DGA pour les agréments AT (Autorisation Transit), IDP (Importation), AGI (Agrément Global Investissement) ni pour la classification matériel de guerre L.2335-1."),
        ("OFAC (US Treasury)", "L'agent fournit un screening indicatif. La décision finale de transaction est de la responsabilité du compliance officer + officier OFAC habilité. Sanctions évoluent en continu."),
        ("Trésor / Direction des Douanes FR", "Validation finale des opérations avec dual-use ou matériel de guerre relève du Trésor + Douanes pour les contrôles EU et FR."),
        ("Compliance Officer Export Control", "L'agent est un outil d'aide à la décision. Le compliance officer habilité reste seul responsable de la décision DIFFUSION OK / BLOQUÉ."),
        ("Juriste défense", "Toute clause contractuelle, accord G2G, cooperation agreement requiert validation par juriste défense indépendant."),
        ("SGDSN — Classification", "L'agent ne classifie PAS les informations Diffusion Restreinte / Secret Défense. Cette classification relève du SGDSN et de l'officier de sécurité."),
        ("Date de fraîcheur des sources", "OFAC SDN, EU consolidated list, BIS Entity List évoluent en continu (mise à jour parfois quotidienne). Date de référence figée : 25/04/2026. À rafraîchir avant TOUTE diffusion sensible."),
        ("Ce que l'agent NE génère PAS", "Ne génère pas de contenu marketing — uniquement audit. Pour rédaction sous supervision, voir AeroTechContent + AeroEventAI."),
        ("Mode démo public", "Scenario-id only. Pas d'ingestion de communication confidentielle réelle sans environnement isolé + NDA + gates server-side."),
        ("Limite v1", "Pas de croisement temps-réel avec les bases consolidées OFAC/EU/UK. v1 utilise un snapshot daté. v2 prévue avec connecteur API officielle."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 70 + 1))
        row += 1


# ────────────────────────────────────────────────────────────────────
# Build main
# ────────────────────────────────────────────────────────────────────


def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb)
    build_referentiel(wb)
    build_cas_inputs(wb)
    build_verdict(wb)
    build_redlines(wb)
    build_ai_disclosure(wb)
    build_limites(wb)

    out = "DefenseCommsGuard_NEURAL.xlsx"
    wb.save(out)
    print(f"OK — {out} written ({len(CASES)} cas, {len(RULES)} règles, {len(REDLINES)} redlines)")


if __name__ == "__main__":
    build()
