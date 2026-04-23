# ═══════════════════════════════════════════════════════════════════════════════
# AGENT LUXETRACEABILITY — NEURAL ENTERPRISE FRAMEWORK
# Supply Chain Luxe — Traçabilité Matières Premières
# Version 1.0 — Avril 2026 — Onglets 1 à 4
# ═══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────
# CONSTANTES — CHARTE GRAPHIQUE LUXE
# ─────────────────────────────────────────────────────────────────────

NOIR_LUXE        = "1A1A1A"
OR_LUXE          = "C9A84C"
BORDEAUX         = "6B1D2A"
BLANC_CASSE      = "F9F6F0"
GRIS_LUXE        = "8C8C8C"
VERT_VALIDATION  = "2E7D32"
ROUGE_ALERTE     = "C62828"

FOND_CRITIQUE    = "FFC7CE"
FOND_MOYEN       = "FFEB9C"
FOND_FAIBLE      = "C6EFCE"
FOND_TRES_FAIBLE = "A8D5A2"

BORDURE_FINE = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────
# FONCTIONS UTILITAIRES
# ─────────────────────────────────────────────────────────────────────

def creer_titre_principal(ws, texte, sous_texte, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    t = ws.cell(row=1, column=1, value=texte)
    t.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    t.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    s = ws.cell(row=2, column=1, value=sous_texte)
    s.font = Font(name="Calibri", size=10, italic=True, color=OR_LUXE)
    s.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28


def appliquer_style_titre_section(ws, row, max_col, texte, couleur=BORDEAUX):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=texte)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDURE_FINE
    ws.row_dimensions[row].height = 30


def appliquer_style_en_tetes(ws, row, headers, largeurs=None):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDURE_FINE
    ws.row_dimensions[row].height = 36
    if largeurs:
        for i, w in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def appliquer_style_donnees(ws, start_row, end_row, max_col):
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.font or cell.font.name is None:
                cell.font = Font(name="Calibri", size=9)
            cell.border = BORDURE_FINE
            if (row_idx - start_row) % 2 == 0:
                cur = cell.fill
                if not cur or cur.fill_type == "none" or (hasattr(cur, 'start_color') and cur.start_color.index in ("00000000", "FFFFFFFF", None)):
                    cell.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")


def ajouter_mfc_statut(ws, col_letter, start_row, end_row):
    plage = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for val, fond, texte in [
        ("Conforme",     FOND_FAIBLE,      "006100"),
        ("Validé",       FOND_FAIBLE,      "006100"),
        ("Actif",        FOND_FAIBLE,      "006100"),
        ("Certifié",     FOND_FAIBLE,      "006100"),
        ("Complet",      FOND_FAIBLE,      "006100"),
        ("En attente",   FOND_MOYEN,       "9C6500"),
        ("En cours",     FOND_MOYEN,       "9C6500"),
        ("Partiel",      FOND_MOYEN,       "9C6500"),
        ("Non conforme", FOND_CRITIQUE,    "9C0006"),
        ("Rejeté",       FOND_CRITIQUE,    "9C0006"),
        ("Suspendu",     FOND_CRITIQUE,    "9C0006"),
        ("Expiré",       FOND_CRITIQUE,    "9C0006"),
        ("Incomplet",    FOND_CRITIQUE,    "9C0006"),
    ]:
        ws.conditional_formatting.add(plage,
            CellIsRule(operator='equal', formula=[f'"{val}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte, bold=(texte != "9C6500"))))


def ajouter_mfc_zone_risque(ws, col_letter, start_row, end_row):
    plage = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for val, fond, texte, bold in [
        ("CRITIQUE",    "FF0000",         "FFFFFF", True),
        ("ÉLEVÉ",       "FF6600",         "FFFFFF", True),
        ("MOYEN",       FOND_MOYEN,       "9C6500", False),
        ("FAIBLE",      FOND_FAIBLE,      "006100", True),
        ("TRÈS FAIBLE", FOND_TRES_FAIBLE, "006100", True),
    ]:
        ws.conditional_formatting.add(plage,
            CellIsRule(operator='equal', formula=[f'"{val}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte, bold=bold)))


def ligne_synthese(ws, row, nb_col, items):
    """
    items = list of (col, value, format, color)
    """
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    ws.cell(row=row, column=1).border = BORDURE_FINE
    for c in range(2, nb_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        ws.cell(row=row, column=c).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws.cell(row=row, column=c).border = BORDURE_FINE
    for (col, val, fmt, color) in items:
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color=color or OR_LUXE)
        cell.border = BORDURE_FINE
        if fmt:
            cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION DU WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def generer_luxetraceability():
    wb = Workbook()
    print("\n  🔨 Démarrage de la génération NEURAL_LuxeTraceability.xlsx\n")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 1 : DASHBOARD
    # ═══════════════════════════════════════════════════════════════════

    ws1 = wb.active
    ws1.title = "1_DASHBOARD"
    ws1.sheet_properties.tabColor = OR_LUXE

    MAX_COL_DASH = 12
    creer_titre_principal(
        ws1,
        "🏛️  AGENT LUXETRACEABILITY — TABLEAU DE BORD CENTRAL",
        "Neural Enterprise Framework — Supply Chain Luxe — Traçabilité Matières Premières — Avril 2026",
        MAX_COL_DASH
    )

    # Bandeau info
    ws1.merge_cells("A3:L3")
    info = ws1.cell(row=3, column=1,
        value="📡 Données temps réel via formules inter-onglets  |  "
              "🔒 CSRD 2024/2862 · CSDDD 2024/1760 · Loi Vigilance 2017-399 · EUDR 2023/1115  |  "
              f"⏱️ Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    info.font = Font(name="Calibri", size=8, italic=True, color=GRIS_LUXE)
    info.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")
    info.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 22

    # ── Section KPIs principaux ──
    appliquer_style_titre_section(ws1, 4, MAX_COL_DASH, "INDICATEURS CLÉS — SUPPLY CHAIN LUXE", BORDEAUX)

    kpi_labels = [
        ("A5", "B5", "MATIÈRES\nRÉFÉRENCÉES",    "=COUNTA('2_MATIERES'!A5:A5000)",       OR_LUXE,   None),
        ("C5", "D5", "FOURNISSEURS\nACTIFS",       "=COUNTIF('3_FOURNISSEURS'!K5:K5000,\"Actif\")", OR_LUXE, None),
        ("E5", "F5", "LOTS EN\nTRACÉ",             "=COUNTA('4_TRACABILITE'!A5:A5000)",    OR_LUXE,   None),
        ("G5", "H5", "TAUX CERT.\nMATIÈRES",       "=IFERROR(COUNTIF('2_MATIERES'!L5:L5000,\"Certifié\")/COUNTA('2_MATIERES'!A5:A5000),0)", OR_LUXE, "0.0%"),
        ("I5", "J5", "FOURNISSEURS\nSUSPENDUS",     "=COUNTIF('3_FOURNISSEURS'!K5:K5000,\"Suspendu\")", "FF0000", None),
        ("K5", "L5", "LOTS\nCOMPLETS",              "=COUNTIF('4_TRACABILITE'!L5:L5000,\"Complet\")", VERT_VALIDATION, None),
    ]

    for label_cell, val_cell, label_text, formula, color, fmt in kpi_labels:
        lc = ws1[label_cell]
        lc.value = label_text
        lc.font = Font(name="Calibri", size=9, bold=True, color=GRIS_LUXE)
        lc.fill = PatternFill(start_color="0D0D0D", end_color="0D0D0D", fill_type="solid")
        lc.alignment = ALIGN_CENTER
        lc.border = BORDURE_FINE
        ws1.row_dimensions[5].height = 32

        vc = ws1[val_cell]
        vc.value = formula
        vc.font = Font(name="Calibri", size=22, bold=True, color=color)
        vc.fill = PatternFill(start_color="0D0D0D", end_color="0D0D0D", fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = BORDURE_FINE
        ws1.row_dimensions[6].height = 48
        if fmt:
            vc.number_format = fmt

    # ── Section risques ──
    appliquer_style_titre_section(ws1, 7, MAX_COL_DASH, "RÉPARTITION PAR NIVEAU DE RISQUE FOURNISSEUR", BORDEAUX)

    risk_headers = ["Niveau de risque", "Nb fournisseurs", "% du total", "Volume achats (€)", "Action requise"]
    risk_largeurs = [22, 18, 14, 22, 32]
    for i, h in enumerate(risk_headers, 1):
        c = ws1.cell(row=8, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = BORDURE_FINE
    ws1.row_dimensions[8].height = 32

    risk_data = [
        ("TRÈS FAIBLE", f'=COUNTIF(\'3_FOURNISSEURS\'!M5:M5000,"TRÈS FAIBLE")',
         f'=IFERROR(B9/COUNTA(\'3_FOURNISSEURS\'!A5:A5000),0)',
         f'=SUMIF(\'3_FOURNISSEURS\'!M5:M5000,"TRÈS FAIBLE",\'3_FOURNISSEURS\'!P5:P5000)',
         "Surveillance standard — Audit triennal"),
        ("FAIBLE",      f'=COUNTIF(\'3_FOURNISSEURS\'!M5:M5000,"FAIBLE")',
         f'=IFERROR(B10/COUNTA(\'3_FOURNISSEURS\'!A5:A5000),0)',
         f'=SUMIF(\'3_FOURNISSEURS\'!M5:M5000,"FAIBLE",\'3_FOURNISSEURS\'!P5:P5000)',
         "Audit annuel — Certification maintenue"),
        ("MOYEN",       f'=COUNTIF(\'3_FOURNISSEURS\'!M5:M5000,"MOYEN")',
         f'=IFERROR(B11/COUNTA(\'3_FOURNISSEURS\'!A5:A5000),0)',
         f'=SUMIF(\'3_FOURNISSEURS\'!M5:M5000,"MOYEN",\'3_FOURNISSEURS\'!P5:P5000)',
         "Plan d'action 90j — Audit renforcé"),
        ("ÉLEVÉ",       f'=COUNTIF(\'3_FOURNISSEURS\'!M5:M5000,"ÉLEVÉ")',
         f'=IFERROR(B12/COUNTA(\'3_FOURNISSEURS\'!A5:A5000),0)',
         f'=SUMIF(\'3_FOURNISSEURS\'!M5:M5000,"ÉLEVÉ",\'3_FOURNISSEURS\'!P5:P5000)',
         "⚠️ Audit urgent 30j — Alternative sourcing"),
        ("CRITIQUE",    f'=COUNTIF(\'3_FOURNISSEURS\'!M5:M5000,"CRITIQUE")',
         f'=IFERROR(B13/COUNTA(\'3_FOURNISSEURS\'!A5:A5000),0)',
         f'=SUMIF(\'3_FOURNISSEURS\'!M5:M5000,"CRITIQUE",\'3_FOURNISSEURS\'!P5:P5000)',
         "🔴 SUSPENSION IMMÉDIATE — Notification Vigilance"),
    ]

    risk_colors = [FOND_TRES_FAIBLE, FOND_FAIBLE, FOND_MOYEN, "FFD9CC", FOND_CRITIQUE]
    risk_txt    = ["006100",         "006100",    "9C6500",   "BF4000", "9C0006"]

    for i, (niv, nb, pct, vol, action) in enumerate(risk_data):
        row = 9 + i
        for col, val in [(1, niv), (2, nb), (3, pct), (4, vol), (5, action)]:
            c = ws1.cell(row=row, column=col, value=val)
            c.font = Font(name="Calibri", size=10, bold=(col == 1), color=risk_txt[i])
            c.fill = PatternFill(start_color=risk_colors[i], end_color=risk_colors[i], fill_type="solid")
            c.alignment = ALIGN_CENTER
            c.border = BORDURE_FINE
            if col == 3:
                c.number_format = "0.0%"
            if col == 4:
                c.number_format = "#,##0 €"
        ws1.row_dimensions[row].height = 22

    # ── Section statut réglementaire ──
    appliquer_style_titre_section(ws1, 15, MAX_COL_DASH, "STATUT CONFORMITÉ RÉGLEMENTAIRE", BORDEAUX)

    reg_headers = ["Réglementation", "Périmètre", "Fournisseurs concernés", "Statut global", "Prochaine échéance"]
    for i, h in enumerate(reg_headers, 1):
        c = ws1.cell(row=16, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        c.alignment = ALIGN_CENTER
        c.border = BORDURE_FINE
    ws1.row_dimensions[16].height = 32

    reglements = [
        ("CSRD 2024/2862", "Reporting durabilité — Toutes entités UE > 250 salariés", "Tous fournisseurs Tier 1", "En cours", "31/12/2026 — Rapport annuel"),
        ("CSDDD 2024/1760", "Devoir de vigilance chaîne de valeur — Due diligence droits humains", "Fournisseurs Tier 1 & 2", "Partiel", "26/07/2027 — Entrée en vigueur"),
        ("Loi Vigilance 2017-399", "Plan vigilance — Risques graves droits humains et environnement", "Fournisseurs stratégiques", "Conforme", "Rapport annuel — Avril 2026"),
        ("EUDR 2023/1115", "Déforestation — Bois, cuir, caoutchouc, cacao, café, palmier, soja", "Ébène + Cuirs bovins", "Partiel", "30/12/2025 — DDS obligatoire ⚠️"),
        ("REACH 1907/2006", "Substances chimiques — Tanneries, traitements cuirs", "Tanneries Tier 1", "Conforme", "Veille continue"),
        ("CITES", "Espèces menacées — Python, crocodile, vigogne, ivoire", "Fournisseurs matières exotiques", "Conforme", "Renouvellement permis 2027"),
    ]

    reg_statut_color = {
        "En cours": (FOND_MOYEN, "9C6500"),
        "Partiel":  (FOND_MOYEN, "9C6500"),
        "Conforme": (FOND_FAIBLE, "006100"),
        "🔴 Non conforme": (FOND_CRITIQUE, "9C0006"),
    }

    for i, (reg, perimetre, fourn, statut, echeance) in enumerate(reglements):
        row = 17 + i
        for col, val in [(1, reg), (2, perimetre), (3, fourn), (4, statut), (5, echeance)]:
            c = ws1.cell(row=row, column=col, value=val)
            fond, txt = reg_statut_color.get(statut, (BLANC_CASSE, NOIR_LUXE))
            c.font = Font(name="Calibri", size=9, color=txt, bold=(col == 4))
            c.fill = PatternFill(start_color=fond if col == 4 else BLANC_CASSE,
                                 end_color=fond if col == 4 else BLANC_CASSE, fill_type="solid")
            c.alignment = ALIGN_CENTER
            c.border = BORDURE_FINE
        ws1.row_dimensions[row].height = 20

    # ── Pied de page ──
    last_row = 24
    ws1.merge_cells(f"A{last_row}:L{last_row}")
    footer = ws1.cell(row=last_row, column=1,
        value="NEURAL — Intelligence Augmentée · Agent LuxeTraceability · Supply Chain Luxe v1.0 · "
              "Confidentiel — Usage interne Maison Aurelia")
    footer.font = Font(name="Calibri", size=8, italic=True, color=GRIS_LUXE)
    footer.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    footer.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[last_row].height = 20

    for col in range(1, MAX_COL_DASH + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    ws1.freeze_panes = "A4"
    print("  ✅ Onglet 1_DASHBOARD créé")


    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 2 : MATIERES
    # ═══════════════════════════════════════════════════════════════════

    ws2 = wb.create_sheet("2_MATIERES")
    ws2.sheet_properties.tabColor = OR_LUXE

    MAX_COL_MAT = 16
    creer_titre_principal(
        ws2,
        "💎  RÉFÉRENTIEL MATIÈRES PREMIÈRES — AGENT LUXETRACEABILITY",
        "Registre centralisé des matières premières luxe — Classification, origine, risques — Avril 2026",
        MAX_COL_MAT
    )
    appliquer_style_titre_section(ws2, 3, MAX_COL_MAT, "SECTION A — INVENTAIRE DES MATIÈRES PREMIÈRES", BORDEAUX)

    mat_headers = [
        "ID_Matière", "Catégorie", "Sous-catégorie", "Désignation complète",
        "Origine géographique", "Pays", "Zone de risque",
        "Fournisseur principal", "Fournisseur backup", "Certification requise",
        "Certification obtenue", "Statut certification", "Score risque (/100)",
        "Prix unitaire (€)", "Unité", "Observations"
    ]
    mat_largeurs = [14, 16, 18, 34, 24, 14, 14, 26, 26, 22, 22, 16, 16, 16, 10, 36]
    appliquer_style_en_tetes(ws2, 4, mat_headers, mat_largeurs)

    matieres = [
        ["MAT-2026-0001","Cuir","Cuir de veau","Cuir de veau pleine fleur — tannage végétal",
         "Toscane","Italie","FAIBLE","Tanneria Veneta SpA","Conceria Walpier",
         "LWG Gold","LWG Gold","Certifié",15,28.50,"dm²",
         "Fournisseur historique — Excellence qualité — Tannage végétal traditionnel"],
        ["MAT-2026-0002","Cuir","Cuir d'agneau","Cuir d'agneau nappa — souplesse premium",
         "Castille","Espagne","FAIBLE","Curtidos Benavent","Peletería Munper",
         "LWG Silver","LWG Silver","Certifié",18,34.00,"dm²",
         "Agneau Entrefino — Traçabilité ferme-tannerie complète"],
        ["MAT-2026-0003","Cuir exotique","Python réticulé","Peau python réticulé — finition brillante",
         "Java","Indonésie","MOYEN","Exotic Skins Malaysia","Python Farm Thailand",
         "CITES Annexe II","CITES Annexe II","Certifié",42,185.00,"peau",
         "Convention CITES — Élevage déclaré — Quotas annuels contrôlés"],
        ["MAT-2026-0004","Cuir exotique","Crocodile du Nil","Peau crocodile Niloticus — flancs",
         "Louisiane","USA","FAIBLE","Louisiana Alligator Farm","Heng Long Leather",
         "CITES + SA8000","CITES","Partiel",38,420.00,"cm largeur",
         "⚠️ SA8000 en cours de renouvellement — Audit planifié Q2 2026"],
        ["MAT-2026-0005","Or","Or 18 carats","Or 18K 750‰ — Certifié responsable",
         "Suisse (raffinage)","Suisse","TRÈS FAIBLE","Valcambi SA","PAMP SA",
         "LBMA + RJC","LBMA + RJC","Certifié",12,58200.00,"kg",
         "Chaîne de garde LBMA — Mine → Raffinerie → Atelier entièrement documentée"],
        ["MAT-2026-0006","Or","Or 24 carats","Or 24K 999‰ — Lingots",
         "Suisse (raffinage)","Suisse","TRÈS FAIBLE","Metalor Technologies","Argor-Heraeus",
         "LBMA","LBMA","Certifié",10,77500.00,"kg",
         "Qualité monétaire — Usage haute joaillerie"],
        ["MAT-2026-0007","Pierres précieuses","Diamant","Diamant naturel — taille brillant",
         "Botswana","Botswana","FAIBLE","De Beers Group","Alrosa (⚠️ sanctions)",
         "Kimberley Process + RJC","Kimberley Process","Partiel",28,12500.00,"carat",
         "⚠️ Alrosa (Russie) exclu — Sanctions UE actives — De Beers fournisseur unique"],
        ["MAT-2026-0008","Pierres précieuses","Rubis","Rubis naturel — qualité Pigeon Blood",
         "Mogok Valley","Myanmar","CRITIQUE","N/A — Sourcing suspendu","Gemfields (Mozambique)",
         "RJC + Origine vérifiée","Aucune","Non conforme",92,8500.00,"carat",
         "🔴 SUSPENDU — Sanctions UE/US Myanmar — Redirection vers Mozambique Gemfields"],
        ["MAT-2026-0009","Pierres précieuses","Saphir","Saphir naturel — bleu Ceylan",
         "Ratnapura","Sri Lanka","MOYEN","Ceylon Sapphire Corp","Gemological Lanka",
         "RJC","RJC","Certifié",30,6200.00,"carat",
         "Extraction artisanale — Programme de soutien communautaire en place"],
        ["MAT-2026-0010","Pierres précieuses","Émeraude","Émeraude naturelle — qualité Muzo",
         "Boyacá","Colombie","MOYEN","Muzo Emerald Colombia","Fura Gems",
         "RJC","RJC","Certifié",35,7800.00,"carat",
         "Mine Muzo — Programme développement communautaire — Traçabilité mine-taille"],
        ["MAT-2026-0011","Textile","Cachemire","Cachemire Hircus — Grade A",
         "Mongolie intérieure","Mongolie","MOYEN","Gobi Cashmere LLC","Erdos Group (⚠️ Chine)",
         "GOTS + Origine vérifiée","GOTS","Partiel",45,180.00,"kg",
         "⚠️ Erdos Group (Xinjiang risk) — Privilégier Gobi (Mongolie) — EU Forced Labour Reg."],
        ["MAT-2026-0012","Textile","Soie","Soie mulberry — 22 momme",
         "Côme","Italie","TRÈS FAIBLE","Seteria Bianchi","Taroni SpA",
         "OEKO-TEX 100","OEKO-TEX 100","Certifié",8,95.00,"mètre",
         "Tissage italien traditionnel — Filature Como — Excellence reconnue"],
        ["MAT-2026-0013","Textile","Laine vigogne","Laine vigogne — Ultra premium",
         "Andes","Pérou","MOYEN","Comunidad Lucanas","N/A — Unique source",
         "CITES Annexe I + Commerce équitable","CITES + FairTrade","Certifié",40,4500.00,"kg",
         "Espèce protégée CITES I — Tonte vivante uniquement — Volume très limité"],
        ["MAT-2026-0014","Perles","Perle Tahiti","Perle de culture Tahiti — 12-15mm — AAA",
         "Polynésie française","France","TRÈS FAIBLE","Robert Wan Pearl","Kamoka Pearl Farm",
         "Label Perle de Tahiti","Label officiel","Certifié",6,850.00,"pièce",
         "Perle noire naturelle — Ferme perlière certifiée — Développement durable"],
        ["MAT-2026-0015","Bois","Ébène de Macassar","Ébène de Macassar — Placage",
         "Sulawesi","Indonésie","ÉLEVÉ","Indonesian Timber Corp","N/A",
         "FSC + EUDR","FSC","Partiel",55,2200.00,"m³",
         "⚠️ EUDR applicable 30/12/2025 — Géolocalisation parcelles requise — En cours"],
        ["MAT-2026-0016","Nacre","Nacre blanche","Nacre white mother-of-pearl",
         "Australie","Australie","FAIBLE","Paspaley Pearling Co","Atlas Pacific",
         "MSC","MSC","Certifié",16,45.00,"pièce",
         "Pêche durable — Programme de gestion marine certifié"],
        ["MAT-2026-0017","Platine","Platine 950","Platine 950‰ — Joaillerie",
         "Bushveld Complex","Afrique du Sud","MOYEN","Anglo American Platinum","Impala Platinum",
         "LBMA + RMI","LBMA + RMI","Certifié",32,31200.00,"kg",
         "Minerai responsable — RMI compliant — Audit annuel site minier"],
        ["MAT-2026-0018","Cuir","Cuir de chèvre","Chèvre Mysore — Finition grainée",
         "Karnataka","Inde","MOYEN","Tata International Leather","Rahman Industries",
         "LWG + SA8000","LWG","Partiel",38,22.00,"dm²",
         "⚠️ SA8000 en attente — Conditions de travail sous surveillance — Audit Q3"],
    ]

    NB_MAT = len(matieres)
    for r, row_data in enumerate(matieres, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDURE_FINE
            if c == 14 and isinstance(val, float):
                cell.number_format = '#,##0.00 "€"'

    appliquer_style_donnees(ws2, 5, 4 + NB_MAT, MAX_COL_MAT)
    ajouter_mfc_statut(ws2, "L", 5, 4 + NB_MAT)
    ajouter_mfc_zone_risque(ws2, "G", 5, 4 + NB_MAT)

    ws2.conditional_formatting.add(f'M5:M{4+NB_MAT}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color=ROUGE_ALERTE))

    # Ligne synthèse
    SR = 4 + NB_MAT + 1
    ws2.cell(row=SR, column=1, value="SYNTHÈSE MATIÈRES")
    ligne_synthese(ws2, SR, MAX_COL_MAT, [
        (2, "Nb matières :",      None,    "FFFFFF"),
        (3, f'=COUNTA(A5:A{4+NB_MAT})', None, OR_LUXE),
        (6, "Certifiées :",       None,    "FFFFFF"),
        (7, f'=COUNTIF(L5:L{4+NB_MAT},"Certifié")', None, OR_LUXE),
        (8, "Taux cert. :",       None,    "FFFFFF"),
        (9, f'=IFERROR(COUNTIF(L5:L{4+NB_MAT},"Certifié")/COUNTA(L5:L{4+NB_MAT}),0)', "0.0%", OR_LUXE),
        (10,"Non conformes :",    None,    "FFFFFF"),
        (11,f'=COUNTIF(L5:L{4+NB_MAT},"Non conforme")', None, "FF0000"),
        (12,"Score risque moy :", None,    "FFFFFF"),
        (13,f'=IFERROR(AVERAGE(M5:M{4+NB_MAT}),0)', "0.0", OR_LUXE),
    ])

    # Validations données
    dv1 = DataValidation(type="list", formula1='"CRITIQUE,ÉLEVÉ,MOYEN,FAIBLE,TRÈS FAIBLE"', allow_blank=False)
    ws2.add_data_validation(dv1); dv1.add("G5:G5000")

    dv2 = DataValidation(type="list", formula1='"Certifié,Partiel,En cours,Non conforme,Expiré,N/A"', allow_blank=False)
    ws2.add_data_validation(dv2); dv2.add("L5:L5000")

    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:P{4+NB_MAT}"
    print("  ✅ Onglet 2_MATIERES créé")


    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 3 : FOURNISSEURS
    # ═══════════════════════════════════════════════════════════════════

    ws3 = wb.create_sheet("3_FOURNISSEURS")
    ws3.sheet_properties.tabColor = OR_LUXE

    MAX_COL_FRN = 18
    creer_titre_principal(
        ws3,
        "🏭  BASE FOURNISSEURS — AGENT LUXETRACEABILITY",
        "Scoring risque automatisé — Devoir de Vigilance — Conformité CSRD/CSDDD — Avril 2026",
        MAX_COL_FRN
    )
    appliquer_style_titre_section(ws3, 3, MAX_COL_FRN, "SECTION A — REGISTRE DES FOURNISSEURS ACTIFS", BORDEAUX)

    fourn_headers = [
        "ID_Fournisseur","Raison sociale","Pays","Ville / Région",
        "Tier","Matières fournies","Certifications détenues",
        "Statut certification","Date dernier audit","Résultat audit",
        "Statut fournisseur","Score risque (/100)","Catégorie risque",
        "Contact principal","Email","Volume annuel (€)",
        "% du total achats","Notes / Actions"
    ]
    fourn_largeurs = [16,28,12,18,8,26,28,16,16,22,14,16,14,22,30,20,14,38]
    appliquer_style_en_tetes(ws3, 4, fourn_headers, fourn_largeurs)

    fournisseurs = [
        ["FRN-2026-0001","Tanneria Veneta SpA","Italie","Arzignano, Vénétie",
         "Tier 1","Cuir de veau pleine fleur","LWG Gold + ISO 14001",
         "Certifié","2026-02-15","Conforme","Actif",12,"FAIBLE",
         "Marco Bianchi","m.bianchi@tanneriaveneta.it",2800000,None,
         "Partenaire premium — 12 ans de collaboration — Excellence constante"],
        ["FRN-2026-0002","Conceria Walpier","Italie","Santa Croce, Toscane",
         "Tier 1","Cuir de veau — tannage végétal","LWG Gold + Pelle Conciata al Vegetale",
         "Certifié","2026-01-20","Conforme","Actif",15,"FAIBLE",
         "Laura Rossi","l.rossi@walpier.it",1950000,None,
         "Tannage végétal Toscane — Savoir-faire ancestral — Backup stratégique"],
        ["FRN-2026-0003","Curtidos Benavent","Espagne","Igualada, Catalogne",
         "Tier 1","Cuir d'agneau nappa","LWG Silver + ISO 9001",
         "Certifié","2025-11-10","Conforme","Actif",20,"FAIBLE",
         "Carlos Benavent","c.benavent@curtidos.es",1200000,None,
         "Spécialiste agneau — Flexibilité couleurs — Livraison J+5"],
        ["FRN-2026-0004","Exotic Skins Malaysia","Malaisie","Penang",
         "Tier 1","Python réticulé","CITES Annexe II + ISO 14001",
         "Certifié","2025-09-22","Conforme (obs. mineures)","Actif",40,"MOYEN",
         "Ahmad Razak","a.razak@exoticskins.my",890000,None,
         "⚠️ 2 observations mineures — Plan d'action Q2 — CITES strict"],
        ["FRN-2026-0005","Louisiana Alligator Farm","USA","Lafayette, Louisiane",
         "Tier 1","Crocodile / Alligator","CITES",
         "Partiel","2025-12-05","Conditionnel","Actif",38,"MOYEN",
         "James Thibodeaux","j.thibodeaux@lafarm.com",1650000,None,
         "⚠️ SA8000 manquant — Audit social programmé — CITES conforme"],
        ["FRN-2026-0006","Valcambi SA","Suisse","Balerna, Tessin",
         "Tier 1","Or 18K / 24K — lingots et granules","LBMA Good Delivery + RJC CoC",
         "Certifié","2026-03-01","Conforme","Actif",8,"TRÈS FAIBLE",
         "Stefan Mueller","s.mueller@valcambi.ch",5200000,None,
         "Leader mondial raffinage responsable — Chaîne de garde complète"],
        ["FRN-2026-0007","Metalor Technologies","Suisse","Marin-Epagnier, Neuchâtel",
         "Tier 1","Or 24K — Lingots","LBMA Good Delivery",
         "Certifié","2026-01-15","Conforme","Actif",10,"TRÈS FAIBLE",
         "Philippe Durand","p.durand@metalor.com",1800000,None,
         "Raffineur historique — Backup or — Qualité constante"],
        ["FRN-2026-0008","De Beers Group","Botswana / UK","Gaborone / Londres",
         "Tier 1","Diamants naturels bruts et taillés","Kimberley Process + RJC + GIA",
         "Certifié","2026-02-28","Conforme","Actif",22,"FAIBLE",
         "Sarah Mitchell","s.mitchell@debeersgroup.com",3100000,None,
         "Sightholders programme — Blockchain traçabilité — Fournisseur unique post-sanctions"],
        ["FRN-2026-0009","Gemfields PLC","Mozambique / UK","Montepuez / Londres",
         "Tier 1","Rubis, Émeraudes","RJC + Fairmined",
         "Certifié","2025-10-18","Conforme","Actif",30,"MOYEN",
         "Isabella Torres","i.torres@gemfields.com",980000,None,
         "Alternative Myanmar pour rubis — Enchères traçables — Programme communautaire"],
        ["FRN-2026-0010","Gobi Cashmere LLC","Mongolie","Oulan-Bator",
         "Tier 1","Cachemire Grade A","GOTS + SFA",
         "Certifié","2025-08-30","Conforme","Actif",35,"MOYEN",
         "Bayasgalan Dorj","b.dorj@gobicashmere.mn",720000,None,
         "Alternative Chine/Xinjiang — Traçabilité berger-filature — Volume limité"],
        ["FRN-2026-0011","Seteria Bianchi","Italie","Côme, Lombardie",
         "Tier 1","Soie mulberry — tissage premium","OEKO-TEX 100 + GOTS",
         "Certifié","2026-01-10","Conforme","Actif",8,"TRÈS FAIBLE",
         "Giulia Marchetti","g.marchetti@seteriabianchi.it",540000,None,
         "Tissage artisanal Como — Excellence — Collaboration exclusive"],
        ["FRN-2026-0012","Anglo American Platinum","Afrique du Sud","Johannesburg",
         "Tier 1","Platine 950‰","LBMA + RMI + ISO 45001",
         "Certifié","2025-12-20","Conforme","Actif",28,"MOYEN",
         "Thabo Molefe","t.molefe@angloamerican.com",1100000,None,
         "Mine Mogalakwena — Développement durable — Audit RMI annuel"],
        ["FRN-2026-0013","Erdos Group","Chine","Ordos, Mongolie intérieure",
         "Tier 2","Cachemire brut","ISO 9001",
         "Non conforme","2025-03-15","Non conforme","Suspendu",85,"CRITIQUE",
         "Wang Li","w.li@erdos.com",0,None,
         "🔴 SUSPENDU — Risque Xinjiang — EU Forced Labour Reg. — Pas de transparence Tier 3"],
        ["FRN-2026-0014","Alrosa PJSC","Russie","Mirny, Yakoutie",
         "Tier 1","Diamants bruts","Kimberley Process (suspendu)",
         "Expiré","2022-02-24","N/A","Suspendu",98,"CRITIQUE",
         "N/A","N/A",0,None,
         "🔴 SUSPENDU DÉFINITIF — Sanctions UE depuis 2022 — Aucun approvisionnement"],
        ["FRN-2026-0015","Robert Wan Pearl","France","Papeete, Polynésie française",
         "Tier 1","Perles de Tahiti — AAA","Label Perle de Tahiti + GIA",
         "Certifié","2026-03-15","Conforme","Actif",6,"TRÈS FAIBLE",
         "Heimana Wan","h.wan@robertwan.com",480000,None,
         "N°1 mondial perle de Tahiti — Ferme propre — Développement durable"],
    ]

    NB_FRN = len(fournisseurs)
    for r, row_data in enumerate(fournisseurs, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDURE_FINE
            if c == 16 and val is not None:
                cell.number_format = '#,##0 "€"'

    # Formules automatiques colonnes M (catégorie risque) et Q (% achats)
    for r in range(5, 5 + NB_FRN):
        ws3.cell(row=r, column=13).value = (
            f'=IF(L{r}>=75,"CRITIQUE",IF(L{r}>=50,"ÉLEVÉ",IF(L{r}>=30,"MOYEN",IF(L{r}>=15,"FAIBLE","TRÈS FAIBLE"))))'
        )
        ws3.cell(row=r, column=13).font = Font(name="Calibri", size=9, bold=True)
        ws3.cell(row=r, column=13).alignment = ALIGN_CENTER
        ws3.cell(row=r, column=13).border = BORDURE_FINE

        ws3.cell(row=r, column=17).value = f'=IFERROR(P{r}/SUM(P$5:P${4+NB_FRN}),0)'
        ws3.cell(row=r, column=17).number_format = '0.0%'
        ws3.cell(row=r, column=17).alignment = ALIGN_CENTER
        ws3.cell(row=r, column=17).border = BORDURE_FINE

    appliquer_style_donnees(ws3, 5, 4 + NB_FRN, MAX_COL_FRN)
    ajouter_mfc_statut(ws3, "H", 5, 4 + NB_FRN)
    ajouter_mfc_statut(ws3, "K", 5, 4 + NB_FRN)
    ajouter_mfc_zone_risque(ws3, "M", 5, 4 + NB_FRN)

    ws3.conditional_formatting.add(f'L5:L{4+NB_FRN}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color=ROUGE_ALERTE))

    SR3 = 4 + NB_FRN + 1
    ws3.cell(row=SR3, column=1, value="SYNTHÈSE FOURNISSEURS")
    ligne_synthese(ws3, SR3, MAX_COL_FRN, [
        (2,  "Total :",           None,    "FFFFFF"),
        (3,  f'=COUNTA(A5:A{4+NB_FRN})', None, OR_LUXE),
        (4,  "Actifs :",          None,    "FFFFFF"),
        (5,  f'=COUNTIF(K5:K{4+NB_FRN},"Actif")', None, OR_LUXE),
        (6,  "Certifiés :",       None,    "FFFFFF"),
        (7,  f'=COUNTIF(H5:H{4+NB_FRN},"Certifié")', None, OR_LUXE),
        (8,  "Suspendus :",       None,    "FFFFFF"),
        (9,  f'=COUNTIF(K5:K{4+NB_FRN},"Suspendu")', None, "FF0000"),
        (10, "Taux cert. :",      None,    "FFFFFF"),
        (11, f'=IFERROR(COUNTIF(H5:H{4+NB_FRN},"Certifié")/COUNTA(H5:H{4+NB_FRN}),0)', "0.0%", OR_LUXE),
        (15, "Volume total :",    None,    "FFFFFF"),
        (16, f'=SUM(P5:P{4+NB_FRN})', '#,##0 "€"', OR_LUXE),
        (17, "Score moy :",       None,    "FFFFFF"),
        (18, f'=IFERROR(AVERAGE(L5:L{4+NB_FRN}),0)', "0.0", OR_LUXE),
    ])

    dv_tier = DataValidation(type="list", formula1='"Tier 1,Tier 2,Tier 3,Tier 4"')
    ws3.add_data_validation(dv_tier); dv_tier.add("E5:E5000")

    dv_sf = DataValidation(type="list", formula1='"Actif,Suspendu,En évaluation,Inactif,Blacklisté"')
    ws3.add_data_validation(dv_sf); dv_sf.add("K5:K5000")

    dv_ra = DataValidation(type="list", formula1='"Conforme,Conforme (obs. mineures),Conditionnel,Non conforme,N/A"')
    ws3.add_data_validation(dv_ra); dv_ra.add("J5:J5000")

    ws3.freeze_panes = "A5"
    ws3.auto_filter.ref = f"A4:R{4+NB_FRN}"
    print("  ✅ Onglet 3_FOURNISSEURS créé")


    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 4 : TRACABILITE
    # ═══════════════════════════════════════════════════════════════════

    ws4 = wb.create_sheet("4_TRACABILITE")
    ws4.sheet_properties.tabColor = VERT_VALIDATION

    MAX_COL_TRC = 18
    creer_titre_principal(
        ws4,
        "🔗  CHAÎNE DE TRAÇABILITÉ — AGENT LUXETRACEABILITY",
        "Documentation lot par lot — Origine → Transformation → Produit fini — Devoir de Vigilance — Avril 2026",
        MAX_COL_TRC
    )
    appliquer_style_titre_section(ws4, 3, MAX_COL_TRC, "SECTION A — REGISTRE DE TRAÇABILITÉ PAR LOT", BORDEAUX)

    trace_headers = [
        "ID_Lot","Date réception","Matière","Sous-catégorie",
        "Fournisseur","ID_Fournisseur","Pays d'origine",
        "Région / Mine / Ferme","N° certificat origine",
        "Étape chaîne actuelle","Point de contrôle",
        "Statut traçabilité","Score complétude (%)",
        "Lien document preuve","N° série interne",
        "Destination (Maison / Atelier)","Date dernière MAJ",
        "Observations"
    ]
    trace_largeurs = [14,14,16,18,26,16,14,26,22,22,20,16,16,24,18,28,16,38]
    appliquer_style_en_tetes(ws4, 4, trace_headers, trace_largeurs)

    lots_trace = [
        ["LOT-2026-0001","2026-03-28","Cuir de veau","Pleine fleur tannage végétal",
         "Tanneria Veneta SpA","FRN-2026-0001","Italie",
         "Arzignano — Élevage Vénétie N°IT-VE-2847","CERT-IT-LWG-2026-4421",
         "Atelier coupe","Contrôle qualité post-tannage",
         "Complet",100,"DOC/2026/CUI/001.pdf","SER-CUI-2026-00847",
         "Atelier Maroquinerie — Paris 3e","2026-04-10",
         "Traçabilité complète : ferme → abattoir → tannerie → atelier — Lot premium"],
        ["LOT-2026-0002","2026-03-15","Cuir d'agneau","Nappa souple",
         "Curtidos Benavent","FRN-2026-0003","Espagne",
         "Igualada — Élevage Castille N°ES-CL-1205","CERT-ES-LWG-2026-0892",
         "Tannerie — finition","Contrôle colorimétrique",
         "Complet",98,"DOC/2026/CUI/002.pdf","SER-AGN-2026-00231",
         "Atelier Prêt-à-Porter — Milan","2026-04-08",
         "Lot agneau Entrefino — Couleur caramel — Finition en cours"],
        ["LOT-2026-0003","2026-02-20","Python réticulé","Brillant grade A",
         "Exotic Skins Malaysia","FRN-2026-0004","Indonésie",
         "Java — Ferme élevage MY-PEN-0034","CITES-MY-2026-PY-00567",
         "Transit douanes UE","Vérification permis CITES",
         "Complet",100,"DOC/2026/EXO/001.pdf","SER-PYT-2026-00089",
         "Atelier Maroquinerie — Paris 3e","2026-04-05",
         "CITES Annexe II conforme — Permis export + import validés — Quarantaine OK"],
        ["LOT-2026-0004","2026-01-10","Or 18 carats","750‰ — Granules",
         "Valcambi SA","FRN-2026-0006","Suisse",
         "Raffinage Balerna — Mine source: Kibali (RDC via LBMA CoC)","LBMA-COC-2026-VAL-03421",
         "Atelier joaillerie","Analyse pureté + poids",
         "Complet",100,"DOC/2026/OR/001.pdf","SER-OR18-2026-00156",
         "Atelier Haute Joaillerie — Place Vendôme","2026-04-12",
         "Chaîne de garde LBMA complète — Mine Kibali → Raffinerie Valcambi → Atelier"],
        ["LOT-2026-0005","2026-03-05","Diamant","Taille brillant — 1.5ct — D/IF",
         "De Beers Group","FRN-2026-0008","Botswana",
         "Mine Jwaneng — Lot Sightholder #BW-JW-2026-M03","KP-BW-2026-08934",
         "Taille — Anvers","Certification GIA en cours",
         "En cours",85,"DOC/2026/DIA/001.pdf","SER-DIA-2026-00034",
         "Atelier Haute Joaillerie — Place Vendôme","2026-04-11",
         "⚠️ Certification GIA attendue avant 25/04 — Traçabilité mine-taille 85% documentée"],
        ["LOT-2026-0006","2026-02-28","Rubis","Qualité Pigeon Blood",
         "Gemfields PLC","FRN-2026-0009","Mozambique",
         "Mine Montepuez — Lot enchère #MZ-MP-2026-FEB-12","RJC-MZ-2026-GEM-01289",
         "Taille — Jaipur (sous-traitant)","Vérification conditions travail Tier 2",
         "Partiel",72,"DOC/2026/PIE/001.pdf","SER-RUB-2026-00012",
         "Atelier Haute Joaillerie — Place Vendôme","2026-04-09",
         "⚠️ Gap Tier 2 : atelier Jaipur — Audit social SA8000 planifié 05/2026"],
        ["LOT-2026-0007","2026-04-01","Cachemire","Grade A — Naturel",
         "Gobi Cashmere LLC","FRN-2026-0010","Mongolie",
         "Coopérative Arkhangai — Bergers nomades certifiés SFA","SFA-MN-2026-GOBI-0234",
         "Filature","Contrôle finesse fibre (< 15.5 microns)",
         "Complet",95,"DOC/2026/TEX/001.pdf","SER-CAS-2026-00078",
         "Atelier Tissage — Biella, Italie","2026-04-10",
         "Traçabilité berger-filature documentée — SFA certifié — Alternative Xinjiang validée"],
        ["LOT-2026-0008","2026-03-20","Crocodile du Nil","Flancs — Grade 1",
         "Louisiana Alligator Farm","FRN-2026-0005","USA",
         "Ferme Lafayette — US-LA-AF-0012","CITES-US-2026-CR-01456",
         "Tannerie (sous-traitant Singapour)","Vérification CITES re-export",
         "Partiel",78,"DOC/2026/EXO/002.pdf","SER-CRO-2026-00023",
         "Atelier Maroquinerie — Paris 3e","2026-04-07",
         "⚠️ Permis re-export Singapour en attente — SA8000 ferme non complété"],
        ["LOT-2026-0009","2026-01-25","Soie mulberry","22 momme — Blanc naturel",
         "Seteria Bianchi","FRN-2026-0011","Italie",
         "Tissage Côme — Filature Como-Silk IT-CO-SB-001","OEKO-IT-2026-SB-00123",
         "Atelier confection","Contrôle tissage terminé",
         "Complet",100,"DOC/2026/TEX/002.pdf","SER-SOI-2026-00091",
         "Atelier Prêt-à-Porter — Paris","2026-04-02",
         "Filière 100% italienne — Traçabilité bombyx → filature → tissage → confection"],
        ["LOT-2026-0010","2026-03-10","Perle Tahiti","13mm — AAA — Paon",
         "Robert Wan Pearl","FRN-2026-0015","France",
         "Ferme perlière Fakarava — PF-FAK-RW-001","LABEL-PF-2026-RW-00456",
         "Appairage — Atelier","Contrôle lustre + nacre",
         "Complet",100,"DOC/2026/PER/001.pdf","SER-PRL-2026-00015",
         "Atelier Haute Joaillerie — Place Vendôme","2026-04-06",
         "Label Perle de Tahiti authentique — Ferme propre Robert Wan — Éco-responsable"],
        ["LOT-2026-0011","2026-04-05","Platine 950","950‰ — Joaillerie",
         "Anglo American Platinum","FRN-2026-0012","Afrique du Sud",
         "Mine Mogalakwena — ZA-LP-AAP-001","LBMA-ZA-2026-AAP-00789",
         "Raffinage","Certification LBMA en cours",
         "En cours",80,"DOC/2026/PLA/001.pdf","SER-PLA-2026-00007",
         "Atelier Haute Joaillerie — Place Vendôme","2026-04-12",
         "⚠️ Certification LBMA lot en cours — Analyse pureté attendue — RMI conforme"],
        ["LOT-2026-0012","2026-04-08","Ébène de Macassar","Placage — Grade select",
         "Indonesian Timber Corp","FRN-2026-EBN","Indonésie",
         "Sulawesi — Parcelle géolocalisée ID-SW-ITC-2026-P07","FSC-ID-2026-ITC-00234",
         "Transit maritime","Vérification EUDR — Géolocalisation",
         "Partiel",65,"DOC/2026/BOI/001.pdf","SER-EBN-2026-00003",
         "Atelier Objets — Faubourg Saint-Honoré","2026-04-12",
         "⚠️ EUDR : géolocalisation parcelle fournie — Vérification déforestation en cours — DDS requis"],
    ]

    NB_TRC = len(lots_trace)
    for r, row_data in enumerate(lots_trace, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDURE_FINE
            if c == 13 and isinstance(val, int):
                cell.number_format = '0"%"'

    appliquer_style_donnees(ws4, 5, 4 + NB_TRC, MAX_COL_TRC)
    ajouter_mfc_statut(ws4, "L", 5, 4 + NB_TRC)

    ws4.conditional_formatting.add(f'M5:M{4+NB_TRC}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color=VERT_VALIDATION))

    # Mise en forme conditionnelle score complétude par couleur
    for seuil, fond, txt in [(90, FOND_FAIBLE, "006100"), (70, FOND_MOYEN, "9C6500"), (0, FOND_CRITIQUE, "9C0006")]:
        op = 'greaterThanOrEqual' if seuil > 0 else 'lessThan'
        val_seuil = str(seuil) if seuil > 0 else "70"
        ws4.conditional_formatting.add(f'M5:M{4+NB_TRC}',
            CellIsRule(operator='greaterThanOrEqual' if seuil >= 70 else 'lessThan',
                       formula=[str(seuil if seuil >= 70 else 70)],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=txt)))

    SR4 = 4 + NB_TRC + 1
    ws4.cell(row=SR4, column=1, value="SYNTHÈSE TRAÇABILITÉ")
    ligne_synthese(ws4, SR4, MAX_COL_TRC, [
        (2,  "Total lots :",    None,   "FFFFFF"),
        (3,  f'=COUNTA(A5:A{4+NB_TRC})', None, OR_LUXE),
        (4,  "Complets :",      None,   "FFFFFF"),
        (5,  f'=COUNTIF(L5:L{4+NB_TRC},"Complet")', None, "006100"),
        (6,  "En cours :",      None,   "FFFFFF"),
        (7,  f'=COUNTIF(L5:L{4+NB_TRC},"En cours")', None, "9C6500"),
        (8,  "Partiels :",      None,   "FFFFFF"),
        (9,  f'=COUNTIF(L5:L{4+NB_TRC},"Partiel")', None, "9C6500"),
        (10, "Incomplets :",    None,   "FFFFFF"),
        (11, f'=COUNTIF(L5:L{4+NB_TRC},"Incomplet")', None, "9C0006"),
        (12, "Score moy :",     None,   "FFFFFF"),
        (13, f'=IFERROR(AVERAGE(M5:M{4+NB_TRC}),0)', "0.0\"%\"", OR_LUXE),
    ])

    dv_statut_trc = DataValidation(
        type="list",
        formula1='"Complet,En cours,Partiel,Incomplet,Suspendu"'
    )
    ws4.add_data_validation(dv_statut_trc)
    dv_statut_trc.add("L5:L5000")

    ws4.freeze_panes = "A5"
    ws4.auto_filter.ref = f"A4:R{4+NB_TRC}"
    print("  ✅ Onglet 4_TRACABILITE créé")


    # ═══════════════════════════════════════════════════════════════════
    # SAUVEGARDE
    # ═══════════════════════════════════════════════════════════════════

    output_path = "C:/Users/Ludo/Desktop/IA projet entreprises/NEURAL_LuxeTraceability.xlsx"
    wb.save(output_path)
    print(f"\n  💾 Fichier sauvegardé : {output_path}")
    print(f"  📊 Onglets générés : 1_DASHBOARD · 2_MATIERES · 3_FOURNISSEURS · 4_TRACABILITE")
    print(f"  📋 Contenu : {NB_MAT} matières · {NB_FRN} fournisseurs · {NB_TRC} lots tracés\n")
    return output_path


if __name__ == "__main__":
    generer_luxetraceability()
