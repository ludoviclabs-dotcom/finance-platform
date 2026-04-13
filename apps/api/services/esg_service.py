from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models.esg import (
    EsgQcControl,
    EsgScoreSnapshot,
    EsgSnapshotResponse,
    MaterialiteIssue,
    MaterialiteSnapshot,
)
from models.vsme import (
    VsmeCompletudeSnapshot,
    VsmeEnvironSnapshot,
    VsmeGovSnapshot,
    VsmeProfileSnapshot,
    VsmeSnapshotResponse,
    VsmeSocialSnapshot,
)

# ---------------------------------------------------------------------------
# Workbook path (same root as carbon_service)
# ---------------------------------------------------------------------------
DEFAULT_WORKBOOK_ROOT = Path(__file__).parent.parent / "data"


def _get_root() -> Path:
    root = os.environ.get("CARBONCO_WORKBOOK_ROOT")
    return Path(root) if root else DEFAULT_WORKBOOK_ROOT


# ---------------------------------------------------------------------------
# Named range → (sheet, cell) fallbacks for ESG workbook
# ---------------------------------------------------------------------------
ESG_FALLBACKS: dict[str, tuple[str, str]] = {
    # VSME Profile
    "CC_VSME_BP1_Raison_Sociale":   ("VSME_Reporting", "D7"),
    "CC_VSME_BP2_Secteur_NAF":      ("VSME_Reporting", "D8"),
    "CC_VSME_BP3_ETP":              ("VSME_Reporting", "D9"),
    "CC_VSME_BP4_CA_Net":           ("VSME_Reporting", "D10"),
    "CC_VSME_BP5_Annee":            ("VSME_Reporting", "D11"),
    "CC_VSME_BP6_Pays":             ("VSME_Reporting", "D12"),
    "CC_VSME_BP7_Perimetre":        ("VSME_Reporting", "D13"),
    # VSME Environnement
    "CC_VSME_E1_Scope1":            ("VSME_Reporting", "D17"),
    "CC_VSME_E2_Scope2_LB":         ("VSME_Reporting", "D18"),
    "CC_VSME_E3_Scope2_MB":         ("VSME_Reporting", "D19"),
    "CC_VSME_E4_Scope3":            ("VSME_Reporting", "D20"),
    "CC_VSME_E5_Total_GES":         ("VSME_Reporting", "D21"),
    "CC_VSME_E6_Intensite_CA":      ("VSME_Reporting", "D22"),
    "CC_VSME_E7_Energie_MWh":       ("VSME_Reporting", "D23"),
    "CC_VSME_E8_Part_ENR":          ("VSME_Reporting", "D24"),
    "CC_VSME_E9_Eau_m3":            ("VSME_Reporting", "D25"),
    "CC_VSME_E10_Dechets_t":        ("VSME_Reporting", "D26"),
    "CC_VSME_E11_Valorisation_Pct": ("VSME_Reporting", "D27"),
    "CC_VSME_E12_Plan_GES":         ("VSME_Reporting", "D28"),
    # VSME Social
    "CC_VSME_S1_Effectif":          ("VSME_Reporting", "D32"),
    "CC_VSME_S2_Pct_CDI":           ("VSME_Reporting", "D33"),
    "CC_VSME_S3_Rotation_Pct":      ("VSME_Reporting", "D34"),
    "CC_VSME_S4_LTIR":              ("VSME_Reporting", "D35"),
    "CC_VSME_S5_Formation_h_ETP":   ("VSME_Reporting", "D36"),
    "CC_VSME_S6_Ecart_Salaire_HF":  ("VSME_Reporting", "D37"),
    "CC_VSME_S7_Pct_Femmes_Mgmt":   ("VSME_Reporting", "D38"),
    "CC_VSME_S8_Diversite":         ("VSME_Reporting", "D39"),
    "CC_VSME_S9_Dialogue_Social":   ("VSME_Reporting", "D40"),
    "CC_VSME_S10_Litiges":          ("VSME_Reporting", "D41"),
    # VSME Gouvernance
    "CC_VSME_G1_Anti_Corruption":   ("VSME_Reporting", "D45"),
    "CC_VSME_G2_Formation_Ethique": ("VSME_Reporting", "D46"),
    "CC_VSME_G3_Whistleblowing":    ("VSME_Reporting", "D47"),
    "CC_VSME_G4_Pct_CA_Independants": ("VSME_Reporting", "D48"),
    "CC_VSME_G5_Protection_Donnees": ("VSME_Reporting", "D49"),
    # Matérialité
    "CC_MAT_E1_Score_Impact":       ("Materialite", "F10"),
    "CC_MAT_E2_Score_Impact":       ("Materialite", "F11"),
    "CC_MAT_E3_Score_Impact":       ("Materialite", "F12"),
    "CC_MAT_E4_Score_Impact":       ("Materialite", "F13"),
    "CC_MAT_E5_Score_Impact":       ("Materialite", "F14"),
    "CC_MAT_S1_Score_Impact":       ("Materialite", "F15"),
    "CC_MAT_S2_Score_Impact":       ("Materialite", "F16"),
    "CC_MAT_S3_Score_Impact":       ("Materialite", "F17"),
    "CC_MAT_S4_Score_Impact":       ("Materialite", "F18"),
    "CC_MAT_G1_Score_Impact":       ("Materialite", "F19"),
    # Synthèse ESG (colonne D = valeurs calculées)
    "CC_ESG_Score_Global":          ("Synthese_ESG", "D10"),
    "CC_ESG_Score_E":               ("Synthese_ESG", "D6"),
    "CC_ESG_Score_S":               ("Synthese_ESG", "D7"),
    "CC_ESG_Score_G":               ("Synthese_ESG", "D8"),
    "CC_ESG_Enjeux_Materiels":      ("Synthese_ESG", "D16"),
}

# ---------------------------------------------------------------------------
# QC controls definition (mirrors Controles_QC_ESG sheet)
# ---------------------------------------------------------------------------
QC_CONTROLS = [
    ("C-E01", "Liaison données renseignée", "Bloquant"),
    ("C-E02", "Score VSME > 50%", "Avertissement"),
    ("C-E03", "Matérialité réalisée", "Bloquant"),
    ("C-E04", "Équilibre E/S/G", "Avertissement"),
    ("C-E05", "GES total renseigné", "Bloquant"),
    ("C-E06", "Social S1 : effectifs", "Avertissement"),
    ("C-E07", "Social S1 : LTIR", "Avertissement"),
    ("C-E08", "Écart salarial H/F", "Avertissement"),
    ("C-E09", "Gouvernance : anti-corruption", "Avertissement"),
    ("C-E10", "Gouvernance : lanceurs d'alerte", "Avertissement"),
    ("C-E11", "Part ENR renseignée", "Info"),
    ("C-E12", "Plan de réduction GES", "Avertissement"),
    ("C-E13", "Score ESG global > 30%", "Bloquant"),
    ("C-E14", "Enjeu E1 climat évalué", "Bloquant"),
    ("C-E15", "Cohérence ETP", "Info"),
]

# ---------------------------------------------------------------------------
# Materiality issue definitions
# ---------------------------------------------------------------------------
MATERIALITE_ISSUES = [
    ("E1", "Changement climatique",       "Environnement", "ESRS E1"),
    ("E2", "Pollution",                   "Environnement", "ESRS E2"),
    ("E3", "Eau et ressources marines",   "Environnement", "ESRS E3"),
    ("E4", "Biodiversité et écosystèmes", "Environnement", "ESRS E4"),
    ("E5", "Économie circulaire",         "Environnement", "ESRS E5"),
    ("S1", "Effectifs propres",           "Social",        "ESRS S1"),
    ("S2", "Travailleurs chaîne valeur",  "Social",        "ESRS S2"),
    ("S3", "Communautés affectées",       "Social",        "ESRS S3"),
    ("S4", "Consommateurs / utilisateurs","Social",        "ESRS S4"),
    ("G1", "Conduite des affaires",       "Gouvernance",   "ESRS G1"),
]


def _read_cell(wb, sheet: str, cell: str) -> Any:
    """Read a single cell value, returning None if sheet/cell missing."""
    try:
        ws = wb[sheet]
        return ws[cell].value
    except Exception:
        return None


def _read_named(wb, name: str) -> Any:
    """Read a CC_* named range; fallback to direct cell lookup."""
    try:
        defn = wb.defined_names.get(name)
        if defn:
            dests = list(defn.destinations)
            if dests:
                sheet_name, coord = dests[0]
                # Strip $ signs
                coord = coord.replace("$", "")
                ws = wb[sheet_name]
                return ws[coord].value
    except Exception:
        pass
    # Fallback
    fb = ESG_FALLBACKS.get(name)
    if fb:
        return _read_cell(wb, fb[0], fb[1])
    return None


def _normalize(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            n = float(s.replace(",", "."))
            return int(n) if n.is_integer() else round(n, 4)
        except Exception:
            return s
    return v


def _is_filled(v: Any) -> bool:
    return v is not None and str(v).strip() not in ("", "0", "0.0")


# ---------------------------------------------------------------------------
# VSME snapshot builder
# ---------------------------------------------------------------------------

def build_vsme_snapshot() -> VsmeSnapshotResponse:
    path = _get_root() / "CarbonCo_ESG_Social.xlsx"
    warnings: list[str] = []

    if not path.exists():
        warnings.append(f"Workbook ESG introuvable : {path}")
        wb = None
    else:
        wb = load_workbook(path, read_only=True, data_only=True)

    def r(name: str) -> Any:
        if wb is None:
            return None
        return _normalize(_read_named(wb, name))

    # Profile
    profile = VsmeProfileSnapshot(
        raisonSociale=r("CC_VSME_BP1_Raison_Sociale"),
        secteurNaf=r("CC_VSME_BP2_Secteur_NAF"),
        etp=r("CC_VSME_BP3_ETP"),
        caNet=r("CC_VSME_BP4_CA_Net"),
        anneeReporting=r("CC_VSME_BP5_Annee"),
        pays=r("CC_VSME_BP6_Pays"),
        perimetre=r("CC_VSME_BP7_Perimetre"),
    )

    # Environnement
    env = VsmeEnvironSnapshot(
        scope1Tco2e=r("CC_VSME_E1_Scope1"),
        scope2LbTco2e=r("CC_VSME_E2_Scope2_LB"),
        scope2MbTco2e=r("CC_VSME_E3_Scope2_MB"),
        scope3Tco2e=r("CC_VSME_E4_Scope3"),
        totalGesTco2e=r("CC_VSME_E5_Total_GES"),
        intensiteCaGes=r("CC_VSME_E6_Intensite_CA"),
        energieMwh=r("CC_VSME_E7_Energie_MWh"),
        partEnrPct=r("CC_VSME_E8_Part_ENR"),
        eauM3=r("CC_VSME_E9_Eau_m3"),
        dechetsTonnes=r("CC_VSME_E10_Dechets_t"),
        valorisationDechetsPct=r("CC_VSME_E11_Valorisation_Pct"),
        planReductionGes=r("CC_VSME_E12_Plan_GES"),
    )

    # Social
    social = VsmeSocialSnapshot(
        effectifTotal=r("CC_VSME_S1_Effectif"),
        pctCdi=r("CC_VSME_S2_Pct_CDI"),
        tauxRotation=r("CC_VSME_S3_Rotation_Pct"),
        ltir=r("CC_VSME_S4_LTIR"),
        formationHEtp=r("CC_VSME_S5_Formation_h_ETP"),
        ecartSalaireHf=r("CC_VSME_S6_Ecart_Salaire_HF"),
        pctFemmesMgmt=r("CC_VSME_S7_Pct_Femmes_Mgmt"),
        diversite=r("CC_VSME_S8_Diversite"),
        dialogueSocial=r("CC_VSME_S9_Dialogue_Social"),
        litigesSociaux=r("CC_VSME_S10_Litiges"),
    )

    # Gouvernance
    gouv = VsmeGovSnapshot(
        antiCorruption=r("CC_VSME_G1_Anti_Corruption"),
        formationEthique=r("CC_VSME_G2_Formation_Ethique"),
        whistleblowing=r("CC_VSME_G3_Whistleblowing"),
        pctCaIndependants=r("CC_VSME_G4_Pct_CA_Independants"),
        protectionDonnees=r("CC_VSME_G5_Protection_Donnees"),
    )

    # Complétude
    all_fields = (
        list(profile.model_dump().values())
        + list(env.model_dump().values())
        + list(social.model_dump().values())
        + list(gouv.model_dump().values())
    )
    total = len(all_fields)
    completed = sum(1 for v in all_fields if _is_filled(v))
    score_pct = round(completed / total * 100, 1) if total else 0.0
    statut = (
        "complet" if score_pct >= 90
        else "avancé" if score_pct >= 60
        else "en cours" if score_pct >= 30
        else "incomplet"
    )

    completude = VsmeCompletudeSnapshot(
        indicateursCompletes=completed,
        totalIndicateurs=total,
        scorePct=score_pct,
        statut=statut,
    )

    if wb:
        wb.close()

    return VsmeSnapshotResponse(
        generatedAt=datetime.now(timezone.utc).isoformat(),
        completude=completude,
        profile=profile,
        environnement=env,
        social=social,
        gouvernance=gouv,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# ESG (materiality + scores + QC) snapshot builder
# ---------------------------------------------------------------------------

def build_esg_snapshot() -> EsgSnapshotResponse:
    path = _get_root() / "CarbonCo_ESG_Social.xlsx"
    warnings: list[str] = []

    if not path.exists():
        warnings.append(f"Workbook ESG introuvable : {path}")
        wb = None
    else:
        wb = load_workbook(path, read_only=True, data_only=True)

    def r(name: str) -> Any:
        if wb is None:
            return None
        return _normalize(_read_named(wb, name))

    # Scores ESG
    scores = EsgScoreSnapshot(
        scoreGlobal=r("CC_ESG_Score_Global"),
        scoreE=r("CC_ESG_Score_E"),
        scoreS=r("CC_ESG_Score_S"),
        scoreG=r("CC_ESG_Score_G"),
        enjeuxMateriels=r("CC_ESG_Enjeux_Materiels"),
    )

    # Matérialité — lecture impact (col F) et probabilité (col G) sur la feuille Materialite
    # Rows: E1=10, E2=11, E3=12, E4=13, E5=14, S1=15, S2=16, S3=17, S4=18, G1=19
    MAT_ROWS = {
        "E1": 10, "E2": 11, "E3": 12, "E4": 13, "E5": 14,
        "S1": 15, "S2": 16, "S3": 17, "S4": 18, "G1": 19,
    }
    mat_impact_map = {
        "E1": "CC_MAT_E1_Score_Impact",
        "E2": "CC_MAT_E2_Score_Impact",
        "E3": "CC_MAT_E3_Score_Impact",
        "E4": "CC_MAT_E4_Score_Impact",
        "E5": "CC_MAT_E5_Score_Impact",
        "S1": "CC_MAT_S1_Score_Impact",
        "S2": "CC_MAT_S2_Score_Impact",
        "S3": "CC_MAT_S3_Score_Impact",
        "S4": "CC_MAT_S4_Score_Impact",
        "G1": "CC_MAT_G1_Score_Impact",
    }
    issues: list[MaterialiteIssue] = []
    for code, label, cat, norme in MATERIALITE_ISSUES:
        score_impact = r(mat_impact_map.get(code, ""))
        # Probabilité : colonne G (pas encore de named range — lecture directe)
        row = MAT_ROWS.get(code)
        score_prob = _normalize(_read_cell(wb, "Materialite", f"G{row}")) if wb and row else None
        # Score total = moyenne impact+probabilité si les deux sont dispo, sinon impact seul
        if isinstance(score_impact, (int, float)) and isinstance(score_prob, (int, float)):
            score_total = round((score_impact + score_prob) / 2, 2)
        elif isinstance(score_impact, (int, float)):
            score_total = float(score_impact)
        else:
            score_total = None
        materiel = (isinstance(score_impact, (int, float)) and score_impact >= 3) if score_impact is not None else None
        issues.append(MaterialiteIssue(
            code=code, label=label, categorie=cat, normeEsrs=norme,
            scoreImpact=score_impact, scoreProbabilite=score_prob,
            scoreImpactTotal=score_total, materiel=materiel,
        ))

    nb_evalue = sum(1 for i in issues if i.scoreImpact is not None)
    nb_materiel = sum(1 for i in issues if i.materiel)
    materialite = MaterialiteSnapshot(
        enjeuxEvalues=nb_evalue,
        enjeuxMateriels=nb_materiel,
        enjeuxNonMateriels=nb_evalue - nb_materiel,
        enjeuxMaterielsE=sum(1 for i in issues if i.materiel and i.categorie == "Environnement"),
        enjeuxMaterielsS=sum(1 for i in issues if i.materiel and i.categorie == "Social"),
        enjeuxMaterielsG=sum(1 for i in issues if i.materiel and i.categorie == "Gouvernance"),
        issues=issues,
    )

    # QC controls — evaluate simple rules
    qc_results: list[EsgQcControl] = []
    for ctrl_id, ctrl_label, criticite in QC_CONTROLS:
        statut = _evaluate_esg_qc(ctrl_id, scores, materialite, wb)
        qc_results.append(EsgQcControl(
            id=ctrl_id, label=ctrl_label, statut=statut, criticite=criticite,
        ))

    if wb:
        wb.close()

    return EsgSnapshotResponse(
        generatedAt=datetime.now(timezone.utc).isoformat(),
        scores=scores,
        materialite=materialite,
        qcControls=qc_results,
        warnings=warnings,
    )


def _evaluate_esg_qc(
    ctrl_id: str,
    scores: EsgScoreSnapshot,
    mat: MaterialiteSnapshot,
    wb: Any,
) -> str:
    """Return OK / WARNING / ERROR for a given QC control."""
    try:
        if ctrl_id == "C-E01":
            # Liaison données: check if CA and ETP are set in Liaison_Donnees
            ca = _normalize(_read_cell(wb, "Liaison_Donnees", "C9")) if wb else None
            etp = _normalize(_read_cell(wb, "Liaison_Donnees", "C10")) if wb else None
            return "OK" if _is_filled(ca) and _is_filled(etp) else "ERROR"
        elif ctrl_id == "C-E02":
            vsme = build_vsme_snapshot()
            return "OK" if vsme.completude.scorePct >= 50 else "WARNING"
        elif ctrl_id == "C-E03":
            return "OK" if mat.enjeuxEvalues >= 5 else "ERROR"
        elif ctrl_id == "C-E04":
            se = scores.scoreE or 0
            ss = scores.scoreS or 0
            sg = scores.scoreG or 0
            return "OK" if min(se, ss, sg) >= 20 else "WARNING"
        elif ctrl_id == "C-E05":
            ges = _normalize(_read_cell(wb, "Liaison_Donnees", "C19")) if wb else None
            return "OK" if _is_filled(ges) else "ERROR"
        elif ctrl_id == "C-E06":
            eff = _normalize(_read_cell(wb, "VSME_Reporting", "D32")) if wb else None
            return "OK" if _is_filled(eff) else "WARNING"
        elif ctrl_id == "C-E13":
            sg = scores.scoreGlobal
            return "OK" if (sg is not None and sg >= 30) else "ERROR"
        elif ctrl_id == "C-E14":
            e1 = next((i for i in mat.issues if i.code == "E1"), None)
            return "OK" if (e1 and e1.scoreImpact is not None) else "ERROR"
        else:
            return "INFO"
    except Exception:
        return "UNKNOWN"
