"""
Build MultiChannelInsur_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'MultiChannelInsur' (branche Assurances — Marketing).

7 onglets : README, Briefs Source, Canal Agent, Canal Courtier,
Canal Direct, Canal Comparateur (DSA/DMA), KPIs Conformité.

3 briefs commerciaux (Auto, MRH, Santé) déclinés sur 4 canaux,
avec mentions légales obligatoires adaptées par canal.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"
BLUE_CANAL = "1D4ED8"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=NEURAL_VIOLET, end_color=NEURAL_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def canal_fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color, end_color=color)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def write_title(ws, text: str, span: int = 7) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 7) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, bold=False, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 7) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 70) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict[int, float]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ────────────────────────────────────────────────────────────────────
# 3 briefs commerciaux sources
# ────────────────────────────────────────────────────────────────────
BRIEFS = [
    {
        "id": "BRIEF-01",
        "produit": "Assurance Auto",
        "cible": "Conducteur 25-35 ans, véhicule récent (< 3 ans), bonus 0.80, cherche meilleur Q/P",
        "message_brut": (
            "Protégez votre véhicule avec notre formule tous risques premium. "
            "Assistance 0 km, véhicule de remplacement, protection du conducteur incluse. "
            "Tarif garanti 12 mois. Souscrivez en ligne en 5 minutes."
        ),
        "enjeux_canal": "Ton expert vs autonomie digitale vs expertise courtier vs transparence comparateur",
        "reglementation": "DDA L.521-2 (identification distributeur), IPID obligatoire avant souscription",
    },
    {
        "id": "BRIEF-02",
        "produit": "Assurance MRH",
        "cible": "Locataire Paris T3, mobilier estimé 35 000 €, besoin couverture dommages + RC locative",
        "message_brut": (
            "Multirisque habitation pour locataires : responsabilité civile, dégâts des eaux, vol, incendie. "
            "Capital mobilier jusqu'à 50 000 €. Garantie valeur à neuf sur 3 ans. "
            "Premier mois offert pour toute souscription avant le 30 juin."
        ),
        "enjeux_canal": "Qualification du besoin réel (agent) vs comparaison panels (courtier) vs offre directe vs ranking comparateur",
        "reglementation": "DDA art. 20 (test d'adéquation), DSA art. 26 (promo temporaire = publicité clairement identifiée)",
    },
    {
        "id": "BRIEF-03",
        "produit": "Complémentaire Santé",
        "cible": "Senior 60-70 ans, forte consommation soins, priorité remboursements dentaire/optique/auditif",
        "message_brut": (
            "Formule Senior Confort : prise en charge à 400% BR dentaire, 600 € optique/an, "
            "aides auditives remboursées jusqu'à 2 000 €. Aucun délai de carence sur hospitalisation. "
            "Devis personnalisé en 2 minutes."
        ),
        "enjeux_canal": "Empathie + pédagogie (agent) vs tableau comparatif (courtier) vs simulateur digital vs algorithme de scoring (comparateur)",
        "reglementation": "DDA art. 20 §5 (IPID santé), ACPR Reco 2024-R-01, résiliation infra-annuelle (loi Lemoine)",
    },
]

# ────────────────────────────────────────────────────────────────────
# Variantes par canal — 3 briefs × 4 canaux
# ────────────────────────────────────────────────────────────────────

# AGENT GÉNÉRAL — ton conseil personnalisé, RDV physique/visio
VARIANTS_AGENT = [
    {
        "brief_id": "BRIEF-01",
        "produit": "Assurance Auto",
        "version": "VA-01",
        "registre": "Conseil personnalisé",
        "accroche": "Votre véhicule mérite plus qu'un tarif — il mérite un conseiller qui connaît votre profil.",
        "corps": (
            "En tant qu'agent agréé [COMPAGNIE], j'analyse votre situation de conducteur pour vous proposer "
            "la formule tous risques vraiment adaptée à votre usage. Assistance 0 km dès la souscription, "
            "véhicule de remplacement sans franchise kilométrique, protection du conducteur avec capital "
            "invalidité permanente de 150 000 €.\n\n"
            "Je vous remets l'IPID et la fiche conseil avant toute signature. Tarif garanti 12 mois, "
            "révision annuelle transparente.\n\n"
            "Prenons 20 minutes pour faire le point."
        ),
        "mentions_legales": (
            "[Agent général de [COMPAGNIE] — ORIAS n° XX.XXX.XXX — www.orias.fr]\n"
            "Rémunération : commissions versées par [COMPAGNIE]. Conseil fondé sur une analyse de vos besoins."
        ),
        "ton_registre": "Relationnel, pédagogue, rassurant",
        "mots_cles_canal": "conseiller, analyse, adapté, IPID, fiche conseil, ORIAS",
        "score_adaptation": 92,
        "score_conformite_dda": 95,
    },
    {
        "brief_id": "BRIEF-02",
        "produit": "Assurance MRH",
        "version": "VA-02",
        "registre": "Conseil personnalisé",
        "accroche": "Votre appartement parisien, vos responsabilités locatives : construisons la couverture juste.",
        "corps": (
            "Je recueille d'abord vos exigences et besoins : surface, valeur mobilier, usage professionnel "
            "ou non, animaux de compagnie, cave ou parking inclus. Ce recueil est documenté et vous est "
            "remis avant toute proposition.\n\n"
            "Sur cette base, je vous propose une MRH avec capital mobilier 35 000 € (valeur à neuf 3 ans), "
            "RC locative illimitée, dégâts des eaux traités sous 48 h. Je vous présente l'IPID et "
            "l'ensemble des exclusions principales pour qu'il n'y ait aucune surprise en cas de sinistre."
        ),
        "mentions_legales": (
            "[Agent général de [COMPAGNIE] — ORIAS n° XX.XXX.XXX]\n"
            "Rémunération : commissions versées par [COMPAGNIE]. Recueil d'exigences et besoins réalisé "
            "préalablement à toute souscription (DDA art. 20)."
        ),
        "ton_registre": "Empathique, méthodique, transparent sur les limites",
        "mots_cles_canal": "recueil besoins, IPID, exclusions, sinistre, RC locative",
        "score_adaptation": 90,
        "score_conformite_dda": 97,
    },
    {
        "brief_id": "BRIEF-03",
        "produit": "Complémentaire Santé",
        "version": "VA-03",
        "registre": "Conseil personnalisé",
        "accroche": "Vos besoins de santé évoluent avec l'âge — votre mutuelle doit en faire autant.",
        "corps": (
            "Je prends le temps de comprendre votre consommation médicale réelle : fréquence des soins "
            "dentaires, port de lunettes ou lentilles, appareillage auditif. Pas de formule générique : "
            "je calibre la prise en charge sur votre situation.\n\n"
            "Formule Senior Confort adaptée à votre profil : dentaire 400% BR, optique 600 €/an, "
            "aides auditives 2 000 €. Hospitalisation sans délai de carence. Résiliation possible "
            "à tout moment après 1 an (loi Lemoine).\n\n"
            "Je vous remets l'IPID santé et la fiche de conseil personnalisée avant signature."
        ),
        "mentions_legales": (
            "[Agent général de [COMPAGNIE] — ORIAS n° XX.XXX.XXX]\n"
            "Rémunération : commissions versées par [COMPAGNIE].\n"
            "Résiliation infra-annuelle possible après 1 an (loi Lemoine 2022)."
        ),
        "ton_registre": "Empathique, pédagogue, orienté bénéfice santé",
        "mots_cles_canal": "consommation médicale, calibre, IPID santé, loi Lemoine",
        "score_adaptation": 88,
        "score_conformite_dda": 96,
    },
]

# COURTIER — expertise comparative, panels, indépendance
VARIANTS_COURTIER = [
    {
        "brief_id": "BRIEF-01",
        "produit": "Assurance Auto",
        "version": "VC-01",
        "registre": "Expertise comparative",
        "accroche": "Nous avons comparé 18 compagnies pour trouver la meilleure couverture tous risques à votre profil.",
        "corps": (
            "En tant que courtier indépendant, nous ne sommes liés à aucune compagnie. Sur les 18 assureurs "
            "de notre panel, nous avons sélectionné 3 offres tous risques premium correspondant exactement "
            "à votre profil : bonus 0.80, véhicule < 3 ans, usage mixte.\n\n"
            "Comparatif détaillé remis : franchises, assistance 0 km, véhicule de remplacement (durée max, "
            "catégorie), protection conducteur (capital invalidité), options valeur à neuf.\n\n"
            "Nous gérons également le sinistre pour vous, de la déclaration au règlement."
        ),
        "mentions_legales": (
            "[Courtier inscrit ORIAS n° XX.XXX.XXX — www.orias.fr]\n"
            "Rémunération : commissions versées par les compagnies partenaires. Notre panel couvre "
            "18 compagnies sur le marché français. Le présent conseil ne constitue pas une couverture "
            "exhaustive du marché."
        ),
        "ton_registre": "Expert, objectif, orienté comparaison et valeur ajoutée",
        "mots_cles_canal": "panel, indépendant, comparatif, sélection, gestion sinistre",
        "score_adaptation": 91,
        "score_conformite_dda": 93,
    },
    {
        "brief_id": "BRIEF-02",
        "produit": "Assurance MRH",
        "version": "VC-02",
        "registre": "Expertise comparative",
        "accroche": "Notre panel de 22 assureurs, filtré sur vos critères : T3 Paris, mobilier 35 000 €, locataire.",
        "corps": (
            "Sur 22 compagnies analysées, 4 proposent une MRH Paris locataire avec capital mobilier "
            "≥ 35 000 € valeur à neuf 3 ans. Nous les avons comparées sur : franchise dégâts des eaux, "
            "délai d'intervention, plafond vol, couverture RC locative illimitée ou plafonnée.\n\n"
            "Notre sélection : [Compagnie A] pour le meilleur ratio prime/couverture mobilier haut de gamme ; "
            "[Compagnie B] pour la gestion sinistre la plus réactive (NPS sinistre 78/100).\n\n"
            "L'offre promotionnelle (premier mois offert) est disponible chez 2 des 4 assureurs jusqu'au 30 juin — "
            "nous vérifions les conditions générales avant engagement."
        ),
        "mentions_legales": (
            "[Courtier inscrit ORIAS n° XX.XXX.XXX]\n"
            "Rémunération : commissions variables selon compagnie (détail disponible sur demande, DDA art. 19).\n"
            "Panel : 22 assureurs. Ne couvre pas l'intégralité du marché français."
        ),
        "ton_registre": "Analytique, factuel, mise en concurrence visible",
        "mots_cles_canal": "panel, NPS sinistre, comparatif conditions générales, commissions sur demande",
        "score_adaptation": 93,
        "score_conformite_dda": 94,
    },
    {
        "brief_id": "BRIEF-03",
        "produit": "Complémentaire Santé",
        "version": "VC-03",
        "registre": "Expertise comparative",
        "accroche": "Tableau comparatif Senior : 12 mutuelles analysées sur dentaire, optique, auditif.",
        "corps": (
            "Nous avons analysé 12 contrats complémentaire santé senior sur 4 critères prioritaires pour "
            "votre profil : remboursement dentaire (% BR), plafond optique annuel, prise en charge aides "
            "auditives, délai de carence hospitalisation.\n\n"
            "Notre recommandation : 2 contrats se distinguent significativement. [Mutuelle A] : dentaire "
            "400% BR, optique 600 €/an, auditif 2 000 €, 0 délai carence hosp. — prime mensuelle 89 €. "
            "[Mutuelle B] : légèrement moins couvrant en optique mais 15% moins cher, à étudier si budget "
            "prioritaire.\n\n"
            "Résiliation infra-annuelle possible après 1 an (loi Lemoine) — vous n'êtes pas piégé."
        ),
        "mentions_legales": (
            "[Courtier inscrit ORIAS n° XX.XXX.XXX]\n"
            "Rémunération : commissions versées par les mutuelles partenaires (détail sur demande).\n"
            "Panel analysé : 12 contrats. Comparatif arrêté au 01/04/2026 — tarifs susceptibles d'évoluer."
        ),
        "ton_registre": "Factuel, chiffré, comparatif tabulaire, neutre entre compagnies",
        "mots_cles_canal": "tableau comparatif, % BR, plafond, panel 12 contrats, loi Lemoine",
        "score_adaptation": 90,
        "score_conformite_dda": 92,
    },
]

# DIRECT / DIGITAL — self-service, autonomie, parcours en ligne
VARIANTS_DIRECT = [
    {
        "brief_id": "BRIEF-01",
        "produit": "Assurance Auto",
        "version": "VD-01",
        "registre": "Self-service digital",
        "accroche": "Assurez votre voiture en ligne en 5 minutes. Tarif garanti. Zéro papier.",
        "corps": (
            "Renseignez votre profil conducteur et votre véhicule : notre simulateur calcule votre tarif "
            "tous risques en temps réel. Assistance 0 km, véhicule de remplacement, protection conducteur "
            "inclus dans la formule premium.\n\n"
            "• Tarif garanti 12 mois, révision à échéance\n"
            "• Attestation d'assurance téléchargeable immédiatement\n"
            "• Déclaration de sinistre 100% en ligne, suivi en temps réel\n"
            "• Résiliation à tout moment depuis votre espace client\n\n"
            "Avant de finaliser, lisez l'IPID (fiche d'information standardisée) disponible en téléchargement."
        ),
        "mentions_legales": (
            "[COMPAGNIE SA — Entreprise d'assurance agréée ACPR — www.acpr.banque-france.fr]\n"
            "Contrat souscrit directement auprès de COMPAGNIE SA, sans intermédiaire.\n"
            "IPID disponible avant souscription. Tarif indicatif, confirmation après validation du dossier."
        ),
        "ton_registre": "Direct, bénéfice-centré, CTA fort, liste à puces",
        "mots_cles_canal": "en ligne, simulateur, immédiatement, espace client, IPID disponible",
        "score_adaptation": 88,
        "score_conformite_dda": 85,
    },
    {
        "brief_id": "BRIEF-02",
        "produit": "Assurance MRH",
        "version": "VD-02",
        "registre": "Self-service digital",
        "accroche": "Votre MRH Paris en 3 minutes. Premier mois offert jusqu'au 30 juin.",
        "corps": (
            "Renseignez : ville, surface, statut locataire, valeur mobilier estimée — et obtenez votre tarif "
            "personnalisé. Notre outil pré-remplit les garanties clés pour un T3 parisien.\n\n"
            "• RC locative illimitée incluse d'office\n"
            "• Capital mobilier jusqu'à 50 000 € (valeur à neuf 3 ans)\n"
            "• Dégâts des eaux gérés sous 48h via l'appli\n"
            "• Vol : déclaration en ligne + avance sur indemnité sous 5 jours\n\n"
            "⚠ Offre promotionnelle : premier mois offert. Offre soumise à conditions. Consultez les "
            "conditions générales et l'IPID avant toute souscription. Offre valable jusqu'au 30/06/2026."
        ),
        "mentions_legales": (
            "[COMPAGNIE SA — Entreprise d'assurance agréée ACPR]\n"
            "Vente directe sans intermédiaire. IPID et conditions générales disponibles en téléchargement.\n"
            "Offre promotionnelle (1er mois offert) : publicité commerciale soumise à conditions générales — "
            "art. 26 DSA : publicité clairement identifiée."
        ),
        "ton_registre": "Promotionnel, CTA clair, transparence promo obligatoire",
        "mots_cles_canal": "tarif personnalisé, appli, 48h, conditions générales, IPID",
        "score_adaptation": 85,
        "score_conformite_dda": 82,
    },
    {
        "brief_id": "BRIEF-03",
        "produit": "Complémentaire Santé",
        "version": "VD-03",
        "registre": "Self-service digital",
        "accroche": "Devis Senior Confort en 2 minutes. Remboursements dentaire, optique, auditif.",
        "corps": (
            "Renseignez votre âge, votre situation (retraité, TNS, salarié) et vos besoins prioritaires. "
            "Notre simulateur calcule votre niveau de remboursement sur les 5 derniers postes de dépenses "
            "déclarés.\n\n"
            "Formule Senior Confort :\n"
            "• Dentaire : jusqu'à 400% BR\n"
            "• Optique : 600 €/an (verres + monture)\n"
            "• Auditif : jusqu'à 2 000 € sur 4 ans\n"
            "• Hospitalisation : aucun délai de carence\n\n"
            "Résiliation possible à tout moment après 1 an de contrat (loi Lemoine).\n"
            "Avant souscription, téléchargez l'IPID santé disponible sur cette page."
        ),
        "mentions_legales": (
            "[COMPAGNIE SA — Organisme complémentaire agréé — Inscription ACPR]\n"
            "Vente directe. IPID et conditions générales disponibles avant souscription.\n"
            "Résiliation infra-annuelle : loi Lemoine — applicable après 1 an de contrat."
        ),
        "ton_registre": "Pédagogique, liste structurée, simulateur mis en avant",
        "mots_cles_canal": "simulateur, postes de dépenses, loi Lemoine, IPID santé",
        "score_adaptation": 87,
        "score_conformite_dda": 84,
    },
]

# COMPARATEUR — DSA/DMA obligations : rémunération, panel, non-exhaustivité
VARIANTS_COMPARATEUR = [
    {
        "brief_id": "BRIEF-01",
        "produit": "Assurance Auto",
        "version": "VCP-01",
        "registre": "Comparateur DSA/DMA",
        "accroche": "Comparez les meilleures assurances auto tous risques — bonus 0.80 — en 2 minutes.",
        "corps": (
            "Résultats de comparaison pour votre profil : conducteur 26 ans, bonus 0.80, "
            "Renault Mégane 2022.\n\n"
            "Les résultats sont classés par note globale (garanties + tarif + avis clients). "
            "Les offres en tête peuvent être des offres commerciales identifiées par l'étiquette "
            "« Partenaire » — elles sont clairement distinguées des résultats organiques.\n\n"
            "Notre panel couvre 23 assureurs auto. Certains acteurs du marché ne participent pas "
            "à notre comparatif (La Mutuelle Générale, MAIF, etc.).\n\n"
            "Ce comparatif ne constitue pas un conseil en assurance. Lisez l'IPID et les conditions "
            "générales de chaque offre avant toute souscription."
        ),
        "mentions_legales": (
            "ℹ Mentions obligatoires DSA/DMA :\n"
            "Rémunération : [NOM COMPARATEUR] perçoit une commission des assureurs présentés lorsqu'un "
            "contrat est souscrit via notre plateforme. Ce mode de rémunération peut influencer l'ordre "
            "d'affichage des offres.\n"
            "Panel : 23 assureurs sur ~180 membres FFSA + mutuelles. Ne couvre pas l'intégralité du marché.\n"
            "Classement : fondé sur un algorithme combinant tarif (40%), garanties (40%), avis clients (20%). "
            "Critères complets disponibles sur notre page Méthodologie.\n"
            "[Intermédiaire d'assurance inscrit ORIAS n° XX.XXX.XXX]"
        ),
        "dsa_remuneration": "OUI — commission/clic ou/acte déclarée",
        "dsa_panel": "OUI — 23/~180 mentionné explicitement",
        "dsa_non_exhaustivite": "OUI — acteurs absents mentionnés par exemple",
        "dsa_criteres_classement": "OUI — algorithme 40/40/20 décrit",
        "dsa_publicite_identifiee": "OUI — label 'Partenaire' sur offres sponsorisées",
        "score_adaptation": 96,
        "score_conformite_dda": 94,
        "score_dsa": 98,
    },
    {
        "brief_id": "BRIEF-02",
        "produit": "Assurance MRH",
        "version": "VCP-02",
        "registre": "Comparateur DSA/DMA",
        "accroche": "MRH locataire Paris — comparez 19 offres en temps réel.",
        "corps": (
            "Résultats pour : locataire, Paris 75, T3 (65 m²), mobilier estimé 35 000 €.\n\n"
            "Critères de classement utilisés pour cette recherche :\n"
            "1. Couverture mobilier (capital et valeur à neuf)\n"
            "2. Franchise dégâts des eaux\n"
            "3. Tarif annuel\n"
            "4. Note service client (source : avis vérifiés Trustpilot)\n\n"
            "Offre promotionnelle « 1er mois offert » : identifier les offres labellisées « Promotion » — "
            "il s'agit d'une communication publicitaire. Vérifiez les conditions avant souscription.\n\n"
            "IPID et conditions générales de chaque offre accessibles via le bouton 'Voir le contrat'."
        ),
        "mentions_legales": (
            "ℹ Mentions obligatoires DSA/DMA :\n"
            "Rémunération : commission versée par les assureurs partenaires à la souscription. "
            "Tarif comparateur gratuit pour l'utilisateur.\n"
            "Panel : 19 assureurs MRH. Ne couvre pas l'intégralité du marché (ex. : certaines mutuelles "
            "régionales en vente directe uniquement).\n"
            "Offres 'Promotion' : publicité commerciale clairement identifiée (DSA art. 26).\n"
            "[Intermédiaire d'assurance inscrit ORIAS n° XX.XXX.XXX]"
        ),
        "dsa_remuneration": "OUI — commission à la souscription déclarée",
        "dsa_panel": "OUI — 19 assureurs, exclusions mentionnées",
        "dsa_non_exhaustivite": "OUI — mutuelles régionales citées comme exemple",
        "dsa_criteres_classement": "OUI — 4 critères listés et pondérés",
        "dsa_publicite_identifiee": "OUI — label 'Promotion' sur offres commerciales",
        "score_adaptation": 95,
        "score_conformite_dda": 91,
        "score_dsa": 97,
    },
    {
        "brief_id": "BRIEF-03",
        "produit": "Complémentaire Santé",
        "version": "VCP-03",
        "registre": "Comparateur DSA/DMA",
        "accroche": "Complémentaire Santé Senior — comparez les remboursements dentaire, optique, auditif.",
        "corps": (
            "Résultats pour votre profil : retraité(e), 63 ans, priorité dentaire + optique + auditif.\n\n"
            "Notre algorithme de scoring santé évalue chaque contrat sur :\n"
            "• Remboursement dentaire (% du tarif de convention)\n"
            "• Plafond optique annuel (monture + verres)\n"
            "• Prise en charge aides auditives (4 ans)\n"
            "• Délai de carence hospitalisation\n"
            "• Prime mensuelle pour votre tranche d'âge\n\n"
            "ℹ Ce comparatif utilise un système de scoring automatisé (IA). "
            "Les résultats sont indicatifs et ne constituent pas un conseil personnalisé en assurance. "
            "(Mention requise — AI Act art. 50, applicable août 2026)\n\n"
            "IPID santé et tableau de garanties disponibles pour chaque offre."
        ),
        "mentions_legales": (
            "ℹ Mentions obligatoires DSA/DMA + AI Act :\n"
            "Rémunération : commission versée par les organismes partenaires à la souscription.\n"
            "Panel : 15 contrats complémentaire santé senior sur ~90 disponibles en France.\n"
            "Non-exhaustivité : de nombreux contrats collectifs et contrats Madelin ne sont pas comparables "
            "sur cette plateforme.\n"
            "Scoring IA : les classements sont générés par un algorithme automatisé. "
            "Ce système n'est pas un conseiller humain (EU AI Act art. 50).\n"
            "[Intermédiaire d'assurance inscrit ORIAS n° XX.XXX.XXX]"
        ),
        "dsa_remuneration": "OUI — commission à la souscription déclarée",
        "dsa_panel": "OUI — 15/~90, sous-représentation collectifs mentionnée",
        "dsa_non_exhaustivite": "OUI — contrats Madelin + collectifs exclus cités",
        "dsa_criteres_classement": "OUI — 5 critères listés",
        "dsa_publicite_identifiee": "OUI (non applicable ici — pas d'offres sponsorisées)",
        "score_adaptation": 94,
        "score_conformite_dda": 90,
        "score_dsa": 95,
    },
]

# ────────────────────────────────────────────────────────────────────
# Couleurs par canal
# ────────────────────────────────────────────────────────────────────
CANAL_COLORS = {
    "agent": "1E3A5F",      # bleu marine foncé
    "courtier": "1D4E4E",   # vert forêt foncé
    "direct": "7C3AED",     # violet NEURAL
    "comparateur": "92400E", # brun foncé
}

CANAL_LABELS = {
    "agent": "AGENT GÉNÉRAL",
    "courtier": "COURTIER INDÉPENDANT",
    "direct": "DIRECT / DIGITAL",
    "comparateur": "COMPARATEUR (DSA/DMA)",
}


# ────────────────────────────────────────────────────────────────────
# Onglet 0 — README
# ────────────────────────────────────────────────────────────────────
def build_readme(ws) -> None:
    write_title(ws, "NEURAL · MultiChannelInsur — Adaptation discours assurance par canal de distribution")
    write_subtitle(ws, 2, "3 briefs commerciaux (Auto, MRH, Santé) déclinés sur 4 canaux — Conformité DDA + DSA/DMA + AI Act art. 50 (comparateurs)")

    ws.row_dimensions[3].height = 10

    write_section_header(ws, 4, "MISSION DE L'AGENT")
    rows_mission = [
        ("Problème", "Un même message commercial assurance doit être radicalement différent selon qu'il passe par un agent général, un courtier, un canal digital direct ou un comparateur. Les registres linguistiques, les mentions légales obligatoires et les contraintes réglementaires divergent par canal."),
        ("Solution NEURAL", "L'agent MultiChannelInsur reçoit un brief commercial brut et génère automatiquement 4 variantes optimisées — ton, structure, mentions légales — avec score d'adaptation et score de conformité DDA/DSA."),
        ("Canaux couverts", "1. Agent général (DDA conseil personnalisé) | 2. Courtier indépendant (DDA + comparaison panels) | 3. Direct/digital (DDA allégé, IPID obligatoire) | 4. Comparateur (DSA/DMA + AI Act art. 50)"),
        ("Produits démontrés", "Assurance Auto (bonus 0.80, véhicule récent) | MRH locataire Paris T3 | Complémentaire Santé Senior"),
    ]
    for i, (k, v) in enumerate(rows_mission):
        row = 5 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font()
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 50

    ws.row_dimensions[9].height = 10

    write_section_header(ws, 10, "RÉFÉRENTIEL RÉGLEMENTAIRE PAR CANAL")
    reg_headers = ["Canal", "Réglementation principale", "Obligation clé", "Sanction en cas de manquement"]
    for col, h in enumerate(reg_headers, 1):
        cell = ws.cell(row=11, column=col, value=h)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[11].height = 28

    reg_data = [
        ("Agent général", "DDA art. 18-20 + Code ass. L.521-2 à L.521-5", "Identification ORIAS, recueil besoins, IPID avant souscription, fiche conseil", "Amende ACPR jusqu'à 5% CA + interdiction d'exercer"),
        ("Courtier indépendant", "DDA art. 18-20 + Code ass. L.521-2 + art. 19 rémunération", "Déclaration nature de rémunération, taille du panel, indépendance affirmée", "Amende ACPR, retrait agrément, responsabilité civile professionnelle"),
        ("Direct / Digital", "DDA art. 20 §5 + Règlement UE 2017/1469 (IPID) + DSA art. 26 (promos)", "IPID accessible avant clic de souscription, mention agréement ACPR, offres promo identifiées", "DGCCRF jusqu'à 2% CA mondial + injonction de mise en conformité"),
        ("Comparateur", "DDA + DSA art. 22-26 + DMA art. 6-7 + AI Act art. 50 (scoring IA)", "Mention rémunération, taille panel, non-exhaustivité, critères classement, publicité identifiée, disclosure IA", "DMA : jusqu'à 10% CA mondial + désignation gatekeeper"),
    ]
    for i, row_data in enumerate(reg_data):
        row = 12 + i
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = alt_fill()
        ws.row_dimensions[row].height = 55

    ws.row_dimensions[16].height = 10

    write_section_header(ws, 17, "MODE D'EMPLOI DES ONGLETS")
    tabs = [
        ("0_README", "Ce guide — mission, réglementation, mode d'emploi"),
        ("1_Briefs_Source", "3 briefs commerciaux bruts à adapter (Auto, MRH, Santé) + enjeux par canal"),
        ("2_Canal_Agent", "Variantes adaptées canal agent général — ton conseil, ORIAS, recueil besoins, IPID"),
        ("3_Canal_Courtier", "Variantes adaptées canal courtier — panels, comparatif, indépendance, rémunération déclarée"),
        ("4_Canal_Direct", "Variantes adaptées canal digital — self-service, CTA, IPID accessible, ACPR"),
        ("5_Canal_Comparateur", "Variantes comparateur — mentions DSA/DMA obligatoires, scoring IA, non-exhaustivité"),
        ("6_KPIs_Conformite", "Matrice récapitulative : score adaptation canal + score conformité DDA + score DSA"),
    ]
    for i, (tab, desc) in enumerate(tabs):
        row = 18 + i
        ws.cell(row=row, column=1, value=tab).font = body_font(bold=True, color=NEURAL_VIOLET)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value=desc).font = body_font()
        ws.row_dimensions[row].height = 22

    ws.row_dimensions[25].height = 10

    write_section_header(ws, 26, "STATUT & LIMITES")
    limits = [
        ("Version", "v1.0 — Démo prospect assureurs / courtiers / comparateurs"),
        ("Données", "Cas fictifs inspirés du marché — toute ressemblance avec des produits réels est fortuite"),
        ("Usage", "Démo commerciale et pitch produit uniquement. Ne constitue pas un audit réglementaire ou un conseil juridique."),
        ("Limite agent", "L'agent ne remplace pas un avocat spécialisé assurance, un compliance officer ACPR ou un juriste DDA/DSA."),
        ("Mise à jour réglementaire", "AI Act art. 50 applicable août 2026 — les obligations de disclosure IA pour comparateurs sont en cours de précision par l'EDPB."),
    ]
    for i, (k, v) in enumerate(limits):
        row = 27 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value=v).font = body_font(color="475569")
        ws.row_dimensions[row].height = 30

    set_widths(ws, {1: 22, 2: 35, 3: 35, 4: 35, 5: 20, 6: 20, 7: 20})


# ────────────────────────────────────────────────────────────────────
# Onglet 1 — Briefs Source
# ────────────────────────────────────────────────────────────────────
def build_briefs(ws) -> None:
    write_title(ws, "NEURAL · MultiChannelInsur — 1_Briefs_Source : Données d'entrée à adapter par canal")
    write_subtitle(ws, 2, "3 briefs commerciaux bruts — tel qu'un marketing manager les produirait, avant optimisation multi-canal")

    headers = ["ID Brief", "Produit", "Cible client", "Message brut (input agent)", "Enjeux d'adaptation par canal", "Réglementation à intégrer"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, brief in enumerate(BRIEFS):
        row = 5 + i
        vals = [brief["id"], brief["produit"], brief["cible"],
                brief["message_brut"], brief["enjeux_canal"], brief["reglementation"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=90)
        ws.cell(row=row, column=1).font = body_font(bold=True, color=NEURAL_VIOLET)

    set_widths(ws, {1: 12, 2: 18, 3: 30, 4: 45, 5: 40, 6: 35})


# ────────────────────────────────────────────────────────────────────
# Onglet générique pour canal agent/courtier/direct
# ────────────────────────────────────────────────────────────────────
def build_canal_tab(ws, canal: str, variants: list, color: str) -> None:
    label = CANAL_LABELS[canal]
    write_title(ws, f"NEURAL · MultiChannelInsur — Canal : {label}", span=7)
    write_subtitle(ws, 2, "Variantes adaptées au registre et aux obligations légales de ce canal de distribution")

    header_data = ["Version", "Produit", "Registre / Ton", "Accroche", "Corps du message", "Mentions légales obligatoires", "Mots-clés canal"]
    style_header_row(ws, 4, len(header_data))
    for col, h in enumerate(header_data, 1):
        ws.cell(row=4, column=col, value=h)

    for i, v in enumerate(variants):
        row = 5 + i
        vals = [
            v["version"], v["produit"], v.get("registre", ""),
            v["accroche"], v["corps"], v["mentions_legales"],
            v.get("mots_cles_canal", ""),
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(header_data), alt=(i % 2 == 1), height=110)
        # Version cell colored per canal
        c = ws.cell(row=row, column=1)
        c.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        c.fill = canal_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[8].height = 10

    # Scores section
    write_section_header(ws, 9, "SCORES D'ADAPTATION — Calculés automatiquement par l'agent MultiChannelInsur", span=7)
    score_headers = ["Version", "Produit", "Score Adaptation Canal (0-100)", "Score Conformité DDA (0-100)", "Interprétation"]
    style_header_row(ws, 10, len(score_headers))
    for col, h in enumerate(score_headers, 1):
        ws.cell(row=10, column=col, value=h)

    for i, v in enumerate(variants):
        row = 11 + i
        s_adapt = v["score_adaptation"]
        s_conf = v["score_conformite_dda"]
        interp = (
            "Excellent — prêt à diffuser" if s_adapt >= 90 and s_conf >= 90
            else "Bon — révision mentions légales recommandée" if s_adapt >= 85
            else "Acceptable — retravailler registre de ton"
        )
        vals = [v["version"], v["produit"], s_adapt, s_conf, interp]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 30

        # Color score cells
        for col, score in [(3, s_adapt), (4, s_conf)]:
            cell = ws.cell(row=row, column=col)
            if score >= 90:
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)
            elif score >= 80:
                cell.fill = amber_fill()
                cell.font = body_font(bold=True, color=AMBER_WARN)
            else:
                cell.fill = red_fill()
                cell.font = body_font(bold=True, color=RED_NO)

    set_widths(ws, {1: 10, 2: 16, 3: 22, 4: 48, 5: 48, 6: 38, 7: 30})


# ────────────────────────────────────────────────────────────────────
# Onglet 5 — Canal Comparateur (avec checklist DSA/DMA)
# ────────────────────────────────────────────────────────────────────
def build_comparateur_tab(ws) -> None:
    color = CANAL_COLORS["comparateur"]
    label = CANAL_LABELS["comparateur"]
    write_title(ws, f"NEURAL · MultiChannelInsur — Canal : {label}", span=9)
    write_subtitle(ws, 2, "Mentions obligatoires DSA art. 22-26 + DMA art. 6-7 + AI Act art. 50 — Scoring IA disclosure")

    header_data = ["Version", "Produit", "Accroche", "Corps du message", "Mentions légales DSA/DMA + AI Act"]
    style_header_row(ws, 4, len(header_data))
    for col, h in enumerate(header_data, 1):
        ws.cell(row=4, column=col, value=h)

    for i, v in enumerate(VARIANTS_COMPARATEUR):
        row = 5 + i
        vals = [v["version"], v["produit"], v["accroche"], v["corps"], v["mentions_legales"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(header_data), alt=(i % 2 == 1), height=130)
        c = ws.cell(row=row, column=1)
        c.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        c.fill = canal_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[8].height = 10

    # DSA Checklist
    write_section_header(ws, 9, "CHECKLIST CONFORMITÉ DSA/DMA PAR VARIANTE — 5 obligations comparateur", span=9)
    dsa_headers = ["Version", "Produit", "Rémunération déclarée", "Panel mentionné", "Non-exhaustivité", "Critères classement", "Publicité identifiée", "Score DSA (0-100)", "Verdict"]
    style_header_row(ws, 10, len(dsa_headers))
    for col, h in enumerate(dsa_headers, 1):
        ws.cell(row=10, column=col, value=h)

    for i, v in enumerate(VARIANTS_COMPARATEUR):
        row = 11 + i
        dsa_checks = [
            v["dsa_remuneration"],
            v["dsa_panel"],
            v["dsa_non_exhaustivite"],
            v["dsa_criteres_classement"],
            v["dsa_publicite_identifiee"],
        ]
        score_dsa = v["score_dsa"]
        verdict = "CONFORME DSA" if score_dsa >= 95 else "RÉVISION REQUISE"
        all_vals = [v["version"], v["produit"]] + dsa_checks + [score_dsa, verdict]
        for col, val in enumerate(all_vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 35

        # Color DSA check cells
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            val_str = str(cell.value or "")
            if val_str.startswith("OUI"):
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)
            elif val_str.startswith("NON"):
                cell.fill = red_fill()
                cell.font = body_font(bold=True, color=RED_NO)

        # Score DSA
        cell_score = ws.cell(row=row, column=8)
        if score_dsa >= 95:
            cell_score.fill = green_fill()
            cell_score.font = body_font(bold=True, color=GREEN_OK)
        elif score_dsa >= 80:
            cell_score.fill = amber_fill()
            cell_score.font = body_font(bold=True, color=AMBER_WARN)
        else:
            cell_score.fill = red_fill()
            cell_score.font = body_font(bold=True, color=RED_NO)

        # Verdict
        cell_verdict = ws.cell(row=row, column=9)
        if verdict == "CONFORME DSA":
            cell_verdict.fill = green_fill()
            cell_verdict.font = body_font(bold=True, color=GREEN_OK)
        else:
            cell_verdict.fill = amber_fill()
            cell_verdict.font = body_font(bold=True, color=AMBER_WARN)

    ws.row_dimensions[14].height = 10

    # AI Act art. 50 block
    write_section_header(ws, 15, "EU AI ACT art. 50 — OBLIGATIONS DISCLOSURE IA POUR COMPARATEURS (applicable août 2026)", span=9)
    ai_act_rows = [
        ("Obligation", "Tout système de scoring automatisé qui classe ou filtre des offres assurance DOIT être identifié comme un système d'IA.", "Base légale : EU AI Act art. 50 §1 — systèmes IA qui interagissent avec des personnes physiques"),
        ("Mention FR requise", "« Les classements présentés sur cette page sont générés par un algorithme automatisé. Ce système n'est pas un conseiller humain. »", "Applicable sur toute page de résultats comparateur utilisant un algorithme de scoring"),
        ("Mention EN requise", "\"Rankings shown on this page are generated by an automated algorithm. This system is not a human advisor.\"", "Required on any comparator results page using a scoring algorithm"),
        ("Systèmes IA à risque", "Les comparateurs utilisant un scoring IA pour des produits financiers (assurance) peuvent être qualifiés de systèmes à 'risque limité' voire 'haut risque' si le scoring influence des décisions de souscription.", "Risque haut : IA qui détermine l'accès à l'assurance ou le tarif proposé"),
        ("Calendrier", "Obligations art. 50 : applicables depuis août 2026. Systèmes haut risque (annexe III) : conformité GPAI requise avant août 2027.", "Source : Règlement UE 2024/1689 (EU AI Act), JO UE 12 juillet 2024"),
    ]
    for i, (k, v, note) in enumerate(ai_act_rows):
        row = 16 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font()
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        note_cell = ws.cell(row=row, column=8, value=note)
        note_cell.font = body_font(size=9, color="475569")
        note_cell.alignment = Alignment(wrap_text=True)
        if i % 2 == 1:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = alt_fill()
        ws.row_dimensions[row].height = 50

    set_widths(ws, {1: 14, 2: 14, 3: 45, 4: 45, 5: 20, 6: 20, 7: 20, 8: 30, 9: 20})


# ────────────────────────────────────────────────────────────────────
# Onglet 6 — KPIs Conformité
# ────────────────────────────────────────────────────────────────────
def build_kpis(ws) -> None:
    write_title(ws, "NEURAL · MultiChannelInsur — 6_KPIs_Conformite : Matrice récapitulative", span=8)
    write_subtitle(ws, 2, "Score adaptation canal + Score conformité DDA + Score DSA (comparateur) — calculés par formule")

    # Matrice principale : 12 variantes × 3 scores
    write_section_header(ws, 4, "MATRICE SCORES — 12 VARIANTES (3 BRIEFS × 4 CANAUX)", span=8)

    kpi_headers = ["Version", "Brief", "Produit", "Canal", "Score Adaptation (0-100)", "Score Conformité DDA (0-100)", "Score DSA (0-100, compar. only)", "Décision"]
    style_header_row(ws, 5, len(kpi_headers))
    for col, h in enumerate(kpi_headers, 1):
        ws.cell(row=5, column=col, value=h)

    all_variants = (
        [(v, "agent") for v in VARIANTS_AGENT] +
        [(v, "courtier") for v in VARIANTS_COURTIER] +
        [(v, "direct") for v in VARIANTS_DIRECT] +
        [(v, "comparateur") for v in VARIANTS_COMPARATEUR]
    )

    for i, (v, canal) in enumerate(all_variants):
        row = 6 + i
        s_adapt = v["score_adaptation"]
        s_conf = v["score_conformite_dda"]
        s_dsa = v.get("score_dsa", 0)
        s_dsa_display = s_dsa if canal == "comparateur" else "N/A"
        decision = (
            "✓ PRÊT" if s_adapt >= 90 and s_conf >= 90
            else "⚠ RÉVISION" if s_adapt >= 80 or s_conf >= 80
            else "✗ RETRAVAILLER"
        )
        vals = [v["version"], v["brief_id"], v["produit"], CANAL_LABELS[canal],
                s_adapt, s_conf, s_dsa_display, decision]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 28

        # Canal color badge
        c = ws.cell(row=row, column=4)
        c.fill = canal_fill(CANAL_COLORS[canal])
        c.font = Font(name=FONT_FAMILY, size=9, bold=True, color="FFFFFF")

        # Score adaptation
        cell_adapt = ws.cell(row=row, column=5)
        if s_adapt >= 90:
            cell_adapt.fill = green_fill()
            cell_adapt.font = body_font(bold=True, color=GREEN_OK)
        elif s_adapt >= 80:
            cell_adapt.fill = amber_fill()
            cell_adapt.font = body_font(bold=True, color=AMBER_WARN)
        else:
            cell_adapt.fill = red_fill()
            cell_adapt.font = body_font(bold=True, color=RED_NO)

        # Score DDA
        cell_dda = ws.cell(row=row, column=6)
        if s_conf >= 90:
            cell_dda.fill = green_fill()
            cell_dda.font = body_font(bold=True, color=GREEN_OK)
        elif s_conf >= 80:
            cell_dda.fill = amber_fill()
            cell_dda.font = body_font(bold=True, color=AMBER_WARN)
        else:
            cell_dda.fill = red_fill()
            cell_dda.font = body_font(bold=True, color=RED_NO)

        # Decision
        cell_dec = ws.cell(row=row, column=8)
        if "PRÊT" in decision:
            cell_dec.fill = green_fill()
            cell_dec.font = body_font(bold=True, color=GREEN_OK)
        elif "RÉVISION" in decision:
            cell_dec.fill = amber_fill()
            cell_dec.font = body_font(bold=True, color=AMBER_WARN)
        else:
            cell_dec.fill = red_fill()
            cell_dec.font = body_font(bold=True, color=RED_NO)

    # Aggregated stats with formulas
    last_row = 6 + len(all_variants)
    ws.row_dimensions[last_row + 1].height = 12

    write_section_header(ws, last_row + 2, "STATISTIQUES AGRÉGÉES (formules)", span=8)

    stats_labels = [
        ("Nb variantes analysées", f"=COUNTA(E6:E{last_row - 1})", ""),
        ("Score adaptation moyen", f"=AVERAGE(E6:E{last_row - 1})", ""),
        ("Score conformité DDA moyen", f"=AVERAGE(F6:F{last_row - 1})", ""),
        ("Nb variantes PRÊT", f'=COUNTIF(H6:H{last_row - 1},"*PRÊT*")', ""),
        ("Nb variantes EN RÉVISION", f'=COUNTIF(H6:H{last_row - 1},"*RÉVISION*")', ""),
        ("Nb variantes À RETRAVAILLER", f'=COUNTIF(H6:H{last_row - 1},"*RETRAV*")', ""),
        ("Score DSA moyen (comparateur)", f"=AVERAGE(G{6+9}:G{6+11})", "Calculé sur les 3 variantes comparateur uniquement"),
    ]

    for i, (label, formula, note) in enumerate(stats_labels):
        row = last_row + 3 + i
        ws.cell(row=row, column=1, value=label).font = body_font(bold=True)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = body_font(bold=True, color="0000FF")  # blue = formula
        cell.alignment = Alignment(horizontal="center")
        if note:
            ws.cell(row=row, column=3, value=note).font = body_font(size=9, color="475569")
        ws.row_dimensions[row].height = 22

    # Légende canaux
    ws.row_dimensions[last_row + 11].height = 12
    write_section_header(ws, last_row + 12, "LÉGENDE CANAUX", span=8)
    for i, (canal, label) in enumerate(CANAL_LABELS.items()):
        row = last_row + 13 + i
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        cell.fill = canal_fill(CANAL_COLORS[canal])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    set_widths(ws, {1: 18, 2: 14, 3: 22, 4: 26, 5: 22, 6: 22, 7: 22, 8: 16})


# ────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────
def main() -> None:
    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = "0_README"
    build_readme(ws_readme)

    ws_briefs = wb.create_sheet("1_Briefs_Source")
    build_briefs(ws_briefs)

    ws_agent = wb.create_sheet("2_Canal_Agent")
    build_canal_tab(ws_agent, "agent", VARIANTS_AGENT, CANAL_COLORS["agent"])

    ws_courtier = wb.create_sheet("3_Canal_Courtier")
    build_canal_tab(ws_courtier, "courtier", VARIANTS_COURTIER, CANAL_COLORS["courtier"])

    ws_direct = wb.create_sheet("4_Canal_Direct")
    build_canal_tab(ws_direct, "direct", VARIANTS_DIRECT, CANAL_COLORS["direct"])

    ws_comp = wb.create_sheet("5_Canal_Comparateur")
    build_comparateur_tab(ws_comp)

    ws_kpis = wb.create_sheet("6_KPIs_Conformite")
    build_kpis(ws_kpis)

    out = "MultiChannelInsur_NEURAL.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
