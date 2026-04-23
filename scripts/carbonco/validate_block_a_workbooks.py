from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_ROOT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\Carbon and Co")

CARBON_FILE = WORKBOOK_ROOT / "CarbonCo_Calcul_Carbone_v2.xlsx"
ESG_FILE = WORKBOOK_ROOT / "CarbonCo_ESG_Social.xlsx"
FINANCE_FILE = WORKBOOK_ROOT / "CarbonCo_Finance_DPP_v1_3.xlsx"


CANONICAL_CC_RANGES = {
    "CC_Raison_Sociale": "'Paramètres'!$B$4",
    "CC_Annee_Reporting": "'Paramètres'!$B$7",
    "CC_CA_Net": "'Paramètres'!$B$9",
    "CC_ETP": "'Paramètres'!$B$11",
    "CC_Secteur_Activite": "'Paramètres'!$B$6",
    "CC_Secteur_NAF": "'Paramètres'!$C$6",
    "CC_Surface_Totale": "'Paramètres'!$B$12",
    "CC_CapEx_Total": "'Paramètres'!$B$14",
    "CC_OpEx_Eligible_Taxo": "'Paramètres'!$B$15",
    "CC_GES_Scope1": "'Synthese_GES'!$C$10",
    "CC_GES_Scope2_LB": "'Synthese_GES'!$C$15",
    "CC_GES_Scope2_MB": "'Synthese_GES'!$C$17",
    "CC_GES_Scope3": "'Synthese_GES'!$C$35",
    "CC_GES_Total_S123": "'Synthese_GES'!$C$47",
    "CC_Intensite_CA": "'Synthese_GES'!$C$50",
    "CC_Intensite_ETP": "'Synthese_GES'!$C$51",
    "CC_Part_Scope1": "'Synthese_GES'!$C$53",
    "CC_Part_Scope2": "'Synthese_GES'!$C$54",
    "CC_Part_Scope3": "'Synthese_GES'!$C$55",
    "CC_Conso_Energie_MWh": "'Energie'!$E$19",
    "CC_Part_ENR": "'Energie'!$E$20",
    "CC_Taxo_CA_Aligne": "'Taxonomie'!$E$27",
    "CC_Taxo_CapEx_Aligne": "'Taxonomie'!$E$28",
    "CC_Taxo_OpEx_Aligne": "'Taxonomie'!$E$29",
    "CC_CBAM_Cout_Estime": "'CBAM'!$M$24",
    "CC_SBTI_Annee_Baseline": "'Trajectoire_SBTi'!$B$4",
    "CC_SBTI_Baseline_S12": "'Trajectoire_SBTi'!$B$5",
    "CC_SBTI_Baseline_S3": "'Trajectoire_SBTi'!$B$6",
    "CC_SBTI_Taux_S12": "'Trajectoire_SBTi'!$B$8",
    "CC_SBTI_Taux_S3": "'Trajectoire_SBTi'!$B$9",
}

REQUIRED_SHEETS = {
    "carbon": {"Couverture", "Sommaire", "Paramètres", "Synthese_GES", "Energie", "Claude Log"},
    "esg": {"Couverture", "Sommaire", "Liaison_Donnees", "Claude Log"},
    "finance": {"Couverture", "Sommaire", "Liaison_Donnees", "Claude Log"},
}

ESG_LIAISON_EXPECTED = {
    "E16": "Synthese_GES!C15",
    "E17": "Synthese_GES!C17",
    "E18": "Synthese_GES!C35",
    "E22": "Energie!E20",
    "E23": "Energie!E19",
    "E25": "Taxonomie!E27",
    "E26": "CBAM!M24",
}

FINANCE_LIAISON_EXPECTED = {
    "B19": "Total S1+S2(LB)+S3 (tCO2e)",
    "E16": "Synthese_GES!C15",
    "E17": "Synthese_GES!C17",
    "E18": "Synthese_GES!C35",
    "E19": "Calcul auto (LB)",
    "E23": "Energie!E20",
    "E24": "Taxonomie!E27",
    "E25": "Taxonomie!E28",
    "E26": "CBAM!M24",
    "E30": "Trajectoire_SBTi!B4",
}


def workbook_summary(path: Path) -> dict:
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet_names = list(wb.sheetnames)
        defined_names = {name: wb.defined_names[name].attr_text for name in wb.defined_names.keys()}
        summary = {
            "path": str(path),
            "sheet_count": len(sheet_names),
            "named_range_count": len(defined_names),
            "sheet_names": sheet_names,
            "defined_names": defined_names,
        }
    finally:
        wb.close()
    return summary


def validate_required_sheets(summary: dict, required: set[str]) -> list[str]:
    current = set(summary["sheet_names"])
    missing = sorted(required - current)
    return [f"Missing required sheet: {sheet}" for sheet in missing]


def validate_named_ranges(summary: dict) -> list[str]:
    failures: list[str] = []
    current = summary["defined_names"]
    for name, expected_target in CANONICAL_CC_RANGES.items():
        actual_target = current.get(name)
        if actual_target is None:
            failures.append(f"Missing named range: {name}")
            continue
        if actual_target != expected_target:
            failures.append(
                f"Named range mismatch: {name} expected {expected_target} got {actual_target}"
            )
    return failures


def validate_liaison_cells(path: Path, expected: dict[str, str]) -> list[str]:
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ws = wb["Liaison_Donnees"]
        failures = []
        for cell, expected_value in expected.items():
            actual_value = ws[cell].value
            if actual_value != expected_value:
                failures.append(
                    f"Liaison mismatch {path.name} {cell}: expected {expected_value!r} got {actual_value!r}"
                )
        return failures
    finally:
        wb.close()


def main() -> None:
    report: dict[str, object] = {
        "status": "ok",
        "checks": [],
        "failures": [],
    }

    paths = {
        "carbon": CARBON_FILE,
        "esg": ESG_FILE,
        "finance": FINANCE_FILE,
    }

    for key, path in paths.items():
        if not path.exists():
            report["failures"].append(f"Missing workbook: {path}")
            continue

        summary = workbook_summary(path)
        report["checks"].append(
            {
                "workbook": key,
                "path": str(path),
                "sheet_count": summary["sheet_count"],
                "named_range_count": summary["named_range_count"],
                "has_claude_log": "Claude Log" in summary["sheet_names"],
            }
        )

        report["failures"].extend(validate_required_sheets(summary, REQUIRED_SHEETS[key]))
        if key == "carbon":
            report["failures"].extend(validate_named_ranges(summary))
        if key == "esg":
            report["failures"].extend(validate_liaison_cells(path, ESG_LIAISON_EXPECTED))
        if key == "finance":
            report["failures"].extend(validate_liaison_cells(path, FINANCE_LIAISON_EXPECTED))

    if report["failures"]:
        report["status"] = "failed"
        print("BLOCK_A_VALIDATION=FAILED")
        print(json.dumps(report, ensure_ascii=True, indent=2))
        sys.exit(1)

    print("BLOCK_A_VALIDATION=OK")
    print(json.dumps(report, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
