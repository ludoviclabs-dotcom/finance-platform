"""
NEURAL_AG005_GreenClaimChecker.xlsx
Detection de claims + matching preuve + scoring risque + gate de conformite.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, STATUS, VERT, VIOLET, WORDING_TYPE, YESNO,
    ajouter_table, attach_dv, cf_stale, convertir_dates, D, DT, dv_liste,
    ecrire_donnees, en_tetes, params_table, param_ligne, section, titre,
    today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_AG005_GreenClaimChecker.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "AG-005  —  GREEN CLAIM CHECKER", "Verification RSE / developpement durable — evidence, risque, gate", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 86

    section(ws, 4, 6, "Mission", NOIR)
    ecrire_donnees(ws, 5, [
        ["", "Role",        "Detecter toute affirmation RSE/ESG, la matcher contre une preuve, scorer le risque, decider PASS / BLOCK."],
        ["", "Input",       "Claims extraits + wording type (ABSOLUTE/QUALIFIED/COMPARATIVE) + juridiction + date cible publication."],
        ["", "Output",      "Decision (PASS / BLOCK / PASS_WITH_REVIEW) + Risk level + Reference evidence."],
        ["", "Gate",        "CLAIM — blocking par defaut. SLA 48h standard / 8h en crise."],
        ["", "Appelants",   "AG-002 (presse), AG-003 (event si claims). Peut etre invoque direct."],
        ["", "Referentiel", "FOUNDATIONS!7_CLAIMS_EVIDENCE_REGISTRY."],
    ], align=AL)

    section(ws, 13, 6, "Logique de risque", BORDEAUX)
    ecrire_donnees(ws, 14, [
        ["", "CRITICAL",    "Pas d'evidence trouvee"],
        ["", "HIGH",        "Evidence existe mais STALE (expiree), OU wording ABSOLUTE sans qualifieur"],
        ["", "MEDIUM",      "Wording QUALIFIED ou COMPARATIVE avec evidence valide"],
        ["", "LOW",         "Evidence valide + wording qualifie + juridiction couverte"],
    ], align=AL)

    section(ws, 21, 6, "Reference reglementaire", VIOLET)
    ecrire_donnees(ws, 22, [
        ["", "EU",         "Directive Green Claims (2024) — affirmations environnementales doivent etre verifiables."],
        ["", "FR",         "Loi Climat & Resilience (Art 12) — interdit d'affirmer qu'un produit est neutre en carbone sans preuve."],
        ["", "UK",         "CMA Green Claims Code (2021)."],
        ["", "US (FTC)",   "Green Guides (revision 2025 en cours)."],
        ["", "Note",       "Voir 10_JURIDICTION_MATRIX pour detail par claim / juridiction."],
    ], align=AL)


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES AGENT", "tblParamsAG005", 3)
    rows = [
        ("AGENT_ID",        "AG-005", ""),
        ("AGENT_NAME",      "GreenClaimChecker", ""),
        ("MAISON_ID",       "M-001", ""),
        ("MAISON_NAME",     "Maison pilote", ""),
        ("LANG_PRIMARY",    "FR", ""),
        ("LANG_SECONDARY",  "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", ""),
        ("DATA_VERSION",    "v2026.04", ""),
        ("PROMPT_VERSION",  "v1.0.0", ""),
        ("SLA_STANDARD_H",  48, "Heures"),
        ("SLA_CRISIS_H",    8,  "Heures"),
        ("CRISIS_MODE_ON",  "NO", "Si YES, SLA reduit"),
        ("BLOCK_IF_STALE",  "YES", "Bloquer si evidence STALE"),
        ("BLOCK_IF_ABSOLUTE_UNQUALIFIED", "YES", "Bloquer si ABSOLUTE sans qualifieur"),
        ("LAST_REFRESH",    D(today_iso()), ""),
    ]
    params_table(ws, "tblParamsAG005", start_row=4, rows=rows)


def onglet_claim_intake(wb):
    ws = wb.create_sheet("2_CLAIM_INTAKE")
    titre(ws, "CLAIM INTAKE", "Claims extraits des documents — arrivent de AG-002 / AG-003", 9)
    en_tetes(ws, 4, ["INTAKE_ID", "MAISON_ID", "DOC_ID", "SOURCE_AGENT", "LANG", "CLAIM_RAW",
                      "CONTEXT", "JURIDICTION_CIBLE", "SUBMITTED_AT"],
             [12, 10, 12, 14, 8, 60, 22, 20, 18])
    intake = [
        ("CL-IN-001", "M-001", "DOC-0001", "AG-002", "FR", "Ciselee a la main par les artisans",       "Presse lifestyle",  "EU",     DT("2026-04-15 09:10")),
        ("CL-IN-002", "M-001", "DOC-0002", "AG-002", "FR", "Eco-responsable et durable",              "Presse business",   "EU",     DT("2026-04-15 10:22")),
        ("CL-IN-003", "M-001", "DOC-0004", "AG-002", "EN", "Fully sustainable collection",            "Press business",    "EU + UK",DT("2026-04-15 13:30")),
        ("CL-IN-004", "M-001", "DOC-0009", "AG-002", "FR", "Or certifie 80% recycle",                 "Presse lifestyle",  "EU",     DT("2026-04-16 08:00")),
        ("CL-IN-005", "M-001", "DOC-0010", "AG-003", "FR", "Zero dechet dans notre atelier",          "Event press",        "FR",     DT("2026-04-16 09:30")),
        ("CL-IN-006", "M-001", "DOC-0011", "AG-002", "FR", "Cuir certifie LWG",                       "Presse trade",      "EU",     DT("2026-04-16 10:00")),
        ("CL-IN-007", "M-001", "DOC-0012", "AG-002", "EN", "Carbon neutral packaging",                "Press trade",        "US",     DT("2026-04-16 11:15")),
        ("CL-IN-008", "M-001", "DOC-0013", "AG-002", "FR", "Fabrique en France (100%)",               "Presse lifestyle",  "FR",     DT("2026-04-16 13:00")),
        ("CL-IN-009", "M-001", "DOC-0014", "AG-002", "FR", "Serie limitee 48 pieces",                 "Presse trade",      "EU",     DT("2026-04-16 14:30")),
        ("CL-IN-010", "M-001", "DOC-0015", "AG-003", "FR", "Diamants sans conflit",                    "Event trade",        "EU",     DT("2026-04-16 15:00")),
        ("CL-IN-011", "M-001", "DOC-0016", "AG-002", "FR", "Cruelty-free pour toute la gamme",        "Presse lifestyle",  "EU",     DT("2026-04-16 16:15")),
        ("CL-IN-012", "M-001", "DOC-0017", "AG-002", "FR", "Empreinte carbone neutre sur cette collection","Presse business","EU", DT("2026-04-16 17:30")),
    ]
    last = ecrire_donnees(ws, 5, intake, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[9])
    ajouter_table(ws, "tblClaimIntake", f"A4:I{last}")
    attach_dv(ws, f"D5:D{last}", ["USER", "AG-002", "AG-003", "CRISIS"])
    attach_dv(ws, f"E5:E{last}", LANG)


def onglet_claim_library(wb):
    ws = wb.create_sheet("3_CLAIM_LIBRARY")
    titre(ws, "CLAIM LIBRARY", "Patterns reconnus — type, autorisation, evidence requise", 8)
    en_tetes(ws, 4, ["LIB_ID", "PATTERN", "CATEGORIE", "WORDING_TYPE", "AUTORISATION", "EVIDENCE_REQUIRED", "JURIDICTIONS_OK", "NOTE"],
             [10, 30, 16, 16, 26, 20, 22, 34])
    lib = [
        ("CL-LIB-001", "recycle",           "ESG",   "QUALIFIED", "AUTORISE_SI_PROUVE",  "YES",  "EU, UK, US, FR", "% + source"),
        ("CL-LIB-002", "durable",            "ESG",   "ABSOLUTE",  "INTERDIT",            "YES",  "-",               "Vague"),
        ("CL-LIB-003", "eco-responsable",    "ESG",   "ABSOLUTE",  "INTERDIT",            "YES",  "-",               "Vague"),
        ("CL-LIB-004", "sustainable",        "ESG",   "ABSOLUTE",  "INTERDIT",            "YES",  "-",               "Vague"),
        ("CL-LIB-005", "fully sustainable",  "ESG",   "ABSOLUTE",  "INTERDIT",            "YES",  "-",               "Vague"),
        ("CL-LIB-006", "zero dechet",        "ESG",   "ABSOLUTE",  "REVIEW",              "YES",  "FR, EU",          "Doit etre prouve par audit"),
        ("CL-LIB-007", "carbon neutral",     "ESG",   "ABSOLUTE",  "REVIEW",              "YES",  "EU, UK",          "Loi Climat France interdit"),
        ("CL-LIB-008", "empreinte carbone",  "ESG",   "QUALIFIED", "AUTORISE_SI_PROUVE",  "YES",  "EU, FR",          "Avec scope + data"),
        ("CL-LIB-009", "LWG certifie",       "ESG",   "QUALIFIED", "AUTORISE_SI_PROUVE",  "YES",  "EU",              "Certificat LWG valide"),
        ("CL-LIB-010", "fabrique en France", "ORIGIN","QUALIFIED", "AUTORISE_SI_PROUVE",  "YES",  "FR, EU",          "Label IGP / atelier"),
        ("CL-LIB-011", "made in France",     "ORIGIN","QUALIFIED", "AUTORISE_SI_PROUVE",  "YES",  "-",               "Version EN"),
        ("CL-LIB-012", "kimberley",          "ESG",   "QUALIFIED", "AUTORISE_SI_PROUVE",  "YES",  "EU, FR, UK",      "Kimberley Process"),
        ("CL-LIB-013", "cruelty-free",       "ESG",   "ABSOLUTE",  "REVIEW",              "YES",  "EU, US",          "Certification PETA/Leaping Bunny"),
        ("CL-LIB-014", "vegan",              "ESG",   "ABSOLUTE",  "REVIEW",              "YES",  "EU, US",          "Certification"),
        ("CL-LIB-015", "serie limitee",      "SCARCITY","QUALIFIED","AUTORISE_SI_PROUVE", "YES",  "EU",              "Chiffre exact requis"),
        ("CL-LIB-016", "fait main",          "TECH",  "ABSOLUTE",  "REVIEW",              "YES",  "EU, FR",          "Video + process"),
        ("CL-LIB-017", "artisanal",          "TECH",  "QUALIFIED", "AUTORISE",            "NO",   "EU, FR",          "Vocable culturel"),
    ]
    last = ecrire_donnees(ws, 5, lib, align=AL)
    ajouter_table(ws, "tblClaimLibrary", f"A4:H{last}")

    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"INTERDIT"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"REVIEW"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"AUTORISE_SI_PROUVE"'], fill=PatternFill(start_color=FOND_BLEU, end_color=FOND_BLEU, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"AUTORISE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))


def onglet_evidence_match(wb):
    ws = wb.create_sheet("4_EVIDENCE_MATCH")
    titre(ws, "EVIDENCE MATCHING", "Matching intake <-> Claims Registry — found / status / expiry", 10)
    en_tetes(ws, 4, ["MATCH_ID", "INTAKE_ID", "CLAIM_ID_REGISTRY", "EVIDENCE_TITLE", "EVIDENCE_DATE", "EVIDENCE_EXPIRY",
                      "EVIDENCE_FOUND", "EVIDENCE_STATUS", "WORDING_TYPE", "NOTE"],
             [10, 12, 18, 36, 14, 14, 14, 16, 14, 38])
    data = [
        ("MT-001", "CL-IN-001", "",        "",                                 None,           None,           "NO",  "",      "",          "Pas de claim ESG"),
        ("MT-002", "CL-IN-002", "",        "",                                 None,           None,           "NO",  "",      "ABSOLUTE",  "Vague, interdit"),
        ("MT-003", "CL-IN-003", "",        "",                                 None,           None,           "NO",  "",      "ABSOLUTE",  "Vague, interdit"),
        ("MT-004", "CL-IN-004", "CLM-001", "Attestation fournisseur LBMA 2026",D("2026-01-15"),D("2027-01-15"),"YES", "",      "QUALIFIED", "Evidence valide"),
        ("MT-005", "CL-IN-005", "CLM-003", "",                                 None,           None,           "NO",  "",      "ABSOLUTE",  "Zero dechet non prouve"),
        ("MT-006", "CL-IN-006", "CLM-002", "Certificat LWG 2025",              D("2025-11-01"),D("2026-11-01"),"YES", "",      "QUALIFIED", "Evidence valide"),
        ("MT-007", "CL-IN-007", "CLM-010", "",                                 None,           None,           "NO",  "",      "ABSOLUTE",  "Carbon neutral = interdit FR"),
        ("MT-008", "CL-IN-008", "CLM-007", "Label IGP atelier",                D("2025-12-01"),D("2027-12-01"),"YES", "",      "QUALIFIED", "Valide"),
        ("MT-009", "CL-IN-009", "CLM-012", "Production book serialise",        D("2026-03-01"),D("2027-03-01"),"YES", "",      "QUALIFIED", "Chiffre confirme"),
        ("MT-010", "CL-IN-010", "CLM-005", "Certificat Kimberley Process",     D("2026-02-20"),D("2027-02-20"),"YES", "",      "QUALIFIED", "Valide"),
        ("MT-011", "CL-IN-011", "CLM-011", "Leaping Bunny",                    D("2025-07-01"),D("2026-07-01"),"YES", "",      "ABSOLUTE",  "Limite REVIEW si couverture gamme entiere"),
        ("MT-012", "CL-IN-012", "CLM-010", "",                                 None,           None,           "NO",  "",      "ABSOLUTE",  "Interdit FR Loi Climat"),
    ]
    last = ecrire_donnees(ws, 5, data, align=AL)
    convertir_dates(ws, 5, last, date_cols=[5, 6])

    # EVIDENCE_STATUS formule : "-" si pas d'evidence, STALE si expiry passee, VALID sinon
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=8,
                     value=f'=IF(G{r}="NO","-",IF(OR(F{r}="",NOT(ISNUMBER(F{r}))),"VALID",IF(F{r}<TODAY(),"STALE","VALID")))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ajouter_table(ws, "tblEvidenceMatch", f"A4:J{last}")
    # R-06
    attach_dv(ws, f"G5:G{last}", YESNO)
    attach_dv(ws, f"I5:I{last}", ["", "ABSOLUTE", "QUALIFIED", "COMPARATIVE"])
    # R-08 : CF J-30
    cf_stale(ws, f"F5:F{last}", days_window=30)

    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"G5:G{last}", CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"STALE"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"VALID"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))


def onglet_risk_engine(wb):
    ws = wb.create_sheet("5_RISK_ENGINE")
    titre(ws, "RISK ENGINE", "Score de risque par claim + decision finale", 11)
    en_tetes(ws, 4, ["RUN_ID", "INTAKE_ID", "MAISON_ID", "TIMESTAMP", "EVIDENCE_FOUND",
                      "EVIDENCE_STATUS", "WORDING_TYPE", "JURIDICTION_OK", "RISK_CLASS", "DECISION", "ESCALATION"],
             [16, 12, 10, 18, 14, 14, 14, 14, 12, 18, 18])

    # Seed runs aligned on intake
    rows = [
        ("R-20260415-010", "CL-IN-001", "M-001", DT("2026-04-15 09:11"), "NO",  "-",     "",            "YES", "", "", ""),
        ("R-20260415-011", "CL-IN-002", "M-001", DT("2026-04-15 10:23"), "NO",  "-",     "ABSOLUTE",    "NO",  "", "", ""),
        ("R-20260415-012", "CL-IN-003", "M-001", DT("2026-04-15 13:31"), "NO",  "-",     "ABSOLUTE",    "NO",  "", "", ""),
        ("R-20260416-001", "CL-IN-004", "M-001", DT("2026-04-16 08:02"), "YES", "VALID", "QUALIFIED",   "YES", "", "", ""),
        ("R-20260416-002", "CL-IN-005", "M-001", DT("2026-04-16 09:32"), "NO",  "-",     "ABSOLUTE",    "NO",  "", "", ""),
        ("R-20260416-003", "CL-IN-006", "M-001", DT("2026-04-16 10:02"), "YES", "VALID", "QUALIFIED",   "YES", "", "", ""),
        ("R-20260416-004", "CL-IN-007", "M-001", DT("2026-04-16 11:17"), "NO",  "-",     "ABSOLUTE",    "NO",  "", "", ""),
        ("R-20260416-005", "CL-IN-008", "M-001", DT("2026-04-16 13:02"), "YES", "VALID", "QUALIFIED",   "YES", "", "", ""),
        ("R-20260416-006", "CL-IN-009", "M-001", DT("2026-04-16 14:32"), "YES", "VALID", "QUALIFIED",   "YES", "", "", ""),
        ("R-20260416-007", "CL-IN-010", "M-001", DT("2026-04-16 15:02"), "YES", "VALID", "QUALIFIED",   "YES", "", "", ""),
        ("R-20260416-008", "CL-IN-011", "M-001", DT("2026-04-16 16:17"), "YES", "VALID", "ABSOLUTE",    "YES", "", "", ""),
        ("R-20260416-009", "CL-IN-012", "M-001", DT("2026-04-16 17:32"), "NO",  "-",     "ABSOLUTE",    "NO",  "", "", ""),
    ]
    last = ecrire_donnees(ws, 5, rows, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[4])

    # Risk engine formula
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=9,
                     value=f'=IFS(E{r}="NO","CRITICAL",F{r}="STALE","HIGH",G{r}="ABSOLUTE","HIGH",G{r}="QUALIFIED","MEDIUM",G{r}="COMPARATIVE","HIGH",TRUE,"LOW")')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

        c = ws.cell(row=r, column=10, value=f'=IF(OR(I{r}="CRITICAL",I{r}="HIGH"),"BLOCK",IF(I{r}="MEDIUM","PASS_WITH_REVIEW","PASS"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

        c = ws.cell(row=r, column=11, value=f'=IF(I{r}="CRITICAL","Legal + ESG Lead",IF(I{r}="HIGH","ESG Lead","-"))')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF

    ajouter_table(ws, "tblRiskEngine", f"A4:K{last}")

    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=PatternFill(start_color=FOND_BLEU, end_color=FOND_BLEU, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"LOW"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"BLOCK"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"J5:J{last}", CellIsRule(operator="equal", formula=['"PASS_WITH_REVIEW"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_review_queue(wb):
    ws = wb.create_sheet("6_REVIEW_QUEUE")
    titre(ws, "REVIEW QUEUE", "Revues humaines ESG + Legal", 9)
    en_tetes(ws, 4, ["REVIEW_ID", "RUN_ID", "INTAKE_ID", "RISK_CLASS", "APPROVER_ROLE", "ASSIGNED_TO", "DUE_AT", "STATUS", "COMMENT"],
             [14, 18, 12, 12, 24, 20, 18, 14, 40])
    rev = [
        ("CREV-001", "R-20260415-011", "CL-IN-002", "CRITICAL", "ESG Lead + Legal", "C. Leblanc", DT("2026-04-17 10:23"), "PENDING", "Eco-responsable non prouve"),
        ("CREV-002", "R-20260415-012", "CL-IN-003", "CRITICAL", "ESG Lead + Legal", "C. Leblanc", DT("2026-04-17 13:31"), "PENDING", "Sustainable vague"),
        ("CREV-003", "R-20260416-002", "CL-IN-005", "CRITICAL", "ESG Lead + Legal", "C. Leblanc", DT("2026-04-18 09:32"), "PENDING", "Zero dechet non audite"),
        ("CREV-004", "R-20260416-004", "CL-IN-007", "CRITICAL", "ESG Lead + Legal", "C. Leblanc", DT("2026-04-18 11:17"), "PENDING", "Carbon neutral interdit FR"),
        ("CREV-005", "R-20260416-008", "CL-IN-011", "MEDIUM",   "ESG Lead",         "C. Leblanc", DT("2026-04-18 16:17"), "PENDING", "Cruelty-free couverture a verifier"),
        ("CREV-006", "R-20260416-009", "CL-IN-012", "CRITICAL", "ESG Lead + Legal", "C. Leblanc", DT("2026-04-18 17:32"), "PENDING", "Empreinte carbone neutre interdit"),
    ]
    last = ecrire_donnees(ws, 5, rev, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[7])
    ajouter_table(ws, "tblReviewQueueClaims", f"A4:I{last}")
    attach_dv(ws, f"D5:D{last}", ["CRITICAL", "HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"H5:H{last}", ["PENDING", "CLOSED", "ESCALATED"])


def onglet_decision_log(wb):
    ws = wb.create_sheet("7_DECISION_LOG")
    titre(ws, "DECISION LOG", "Journal exhaustif — export vers MASTER", 14)
    en_tetes(ws, 4, ["RUN_ID", "AGENT_ID", "MAISON_ID", "DOC_ID", "TIMESTAMP", "LANG",
                      "FINAL_STATUS", "RISK_LEVEL", "GATE_TRIGGERED", "SLA_MET",
                      "TOTAL_MINUTES", "TOKENS_IN", "TOKENS_OUT", "COST_USD"],
             [16, 10, 10, 12, 18, 8, 14, 12, 12, 10, 12, 12, 12, 12])
    rows = [
        ("R-20260415-010", "AG-005", "M-001", "DOC-0001", DT("2026-04-15 09:11"), "FR", "APPROVED", "LOW",      "CLAIM", "YES", 2.1, 800, 300, 0.010),
        ("R-20260415-011", "AG-005", "M-001", "DOC-0002", DT("2026-04-15 10:23"), "FR", "REJECTED", "CRITICAL", "CLAIM", "YES", 3.8, 1200,600, 0.018),
        ("R-20260415-012", "AG-005", "M-001", "DOC-0004", DT("2026-04-15 13:31"), "EN", "REJECTED", "CRITICAL", "CLAIM", "YES", 3.9, 1300,650, 0.019),
        ("R-20260416-001", "AG-005", "M-001", "DOC-0009", DT("2026-04-16 08:02"), "FR", "APPROVED", "MEDIUM",   "CLAIM", "YES", 4.5, 1400,700, 0.022),
        ("R-20260416-002", "AG-005", "M-001", "DOC-0010", DT("2026-04-16 09:32"), "FR", "REJECTED", "CRITICAL", "CLAIM", "YES", 3.2, 1100,500, 0.016),
        ("R-20260416-003", "AG-005", "M-001", "DOC-0011", DT("2026-04-16 10:02"), "FR", "APPROVED", "MEDIUM",   "CLAIM", "YES", 4.0, 1300,650, 0.021),
        ("R-20260416-004", "AG-005", "M-001", "DOC-0012", DT("2026-04-16 11:17"), "EN", "REJECTED", "CRITICAL", "CLAIM", "YES", 3.7, 1250,600, 0.019),
        ("R-20260416-005", "AG-005", "M-001", "DOC-0013", DT("2026-04-16 13:02"), "FR", "APPROVED", "MEDIUM",   "CLAIM", "YES", 3.5, 1200,600, 0.018),
        ("R-20260416-006", "AG-005", "M-001", "DOC-0014", DT("2026-04-16 14:32"), "FR", "APPROVED", "MEDIUM",   "CLAIM", "YES", 2.9, 1000,500, 0.014),
        ("R-20260416-007", "AG-005", "M-001", "DOC-0015", DT("2026-04-16 15:02"), "FR", "APPROVED", "MEDIUM",   "CLAIM", "YES", 3.3, 1100,550, 0.016),
        ("R-20260416-008", "AG-005", "M-001", "DOC-0016", DT("2026-04-16 16:17"), "FR", "IN_REVIEW","MEDIUM",   "CLAIM", "YES", 3.1, 1100,550, 0.016),
        ("R-20260416-009", "AG-005", "M-001", "DOC-0017", DT("2026-04-16 17:32"), "FR", "REJECTED", "CRITICAL", "CLAIM", "YES", 4.1, 1400,700, 0.022),
    ]
    last = ecrire_donnees(ws, 5, rows, fmt_cols={11: "0.00", 12: "0", 13: "0", 14: "0.000"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5])
    ajouter_table(ws, "tblDecisionLog", f"A4:N{last}")


def onglet_dashboard(wb):
    ws = wb.create_sheet("8_DASHBOARD")
    titre(ws, "DASHBOARD", "Claims par risque, coverage evidence, taux BLOCK", 4)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18

    section(ws, 4, 4, "KPIs claims", NOIR)
    kpis = [
        ("Claims CRITICAL",                 '=COUNTIFS(tblRiskEngine[RISK_CLASS],"CRITICAL")',          "0"),
        ("Claims HIGH",                     '=COUNTIFS(tblRiskEngine[RISK_CLASS],"HIGH")',              "0"),
        ("Claims MEDIUM",                   '=COUNTIFS(tblRiskEngine[RISK_CLASS],"MEDIUM")',            "0"),
        ("Claims LOW",                      '=COUNTIFS(tblRiskEngine[RISK_CLASS],"LOW")',               "0"),
        ("Coverage evidence (found/total)", '=IFERROR(COUNTIFS(tblRiskEngine[EVIDENCE_FOUND],"YES")/COUNTA(tblRiskEngine[RUN_ID]),0)', "0.0%"),
        ("Taux BLOCK",                      '=IFERROR(COUNTIFS(tblRiskEngine[DECISION],"BLOCK")/COUNTA(tblRiskEngine[RUN_ID]),0)',  "0.0%"),
        ("Taux PASS",                       '=IFERROR(COUNTIFS(tblRiskEngine[DECISION],"PASS")/COUNTA(tblRiskEngine[RUN_ID]),0)',    "0.0%"),
        ("Cout USD cumule",                 '=IFERROR(SUM(tblDecisionLog[COST_USD]),0)',                "0.000"),
    ]
    for i, (lbl, formula, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = fmt

    # Chart : distribution risque
    section(ws, 15, 4, "Distribution risque", BORDEAUX)
    for i, lvl in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW"]):
        r = 16 + i
        ws.cell(row=r, column=2, value=lvl).border = BF
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True)
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(tblRiskEngine[RISK_CLASS],"{lvl}")')
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch = PieChart()
    data = Reference(ws, min_col=3, min_row=16, max_row=19)
    labels = Reference(ws, min_col=2, min_row=16, max_row=19)
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(labels)
    ch.title = "Claims par niveau de risque"
    ch.height = 8
    ch.width = 14
    ws.add_chart(ch, "E15")


def onglet_testset(wb):
    ws = wb.create_sheet("9_TESTSET")
    titre(ws, "TESTSET CLAIMS", "Cas de reference AG-005 — EXPECTED vs OBSERVED avec PASS/FAIL", 8)
    en_tetes(ws, 4, ["TEST_ID", "CLAIM_TEST", "WORDING_TYPE", "EVIDENCE_EXPECTED", "JURI",
                      "EXPECTED_DECISION", "OBSERVED_DECISION", "PASS_FAIL"],
             [10, 40, 16, 22, 16, 20, 20, 12])
    tests = [
        ("T-001", "Or certifie 80% recycle, audit LBMA 2026",         "QUALIFIED", "CLM-001 valide",        "EU",     "PASS",             "", ""),
        ("T-002", "Eco-responsable",                                   "ABSOLUTE",  "-",                     "EU",     "BLOCK",            "", ""),
        ("T-003", "Sustainable collection",                            "ABSOLUTE",  "-",                     "EU",     "BLOCK",            "", ""),
        ("T-004", "Empreinte carbone neutre sur cette collection",    "ABSOLUTE",  "-",                     "FR",     "BLOCK",            "", ""),
        ("T-005", "Cuir certifie LWG (certif 2025)",                   "QUALIFIED", "CLM-002 valide",        "EU",     "PASS_WITH_REVIEW", "", ""),
        ("T-006", "Serie limitee 48 pieces",                           "QUALIFIED", "CLM-012 valide",        "EU",     "PASS_WITH_REVIEW", "", ""),
        ("T-007", "Fully sustainable (EN)",                            "ABSOLUTE",  "-",                     "EU + UK","BLOCK",            "", ""),
        ("T-008", "Fabrique en France avec label IGP",                 "QUALIFIED", "CLM-007 valide",        "FR",     "PASS_WITH_REVIEW", "", ""),
        ("T-009", "Cruelty-free sur toute la gamme",                   "ABSOLUTE",  "CLM-011 valide partiel","EU",     "PASS_WITH_REVIEW", "", ""),
        ("T-010", "Diamants sans conflit (Kimberley)",                 "QUALIFIED", "CLM-005 valide",        "EU",     "PASS_WITH_REVIEW", "", ""),
    ]
    last = ecrire_donnees(ws, 5, tests, align=AL)
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=8, value=f'=IF(G{r}="","-",IF(F{r}=G{r},"PASS","FAIL"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS
    ajouter_table(ws, "tblTestsetClaims", f"A4:H{last}")
    attach_dv(ws, f"C5:C{last}", WORDING_TYPE)
    attach_dv(ws, f"F5:F{last}", ["PASS", "PASS_WITH_REVIEW", "BLOCK"])
    attach_dv(ws, f"G5:G{last}", ["", "PASS", "PASS_WITH_REVIEW", "BLOCK"])
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_juridiction(wb):
    ws = wb.create_sheet("10_JURIDICTION_MATRIX")
    titre(ws, "JURIDICTION MATRIX", "Mapping regle par claim par juridiction — ajout v1 suite analyse", 7)
    en_tetes(ws, 4, ["LIB_ID", "CLAIM_PATTERN", "EU", "FR", "UK", "US", "CH"],
             [12, 28, 18, 18, 18, 18, 18])
    matrix = [
        ("CL-LIB-001", "recycle",           "OK si %",          "OK si %",                        "OK si %",                       "OK si %",                     "OK si %"),
        ("CL-LIB-002", "durable",            "INTERDIT vague",   "INTERDIT vague",                 "INTERDIT (CMA)",                "A QUALIFIER (FTC)",           "INTERDIT vague"),
        ("CL-LIB-003", "eco-responsable",    "INTERDIT vague",   "INTERDIT vague (Loi Climat)",   "INTERDIT (CMA)",                "N/A",                         "INTERDIT vague"),
        ("CL-LIB-004", "sustainable",        "INTERDIT vague",   "INTERDIT vague",                 "INTERDIT (CMA)",                "REVIEW FTC",                  "INTERDIT"),
        ("CL-LIB-007", "carbon neutral",     "REVIEW",           "INTERDIT (Loi Climat 2023)",    "REVIEW (ASA)",                  "A QUALIFIER",                 "REVIEW"),
        ("CL-LIB-008", "empreinte carbone",  "OK si scope+data", "OK si scope+data",               "OK si scope+data",              "OK si scope+data",            "OK si scope+data"),
        ("CL-LIB-009", "LWG certifie",       "OK si certif",     "OK si certif",                   "OK si certif",                  "OK si certif",                "OK si certif"),
        ("CL-LIB-010", "fabrique en France", "OK si IGP/atelier","OK si IGP/atelier (DGCCRF)",    "OK",                            "OK",                          "OK"),
        ("CL-LIB-013", "cruelty-free",       "REVIEW",           "REVIEW",                         "REVIEW",                        "REVIEW (FDA)",                "REVIEW"),
        ("CL-LIB-014", "vegan",              "REVIEW",           "REVIEW",                         "REVIEW",                        "OK si certif",                "REVIEW"),
    ]
    last = ecrire_donnees(ws, 5, matrix, align=AL)
    ajouter_table(ws, "tblJuridiction", f"A4:G{last}")


def onglet_learnings(wb):
    """R-13 : boucle d'apprentissage AG-005."""
    ws = wb.create_sheet("11_LEARNINGS")
    titre(ws, "LEARNINGS — AG-005", "Capitalisation post-run : motifs BLOCK, ajustements claim library / juridiction", 8)
    en_tetes(ws, 4, ["LEARN_ID", "DATE", "SOURCE_RUN", "MOTIF", "FREQUENCE_MOIS", "ACTION_PROPOSEE", "OWNER", "STATUT"],
             [12, 12, 18, 40, 16, 44, 18, 14])
    lrn = [
        ("LRN-C-001", D(today_iso()), "R-20260416-004", "'carbon neutral' non intercepte avant draft", 4, "Pre-hook AG-002 + Hard-fail EN",      "AI Ops",  "OPEN"),
        ("LRN-C-002", D(today_iso()), "R-20260416-002", "'zero dechet' ABSOLUTE systematiquement BLOCK, alors qu'audit existe",  2, "Ajouter CL-LIB-006 evidence pathway", "ESG Lead","OPEN"),
        ("LRN-C-003", D(today_iso()), "-",              "Couverture juridiction CH pas encore enseignee", 0, "Seed col CH dans 10_JURIDICTION",    "Legal",   "OPEN"),
    ]
    last = ecrire_donnees(ws, 5, lrn, align=AL)
    convertir_dates(ws, 5, last, date_cols=[2])
    ajouter_table(ws, "tblLearningsAG005", f"A4:H{last}")
    attach_dv(ws, f"H5:H{last}", ["OPEN", "IN_REVIEW", "ACCEPTED", "REJECTED", "DONE"])


def build():
    wb = Workbook()
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_claim_intake(wb)
    onglet_claim_library(wb)
    onglet_evidence_match(wb)
    onglet_risk_engine(wb)
    onglet_review_queue(wb)
    onglet_decision_log(wb)
    onglet_dashboard(wb)
    onglet_testset(wb)
    onglet_juridiction(wb)
    onglet_learnings(wb)

    order = ["0_README", "1_PARAMS", "2_CLAIM_INTAKE", "3_CLAIM_LIBRARY",
             "4_EVIDENCE_MATCH", "5_RISK_ENGINE", "6_REVIEW_QUEUE",
             "7_DECISION_LOG", "8_DASHBOARD", "9_TESTSET", "10_JURIDICTION_MATRIX", "11_LEARNINGS"]
    wb._sheets = [wb[n] for n in order]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
