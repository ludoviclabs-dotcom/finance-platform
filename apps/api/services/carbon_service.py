from __future__ import annotations

import os
import json
import re
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class CarbonServiceError(Exception):
    """Base exception for Carbon backend service failures."""


DEFAULT_WORKBOOK_ROOT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\Carbon and Co")

WORKBOOK_FILES = {
    "carbon": "CarbonCo_Calcul_Carbone_v2.xlsx",
    "esg": "CarbonCo_ESG_Social.xlsx",
    "finance": "CarbonCo_Finance_DPP_v1_3.xlsx",
}

REQUIRED_SHEETS = {
    "carbon": {"Couverture", "Sommaire", "Paramètres", "Synthese_GES", "Energie", "Claude Log"},
    "esg": {"Couverture", "Sommaire", "Liaison_Donnees", "Claude Log"},
    "finance": {"Couverture", "Sommaire", "Liaison_Donnees", "Claude Log"},
}

ESG_LIAISON_EXPECTED = {
    "E16": "Synthese_GES!C15",
    "E17": "Synthese_GES!C17",
    "E18": "Synthese_GES!C35",
    "E22": "Energie!E20",
    "E23": "Energie!E19",
    "E25": "Taxonomie!E27",
    "E26": "CBAM!M24",
}

FINANCE_LIAISON_EXPECTED = {
    "B19": "Total S1+S2(LB)+S3 (tCO2e)",
    "E16": "Synthese_GES!C15",
    "E17": "Synthese_GES!C17",
    "E18": "Synthese_GES!C35",
    "E19": "Calcul auto (LB)",
    "E23": "Energie!E20",
    "E24": "Taxonomie!E27",
    "E25": "Taxonomie!E28",
    "E26": "CBAM!M24",
    "E30": "Trajectoire_SBTi!B4",
}

CANONICAL_CC_RANGES = {
    "CC_Raison_Sociale": "'Paramètres'!$B$4",
    "CC_Annee_Reporting": "'Paramètres'!$B$7",
    "CC_CA_Net": "'Paramètres'!$B$9",
    "CC_ETP": "'Paramètres'!$B$11",
    "CC_Secteur_Activite": "'Paramètres'!$B$6",
    "CC_Secteur_NAF": "'Paramètres'!$C$6",
    "CC_Surface_Totale": "'Paramètres'!$B$12",
    "CC_CapEx_Total": "'Paramètres'!$B$14",
    "CC_OpEx_Eligible_Taxo": "'Paramètres'!$B$15",
    "CC_GES_Scope1": "'Synthese_GES'!$C$10",
    "CC_GES_Scope2_LB": "'Synthese_GES'!$C$15",
    "CC_GES_Scope2_MB": "'Synthese_GES'!$C$17",
    "CC_GES_Scope3": "'Synthese_GES'!$C$35",
    "CC_GES_Total_S123": "'Synthese_GES'!$C$47",
    "CC_Intensite_CA": "'Synthese_GES'!$C$50",
    "CC_Intensite_ETP": "'Synthese_GES'!$C$51",
    "CC_Part_Scope1": "'Synthese_GES'!$C$53",
    "CC_Part_Scope2": "'Synthese_GES'!$C$54",
    "CC_Part_Scope3": "'Synthese_GES'!$C$55",
    "CC_Conso_Energie_MWh": "'Energie'!$E$19",
    "CC_Part_ENR": "'Energie'!$E$20",
    "CC_Taxo_CA_Aligne": "'Taxonomie'!$E$27",
    "CC_Taxo_CapEx_Aligne": "'Taxonomie'!$E$28",
    "CC_Taxo_OpEx_Aligne": "'Taxonomie'!$E$29",
    "CC_CBAM_Cout_Estime": "'CBAM'!$M$24",
    "CC_SBTI_Annee_Baseline": "'Trajectoire_SBTi'!$B$4",
    "CC_SBTI_Baseline_S12": "'Trajectoire_SBTi'!$B$5",
    "CC_SBTI_Baseline_S3": "'Trajectoire_SBTi'!$B$6",
    "CC_SBTI_Taux_S12": "'Trajectoire_SBTi'!$B$8",
    "CC_SBTI_Taux_S3": "'Trajectoire_SBTi'!$B$9",
}

FALLBACK_CELLS = {
    "CC_Raison_Sociale": ("Paramètres", "B4"),
    "CC_Annee_Reporting": ("Paramètres", "B7"),
    "CC_CA_Net": ("Paramètres", "B9"),
    "CC_ETP": ("Paramètres", "B11"),
    "CC_Secteur_Activite": ("Paramètres", "B6"),
    "CC_Secteur_NAF": ("Paramètres", "C6"),
    "CC_Surface_Totale": ("Paramètres", "B12"),
    "CC_CapEx_Total": ("Paramètres", "B14"),
    "CC_OpEx_Eligible_Taxo": ("Paramètres", "B15"),
    "CC_GES_Scope1": ("Synthese_GES", "C10"),
    "CC_GES_Scope2_LB": ("Synthese_GES", "C15"),
    "CC_GES_Scope2_MB": ("Synthese_GES", "C17"),
    "CC_GES_Scope3": ("Synthese_GES", "C35"),
    "CC_GES_Total_S123": ("Synthese_GES", "C47"),
    "CC_Intensite_CA": ("Synthese_GES", "C50"),
    "CC_Intensite_ETP": ("Synthese_GES", "C51"),
    "CC_Part_Scope1": ("Synthese_GES", "C53"),
    "CC_Part_Scope2": ("Synthese_GES", "C54"),
    "CC_Part_Scope3": ("Synthese_GES", "C55"),
    "CC_Conso_Energie_MWh": ("Energie", "E19"),
    "CC_Part_ENR": ("Energie", "E20"),
    "CC_Taxo_CA_Aligne": ("Taxonomie", "E27"),
    "CC_Taxo_CapEx_Aligne": ("Taxonomie", "E28"),
    "CC_Taxo_OpEx_Aligne": ("Taxonomie", "E29"),
    "CC_CBAM_Cout_Estime": ("CBAM", "M24"),
    "CC_SBTI_Annee_Baseline": ("Trajectoire_SBTi", "B4"),
    "CC_SBTI_Baseline_S12": ("Trajectoire_SBTi", "B5"),
    "CC_SBTI_Baseline_S3": ("Trajectoire_SBTi", "B6"),
    "CC_SBTI_Taux_S12": ("Trajectoire_SBTi", "B8"),
    "CC_SBTI_Taux_S3": ("Trajectoire_SBTi", "B9"),
}

SNAPSHOT_FIELD_TO_KEY = {
    "company.name": "CC_Raison_Sociale",
    "company.reportingYear": "CC_Annee_Reporting",
    "company.sectorActivity": "CC_Secteur_Activite",
    "company.nafCode": "CC_Secteur_NAF",
    "company.revenueNetEur": "CC_CA_Net",
    "company.fte": "CC_ETP",
    "company.surfaceSqm": "CC_Surface_Totale",
    "company.capexTotalEur": "CC_CapEx_Total",
    "company.opexEligibleTaxoEur": "CC_OpEx_Eligible_Taxo",
    "carbon.scope1Tco2e": "CC_GES_Scope1",
    "carbon.scope2LbTco2e": "CC_GES_Scope2_LB",
    "carbon.scope2MbTco2e": "CC_GES_Scope2_MB",
    "carbon.scope3Tco2e": "CC_GES_Scope3",
    "carbon.totalS123Tco2e": "CC_GES_Total_S123",
    "carbon.intensityRevenueTco2ePerMEur": "CC_Intensite_CA",
    "carbon.intensityFteTco2ePerFte": "CC_Intensite_ETP",
    "carbon.shareScope1Pct": "CC_Part_Scope1",
    "carbon.shareScope2Pct": "CC_Part_Scope2",
    "carbon.shareScope3Pct": "CC_Part_Scope3",
    "energy.consumptionMWh": "CC_Conso_Energie_MWh",
    "energy.renewableSharePct": "CC_Part_ENR",
    "taxonomy.turnoverAlignedPct": "CC_Taxo_CA_Aligne",
    "taxonomy.capexAlignedPct": "CC_Taxo_CapEx_Aligne",
    "taxonomy.opexAlignedPct": "CC_Taxo_OpEx_Aligne",
    "cbam.estimatedCostEur": "CC_CBAM_Cout_Estime",
    "sbti.baselineYear": "CC_SBTI_Annee_Baseline",
    "sbti.baselineS12Tco2e": "CC_SBTI_Baseline_S12",
    "sbti.baselineS3Tco2e": "CC_SBTI_Baseline_S3",
    "sbti.targetReductionS12Pct": "CC_SBTI_Taux_S12",
    "sbti.targetReductionS3Pct": "CC_SBTI_Taux_S3",
}

REQUIRED_SNAPSHOT_FIELDS = {
    "company.name",
    "company.reportingYear",
    "company.revenueNetEur",
    "carbon.totalS123Tco2e",
    "carbon.intensityRevenueTco2ePerMEur",
    "energy.renewableSharePct",
    "taxonomy.turnoverAlignedPct",
}

CELL_REF_RE = re.compile(
    r"(?:(?:'[^']+'|[A-Za-zÀ-ÿ_][A-Za-zÀ-ÿ0-9_ ]*)!)?\$?[A-Z]{1,3}\$?\d+"
)


def get_workbook_root() -> Path:
    root = os.environ.get("CARBONCO_WORKBOOK_ROOT")
    return Path(root) if root else DEFAULT_WORKBOOK_ROOT


def get_workbook_paths() -> dict[str, Path]:
    root = get_workbook_root()
    return {key: root / filename for key, filename in WORKBOOK_FILES.items()}


def _load_workbook(path: Path, *, data_only: bool) -> Workbook:
    return load_workbook(path, read_only=True, data_only=data_only)


def _normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            numeric = float(stripped.replace(",", "."))
            if numeric.is_integer():
                return int(numeric)
            return round(numeric, 4)
        except Exception:
            return stripped
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            return round(value, 4)
        return value
    return value


def _get_defined_name_formula(workbook: Workbook, name: str) -> str | None:
    defined = workbook.defined_names.get(name)
    if defined is None:
        return None
    return getattr(defined, "attr_text", None)


def _get_named_value(workbook: Workbook, name: str) -> Any:
    defined = workbook.defined_names.get(name)
    if defined is None:
        return None
    destinations = list(defined.destinations)
    if not destinations:
        return None
    sheet_name, cell_ref = destinations[0]
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name][cell_ref].value


def _get_cell_value(workbook: Workbook, sheet_name: str, cell_ref: str) -> Any:
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name][cell_ref].value


def _split_excel_args(expr: str) -> list[str]:
    args: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    i = 0
    while i < len(expr):
        char = expr[i]
        if char == '"':
            in_string = not in_string
            current.append(char)
        elif not in_string and char == "(":
            depth += 1
            current.append(char)
        elif not in_string and char == ")":
            depth -= 1
            current.append(char)
        elif not in_string and depth == 0 and char == ",":
            args.append("".join(current).strip())
            current = []
        else:
            current.append(char)
        i += 1
    if current:
        args.append("".join(current).strip())
    return args


def _strip_sheet_name(sheet_ref: str) -> str:
    cleaned = sheet_ref.strip()
    if cleaned.startswith("'") and cleaned.endswith("'"):
        return cleaned[1:-1]
    return cleaned


def _split_cell_reference(ref: str, current_sheet: str) -> tuple[str, str]:
    cleaned = ref.replace("$", "")
    if "!" in cleaned:
        sheet_name, cell_ref = cleaned.split("!", 1)
        return _strip_sheet_name(sheet_name), cell_ref
    return current_sheet, cleaned


def _coerce_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except Exception:
        return 0.0


def _replace_excel_equals(expr: str) -> str:
    expr = expr.replace("<>", "!=")
    return re.sub(r"(?<![<>=])=(?!=)", "==", expr)


def _looks_like_condition(expr: str) -> bool:
    stripped = expr.strip()
    if stripped.startswith(("IF(", "SUM(")):
        return False
    if stripped.startswith(("OR(", "AND(")):
        return True
    return any(operator in stripped for operator in ("<>", ">=", "<=", "=", ">", "<"))


def _evaluate_formula_cell(
    workbook: Workbook,
    sheet_name: str,
    cell_ref: str,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
) -> Any:
    key = (sheet_name, cell_ref)
    if key in cache:
        return cache[key]
    if key in stack:
        raise CarbonServiceError(f"Circular formula reference detected at {sheet_name}!{cell_ref}")

    stack.add(key)
    try:
        raw_value = _get_cell_value(workbook, sheet_name, cell_ref)
        result = _evaluate_excel_value(workbook, sheet_name, raw_value, cache, stack)
        cache[key] = result
        return result
    finally:
        stack.remove(key)


def _evaluate_range_sum(
    workbook: Workbook,
    sheet_name: str,
    start_ref: str,
    end_ref: str,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
) -> float:
    ws = workbook[sheet_name]
    start = ws[start_ref]
    end = ws[end_ref]
    total = 0.0
    for row in ws.iter_rows(
        min_row=start.row,
        max_row=end.row,
        min_col=start.column,
        max_col=end.column,
    ):
        for cell in row:
            total += _coerce_number(
                _evaluate_formula_cell(workbook, sheet_name, cell.coordinate, cache, stack)
            )
    return total


def _evaluate_range_values(
    workbook: Workbook,
    sheet_name: str,
    start_ref: str,
    end_ref: str,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
) -> list[float]:
    ws = workbook[sheet_name]
    start = ws[start_ref]
    end = ws[end_ref]
    values: list[float] = []
    for row in ws.iter_rows(
        min_row=start.row,
        max_row=end.row,
        min_col=start.column,
        max_col=end.column,
    ):
        for cell in row:
            values.append(
                _coerce_number(
                    _evaluate_formula_cell(workbook, sheet_name, cell.coordinate, cache, stack)
                )
            )
    return values


def _evaluate_condition(
    workbook: Workbook,
    sheet_name: str,
    condition_expr: str,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
) -> bool:
    stripped = condition_expr.strip()
    if stripped.startswith("OR(") and stripped.endswith(")"):
        args = _split_excel_args(stripped[3:-1])
        return any(_evaluate_condition(workbook, sheet_name, arg, cache, stack) for arg in args)
    if stripped.startswith("AND(") and stripped.endswith(")"):
        args = _split_excel_args(stripped[4:-1])
        return all(_evaluate_condition(workbook, sheet_name, arg, cache, stack) for arg in args)

    prepared = _replace_excel_equals(condition_expr)
    prepared = prepared.replace("TRUE", "True").replace("FALSE", "False")

    def repl(match: re.Match[str]) -> str:
        ref = match.group(0)
        ref_sheet, ref_cell = _split_cell_reference(ref, sheet_name)
        value = _evaluate_formula_cell(workbook, ref_sheet, ref_cell, cache, stack)
        if isinstance(value, str):
            return repr(value)
        if value is None:
            return repr("")
        return repr(value)

    prepared = CELL_REF_RE.sub(repl, prepared)
    return bool(eval(prepared, {"__builtins__": {}}, {}))


def _evaluate_excel_formula(
    workbook: Workbook,
    sheet_name: str,
    formula: str,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
) -> Any:
    expr = formula[1:].strip()

    if expr.startswith("IF(") and expr.endswith(")"):
        args = _split_excel_args(expr[3:-1])
        if len(args) != 3:
            raise CarbonServiceError(f"Unsupported IF formula: {formula}")
        condition = _evaluate_condition(workbook, sheet_name, args[0], cache, stack)
        branch = args[1] if condition else args[2]
        return _evaluate_excel_value(workbook, sheet_name, branch, cache, stack)

    if expr.startswith(("SUM(", "MAX(", "MIN(")) and expr.endswith(")"):
        function_name = expr[:3]
        inner = expr[4:-1]
        if expr.startswith("MAX("):
            function_name = "MAX"
            inner = expr[4:-1]
        elif expr.startswith("MIN("):
            function_name = "MIN"
            inner = expr[4:-1]
        parts = _split_excel_args(expr[4:-1])
        values: list[float] = []
        for part in parts:
            if ":" in part:
                start_ref, end_ref = [segment.strip() for segment in part.split(":", 1)]
                start_sheet, start_cell = _split_cell_reference(start_ref, sheet_name)
                end_sheet, end_cell = _split_cell_reference(end_ref, sheet_name)
                if start_sheet != end_sheet:
                    raise CarbonServiceError(f"Unsupported cross-sheet {function_name} range: {formula}")
                values.extend(
                    _evaluate_range_values(workbook, start_sheet, start_cell, end_cell, cache, stack)
                )
            else:
                if CELL_REF_RE.fullmatch(part.strip()):
                    ref_sheet, ref_cell = _split_cell_reference(part, sheet_name)
                    values.append(
                        _evaluate_formula_cell(workbook, ref_sheet, ref_cell, cache, stack)
                    )
                else:
                    values.append(
                        _evaluate_excel_value(workbook, sheet_name, part.strip(), cache, stack)
                    )
        numeric_values = [_coerce_number(value) for value in values]
        if function_name == "SUM":
            return sum(numeric_values)
        if function_name == "MAX":
            return max(numeric_values) if numeric_values else 0.0
        if function_name == "MIN":
            return min(numeric_values) if numeric_values else 0.0

    if expr.startswith("OR(") and expr.endswith(")"):
        args = _split_excel_args(expr[3:-1])
        return any(_evaluate_condition(workbook, sheet_name, arg, cache, stack) for arg in args)

    if expr.startswith("AND(") and expr.endswith(")"):
        args = _split_excel_args(expr[4:-1])
        return all(_evaluate_condition(workbook, sheet_name, arg, cache, stack) for arg in args)

    prepared = expr

    def repl(match: re.Match[str]) -> str:
        ref = match.group(0)
        ref_sheet, ref_cell = _split_cell_reference(ref, sheet_name)
        value = _evaluate_formula_cell(workbook, ref_sheet, ref_cell, cache, stack)
        return str(_coerce_number(value))

    prepared = CELL_REF_RE.sub(repl, prepared)
    prepared = prepared.replace("TRUE", "True").replace("FALSE", "False")
    return eval(prepared, {"__builtins__": {}}, {})


def _evaluate_excel_value(
    workbook: Workbook,
    sheet_name: str,
    raw_value: Any,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
) -> Any:
    if not isinstance(raw_value, str) or not raw_value.startswith("="):
        return raw_value

    stripped = raw_value.strip()
    if stripped.startswith('="') and stripped.endswith('"'):
        return stripped[2:-1]
    if _looks_like_condition(stripped[1:]):
        return _evaluate_condition(workbook, sheet_name, stripped[1:], cache, stack)
    return _evaluate_excel_formula(workbook, sheet_name, stripped, cache, stack)


def _evaluate_contract_key(workbook: Workbook, contract_key: str) -> Any:
    fallback = FALLBACK_CELLS.get(contract_key)
    if fallback is None:
        return None
    cache: dict[tuple[str, str], Any] = {}
    return _evaluate_formula_cell(workbook, fallback[0], fallback[1], cache, set())


def _collect_values_with_excel_com(path: Path, key_to_cell: dict[str, tuple[str, str]]) -> dict[str, Any]:
    """Read workbook values through installed Excel when formula caches are empty.

    This is intentionally a Windows-local fallback for the user's current setup.
    It is only used when openpyxl cannot resolve calculated values.
    """
    if not key_to_cell:
        return {}

    cell_map_json = json.dumps(
        {key: {"sheet": sheet, "cell": cell} for key, (sheet, cell) in key_to_cell.items()},
        ensure_ascii=False,
    )
    script = f"""
$ErrorActionPreference = 'Stop'
$cellMap = @'
{cell_map_json}
'@ | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$excel.AskToUpdateLinks = $false
$excel.EnableEvents = $false
$excel.ScreenUpdating = $false
$excel.Calculation = -4105
$wb = $excel.Workbooks.Open('{str(path).replace("'", "''")}', $null, $true)
try {{
  $result = [ordered]@{{}}
  foreach ($entry in $cellMap.PSObject.Properties) {{
    $sheet = $entry.Value.sheet
    $cell = $entry.Value.cell
    $value = $wb.Worksheets.Item($sheet).Range($cell).Value2
    $result[$entry.Name] = $value
  }}
  $result | ConvertTo-Json -Compress -Depth 5
}}
finally {{
  $wb.Close($false)
  $excel.Quit()
}}
"""
    completed = subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-NonInteractive",
            "-Command",
            "-",
        ],
        input=script,
        capture_output=True,
        text=True,
        check=False,
    )
    if completed.returncode != 0:
        raise CarbonServiceError(
            "Excel desktop fallback failed: " + (completed.stderr.strip() or completed.stdout.strip())
        )

    stdout = completed.stdout.strip()
    if not stdout:
        return {}
    return json.loads(stdout)


def _set_nested(target: dict[str, Any], path: str, value: Any) -> None:
    parts = path.split(".")
    current = target
    for part in parts[:-1]:
        current = current.setdefault(part, {})
    current[parts[-1]] = value


def _validate_required_sheets(summary: dict[str, Any], required: set[str]) -> list[str]:
    current = set(summary["sheet_names"])
    missing = sorted(required - current)
    return [f"Missing required sheet in {summary['workbook']}: {sheet}" for sheet in missing]


def _validate_named_ranges(summary: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    current = summary["defined_names"]
    for name, expected_target in CANONICAL_CC_RANGES.items():
        actual_target = current.get(name)
        if actual_target is None:
            failures.append(f"Missing named range: {name}")
            continue
        if actual_target != expected_target:
            failures.append(
                f"Named range mismatch: {name} expected {expected_target} got {actual_target}"
            )
    return failures


def _validate_liaison_cells(path: Path, expected: dict[str, str]) -> list[str]:
    workbook = _load_workbook(path, data_only=False)
    try:
        worksheet = workbook["Liaison_Donnees"]
        failures = []
        for cell_ref, expected_value in expected.items():
            actual_value = worksheet[cell_ref].value
            if actual_value != expected_value:
                failures.append(
                    f"Liaison mismatch {path.name} {cell_ref}: expected {expected_value!r} got {actual_value!r}"
                )
        return failures
    finally:
        workbook.close()


def validate_master_workbooks() -> dict[str, Any]:
    paths = get_workbook_paths()
    checks: list[dict[str, Any]] = []
    failures: list[str] = []
    warnings: list[str] = []

    for key, path in paths.items():
        if not path.exists():
            failures.append(f"Missing workbook: {path}")
            continue

        workbook = _load_workbook(path, data_only=False)
        try:
            sheet_names = list(workbook.sheetnames)
            defined_names = {
                name: _get_defined_name_formula(workbook, name)
                for name in workbook.defined_names.keys()
            }
            summary = {
                "workbook": key,
                "path": str(path),
                "sheet_names": sheet_names,
                "defined_names": defined_names,
                "sheet_count": len(sheet_names),
                "named_range_count": len(defined_names),
                "has_claude_log": "Claude Log" in sheet_names,
            }
        finally:
            workbook.close()

        checks.append(
            {
                "workbook": key,
                "path": str(path),
                "sheet_count": summary["sheet_count"],
                "named_range_count": summary["named_range_count"],
                "has_claude_log": summary["has_claude_log"],
            }
        )
        failures.extend(_validate_required_sheets(summary, REQUIRED_SHEETS[key]))
        if key == "carbon":
            failures.extend(_validate_named_ranges(summary))
        elif key == "esg":
            failures.extend(_validate_liaison_cells(path, ESG_LIAISON_EXPECTED))
        elif key == "finance":
            failures.extend(_validate_liaison_cells(path, FINANCE_LIAISON_EXPECTED))

    status = "ok" if not failures else "failed"
    return {
        "status": status,
        "checks": checks,
        "failures": failures,
        "warnings": warnings,
    }


def _build_source_statuses(validation_result: dict[str, Any]) -> dict[str, dict[str, Any]]:
    checks_by_key = {item["workbook"]: item for item in validation_result["checks"]}
    paths = get_workbook_paths()
    statuses: dict[str, dict[str, Any]] = {}
    for key, path in paths.items():
        check = checks_by_key.get(key)
        status = "ok" if check else "missing"
        statuses[key] = {
            "filename": path.name,
            "status": status,
            "path": str(path),
            "sheet_count": check.get("sheet_count") if check else None,
            "named_range_count": check.get("named_range_count") if check else None,
            "has_claude_log": check.get("has_claude_log") if check else None,
        }
    return statuses


def build_carbon_snapshot() -> dict[str, Any]:
    validation_result = validate_master_workbooks()
    paths = get_workbook_paths()
    carbon_path = paths["carbon"]

    if not carbon_path.exists():
        raise CarbonServiceError(f"Carbon workbook not found: {carbon_path}")

    workbook_values = _load_workbook(carbon_path, data_only=True)
    workbook_formulas = _load_workbook(carbon_path, data_only=False)
    try:
        snapshot_data: dict[str, Any] = {
            "company": {},
            "carbon": {},
            "energy": {},
            "taxonomy": {},
            "cbam": {},
            "sbti": {},
        }
        warnings = list(validation_result["warnings"])
        missing_contract_keys: set[str] = set()
        formula_eval_keys: set[str] = set()

        for field_path, contract_key in SNAPSHOT_FIELD_TO_KEY.items():
            value = _get_named_value(workbook_values, contract_key)
            source_kind = "named_range"
            if value is None:
                fallback = FALLBACK_CELLS.get(contract_key)
                if fallback is not None:
                    source_kind = "fallback_cell"
                    value = _get_cell_value(workbook_values, fallback[0], fallback[1])
            if value is None:
                try:
                    source_kind = "formula_eval"
                    value = _evaluate_contract_key(workbook_formulas, contract_key)
                except Exception as exc:
                    warnings.append(f"Formula evaluation failed for {contract_key}: {exc}")
            if value is None:
                missing_contract_keys.add(contract_key)
            normalized = _normalize_value(value)
            _set_nested(snapshot_data, field_path, normalized)
            if normalized is not None and source_kind == "fallback_cell":
                warnings.append(f"Fallback cell used for {contract_key}")
            elif normalized is not None and source_kind == "formula_eval":
                formula_eval_keys.add(contract_key)

        if missing_contract_keys:
            fallback_cells = {
                key: FALLBACK_CELLS[key]
                for key in sorted(missing_contract_keys)
                if key in FALLBACK_CELLS
            }
            if fallback_cells:
                try:
                    com_values = _collect_values_with_excel_com(carbon_path, fallback_cells)
                except CarbonServiceError as exc:
                    warnings.append(str(exc))
                    com_values = {}
                if com_values:
                    warnings.append("Excel desktop fallback used to resolve calculated workbook values.")
                for field_path, contract_key in SNAPSHOT_FIELD_TO_KEY.items():
                    if contract_key not in com_values:
                        continue
                    normalized = _normalize_value(com_values[contract_key])
                    _set_nested(snapshot_data, field_path, normalized)

        if formula_eval_keys:
            warnings.append(
                "Formula evaluation fallback used for calculated workbook cells: "
                + ", ".join(sorted(formula_eval_keys))
            )

        if snapshot_data["company"].get("name") is None:
            snapshot_data["company"]["name"] = "Entreprise non renseignee"
            warnings.append(
                "Company name is empty in Paramètres!B4, default placeholder used in snapshot."
            )

        for field_path, contract_key in SNAPSHOT_FIELD_TO_KEY.items():
            normalized = snapshot_data
            for part in field_path.split("."):
                normalized = normalized.get(part) if isinstance(normalized, dict) else None
            if normalized is None:
                message = f"Missing snapshot value for {field_path} via {contract_key}"
                if field_path in REQUIRED_SNAPSHOT_FIELDS:
                    validation_result["failures"].append(message)
                else:
                    warnings.append(message)

        validation_status = "ok"
        if validation_result["failures"]:
            validation_status = "failed"
        elif warnings:
            validation_status = "warning"

        source_statuses = _build_source_statuses(validation_result)
        return {
            "snapshotVersion": "v1",
            "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "source": {
                "carbonWorkbook": source_statuses["carbon"],
                "esgWorkbook": source_statuses["esg"],
                "financeWorkbook": source_statuses["finance"],
            },
            "validation": {
                "status": validation_status,
                "failures": validation_result["failures"],
                "warnings": warnings,
            },
            **snapshot_data,
        }
    finally:
        workbook_values.close()
        workbook_formulas.close()
