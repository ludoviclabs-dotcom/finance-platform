"""
seed_factors.py — Peuplement de la table emission_factors.

Utilise les valeurs ADEME Base Empreinte v2025.0 (réf. publique).
Si le fichier Facteurs_Emission.xlsx est présent dans apps/api/data/factors/,
il est parsé en priorité. Sinon, les données embarquées sont utilisées.

Usage :
    python apps/api/scripts/seed_factors.py [--dry-run] [--version v2025.0]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from pathlib import Path

# Ajouter le répertoire api au path pour les imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from db.database import db_available, get_db

VERSION = "v2025.0"
SOURCE = "ADEME Base Empreinte"

# ---------------------------------------------------------------------------
# Données ADEME embarquées — valeurs de référence publiques v2025.0
# Format : (ef_code, label, scope, category, factor_kgco2e, unit)
# ---------------------------------------------------------------------------

_FACTORS_ENERGY_SCOPE1 = [
    # Combustibles fossiles — Scope 1
    ("ADEME.2025.COMBUST.GAZ_NAT",         "Gaz naturel (réseau)",                     1, "energy",    0.2014,  "kWh"),
    ("ADEME.2025.COMBUST.GAZ_NAT_MJ",      "Gaz naturel (énergie primaire)",            1, "energy",   55.6,    "MJ"),
    ("ADEME.2025.COMBUST.FIOUL_DOM",        "Fioul domestique",                          1, "energy",    0.3247,  "kWh"),
    ("ADEME.2025.COMBUST.FIOUL_DOM_L",      "Fioul domestique",                          1, "energy", 3188.0,    "m3"),
    ("ADEME.2025.COMBUST.FIOUL_LOURD",      "Fioul lourd (HFO)",                         1, "energy",    0.3375,  "kWh"),
    ("ADEME.2025.COMBUST.CHARBON",          "Charbon (combustion)",                      1, "energy",    0.3415,  "kWh"),
    ("ADEME.2025.COMBUST.CHARBON_T",        "Charbon (par tonne)",                       1, "energy", 2413.0,    "t"),
    ("ADEME.2025.COMBUST.LIGNITE",          "Lignite",                                   1, "energy",    0.3640,  "kWh"),
    ("ADEME.2025.COMBUST.GPL",              "GPL (propane/butane)",                      1, "energy",    0.2271,  "kWh"),
    ("ADEME.2025.COMBUST.GPL_KG",           "GPL (par kg)",                              1, "energy",    2.950,   "kg"),
    ("ADEME.2025.COMBUST.PROPANE",          "Propane",                                   1, "energy",    0.2271,  "kWh"),
    ("ADEME.2025.COMBUST.BUTANE",           "Butane",                                    1, "energy",    0.2271,  "kWh"),
    ("ADEME.2025.COMBUST.COKE",             "Coke de pétrole",                           1, "energy",    0.3937,  "kWh"),
    ("ADEME.2025.COMBUST.BIOMASSE_BOIS",    "Biomasse bois (bûches)",                    1, "energy",    0.0390,  "kWh"),
    ("ADEME.2025.COMBUST.BIOMASSE_GRANULE", "Granulés bois (pellets)",                   1, "energy",    0.0280,  "kWh"),
    ("ADEME.2025.COMBUST.BIOMASSE_PLAQUETTE","Plaquettes forestières",                   1, "energy",    0.0310,  "kWh"),
    ("ADEME.2025.COMBUST.GAZ_VILLE",        "Gaz de ville",                              1, "energy",    0.2250,  "kWh"),
    ("ADEME.2025.COMBUST.BIOGAS",           "Biogaz (valorisation déchets)",             1, "energy",    0.0495,  "kWh"),
    ("ADEME.2025.COMBUST.H2_ELEC",          "Hydrogène électrolyse (mix FR)",            1, "energy",    0.0320,  "kg"),
    ("ADEME.2025.COMBUST.H2_SMR",           "Hydrogène reformage gaz naturel",           1, "energy",    9.3,     "kg"),
    ("ADEME.2025.COMBUST.TOURBE",           "Tourbe",                                    1, "energy",    0.3820,  "kWh"),
    ("ADEME.2025.COMBUST.ANTHRACITE",       "Anthracite",                                1, "energy",    0.3550,  "kWh"),
    ("ADEME.2025.COMBUST.COKE_CHARBON",     "Coke de charbon",                           1, "energy",    0.4050,  "kWh"),
    ("ADEME.2025.COMBUST.GAZ_HAUT_FOURNEAU","Gaz de haut fourneau",                     1, "energy",    0.2600,  "kWh"),
]

_FACTORS_ENERGY_SCOPE2 = [
    # Électricité par pays — Scope 2 (marché résiduel, AIB 2024)
    ("ADEME.2025.ELEC.FR",     "Électricité France",           2, "energy",  0.0512, "kWh"),
    ("ADEME.2025.ELEC.DE",     "Électricité Allemagne",        2, "energy",  0.3850, "kWh"),
    ("ADEME.2025.ELEC.ES",     "Électricité Espagne",          2, "energy",  0.1850, "kWh"),
    ("ADEME.2025.ELEC.IT",     "Électricité Italie",           2, "energy",  0.2330, "kWh"),
    ("ADEME.2025.ELEC.BE",     "Électricité Belgique",         2, "energy",  0.1650, "kWh"),
    ("ADEME.2025.ELEC.NL",     "Électricité Pays-Bas",         2, "energy",  0.2900, "kWh"),
    ("ADEME.2025.ELEC.PL",     "Électricité Pologne",          2, "energy",  0.7150, "kWh"),
    ("ADEME.2025.ELEC.CZ",     "Électricité Tchéquie",         2, "energy",  0.4120, "kWh"),
    ("ADEME.2025.ELEC.AT",     "Électricité Autriche",         2, "energy",  0.1020, "kWh"),
    ("ADEME.2025.ELEC.CH",     "Électricité Suisse",           2, "energy",  0.0230, "kWh"),
    ("ADEME.2025.ELEC.GB",     "Électricité Royaume-Uni",      2, "energy",  0.2070, "kWh"),
    ("ADEME.2025.ELEC.SE",     "Électricité Suède",            2, "energy",  0.0130, "kWh"),
    ("ADEME.2025.ELEC.NO",     "Électricité Norvège",          2, "energy",  0.0080, "kWh"),
    ("ADEME.2025.ELEC.DK",     "Électricité Danemark",         2, "energy",  0.1650, "kWh"),
    ("ADEME.2025.ELEC.FI",     "Électricité Finlande",         2, "energy",  0.0710, "kWh"),
    ("ADEME.2025.ELEC.PT",     "Électricité Portugal",         2, "energy",  0.1780, "kWh"),
    ("ADEME.2025.ELEC.HU",     "Électricité Hongrie",          2, "energy",  0.2270, "kWh"),
    ("ADEME.2025.ELEC.RO",     "Électricité Roumanie",         2, "energy",  0.2710, "kWh"),
    ("ADEME.2025.ELEC.GR",     "Électricité Grèce",            2, "energy",  0.3710, "kWh"),
    ("ADEME.2025.ELEC.SK",     "Électricité Slovaquie",        2, "energy",  0.1440, "kWh"),
    ("ADEME.2025.ELEC.HR",     "Électricité Croatie",          2, "energy",  0.1890, "kWh"),
    ("ADEME.2025.ELEC.BG",     "Électricité Bulgarie",         2, "energy",  0.4080, "kWh"),
    ("ADEME.2025.ELEC.IE",     "Électricité Irlande",          2, "energy",  0.2950, "kWh"),
    ("ADEME.2025.ELEC.LT",     "Électricité Lituanie",         2, "energy",  0.1230, "kWh"),
    ("ADEME.2025.ELEC.LV",     "Électricité Lettonie",         2, "energy",  0.0820, "kWh"),
    ("ADEME.2025.ELEC.EE",     "Électricité Estonie",          2, "energy",  0.5220, "kWh"),
    ("ADEME.2025.ELEC.SI",     "Électricité Slovénie",         2, "energy",  0.2280, "kWh"),
    ("ADEME.2025.ELEC.LU",     "Électricité Luxembourg",       2, "energy",  0.0730, "kWh"),
    ("ADEME.2025.ELEC.US",     "Électricité États-Unis",       2, "energy",  0.3860, "kWh"),
    ("ADEME.2025.ELEC.CN",     "Électricité Chine",            2, "energy",  0.5810, "kWh"),
    ("ADEME.2025.ELEC.IN",     "Électricité Inde",             2, "energy",  0.7080, "kWh"),
    ("ADEME.2025.ELEC.BR",     "Électricité Brésil",           2, "energy",  0.0750, "kWh"),
    ("ADEME.2025.ELEC.JP",     "Électricité Japon",            2, "energy",  0.4300, "kWh"),
    ("ADEME.2025.ELEC.AU",     "Électricité Australie",        2, "energy",  0.6530, "kWh"),
    ("ADEME.2025.ELEC.CA",     "Électricité Canada",           2, "energy",  0.1300, "kWh"),
    ("ADEME.2025.ELEC.ZA",     "Électricité Afrique du Sud",   2, "energy",  0.9010, "kWh"),
    ("ADEME.2025.ELEC.MX",     "Électricité Mexique",          2, "energy",  0.4120, "kWh"),
    ("ADEME.2025.ELEC.TR",     "Électricité Turquie",          2, "energy",  0.4680, "kWh"),
    ("ADEME.2025.ELEC.EU_MOY", "Électricité Europe (moyenne)", 2, "energy",  0.2330, "kWh"),
    # Chaleur / vapeur
    ("ADEME.2025.CHALEUR.RESEAU_FR",  "Chaleur réseau urbain France",    2, "energy",  0.0678, "kWh"),
    ("ADEME.2025.CHALEUR.VAPEUR_GAZ", "Vapeur production gaz naturel",   2, "energy",  0.1870, "kWh"),
    ("ADEME.2025.CHALEUR.VAPEUR_CHAR","Vapeur production charbon",        2, "energy",  0.3020, "kWh"),
    ("ADEME.2025.FROID.RESEAU_FR",    "Réseau de froid urbain France",   2, "energy",  0.0234, "kWh"),
]

_FACTORS_TRANSPORT_SCOPE3 = [
    # Transport routier — Scope 3
    ("ADEME.2025.TRANSP.VP_THER_MOY",     "Voiture particulière thermique moyenne",     3, "transport",  0.2170, "km"),
    ("ADEME.2025.TRANSP.VP_THER_PETITE",  "Voiture particulière thermique petite",      3, "transport",  0.1560, "km"),
    ("ADEME.2025.TRANSP.VP_THER_GRANDE",  "Voiture particulière thermique grande",      3, "transport",  0.2580, "km"),
    ("ADEME.2025.TRANSP.VP_ELEC_MOY",     "Voiture électrique (mix FR)",               3, "transport",  0.0330, "km"),
    ("ADEME.2025.TRANSP.VP_HYBRIDE",      "Voiture hybride rechargeable",              3, "transport",  0.1050, "km"),
    ("ADEME.2025.TRANSP.MOTO_GROSSE",     "Moto > 250 cm³",                            3, "transport",  0.1910, "km"),
    ("ADEME.2025.TRANSP.MOTO_PETITE",     "Scooter / moto < 125 cm³",                  3, "transport",  0.0750, "km"),
    ("ADEME.2025.TRANSP.VELO_ELEC",       "Vélo électrique",                           3, "transport",  0.0110, "km"),
    ("ADEME.2025.TRANSP.BUS_DIESEL",      "Bus diesel (passager)",                     3, "transport",  0.1030, "km"),
    ("ADEME.2025.TRANSP.BUS_ELEC",        "Bus électrique (mix FR)",                   3, "transport",  0.0159, "km"),
    ("ADEME.2025.TRANSP.BUS_GNV",         "Bus GNV",                                   3, "transport",  0.0890, "km"),
    ("ADEME.2025.TRANSP.METRO",           "Métro / RER (passager)",                    3, "transport",  0.0037, "km"),
    ("ADEME.2025.TRANSP.TRAMWAY",         "Tramway (passager)",                        3, "transport",  0.0029, "km"),
    ("ADEME.2025.TRANSP.TER",             "Train régional TER (passager)",             3, "transport",  0.0289, "km"),
    ("ADEME.2025.TRANSP.TGV",             "TGV grande vitesse (passager)",             3, "transport",  0.0023, "km"),
    ("ADEME.2025.TRANSP.INTERCITES",      "Intercités (passager)",                     3, "transport",  0.0042, "km"),
    ("ADEME.2025.TRANSP.FRET_CAMION_PL",  "Camion poids lourd (tonne.km)",             3, "transport",  0.0952, "t.km"),
    ("ADEME.2025.TRANSP.FRET_CAMION_VUL", "Véhicule utilitaire léger (tonne.km)",      3, "transport",  0.1680, "t.km"),
    ("ADEME.2025.TRANSP.FRET_TRAIN",      "Train fret (tonne.km)",                     3, "transport",  0.0025, "t.km"),
    ("ADEME.2025.TRANSP.FRET_AVION_COURT","Avion fret court-courrier (tonne.km)",      3, "transport",  0.9540, "t.km"),
    ("ADEME.2025.TRANSP.FRET_AVION_LONG", "Avion fret long-courrier (tonne.km)",       3, "transport",  0.5420, "t.km"),
    ("ADEME.2025.TRANSP.FRET_MARITIME",   "Transport maritime conteneur (tonne.km)",   3, "transport",  0.0118, "t.km"),
    ("ADEME.2025.TRANSP.FRET_FLUVIAL",    "Fluvial (tonne.km)",                        3, "transport",  0.0327, "t.km"),
    # Aviation passagers
    ("ADEME.2025.AVION.COURT_ECO",        "Avion court-courrier économique",           3, "transport",  0.2580, "km"),
    ("ADEME.2025.AVION.COURT_BUS",        "Avion court-courrier business",             3, "transport",  0.5160, "km"),
    ("ADEME.2025.AVION.MOYEN_ECO",        "Avion moyen-courrier économique",           3, "transport",  0.1860, "km"),
    ("ADEME.2025.AVION.MOYEN_BUS",        "Avion moyen-courrier business",             3, "transport",  0.5590, "km"),
    ("ADEME.2025.AVION.LONG_ECO",         "Avion long-courrier économique",            3, "transport",  0.1520, "km"),
    ("ADEME.2025.AVION.LONG_BUS",         "Avion long-courrier business",              3, "transport",  0.4560, "km"),
    ("ADEME.2025.AVION.LONG_FIRST",       "Avion long-courrier première classe",       3, "transport",  0.6080, "km"),
    # Ferry / bateau
    ("ADEME.2025.FERRY.PASSAGER",         "Ferry passagers (transmanche)",             3, "transport",  0.1140, "km"),
    ("ADEME.2025.FERRY.CROISIERE",        "Croisière passager",                        3, "transport",  0.2560, "km"),
]

_FACTORS_MATERIALS = [
    # Matériaux de construction
    ("ADEME.2025.MAT.BETON_STD",          "Béton standard C25/30",                    3, "materials",  350.0,    "t"),
    ("ADEME.2025.MAT.BETON_HPC",          "Béton haute performance",                  3, "materials",  430.0,    "t"),
    ("ADEME.2025.MAT.BETON_RECYC",        "Béton avec agrégats recyclés",             3, "materials",  280.0,    "t"),
    ("ADEME.2025.MAT.ACIER_VIERGE",       "Acier vierge (bas fourneau)",              3, "materials", 1890.0,    "t"),
    ("ADEME.2025.MAT.ACIER_RECYCLE",      "Acier recyclé (four électrique)",          3, "materials",  560.0,    "t"),
    ("ADEME.2025.MAT.ACIER_INOX",         "Acier inoxydable",                         3, "materials", 2150.0,    "t"),
    ("ADEME.2025.MAT.ALUMINIUM_VIERGE",   "Aluminium primaire",                       3, "materials", 8900.0,    "t"),
    ("ADEME.2025.MAT.ALUMINIUM_RECYCLE",  "Aluminium recyclé",                        3, "materials",  590.0,    "t"),
    ("ADEME.2025.MAT.CUIVRE_VIERGE",      "Cuivre primaire",                          3, "materials", 2900.0,    "t"),
    ("ADEME.2025.MAT.CUIVRE_RECYCLE",     "Cuivre recyclé",                           3, "materials",  980.0,    "t"),
    ("ADEME.2025.MAT.PVC",                "PVC (chlorure de polyvinyle)",             3, "materials", 2420.0,    "t"),
    ("ADEME.2025.MAT.PE_HD",              "Polyéthylène haute densité",               3, "materials", 1850.0,    "t"),
    ("ADEME.2025.MAT.PE_LD",              "Polyéthylène basse densité",               3, "materials", 1920.0,    "t"),
    ("ADEME.2025.MAT.PP",                 "Polypropylène",                            3, "materials", 1840.0,    "t"),
    ("ADEME.2025.MAT.PS",                 "Polystyrène",                              3, "materials", 2270.0,    "t"),
    ("ADEME.2025.MAT.PET",                "PET (polyéthylène téréphtalate)",          3, "materials", 2730.0,    "t"),
    ("ADEME.2025.MAT.VERRE_CREUX",        "Verre creux (bouteilles/bocaux)",          3, "materials",  870.0,    "t"),
    ("ADEME.2025.MAT.VERRE_PLAT",         "Verre plat",                               3, "materials",  950.0,    "t"),
    ("ADEME.2025.MAT.PAPIER_VIERGE",      "Papier/carton vierge",                     3, "materials",  920.0,    "t"),
    ("ADEME.2025.MAT.PAPIER_RECYCLE",     "Papier/carton recyclé",                    3, "materials",  610.0,    "t"),
    ("ADEME.2025.MAT.CARTON_ONDUL",       "Carton ondulé",                            3, "materials",  750.0,    "t"),
    ("ADEME.2025.MAT.BOIS_MASSIF",        "Bois massif (construction)",               3, "materials",  132.0,    "t"),
    ("ADEME.2025.MAT.OSB",                "Panneau OSB",                              3, "materials",  485.0,    "t"),
    ("ADEME.2025.MAT.MDF",                "Panneau MDF",                              3, "materials",  520.0,    "t"),
    ("ADEME.2025.MAT.LAINE_VERRE",        "Laine de verre (isolation)",               3, "materials", 1230.0,    "t"),
    ("ADEME.2025.MAT.LAINE_ROCHE",        "Laine de roche (isolation)",               3, "materials", 1430.0,    "t"),
    ("ADEME.2025.MAT.XPS",                "Polystyrène extrudé XPS",                  3, "materials", 3300.0,    "t"),
    ("ADEME.2025.MAT.LIEGE",              "Liège (isolation)",                        3, "materials",  480.0,    "t"),
    ("ADEME.2025.MAT.CIMENT",             "Ciment Portland CEM I",                    3, "materials",  830.0,    "t"),
    ("ADEME.2025.MAT.CHAUX",              "Chaux",                                    3, "materials",  750.0,    "t"),
    ("ADEME.2025.MAT.BRIQUE_TERRECUITE",  "Brique terre cuite",                       3, "materials",  350.0,    "t"),
    ("ADEME.2025.MAT.TUILE",              "Tuile céramique",                          3, "materials",  420.0,    "t"),
    ("ADEME.2025.MAT.PLATRE",             "Plâtre",                                   3, "materials",  310.0,    "t"),
    ("ADEME.2025.MAT.MORTIER",            "Mortier de ciment",                        3, "materials",  270.0,    "t"),
    ("ADEME.2025.MAT.GRANULAT",           "Granulat naturel",                         3, "materials",   10.0,    "t"),
    ("ADEME.2025.MAT.SABLE",              "Sable naturel",                            3, "materials",    5.0,    "t"),
    ("ADEME.2025.MAT.ZINC",               "Zinc",                                     3, "materials", 3500.0,    "t"),
    ("ADEME.2025.MAT.PLOMB",              "Plomb",                                    3, "materials", 1980.0,    "t"),
    ("ADEME.2025.MAT.NICKEL",             "Nickel",                                   3, "materials",11900.0,    "t"),
    ("ADEME.2025.MAT.TITANE",             "Titane",                                   3, "materials", 7200.0,    "t"),
    # Textiles
    ("ADEME.2025.MAT.COTON_CONV",         "Coton conventionnel",                      3, "materials", 5890.0,    "t"),
    ("ADEME.2025.MAT.COTON_BIO",          "Coton biologique",                         3, "materials", 3600.0,    "t"),
    ("ADEME.2025.MAT.POLYESTER_VIERGE",   "Polyester vierge",                         3, "materials", 5500.0,    "t"),
    ("ADEME.2025.MAT.POLYESTER_RECYCLE",  "Polyester recyclé (rPET)",                 3, "materials", 1800.0,    "t"),
    ("ADEME.2025.MAT.NYLON_66",           "Nylon 6,6",                                3, "materials", 7900.0,    "t"),
    ("ADEME.2025.MAT.LAINE_VIERGE",       "Laine vierge",                             3, "materials", 8250.0,    "t"),
    ("ADEME.2025.MAT.LIN",                "Lin (fibre)",                              3, "materials",  590.0,    "t"),
    ("ADEME.2025.MAT.VISCOSE",            "Viscose (rayonne)",                        3, "materials", 4400.0,    "t"),
    ("ADEME.2025.MAT.ELASTHANE",          "Élasthane (spandex)",                      3, "materials",25000.0,    "t"),
]

_FACTORS_WASTE = [
    # Traitement des déchets
    ("ADEME.2025.DECH.DIB_ENFOUIS",       "DIB enfouis (ISDD)",                       3, "waste",     560.0,    "t"),
    ("ADEME.2025.DECH.DIB_INCINERE",      "DIB incinérés avec récupération énergie",  3, "waste",     304.0,    "t"),
    ("ADEME.2025.DECH.DIB_RECYCLE",       "DIB recyclés",                             3, "waste",     180.0,    "t"),
    ("ADEME.2025.DECH.DECHETS_MENAGERS",  "Déchets ménagers enfouis",                 3, "waste",     470.0,    "t"),
    ("ADEME.2025.DECH.PAPIER_ENFOUI",     "Papier/carton enfoui",                     3, "waste",     500.0,    "t"),
    ("ADEME.2025.DECH.PAPIER_INCINER",    "Papier/carton incinéré",                   3, "waste",     260.0,    "t"),
    ("ADEME.2025.DECH.PLASTIQUE_ENFOUI",  "Plastique enfoui",                         3, "waste",     110.0,    "t"),
    ("ADEME.2025.DECH.PLASTIQUE_INCINER", "Plastique incinéré",                       3, "waste",    2100.0,    "t"),
    ("ADEME.2025.DECH.PLASTIQUE_RECYCLE", "Plastique recyclé",                        3, "waste",      70.0,    "t"),
    ("ADEME.2025.DECH.VERRE_RECYCLE",     "Verre recyclé",                            3, "waste",      21.0,    "t"),
    ("ADEME.2025.DECH.ACIER_RECYCLE",     "Acier recyclé (déchet)",                   3, "waste",      28.0,    "t"),
    ("ADEME.2025.DECH.ALUM_RECYCLE",      "Aluminium recyclé (déchet)",               3, "waste",      40.0,    "t"),
    ("ADEME.2025.DECH.ORGANIQUE_COMPOST", "Déchets organiques compostage",            3, "waste",     130.0,    "t"),
    ("ADEME.2025.DECH.ORGANIQUE_METHAN",  "Déchets organiques méthanisation",         3, "waste",      80.0,    "t"),
    ("ADEME.2025.DECH.BOUE_EPURATION",    "Boues d'épuration (épandage)",             3, "waste",     380.0,    "t"),
    ("ADEME.2025.DECH.DECHETS_ELEC",      "DEEE (déchets électriques)",               3, "waste",     450.0,    "t"),
    ("ADEME.2025.DECH.DECHETS_BATIS",     "Déchets de démolition inertes",            3, "waste",      10.0,    "t"),
    ("ADEME.2025.DECH.EFFLUENTS_INDUST",  "Effluents industriels traitement",         3, "waste",     150.0,    "t"),
    ("ADEME.2025.DECH.HUILE_USAGEE",      "Huile usagée — incinération",              3, "waste",    2700.0,    "t"),
    ("ADEME.2025.DECH.SOLVANT_HALOG",     "Solvants halogénés — incinération",        3, "waste",    1800.0,    "t"),
]

_FACTORS_FOOD = [
    # Alimentation (ACV, de la ferme à l'assiette)
    ("ADEME.2025.ALIM.BOEUF_FR",          "Bœuf France (conventionnel)",              3, "food",     27400.0,   "t"),
    ("ADEME.2025.ALIM.BOEUF_BIO",         "Bœuf biologique France",                   3, "food",     22100.0,   "t"),
    ("ADEME.2025.ALIM.VEAU",              "Veau",                                     3, "food",     13800.0,   "t"),
    ("ADEME.2025.ALIM.PORC",              "Porc",                                     3, "food",      5700.0,   "t"),
    ("ADEME.2025.ALIM.VOLAILLE",          "Volaille (poulet, dinde)",                 3, "food",      4600.0,   "t"),
    ("ADEME.2025.ALIM.AGNEAU",            "Agneau",                                   3, "food",     24400.0,   "t"),
    ("ADEME.2025.ALIM.LAIT_VACHE",        "Lait de vache (conventionnel)",            3, "food",      2700.0,   "t"),
    ("ADEME.2025.ALIM.FROMAGE",           "Fromage à pâte dure",                      3, "food",     13500.0,   "t"),
    ("ADEME.2025.ALIM.OEUF",              "Œufs de poule",                            3, "food",      4100.0,   "t"),
    ("ADEME.2025.ALIM.BLE_TENDRE",        "Blé tendre (farine)",                      3, "food",       780.0,   "t"),
    ("ADEME.2025.ALIM.RIZ_BLANC",         "Riz blanc",                                3, "food",      2300.0,   "t"),
    ("ADEME.2025.ALIM.MAIS",              "Maïs grain",                               3, "food",       530.0,   "t"),
    ("ADEME.2025.ALIM.SOJA",              "Soja (tourteau)",                          3, "food",       540.0,   "t"),
    ("ADEME.2025.ALIM.POMME_DE_TERRE",    "Pommes de terre",                          3, "food",       210.0,   "t"),
    ("ADEME.2025.ALIM.TOMATE_SERRE",      "Tomates sous serre (chauffée)",            3, "food",      2700.0,   "t"),
    ("ADEME.2025.ALIM.TOMATE_PLEIN_AIR",  "Tomates plein air",                        3, "food",       320.0,   "t"),
    ("ADEME.2025.ALIM.LEGUMES_FEUILLES",  "Légumes feuilles (salade, épinards...)",   3, "food",       700.0,   "t"),
    ("ADEME.2025.ALIM.FRUITS_LOCAUX",     "Fruits locaux de saison",                  3, "food",       360.0,   "t"),
    ("ADEME.2025.ALIM.BANANE",            "Banane (importée)",                        3, "food",       840.0,   "t"),
    ("ADEME.2025.ALIM.CAFE_TORREFIE",     "Café torréfié",                            3, "food",     17000.0,   "t"),
    ("ADEME.2025.ALIM.CHOCOLAT_NOIR",     "Chocolat noir 70%",                        3, "food",      4600.0,   "t"),
    ("ADEME.2025.ALIM.HUILE_OLIVE",       "Huile d'olive",                            3, "food",      3300.0,   "t"),
    ("ADEME.2025.ALIM.HUILE_TOURNESOL",   "Huile de tournesol",                       3, "food",      1300.0,   "t"),
    ("ADEME.2025.ALIM.SUCRE_BLANC",       "Sucre blanc",                              3, "food",       570.0,   "t"),
    ("ADEME.2025.ALIM.PAIN_BLANC",        "Pain blanc",                               3, "food",       890.0,   "t"),
    ("ADEME.2025.ALIM.SAUMON_ELEV",       "Saumon d'élevage",                         3, "food",      5400.0,   "t"),
    ("ADEME.2025.ALIM.THON_PECHE",        "Thon pêché",                               3, "food",      9600.0,   "t"),
    ("ADEME.2025.ALIM.CREVETTE_ELEV",     "Crevette d'élevage",                       3, "food",     12500.0,   "t"),
    ("ADEME.2025.ALIM.EAU_MINERAL",       "Eau minérale en bouteille plastique",      3, "food",       580.0,   "t"),
    ("ADEME.2025.ALIM.VIN",               "Vin (AOC France)",                         3, "food",      1400.0,   "t"),
    ("ADEME.2025.ALIM.BIERE",             "Bière (brasserie artisanale)",             3, "food",       500.0,   "t"),
]

_FACTORS_WATER_SERVICES = [
    # Eau
    ("ADEME.2025.EAU.POTABLE_RESEAU",     "Eau potable réseau (traitement + distrib)", 3, "water",    0.3440, "m3"),
    ("ADEME.2025.EAU.PLUIE_TOITURE",      "Eau de pluie récupérée",                   3, "water",    0.0090, "m3"),
    ("ADEME.2025.EAU.EAUX_USEES",         "Traitement eaux usées",                    3, "water",    0.7080, "m3"),
    ("ADEME.2025.EAU.DESSALEMENT",        "Eau dessalée (osmose inverse)",            3, "water",    2.1000, "m3"),
    # Numérique
    ("ADEME.2025.NUM.SMARTPHONE",          "Smartphone (fabrication + usage 3 ans)",   3, "digital",   39.1,   "unité"),
    ("ADEME.2025.NUM.LAPTOP",              "Ordinateur portable (fabrication + usage)", 3, "digital",  156.0,   "unité"),
    ("ADEME.2025.NUM.DESKTOP",             "Ordinateur fixe + écran",                  3, "digital",  296.0,   "unité"),
    ("ADEME.2025.NUM.ECRAN_24",            "Écran 24 pouces",                          3, "digital",  128.0,   "unité"),
    ("ADEME.2025.NUM.TABLETTE",            "Tablette",                                 3, "digital",   63.0,   "unité"),
    ("ADEME.2025.NUM.SERVEUR_1U",          "Serveur 1U (fabrication)",                 3, "digital", 1050.0,   "unité"),
    ("ADEME.2025.NUM.STREAMING_VIDEO",     "Streaming vidéo HD (1h)",                  3, "digital",    0.036, "heure"),
    ("ADEME.2025.NUM.EMAIL_SIMPLE",        "Email sans pièce jointe",                  3, "digital",    0.000019, "unité"),
    ("ADEME.2025.NUM.EMAIL_PJ",            "Email avec pièce jointe (1 Mo)",           3, "digital",    0.000035, "unité"),
    ("ADEME.2025.NUM.WEB_PAGE",            "Consultation page web (50 Ko)",            3, "digital",    0.000002, "unité"),
    ("ADEME.2025.NUM.DATACENTER_FR",       "Datacenter France (par kWh IT)",           3, "digital",    0.0350, "kWh"),
    # Services financiers
    ("ADEME.2025.SRV.BANQUE_COMPTE",       "Compte bancaire particulier (an)",         3, "services",   8.3,   "an"),
    ("ADEME.2025.SRV.ASSURANCE_AUTO",      "Assurance auto (an)",                      3, "services",  42.0,   "an"),
    ("ADEME.2025.SRV.HOTEL_1NUIT_FR",      "Nuit d'hôtel France (chambre standard)",   3, "services",  14.2,   "nuit"),
    ("ADEME.2025.SRV.RESTAURANT_REPAS",    "Repas au restaurant (menu standard)",      3, "services",   1.4,   "repas"),
    ("ADEME.2025.SRV.COIFFEUR",            "Coupe de cheveux",                         3, "services",   0.8,   "visite"),
    # Construction / immobilier
    ("ADEME.2025.IMMO.LOGEMENT_NEUF_BBC",  "Construction logement BBC neuf (m²)",      3, "buildings", 580.0,  "m2"),
    ("ADEME.2025.IMMO.RENOVATION_THERM",   "Rénovation thermique (m²)",                3, "buildings", 140.0,  "m2"),
    ("ADEME.2025.IMMO.BUREAU_TERTIAIRE",   "Construction bureau tertiaire (m²)",       3, "buildings", 720.0,  "m2"),
    ("ADEME.2025.IMMO.ENTREPOT",           "Construction entrepôt logistique (m²)",    3, "buildings", 310.0,  "m2"),
    ("ADEME.2025.IMMO.ROUTE_BITUMEE",      "Route bitumée (km, 2 voies)",              3, "buildings",9200.0,  "km"),
]

_FACTORS_AGRICULTURE = [
    # Agriculture et intrants
    ("ADEME.2025.AGRI.ENGRAIS_AZOTE",      "Engrais azoté (urée)",                    3, "agriculture", 5300.0,  "t"),
    ("ADEME.2025.AGRI.ENGRAIS_PHOS",       "Engrais phosphaté (TSP)",                 3, "agriculture",  940.0,  "t"),
    ("ADEME.2025.AGRI.ENGRAIS_POTAS",      "Engrais potassique (KCl)",                3, "agriculture",  500.0,  "t"),
    ("ADEME.2025.AGRI.FUMIER_BOVIN",       "Fumier bovin (stockage couvert)",         3, "agriculture",  113.0,  "t"),
    ("ADEME.2025.AGRI.LISIER_PORC",        "Lisier de porc (stockage)",               3, "agriculture",   22.0,  "t"),
    ("ADEME.2025.AGRI.DIGESTAT_METH",      "Digestat méthanisation",                  3, "agriculture",   12.0,  "t"),
    ("ADEME.2025.AGRI.PESTICIDE_FOND",     "Pesticide fongicide",                     3, "agriculture", 9200.0,  "t"),
    ("ADEME.2025.AGRI.HERBICIDE",          "Herbicide",                               3, "agriculture", 7800.0,  "t"),
    ("ADEME.2025.AGRI.INSECTICIDE",        "Insecticide",                             3, "agriculture", 8100.0,  "t"),
    ("ADEME.2025.AGRI.IRRIGATION_SURF",    "Irrigation (eau de surface)",             3, "agriculture",    0.7,  "m3"),
    ("ADEME.2025.AGRI.GASOIL_AGRICOLE",    "Gasoil agricole (machines)",              1, "agriculture",    2.84, "L"),
    ("ADEME.2025.AGRI.FILM_PLASTIQUE",     "Film plastique agricole (PE)",            3, "agriculture", 1920.0,  "t"),
    ("ADEME.2025.AGRI.SEMENCES_BLE",       "Semences blé certifiées",                 3, "agriculture",  680.0,  "t"),
    ("ADEME.2025.AGRI.ELECTRICITE_SERRE",  "Électricité serre maraîchage (mix FR)",   2, "agriculture",    0.0512,"kWh"),
    ("ADEME.2025.AGRI.GAZ_SERRE",          "Gaz serre maraîchage (chauffage)",        1, "agriculture",    0.2014,"kWh"),
    # Élevage
    ("ADEME.2025.AGRI.VACHE_LAIT_AN",      "Vache laitière (émissions annuelles)",    1, "agriculture", 6800.0,  "tête"),
    ("ADEME.2025.AGRI.VACHE_VIANDE_AN",    "Vache à viande (émissions annuelles)",    1, "agriculture", 4200.0,  "tête"),
    ("ADEME.2025.AGRI.PORC_AN",            "Porc à l'engraissement",                  1, "agriculture",  140.0,  "tête"),
    ("ADEME.2025.AGRI.POULET_AN",          "Poulet de chair (bande)",                 1, "agriculture",    3.2,  "tête"),
    ("ADEME.2025.AGRI.BREBIS_AN",          "Brebis allaitante",                       1, "agriculture",  160.0,  "tête"),
]

_FACTORS_REFRIGERANTS = [
    # Fluides frigorigènes — PRG (GWP100)
    ("ADEME.2025.REFRIG.R134A",            "R-134a (HFC, PRG 1430)",                  1, "refrigerants", 1430.0, "kg"),
    ("ADEME.2025.REFRIG.R410A",            "R-410A (HFC, PRG 2088)",                  1, "refrigerants", 2088.0, "kg"),
    ("ADEME.2025.REFRIG.R404A",            "R-404A (HFC, PRG 3922)",                  1, "refrigerants", 3922.0, "kg"),
    ("ADEME.2025.REFRIG.R407C",            "R-407C (HFC, PRG 1774)",                  1, "refrigerants", 1774.0, "kg"),
    ("ADEME.2025.REFRIG.R32",              "R-32 (HFC, PRG 675)",                     1, "refrigerants",  675.0, "kg"),
    ("ADEME.2025.REFRIG.R1234YF",          "R-1234yf (HFO, PRG 4)",                   1, "refrigerants",    4.0, "kg"),
    ("ADEME.2025.REFRIG.R744",             "R-744 (CO2 transcritique, PRG 1)",        1, "refrigerants",    1.0, "kg"),
    ("ADEME.2025.REFRIG.R717",             "R-717 (ammoniac, PRG 0)",                 1, "refrigerants",    0.0, "kg"),
    ("ADEME.2025.REFRIG.R290",             "R-290 (propane, PRG 3)",                  1, "refrigerants",    3.0, "kg"),
    ("ADEME.2025.REFRIG.SF6",              "SF6 (hexafluorure de soufre, PRG 23900)", 1, "refrigerants",23900.0, "kg"),
    ("ADEME.2025.REFRIG.HFC23",            "HFC-23 (PRG 14800)",                      1, "refrigerants",14800.0, "kg"),
    ("ADEME.2025.REFRIG.PFC14",            "PFC-14 (PRG 7390)",                       1, "refrigerants", 7390.0, "kg"),
    ("ADEME.2025.REFRIG.N2O",              "Protoxyde d'azote N2O (PRG 265)",         1, "refrigerants",  265.0, "kg"),
    ("ADEME.2025.REFRIG.R22",              "R-22 (HCFC, PRG 1810)",                   1, "refrigerants", 1810.0, "kg"),
    ("ADEME.2025.REFRIG.R123",             "R-123 (HCFC, PRG 79)",                    1, "refrigerants",   79.0, "kg"),
]

_FACTORS_CHEMICALS = [
    # Produits chimiques et industriels
    ("ADEME.2025.CHIM.ETHYLENE",           "Éthylène",                                3, "chemicals", 1740.0, "t"),
    ("ADEME.2025.CHIM.AMMONIAC",           "Ammoniac (synthèse Haber-Bosch)",         3, "chemicals", 1890.0, "t"),
    ("ADEME.2025.CHIM.ACIDE_SULFURIQUE",   "Acide sulfurique",                        3, "chemicals",  144.0, "t"),
    ("ADEME.2025.CHIM.ACIDE_CHLORHYDRIQUE","Acide chlorhydrique",                     3, "chemicals",  590.0, "t"),
    ("ADEME.2025.CHIM.SOUDE",              "Soude caustique NaOH",                    3, "chemicals",  820.0, "t"),
    ("ADEME.2025.CHIM.CHLORE",             "Chlore (électrolyse)",                    3, "chemicals",  590.0, "t"),
    ("ADEME.2025.CHIM.METHANOL",           "Méthanol",                                3, "chemicals",  920.0, "t"),
    ("ADEME.2025.CHIM.ETHANOL_AGRI",       "Éthanol agricole",                        3, "chemicals", 1490.0, "t"),
    ("ADEME.2025.CHIM.ACETONE",            "Acétone",                                 3, "chemicals", 1590.0, "t"),
    ("ADEME.2025.CHIM.TOLUENE",            "Toluène",                                 3, "chemicals", 2150.0, "t"),
    ("ADEME.2025.CHIM.EPOXY_RESINE",       "Résine époxy",                            3, "chemicals", 6200.0, "t"),
    ("ADEME.2025.CHIM.POLYURETHANE",       "Polyuréthane (mousse rigide)",            3, "chemicals", 3800.0, "t"),
    ("ADEME.2025.CHIM.SILICONE",           "Silicone",                                3, "chemicals", 4700.0, "t"),
    ("ADEME.2025.CHIM.CARBURE_DE_SILICIUM","Carbure de silicium",                     3, "chemicals",19000.0, "t"),
    ("ADEME.2025.CHIM.ACIDE_NITRIQUE",     "Acide nitrique",                          3, "chemicals",  800.0, "t"),
    ("ADEME.2025.CHIM.UREE_ENGRAIS",       "Urée (engrais)",                          3, "chemicals", 5300.0, "t"),
    ("ADEME.2025.CHIM.PEINTURE_SOLV",      "Peinture à solvant",                      3, "chemicals", 3200.0, "t"),
    ("ADEME.2025.CHIM.PEINTURE_EAU",       "Peinture à l'eau",                        3, "chemicals", 1600.0, "t"),
    ("ADEME.2025.CHIM.VERNIS_BOIS",        "Vernis bois polyuréthane",                3, "chemicals", 2900.0, "t"),
    ("ADEME.2025.CHIM.COLLE_POLYESTER",    "Colle polyester",                         3, "chemicals", 4100.0, "t"),
]

_FACTORS_ENERGY_SCOPE3 = [
    # Combustibles fossiles achetés — Scope 3 amont (extraction + transport)
    ("ADEME.2025.AMONT.PETROL_BRUT",       "Pétrole brut (amont extraction)",         3, "energy",    0.0420, "kWh"),
    ("ADEME.2025.AMONT.GAZ_NAT_AMONT",     "Gaz naturel (amont extraction)",          3, "energy",    0.0280, "kWh"),
    ("ADEME.2025.AMONT.CHARBON_AMONT",     "Charbon (amont extraction + transport)",  3, "energy",    0.0320, "kWh"),
    ("ADEME.2025.AMONT.FIOUL_AMONT",       "Fioul (raffinerie + transport)",          3, "energy",    0.0380, "kWh"),
    ("ADEME.2025.AMONT.ESSENCE_AMONT",     "Essence (raffinerie + transport)",        3, "energy",    0.0400, "kWh"),
    ("ADEME.2025.AMONT.DIESEL_AMONT",      "Diesel (raffinerie + transport)",         3, "energy",    0.0360, "kWh"),
    ("ADEME.2025.AMONT.GNL_AMONT",         "GNL (liquéfaction + transport)",          3, "energy",    0.0580, "kWh"),
    ("ADEME.2025.AMONT.BIOGAZ_AMONT",      "Biogaz (production + distribution)",      3, "energy",    0.0120, "kWh"),
    # Énergies renouvelables (ACV complète)
    ("ADEME.2025.ENR.SOLAIRE_PV_TOITURE",  "Solaire PV toiture (ACV)",               None, "energy",  0.0550, "kWh"),
    ("ADEME.2025.ENR.SOLAIRE_PV_AU_SOL",   "Solaire PV au sol (ACV)",                None, "energy",  0.0480, "kWh"),
    ("ADEME.2025.ENR.EOLIEN_TERRESTRE",    "Éolien terrestre (ACV)",                 None, "energy",  0.0130, "kWh"),
    ("ADEME.2025.ENR.EOLIEN_OFFSHORE",     "Éolien offshore (ACV)",                  None, "energy",  0.0150, "kWh"),
    ("ADEME.2025.ENR.HYDRAULIQUE_FR",      "Hydraulique France (ACV)",               None, "energy",  0.0060, "kWh"),
    ("ADEME.2025.ENR.NUCLEAIRE_FR",        "Nucléaire France (ACV)",                 None, "energy",  0.0040, "kWh"),
    ("ADEME.2025.ENR.GEOTHERMIQUE",        "Géothermique (vapeur)",                  None, "energy",  0.0380, "kWh"),
    ("ADEME.2025.ENR.BIOMASSE_CHALEUR",    "Biomasse chaleur (réseau, ACV)",         None, "energy",  0.0250, "kWh"),
]

_FACTORS_BUSINESS_TRAVEL = [
    # Déplacements professionnels
    ("ADEME.2025.BTRAVEL.TAXI",            "Taxi (course urbaine)",                   3, "business_travel",  0.1410, "km"),
    ("ADEME.2025.BTRAVEL.VTC",             "VTC (course urbaine)",                    3, "business_travel",  0.1700, "km"),
    ("ADEME.2025.BTRAVEL.TAXI_ELEC",       "Taxi électrique",                         3, "business_travel",  0.0480, "km"),
    ("ADEME.2025.BTRAVEL.COVOITURAGE_2",   "Covoiturage 2 personnes (thermique)",     3, "business_travel",  0.1090, "km"),
    ("ADEME.2025.BTRAVEL.COVOITURAGE_3",   "Covoiturage 3 personnes (thermique)",     3, "business_travel",  0.0720, "km"),
    ("ADEME.2025.BTRAVEL.LOC_VOITURE",     "Location voiture compacte thermique",     3, "business_travel",  0.2170, "km"),
    ("ADEME.2025.BTRAVEL.LOC_VOITURE_E",   "Location voiture électrique (mix FR)",   3, "business_travel",  0.0450, "km"),
    ("ADEME.2025.BTRAVEL.EUROSTAR",        "Eurostar Paris-Londres",                  3, "business_travel",  0.0041, "km"),
    ("ADEME.2025.BTRAVEL.THALYS",          "Thalys Paris-Amsterdam",                  3, "business_travel",  0.0034, "km"),
    ("ADEME.2025.BTRAVEL.ICE_DE",          "TGV Allemagne ICE",                       3, "business_travel",  0.0048, "km"),
    ("ADEME.2025.BTRAVEL.AVION_INTERNE_FR","Vol domestique France (toutes classes)",  3, "business_travel",  0.2570, "km"),
    ("ADEME.2025.BTRAVEL.HOTEL_AFFAIRES",  "Hôtel d'affaires France (nuit)",          3, "business_travel",  18.4,   "nuit"),
    ("ADEME.2025.BTRAVEL.HOTEL_AFFAIRES_INT","Hôtel d'affaires international (nuit)", 3, "business_travel",  31.2,   "nuit"),
    ("ADEME.2025.BTRAVEL.CONF_EN_LIGNE",   "Vidéoconférence 1h (par participant)",    3, "business_travel",  0.0044, "heure"),
    ("ADEME.2025.BTRAVEL.TAXI_MOTO",       "Mototaxi / coursier",                     3, "business_travel",  0.0890, "km"),
    ("ADEME.2025.BTRAVEL.TELEPORTATION",   "Déplacement 0 km (télétravail)",          3, "business_travel",  0.0,    "km"),
]

_FACTORS_FUELS_SCOPE1_EXTRA = [
    # Carburants routiers — Scope 1 (combustion directe véhicules entreprise)
    ("ADEME.2025.CARB.ESSENCE_SP95",       "Essence SP95 (combustion)",               1, "fuels",    2.290, "L"),
    ("ADEME.2025.CARB.ESSENCE_SP98",       "Essence SP98 (combustion)",               1, "fuels",    2.310, "L"),
    ("ADEME.2025.CARB.ESSENCE_E10",        "Essence E10 (combustion)",                1, "fuels",    2.160, "L"),
    ("ADEME.2025.CARB.DIESEL_B7",          "Diesel B7 (combustion)",                  1, "fuels",    2.510, "L"),
    ("ADEME.2025.CARB.DIESEL_B30",         "Diesel B30 (combustion)",                 1, "fuels",    2.260, "L"),
    ("ADEME.2025.CARB.DIESEL_HVO",         "HVO (huile végétale hydrogénée)",         1, "fuels",    0.650, "L"),
    ("ADEME.2025.CARB.GNV_CNG",            "GNV (gaz naturel véhicule, CNG)",         1, "fuels",    2.050, "kg"),
    ("ADEME.2025.CARB.GNV_LNG",            "GNL (gaz naturel liquéfié, LNG)",         1, "fuels",    2.750, "kg"),
    ("ADEME.2025.CARB.BIOETHANOL",         "Bioéthanol E85",                          1, "fuels",    0.590, "L"),
    ("ADEME.2025.CARB.BIODIESEL_B100",     "Biodiesel B100 EMHV",                     1, "fuels",    0.570, "L"),
    ("ADEME.2025.CARB.KEROSENE_JET_A1",    "Kérosène Jet A-1 (aviation)",             1, "fuels",    2.540, "L"),
    ("ADEME.2025.CARB.KEROSENE_SAF",       "SAF (carburant aviation durable 30%)",    1, "fuels",    1.780, "L"),
    ("ADEME.2025.CARB.FIOUL_MARINE_HFO",   "Fioul marin HFO (bunker)",               1, "fuels",    3.140, "L"),
    ("ADEME.2025.CARB.FIOUL_MARINE_MDO",   "Distillat marin MDO",                     1, "fuels",    2.890, "L"),
    ("ADEME.2025.CARB.FIOUL_MARINE_LNG",   "GNL marin (moteur dual-fuel)",           1, "fuels",    2.760, "kg"),
    ("ADEME.2025.CARB.METHANOL_MARIN",     "Méthanol marin",                          1, "fuels",    0.480, "L"),
    ("ADEME.2025.CARB.AMMONIAC_MARIN",     "Ammoniac comme carburant marin",          1, "fuels",    0.000, "kg"),  # 0 CO2 direct
]

_FACTORS_PACKAGING = [
    # Emballages — ACV (production + fin de vie moyenne)
    ("ADEME.2025.EMBAL.BOUTEILLE_PET_33CL",   "Bouteille PET 33 cl vierge",              3, "packaging",   0.0480, "unité"),
    ("ADEME.2025.EMBAL.BOUTEILLE_PET_1L",     "Bouteille PET 1L vierge",                 3, "packaging",   0.1120, "unité"),
    ("ADEME.2025.EMBAL.BOUTEILLE_VERRE_75CL", "Bouteille verre 75 cl vierge",            3, "packaging",   0.6900, "unité"),
    ("ADEME.2025.EMBAL.BOUTEILLE_VERRE_REU",  "Bouteille verre 75 cl réutilisable",      3, "packaging",   0.1400, "unité"),
    ("ADEME.2025.EMBAL.CANETTE_ALUM_33CL",    "Canette aluminium 33 cl",                 3, "packaging",   0.1180, "unité"),
    ("ADEME.2025.EMBAL.CANETTE_ACIER_33CL",   "Canette acier 33 cl",                     3, "packaging",   0.0930, "unité"),
    ("ADEME.2025.EMBAL.BOITE_CARTON_250G",    "Boite carton pliante 250 g",              3, "packaging",   0.0620, "unité"),
    ("ADEME.2025.EMBAL.SAC_PLASTIQUE_USAGE",  "Sac plastique usage unique (5-10 g)",     3, "packaging",   0.0165, "unité"),
    ("ADEME.2025.EMBAL.SAC_COTON_REUTILIS",   "Sac coton réutilisable",                  3, "packaging",   0.1490, "unité"),
    ("ADEME.2025.EMBAL.SAC_PAPIER_KRAFT",     "Sac papier kraft 80g",                    3, "packaging",   0.0280, "unité"),
    ("ADEME.2025.EMBAL.FILM_ETIRABLE",        "Film étirable polyéthylène (kg)",         3, "packaging",   1.9200, "kg"),
    ("ADEME.2025.EMBAL.BARQUETTE_PS",         "Barquette polystyrène alimentaire",       3, "packaging",   0.0082, "unité"),
    ("ADEME.2025.EMBAL.BARQUETTE_PP",         "Barquette polypropylène",                 3, "packaging",   0.0074, "unité"),
    ("ADEME.2025.EMBAL.BARQUETTE_PLA",        "Barquette PLA (bioplastique)",            3, "packaging",   0.0095, "unité"),
    ("ADEME.2025.EMBAL.PALETTE_BOIS",         "Palette bois EPAL 4 voies",               3, "packaging",  14.8000, "unité"),
    ("ADEME.2025.EMBAL.PALETTE_PLASTIQUE",    "Palette plastique HDPE",                  3, "packaging",  18.6000, "unité"),
    ("ADEME.2025.EMBAL.CARTON_TRANSPORT",     "Carton de transport (kg)",                3, "packaging",   0.7500, "kg"),
    ("ADEME.2025.EMBAL.CALAGE_POLYSTYRENE",   "Calage polystyrène expansé (kg)",         3, "packaging",   3.2000, "kg"),
    ("ADEME.2025.EMBAL.CALAGE_PAPIER",        "Calage papier kraft (kg)",                3, "packaging",   0.9200, "kg"),
    ("ADEME.2025.EMBAL.ETIQUETTE_PP",         "Étiquette PP autocollante (100 unités)",  3, "packaging",   0.0420, "lot"),
    ("ADEME.2025.EMBAL.BOUCHON_PE",           "Bouchon polyéthylène",                    3, "packaging",   0.0023, "unité"),
    ("ADEME.2025.EMBAL.CAPSULE_ALUMINIUM",    "Capsule aluminium",                       3, "packaging",   0.0031, "unité"),
    ("ADEME.2025.EMBAL.TETRA_PACK_1L",        "Tetra Pak 1L",                            3, "packaging",   0.0940, "unité"),
    ("ADEME.2025.EMBAL.BOITE_CONSERVE_400G",  "Boîte conserve acier 400g",               3, "packaging",   0.1620, "unité"),
    ("ADEME.2025.EMBAL.FLACON_VERRE_500ML",   "Flacon verre 500 ml",                     3, "packaging",   0.4100, "unité"),
    ("ADEME.2025.EMBAL.TUBE_PLASTIQUE_100G",  "Tube plastique 100g",                     3, "packaging",   0.0510, "unité"),
    ("ADEME.2025.EMBAL.BLISTER_PVC",          "Blister PVC thermoformé",                 3, "packaging",   0.0380, "unité"),
    ("ADEME.2025.EMBAL.FILM_RETRACTABLE",     "Film rétractable PE (kg)",                3, "packaging",   1.8500, "kg"),
    ("ADEME.2025.EMBAL.FILET_PP",             "Filet polypropylène emballage",           3, "packaging",   0.0280, "kg"),
    ("ADEME.2025.EMBAL.EMBAL_VIDE_PLASTIQUE", "Emballage vide plastique multicouches",   3, "packaging",   0.0350, "unité"),
]

_FACTORS_ELECTRONICS = [
    # Équipements électroniques et industriels (fabrication + usage 5 ans)
    ("ADEME.2025.ELEC_EQUIP.TV_55POUCES",     "Téléviseur 55 pouces LED",                3, "electronics",  490.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.IMPRIMANTE_LASER","Imprimante laser A4",                     3, "electronics",  210.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.IMPRIMANTE_JET",  "Imprimante jet d'encre",                  3, "electronics",   77.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.ROUTEUR_WIFI",    "Routeur Wi-Fi professionnel",             3, "electronics",   32.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.SWITCH_24P",      "Switch réseau 24 ports",                  3, "electronics",  148.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.ONDULEUR_1KVA",   "Onduleur 1 kVA",                          3, "electronics",   85.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CLIMATISEUR_3KW", "Climatiseur split 3 kW",                  3, "electronics",  680.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CHAUDIERE_GAZ",   "Chaudière gaz condensation 25 kW",        3, "electronics",  890.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.PAC_AIR_AIR",     "Pompe à chaleur air/air 5 kW",            3, "electronics",  780.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.PAC_AIR_EAU",     "Pompe à chaleur air/eau 12 kW",           3, "electronics", 1250.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CHAUFFE_EAU_ELEC","Chauffe-eau électrique 200L",             3, "electronics",  320.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CHAUFFE_EAU_THER","Chauffe-eau thermodynamique 270L",        3, "electronics",  560.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.MOTEUR_ELEC_15KW","Moteur électrique industriel 15 kW",      3, "electronics",  245.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.COMPRESSEUR_AIR",  "Compresseur d'air industriel 11 kW",     3, "electronics",  420.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.ENROULEUR_CONVEY", "Convoyeur à bande (ml)",                 3, "electronics",  180.0, "ml"),
    ("ADEME.2025.ELEC_EQUIP.BATTERIE_LION_1KWH","Batterie Li-ion 1 kWh (fabrication)",   3, "electronics",  150.0, "kWh"),
    ("ADEME.2025.ELEC_EQUIP.PANNEAU_PV_1KWP",  "Module solaire PV 1 kWc (fabrication)",  3, "electronics",  550.0, "kWc"),
    ("ADEME.2025.ELEC_EQUIP.EOLIENNE_2MW",     "Éolienne 2 MW (fabrication)",            3, "electronics",  760000.0,"unité"),
    ("ADEME.2025.ELEC_EQUIP.VEHICULE_ELEC",    "Véhicule électrique (fabrication)",      3, "electronics", 9800.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.VEHICULE_THER",    "Véhicule thermique (fabrication)",       3, "electronics", 7400.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.VEHICULE_HYBR",    "Véhicule hybride (fabrication)",         3, "electronics", 8600.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CAMION_PL_18T",    "Camion poids lourd 18T (fabrication)",   3, "electronics",38000.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.SMARTPHONE_REPAR", "Smartphone reconditionné",               3, "electronics",  8.2,   "unité"),
    ("ADEME.2025.ELEC_EQUIP.BORNE_RECHARGE",   "Borne de recharge VE 22 kW",             3, "electronics",  480.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CABLE_CUIVRE",     "Câble cuivre 50 mm² (mètre)",            3, "electronics",  0.9,   "m"),
    ("ADEME.2025.ELEC_EQUIP.TRANSFORMATEUR",   "Transformateur BT/HTA 630 kVA",          3, "electronics", 4200.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.SERVEUR_RACK_42U", "Baie serveur 42U complète",              3, "electronics",18000.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.CLOUD_1TB_AN",     "Stockage cloud 1 To (par an, mix FR)",   3, "electronics",   87.0, "an"),
    ("ADEME.2025.ELEC_EQUIP.DETECTEUR_FUMEE",  "Détecteur de fumée",                     3, "electronics",    3.1, "unité"),
    ("ADEME.2025.ELEC_EQUIP.ECLAIRAGE_LED_9W", "Ampoule LED 9W (fabrication)",           3, "electronics",    2.8, "unité"),
    ("ADEME.2025.ELEC_EQUIP.LUMINAIRE_BUREAU", "Luminaire bureau LED 40W",               3, "electronics",   18.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.ASCENSEUR_6P",     "Ascenseur 6 personnes (fabrication)",    3, "electronics",12500.0, "unité"),
    ("ADEME.2025.ELEC_EQUIP.TELESCOPE_SOLAIRE","Capteur solaire thermique 2m²",          3, "electronics",  210.0, "unité"),
]

_FACTORS_OFFICES = [
    # Fournitures et mobilier de bureau
    ("ADEME.2025.BUREAU.RAME_PAPIER_A4",   "Rame papier A4 80g (500 feuilles)",       3, "office",    3.020, "rame"),
    ("ADEME.2025.BUREAU.RAME_PAPIER_RECYC","Rame papier A4 recyclé (500 feuilles)",   3, "office",    1.580, "rame"),
    ("ADEME.2025.BUREAU.STYLO_BILLE",      "Stylo bille plastique",                   3, "office",    0.017, "unité"),
    ("ADEME.2025.BUREAU.CRAYON_PAPIER",    "Crayon à papier",                         3, "office",    0.010, "unité"),
    ("ADEME.2025.BUREAU.CHEMISE_CARTON",   "Chemise cartonnée",                       3, "office",    0.085, "unité"),
    ("ADEME.2025.BUREAU.CLASSEUR_A4",      "Classeur A4 en carton",                   3, "office",    0.340, "unité"),
    ("ADEME.2025.BUREAU.ENVELOPPE_C5",     "Enveloppe C5 avec fenêtre",               3, "office",    0.011, "unité"),
    ("ADEME.2025.BUREAU.TONER_LASER_NOIR", "Toner laser noir (cartouche)",            3, "office",    3.800, "unité"),
    ("ADEME.2025.BUREAU.TONER_LASER_COLOR","Toner laser couleur (cartouche)",         3, "office",    4.200, "unité"),
    ("ADEME.2025.BUREAU.CHAISE_BUREAU",    "Chaise de bureau ergonomique",            3, "office",   72.000, "unité"),
    ("ADEME.2025.BUREAU.BUREAU_STANDARD",  "Bureau de travail standard",              3, "office",  120.000, "unité"),
    ("ADEME.2025.BUREAU.MEUBLE_RANGEMENT", "Meuble de rangement métal",               3, "office",  185.000, "unité"),
    ("ADEME.2025.BUREAU.TABLE_CONF",       "Table de conférence (8 places)",          3, "office",  310.000, "unité"),
    ("ADEME.2025.BUREAU.TABLEAU_BLANC",    "Tableau blanc 120x90",                    3, "office",   28.000, "unité"),
    ("ADEME.2025.BUREAU.CAFETIERE_CAPSULE","Cafetière à capsules",                    3, "office",   24.000, "unité"),
    ("ADEME.2025.BUREAU.CAPSULE_CAFE",     "Capsule café aluminium",                  3, "office",    0.043, "unité"),
    ("ADEME.2025.BUREAU.CAPSULE_COMPOST",  "Capsule café compostable",                3, "office",    0.031, "unité"),
    ("ADEME.2025.BUREAU.MACHINE_EXPRESSO", "Machine expresso professionnelle",        3, "office",  380.000, "unité"),
    ("ADEME.2025.BUREAU.DOSETTE_CAFE",     "Dosette café papier (sachet 7g)",         3, "office",    0.038, "unité"),
    ("ADEME.2025.BUREAU.POUBELLE_PLASTIQUE","Poubelle plastique bureau 15L",          3, "office",    1.800, "unité"),
    ("ADEME.2025.BUREAU.SAC_POUBELLE_50L", "Sac poubelle polyéthylène 50L",          3, "office",    0.180, "unité"),
    ("ADEME.2025.BUREAU.NETTOYANT_MULTI",  "Nettoyant multi-surfaces (500ml)",        3, "office",    1.200, "unité"),
    ("ADEME.2025.BUREAU.PAPIER_HYGIENE",   "Papier hygiénique (rouleau)",             3, "office",    0.083, "rouleau"),
    ("ADEME.2025.BUREAU.ESSUIE_MAIN_PAPIER","Essuie-mains papier (100 feuilles)",     3, "office",    0.450, "lot"),
    ("ADEME.2025.BUREAU.REPAS_CANTINE_VEG","Repas végétarien cantine entreprise",    3, "office",    1.100, "repas"),
    ("ADEME.2025.BUREAU.REPAS_CANTINE_VIA","Repas avec viande cantine entreprise",   3, "office",    3.600, "repas"),
    ("ADEME.2025.BUREAU.FONTAINE_EAU",     "Fontaine à eau réfrigérante (an)",        3, "office",  145.000, "an"),
    ("ADEME.2025.BUREAU.BOUTEILLE_EAU_33",  "Bouteille eau minérale 33cl plastique", 3, "office",    0.110, "unité"),
    ("ADEME.2025.BUREAU.GOURDE_INOX",      "Gourde inox réutilisable 500ml",          3, "office",    3.200, "unité"),
    ("ADEME.2025.BUREAU.BADGE_RFID",       "Badge RFID accès",                        3, "office",    0.420, "unité"),
]

_FACTORS_INDUSTRY_PROCESS = [
    # Émissions de procédé industriel — Scope 1
    ("ADEME.2025.PROC.ACIER_HAUT_FOURNEAU", "Production acier haut fourneau (procédé)", 1, "process",  1480.0, "t"),
    ("ADEME.2025.PROC.CIMENT_CLINKER",      "Calcination clinker ciment",               1, "process",   523.0, "t"),
    ("ADEME.2025.PROC.ALUMINIUM_ANODE",     "Effet anodique aluminium (PFC)",           1, "process",   163.0, "t"),
    ("ADEME.2025.PROC.VERRE_FUSION",        "Fusion verre float (procédé)",             1, "process",   220.0, "t"),
    ("ADEME.2025.PROC.CERAMIQUE_CUISSON",   "Cuisson céramique (procédé)",              1, "process",   140.0, "t"),
    ("ADEME.2025.PROC.PAPIER_KRAFT",        "Kraft pulping (procédé)",                  1, "process",    35.0, "t"),
    ("ADEME.2025.PROC.RAFFINAGE_PETROL",    "Raffinage pétrole brut",                   1, "process",    62.0, "t"),
    ("ADEME.2025.PROC.PRODUCTION_FONTE",    "Production fonte de fer",                  1, "process",  1620.0, "t"),
    ("ADEME.2025.PROC.CALCINATION_CHAUX",   "Calcination chaux vive",                   1, "process",   750.0, "t"),
    ("ADEME.2025.PROC.FERMENTATION_ALCOOL", "Fermentation alcoolique (CO2)",            1, "process",   511.0, "t"),
    ("ADEME.2025.PROC.METHANISATION_BIOGAZ","Méthanisation (pertes méthane)",           1, "process",    28.5, "t"),
    ("ADEME.2025.PROC.TRAITEMENT_EAUX_IND", "Station épuration industrielle",           1, "process",   195.0, "t"),
    ("ADEME.2025.PROC.ELECTROLYSE_CHLORE",  "Électrolyse chlore-soude",                1, "process",    48.0, "t"),
    ("ADEME.2025.PROC.PRODUCTION_ACIDE_NIT","Production acide nitrique (N2O)",         1, "process",   900.0, "t"),
    ("ADEME.2025.PROC.PRODUCTION_CAPROLACT","Production caprolactame (N2O)",            1, "process",  2850.0, "t"),
    ("ADEME.2025.PROC.INCIN_DECHETS_INDUST","Incinération déchets industriels spéc.",  1, "process",  1600.0, "t"),
    ("ADEME.2025.PROC.STOCKAGE_DECHETS",    "CET (décharge) méthane diffus",           1, "process",   310.0, "t"),
    ("ADEME.2025.PROC.EXTRACTION_TOURBE",   "Extraction tourbe (drainage sol)",         1, "process",    45.0, "t"),
    ("ADEME.2025.PROC.SECHAGE_BOIS_GAZ",    "Séchage bois au gaz (procédé)",           1, "process",   180.0, "t"),
    ("ADEME.2025.PROC.SOUDAGE_ARC",         "Soudage à l'arc électrique",              1, "process",     1.2, "heure"),
    ("ADEME.2025.PROC.PEINTURE_CABINE",     "Peinture industrielle en cabine (kg pdr)", 1, "process",    3.8, "kg"),
    ("ADEME.2025.PROC.TRAITEMENT_SURFACE",  "Traitement de surface (dégraissage)",      1, "process",    2.1, "m2"),
    ("ADEME.2025.PROC.GALVANISATION",       "Galvanisation à chaud (par kg zinc)",      1, "process",    3.5, "kg"),
    ("ADEME.2025.PROC.ANODISATION",         "Anodisation aluminium (par m²)",           1, "process",    4.2, "m2"),
    ("ADEME.2025.PROC.USINAGE_TOURNAGE",    "Usinage tournage (énergie électrique)",    1, "process",    0.0512, "kWh"),
    ("ADEME.2025.PROC.IMPRESSION_3D_FDM",   "Impression 3D FDM (kWh + filament)",      3, "process",    1.8, "kg"),
    ("ADEME.2025.PROC.IMPRESSION_3D_SLA",   "Impression 3D SLA résine photopolymère",  3, "process",    5.3, "kg"),
    ("ADEME.2025.PROC.SOUFFLAGE_PLASTIQUE", "Soufflage extrusion plastique",            1, "process",    0.28, "kg"),
    ("ADEME.2025.PROC.INJECTION_PLASTIQUE", "Injection plastique",                      1, "process",    0.31, "kg"),
    ("ADEME.2025.PROC.THERMOFORMAGE",       "Thermoformage plastique",                  1, "process",    0.25, "kg"),
]


_FACTORS_SCOPE3_UPSTREAM = [
    # Scope 3 amont — biens et services achetés (catégorie 1)
    ("ADEME.2025.ACHAT.MOBILIER_BOIS",       "Mobilier en bois (achat neuf, kg)",       3, "purchased_goods",    1.85, "kg"),
    ("ADEME.2025.ACHAT.MOBILIER_ACIER",      "Mobilier en acier (achat neuf, kg)",      3, "purchased_goods",    4.20, "kg"),
    ("ADEME.2025.ACHAT.VETEMENT_POLOS",      "Vêtement de travail polyester",           3, "purchased_goods",   12.80, "unité"),
    ("ADEME.2025.ACHAT.VETEMENT_COTON",      "Vêtement de travail coton",               3, "purchased_goods",   16.50, "unité"),
    ("ADEME.2025.ACHAT.CHAUSSURES_CUIR",     "Chaussures cuir (ACV)",                   3, "purchased_goods",   17.40, "paire"),
    ("ADEME.2025.ACHAT.EQUIPEMENT_SECURITE", "EPI standard (casque + gants + lunettes)", 3,"purchased_goods",    4.80, "lot"),
    ("ADEME.2025.ACHAT.FOURNITURES_BUREAU",  "Fournitures bureau (€ dépensé)",          3, "purchased_goods",    0.28, "€"),
    ("ADEME.2025.ACHAT.SERVICES_IT",         "Services informatiques (€ dépensé)",      3, "purchased_goods",    0.19, "€"),
    ("ADEME.2025.ACHAT.SERVICES_NETTOYAGE",  "Services nettoyage locaux (€ dépensé)",   3, "purchased_goods",    0.22, "€"),
    ("ADEME.2025.ACHAT.SERVICES_SECURITE",   "Services gardiennage (€ dépensé)",        3, "purchased_goods",    0.15, "€"),
    ("ADEME.2025.ACHAT.PUBLICITE_PAPIER",    "Publicité imprimée (kg papier)",          3, "purchased_goods",    0.92, "kg"),
    ("ADEME.2025.ACHAT.PUBLICITE_NUMERIQUE", "Publicité numérique (€ dépensé)",         3, "purchased_goods",    0.04, "€"),
    ("ADEME.2025.ACHAT.FORMATION",           "Formation présentielle (jour/stagiaire)", 3, "purchased_goods",   14.20, "jour"),
    ("ADEME.2025.ACHAT.CLOUD_COMPUTING",     "Cloud computing AWS/Azure (€ dépensé)",   3, "purchased_goods",    0.13, "€"),
    ("ADEME.2025.ACHAT.LOGICIEL_SAAS",       "Logiciel SaaS (€ dépensé, an)",           3, "purchased_goods",    0.09, "€"),
    ("ADEME.2025.ACHAT.TELEPHONE_FIXE",      "Téléphonie fixe (€ dépensé, an)",         3, "purchased_goods",    0.11, "€"),
    ("ADEME.2025.ACHAT.INTERNET_FIBRE",      "Accès internet fibre (an)",               3, "purchased_goods",   26.0,  "an"),
    ("ADEME.2025.ACHAT.ELECTRICITE_ACHETEE", "Électricité achetée (facteur S3, mix FR)",3, "purchased_goods",    0.0512,"kWh"),
    ("ADEME.2025.ACHAT.GAZ_ACHETE",          "Gaz naturel acheté (facteur S3)",         3, "purchased_goods",    0.2290,"kWh"),
    # Scope 3 aval — utilisation des produits vendus (catégorie 11)
    ("ADEME.2025.AVAL.UTILISATION_APPAREL",  "Usage appareils électroménagers (kWh)",   3, "product_use",    0.0512, "kWh"),
    ("ADEME.2025.AVAL.UTILISATION_VEHICULE", "Usage véhicule vendu (km parcouru)",      3, "product_use",    0.2170, "km"),
    ("ADEME.2025.AVAL.FIN_VIE_APPAREL",      "Fin de vie appareils électroménagers",    3, "end_of_life",  380.0,    "t"),
    ("ADEME.2025.AVAL.FIN_VIE_VEHICULE",     "Fin de vie véhicule (démontage+broyage)", 3, "end_of_life",  620.0,    "t"),
    ("ADEME.2025.AVAL.FIN_VIE_TEXTILE",      "Fin de vie textile (décharge)",           3, "end_of_life",  280.0,    "t"),
    ("ADEME.2025.AVAL.FIN_VIE_EMBALLAGE",    "Fin de vie emballages consommateurs",     3, "end_of_life",  145.0,    "t"),
    # Catégorie 12 — traitement fin de vie des produits vendus
    ("ADEME.2025.CAT12.DEEE_COLLECTE",       "Collecte DEEE (logistique inverse)",      3, "end_of_life",   45.0,    "t"),
    ("ADEME.2025.CAT12.DEEE_DEMONTAGE",      "Démontage DEEE (main d'œuvre)",           3, "end_of_life",  120.0,    "t"),
    ("ADEME.2025.CAT12.DEEE_VALORISATION",   "Valorisation DEEE (broyage+tri)",         3, "end_of_life",   78.0,    "t"),
    # Catégorie 13 — voyages d'affaires (complément)
    ("ADEME.2025.CAT13.SALON_PROFESSIONNEL", "Participation salon professionnel (jour)", 3, "business_travel", 32.0, "jour"),
    ("ADEME.2025.CAT13.SEMINAIRE_HOTEL",     "Séminaire résidentiel (jour/personne)",    3, "business_travel", 48.0, "jour"),
    # Catégorie 15 — investissements financiers
    ("ADEME.2025.CAT15.INVESTISSEMENT_PEA",  "Investissement actions cotées (€)",       3, "investments",   0.000082, "€"),
    ("ADEME.2025.CAT15.INVESTISSEMENT_OBLIG","Investissement obligations (€)",          3, "investments",   0.000031, "€"),
    ("ADEME.2025.CAT15.CREDIT_IMMOBILIER",   "Crédit immobilier accordé (€)",           3, "investments",   0.000097, "€"),
    ("ADEME.2025.CAT15.CREDIT_AUTO",         "Crédit automobile accordé (€)",           3, "investments",   0.000145, "€"),
    ("ADEME.2025.CAT15.ASSURANCE_VIE",       "Provisions assurance-vie (€)",            3, "investments",   0.000028, "€"),
    # Scope 3 — eau et énergie dans le bâtiment (catégorie 3)
    ("ADEME.2025.BATIMENT.GAZ_CHAUFF_S3",    "Gaz chauffage bâtiment loué (S3 cat3)",   3, "energy",    0.2014, "kWh"),
    ("ADEME.2025.BATIMENT.ELEC_LOCAT_S3",    "Électricité bâtiment loué (S3 cat3)",     3, "energy",    0.0512, "kWh"),
    ("ADEME.2025.BATIMENT.CLIMATISATION_S3", "Climatisation bâtiment loué (S3 cat3)",   3, "energy",    0.0512, "kWh"),
    ("ADEME.2025.BATIMENT.CHAUFF_FIOUL_S3",  "Fioul chauffage bâtiment loué (S3)",      3, "energy",    0.3247, "kWh"),
    ("ADEME.2025.BATIMENT.EAU_CHAUDE_S3",    "Eau chaude sanitaire bâtiment loué",      3, "water",     0.3440, "m3"),
    # Scope 3 — déchets générés (catégorie 5)
    ("ADEME.2025.CAT5.CARTON_BUREAU_BENNE",  "Carton bureau en benne (collecte)",       3, "waste",    610.0, "t"),
    ("ADEME.2025.CAT5.PLASTIQUE_BUREAU",     "Plastiques divers bureau (collecte)",     3, "waste",    180.0, "t"),
    ("ADEME.2025.CAT5.DECHETS_RESTO_ENTR",   "Déchets restauration entreprise",         3, "waste",    430.0, "t"),
    ("ADEME.2025.CAT5.DIB_PME",              "DIB entreprise < 50 sal. (collecte)",     3, "waste",    350.0, "t"),
    # Scope 3 — actifs loués amont (catégorie 8)
    ("ADEME.2025.CAT8.VOITURE_LOCATION",     "Voiture de société en LOA (km)",          3, "leased_assets", 0.2170, "km"),
    ("ADEME.2025.CAT8.LOCAUX_LOUES_M2",      "Locaux professionnels loués (m², an)",    3, "leased_assets",  9.8, "m2"),
    # Scope 3 — franchises (catégorie 14)
    ("ADEME.2025.CAT14.FRANCHISE_ENERGIE",   "Franchise — énergie magasin (€ CA)",      3, "franchises",   0.000065, "€"),
    ("ADEME.2025.CAT14.FRANCHISE_TRANSPORT", "Franchise — transport livraison (€ CA)",  3, "franchises",   0.000091, "€"),
    # Scope 1 — fugitives (pompage, compression)
    ("ADEME.2025.FUGIT.METHANE_POMPE_GAZ",   "Fuite méthane pompe centrifuge gaz",      1, "fugitives",  28.0, "kg"),
    ("ADEME.2025.FUGIT.CO2_PUITS_PETROL",    "Dégazage CO2 extraction pétrolière",      1, "fugitives",   1.0, "kg"),
    ("ADEME.2025.FUGIT.CH4_DISTRIB_GAZ",     "Fuite réseau de distribution gaz",        1, "fugitives",  28.0, "kg"),
]


def _collect_all_factors() -> list[tuple]:
    return (
        _FACTORS_ENERGY_SCOPE1
        + _FACTORS_ENERGY_SCOPE2
        + _FACTORS_TRANSPORT_SCOPE3
        + _FACTORS_MATERIALS
        + _FACTORS_WASTE
        + _FACTORS_FOOD
        + _FACTORS_WATER_SERVICES
        + _FACTORS_AGRICULTURE
        + _FACTORS_REFRIGERANTS
        + _FACTORS_CHEMICALS
        + _FACTORS_ENERGY_SCOPE3
        + _FACTORS_BUSINESS_TRAVEL
        + _FACTORS_FUELS_SCOPE1_EXTRA
        + _FACTORS_PACKAGING
        + _FACTORS_ELECTRONICS
        + _FACTORS_OFFICES
        + _FACTORS_INDUSTRY_PROCESS
        + _FACTORS_SCOPE3_UPSTREAM
    )


def _try_parse_xlsx(xlsx_path: Path, version: str) -> list[tuple] | None:
    """Parse le fichier ADEME Base Empreinte s'il est disponible."""
    try:
        from openpyxl import load_workbook as lw
    except ImportError:
        return None
    if not xlsx_path.exists():
        return None
    try:
        wb = lw(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[WARN] Impossible de lire {xlsx_path}: {exc}")
        return None

    factors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row is None or len(row) < 5:
                continue
            ef_code, label, scope, category, factor_val, unit = (
                str(row[0] or "").strip(),
                str(row[1] or "").strip(),
                row[2],
                str(row[3] or "").strip(),
                row[4],
                str(row[5] or "kg") if len(row) > 5 else "kg",
            )
            if not ef_code or not label or factor_val is None:
                continue
            try:
                fv = float(factor_val)
                scope_int = int(scope) if scope else None
            except (TypeError, ValueError):
                continue
            factors.append((ef_code, label, scope_int, category, fv, unit))
    wb.close()
    return factors if len(factors) >= 10 else None


def seed(version: str = VERSION, dry_run: bool = False) -> int:
    """Insère les emission_factors. Retourne le nombre de lignes insérées."""
    xlsx_path = Path(__file__).parent.parent / "data" / "factors" / "Facteurs_Emission.xlsx"
    factors = _try_parse_xlsx(xlsx_path, version)
    if factors:
        print(f"[INFO] Fichier ADEME trouvé — {len(factors)} facteurs à importer depuis {xlsx_path.name}")
    else:
        factors = _collect_all_factors()
        print(f"[INFO] Fichier ADEME absent — utilisation des {len(factors)} facteurs embarqués (ADEME Base Empreinte {version})")

    if dry_run:
        print(f"[DRY-RUN] {len(factors)} facteurs seraient insérés.")
        return len(factors)

    if not db_available():
        print("[WARN] Base de données non disponible — skip insertion.")
        return 0

    inserted = 0
    with get_db() as conn:
        with conn.cursor() as cur:
            # S'assurer que la table existe
            cur.execute(
                "SELECT to_regclass('public.emission_factors')::text"
            )
            if cur.fetchone()[0] is None:
                migration_path = (
                    Path(__file__).parent.parent
                    / "db"
                    / "migrations"
                    / "001_emission_factors.sql"
                )
                cur.execute(migration_path.read_text(encoding="utf-8"))
                print("[INFO] Table emission_factors créée.")

            for ef_code, label, scope, category, factor_kgco2e, unit in factors:
                try:
                    cur.execute(
                        """
                        INSERT INTO emission_factors
                            (ef_code, label, scope, category, factor_kgco2e, unit, source, version)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (ef_code, version) DO UPDATE
                            SET label=EXCLUDED.label,
                                factor_kgco2e=EXCLUDED.factor_kgco2e,
                                unit=EXCLUDED.unit
                        """,
                        (ef_code, label, scope, category, factor_kgco2e, unit, SOURCE, version),
                    )
                    inserted += 1
                except Exception as exc:
                    print(f"[WARN] Skip {ef_code}: {exc}")
    print(f"[OK] {inserted} facteurs insérés/mis à jour (version {version}).")
    return inserted


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seed ADEME emission_factors")
    parser.add_argument("--dry-run", action="store_true", help="Simuler sans écrire en DB")
    parser.add_argument("--version", default=VERSION, help=f"Version des facteurs (défaut: {VERSION})")
    args = parser.parse_args()
    count = seed(version=args.version, dry_run=args.dry_run)
    sys.exit(0 if count >= 0 else 1)
