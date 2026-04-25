"""Shared helpers for NEURAL / Banque / Marketing workbooks."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = REPO_ROOT / "apps" / "neural" / "data" / "bank-marketing"
REF_DATE = date(2026, 4, 25)

INK = "0D1B2A"
BLUE = "1D4ED8"
BLUE_LIGHT = "DBEAFE"
GREEN = "166534"
GREEN_LIGHT = "DCFCE7"
AMBER = "B45309"
AMBER_LIGHT = "FEF3C7"
RED = "B91C1C"
RED_LIGHT = "FEE2E2"
GREY_LIGHT = "F8FAFC"
WHITE = "FFFFFF"

THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)

YES_NO = ["YES", "NO"]
YES_NO_NA = ["YES", "NO", "NA", "REVIEW"]
VERDICTS = ["PASS", "PASS_WITH_REVIEW", "BLOCK"]
SEVERITY = ["INFO", "LOW", "MEDIUM", "HIGH", "CRITICAL"]
SOURCE_STATUS = ["ACTIVE", "STALE", "REJECTED"]
PRIORITY = ["MVP", "V1", "V2"]
AGENT_STATUS = ["planned", "demo", "live"]

FILES = {
    "foundations": "NEURAL_BANK_MARKETING_FOUNDATIONS.xlsx",
    "master": "NEURAL_BANK_MARKETING_MASTER.xlsx",
    "agbm001": "NEURAL_AGBM001_BankMarketingComplianceGuard.xlsx",
    "agbm002": "NEURAL_AGBM002_FinLiteracyContent.xlsx",
    "agbm003": "NEURAL_AGBM003_SegmentedBankMarketing.xlsx",
    "agbm004": "NEURAL_AGBM004_MiFIDProductMarketingGuard.xlsx",
}

MVP_GATES = [
    "GATE-AMF-ACPR-CLEAR-NOT-MISLEADING",
    "GATE-RISK-BENEFIT-BALANCE",
    "GATE-SOURCE-ACTIVE",
    "GATE-NUM-VALIDATED",
    "GATE-MIFID-TARGET-MARKET",
    "GATE-PRIIPS-KID-CONSISTENCY",
    "GATE-MICA-CRYPTO-MARKETING",
    "GATE-GDPR-CONSENT-PROFILING",
    "GATE-AI-ACT-DISCLOSURE",
    "GATE-HITL-COMPLIANCE",
]


def make_workbook(title_text: str, subject: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = title_text
    wb.properties.subject = subject
    wb.properties.creator = "NEURAL"
    wb.properties.category = "Banque - Marketing"
    wb.properties.description = "Synthetic portfolio workbook generated for NEURAL Banque Marketing."
    return wb


def save_workbook(wb: Workbook, filename: str) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / filename
    wb.save(path)
    return path


def title(ws, main: str, subtitle: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, main)
    c.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = AC
    ws.row_dimensions[1].height = 38
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(name="Calibri", size=10, italic=True, color=BLUE_LIGHT)
    s.fill = PatternFill("solid", fgColor=INK)
    s.alignment = AC
    ws.row_dimensions[2].height = 24


def set_widths(ws, widths: Iterable[float]) -> None:
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def headers(ws, row: int, values: list[str], widths: Iterable[float] | None = None) -> None:
    if widths:
        set_widths(ws, widths)
    for idx, value in enumerate(values, 1):
        c = ws.cell(row, idx, value)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = AC
        c.border = THIN
    ws.row_dimensions[row].height = 32


def write_rows(ws, start_row: int, rows: list[tuple], align=AL) -> int:
    for r_idx, row in enumerate(rows, start_row):
        fill = GREY_LIGHT if (r_idx - start_row) % 2 == 0 else WHITE
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Calibri", size=9)
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = align
            c.border = THIN
    return start_row + len(rows) - 1 if rows else start_row - 1


def make_table(ws, table_name: str, start_row: int, end_row: int, end_col: int) -> None:
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_list_validation(ws, col: int, start_row: int, end_row: int, values: list[str]) -> None:
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    dv.error = "Value outside list"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    letter = get_column_letter(col)
    dv.add(f"{letter}{start_row}:{letter}{end_row}")


def add_status_formatting(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=GREEN_LIGHT)))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"PASS_WITH_REVIEW"'], fill=PatternFill("solid", fgColor=AMBER_LIGHT)))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"BLOCK"'], fill=PatternFill("solid", fgColor=RED_LIGHT)))


def finish_sheet(ws, freeze: str = "A5") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False


def readme_sheet(wb: Workbook, workbook_name: str, role: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("0_README")
    title(ws, f"NEURAL / BANQUE / MARKETING - {workbook_name}", role, 4)
    headers(ws, 4, ["#", "Topic", "Detail", "Status"], [6, 28, 94, 18])
    data = [(idx, topic, detail, "synthetic") for idx, (topic, detail) in enumerate(rows, 1)]
    last = write_rows(ws, 5, data)
    make_table(ws, f"tbl{workbook_name.replace('-', '').replace(' ', '')}Readme", 4, last, 4)
    finish_sheet(ws, "B5")


def param_sheet(wb: Workbook, table_name: str, rows: list[tuple]) -> None:
    ws = wb.create_sheet("1_PARAMS")
    title(ws, "PARAMS", "Workbook parameters consumed by later sync/parsers", 3)
    headers(ws, 4, ["KEY", "VALUE", "NOTE"], [34, 36, 80])
    last = write_rows(ws, 5, rows)
    make_table(ws, table_name, 4, last, 3)
    finish_sheet(ws)


def build_agent_workbook(
    *,
    filename: str,
    agent_id: str,
    agent_name: str,
    mission: str,
    owner: str,
    scenarios: list[tuple],
    gate_rows: list[tuple],
    signal_rows: list[tuple],
    output_rows: list[tuple],
    learning_rows: list[tuple],
) -> Path:
    wb = make_workbook(f"NEURAL - {agent_name}", mission)
    short_id = agent_id.replace("-", "")

    readme_sheet(
        wb,
        agent_id,
        mission,
        [
            ("Agent", f"{agent_name} ({agent_id})"),
            ("Owner", owner),
            ("Mode", "Scenario-id only, portfolio workbook, no production personal data."),
            ("Data", "Synthetic recruiter demo as of 2026-04-25."),
        ],
    )
    param_sheet(
        wb,
        f"tblParams{short_id}",
        [
            ("AGENT_ID", agent_id, ""),
            ("AGENT_NAME", agent_name, ""),
            ("OWNER", owner, ""),
            ("DATA_VERSION", "v2026.04.25", ""),
            ("REFRESH_DATE", REF_DATE, ""),
            ("INPUT_MODE", "scenario-id-only", ""),
            ("HITL_REQUIRED", "YES", "Compliance, legal or DPO final decision."),
        ],
    )

    scenario_headers = [
        "SCENARIO_ID", "LABEL", "PRODUCT", "CHANNEL", "SEGMENT", "JURISDICTION",
        "CLEAR_NOT_MISLEADING", "RISK_BALANCED", "SOURCE_ACTIVE", "NUM_VALIDATED",
        "MIFID_TARGET_MARKET", "PRIIPS_KID_CONSISTENT", "MICA_CRYPTO_MARKETING",
        "GDPR_CONSENT_PROFILING", "AI_ACT_DISCLOSURE", "HITL_COMPLIANCE",
        "EXPECTED_VERDICT", "NOTE",
    ]
    ws = wb.create_sheet("2_INPUT_SCENARIOS")
    title(ws, "INPUT SCENARIOS", "Frozen examples used to exercise PASS / REVIEW / BLOCK decisions", len(scenario_headers))
    headers(ws, 4, scenario_headers, [18, 36, 22, 18, 22, 14, 22, 18, 16, 16, 22, 24, 24, 24, 20, 20, 20, 82])
    last = write_rows(ws, 5, scenarios)
    make_table(ws, f"tbl{short_id}Scenarios", 4, last, len(scenario_headers))
    for col in range(7, 17):
        add_list_validation(ws, col, 5, last, YES_NO_NA)
    add_list_validation(ws, 17, 5, last, VERDICTS)
    add_status_formatting(ws, f"Q5:Q{last}")
    finish_sheet(ws)

    ws = wb.create_sheet("3_POLICY_GATES")
    title(ws, "POLICY GATES", "Deterministic gates that override any LLM output", 7)
    gate_headers = ["GATE_ID", "LABEL", "SEVERITY", "BLOCKING", "SOURCE_REF", "OWNER", "NOTE"]
    headers(ws, 4, gate_headers, [38, 50, 14, 12, 28, 24, 74])
    last = write_rows(ws, 5, gate_rows)
    make_table(ws, f"tbl{short_id}Gates", 4, last, len(gate_headers))
    add_list_validation(ws, 3, 5, last, SEVERITY)
    add_list_validation(ws, 4, 5, last, YES_NO)
    finish_sheet(ws)

    ws = wb.create_sheet("4_RESTRICTED_SIGNALS")
    title(ws, "RESTRICTED SIGNALS", "Terms, patterns and situations that trigger review or block", 6)
    sig_headers = ["SIGNAL_ID", "PATTERN", "TYPE", "SEVERITY", "ACTION", "RATIONALE"]
    headers(ws, 4, sig_headers, [18, 38, 20, 14, 18, 86])
    last = write_rows(ws, 5, signal_rows)
    make_table(ws, f"tbl{short_id}Signals", 4, last, len(sig_headers))
    add_list_validation(ws, 4, 5, last, SEVERITY)
    finish_sheet(ws)

    ws = wb.create_sheet("5_SCORING_ENGINE")
    title(ws, "SCORING ENGINE", "Excel formulas mirror the first deterministic gate implementation", 16)
    score_headers = [
        "SCENARIO_ID", "EXPECTED_VERDICT", "G_CLEAR", "G_BALANCE", "G_SOURCE", "G_NUM",
        "G_MIFID", "G_PRIIPS", "G_MICA", "G_GDPR", "G_AI", "G_HITL",
        "BLOCK_COUNT", "REVIEW_COUNT", "OBSERVED_DECISION", "PASS_FAIL",
    ]
    headers(ws, 4, score_headers, [18, 20, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 20, 14])
    for row in range(5, 5 + len(scenarios)):
        ws.cell(row, 1, f"='2_INPUT_SCENARIOS'!A{row}")
        ws.cell(row, 2, f"='2_INPUT_SCENARIOS'!Q{row}")
        for idx, src_col in enumerate("GHIJKLMNOP", 3):
            ws.cell(row, idx, f'=IF(\'2_INPUT_SCENARIOS\'!{src_col}{row}="NO","FAIL",IF(\'2_INPUT_SCENARIOS\'!{src_col}{row}="REVIEW","REVIEW","PASS"))')
        ws.cell(row, 13, f'=COUNTIF(C{row}:L{row},"FAIL")')
        ws.cell(row, 14, f'=COUNTIF(C{row}:L{row},"REVIEW")')
        ws.cell(row, 15, f'=IF(M{row}>0,"BLOCK",IF(N{row}>0,"PASS_WITH_REVIEW","PASS"))')
        ws.cell(row, 16, f'=IF(O{row}=B{row},"OK","FAIL")')
        for col in range(1, 17):
            ws.cell(row, col).border = THIN
            ws.cell(row, col).alignment = AC
    last = 4 + len(scenarios)
    make_table(ws, f"tbl{short_id}Scoring", 4, last, len(score_headers))
    add_status_formatting(ws, f"O5:O{last}")
    finish_sheet(ws)

    ws = wb.create_sheet("6_OUTPUT_PACK")
    title(ws, "OUTPUT PACK", "Portfolio-safe outputs for later site integration", 8)
    out_headers = ["OUTPUT_ID", "SCENARIO_ID", "PACK_SECTION", "DECISION_REF", "CONTENT", "EXPORT_READY", "REVIEW_OWNER", "HASH_SEED"]
    headers(ws, 4, out_headers, [18, 18, 22, 18, 88, 14, 24, 34])
    rows = []
    for idx, row in enumerate(output_rows, 5):
        output_id, scenario_id, section_name, content, review_owner = row
        rows.append((output_id, scenario_id, section_name, f"='5_SCORING_ENGINE'!O{idx}", content, "YES", review_owner, f"{agent_id}-{scenario_id}-v20260425"))
    last = write_rows(ws, 5, rows)
    make_table(ws, f"tbl{short_id}Outputs", 4, last, len(out_headers))
    add_list_validation(ws, 6, 5, last, YES_NO)
    finish_sheet(ws)

    ws = wb.create_sheet("7_DASHBOARD")
    title(ws, "DASHBOARD", "Executive portfolio dashboard", 4)
    headers(ws, 4, ["KPI", "VALUE", "FORMULA_SOURCE", "STATUS"], [34, 18, 52, 18])
    dashboard_rows = [
        ("Scenarios", "=COUNTA('2_INPUT_SCENARIOS'!A5:A50)", "2_INPUT_SCENARIOS", "OK"),
        ("PASS", '=COUNTIF(\'5_SCORING_ENGINE\'!O:O,"PASS")', "5_SCORING_ENGINE", "OK"),
        ("PASS_WITH_REVIEW", '=COUNTIF(\'5_SCORING_ENGINE\'!O:O,"PASS_WITH_REVIEW")', "5_SCORING_ENGINE", "OK"),
        ("BLOCK", '=COUNTIF(\'5_SCORING_ENGINE\'!O:O,"BLOCK")', "5_SCORING_ENGINE", "OK"),
        ("Formula tests OK", '=COUNTIF(\'5_SCORING_ENGINE\'!P:P,"OK")', "5_SCORING_ENGINE", "OK"),
    ]
    last = write_rows(ws, 5, dashboard_rows)
    make_table(ws, f"tbl{short_id}Dashboard", 4, last, 4)
    finish_sheet(ws)

    ws = wb.create_sheet("8_TESTSET")
    title(ws, "TESTSET", "Expected decisions used by verify_workbooks.py", 5)
    headers(ws, 4, ["SCENARIO_ID", "EXPECTED", "OBSERVED", "PASS_FAIL", "NOTE"], [18, 20, 20, 14, 74])
    for row in range(5, 5 + len(scenarios)):
        ws.cell(row, 1, f"='2_INPUT_SCENARIOS'!A{row}")
        ws.cell(row, 2, f"='2_INPUT_SCENARIOS'!Q{row}")
        ws.cell(row, 3, f"='5_SCORING_ENGINE'!O{row}")
        ws.cell(row, 4, f"='5_SCORING_ENGINE'!P{row}")
        ws.cell(row, 5, "Scenario exercises one MVP Banque Marketing gate path.")
        for col in range(1, 6):
            ws.cell(row, col).border = THIN
            ws.cell(row, col).alignment = AL
    last = 4 + len(scenarios)
    make_table(ws, f"tbl{short_id}Testset", 4, last, 5)
    finish_sheet(ws)

    ws = wb.create_sheet("9_LEARNINGS")
    title(ws, "LEARNINGS", "Recruiter-facing implementation notes and next steps", 5)
    headers(ws, 4, ["ID", "TOPIC", "LEARNING", "NEXT_STEP", "STATUS"], [12, 24, 74, 74, 16])
    last = write_rows(ws, 5, learning_rows)
    make_table(ws, f"tbl{short_id}Learnings", 4, last, 5)
    finish_sheet(ws)

    return save_workbook(wb, filename)
