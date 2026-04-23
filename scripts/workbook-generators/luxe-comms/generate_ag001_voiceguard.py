"""
NEURAL_AG001_MaisonVoiceGuard.xlsx
Moteur central de scoring brand — hard-fail, review, journalisation, learnings.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _styles import (
    AC, AL, BF, BF_EPAIS, BLANC_CASSE, BLEU, BORDEAUX, FOND_BLEU,
    FOND_JAUNE, FOND_ROUGE, FOND_VERT, FOND_VIOLET, GRIS, LANG, NOIR, OR,
    ORANGE, ROUGE, STATUS, VERT, VIOLET, YESNO,
    ajouter_table, attach_dv, convertir_dates, D, DT, dv_liste, ecrire_donnees,
    en_tetes, params_table, param_ligne, section, titre, today_iso,
)

OUT = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication\NEURAL_AG001_MaisonVoiceGuard.xlsx")


def onglet_readme(wb):
    ws = wb.create_sheet("0_README")
    titre(ws, "AG-001  —  MAISONVOICEGUARD", "Moteur de scoring brand. Toute sortie externe passe par ici.", 6)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 86

    section(ws, 4, 6, "Mission", NOIR)
    ecrire_donnees(ws, 5, [
        ["", "Role",        "Scorer chaque texte communication sur conformite charte. Refus automatique si score < seuil ou hard-fail detecte."],
        ["", "Input",       "Texte (FR ou EN), LANG, CONTEXTE (presse / event / social), appelant (USER / AG-002 / AG-003)."],
        ["", "Output",      "Score /100 + Decision (APPROVE / REWORK / REJECT) + Feedback structure."],
        ["", "Gate",        "BRAND — blocking par defaut. Fast-track 4h si CRISIS_MODE_ON=YES."],
        ["", "Appelants",   "AG-002 (presse), AG-003 (event), USER direct, CRISIS."],
        ["", "Referentiel", "FOUNDATIONS!2_BRAND_CHARTER, 3_BRAND_VOCAB_FR, 4_BRAND_VOCAB_EN."],
    ], align=AL)

    section(ws, 13, 6, "Logique de scoring", BORDEAUX)
    ecrire_donnees(ws, 14, [
        ["", "Score de base", "100"],
        ["", "Penalites",     "Tone, Forbidden_term_count, Preferred_missing, Structure, Identite, Claim_unproven"],
        ["", "Hard fail",     "Si HARD_FAIL_COUNT > 0 -> DECISION = REJECT immediat"],
        ["", "Seuil FR",      "Parametrable 1_PARAMS!B10 (defaut 75)"],
        ["", "Seuil EN",      "Parametrable 1_PARAMS!B11 (defaut 75)"],
        ["", "Decision",      "REJECT si HARD_FAIL>0, REWORK si SCORE < seuil, APPROVE sinon"],
    ], align=AL)

    section(ws, 22, 6, "Boucle d'apprentissage (10_LEARNINGS)", VIOLET)
    ecrire_donnees(ws, 23, [
        ["", "Principe",      "Chaque REWORK/REJECT est logue avec motif. Revue trimestrielle -> recalibration regles."],
        ["", "Trigger",       "Si un motif depasse 10 occurrences / mois -> flag pour ajustement regle."],
        ["", "Output",        "Ajustements pondus dans 3_BRAND_RULES + seed nouveaux termes 3_BRAND_VOCAB_FR."],
    ], align=AL)


def onglet_params(wb):
    ws = wb.create_sheet("1_PARAMS")
    titre(ws, "PARAMETRES AGENT", "tblParamsAG001 consommee via XLOOKUP par le moteur de scoring", 3)
    rows = [
        ("AGENT_ID",                "AG-001", ""),
        ("AGENT_NAME",              "MaisonVoiceGuard", ""),
        ("MAISON_ID",               "M-001", ""),
        ("MAISON_NAME",             "Maison pilote", ""),
        ("LANG_PRIMARY",            "FR", ""),
        ("LANG_SECONDARY",          "EN", ""),
        ("REVIEW_REQUIRED_DEFAULT", "YES", "Revue humaine obligatoire"),
        ("DATA_VERSION",            "v2026.04", ""),
        ("PROMPT_VERSION",          "v1.0.0", ""),
        ("THRESHOLD_FR",            75, "Score mini FR"),
        ("THRESHOLD_EN",            75, "Score mini EN"),
        ("CRISIS_MODE_ON",          "NO", "Si YES, SLA reduit a 4h"),
        ("SLA_STANDARD_H",          24, "Heures, standard"),
        ("SLA_CRISIS_H",            4,  "Heures, mode crise"),
        ("ESCALATION_ROLE",         "CMO", ""),
        ("LAST_REFRESH",            D(today_iso()), ""),
    ]
    params_table(ws, "tblParamsAG001", start_row=4, rows=rows)


def onglet_input_queue(wb):
    ws = wb.create_sheet("2_INPUT_QUEUE")
    titre(ws, "INPUT QUEUE", "File des textes entrants a scorer — alimentee par USER, AG-002, AG-003", 9)
    en_tetes(ws, 4, ["DOC_ID", "MAISON_ID", "SOURCE_AGENT", "LANG", "CONTEXTE", "PRIORITE",
                      "TEXTE", "SUBMITTED_AT", "STATUS"],
             [12, 10, 14, 8, 16, 10, 70, 18, 12])

    demo = [
        ("DOC-0001", "M-001", "AG-002", "FR", "Presse lifestyle", "P1",
         "La maison devoile une nouvelle collection de haute joaillerie, inspiree des archives de 1920. Chaque piece est ciselee a la main par les artisans de l'atelier parisien.",
         DT("2026-04-15 09:10"), "PENDING"),
        ("DOC-0002", "M-001", "AG-002", "FR", "Presse business", "P1",
         "Le meilleur groupe de luxe francais annonce une croissance unique au monde, grace a une strategie discount revolutionnaire.",
         DT("2026-04-15 10:22"), "PENDING"),
        ("DOC-0003", "M-001", "AG-003", "FR", "Event social", "P2",
         "Notre defile haute couture printemps souligne un savoir-faire d'exception, transmis de generation en generation.",
         DT("2026-04-15 11:04"), "PENDING"),
        ("DOC-0004", "M-001", "AG-002", "EN", "Press business", "P1",
         "The maison announces a fully sustainable collection crafted in our atelier, with the best materials ever used.",
         DT("2026-04-15 13:30"), "PENDING"),
        ("DOC-0005", "M-001", "USER",   "FR", "Social post", "P3",
         "Une piece rare, issue d'un geste artisanal preserve depuis 1923.",
         DT("2026-04-15 14:15"), "PENDING"),
        ("DOC-0006", "M-001", "AG-003", "FR", "Invitation VIP", "P2",
         "Nous vous invitons a notre cocktail exclusif, en l'honneur de la collection unique au monde.",
         DT("2026-04-15 16:00"), "PENDING"),
        ("DOC-0007", "M-001", "AG-002", "EN", "Press lifestyle", "P2",
         "A piece crafted in our Parisian atelier, a testament to enduring craftsmanship.",
         DT("2026-04-15 17:12"), "PENDING"),
        ("DOC-0008", "M-001", "CRISIS", "FR", "Reponse crise", "P0",
         "Suite aux interrogations mediatiques, la maison tient a preciser les conditions de production de sa collection automne.",
         DT("2026-04-15 18:00"), "PENDING"),
    ]
    last = ecrire_donnees(ws, 5, demo, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[8])
    ajouter_table(ws, "tblInputQueue", f"A4:I{last}")
    # R-06 : DV
    attach_dv(ws, f"C5:C{last}", ["USER", "AG-002", "AG-003", "CRISIS"])
    attach_dv(ws, f"D5:D{last}", LANG)
    attach_dv(ws, f"F5:F{last}", ["P0", "P1", "P2", "P3"])
    attach_dv(ws, f"I5:I{last}", ["PENDING", "IN_PROGRESS", "DONE", "CANCELLED"])

    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"P0"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"F5:F{last}", CellIsRule(operator="equal", formula=['"P1"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def onglet_brand_rules(wb):
    ws = wb.create_sheet("3_BRAND_RULES")
    titre(ws, "BRAND RULES (pondereees)", "Table de regles lues par le moteur — miroir de FOUNDATIONS mais pondere", 8)
    en_tetes(ws, 4, ["RULE_ID", "CATEGORIE", "REGLE", "NIVEAU", "POIDS", "LANG", "DETECTION_MODE", "ENABLED"],
             [12, 18, 60, 14, 10, 10, 22, 12])
    rules = [
        ("R-001", "Tone",         "Registre allusif, jamais declamatoire",                       "CRITICAL", 25, "FR", "SEMANTIC+VOCAB",  "YES"),
        ("R-002", "Tone",         "Pas de superlatif absolu",                                     "HIGH",     15, "FR", "VOCAB",           "YES"),
        ("R-003", "Structure",    "Accroche <140 caracteres + factuel",                           "MEDIUM",   10, "FR", "REGEX",           "YES"),
        ("R-004", "Identite",     "Pas de nom concurrent / comparatif",                           "CRITICAL", 25, "FR", "REGEX+VOCAB",     "YES"),
        ("R-005", "Claim",        "Affirmation RSE = prouvee (handoff AG-005)",                   "CRITICAL", 25, "FR", "VOCAB+HANDOFF",   "YES"),
        ("R-006", "Heritage",     "Reference historique datee doit etre sourcee",                 "HIGH",     15, "FR", "REGEX",           "YES"),
        ("R-007", "Formulation",  "Present de narration, pas futur vendeur",                      "MEDIUM",   10, "FR", "SEMANTIC",        "YES"),
        ("R-008", "Formulation",  "Pas d'exhortation commerciale",                                "HIGH",     15, "FR", "VOCAB",           "YES"),
        ("R-009", "Legal",        "Pas de revendication sante / medicale",                       "CRITICAL", 25, "FR", "VOCAB",           "YES"),
        ("R-010", "Tone",         "Allusion > revendication directe",                             "MEDIUM",   10, "FR", "SEMANTIC",        "YES"),
        ("R-011", "Identite",     "Coherence vocabulaire",                                         "HIGH",     15, "FR", "VOCAB",           "YES"),
        ("R-012", "Inclusive",    "Formulation inclusive",                                         "MEDIUM",   10, "FR", "SEMANTIC",        "YES"),
        ("R-013", "Tone",         "Neutral prestige tone in EN",                                  "CRITICAL", 25, "EN", "SEMANTIC+VOCAB",  "YES"),
        ("R-014", "Claim",        "'sustainable' requires evidence",                              "CRITICAL", 25, "EN", "VOCAB+HANDOFF",   "YES"),
        ("R-015", "Crisis",       "Crisis mode — fast-track mais review maintenue",               "HIGH",     15, "FR", "FLAG",            "YES"),
    ]
    last = ecrire_donnees(ws, 5, rules, fmt_cols={5: "0"}, align=AL)
    ajouter_table(ws, "tblBrandRulesAG001", f"A4:H{last}")
    # R-06 : DV
    attach_dv(ws, f"D5:D{last}", ["CRITICAL", "HIGH", "MEDIUM", "LOW"])
    attach_dv(ws, f"F5:F{last}", LANG)
    attach_dv(ws, f"G5:G{last}", ["VOCAB", "REGEX", "SEMANTIC", "SEMANTIC+VOCAB", "REGEX+VOCAB", "VOCAB+HANDOFF", "FLAG"])
    attach_dv(ws, f"H5:H{last}", YESNO)

    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=PatternFill(start_color=FOND_BLEU, end_color=FOND_BLEU, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_hard_fail(wb):
    ws = wb.create_sheet("4_HARD_FAIL_RULES")
    titre(ws, "HARD FAIL RULES", "Mots / patterns qui declenchent un REJECT immediat — pas de tolerance", 6)
    en_tetes(ws, 4, ["HF_ID", "PATTERN", "TYPE", "LANG", "CATEGORIE", "NOTE"],
             [10, 32, 14, 10, 18, 48])
    hf = [
        ("HF-001", "discount",              "LITERAL", "FR", "Commercial",   "Jamais tolere en luxe"),
        ("HF-002", "promo",                 "LITERAL", "FR", "Commercial",   "Jamais tolere en luxe"),
        ("HF-003", "unique au monde",       "LITERAL", "FR", "Superlatif",   "Superlatif absolu"),
        ("HF-004", "le meilleur",           "LITERAL", "FR", "Superlatif",   "Superlatif absolu"),
        ("HF-005", "usine",                 "LITERAL", "FR", "Identite",     "Hors charte"),
        ("HF-006", "no 1 mondial",          "LITERAL", "FR", "Superlatif",   "Comparatif prohibe"),
        ("HF-007", "anti-age",              "LITERAL", "FR", "Claim sante",  "Revendication medicale"),
        ("HF-008", "eco-responsable",       "LITERAL", "FR", "Claim RSE",    "Declenche gate AG-005 imperatif"),
        ("HF-009", r"[Cc]oncurrent\s+\w+",  "REGEX",   "FR", "Identite",     "Comparatif concurrent"),
        ("HF-010", "discount",              "LITERAL", "EN", "Commercial",   "Never"),
        ("HF-011", "sale",                  "LITERAL", "EN", "Commercial",   "Never"),
        ("HF-012", "the best",              "LITERAL", "EN", "Superlative",  "Absolute"),
        ("HF-013", "sustainable",           "LITERAL", "EN", "ESG claim",    "Triggers AG-005"),
        ("HF-014", "eco-friendly",          "LITERAL", "EN", "ESG claim",    "Triggers AG-005"),
        ("HF-015", "factory",               "LITERAL", "EN", "Identite",     "Off-brand"),
        ("HF-016", "anti-aging",            "LITERAL", "EN", "Health claim", "Medical claim"),
        ("HF-017", "number one",            "LITERAL", "EN", "Superlative",  "Comparative"),
    ]
    last = ecrire_donnees(ws, 5, hf, align=AL)
    ajouter_table(ws, "tblHardFail", f"A4:F{last}")


def onglet_scoring(wb):
    ws = wb.create_sheet("5_SCORING_ENGINE")
    titre(ws, "SCORING ENGINE", "1 ligne par DOC_ID — calcul score, hard-fail, decision, SLA", 19)
    # R-07 : ajout SLA_TARGET_H (branche sur CRISIS_MODE_ON) et SLA_MET calcule
    headers = ["RUN_ID", "DOC_ID", "MAISON_ID", "LANG", "TIMESTAMP",
               "PEN_TONE", "PEN_FORBIDDEN", "PEN_PREFERRED_MISSING", "PEN_STRUCTURE",
               "PEN_IDENTITY", "PEN_CLAIM", "HARD_FAIL_COUNT",
               "SCORE", "THRESHOLD", "DECISION", "FEEDBACK_SHORT",
               "DECIDED_AT", "SLA_TARGET_H", "SLA_MET"]
    en_tetes(ws, 4, headers, [16, 12, 10, 8, 18, 10, 12, 16, 12, 12, 12, 12, 10, 10, 12, 40, 18, 14, 10])

    # R-01 : timestamps en datetime + DECIDED_AT en datetime
    demo = [
        ("R-20260415-001", "DOC-0001", "M-001", "FR", DT("2026-04-15 09:12"), 0,  0,  0, 0, 0,  0, 0, "", "", "", "Draft conforme.",                             DT("2026-04-15 09:16"), "", ""),
        ("R-20260415-002", "DOC-0002", "M-001", "FR", DT("2026-04-15 10:25"), 25, 50, 0, 0, 25, 25, 4, "", "", "", "Superlatif + discount + concurrent.",          DT("2026-04-15 10:40"), "", ""),
        ("R-20260415-003", "DOC-0003", "M-001", "FR", DT("2026-04-15 11:06"), 0,  0,  0, 0, 0,  0,  0, "", "", "", "Conforme charte.",                             DT("2026-04-15 11:10"), "", ""),
        ("R-20260415-004", "DOC-0004", "M-001", "EN", DT("2026-04-15 13:33"), 25, 25, 0, 0, 0,  25, 2, "", "", "", "sustainable + best -> AG-005 + reject.",       DT("2026-04-15 13:50"), "", ""),
        ("R-20260415-005", "DOC-0005", "M-001", "FR", DT("2026-04-15 14:17"), 0,  0,  0, 0, 0,  0,  0, "", "", "", "Conforme.",                                    DT("2026-04-15 14:20"), "", ""),
        ("R-20260415-006", "DOC-0006", "M-001", "FR", DT("2026-04-15 16:03"), 10, 25, 0, 0, 0,  0,  1, "", "", "", "unique au monde hard-fail.",                    DT("2026-04-15 16:10"), "", ""),
        ("R-20260415-007", "DOC-0007", "M-001", "EN", DT("2026-04-15 17:15"), 0,  0,  0, 0, 0,  0,  0, "", "", "", "Clean EN draft.",                              DT("2026-04-15 17:20"), "", ""),
        ("R-20260415-008", "DOC-0008", "M-001", "FR", DT("2026-04-15 18:02"), 0,  0,  0, 0, 0,  0,  0, "", "", "", "Crisis fast-track, pas de hard fail.",          DT("2026-04-15 19:40"), "", ""),
    ]
    last = ecrire_donnees(ws, 5, demo, fmt_cols={6: "0", 7: "0", 8: "0", 9: "0", 10: "0", 11: "0", 12: "0"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5, 17])

    # R-05 : formules THRESHOLD via XLOOKUP sur tblParamsAG001 (plus de refs absolues)
    # R-07 : SLA_TARGET_H branche sur CRISIS_MODE_ON
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=13, value=f"=MAX(0,100-(F{r}+G{r}+H{r}+I{r}+J{r}+K{r}))")
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = "0"

        # Threshold via XLOOKUP (R-05)
        c = ws.cell(row=r, column=14,
                     value=f'=IF(D{r}="FR",XLOOKUP("THRESHOLD_FR",tblParamsAG001[KEY],tblParamsAG001[VALUE]),XLOOKUP("THRESHOLD_EN",tblParamsAG001[KEY],tblParamsAG001[VALUE]))')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF
        c.number_format = "0"

        c = ws.cell(row=r, column=15, value=f'=IF(L{r}>0,"REJECT",IF(M{r}<N{r},"REWORK","APPROVE"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

        # R-07 : SLA_TARGET_H = SLA_CRISIS_H si CRISIS_MODE_ON=YES sinon SLA_STANDARD_H
        c = ws.cell(row=r, column=18,
                     value=f'=IF(XLOOKUP("CRISIS_MODE_ON",tblParamsAG001[KEY],tblParamsAG001[VALUE])="YES",XLOOKUP("SLA_CRISIS_H",tblParamsAG001[KEY],tblParamsAG001[VALUE]),XLOOKUP("SLA_STANDARD_H",tblParamsAG001[KEY],tblParamsAG001[VALUE]))')
        c.font = Font(name="Calibri", size=9)
        c.alignment = AC
        c.border = BF
        c.number_format = "0"

        # R-07 : SLA_MET calcule = (DECIDED_AT - TIMESTAMP) en heures <= SLA_TARGET_H ?
        c = ws.cell(row=r, column=19,
                     value=f'=IF(OR(E{r}="",Q{r}=""),"-",IF((Q{r}-E{r})*24<=R{r},"YES","NO"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ajouter_table(ws, "tblScoring", f"A4:S{last}")
    attach_dv(ws, f"D5:D{last}", LANG)

    # Conditional formatting sur DECISION
    ws.conditional_formatting.add(f"O5:O{last}", CellIsRule(operator="equal", formula=['"APPROVE"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"O5:O{last}", CellIsRule(operator="equal", formula=['"REWORK"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"O5:O{last}", CellIsRule(operator="equal", formula=['"REJECT"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    # Score colored
    ws.conditional_formatting.add(f"M5:M{last}", FormulaRule(formula=[f"M5<50"], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"M5:M{last}", FormulaRule(formula=[f"AND(M5>=50,M5<75)"], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"M5:M{last}", FormulaRule(formula=[f"M5>=75"], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))


def onglet_review_queue(wb):
    ws = wb.create_sheet("6_REVIEW_QUEUE")
    titre(ws, "REVIEW QUEUE", "File d'attente humaine — REWORK et REJECT", 10)
    en_tetes(ws, 4, ["REVIEW_ID", "RUN_ID", "DOC_ID", "DECISION_AGENT", "APPROVER_ROLE", "ASSIGNED_TO",
                      "DUE_AT", "STATUS", "COMMENT_APPROVER", "CLOSED_AT"],
             [14, 16, 12, 14, 22, 20, 18, 14, 40, 18])
    rev = [
        ("REV-0001", "R-20260415-002", "DOC-0002", "REJECT", "Brand Director", "A. Dupont",  DT("2026-04-16 10:25"), "PENDING", "",                       None),
        ("REV-0002", "R-20260415-004", "DOC-0004", "REJECT", "Brand Director", "A. Dupont",  DT("2026-04-16 13:33"), "PENDING", "",                       None),
        ("REV-0003", "R-20260415-006", "DOC-0006", "REJECT", "Brand Director", "A. Dupont",  DT("2026-04-16 16:03"), "PENDING", "",                       None),
        ("REV-0004", "R-20260415-008", "DOC-0008", "APPROVE","CMO",            "G. Petit",   DT("2026-04-15 20:02"), "CLOSED",  "Approuve fast-track.",   DT("2026-04-15 19:45")),
    ]
    last = ecrire_donnees(ws, 5, rev, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[7, 10])
    ajouter_table(ws, "tblReviewQueue", f"A4:J{last}")
    attach_dv(ws, f"D5:D{last}", ["APPROVE", "REWORK", "REJECT", "PENDING"])
    attach_dv(ws, f"H5:H{last}", ["PENDING", "CLOSED", "ESCALATED"])

    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"PENDING"'], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))
    ws.conditional_formatting.add(f"H5:H{last}", CellIsRule(operator="equal", formula=['"CLOSED"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))


def onglet_outputs_log(wb):
    ws = wb.create_sheet("7_OUTPUTS_LOG")
    titre(ws, "OUTPUTS LOG", "Journal exhaustif — exporte vers MASTER!8_GLOBAL_LOG_IMPORT", 13)
    en_tetes(ws, 4, ["RUN_ID", "AGENT_ID", "MAISON_ID", "DOC_ID", "TIMESTAMP", "LANG",
                      "FINAL_STATUS", "RISK_LEVEL", "GATE_TRIGGERED", "SLA_MET",
                      "TOTAL_MINUTES", "TOKENS_IN", "TOKENS_OUT", "COST_USD"],
             [16, 10, 10, 12, 18, 8, 14, 12, 14, 10, 12, 12, 12, 12])
    rows = [
        ("R-20260415-001", "AG-001", "M-001", "DOC-0001", DT("2026-04-15 09:12"), "FR", "APPROVED", "LOW",      "BRAND", "YES", 4.2,  1800, 900, 0.022),
        ("R-20260415-002", "AG-001", "M-001", "DOC-0002", DT("2026-04-15 10:25"), "FR", "REJECTED", "HIGH",     "BRAND", "YES", 2.9,  1400, 400, 0.012),
        ("R-20260415-003", "AG-001", "M-001", "DOC-0003", DT("2026-04-15 11:06"), "FR", "APPROVED", "LOW",      "BRAND", "YES", 3.1,  1700, 800, 0.020),
        ("R-20260415-004", "AG-001", "M-001", "DOC-0004", DT("2026-04-15 13:33"), "EN", "REJECTED", "HIGH",     "BRAND", "YES", 3.3,  1600, 500, 0.017),
        ("R-20260415-005", "AG-001", "M-001", "DOC-0005", DT("2026-04-15 14:17"), "FR", "APPROVED", "LOW",      "BRAND", "YES", 2.0,  900,  400, 0.011),
        ("R-20260415-006", "AG-001", "M-001", "DOC-0006", DT("2026-04-15 16:03"), "FR", "REJECTED", "MEDIUM",   "BRAND", "YES", 2.5,  1100, 300, 0.010),
        ("R-20260415-007", "AG-001", "M-001", "DOC-0007", DT("2026-04-15 17:15"), "EN", "APPROVED", "LOW",      "BRAND", "YES", 2.2,  1000, 500, 0.013),
        ("R-20260415-008", "AG-001", "M-001", "DOC-0008", DT("2026-04-15 18:02"), "FR", "IN_REVIEW","LOW",      "BRAND", "YES", 1.8,  900,  400, 0.011),
    ]
    last = ecrire_donnees(ws, 5, rows, fmt_cols={11: "0.00", 12: "0", 13: "0", 14: "0.000"}, align=AL)
    convertir_dates(ws, 5, last, date_cols=[], datetime_cols=[5])
    ajouter_table(ws, "tblOutputsLog", f"A4:N{last}")


def onglet_kpi(wb):
    ws = wb.create_sheet("8_KPI_DASHBOARD")
    titre(ws, "KPI DASHBOARD", "Score moyen, taux approve, hard-fails, cout", 4)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18

    section(ws, 4, 4, "Indicateurs", NOIR)
    kpis = [
        ("Score moyen",                       '=IFERROR(AVERAGE(tblScoring[SCORE]),0)',          "0.0"),
        ("Taux APPROVE",                      '=IFERROR(COUNTIFS(tblScoring[DECISION],"APPROVE")/COUNTA(tblScoring[RUN_ID]),0)', "0.0%"),
        ("Taux REWORK",                       '=IFERROR(COUNTIFS(tblScoring[DECISION],"REWORK")/COUNTA(tblScoring[RUN_ID]),0)', "0.0%"),
        ("Taux REJECT",                       '=IFERROR(COUNTIFS(tblScoring[DECISION],"REJECT")/COUNTA(tblScoring[RUN_ID]),0)', "0.0%"),
        ("Docs avec hard fail >=1",           '=COUNTIFS(tblScoring[HARD_FAIL_COUNT],">0")',                                   "0"),
        ("Runs total",                        '=COUNTA(tblScoring[RUN_ID])',                      "0"),
        ("Cout USD cumule",                   '=IFERROR(SUM(tblOutputsLog[COST_USD]),0)',         "0.000"),
        ("Cout moyen USD / run",              '=IFERROR(AVERAGE(tblOutputsLog[COST_USD]),0)',     "0.000"),
    ]
    for i, (lbl, formula, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = AL
        ws.cell(row=r, column=2).border = BF
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
        c.alignment = AC
        c.border = BF_EPAIS
        c.number_format = fmt

    section(ws, 15, 4, "Distribution des scores", BORDEAUX)
    buckets = [("<50",    f'=COUNTIFS(tblScoring[SCORE],"<50")'),
               ("50-74",  f'=COUNTIFS(tblScoring[SCORE],">=50",tblScoring[SCORE],"<75")'),
               ("75-89",  f'=COUNTIFS(tblScoring[SCORE],">=75",tblScoring[SCORE],"<90")'),
               (">=90",   f'=COUNTIFS(tblScoring[SCORE],">=90")')]
    for i, (lbl, fm) in enumerate(buckets):
        r = 16 + i
        ws.cell(row=r, column=2, value=lbl).border = BF
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True)
        c = ws.cell(row=r, column=3, value=fm)
        c.border = BF; c.alignment = AC; c.number_format = "0"

    ch = BarChart()
    data = Reference(ws, min_col=3, min_row=16, max_row=19)
    labels = Reference(ws, min_col=2, min_row=16, max_row=19)
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(labels)
    ch.title = "Distribution des scores"
    ch.height = 8
    ch.width = 14
    ws.add_chart(ch, "E15")


def onglet_testset(wb):
    ws = wb.create_sheet("9_TESTSET")
    titre(ws, "TESTSET", "Cas de test — EXPECTED vs OBSERVED, calcul PASS/FAIL", 10)
    # R-12 : colonnes OBSERVED + PASS
    en_tetes(ws, 4, ["TEST_ID", "CATEGORIE", "LANG", "TEXTE_TEST",
                      "EXPECTED_DECISION", "EXPECTED_HARDFAIL",
                      "OBSERVED_DECISION", "OBSERVED_HARDFAIL", "PASS_FAIL", "NOTE"],
             [10, 16, 8, 60, 18, 18, 18, 18, 12, 28])
    tests = [
        ("T-001", "Golden",         "FR", "Une piece rare issue d'un savoir-faire transmis depuis 1923.",         "APPROVE", 0, "", "", "", "Vocab preferred + heritage"),
        ("T-002", "Hardfail",       "FR", "Notre nouvelle collection en discount pour un week-end.",              "REJECT",  1, "", "", "", "discount"),
        ("T-003", "Hardfail",       "FR", "Le meilleur bijou du monde, unique au monde.",                         "REJECT",  2, "", "", "", "2 hardfails"),
        ("T-004", "Claim",          "FR", "Notre collection entierement eco-responsable est sans precedent.",     "REJECT",  1, "", "", "", "eco-responsable + sans precedent"),
        ("T-005", "Hardfail",       "FR", "Une piece fabriquee dans notre usine avec soin.",                       "REJECT",  1, "", "", "", "usine"),
        ("T-006", "Golden",         "EN", "A piece crafted in our Parisian atelier.",                              "APPROVE", 0, "", "", "", "Clean"),
        ("T-007", "Hardfail",       "EN", "The best sustainable product from our factory.",                        "REJECT",  3, "", "", "", "3 hardfails"),
        ("T-008", "Legal",          "FR", "Effet anti-age prouve scientifiquement.",                                "REJECT",  1, "", "", "", "anti-age medical"),
        ("T-009", "Golden",         "EN", "A limited run of 48 numbered pieces.",                                  "APPROVE", 0, "", "", "", "Scarcity OK"),
        ("T-010", "Edge",           "FR", "Piece exclusive, reservee a nos clients prives.",                       "REWORK",  0, "", "", "", "exclusif a reviewer"),
        ("T-011", "Crisis",         "FR", "La maison tient a clarifier les conditions de production.",              "APPROVE", 0, "", "", "", "Crisis ton factuel"),
        ("T-012", "Heritage",       "FR", "Inspiree des archives maison 1987 (cote A-1987-032).",                  "APPROVE", 0, "", "", "", "Source citee"),
    ]
    last = ecrire_donnees(ws, 5, tests, fmt_cols={6: "0", 8: "0"}, align=AL)
    # R-12 : PASS_FAIL formule
    for r in range(5, last + 1):
        c = ws.cell(row=r, column=9,
                     value=f'=IF(OR(G{r}="",H{r}=""),"-",IF(AND(E{r}=G{r},F{r}=H{r}),"PASS","FAIL"))')
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = AC
        c.border = BF_EPAIS

    ajouter_table(ws, "tblTestset", f"A4:J{last}")
    attach_dv(ws, f"E5:E{last}", ["APPROVE", "REWORK", "REJECT"])
    attach_dv(ws, f"G5:G{last}", ["", "APPROVE", "REWORK", "REJECT"])
    attach_dv(ws, f"C5:C{last}", LANG)

    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill(start_color=FOND_VERT, end_color=FOND_VERT, fill_type="solid")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))


def onglet_learnings(wb):
    ws = wb.create_sheet("10_LEARNINGS")
    titre(ws, "LEARNINGS — BOUCLE D'APPRENTISSAGE", "Capitalisation post-run : motifs de REWORK/REJECT, ajustements", 8)
    en_tetes(ws, 4, ["LEARN_ID", "DATE", "SOURCE_RUN", "MOTIF", "FREQUENCE_MOIS", "ACTION_PROPOSEE", "OWNER", "STATUT"],
             [12, 12, 18, 40, 16, 44, 18, 14])
    lrn = [
        ("LRN-001", D(today_iso()), "R-20260415-002", "Superlatif + concurrent (hardfail R-004 frequent)",  3, "Ajouter regex concurrent multi-mots",        "Brand Dir",  "OPEN"),
        ("LRN-002", D(today_iso()), "R-20260415-004", "'sustainable' en EN mal intercepte avant draft",     4, "Pre-hook dans AG-002 avant passage",         "AI Ops",     "OPEN"),
        ("LRN-003", D(today_iso()), "R-20260415-006", "'exclusif' non bloque mais trop utilise",            5, "Ajouter regle REVIEW pondere dans 3_BRAND_RULES", "Brand Dir", "OPEN"),
        ("LRN-004", D(today_iso()), "-",              "PEN_STRUCTURE sous-utilise (0 detection 7j)",         0, "Activer regex structure",                    "AI Ops",     "OPEN"),
    ]
    last = ecrire_donnees(ws, 5, lrn, align=AL)
    convertir_dates(ws, 5, last, date_cols=[2])
    ajouter_table(ws, "tblLearnings", f"A4:H{last}")
    attach_dv(ws, f"H5:H{last}", ["OPEN", "IN_REVIEW", "ACCEPTED", "REJECTED", "DONE"])

    ws.conditional_formatting.add(f"E5:E{last}", FormulaRule(formula=[f"E5>=10"], fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid")))
    ws.conditional_formatting.add(f"E5:E{last}", FormulaRule(formula=[f"AND(E5>=5,E5<10)"], fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")))


def build():
    wb = Workbook()
    wb.remove(wb.active)
    onglet_readme(wb)
    onglet_params(wb)
    onglet_input_queue(wb)
    onglet_brand_rules(wb)
    onglet_hard_fail(wb)
    onglet_scoring(wb)
    onglet_review_queue(wb)
    onglet_outputs_log(wb)
    onglet_kpi(wb)
    onglet_testset(wb)
    onglet_learnings(wb)

    order = ["0_README", "1_PARAMS", "2_INPUT_QUEUE", "3_BRAND_RULES",
             "4_HARD_FAIL_RULES", "5_SCORING_ENGINE", "6_REVIEW_QUEUE",
             "7_OUTPUTS_LOG", "8_KPI_DASHBOARD", "9_TESTSET", "10_LEARNINGS"]
    wb._sheets = [wb[n] for n in order]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    build()
