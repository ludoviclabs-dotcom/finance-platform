"""
Build DDA_MarketingGuard_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'DDA_MarketingGuard' (branche Assurances — Marketing).

7 onglets : README, Checklist DDA 12 points, Cas communications, Verdict
par cas, Redlines, AI Act Disclosure, Limites.
6 communications marketing à auditer (auto, MRH, santé, prévoyance, vie, comparateur).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=NEURAL_VIOLET, end_color=NEURAL_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def yellow_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def write_title(ws, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, bold=False, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict[int, float]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ────────────────────────────────────────────────────────────────────
# Référentiel DDA — 12 points devoir de conseil + clarté contractuelle
# ────────────────────────────────────────────────────────────────────
DDA_CHECKLIST = [
    {
        "id": "P01",
        "axe": "Identification",
        "regle": "Identité du distributeur clairement mentionnée (nom, statut, ORIAS)",
        "ref": "DDA art. 18 + Code des assurances L.521-2",
        "obligatoire": "Oui",
    },
    {
        "id": "P02",
        "axe": "Identification",
        "regle": "Statut (mandataire, agent général, courtier, intermédiaire à titre accessoire)",
        "ref": "DDA art. 18 + Code des assurances L.521-2",
        "obligatoire": "Oui",
    },
    {
        "id": "P03",
        "axe": "Rémunération",
        "regle": "Nature de la rémunération (commission / honoraires / mixte) divulguée",
        "ref": "DDA art. 19 + Code des assurances L.521-2",
        "obligatoire": "Oui",
    },
    {
        "id": "P04",
        "axe": "Conseil",
        "regle": "Recueil des exigences et besoins du client documenté",
        "ref": "DDA art. 20 + Code des assurances L.521-4",
        "obligatoire": "Oui",
    },
    {
        "id": "P05",
        "axe": "Conseil",
        "regle": "Justification de l'adéquation produit/besoin (test d'adéquation)",
        "ref": "DDA art. 20 + Code des assurances L.521-4",
        "obligatoire": "Oui",
    },
    {
        "id": "P06",
        "axe": "Information",
        "regle": "IPID fourni avant souscription (Insurance Product Information Document)",
        "ref": "DDA art. 20 §5 + Règlement d'exécution UE 2017/1469",
        "obligatoire": "Oui",
    },
    {
        "id": "P07",
        "axe": "Clarté",
        "regle": "Langage clair, non trompeur, exact (Acte délégué clarté contractuelle 2025)",
        "ref": "Acte délégué (UE) 2025/xxx + ACPR Recommandation 2024-R-01",
        "obligatoire": "Oui",
    },
    {
        "id": "P08",
        "axe": "Clarté",
        "regle": "Distinction visuelle claims/promesses vs. garanties contractuelles",
        "ref": "ACPR Recommandation 2024-R-01 + DGCCRF",
        "obligatoire": "Oui",
    },
    {
        "id": "P09",
        "axe": "Greenwashing",
        "regle": "Allégations ESG/climat sourcées et vérifiables (anti-greenwashing)",
        "ref": "ACPR communiqué mars 2026 + Règlement (UE) 2024/1760 CSRD",
        "obligatoire": "Oui",
    },
    {
        "id": "P10",
        "axe": "Renonciation",
        "regle": "Mention du droit de renonciation (14 jours assurance non-vie / 30 jours vie)",
        "ref": "Code des assurances L.112-2-1 + L.132-5-1",
        "obligatoire": "Oui",
    },
    {
        "id": "P11",
        "axe": "Réclamation",
        "regle": "Modalités de réclamation et coordonnées du médiateur",
        "ref": "ACPR Recommandation 2016-R-02 + Code des assurances",
        "obligatoire": "Oui",
    },
    {
        "id": "P12",
        "axe": "IA",
        "regle": "Mention « contenu généré par IA » si applicable (EU AI Act art. 50)",
        "ref": "Règlement (UE) 2024/1689 EU AI Act, art. 50 — applicable août 2026",
        "obligatoire": "Oui (août 2026)",
    },
]


# ────────────────────────────────────────────────────────────────────
# Cas — 6 communications marketing à auditer
# ────────────────────────────────────────────────────────────────────
CASES = [
    {
        "id": "COMM-01",
        "branche": "Auto",
        "canal": "Email acquisition",
        "format": "Email B2C",
        "titre": "« Économisez 30% sur votre assurance auto en 5 minutes »",
        "contenu": (
            "Découvrez notre nouvelle offre auto. Économisez jusqu'à 30%* sur votre prime "
            "annuelle ! Devis en 5 minutes. Sans engagement. Cliquez ici pour souscrire."
            "\n\n*Économies constatées sur la base d'études internes 2025."
        ),
        # Verdicts par règle DDA (P01..P12) — 'OK' / 'KO' / 'WARN'
        "verdicts": {
            "P01": "KO", "P02": "KO", "P03": "KO", "P04": "WARN", "P05": "WARN",
            "P06": "KO", "P07": "WARN", "P08": "KO", "P09": "OK",
            "P10": "KO", "P11": "KO", "P12": "KO",
        },
        "redlines": [
            ("P01-P02", "« Découvrez notre nouvelle offre auto »",
             "« [Nom de l'assureur], [statut ORIAS n° XXXXX], vous propose son offre auto »"),
            ("P08", "« Économisez jusqu'à 30%* sur votre prime annuelle ! »",
             "« Selon votre profil, votre prime peut être jusqu'à 30% inférieure à celle de votre contrat actuel.* »"),
            ("P06", "(absent)",
             "Ajouter avant le bouton « souscrire » : « Avant souscription, consultez l'IPID disponible ici → [lien] »"),
            ("P10", "(absent)",
             "« Vous bénéficiez d'un droit de renonciation de 14 jours après signature, sans frais ni motif. »"),
            ("P12", "(absent)",
             "Mention obligatoire dès août 2026 si génération IA : « Cet email a été rédigé avec une assistance IA, validé par notre équipe marketing. »"),
        ],
    },
    {
        "id": "COMM-02",
        "branche": "MRH",
        "canal": "Landing page",
        "format": "Page web",
        "titre": "« L'assurance habitation la plus complète du marché »",
        "contenu": (
            "L'assurance habitation la plus complète du marché. Couverture totale, "
            "sans franchise, indemnisation en 24h. Notre IA analyse votre dossier et "
            "vous garantit la meilleure offre. Souscrivez en ligne maintenant."
        ),
        "verdicts": {
            "P01": "WARN", "P02": "WARN", "P03": "KO", "P04": "KO", "P05": "KO",
            "P06": "KO", "P07": "KO", "P08": "KO", "P09": "OK",
            "P10": "KO", "P11": "KO", "P12": "WARN",
        },
        "redlines": [
            ("P07-P08", "« L'assurance habitation la plus complète du marché »",
             "Allégation absolue interdite (DGCCRF). Reformuler : « Une couverture habitation conçue pour [persona cible] » avec preuves comparatives sourcées."),
            ("P07", "« Couverture totale, sans franchise »",
             "Faux dans 99% des contrats. Préciser les exclusions et la franchise réelle (renvoi CG art. X)."),
            ("P07", "« indemnisation en 24h »",
             "Promesse contractuelle implicite. Reformuler : « Délai d'indemnisation moyen constaté : X jours sur les sinistres simples 2025. »"),
            ("P05-P06", "« Notre IA vous garantit la meilleure offre »",
             "Pas de garantie possible sans recueil besoins. Ajouter parcours « test d'adéquation » + IPID avant souscription."),
            ("P12", "Mention IA absente",
             "Si l'analyse de dossier est faite par IA, mention obligatoire : « Décision assistée par un système d'IA, supervisée par un conseiller. »"),
        ],
    },
    {
        "id": "COMM-03",
        "branche": "Santé",
        "canal": "Post LinkedIn",
        "format": "Réseau social",
        "titre": "« Notre nouvelle complémentaire santé éco-responsable »",
        "contenu": (
            "Fiers d'annoncer notre complémentaire santé verte 🌱 ! 100% durable, "
            "investissements ESG, neutre en carbone. Pour chaque contrat souscrit, "
            "nous plantons un arbre. Rejoignez la révolution responsable. #ESG #Santé"
        ),
        "verdicts": {
            "P01": "KO", "P02": "KO", "P03": "KO", "P04": "WARN", "P05": "WARN",
            "P06": "KO", "P07": "WARN", "P08": "WARN", "P09": "KO",
            "P10": "KO", "P11": "KO", "P12": "WARN",
        },
        "redlines": [
            ("P09", "« 100% durable, neutre en carbone »",
             "Allégations vertes interdites sans preuve (ACPR mars 2026 + CSRD). Remplacer par : « Notre fonds investit X% dans des actifs labellisés [label précis] — voir notre rapport de durabilité 2025. »"),
            ("P09", "« investissements ESG »",
             "Préciser le label/règlement applicable : SFDR art. 8 ou art. 9, taxonomie UE, etc."),
            ("P09", "« nous plantons un arbre »",
             "Préciser : partenaire (nom), nombre d'arbres certifiés (label Bord-de-Forêt, Reforest'Action), et reportage de suivi."),
            ("P01-P02", "(absent)",
             "Identification de l'assureur + statut + ORIAS obligatoire même sur réseau social."),
            ("P06", "(absent)",
             "Lien vers IPID obligatoire avant tout CTA souscription."),
        ],
    },
    {
        "id": "COMM-04",
        "branche": "Prévoyance",
        "canal": "Mailing courrier",
        "format": "Lettre papier",
        "titre": "« Protégez vos proches en cas de coup dur »",
        "contenu": (
            "Madame, Monsieur,\n\nProtégez vos proches en cas de coup dur. Notre garantie "
            "décès vous offre jusqu'à 500 000€ de capital pour votre famille, sans "
            "questionnaire médical*, à partir de 12€/mois.\n\nRetournez le coupon-réponse "
            "ci-joint avant le 31/05/2026.\n\n*Sous réserve d'acceptation par nos services."
        ),
        "verdicts": {
            "P01": "OK", "P02": "WARN", "P03": "KO", "P04": "WARN", "P05": "WARN",
            "P06": "WARN", "P07": "WARN", "P08": "KO", "P09": "OK",
            "P10": "KO", "P11": "KO", "P12": "OK",
        },
        "redlines": [
            ("P07-P08", "« sans questionnaire médical* »",
             "L'astérisque renvoie à « sous réserve d'acceptation » — c'est trompeur. Reformuler : « Sans questionnaire de santé pour les capitaux jusqu'à X€. Au-delà, un questionnaire simplifié vous sera demandé. »"),
            ("P10", "(absent)",
             "Droit de renonciation 30 jours obligatoire pour assurance vie/décès. Mention manquante."),
            ("P11", "(absent)",
             "Coordonnées service réclamations + médiateur obligatoires en pied de courrier."),
            ("P03", "(absent)",
             "Si commercialisé via courtier, mention de la commission/rémunération obligatoire."),
            ("P05", "« Protégez vos proches en cas de coup dur »",
             "Pas de recueil besoins en mailing. Le CTA doit renvoyer vers un parcours conseil avec test d'adéquation, pas vers une souscription directe."),
        ],
    },
    {
        "id": "COMM-05",
        "branche": "Vie épargne",
        "canal": "Brochure print",
        "format": "Brochure agence",
        "titre": "« Notre contrat vie multisupport — 4,5% de rendement 2025 »",
        "contenu": (
            "Notre contrat d'assurance vie multisupport vous offre un rendement attractif. "
            "Performance fonds euros 2025 : 4,5%*. Diversifiez vos placements en unités de "
            "compte avec un accompagnement personnalisé. Frais d'entrée offerts pour toute "
            "souscription avant le 30/06/2026.\n\n*Performance nette de frais de gestion. "
            "Les performances passées ne préjugent pas des performances futures."
        ),
        "verdicts": {
            "P01": "WARN", "P02": "WARN", "P03": "KO", "P04": "WARN", "P05": "WARN",
            "P06": "OK", "P07": "WARN", "P08": "WARN", "P09": "OK",
            "P10": "OK", "P11": "WARN", "P12": "OK",
        },
        "redlines": [
            ("P07-P08", "« 4,5% de rendement »",
             "Préciser : rendement fonds euros uniquement, NET de frais sur encours, BRUT de prélèvements sociaux et fiscalité. Cible : indication claire « net/brut »."),
            ("P08", "« Diversifiez vos placements en unités de compte »",
             "UC = risque de perte en capital. Mention obligatoire : « Les unités de compte présentent un risque de perte en capital. »"),
            ("P03", "(absent)",
             "Mention de la commission ou des frais sur versement (même si « offerts ») obligatoire."),
            ("P05", "« accompagnement personnalisé »",
             "Doit être lié à un test d'adéquation MIF II / DDA documenté."),
            ("P11", "(absent ou trop discret)",
             "Coordonnées du médiateur de l'assurance + service réclamation interne obligatoires."),
        ],
    },
    {
        "id": "COMM-06",
        "branche": "Auto",
        "canal": "Comparateur",
        "format": "Fiche comparateur",
        "titre": "« Assurance auto à partir de 19€/mois »",
        "contenu": (
            "Assurance auto à partir de 19€/mois. Devis instantané. Garanties au choix. "
            "Note client : 4,8/5 (15 234 avis). [Souscrire en 1 clic]"
        ),
        "verdicts": {
            "P01": "WARN", "P02": "WARN", "P03": "KO", "P04": "KO", "P05": "KO",
            "P06": "KO", "P07": "WARN", "P08": "WARN", "P09": "OK",
            "P10": "KO", "P11": "KO", "P12": "WARN",
        },
        "redlines": [
            ("P03 (DSA)", "(absent)",
             "DSA + DMA : un comparateur doit divulguer (a) sa rémunération par l'assureur, (b) si l'assureur est dans son panel partenaire, (c) l'exhaustivité du panel comparé. Bandeau obligatoire."),
            ("P07", "« à partir de 19€/mois »",
             "Préciser le profil de référence (âge, ville, véhicule, bonus) sur lequel le tarif est calculé. Sans cela = trompeur."),
            ("P05-P06", "« Souscrire en 1 clic »",
             "Pas de souscription possible sans recueil besoins (DDA art. 20) + remise IPID. Le « 1 clic » doit ouvrir un parcours de qualification, pas un contrat signé."),
            ("P08", "« Note client : 4,8/5 »",
             "Préciser source, méthode de collecte et période. Avis modérés ? Vérifiés ? (Loi 2023 sur les avis en ligne)."),
            ("P12", "(absent)",
             "Si recommandation algorithmique, mention IA obligatoire (août 2026)."),
        ],
    },
]


# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────
def build_readme(ws) -> None:
    write_title(ws, "DDA_MarketingGuard — Démo NEURAL · Assurances/Marketing")
    write_subtitle(
        ws,
        2,
        "Audit de conformité des communications marketing assurance contre la "
        "Directive Distribution Assurance (DDA) + Acte délégué clarté contractuelle "
        "+ ACPR + EU AI Act art. 50.",
    )

    write_section_header(ws, 4, "Mission")
    ws["A5"] = (
        "Vérifier que chaque communication marketing (email, landing page, post réseau "
        "social, mailing, brochure, fiche comparateur) respecte les 12 points obligatoires "
        "de la DDA et de la réglementation associée — avant diffusion. Génère un verdict "
        "par règle, des suggestions de correction (redlines), et un rapport auditable."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A5"].font = body_font()
    ws.merge_cells("A5:F5")
    ws.row_dimensions[5].height = 70

    write_section_header(ws, 7, "Référentiel d'audit (12 points)")
    ws["A8"] = (
        "DDA (Directive UE 2016/97 transposée Code des assurances L.521 et suivants) — "
        "Acte délégué clarté contractuelle (UE) 2025 — ACPR Recommandation 2024-R-01 — "
        "ACPR communiqué greenwashing mars 2026 — EU AI Act (UE) 2024/1689 art. 50 "
        "(applicable août 2026) — DGCCRF — DSA/DMA pour comparateurs."
    )
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A8"].font = body_font()
    ws.merge_cells("A8:F8")
    ws.row_dimensions[8].height = 60

    write_section_header(ws, 10, "Mode d'emploi des onglets")
    onglets = [
        ("0_README", "Mission, référentiel, mode d'emploi (cet onglet)"),
        ("1_Checklist_DDA", "Les 12 points DDA — règle, axe, référence légale"),
        ("2_Cas_Communications", "6 communications marketing à auditer (auto/MRH/santé/prévoyance/vie/comparateur)"),
        ("3_Verdict_Par_Cas", "Matrice 12 règles × 6 cas, score conformité calculé par formule"),
        ("4_Redlines", "Suggestions de correction concrètes par cas et par règle"),
        ("5_AI_Act_Disclosure", "Modèles de mention obligatoire EU AI Act art. 50 (août 2026)"),
        ("6_Limites", "Truth layer : ce que l'agent ne fait pas + validations humaines"),
    ]
    row = 11
    for name, desc in onglets:
        ws.cell(row=row, column=1, value=name).font = body_font(bold=True, color=NEURAL_VIOLET)
        ws.cell(row=row, column=2, value=desc).font = body_font()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1

    write_section_header(ws, 19, "Échelle de verdict")
    legend = [
        ("OK", "La règle est satisfaite — diffusion possible", "Vert"),
        ("WARN", "Partiel ou ambigu — relecture compliance avant diffusion", "Ambre"),
        ("KO", "Non conforme — bloquant, redline obligatoire", "Rouge"),
    ]
    for i, (v, desc, color) in enumerate(legend, start=20):
        ws.cell(row=i, column=1, value=v).font = body_font(bold=True)
        ws.cell(row=i, column=2, value=desc).font = body_font()
        ws.cell(row=i, column=3, value=color).font = body_font()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        if v == "OK":
            ws.cell(row=i, column=1).fill = green_fill()
        elif v == "WARN":
            ws.cell(row=i, column=1).fill = amber_fill()
        else:
            ws.cell(row=i, column=1).fill = red_fill()
        ws.row_dimensions[i].height = 22

    write_section_header(ws, 24, "Statut")
    ws["A25"] = "Démo orchestrée · audit indicatif · validation compliance humaine obligatoire avant publication"
    ws["A25"].font = body_font(bold=True, color=AMBER_WARN)
    ws["A25"].fill = yellow_fill()
    ws.merge_cells("A25:F25")
    ws.row_dimensions[25].height = 26

    set_widths(ws, {1: 24, 2: 24, 3: 18, 4: 18, 5: 18, 6: 18})


def build_checklist(ws) -> None:
    write_title(ws, "1 · Checklist DDA — 12 points obligatoires", span=5)
    write_subtitle(
        ws,
        2,
        "Référentiel utilisé par l'agent pour auditer chaque communication marketing. "
        "Chaque point cite la base légale exacte.",
        span=5,
    )

    headers = ["ID", "Axe", "Règle obligatoire", "Référence légale", "Obligatoire"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    for i, item in enumerate(DDA_CHECKLIST):
        row = 5 + i
        ws.cell(row=row, column=1, value=item["id"])
        ws.cell(row=row, column=2, value=item["axe"])
        ws.cell(row=row, column=3, value=item["regle"])
        ws.cell(row=row, column=4, value=item["ref"])
        ws.cell(row=row, column=5, value=item["obligatoire"])
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=44)
        ws.cell(row=row, column=1).font = body_font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    set_widths(ws, {1: 8, 2: 18, 3: 60, 4: 50, 5: 16})


def build_cases(ws) -> None:
    write_title(ws, "2 · Cas — Communications marketing à auditer")
    write_subtitle(
        ws,
        2,
        "6 communications réelles ou inspirées du marché (anonymisées) couvrant "
        "auto, MRH, santé, prévoyance, vie épargne, comparateur.",
    )

    headers = ["ID", "Branche", "Canal", "Format", "Titre / accroche", "Contenu intégral"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    for i, case in enumerate(CASES):
        row = 5 + i
        ws.cell(row=row, column=1, value=case["id"])
        ws.cell(row=row, column=2, value=case["branche"])
        ws.cell(row=row, column=3, value=case["canal"])
        ws.cell(row=row, column=4, value=case["format"])
        ws.cell(row=row, column=5, value=case["titre"])
        ws.cell(row=row, column=6, value=case["contenu"])
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=140)
        ws.cell(row=row, column=1).font = body_font(bold=True)

    set_widths(ws, {1: 12, 2: 14, 3: 18, 4: 18, 5: 38, 6: 70})


def build_verdict(ws) -> None:
    write_title(ws, "3 · Verdict par cas — matrice conformité 12 × 6")
    write_subtitle(
        ws,
        2,
        "Verdict OK / WARN / KO pour chaque règle DDA × cas. Score de conformité "
        "calculé par formule (OK = 1, WARN = 0.5, KO = 0).",
    )

    # Header : 12 règles × 6 cas
    case_ids = [c["id"] for c in CASES]
    headers = ["Règle DDA"] + case_ids + ["Total OK", "Total WARN", "Total KO"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    rule_ids = [r["id"] for r in DDA_CHECKLIST]

    # Matrice : 12 lignes (règles) × 6 colonnes (cas)
    for i, rule in enumerate(DDA_CHECKLIST):
        row = 5 + i
        ws.cell(row=row, column=1, value=f"{rule['id']} · {rule['axe']}")
        ws.cell(row=row, column=1).font = body_font(bold=True)
        for j, case in enumerate(CASES):
            verdict = case["verdicts"][rule["id"]]
            cell = ws.cell(row=row, column=2 + j, value=verdict)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = body_font(bold=True, color=("16A34A" if verdict == "OK" else "D97706" if verdict == "WARN" else "DC2626"))
            if verdict == "OK":
                cell.fill = green_fill()
            elif verdict == "WARN":
                cell.fill = amber_fill()
            else:
                cell.fill = red_fill()
            cell.border = THIN_BORDER

        # Totaux par ligne (par règle, sur tous les cas)
        first_col = "B"
        last_col = get_column_letter(1 + len(CASES))
        ws.cell(row=row, column=2 + len(CASES), value=f'=COUNTIF({first_col}{row}:{last_col}{row},"OK")')
        ws.cell(row=row, column=3 + len(CASES), value=f'=COUNTIF({first_col}{row}:{last_col}{row},"WARN")')
        ws.cell(row=row, column=4 + len(CASES), value=f'=COUNTIF({first_col}{row}:{last_col}{row},"KO")')
        for c_idx in (2 + len(CASES), 3 + len(CASES), 4 + len(CASES)):
            cell = ws.cell(row=row, column=c_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = body_font(bold=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 24

    # Lignes agrégat par cas (en bas)
    n = len(DDA_CHECKLIST)
    score_row = 5 + n + 1
    ws.cell(row=score_row, column=1, value="Score conformité (%)").font = body_font(bold=True, color="FFFFFF")
    ws.cell(row=score_row, column=1).fill = title_fill()
    ws.cell(row=score_row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for j, case in enumerate(CASES):
        col_letter = get_column_letter(2 + j)
        # Score = (OK + 0.5*WARN) / 12
        ws.cell(
            row=score_row,
            column=2 + j,
            value=(
                f'=(COUNTIF({col_letter}5:{col_letter}{4+n},"OK")'
                f'+0.5*COUNTIF({col_letter}5:{col_letter}{4+n},"WARN"))'
                f'/{n}'
            ),
        )
        cell = ws.cell(row=score_row, column=2 + j)
        cell.number_format = "0.0%"
        cell.font = body_font(bold=True, color="FFFFFF")
        cell.fill = title_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # KO count par cas
    ko_row = score_row + 1
    ws.cell(row=ko_row, column=1, value="Nombre de KO bloquants").font = body_font(bold=True)
    for j, case in enumerate(CASES):
        col_letter = get_column_letter(2 + j)
        ws.cell(row=ko_row, column=2 + j, value=f'=COUNTIF({col_letter}5:{col_letter}{4+n},"KO")')
        cell = ws.cell(row=ko_row, column=2 + j)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = body_font(bold=True, color=RED_NO)
        cell.fill = red_fill()
        cell.border = THIN_BORDER

    # Décision diffusion
    decision_row = score_row + 2
    ws.cell(row=decision_row, column=1, value="Décision diffusion").font = body_font(bold=True)
    for j, case in enumerate(CASES):
        col_letter = get_column_letter(2 + j)
        # Décision : si KO > 0 → BLOQUÉ ; si score > 80% → OK ; sinon RELECTURE
        ws.cell(
            row=decision_row,
            column=2 + j,
            value=(
                f'=IF(COUNTIF({col_letter}5:{col_letter}{4+n},"KO")>0,'
                f'"BLOQUÉ",'
                f'IF((COUNTIF({col_letter}5:{col_letter}{4+n},"OK")'
                f'+0.5*COUNTIF({col_letter}5:{col_letter}{4+n},"WARN"))/{n}>0.8,'
                f'"DIFFUSION OK","RELECTURE"))'
            ),
        )
        cell = ws.cell(row=decision_row, column=2 + j)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = body_font(bold=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[score_row].height = 28
    ws.row_dimensions[ko_row].height = 24
    ws.row_dimensions[decision_row].height = 28

    # Largeurs colonnes
    widths = {1: 26}
    for j in range(len(CASES)):
        widths[2 + j] = 14
    widths[2 + len(CASES)] = 11
    widths[3 + len(CASES)] = 11
    widths[4 + len(CASES)] = 11
    set_widths(ws, widths)


def build_redlines(ws) -> None:
    write_title(ws, "4 · Redlines — suggestions de correction")
    write_subtitle(
        ws,
        2,
        "Pour chaque cas et règle violée, suggestion concrète de reformulation. "
        "Les redlines sont indicatives — validation compliance + juridique obligatoire.",
    )

    headers = ["ID cas", "Branche", "Règle DDA", "Texte original", "Reformulation suggérée"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    for case_idx, case in enumerate(CASES):
        for rule_id, orig, refor in case["redlines"]:
            ws.cell(row=row, column=1, value=case["id"])
            ws.cell(row=row, column=2, value=case["branche"])
            ws.cell(row=row, column=3, value=rule_id)
            ws.cell(row=row, column=4, value=orig)
            ws.cell(row=row, column=5, value=refor)
            style_body_row(ws, row, len(headers), alt=(case_idx % 2 == 1), height=80)
            ws.cell(row=row, column=1).font = body_font(bold=True)
            ws.cell(row=row, column=3).font = body_font(bold=True, color=NEURAL_VIOLET)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="top")
            row += 1

    set_widths(ws, {1: 12, 2: 14, 3: 14, 4: 55, 5: 60})


def build_ai_disclosure(ws) -> None:
    write_title(ws, "5 · AI Act Disclosure — modèles de mention IA")
    write_subtitle(
        ws,
        2,
        "EU AI Act art. 50 (applicable août 2026) — toute communication générée ou "
        "assistée par IA destinée aux assurés doit porter une mention explicite. "
        "Modèles ci-dessous, par canal et par niveau d'assistance.",
    )

    write_section_header(ws, 4, "Cadre légal")
    ws["A5"] = (
        "Règlement (UE) 2024/1689 « EU AI Act », article 50 « Obligations de transparence "
        "pour les fournisseurs et déployeurs de certains systèmes d'IA » — applicable à "
        "compter du 2 août 2026. Sanctions : jusqu'à 15 M€ ou 3% du CA mondial annuel "
        "(article 99 §4)."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A5"].font = body_font()
    ws.merge_cells("A5:F5")
    ws.row_dimensions[5].height = 70

    headers = ["Canal", "Niveau d'assistance IA", "Mention obligatoire (FR)", "Mention obligatoire (EN)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=7, column=col, value=h)
    style_header_row(ws, 7, len(headers))

    disclosures = [
        (
            "Email B2C",
            "Rédaction assistée IA, validation humaine",
            "Cet email a été rédigé avec une assistance IA et validé par notre équipe. "
            "Pour tout conseil personnalisé, contactez votre conseiller au [numéro].",
            "This email was drafted with AI assistance and reviewed by our team. "
            "For personalized advice, contact your advisor at [number].",
        ),
        (
            "Landing page",
            "Recommandation algorithmique",
            "Les recommandations affichées sur cette page sont produites par un système "
            "d'intelligence artificielle, en complément du conseil humain.",
            "Recommendations on this page are produced by an artificial intelligence system, "
            "as a complement to human advice.",
        ),
        (
            "Post réseau social",
            "Contenu visuel ou textuel généré IA",
            "Visuel / texte généré avec assistance IA. Validé par notre direction de la "
            "communication.",
            "Visual / text generated with AI assistance. Reviewed by our communications team.",
        ),
        (
            "Mailing courrier",
            "Personnalisation IA (segmentation)",
            "Cette communication a été personnalisée à l'aide d'un système d'IA "
            "(segmentation marketing). Aucune décision contractuelle automatique.",
            "This communication has been personalized using an AI system (marketing segmentation). "
            "No automated contractual decisions.",
        ),
        (
            "Brochure print",
            "Aucune assistance IA",
            "(Aucune mention requise si aucune génération IA — sauf si décision "
            "automatisée par ailleurs sur le parcours.)",
            "(No mention required if no AI generation — unless automated decision elsewhere "
            "in the journey.)",
        ),
        (
            "Comparateur / fiche produit",
            "Tri & ranking algorithmique",
            "Le classement des offres présentées est produit par un algorithme. Critères "
            "de tri : [prix / popularité / partenariat]. Détails : [lien méthodologie].",
            "The ranking of offers shown is produced by an algorithm. Sort criteria: "
            "[price / popularity / partnership]. Details: [methodology link].",
        ),
        (
            "Chatbot / assistant virtuel",
            "Interaction conversationnelle IA",
            "Vous échangez avec un assistant virtuel propulsé par une IA. Pour parler à "
            "un conseiller humain, tapez « conseiller ».",
            "You are chatting with a virtual assistant powered by AI. To speak with a human "
            "advisor, type « advisor ».",
        ),
        (
            "Réclamation / lettre",
            "Décision automatisée (refus, indemnisation)",
            "Cette décision a été prise avec l'aide d'un système d'IA, sous supervision "
            "humaine. Vous pouvez demander un réexamen humain en répondant à ce courrier.",
            "This decision was made with the help of an AI system, under human supervision. "
            "You may request a human review by replying to this letter.",
        ),
    ]

    for i, (canal, niveau, fr, en) in enumerate(disclosures):
        row = 8 + i
        ws.cell(row=row, column=1, value=canal)
        ws.cell(row=row, column=2, value=niveau)
        ws.cell(row=row, column=3, value=fr)
        ws.cell(row=row, column=4, value=en)
        style_body_row(ws, row, 4, alt=(i % 2 == 1), height=70)
        ws.cell(row=row, column=1).font = body_font(bold=True)

    set_widths(ws, {1: 22, 2: 30, 3: 60, 4: 60})


def build_limites(ws) -> None:
    write_title(ws, "6 · Limites — truth layer")
    write_subtitle(
        ws,
        2,
        "Ce que l'agent DDA_MarketingGuard ne fait pas (et ne prétend pas faire). "
        "À diffuser aux équipes Marketing + Compliance avant tout pilote.",
    )

    sections = [
        (
            "Ce que l'agent ne remplace pas",
            [
                "Un compliance officer / juriste assurance : le verdict est indicatif.",
                "Un Direction Conformité interne : la décision finale de diffusion est humaine.",
                "L'avocat spécialisé : pas d'analyse juridique personnalisée d'un litige.",
                "Le contrôle ACPR : audit interne ≠ contrôle prudentiel.",
            ],
        ),
        (
            "Ce que l'agent ne génère pas",
            [
                "De nouveaux contrats ou clauses contractuelles.",
                "Des décisions de souscription, refus, indemnisation.",
                "Une analyse personnalisée du devoir de conseil pour un client donné.",
                "Des verdicts sur des sujets hors DDA (ex : RGPD avancé, droit de la consommation général).",
            ],
        ),
        (
            "Validations humaines obligatoires",
            [
                "Compliance : revue de chaque verdict avant publication / diffusion.",
                "Direction Marketing : validation des reformulations vs. brand voice.",
                "Direction Juridique : validation des redlines à enjeu (claims, ESG, vie).",
                "Direction Communication : validation du ton et du wording final.",
            ],
        ),
        (
            "Couverture réglementaire — limites de scope",
            [
                "Cible v1 : DDA + Acte délégué clarté + ACPR + EU AI Act art. 50.",
                "Hors scope v1 : RGPD avancé (DPIA), DDA pour PRIIPs, MIF II détaillé.",
                "Hors scope v1 : droit comparé multi-juridictionnel (UK, Suisse, etc.).",
                "Hors scope v1 : analyse iconographique / visuelle (audit textuel uniquement).",
            ],
        ),
        (
            "Limites techniques v1 (à industrialiser v2)",
            [
                "Verdicts encodés manuellement v1 — automatiser via prompts Claude + RAG checklist.",
                "Pas de mise à jour automatique du référentiel sur évolution ACPR / EIOPA.",
                "Pas de versioning fin entre versions de la communication marketing.",
                "Détection greenwashing limitée aux mots-clés évidents (renforcer via taxonomie).",
            ],
        ),
        (
            "Disclaimers obligatoires sur les outputs",
            [
                "« Audit indicatif — ne se substitue pas à une revue compliance humaine. »",
                "« Verdicts générés avec assistance IA — relecture humaine obligatoire (EU AI Act). »",
                "« Référentiel à jour au [date] — vérifier les évolutions réglementaires depuis. »",
                "Conservation 5 ans minimum des audits + décisions humaines (traçabilité ACPR).",
            ],
        ),
    ]

    row = 4
    for title, items in sections:
        write_section_header(ws, row, title)
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=f"• {item}").font = body_font()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.row_dimensions[row].height = 32
            row += 1
        row += 1

    set_widths(ws, {1: 24, 2: 24, 3: 18, 4: 18, 5: 18, 6: 18})


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_readme = wb.create_sheet("0_README")
    build_readme(ws_readme)

    ws_checklist = wb.create_sheet("1_Checklist_DDA")
    build_checklist(ws_checklist)

    ws_cases = wb.create_sheet("2_Cas_Communications")
    build_cases(ws_cases)

    ws_verdict = wb.create_sheet("3_Verdict_Par_Cas")
    build_verdict(ws_verdict)

    ws_redlines = wb.create_sheet("4_Redlines")
    build_redlines(ws_redlines)

    ws_disclosure = wb.create_sheet("5_AI_Act_Disclosure")
    build_ai_disclosure(ws_disclosure)

    ws_limites = wb.create_sheet("6_Limites")
    build_limites(ws_limites)

    output_path = "C:/Users/Ludo/finance-platform/DDA_MarketingGuard_NEURAL.xlsx"
    wb.save(output_path)
    print(f"WROTE {output_path}")


if __name__ == "__main__":
    main()
