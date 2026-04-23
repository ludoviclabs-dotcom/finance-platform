# ═══════════════════════════════════════════════════════════════════════════════
# AGENT LUXETRACEABILITY — NEURAL ENTERPRISE FRAMEWORK
# Supply Chain Luxe — Onglets 5 à 11
# Version 1.0 — Avril 2026
# ═══════════════════════════════════════════════════════════════════════════════

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────
# CONSTANTES — CHARTE GRAPHIQUE LUXE
# ─────────────────────────────────────────────────────────────────────

NOIR_LUXE        = "1A1A1A"
OR_LUXE          = "C9A84C"
BORDEAUX         = "6B1D2A"
BLANC_CASSE      = "F9F6F0"
GRIS_LUXE        = "8C8C8C"
GRIS_FONCE       = "444444"
VERT_VALIDATION  = "2E7D32"
ROUGE_ALERTE     = "C62828"
ORANGE_ATTENTION = "E65100"
BLEU_INFO        = "1565C0"

FOND_CRITIQUE    = "FFC7CE"
FOND_MOYEN       = "FFEB9C"
FOND_FAIBLE      = "C6EFCE"
FOND_TRES_FAIBLE = "A8D5A2"

BORDURE_FINE = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────
# FONCTIONS UTILITAIRES
# ─────────────────────────────────────────────────────────────────────

def creer_titre_principal(ws, texte, sous_texte, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    t = ws.cell(row=1, column=1, value=texte)
    t.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    t.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    s = ws.cell(row=2, column=1, value=sous_texte)
    s.font = Font(name="Calibri", size=10, italic=True, color=OR_LUXE)
    s.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28


def appliquer_style_titre_section(ws, row, max_col, texte, couleur=BORDEAUX):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=texte)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDURE_FINE
    ws.row_dimensions[row].height = 30


def appliquer_style_en_tetes(ws, row, headers, largeurs=None):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDURE_FINE
    ws.row_dimensions[row].height = 36
    if largeurs:
        for i, w in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def appliquer_style_donnees(ws, start_row, end_row, max_col):
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.font or cell.font.name is None:
                cell.font = Font(name="Calibri", size=9)
            cell.border = BORDURE_FINE
            if (row_idx - start_row) % 2 == 0:
                cur = cell.fill
                if not cur or cur.fill_type == "none" or (
                    hasattr(cur, 'start_color') and cur.start_color.index in ("00000000", "FFFFFFFF", None)
                ):
                    cell.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")


def ajouter_mise_en_forme_conditionnelle_statut(ws, col_letter, start_row, end_row):
    plage = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for val, fond, texte in [
        ("Conforme",      FOND_FAIBLE,   "006100"),
        ("Validé",        FOND_FAIBLE,   "006100"),
        ("Actif",         FOND_FAIBLE,   "006100"),
        ("Certifié",      FOND_FAIBLE,   "006100"),
        ("Complet",       FOND_FAIBLE,   "006100"),
        ("En attente",    FOND_MOYEN,    "9C6500"),
        ("En cours",      FOND_MOYEN,    "9C6500"),
        ("Partiel",       FOND_MOYEN,    "9C6500"),
        ("Non conforme",  FOND_CRITIQUE, "9C0006"),
        ("Rejeté",        FOND_CRITIQUE, "9C0006"),
        ("Suspendu",      FOND_CRITIQUE, "9C0006"),
        ("Expiré",        FOND_CRITIQUE, "9C0006"),
        ("Incomplet",     FOND_CRITIQUE, "9C0006"),
    ]:
        ws.conditional_formatting.add(plage,
            CellIsRule(
                operator='equal', formula=[f'"{val}"'],
                fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                font=Font(color=texte, bold=(texte != "9C6500"))
            ))


def ligne_synthese(ws, row, nb_col, items):
    ws.cell(row=row, column=1).border = BORDURE_FINE
    ws.cell(row=row, column=1).fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    for c in range(2, nb_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        ws.cell(row=row, column=c).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws.cell(row=row, column=c).border = BORDURE_FINE
    for (col, val, fmt, color) in items:
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color=color or OR_LUXE)
        cell.border = BORDURE_FINE
        if fmt:
            cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION — ONGLETS 5 À 11
# ═══════════════════════════════════════════════════════════════════════════════

def ajouter_onglets_5_11():
    input_path = "C:/Users/Ludo/Desktop/IA projet entreprises/NEURAL_LuxeTraceability.xlsx"
    wb = load_workbook(input_path)

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 5 : CONFORMITE
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("5_CONFORMITE")
    ws.sheet_properties.tabColor = BLEU_INFO

    creer_titre_principal(
        ws,
        "CONFORMITE REGLEMENTAIRE — AGENT LUXETRACEABILITY",
        "Devoir de Vigilance — CSRD 2024/2862 — CSDDD 2024/1760 — EUDR 2023/1115 — Points de controle — Avril 2026",
        14
    )

    appliquer_style_titre_section(ws, 3, 14, "SECTION A — MATRICE DE CONFORMITE PAR EXIGENCE REGLEMENTAIRE", BORDEAUX)

    conf_headers = [
        "ID_Controle", "Reglementation", "Article / Exigence",
        "Description de l'obligation", "Perimetre (matieres)",
        "Frequence controle", "Statut conformite",
        "Date dernier controle", "Prochain controle",
        "Responsable", "Preuves / Documents",
        "Score conformite (%)", "Risque si non-conforme",
        "Actions correctives"
    ]
    conf_largeurs = [14, 22, 22, 40, 22, 16, 16, 16, 16, 20, 24, 16, 30, 36]
    appliquer_style_en_tetes(ws, 4, conf_headers, conf_largeurs)

    points_conformite = [
        ["CONF-001", "Loi Vigilance (2017-399)", "Art. L225-102-4 — Plan de vigilance",
         "Etablir et publier un plan de vigilance incluant cartographie des risques, procedures d'evaluation des filiales et fournisseurs, actions d'attenuation, mecanisme d'alerte, dispositif de suivi",
         "Toutes matieres", "Annuel", "Conforme",
         "2026-01-15", "2027-01-15", "Dir. Juridique / Dir. RSE",
         "Plan Vigilance 2026 publie — Rapport annuel p.45-72", 95,
         "Amende jusqu'a 30M EUR + responsabilite civile — Risque reputationnel majeur",
         "Mise a jour cartographie risques en cours — Integration CSDDD"],

        ["CONF-002", "CSRD (2024/2862)", "ESRS S2 — Workers in the value chain",
         "Reporting sur les conditions de travail dans la chaine de valeur — Due diligence — Double materialite — Indicateurs quantitatifs obligatoires",
         "Toutes matieres (focus cuir, textile)", "Annuel", "Conforme",
         "2026-03-01", "2027-03-01", "Dir. RSE",
         "Rapport CSRD 2025 depose — Assurance limitee EY", 90,
         "Sanctions AMF — Non-publication rapport durabilite — Penalites financieres",
         "Renforcement indicateurs S2 pour exercice 2026 — Audit Tier 2"],

        ["CONF-003", "CSDDD (2024/1760)", "Art. 7-8 — Due diligence obligations",
         "Identifier et prevenir les incidences negatives reelles et potentielles sur les droits de l'homme et l'environnement dans la chaine de valeur",
         "Toutes matieres — Chaine complete", "Continu", "En cours",
         "2026-04-01", "2026-07-01", "Dir. Juridique",
         "Cartographie en cours — Transposition nationale attendue 07/2026", 70,
         "Penalites jusqu'a 5% du CA mondial — Responsabilite civile — Exclusion marches publics",
         "Transposition francaise en cours — Anticiper exigences — Mapping Tier 1-2-3 en cours"],

        ["CONF-004", "EUDR (2023/1115)", "Art. 3-4 — Tracabilite geolocalisation",
         "Declaration de diligence raisonnee (DDS) pour produits contenant cuir, bois, caoutchouc — Geolocalisation des parcelles de production",
         "Cuir bovins, Bois, Caoutchouc", "Par lot / importation", "Partiel",
         "2026-04-08", "2026-04-30", "Dir. Supply Chain",
         "DDS cuir Italie OK — DDS bois Indonesie en cours — Geolocalisation partielle", 65,
         "Interdiction mise sur marche UE — Saisie douanes — Amendes",
         "EUDR applicable depuis 30/12/2025 — Completer geolocalisation ebene Sulawesi URGENT"],

        ["CONF-005", "CITES", "Convention Washington — Annexes I, II, III",
         "Permis d'importation/exportation pour especes protegees — Python, crocodile, vigogne, corail — Quotas annuels",
         "Python, Crocodile, Vigogne, Corail, Nacre", "Par transaction", "Conforme",
         "2026-04-05", "Continu", "Dir. Achats / Douanes",
         "Tous permis CITES a jour — Registre CITES centralise", 100,
         "Saisie et destruction des marchandises — Poursuites penales — Sanctions UE",
         "RAS — Permis valides — Prochaine echeance permis python : 09/2026"],

        ["CONF-006", "Kimberley Process (KPCS)", "Schema de certification KP",
         "Certification que les diamants bruts ne proviennent pas de zones de conflit — Certificat KP obligatoire par lot",
         "Diamants", "Par lot", "Conforme",
         "2026-02-28", "Continu", "Dir. Achats Joaillerie",
         "Certificats KP De Beers a jour — Alrosa exclu (sanctions)", 100,
         "Interdiction commercialisation — Sanctions penales — Embargo",
         "Fournisseur unique De Beers post-exclusion Russie — Monitoring continu"],

        ["CONF-007", "LBMA Good Delivery", "Rules for Good Delivery",
         "Chaine de garde (Chain of Custody) pour or et platine — Audit annuel raffineurs — Reporting origine",
         "Or, Platine, Argent", "Annuel (par raffineur)", "Conforme",
         "2026-03-01", "2027-03-01", "Dir. Joaillerie",
         "Valcambi + Metalor certifies LBMA GD — Audit 2026 passe", 98,
         "Exclusion liste LBMA — Impossibilite sourcing or responsable — Risque reputationnel",
         "RAS — Renouvellement automatique — Prochain audit Metalor : 01/2027"],

        ["CONF-008", "RJC (Responsible Jewellery Council)", "Code of Practices 2019",
         "Certification des pratiques responsables en joaillerie — Droits humains, environnement, ethique — Audit tierce partie",
         "Or, Diamants, Pierres precieuses, Platine", "Triennal", "Conforme",
         "2025-06-15", "2028-06-15", "Dir. RSE Joaillerie",
         "Certification RJC CoP renouvelee 2025 — Valide 3 ans", 92,
         "Perte certification — Exclusion communaute joaillerie responsable — Risque client B2B",
         "Prochain audit de recertification : 06/2028 — Preparation Q1 2028"],

        ["CONF-009", "REACH (1907/2006)", "Annexe XVII — Restrictions substances",
         "Verification absence de substances restreintes dans cuirs et textiles — Chrome VI, formaldehyde, AZO colorants, PFAS",
         "Cuir, Textiles", "Par lot (analyse labo)", "Conforme",
         "2026-04-01", "2026-07-01", "Dir. Qualite",
         "Analyses SGS trimestrielles — Tous lots conformes 2026", 100,
         "Retrait du marche — Rappel produits — Amendes ECHA",
         "RAS — Prochaine analyse labo : 07/2026 — PFAS nouvelle restriction a anticiper"],

        ["CONF-010", "EU Forced Labour Regulation (2024/3015)", "Art. 3 — Interdiction",
         "Interdiction de mise sur le marche UE de produits issus du travail force — Investigation et preuve de non-recours",
         "Toutes matieres (focus textile Chine)", "Continu", "En cours",
         "2026-04-01", "2027-01-01", "Dir. Juridique / Dir. RSE",
         "Mapping fournisseurs Tier 2-3 en cours — Exclusion Xinjiang effective", 60,
         "Saisie produits — Interdiction marche UE — Amendes — Risque mediatique majeur",
         "Application progressive 2027 — Anticiper — Erdos Group suspendu — Audit Tier 3 cachemire"],

        ["CONF-011", "LWG (Leather Working Group)", "Protocole audit environnemental",
         "Certification environnementale des tanneries — Gestion eau, energie, dechets, produits chimiques — Niveaux Gold/Silver/Bronze",
         "Cuir (toutes categories)", "Biennal", "Conforme",
         "2026-02-15", "2028-02-15", "Dir. Achats Cuir",
         "Tanneria Veneta Gold — Curtidos Benavent Silver — Walpier Gold", 95,
         "Impossibilite certification cuir responsable — Perte clients B2B premium",
         "RAS — Tous fournisseurs Tier 1 cuir certifies LWG — Prochain renouvellement 2028"],

        ["CONF-012", "SA8000 (Social Accountability)", "Norme SA8000:2014",
         "Conditions de travail decentes — Travail enfants, travail force, sante securite, liberte syndicale — Audit social",
         "Cuir exotique, Textile (Inde, Bangladesh)", "Annuel", "Partiel",
         "2025-12-05", "2026-06-05", "Dir. RSE",
         "Louisiana Farm : audit planifie — Tata Leather : en attente", 55,
         "Risque reputationnel — Investigations ONG — Boycott consommateurs",
         "2 fournisseurs sans SA8000 — Audits sociaux programmes Q2-Q3 2026"],
    ]

    for row_idx, conf_line in enumerate(points_conformite, 5):
        for col_idx, val in enumerate(conf_line, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [4, 13, 14] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
            if col_idx == 12 and isinstance(val, (int, float)):
                cell.number_format = '0"%"'
        ws.row_dimensions[row_idx].height = 52

    nb_conf = len(points_conformite)
    appliquer_style_donnees(ws, 5, 4 + nb_conf, 14)

    ajouter_mise_en_forme_conditionnelle_statut(ws, "G", 5, 4 + nb_conf)
    ws.conditional_formatting.add(f'G5:G{4+nb_conf}',
        CellIsRule(operator='equal', formula=['"Partiel"'],
                   fill=PatternFill(start_color=FOND_MOYEN, end_color=FOND_MOYEN, fill_type="solid"),
                   font=Font(color="9C6500")))
    ws.conditional_formatting.add(f'L5:L{4+nb_conf}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color=BLEU_INFO))

    total_conf_row = 4 + nb_conf + 1
    ligne_synthese(ws, total_conf_row, 14, [
        (1,  "SYNTHESE CONFORMITE",                   None,      "FFFFFF"),
        (2,  "Total controles :",                      None,      "FFFFFF"),
        (3,  f'=COUNTA(A5:A{4+nb_conf})',              None,      OR_LUXE),
        (5,  "Conformes :",                            None,      "FFFFFF"),
        (6,  f'=COUNTIF(G5:G{4+nb_conf},"Conforme")', None,      "C6EFCE"),
        (7,  "En cours :",                             None,      "FFFFFF"),
        (8,  f'=COUNTIF(G5:G{4+nb_conf},"En cours")', None,      FOND_MOYEN),
        (9,  "Partiels :",                             None,      "FFFFFF"),
        (10, f'=COUNTIF(G5:G{4+nb_conf},"Partiel")',  None,      FOND_MOYEN),
        (11, "Score moyen :",                          None,      "FFFFFF"),
        (12, f'=IFERROR(AVERAGE(L5:L{4+nb_conf}),0)', '0.0"%"',  OR_LUXE),
    ])
    ws.cell(row=total_conf_row + 1, column=2, value="Taux conformite globale :")
    ws.cell(row=total_conf_row + 1, column=2).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws.cell(row=total_conf_row + 1, column=2).fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")
    ws.cell(row=total_conf_row + 1, column=3).value = f'=IFERROR(COUNTIF(G5:G{4+nb_conf},"Conforme")/COUNTA(G5:G{4+nb_conf}),0)'
    ws.cell(row=total_conf_row + 1, column=3).number_format = '0.0%'
    ws.cell(row=total_conf_row + 1, column=3).font = Font(name="Calibri", size=12, bold=True, color=OR_LUXE)
    ws.cell(row=total_conf_row + 1, column=3).fill = PatternFill(start_color=NOIR_LUXE, end_color=NOIR_LUXE, fill_type="solid")

    dv_statut_conf = DataValidation(type="list",
        formula1='"Conforme,En cours,Partiel,Non conforme,N/A"', allow_blank=False)
    ws.add_data_validation(dv_statut_conf)
    dv_statut_conf.add('G5:G5000')
    dv_freq = DataValidation(type="list",
        formula1='"Continu,Par lot,Par transaction,Mensuel,Trimestriel,Semestriel,Annuel,Biennal,Triennal"',
        allow_blank=False)
    ws.add_data_validation(dv_freq)
    dv_freq.add('F5:F5000')

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{4+nb_conf}"
    print("  ✅ Onglet 5_CONFORMITE cree")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 6 : CERTIFICATIONS
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("6_CERTIFICATIONS")
    ws.sheet_properties.tabColor = VERT_VALIDATION

    creer_titre_principal(
        ws,
        "SUIVI DES CERTIFICATIONS — AGENT LUXETRACEABILITY",
        "Registre certifications par matiere et fournisseur — Alertes expiration — Avril 2026",
        16
    )
    appliquer_style_titre_section(ws, 3, 16, "SECTION A — REGISTRE DES CERTIFICATIONS ACTIVES", BORDEAUX)

    cert_headers = [
        "ID_Certification", "Type de certification", "Organisme emetteur",
        "N Certificat", "Fournisseur concerne", "Matiere(s) couverte(s)",
        "Date emission", "Date expiration", "Statut",
        "Jours restants", "Niveau / Grade",
        "Perimetre audite", "Document attache",
        "Dernier audit", "Prochain renouvellement",
        "Observations"
    ]
    cert_largeurs = [16, 24, 20, 20, 26, 22, 14, 14, 12, 14, 14, 24, 22, 14, 16, 34]
    appliquer_style_en_tetes(ws, 4, cert_headers, cert_largeurs)

    certifications_data = [
        ["CRT-2026-001", "LWG — Leather Working Group", "LWG",
         "LWG-IT-2026-Gold-4421", "Tanneria Veneta SpA", "Cuir de veau",
         "2025-02-15", "2027-02-15", "Actif", None, "Gold",
         "Site Arzignano — Processus complet", "CRT/LWG/TVS_2026.pdf",
         "2025-02-15", "2027-02-15", "Certification Gold renouvelee — Score audit 94/100"],

        ["CRT-2026-002", "LWG — Leather Working Group", "LWG",
         "LWG-IT-2026-Gold-3892", "Conceria Walpier", "Cuir de veau tannage vegetal",
         "2025-06-20", "2027-06-20", "Actif", None, "Gold",
         "Site Santa Croce — Tannage vegetal", "CRT/LWG/CW_2026.pdf",
         "2025-06-20", "2027-06-20", "Specialite Pelle Conciata al Vegetale — Double certification"],

        ["CRT-2026-003", "LWG — Leather Working Group", "LWG",
         "LWG-ES-2025-Silver-0892", "Curtidos Benavent", "Cuir d'agneau",
         "2024-11-10", "2026-11-10", "Actif", None, "Silver",
         "Site Igualada — Processus agneau", "CRT/LWG/CB_2025.pdf",
         "2024-11-10", "2026-11-10", "Upgrade Gold en cours — Audit prevu Q3 2026"],

        ["CRT-2026-004", "CITES — Annexe II", "CITES / KLHK Indonesie",
         "CITES-MY-2026-PY-00567", "Exotic Skins Malaysia", "Python reticulé",
         "2026-01-01", "2026-12-31", "Actif", None, "Annexe II",
         "Ferme elevage Java — Quota annuel", "CRT/CITES/ESM_2026.pdf",
         "2026-01-01", "2026-12-31", "Quota 2026 : 500 peaux — 312 utilisees au 10/04/2026"],

        ["CRT-2026-005", "CITES — Annexe II", "CITES / USFWS",
         "CITES-US-2026-CR-01456", "Louisiana Alligator Farm", "Crocodile / Alligator",
         "2026-01-01", "2026-12-31", "Actif", None, "Annexe II",
         "Ferme Lafayette — Elevage controle", "CRT/CITES/LAF_2026.pdf",
         "2026-01-01", "2026-12-31", "Quota 2026 : 200 peaux — 87 utilisees au 10/04/2026"],

        ["CRT-2026-006", "LBMA Good Delivery", "LBMA",
         "LBMA-COC-2026-VAL-03421", "Valcambi SA", "Or 18K / 24K",
         "2025-03-01", "2027-03-01", "Actif", None, "Good Delivery + CoC",
         "Site Balerna — Raffinage complet", "CRT/LBMA/VAL_2026.pdf",
         "2025-03-01", "2027-03-01", "Chain of Custody complete — Mine-raffinerie-atelier documentee"],

        ["CRT-2026-007", "LBMA Good Delivery", "LBMA",
         "LBMA-GD-2026-MET-00891", "Metalor Technologies", "Or 24K",
         "2025-01-15", "2027-01-15", "Actif", None, "Good Delivery",
         "Site Marin-Epagnier — Raffinage", "CRT/LBMA/MET_2026.pdf",
         "2025-01-15", "2027-01-15", "Certificat Good Delivery — Sans Chain of Custody (backup)"],

        ["CRT-2026-008", "Kimberley Process (KPCS)", "KP Secretariat / Botswana",
         "KP-BW-2026-08934", "De Beers Group", "Diamants bruts",
         "2026-01-01", "2026-12-31", "Actif", None, "Certificat KP",
         "Mines Jwaneng + Orapa — Botswana", "CRT/KP/DB_2026.pdf",
         "2026-01-01", "2026-12-31", "Fournisseur unique diamants post-sanctions Russie — KP renouvele annuellement"],

        ["CRT-2026-009", "RJC — Code of Practices", "RJC",
         "RJC-CoP-2025-FR-00234", "Maison (interne)", "Or, Diamants, Pierres, Platine",
         "2025-06-15", "2028-06-15", "Actif", None, "COP Certified",
         "Ensemble de la chaine joaillerie", "CRT/RJC/MAISON_2025.pdf",
         "2025-06-15", "2028-06-15", "Certification RJC CoP de la Maison — Valide 3 ans — Prochain audit 2028"],

        ["CRT-2026-010", "GOTS — Global Organic Textile", "GOTS / Control Union",
         "GOTS-MN-2025-GOBI-0234", "Gobi Cashmere LLC", "Cachemire Grade A",
         "2025-08-30", "2026-08-30", "Actif", None, "Scope Certificate",
         "Filature + Triage — Oulan-Bator", "CRT/GOTS/GC_2025.pdf",
         "2025-08-30", "2026-08-30", "Renouvellement annuel — Audit prevu 07/2026"],

        ["CRT-2026-011", "OEKO-TEX Standard 100", "OEKO-TEX / Hohenstein",
         "OEKO-IT-2026-SB-00123", "Seteria Bianchi", "Soie mulberry",
         "2026-01-10", "2027-01-10", "Actif", None, "Classe I (contact peau)",
         "Tissage + Teinture — Come", "CRT/OEKO/SB_2026.pdf",
         "2026-01-10", "2027-01-10", "Classe I — Plus haut niveau de securite — Contact peau nourrisson valide"],

        ["CRT-2026-012", "RMI — Responsible Minerals", "RMI / Assent",
         "RMI-ZA-2025-AAP-00789", "Anglo American Platinum", "Platine 950",
         "2025-12-20", "2026-12-20", "Actif", None, "Conformant",
         "Mine Mogalakwena — Extraction + Raffinage", "CRT/RMI/AAP_2025.pdf",
         "2025-12-20", "2026-12-20", "Minerai responsable — Pas de conflit — Audit annuel RMI site minier"],

        ["CRT-2026-013", "FSC — Forest Stewardship Council", "FSC / SGS",
         "FSC-ID-2025-ITC-00234", "Indonesian Timber Corp", "Ebene de Macassar",
         "2025-05-15", "2026-05-15", "Actif", None, "FSC Mix",
         "Concession forestiere Sulawesi", "CRT/FSC/ITC_2025.pdf",
         "2025-05-15", "2026-05-15", "Renouvellement dans 35j — EUDR exige geolocalisation complementaire"],

        ["CRT-2026-014", "SA8000 — Social Accountability", "SAI / SGS",
         "SA8000-IT-2024-TV-00891", "Tanneria Veneta SpA", "Cuir de veau",
         "2024-07-01", "2027-07-01", "Actif", None, "SA8000:2014",
         "Site Arzignano — 142 employes", "CRT/SA8K/TV_2024.pdf",
         "2024-07-01", "2027-07-01", "Derniere non-conformite mineure resolue — Heures supplementaires"],

        ["CRT-2026-015", "Label Perle de Tahiti", "GIE Poe Rava Nui",
         "LABEL-PF-2026-RW-00456", "Robert Wan Pearl", "Perles de Tahiti",
         "2026-01-01", "2026-12-31", "Actif", None, "Label Officiel",
         "Ferme Fakarava — Perles noires naturelles", "CRT/LABEL/RW_2026.pdf",
         "2026-01-01", "2026-12-31", "Label officiel Polynesie francaise — Garantie origine et qualite"],

        ["CRT-2026-016", "Kimberley Process (KPCS)", "KP Secretariat / Russie",
         "KP-RU-2022-ALROSA-REVOQUE", "Alrosa PJSC", "Diamants bruts",
         "2022-01-01", "2022-12-31", "Expiré", None, "REVOQUE",
         "N/A — Sanctions", "N/A",
         "2022-02-24", "N/A", "REVOQUE — Sanctions UE — Aucun approvisionnement autorise"],
    ]

    for row_idx, cert_line in enumerate(certifications_data, 5):
        for col_idx, val in enumerate(cert_line, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDURE_FINE
        ws.row_dimensions[row_idx].height = 40

    nb_cert = len(certifications_data)
    appliquer_style_donnees(ws, 5, 4 + nb_cert, 16)

    # Jours restants (col J)
    for r in range(5, 5 + nb_cert):
        ws.cell(row=r, column=10).value = f'=IFERROR(IF(I{r}="Expire",0,IF(H{r}="","",MAX(0,H{r}-TODAY()))),"-")'
        ws.cell(row=r, column=10).number_format = '0'
        ws.cell(row=r, column=10).font = Font(name="Calibri", size=9, bold=True)

    ajouter_mise_en_forme_conditionnelle_statut(ws, "I", 5, 4 + nb_cert)
    ws.conditional_formatting.add(f'J5:J{4+nb_cert}',
        CellIsRule(operator='lessThanOrEqual', formula=['30'],
                   fill=PatternFill(start_color=FOND_CRITIQUE, end_color=FOND_CRITIQUE, fill_type="solid"),
                   font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(f'J5:J{4+nb_cert}',
        CellIsRule(operator='between', formula=['31', '90'],
                   fill=PatternFill(start_color=FOND_MOYEN, end_color=FOND_MOYEN, fill_type="solid"),
                   font=Font(color="9C6500", bold=True)))
    ws.conditional_formatting.add(f'J5:J{4+nb_cert}',
        CellIsRule(operator='greaterThan', formula=['90'],
                   fill=PatternFill(start_color=FOND_FAIBLE, end_color=FOND_FAIBLE, fill_type="solid"),
                   font=Font(color="006100")))

    total_cert_row = 4 + nb_cert + 1
    ligne_synthese(ws, total_cert_row, 16, [
        (1,  "SYNTHESE CERTIFICATIONS",                   None,   "FFFFFF"),
        (2,  "Total certifications :",                    None,   "FFFFFF"),
        (3,  f'=COUNTA(A5:A{4+nb_cert})',                 None,   OR_LUXE),
        (5,  "Actives :",                                 None,   "FFFFFF"),
        (6,  f'=COUNTIF(I5:I{4+nb_cert},"Actif")',        None,   "C6EFCE"),
        (7,  "Expirees :",                                None,   "FFFFFF"),
        (8,  f'=COUNTIF(I5:I{4+nb_cert},"Expire")',       None,   FOND_CRITIQUE),
        (9,  "Taux actif :",                              None,   "FFFFFF"),
        (10, f'=IFERROR(COUNTIF(I5:I{4+nb_cert},"Actif")/COUNTA(I5:I{4+nb_cert}),0)',
                                                          '0.0%', OR_LUXE),
        (11, "Expire <30j :",                             None,   "FFFFFF"),
        (12, f'=COUNTIFS(J5:J{4+nb_cert},"<="&30,J5:J{4+nb_cert},">"&0)',
                                                          None,   FOND_CRITIQUE),
    ])

    dv_statut_cert = DataValidation(type="list",
        formula1='"Actif,Expire,Suspendu,En renouvellement,Revoque"', allow_blank=False)
    ws.add_data_validation(dv_statut_cert)
    dv_statut_cert.add('I5:I5000')
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:P{4+nb_cert}"
    print("  ✅ Onglet 6_CERTIFICATIONS cree")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 7 : AUDIT_JOURNAL
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("7_AUDIT")
    ws.sheet_properties.tabColor = GRIS_FONCE

    creer_titre_principal(
        ws,
        "JOURNAL D'AUDIT — AGENT LUXETRACEABILITY",
        "Historique complet des actions, verifications et modifications — Piste d'audit reglementaire — Avril 2026",
        10
    )
    appliquer_style_titre_section(ws, 3, 10, "REGISTRE D'AUDIT — TRACABILITE DES ACTIONS", BORDEAUX)

    audit_headers = [
        "ID_Action", "Date", "Heure", "Utilisateur",
        "Type d'action", "Objet concerne",
        "Description de l'action", "Statut validation",
        "Niveau completude", "Observations"
    ]
    audit_largeurs = [14, 14, 10, 20, 22, 24, 44, 16, 16, 36]
    appliquer_style_en_tetes(ws, 4, audit_headers, audit_largeurs)

    logs = [
        ["AUD-2026-0001", "2026-01-15", "09:00", "Dir. Juridique",
         "Publication", "Plan de Vigilance 2026",
         "Publication du plan de vigilance annuel — Cartographie des risques mise a jour — Parties prenantes consultees",
         "Valide", "Complet", "Conforme Art. L225-102-4 — Publie rapport annuel p.45-72"],

        ["AUD-2026-0002", "2026-01-20", "14:30", "Resp. Tracabilite",
         "Enregistrement fournisseur", "Conceria Walpier (FRN-2026-0002)",
         "Ajout fournisseur backup cuir tannage vegetal — Certification LWG Gold verifiee — Dossier complet",
         "Valide", "Complet", "Fournisseur Tier 1 valide — Score risque 15/100"],

        ["AUD-2026-0003", "2026-02-15", "10:15", "Auditeur externe (SGS)",
         "Audit fournisseur", "Tanneria Veneta SpA (FRN-2026-0001)",
         "Audit LWG Gold sur site Arzignano — Score 94/100 — 0 non-conformite majeure — 1 observation mineure (traitement effluents)",
         "Valide", "Complet", "Renouvellement Gold confirme — Observation mineure en cours de resolution"],

        ["AUD-2026-0004", "2026-02-24", "08:00", "Dir. Achats",
         "Suspension fournisseur", "Alrosa PJSC (FRN-2026-0014)",
         "Maintien suspension definitive — Sanctions UE toujours actives — Aucun approvisionnement diamants Russie",
         "Valide", "Complet", "Confirme par Dir. Juridique et Compliance — Pas de contournement detecte"],

        ["AUD-2026-0005", "2026-02-28", "11:00", "Dir. Joaillerie",
         "Reception certificat", "KP Botswana (CRT-2026-008)",
         "Reception et verification certificat Kimberley Process lot diamants De Beers — Lot Sightholder #BW-JW-2026-M03 conforme",
         "Valide", "Complet", "Certificat original scanne et archive — Verification numero de serie OK"],

        ["AUD-2026-0006", "2026-03-01", "09:30", "Dir. RSE",
         "Rapport CSRD", "Rapport durabilite exercice 2025",
         "Depot rapport CSRD 2025 — Assurance limitee par EY — Indicateurs ESRS S2 conformes — Double materialite documentee",
         "Valide", "Complet", "Rapport publie — Assurance EY ref. FR-2026-CSRD-00123"],

        ["AUD-2026-0007", "2026-03-15", "16:45", "Dir. RSE",
         "Investigation alerte", "Alerte conditions elevage (ALR-2026-006)",
         "Investigation rapport ONG sur ferme crocodile Louisiane — Audit independant mandate — Resultats : allegations non fondees",
         "Valide", "Verifie", "Alerte cloturee 28/03 — Rapport audit independant archive — Pas de non-conformite"],

        ["AUD-2026-0008", "2026-03-28", "10:00", "Resp. Tracabilite",
         "Reception lot", "LOT-2026-0001 — Cuir veau Tanneria Veneta",
         "Reception lot cuir veau — Controle qualite OK — Documents tracabilite complets — Certificat LWG verifie",
         "Valide", "Complet", "Lot affecte Atelier Maroquinerie Paris 3e — Serie SER-CUI-2026-00847"],

        ["AUD-2026-0009", "2026-04-01", "08:15", "Systeme automatique",
         "Detection alerte", "Certificat LWG Tanneria Veneta (ALR-2026-001)",
         "Alerte automatique : le systeme a confondu ancien certificat (CRT-2026-017 expire) avec le nouveau (CRT-2026-001 actif) — Faux positif resolu",
         "Valide", "Verifie", "Faux positif — Ancien certificat archive — Parametrage mis a jour"],

        ["AUD-2026-0010", "2026-04-03", "14:30", "Resp. Tracabilite",
         "Detection gap tracabilite", "Or 18K — Gap mine-raffinerie (ALR-2026-002)",
         "Gap identifie dans la tracabilite or entre mine artisanale RDC et raffinerie Valcambi — Documentation intermediaire manquante",
         "En attente", "Partiel", "Chain of Custody LBMA exige documentation complete — Demande envoyee a Valcambi"],

        ["AUD-2026-0011", "2026-04-05", "09:45", "Dir. Supply Chain",
         "Alerte geopolitique", "Cachemire — Region Xinjiang (ALR-2026-003)",
         "Nouvelles sanctions EU sur importations textiles Xinjiang — Verification impact sur fournisseurs Tier 2-3 cachemire",
         "En attente", "Partiel", "Erdos Group deja suspendu — Verification Gobi Cashmere supply chain en cours"],

        ["AUD-2026-0012", "2026-04-06", "11:00", "Resp. Qualite",
         "Rapport anomalie", "Python — Taux rebut anormal (ALR-2026-004)",
         "Taux de rebut lot #PY-2026-089 a 18% vs 5% standard — Defauts de finition detectes — Rapport qualite demande au fournisseur",
         "En attente", "Partiel", "Attente rapport qualite Exotic Skins Malaysia — Impact financier estime 15K EUR"],

        ["AUD-2026-0013", "2026-04-08", "15:00", "Dir. Supply Chain",
         "Verification EUDR", "Ebene Macassar — Geolocalisation (LOT-2026-0012)",
         "Verification Due Diligence Statement EUDR — Geolocalisation parcelle fournie — Verification base donnees deforestation en cours",
         "En attente", "Partiel", "EUDR obligatoire depuis 30/12/2025 — DDS a completer avant mise sur marche UE"],

        ["AUD-2026-0014", "2026-04-10", "07:30", "Systeme automatique",
         "Alerte prix matiere", "Platine — Variation +12% (ALR-2026-007)",
         "Alerte automatique : prix platine a depasse le seuil de +10% sur 30 jours — Impact marge collections joaillerie",
         "Valide", "Complet", "Alerte transmise Dir. Finance — Revision pricing en cours — Hedging recommande"],

        ["AUD-2026-0015", "2026-04-12", "09:00", "Dir. Transformation",
         "Revue mensuelle KPI", "Dashboard Agent LuxeTraceability",
         "Revue mensuelle des KPIs — Taux tracabilite 89% (cible 95%) — 2 alertes critiques ouvertes — ROI a +73% vs baseline",
         "Valide", "Complet", "Presentation COMEX — Actions correctives definies — Prochaine revue 12/05"],
    ]

    for row_idx, log_line in enumerate(logs, 5):
        for col_idx, val in enumerate(log_line, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [7, 10] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
        ws.row_dimensions[row_idx].height = 48

    nb_audit = len(logs)
    appliquer_style_donnees(ws, 5, 4 + nb_audit, 10)
    ajouter_mise_en_forme_conditionnelle_statut(ws, "H", 5, 4 + nb_audit)
    ajouter_mise_en_forme_conditionnelle_statut(ws, "I", 5, 4 + nb_audit)

    total_row = 4 + nb_audit + 1
    ligne_synthese(ws, total_row, 10, [
        (1, "SYNTHESE AUDIT",                                       None, "FFFFFF"),
        (2, "Total actions :",                                       None, "FFFFFF"),
        (3, f'=COUNTA(A5:A{4+nb_audit})',                            None, OR_LUXE),
        (5, "Validees :",                                            None, "FFFFFF"),
        (6, f'=COUNTIF(H5:H{4+nb_audit},"Valide")',                  None, "C6EFCE"),
        (7, "En attente :",                                          None, "FFFFFF"),
        (8, f'=COUNTIF(H5:H{4+nb_audit},"En attente")',              None, FOND_MOYEN),
        (9, "Taux validation :",                                     None, "FFFFFF"),
        (10, f'=IFERROR(COUNTIF(H5:H{4+nb_audit},"Valide")/COUNTA(H5:H{4+nb_audit}),0)',
                                                                    '0.0%', OR_LUXE),
    ])

    dv_statut_audit = DataValidation(type="list",
        formula1='"Valide,En attente,Rejete,Archive"', allow_blank=False)
    ws.add_data_validation(dv_statut_audit)
    dv_statut_audit.add('H5:H5000')
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{4+nb_audit}"
    print("  ✅ Onglet 7_AUDIT cree")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 8 : KPI_ROI
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("8_KPI_ROI")
    ws.sheet_properties.tabColor = OR_LUXE

    creer_titre_principal(
        ws,
        "KPI & ROI — AGENT LUXETRACEABILITY",
        "Indicateurs de performance — Retour sur investissement — Tableau de bord executif — Avril 2026",
        10
    )
    appliquer_style_titre_section(ws, 3, 10, "SECTION A — KPI OPERATIONNELS SUPPLY CHAIN", BORDEAUX)

    kpi_headers = ["Indicateur", "Categorie", "Valeur actuelle", "Cible", "Ecart",
                   "Tendance", "Periode", "Source donnee", "Commentaire", "Priorite"]
    kpi_largeurs = [34, 20, 18, 14, 14, 12, 14, 20, 34, 12]
    appliquer_style_en_tetes(ws, 4, kpi_headers, kpi_largeurs)

    kpis = [
        ["Taux de tracabilite complete (%)",
         "Tracabilite", 89.0, 95.0, None, "En hausse", "Avril 2026",
         "4_TRACABILITE col M", "Objectif 95% fin T2 2026 — Gap 6 pts sur lots bois/cachemire", "CRITIQUE"],
        ["Nombre d'alertes critiques ouvertes",
         "Risque", 2, 0, None, "Stable", "Avril 2026",
         "1_DASHBOARD", "ALR-2026-002 (or) + ALR-2026-003 (xinjiang) — Plans d'action actifs", "CRITIQUE"],
        ["Score conformite reglementaire moyen (%)",
         "Conformite", None, 95.0, None, "En hausse", "Avril 2026",
         "5_CONFORMITE col L", "Formule inter-onglets — Score consolide 12 referentiels", "ELEVE"],
        ["Taux de certification active (%)",
         "Certifications", None, 90.0, None, "Stable", "Avril 2026",
         "6_CERTIFICATIONS col I", "Formule inter-onglets — 15 certifs actives / 17 total", "ELEVE"],
        ["Delai moyen tracabilite lot (jours)",
         "Efficacite", 4.2, 3.0, None, "En baisse", "T1 2026",
         "4_TRACABILITE", "Amelioration attendue apres digitalisation ébène Sulawesi", "MOYEN"],
        ["Nombre de fournisseurs Tier 1 certifies LWG/LBMA/RJC",
         "Fournisseurs", 8, 10, None, "Stable", "Avril 2026",
         "3_FOURNISSEURS", "2 fournisseurs cuir exotique en cours de certification SA8000", "MOYEN"],
        ["Taux incidents qualite matieres (%)",
         "Qualite", 3.8, 2.0, None, "En hausse", "T1 2026",
         "3_FOURNISSEURS", "Pic lot python (rebut 18%) — Plan qualite Exotic Skins engage", "CRITIQUE"],
        ["Couverture audit social fournisseurs Tier 1 (%)",
         "Social", 73.0, 100.0, None, "En hausse", "Avril 2026",
         "5_CONFORMITE CONF-012", "11/15 fournisseurs Tier 1 audites — 4 audits Q2-Q3 2026", "ELEVE"],
        ["Nombre de non-conformites CITES detectees",
         "Conformite", 0, 0, None, "Stable", "T1 2026",
         "5_CONFORMITE CONF-005", "Aucune non-conformite — Tous permis valides", "FAIBLE"],
        ["Taux documents traçabilite numerises (%)",
         "Digitalisation", 78.0, 100.0, None, "En hausse", "Avril 2026",
         "4_TRACABILITE", "Objectif 100% T4 2026 — Priorite ébène et cachemire", "MOYEN"],
        ["Delai moyen reponse alerte fournisseur (heures)",
         "Reactivite", 18.5, 12.0, None, "En baisse", "T1 2026",
         "7_AUDIT", "Amelioration process alerte — SLA fournisseurs renforce", "ELEVE"],
        ["Score ESG moyen fournisseurs actifs (/100)",
         "RSE", 68.4, 75.0, None, "En hausse", "Avril 2026",
         "3_FOURNISSEURS col H", "Calcul moyen 13 fournisseurs actifs — Exclus 2 suspendus", "ELEVE"],
    ]

    for row_idx, kpi in enumerate(kpis, 5):
        for col_idx, val in enumerate(kpi, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [1, 9] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
            if col_idx == 3 and isinstance(val, float):
                cell.number_format = '0.0'
        # Formule ecart (col E)
        ws.cell(row=row_idx, column=5).value = f'=IF(AND(ISNUMBER(C{row_idx}),ISNUMBER(D{row_idx})),C{row_idx}-D{row_idx},"")'
        ws.cell(row=row_idx, column=5).number_format = '0.0'
        ws.row_dimensions[row_idx].height = 42

    nb_kpi = len(kpis)
    appliquer_style_donnees(ws, 5, 4 + nb_kpi, 10)

    # MFC ecart (col E) — rouge si negatif, vert si positif
    plage_ecart = f"E5:E{4+nb_kpi}"
    ws.conditional_formatting.add(plage_ecart,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color=FOND_CRITIQUE, end_color=FOND_CRITIQUE, fill_type="solid"),
                   font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(plage_ecart,
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=PatternFill(start_color=FOND_FAIBLE, end_color=FOND_FAIBLE, fill_type="solid"),
                   font=Font(color="006100")))

    # MFC priorite (col J)
    for val, fond, texte, bold in [
        ("CRITIQUE", "FF0000", "FFFFFF", True),
        ("ELEVE",    "FF6600", "FFFFFF", True),
        ("MOYEN",    FOND_MOYEN, "9C6500", False),
        ("FAIBLE",   FOND_FAIBLE, "006100", True),
    ]:
        ws.conditional_formatting.add(f"J5:J{4+nb_kpi}",
            CellIsRule(operator='equal', formula=[f'"{val}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte, bold=bold)))

    # SECTION B — ROI
    roi_start = 4 + nb_kpi + 3
    appliquer_style_titre_section(ws, roi_start, 10, "SECTION B — ANALYSE ROI — INVESTISSEMENT AGENT LUXETRACEABILITY", BORDEAUX)

    roi_headers_row = roi_start + 1
    roi_headers = ["Poste", "Type", "Valeur avant agent (EUR/an)",
                   "Valeur apres agent (EUR/an)", "Gain annuel (EUR)",
                   "Gain (%)", "Investissement (EUR)", "Delai ROI (mois)",
                   "Methodologie", "Confiance"]
    appliquer_style_en_tetes(ws, roi_headers_row, roi_headers, kpi_largeurs)

    roi_data = [
        ["Amendes reglementaires evitees", "Reduction risque", 180000, 25000,
         155000, None, 0, None,
         "Historique amendes secteur luxe 2020-2025 — EUDR + CITES + CSRD", "Elevee"],
        ["Couts audits fournisseurs", "Efficacite", 95000, 52000,
         43000, None, 0, None,
         "Reduction visites terrain grace aux alertes automatisees", "Moyenne"],
        ["Pertes matieres non conformes", "Qualite", 68000, 28000,
         40000, None, 0, None,
         "Reduction rebuts et rejets douaniers — Tracabilite anticipee", "Elevee"],
        ["Couts renégociation fournisseurs", "Negociation", 35000, 12000,
         23000, None, 0, None,
         "Levier benchmark ESG et certifications dans les negotiations", "Moyenne"],
        ["Risques reputationnels (val. brand)", "Marque", 500000, 100000,
         400000, None, 0, None,
         "Estimation impact media negatif sur marques luxe — 1 incident = -5% brand value", "Indicative"],
        ["Investissement agent (annualise)", "Investissement", 0, 0,
         -85000, None, 85000, None,
         "Developpement + maintenance + formation — 3 ans amortissement", "Certaine"],
    ]

    roi_data_start = roi_headers_row + 1
    for row_idx, roi in enumerate(roi_data, roi_data_start):
        for col_idx, val in enumerate(roi, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [1, 9] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
            if col_idx in [3, 4, 5, 7]:
                cell.number_format = '#,##0" EUR"'
        # Gain %
        ws.cell(row=row_idx, column=6).value = f'=IFERROR(E{row_idx}/C{row_idx},"")'
        ws.cell(row=row_idx, column=6).number_format = '0.0%'
        # Delai ROI
        ws.cell(row=row_idx, column=8).value = f'=IFERROR(G{row_idx}/E{row_idx}*12,"")'
        ws.cell(row=row_idx, column=8).number_format = '0.0'
        ws.row_dimensions[row_idx].height = 36

    nb_roi = len(roi_data)
    appliquer_style_donnees(ws, roi_data_start, roi_data_start + nb_roi - 1, 10)

    # Ligne total ROI
    total_roi_row = roi_data_start + nb_roi
    ligne_synthese(ws, total_roi_row, 10, [
        (1,  "TOTAL ROI ANNUEL ESTIME",                               None,           "FFFFFF"),
        (3,  f'=SUM(C{roi_data_start}:C{roi_data_start+nb_roi-1})',   '#,##0" EUR"',  OR_LUXE),
        (4,  f'=SUM(D{roi_data_start}:D{roi_data_start+nb_roi-1})',   '#,##0" EUR"',  OR_LUXE),
        (5,  f'=SUM(E{roi_data_start}:E{roi_data_start+nb_roi-1})',   '#,##0" EUR"',  OR_LUXE),
        (6,  f'=IFERROR(E{total_roi_row}/C{total_roi_row},"")',       '0.0%',         OR_LUXE),
        (7,  f'=SUM(G{roi_data_start}:G{roi_data_start+nb_roi-1})',   '#,##0" EUR"',  OR_LUXE),
    ])

    ws.freeze_panes = "A5"
    print("  ✅ Onglet 8_KPI_ROI cree")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 9 : RISQUES_PAYS
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("9_RISQUES_PAYS")
    ws.sheet_properties.tabColor = ROUGE_ALERTE

    creer_titre_principal(
        ws,
        "CARTOGRAPHIE DES RISQUES PAYS — AGENT LUXETRACEABILITY",
        "Risques geopolitiques, reglementaires et ESG par pays d'origine des matieres premieres — Avril 2026",
        12
    )
    appliquer_style_titre_section(ws, 3, 12, "MATRICE DE RISQUE PAR PAYS FOURNISSEUR", BORDEAUX)

    rp_headers = [
        "Pays", "Region", "Matieres concernees",
        "Risque geopolitique", "Risque social / Droits humains",
        "Risque environnemental", "Risque reglementaire UE",
        "Score risque global (/100)", "Niveau risque",
        "Sanctions actives", "Actions de mitigation", "Statut approvisionnement"
    ]
    rp_largeurs = [20, 16, 28, 16, 18, 18, 18, 16, 14, 20, 36, 20]
    appliquer_style_en_tetes(ws, 4, rp_headers, rp_largeurs)

    risques_pays = [
        ["Italie", "Europe", "Cuir de veau, Cuir agneau, Soie, Tissus",
         "FAIBLE", "FAIBLE", "MOYEN", "FAIBLE", 18, "FAIBLE",
         "Aucune", "RAS — Suivi standard LWG + OEKO-TEX", "Actif — Prioritaire"],

        ["France", "Europe", "Fourrure (legislation), Fournisseurs ateliers",
         "FAIBLE", "FAIBLE", "FAIBLE", "FAIBLE", 12, "FAIBLE",
         "Aucune", "RAS — Legislation fur tres stricte FR", "Actif"],

        ["Espagne", "Europe", "Cuir d'agneau",
         "FAIBLE", "FAIBLE", "FAIBLE", "FAIBLE", 15, "FAIBLE",
         "Aucune", "RAS — Certification LWG Silver — Upgrade Gold en cours", "Actif"],

        ["Suisse", "Europe", "Or (raffinage), Platine (raffinage)",
         "FAIBLE", "FAIBLE", "FAIBLE", "FAIBLE", 8, "TRES FAIBLE",
         "Aucune", "RAS — LBMA Gold + Chain of Custody — Standard referentiel", "Actif — Critique"],

        ["Botswana", "Afrique Australe", "Diamants bruts",
         "MOYEN", "MOYEN", "MOYEN", "FAIBLE", 35, "MOYEN",
         "Aucune", "Kimberley Process — Fournisseur exclusif post-Russie — Monitoring De Beers", "Actif — Surveille"],

        ["Afrique du Sud", "Afrique Australe", "Platine 950",
         "MOYEN", "MOYEN", "MOYEN", "FAIBLE", 38, "MOYEN",
         "Aucune", "RMI Conformant — Audit annuel site minier Mogalakwena — CSRD S2", "Actif"],

        ["Indonesie", "Asie du Sud-Est", "Python, Ebene de Macassar",
         "MOYEN", "ELEVE", "ELEVE", "ELEVE", 72, "ELEVE",
         "Aucune (EUDR a risque)", "CITES Annexe II — EUDR DDS incomplet — Geolocalisation en cours",
         "Sous surveillance — EUDR incomplet"],

        ["Malaisie", "Asie du Sud-Est", "Python reticulé (elevage)",
         "FAIBLE", "MOYEN", "MOYEN", "MOYEN", 42, "MOYEN",
         "Aucune", "CITES quota annuel — Audit ferme elevage Java — Quota 2026 surveille", "Actif — Surveille"],

        ["Mongolie", "Asie Centrale", "Cachemire Grade A",
         "MOYEN", "MOYEN", "MOYEN", "FAIBLE", 40, "MOYEN",
         "Aucune", "GOTS certifie — Gobi Cashmere audite — Pas de lien Xinjiang confirme", "Actif"],

        ["Chine (Xinjiang)", "Asie", "Cachemire (Tier 2-3 potentiel)",
         "ELEVE", "CRITIQUE", "ELEVE", "CRITIQUE", 92, "CRITIQUE",
         "Sanctions EU travail force — Reg. 2024/3015",
         "Erdos Group suspendu — Verification Gobi supply chain Tier 2-3 — Exclusion totale visee",
         "Suspendu — Investigation"],

        ["Russie", "Europe de l'Est", "Diamants (Alrosa — suspendu)",
         "CRITIQUE", "CRITIQUE", "ELEVE", "CRITIQUE", 98, "CRITIQUE",
         "Sanctions UE completes depuis 02/2022 — Reg. 833/2014",
         "AUCUN approvisionnement — Alrosa PJSC sanctionne — Surveillance renforcee anti-contournement",
         "Suspendu — Definitif"],

        ["Etats-Unis (Louisiane)", "Amerique du Nord", "Crocodile / Alligator (elevage)",
         "FAIBLE", "MOYEN", "FAIBLE", "FAIBLE", 22, "FAIBLE",
         "Aucune", "CITES Annexe II — USFWS quota — Alerte ONG resolue en 03/2026 — Audit independant", "Actif"],

        ["Inde", "Asie du Sud", "Cuir tannage (Tier 2), Textiles",
         "MOYEN", "ELEVE", "ELEVE", "MOYEN", 58, "ELEVE",
         "Aucune (risque sanction potentielle)",
         "Tata Leather SA8000 en attente — Audit social programme Q3 2026 — Suivi CSDDD Tier 2",
         "Sous surveillance — Audit requis"],

        ["RDC", "Afrique Centrale", "Or artisanal (ASM potentiel)",
         "CRITIQUE", "CRITIQUE", "ELEVE", "ELEVE", 87, "CRITIQUE",
         "Aucune directe — RMI requis",
         "Gap tracabilite detecte (AUD-2026-0010) — Chain of Custody LBMA incomplete — Valcambi interpele",
         "Sous surveillance — Chain of Custody incomplete"],

        ["Polynesie francaise", "Oceanie", "Perles de Tahiti",
         "FAIBLE", "FAIBLE", "FAIBLE", "FAIBLE", 10, "TRES FAIBLE",
         "Aucune", "Label Perle de Tahiti — Origine garantie — Territoire francais", "Actif"],
    ]

    for row_idx, rp in enumerate(risques_pays, 5):
        for col_idx, val in enumerate(rp, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [3, 11] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
        ws.row_dimensions[row_idx].height = 48

    nb_rp = len(risques_pays)
    appliquer_style_donnees(ws, 5, 4 + nb_rp, 12)

    # MFC risque global col D/E/F/G
    for col_l in ["D", "E", "F", "G"]:
        for val, fond, texte, bold in [
            ("CRITIQUE",    "FF0000",      "FFFFFF", True),
            ("ELEVE",       "FF6600",      "FFFFFF", True),
            ("MOYEN",       FOND_MOYEN,    "9C6500", False),
            ("FAIBLE",      FOND_FAIBLE,   "006100", True),
            ("TRES FAIBLE", FOND_TRES_FAIBLE, "006100", True),
        ]:
            ws.conditional_formatting.add(f"{col_l}5:{col_l}{4+nb_rp}",
                CellIsRule(operator='equal', formula=[f'"{val}"'],
                           fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                           font=Font(color=texte, bold=bold)))

    # Score risque databar (col H)
    ws.conditional_formatting.add(f'H5:H{4+nb_rp}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color=ROUGE_ALERTE))

    # MFC niveau risque (col I)
    for val, fond, texte, bold in [
        ("CRITIQUE",    "FF0000",      "FFFFFF", True),
        ("ELEVE",       "FF6600",      "FFFFFF", True),
        ("MOYEN",       FOND_MOYEN,    "9C6500", False),
        ("FAIBLE",      FOND_FAIBLE,   "006100", True),
        ("TRES FAIBLE", FOND_TRES_FAIBLE, "006100", True),
    ]:
        ws.conditional_formatting.add(f"I5:I{4+nb_rp}",
            CellIsRule(operator='equal', formula=[f'"{val}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte, bold=bold)))

    # MFC statut approvisionnement (col L)
    for keyword, fond, texte in [
        ("Actif", FOND_FAIBLE, "006100"),
        ("Suspendu", FOND_CRITIQUE, "9C0006"),
        ("surveillance", FOND_MOYEN, "9C6500"),
    ]:
        ws.conditional_formatting.add(f"L5:L{4+nb_rp}",
            CellIsRule(operator='containsText' if keyword == "surveillance" else 'equal',
                       formula=[f'"*{keyword}*"' if keyword == "surveillance" else f'"{keyword}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte)))

    total_rp_row = 4 + nb_rp + 1
    ligne_synthese(ws, total_rp_row, 12, [
        (1,  "SYNTHESE PAYS",                                           None, "FFFFFF"),
        (2,  "Total pays :",                                             None, "FFFFFF"),
        (3,  f'=COUNTA(A5:A{4+nb_rp})',                                  None, OR_LUXE),
        (4,  "Critiques :",                                              None, "FFFFFF"),
        (5,  f'=COUNTIF(I5:I{4+nb_rp},"CRITIQUE")',                     None, "FFC7CE"),
        (6,  "Eleves :",                                                 None, "FFFFFF"),
        (7,  f'=COUNTIF(I5:I{4+nb_rp},"ELEVE")',                        None, FOND_MOYEN),
        (8,  "Score moy :",                                              None, "FFFFFF"),
        (9,  f'=IFERROR(AVERAGE(H5:H{4+nb_rp}),0)',                     '0.0', OR_LUXE),
        (10, "Suspendus :",                                              None, "FFFFFF"),
        (11, f'=COUNTIF(L5:L{4+nb_rp},"Suspendu*")',                    None, "FFC7CE"),
    ])

    dv_risque = DataValidation(type="list",
        formula1='"CRITIQUE,ELEVE,MOYEN,FAIBLE,TRES FAIBLE"', allow_blank=False)
    ws.add_data_validation(dv_risque)
    for col_l in ["D5:D5000", "E5:E5000", "F5:F5000", "G5:G5000", "I5:I5000"]:
        dv_risque.add(col_l)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{4+nb_rp}"
    print("  ✅ Onglet 9_RISQUES_PAYS cree")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 10 : PARAMETRES
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("10_PARAMETRES")
    ws.sheet_properties.tabColor = GRIS_LUXE

    creer_titre_principal(
        ws,
        "PARAMETRES — AGENT LUXETRACEABILITY",
        "Configuration de l'agent — Seuils alertes — Referentiels — Tables de donnees — Avril 2026",
        8
    )

    # Section A — Seuils alertes
    appliquer_style_titre_section(ws, 3, 8, "SECTION A — SEUILS D'ALERTES PARAMETRABLES", BORDEAUX)
    seuils_headers = ["Parametre", "Valeur actuelle", "Unité", "Min recommande",
                      "Max recommande", "Description", "Impact si depasse", "Modifiable"]
    appliquer_style_en_tetes(ws, 4, seuils_headers,
        [34, 16, 12, 16, 16, 36, 30, 12])

    seuils = [
        ["Seuil alerte expiration certification (jours)", 30, "Jours", 15, 90,
         "Nombre de jours avant expiration declenchant une alerte rouge", "Rupture certification — Non-conformite", "Oui"],
        ["Seuil alerte expiration certification (avertissement)", 90, "Jours", 30, 180,
         "Nombre de jours avant expiration declenchant un avertissement orange", "Preparation renouvellement", "Oui"],
        ["Taux de rebut matiere — Seuil alerte (%)", 5.0, "%", 1.0, 15.0,
         "Taux de rebut par lot au-dela duquel une alerte qualite est generee", "Impact marge + risque conformite", "Oui"],
        ["Variation prix matiere — Seuil alerte (%)", 10.0, "%", 5.0, 25.0,
         "Variation du prix matiere sur 30 jours declenchant une alerte", "Impact marge collections", "Oui"],
        ["Score risque fournisseur — Seuil suspension auto", 80, "/100", 60, 90,
         "Score de risque au-dela duquel la suspension automatique est proposee", "Interruption approvisionnement", "Oui"],
        ["Taux tracabilite minimum requis (%)", 95.0, "%", 80.0, 100.0,
         "Taux de tracabilite complete minimum pour les lots en production", "Blocage mise sur marche UE", "Oui"],
        ["Frequence mise a jour dashboard (minutes)", 60, "Min", 15, 1440,
         "Frequence de recalcul automatique des KPIs dashboard", "Fraicheur des donnees", "Oui"],
        ["Retention historique audit (annees)", 10, "Annees", 5, 20,
         "Duree de conservation des logs d'audit (exigence reglementaire CSRD : 10 ans)", "Conformite archivage", "Non"],
    ]

    for row_idx, s in enumerate(seuils, 5):
        for col_idx, val in enumerate(s, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [1, 6, 7] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
        ws.row_dimensions[row_idx].height = 40
    nb_s = len(seuils)
    appliquer_style_donnees(ws, 5, 4 + nb_s, 8)

    # Section B — Referentiel maisons
    appliquer_style_titre_section(ws, 4 + nb_s + 2, 8,
        "SECTION B — REFERENTIEL MAISONS CONFIGURABLES", BORDEAUX)
    maisons_h_row = 4 + nb_s + 3
    maisons_headers = ["Code Maison", "Nom Maison", "Categorie principale",
                       "Marches principaux", "Standards applicables",
                       "Fournisseurs Tier 1 dedies", "Contact RSE", "Statut"]
    appliquer_style_en_tetes(ws, maisons_h_row, maisons_headers,
        [14, 22, 20, 20, 30, 20, 22, 12])

    maisons = [
        ["MAI-001", "Maison Aurelia", "Maroquinerie / Cuir",
         "Europe, Asie, Amerique", "LWG+REACH+CITES+Loi Vigilance+CSRD", "Tanneria Veneta, Walpier", "rse@aurelia.com", "Actif"],
        ["MAI-002", "Maison Lumiere", "Joaillerie / Haute Couture",
         "Europe, Moyen-Orient, Japon", "LBMA+KPCS+RJC+CITES+CSRD", "Valcambi, De Beers, Anglo American", "rse@lumiere.com", "Actif"],
        ["MAI-003", "Maison Soir", "Pret-a-porter Luxe",
         "Europe, Asie", "GOTS+OEKO-TEX+SA8000+CSRD+CSDDD", "Gobi Cashmere, Seteria Bianchi", "rse@soir.com", "Actif"],
        ["MAI-004", "Maison Exotique", "Maroquinerie Exotique",
         "Europe, Asie, Golfe", "CITES+LWG+SA8000+Loi Vigilance", "Exotic Skins Malaysia, Louisiana Farm", "rse@exotique.com", "Sous surveillance"],
    ]

    for row_idx, m in enumerate(maisons, maisons_h_row + 1):
        for col_idx, val in enumerate(m, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDURE_FINE
        ws.row_dimensions[row_idx].height = 36
    nb_m = len(maisons)
    appliquer_style_donnees(ws, maisons_h_row + 1, maisons_h_row + nb_m, 8)
    ajouter_mise_en_forme_conditionnelle_statut(ws, "H",
        maisons_h_row + 1, maisons_h_row + nb_m)

    ws.freeze_panes = "A5"
    print("  ✅ Onglet 10_PARAMETRES cree")

    # ═══════════════════════════════════════════════════════════════════
    # ONGLET 11 : JOURNAL
    # ═══════════════════════════════════════════════════════════════════

    ws = wb.create_sheet("11_JOURNAL")
    ws.sheet_properties.tabColor = NOIR_LUXE

    creer_titre_principal(
        ws,
        "JOURNAL SYSTEME — AGENT LUXETRACEABILITY",
        "Log technique des imports, exports, synchronisations et erreurs — Trace complete — Avril 2026",
        9
    )
    appliquer_style_titre_section(ws, 3, 9, "JOURNAL TECHNIQUE — EVENEMENTS SYSTEME", BORDEAUX)

    jrnl_headers = [
        "ID_Event", "Timestamp", "Type evenement", "Composant",
        "Description", "Donnees entree", "Donnees sortie",
        "Statut execution", "Duree (ms)"
    ]
    jrnl_largeurs = [16, 20, 18, 18, 44, 24, 24, 14, 12]
    appliquer_style_en_tetes(ws, 4, jrnl_headers, jrnl_largeurs)

    events = [
        ["EVT-2026-000001", "2026-01-15 09:00:12", "IMPORT", "Parser Excel",
         "Import initial fichier NEURAL_LuxeTraceability.xlsx — Chargement 4 onglets de base",
         "NEURAL_LuxeTraceability.xlsx v1.0", "18 matieres / 15 fournisseurs / 12 lots",
         "Succes", 1240],

        ["EVT-2026-000002", "2026-01-15 09:00:14", "CALCUL", "Dashboard KPI",
         "Premier calcul KPIs dashboard — Etablissement baseline metriques",
         "Donnees onglets 2-3-4", "KPIs initialises — Taux tracabilite 76% (baseline)",
         "Succes", 380],

        ["EVT-2026-000003", "2026-02-15 10:18:43", "AUDIT", "Module Conformite",
         "Enregistrement audit LWG Tanneria Veneta — Score 94/100 — Certification Gold renouvelee",
         "Rapport SGS audit #2026-LWG-TV-001", "CRT-2026-001 mise a jour — AUD-2026-0003 cree",
         "Succes", 215],

        ["EVT-2026-000004", "2026-02-24 08:01:05", "MISE A JOUR", "Registre Fournisseurs",
         "Mise a jour statut Alrosa PJSC — Suspension maintenue — Sanctions UE confirmees",
         "Decision Dir. Achats + Dir. Juridique", "FRN-2026-0014 statut='Suspendu definitif'",
         "Succes", 89],

        ["EVT-2026-000005", "2026-03-28 10:05:33", "RECEPTION", "Module Tracabilite",
         "Reception et enregistrement lot cuir veau LOT-2026-0001 — Controle qualite integre",
         "BL transport + certificat LWG + rapport qualite", "LOT-2026-0001 statut='Complet' — Score 97%",
         "Succes", 567],

        ["EVT-2026-000006", "2026-04-01 00:00:01", "ALERTE AUTO", "Moniteur Certifications",
         "Execution planifiee : verification certifications expirant dans 30 jours — Scan complet registre",
         "6_CERTIFICATIONS — 17 certifications", "1 alerte rouge generee (FSC ITC expire dans 35j)",
         "Succes", 423],

        ["EVT-2026-000007", "2026-04-03 14:31:18", "ALERTE", "Analyseur Tracabilite",
         "Detection gap tracabilite : lot or 18K — segment mine-raffinerie non documente",
         "LOT-2026-0008 historique tracabilite", "ALR-2026-002 cree — Notification envoyee Dir. Supply Chain",
         "Succes", 312],

        ["EVT-2026-000008", "2026-04-05 09:46:02", "VEILLE", "Module Geopolitique",
         "Veille reglementaire : detection nouvelles sanctions EU textiles Xinjiang — Reg. (EU) 2024/3015",
         "Feed EUR-Lex + Journal Officiel UE", "ALR-2026-003 cree — 2 fournisseurs identifies en risque Tier 2-3",
         "Succes", 1890],

        ["EVT-2026-000009", "2026-04-06 11:01:44", "QUALITE", "Analyseur Lots",
         "Detection anomalie taux rebut lot python #PY-2026-089 — 18% vs seuil 5%",
         "Rapport qualite reception atelier", "ALR-2026-004 cree — Email automatique Exotic Skins Malaysia",
         "Succes", 198],

        ["EVT-2026-000010", "2026-04-08 15:02:11", "CONFORMITE", "Module EUDR",
         "Verification DDS EUDR lot ebene LOT-2026-0012 — Geolocalisation parcelle incomplete",
         "Lot LOT-2026-0012 + base ForestMapper", "Statut DDS='Partiel' — Blocage mise sur marche UE signale",
         "Avertissement", 2340],

        ["EVT-2026-000011", "2026-04-10 00:00:01", "PRIX", "Moniteur Marche",
         "Execution planifiee : verification variations prix matieres — Platine +12% sur 30 jours detecte",
         "Feed London Metal Exchange + LBMA", "ALR-2026-007 cree — Notification Dir. Finance + Dir. Achats",
         "Succes", 876],

        ["EVT-2026-000012", "2026-04-12 09:01:33", "EXPORT", "Module Reporting",
         "Generation rapport mensuel KPIs pour COMEX — Consolidation tous onglets",
         "Onglets 1-11 — Donnees avril 2026", "Rapport COMEX-2026-04.pdf — 12 KPIs consolides",
         "Succes", 4521],

        ["EVT-2026-000013", "2026-04-12 09:05:44", "ERREUR", "Module CSDDD",
         "Tentative de mapping Tier 3 pour fournisseur Gobi Cashmere — API externe non disponible",
         "API EcoVadis supply chain mapping", "Erreur connexion timeout — Retry programme dans 24h",
         "Erreur", 30000],

        ["EVT-2026-000014", "2026-04-12 09:05:45", "RETRY", "Module CSDDD",
         "Retry planifie mapping Tier 3 Gobi Cashmere suite erreur EVT-2026-000013",
         "EVT-2026-000013 — Retry #1/3", "En attente execution 13/04/2026 09:05",
         "En attente", 0],

        ["EVT-2026-000015", "2026-04-12 10:00:00", "SAUVEGARDE", "Systeme",
         "Sauvegarde automatique quotidienne — Backup complet workbook + logs",
         "NEURAL_LuxeTraceability.xlsx + 7_AUDIT + 11_JOURNAL",
         "Backup_LuxeTrace_2026-04-12.xlsx — Archive cloud + local",
         "Succes", 1120],
    ]

    for row_idx, ev in enumerate(events, 5):
        for col_idx, val in enumerate(ev, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [5, 6, 7] else "center",
                vertical="center", wrap_text=True
            )
            cell.border = BORDURE_FINE
            if col_idx == 9 and isinstance(val, int):
                cell.number_format = '#,##0'
        ws.row_dimensions[row_idx].height = 44

    nb_ev = len(events)
    appliquer_style_donnees(ws, 5, 4 + nb_ev, 9)

    # MFC type evenement (col C)
    for val, fond, texte, bold in [
        ("ERREUR",      FOND_CRITIQUE, "9C0006", True),
        ("ALERTE AUTO", FOND_MOYEN,    "9C6500", True),
        ("ALERTE",      FOND_MOYEN,    "9C6500", True),
        ("IMPORT",      FOND_FAIBLE,   "006100", False),
        ("EXPORT",      FOND_FAIBLE,   "006100", False),
        ("SUCCES",      FOND_FAIBLE,   "006100", True),
    ]:
        ws.conditional_formatting.add(f"C5:C{4+nb_ev}",
            CellIsRule(operator='equal', formula=[f'"{val}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte, bold=bold)))

    # MFC statut execution (col H)
    ajouter_mise_en_forme_conditionnelle_statut(ws, "H", 5, 4 + nb_ev)
    for val, fond, texte in [
        ("Succes",      FOND_FAIBLE,   "006100"),
        ("Erreur",      FOND_CRITIQUE, "9C0006"),
        ("Avertissement", FOND_MOYEN,  "9C6500"),
        ("En attente",  FOND_MOYEN,    "9C6500"),
    ]:
        ws.conditional_formatting.add(f"H5:H{4+nb_ev}",
            CellIsRule(operator='equal', formula=[f'"{val}"'],
                       fill=PatternFill(start_color=fond, end_color=fond, fill_type="solid"),
                       font=Font(color=texte, bold=True)))

    total_ev_row = 4 + nb_ev + 1
    ligne_synthese(ws, total_ev_row, 9, [
        (1, "SYNTHESE JOURNAL",                                              None,  "FFFFFF"),
        (2, "Total evenements :",                                             None,  "FFFFFF"),
        (3, f'=COUNTA(A5:A{4+nb_ev})',                                        None,  OR_LUXE),
        (4, "Succes :",                                                        None,  "FFFFFF"),
        (5, f'=COUNTIF(H5:H{4+nb_ev},"Succes")',                              None,  "C6EFCE"),
        (6, "Erreurs :",                                                       None,  "FFFFFF"),
        (7, f'=COUNTIF(H5:H{4+nb_ev},"Erreur")',                              None,  FOND_CRITIQUE),
        (8, "Duree moy (ms) :",                                               None,  "FFFFFF"),
        (9, f'=IFERROR(AVERAGEIF(H5:H{4+nb_ev},"Succes",I5:I{4+nb_ev}),0)',  '#,##0', OR_LUXE),
    ])

    dv_type_ev = DataValidation(type="list",
        formula1='"IMPORT,EXPORT,CALCUL,ALERTE,ALERTE AUTO,AUDIT,RECEPTION,QUALITE,CONFORMITE,PRIX,VEILLE,MISE A JOUR,ERREUR,RETRY,SAUVEGARDE"',
        allow_blank=False)
    ws.add_data_validation(dv_type_ev)
    dv_type_ev.add('C5:C5000')

    dv_statut_ev = DataValidation(type="list",
        formula1='"Succes,Erreur,Avertissement,En attente,Annule"', allow_blank=False)
    ws.add_data_validation(dv_statut_ev)
    dv_statut_ev.add('H5:H5000')

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{4+nb_ev}"
    print("  ✅ Onglet 11_JOURNAL cree")

    # ═══════════════════════════════════════════════════════════════════
    # SAUVEGARDE
    # ═══════════════════════════════════════════════════════════════════

    import os
    output_path = "C:/Users/Ludo/Desktop/IA projet entreprises/NEURAL_LuxeTraceability.xlsx"
    temp_path   = "C:/Users/Ludo/Desktop/IA projet entreprises/NEURAL_LuxeTraceability_NEW.xlsx"
    wb.save(temp_path)
    # Remplacer l'original si possible, sinon garder le _NEW
    try:
        if os.path.exists(output_path):
            os.replace(temp_path, output_path)
        else:
            os.rename(temp_path, output_path)
    except PermissionError:
        output_path = temp_path
        print(f"  INFO: Fichier original verrouillle — sauvegarde sous nom alternatif")
    print(f"\n  Fichier mis a jour : {output_path}")
    print(f"  Onglets ajoutes : 5_CONFORMITE · 6_CERTIFICATIONS · 7_AUDIT · 8_KPI_ROI · 9_RISQUES_PAYS · 10_PARAMETRES · 11_JOURNAL")
    print(f"  Total workbook : 11 onglets complets\n")


if __name__ == "__main__":
    ajouter_onglets_5_11()
