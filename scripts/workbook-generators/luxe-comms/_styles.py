"""Shared styling + helpers for NEURAL / LUXE / Communication workbooks."""
from datetime import date, datetime

from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# Palette LUXE (aligne sur les autres workbooks NEURAL)
NOIR = "1A1A1A"
OR = "C9A84C"
BORDEAUX = "6B1D2A"
BLANC_CASSE = "F9F6F0"
GRIS = "8C8C8C"
VIOLET = "7C3AED"   # Violet IA (identite NEURAL)
BLEU_NUIT = "0A1628"
VERT = "2E7D32"
BLEU = "1565C0"
ROUGE = "C62828"
ORANGE = "E65100"
FOND_ROUGE = "FFC7CE"
FOND_JAUNE = "FFEB9C"
FOND_VERT = "C6EFCE"
FOND_BLEU = "D6E4F0"
FOND_VIOLET = "EDE9FE"

BF = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BF_EPAIS = Border(
    left=Side(style="medium", color=NOIR),
    right=Side(style="medium", color=NOIR),
    top=Side(style="medium", color=NOIR),
    bottom=Side(style="medium", color=NOIR),
)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)


def titre(ws, t1, t2, mc):
    """Bandeau titre sur 2 lignes, fond noir, accent or."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
    c = ws.cell(row=1, column=1, value=t1)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 44
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)
    c2 = ws.cell(row=2, column=1, value=t2)
    c2.font = Font(name="Calibri", size=10, italic=True, color=OR)
    c2.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26


def section(ws, row, mc, texte, couleur=BORDEAUX):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1, value=texte)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def en_tetes(ws, row, headers, largeurs=None, couleur=NOIR):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
        c.alignment = AC
        c.border = BF
    ws.row_dimensions[row].height = 34
    if largeurs:
        for i, w in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def ecrire_donnees(ws, start_row, data, fmt_cols=None, align=None):
    for ri, ligne in enumerate(data, start_row):
        for ci, val in enumerate(ligne, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=9)
            c.alignment = align if align else AC
            c.border = BF
            if (ri - start_row) % 2 == 0:
                c.fill = PatternFill(start_color=BLANC_CASSE, end_color=BLANC_CASSE, fill_type="solid")
            if fmt_cols and ci in fmt_cols:
                c.number_format = fmt_cols[ci]
    return start_row + len(data) - 1 if data else start_row - 1


def param_ligne(ws, row, label, value, note=""):
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c1.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
    c1.alignment = AL
    c1.border = BF
    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = Font(name="Calibri", size=10, bold=True)
    c2.fill = PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")
    c2.alignment = AC
    c2.border = BF_EPAIS
    if note:
        c3 = ws.cell(row=row, column=3, value=note)
        c3.font = Font(name="Calibri", size=9, italic=True, color=GRIS)
        c3.alignment = AL
        c3.border = BF


def ajouter_table(ws, nom, ref, style="TableStyleMedium9"):
    """Wrap un range en Excel Table structuree."""
    try:
        t = Table(displayName=nom, ref=ref)
        t.tableStyleInfo = TableStyleInfo(
            name=style, showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(t)
    except Exception as e:
        print(f"[warn] Table {nom} non creee: {e}")


def dv_liste(values):
    """DataValidation list-based."""
    lst = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=lst, allow_blank=True)
    dv.error = "Valeur hors liste"
    dv.errorTitle = "Entree invalide"
    return dv


def today_iso():
    return datetime.now().strftime("%Y-%m-%d")


def ajuster_hauteurs(ws, start_row, end_row, h=30):
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = h


# Enums communs figes par l'architecture
STATUS = ["DRAFT", "IN_REVIEW", "APPROVED", "REWORK", "REJECTED", "BLOCKED"]
RISK_LEVEL = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
GATE_TYPE = ["BRAND", "CLAIM", "HERITAGE", "LEGAL", "EVENT", "CRISIS"]
LANG = ["FR", "EN", "IT", "DE", "JA", "ZH"]
DECISION = ["APPROVE", "REWORK", "REJECT", "PENDING"]
YESNO = ["YES", "NO"]
SEVERITY = ["INFO", "LOW", "MEDIUM", "HIGH", "CRITICAL"]
PRIORITE = ["P0", "P1", "P2", "P3"]
TERM_TYPE = ["PREFERRED", "FORBIDDEN", "REVIEW"]
WORDING_TYPE = ["ABSOLUTE", "QUALIFIED", "COMPARATIVE"]


# ---------------------------------------------------------------------------
# Helpers introduits par Sprint correctif (audit)
# ---------------------------------------------------------------------------

def D(s):
    """Convertit une string 'YYYY-MM-DD' en date. None si vide."""
    if isinstance(s, (date, datetime)):
        return s
    if s is None or s == "":
        return None
    return date.fromisoformat(str(s)[:10])


def DT(s):
    """Convertit 'YYYY-MM-DD HH:MM' ou 'YYYY-MM-DD HH:MM:SS' en datetime."""
    if isinstance(s, datetime):
        return s
    if s is None or s == "":
        return None
    ss = str(s).strip()
    if len(ss) <= 10:
        return datetime.fromisoformat(ss + "T00:00:00")
    if len(ss) == 16:
        return datetime.fromisoformat(ss.replace(" ", "T") + ":00")
    return datetime.fromisoformat(ss.replace(" ", "T"))


def convertir_dates(ws, start_row, end_row, date_cols, datetime_cols=None):
    """
    Apres `ecrire_donnees`, convertit les cellules de dates string -> date/datetime
    et applique le bon number_format. Ne touche que les colonnes listees.
    """
    datetime_cols = datetime_cols or []
    for r in range(start_row, end_row + 1):
        for col in date_cols:
            c = ws.cell(row=r, column=col)
            if isinstance(c.value, str) and c.value.strip():
                try:
                    c.value = D(c.value)
                    c.number_format = "yyyy-mm-dd"
                except Exception:
                    pass
            elif isinstance(c.value, (date, datetime)) and not isinstance(c.value, datetime):
                c.number_format = "yyyy-mm-dd"
        for col in datetime_cols:
            c = ws.cell(row=r, column=col)
            if isinstance(c.value, str) and c.value.strip():
                try:
                    c.value = DT(c.value)
                    c.number_format = "yyyy-mm-dd hh:mm"
                except Exception:
                    pass
            elif isinstance(c.value, datetime):
                c.number_format = "yyyy-mm-dd hh:mm"


def params_table(ws, table_name, start_row, rows, col_widths=(30, 28, 60)):
    """
    Ecrit header KEY | VALUE | NOTE en `start_row`, les donnees en-dessous,
    wrap en Excel Table. Retourne last_row. Pour AG-001 qui utilise XLOOKUP
    sur 1_PARAMS, on passe par le nom de table (pas des references absolues).
    """
    # Widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Header
    headers = ["KEY", "VALUE", "NOTE"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
        c.alignment = AC
        c.border = BF
    # Rows
    last = start_row
    for i, row in enumerate(rows):
        r = start_row + 1 + i
        last = r
        k, v, note = row
        c1 = ws.cell(row=r, column=1, value=k)
        c1.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c1.fill = PatternFill(start_color=NOIR, end_color=NOIR, fill_type="solid")
        c1.alignment = AL
        c1.border = BF
        c2 = ws.cell(row=r, column=2, value=v)
        c2.font = Font(name="Calibri", size=10, bold=True)
        c2.fill = PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid")
        c2.alignment = AC
        c2.border = BF_EPAIS
        if isinstance(v, (date, datetime)) and not isinstance(v, datetime):
            c2.number_format = "yyyy-mm-dd"
        elif isinstance(v, datetime):
            c2.number_format = "yyyy-mm-dd hh:mm"
        c3 = ws.cell(row=r, column=3, value=note)
        c3.font = Font(name="Calibri", size=9, italic=True, color=GRIS)
        c3.alignment = AL
        c3.border = BF
    # Table wrap
    try:
        t = Table(displayName=table_name, ref=f"A{start_row}:C{last}")
        t.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False,
        )
        ws.add_table(t)
    except Exception as e:
        print(f"[warn] params_table {table_name} : {e}")
    return last


def attach_dv(ws, range_ref, values):
    """Attache une DataValidation liste sur un range."""
    lst = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=lst, allow_blank=True,
                        showErrorMessage=True, errorTitle="Valeur hors liste",
                        error="La valeur saisie n'est pas dans la liste autorisee.")
    ws.add_data_validation(dv)
    dv.add(range_ref)
    return dv


def cf_stale(ws, range_ref, days_window=30):
    """CF : cellule jaune si date < TODAY()+days_window et >= TODAY() ; rouge si date < TODAY()."""
    from openpyxl.formatting.rule import FormulaRule
    # Trouver la 1ere cellule du range pour la formule relative
    first = range_ref.split(":")[0]
    ws.conditional_formatting.add(
        range_ref,
        FormulaRule(formula=[f'AND(ISNUMBER({first}),{first}<TODAY())'],
                    fill=PatternFill(start_color=FOND_ROUGE, end_color=FOND_ROUGE, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        range_ref,
        FormulaRule(formula=[f'AND(ISNUMBER({first}),{first}>=TODAY(),{first}<TODAY()+{days_window})'],
                    fill=PatternFill(start_color=FOND_JAUNE, end_color=FOND_JAUNE, fill_type="solid"))
    )
