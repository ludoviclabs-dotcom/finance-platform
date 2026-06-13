"""Tests T3.3 — export Rapport VSME (PDF + Excel + ZIP auditable).

Pur (sans DB) : rendu PDF/Excel depuis un mapping, structure du ZIP, déterminisme
du manifest_hash, CHECKSUMS valides, garde auth.
"""

from __future__ import annotations

import hashlib
import io
import zipfile

from fastapi.testclient import TestClient

from main import app
from services import vsme_export
from services import vsme_mapping_service as svc

client = TestClient(app)

SNAPSHOT = {
    "profile": {"raisonSociale": "Exemplia", "secteurNaf": "62.01", "etp": 50, "caNet": 1_000_000,
                "anneeReporting": 2025, "pays": "FR", "perimetre": "mono"},
    "environnement": {"scope1Tco2e": 100, "scope2LbTco2e": 50, "scope3Tco2e": 200, "totalGesTco2e": 350,
                      "intensiteCaGes": 35, "energieMwh": 500, "partEnrPct": 30, "eauM3": 1000,
                      "dechetsTonnes": 10, "planReductionGes": "Plan 2030"},
    "social": {}, "gouvernance": {},
}


def _mapping():
    rows = svc.map_datapoints(SNAPSHOT, {})
    return {"version": "v2024.12", "completeness": svc.completeness(rows), "datapoints": rows}


class TestRenderers:
    def test_pdf_is_pdf(self) -> None:
        b = vsme_export.build_vsme_pdf(company_name="Exemplia", mapping=_mapping(), generated_at="13/06/2026")
        assert b[:5] == b"%PDF-"
        assert len(b) > 1000

    def test_xlsx_opens_with_module_sheets(self) -> None:
        b = vsme_export.build_vsme_xlsx(company_name="Exemplia", mapping=_mapping())
        assert b[:4] == b"PK\x03\x04"
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(b))
        assert "Synthèse" in wb.sheetnames
        assert "B3" in wb.sheetnames

    def test_xlsx_has_row_hashes(self) -> None:
        b = vsme_export.build_vsme_xlsx(company_name="Exemplia", mapping=_mapping())
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(b))
        ws = wb["B3"]
        # En-tête + au moins une ligne ; dernière colonne = hash 64 hex
        assert ws.cell(row=1, column=8).value.startswith("Hash")
        assert len(ws.cell(row=2, column=8).value) == 64

    def test_xlsx_reflects_mixed_statuses(self) -> None:
        # Mélange auto (E1) + manuel (override) + na (justifié) dans le même export.
        rows = svc.map_datapoints(SNAPSHOT, {
            "B4-1": {"is_applicable": False, "na_justification": "Non matériel pour le secteur"},
            "B10-2": {"value": 12, "is_applicable": True},
        })
        mapping = {"version": "v2024.12", "completeness": svc.completeness(rows), "datapoints": rows}
        b = vsme_export.build_vsme_xlsx(company_name="Exemplia", mapping=mapping)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(b))
        cells: dict[str, tuple] = {}
        for sheet in wb.sheetnames:
            if sheet == "Synthèse":
                continue
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                cells[row[0]] = (row[5], row[4])  # code -> (statut, valeur)
        assert cells["B4-1"] == ("na", "Non applicable")
        assert cells["B10-2"][0] == "manuel"
        assert cells["B3-1"][0] == "auto"  # Scope 1 auto depuis E1


class TestPackage:
    def test_zip_contents(self) -> None:
        res = vsme_export.build_vsme_report(company_id=1, company_name="Exemplia", mapping=_mapping())
        with zipfile.ZipFile(io.BytesIO(res["zip_bytes"])) as zf:
            names = set(zf.namelist())
        assert {"manifest.json", "rapport-vsme.pdf", "annexe-vsme.xlsx", "CHECKSUMS.sha256", "README.txt"} <= names
        assert len(res["package_hash"]) == 64
        assert len(res["manifest_hash"]) == 64

    def test_manifest_hash_deterministic(self) -> None:
        m = _mapping()
        a = vsme_export.build_vsme_report(company_id=1, company_name="Exemplia", mapping=m)
        b = vsme_export.build_vsme_report(company_id=1, company_name="Exemplia", mapping=m)
        assert a["manifest_hash"] == b["manifest_hash"]

    def test_checksums_match(self) -> None:
        res = vsme_export.build_vsme_report(company_id=1, company_name="Exemplia", mapping=_mapping())
        with zipfile.ZipFile(io.BytesIO(res["zip_bytes"])) as zf:
            checksums = zf.read("CHECKSUMS.sha256").decode("utf-8")
            for line in checksums.strip().splitlines():
                sha, name = line.split("  ", 1)
                assert hashlib.sha256(zf.read(name)).hexdigest() == sha


class TestEndpoint:
    def test_requires_auth(self) -> None:
        assert client.post("/vsme/report").status_code in (401, 403)
