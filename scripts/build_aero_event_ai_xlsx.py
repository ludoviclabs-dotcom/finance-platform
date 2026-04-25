"""
Build AeroEventAI_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'AeroEventAI' (branche Aéronautique — Marketing).

7 onglets : README, Calendrier salons 2026, Pack événementiel par salon,
Verdict par pack, Redlines, AI Act Disclosure, Limites.
4 salons réels Q2-Q4 2026 (Farnborough, ILA Berlin, Eurosatory, MEBAA).

Date de référence : 25/04/2026.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"
AGENT_VIOLET = "7C3AED"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=AGENT_VIOLET, end_color=AGENT_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 60) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def verdict_fill(verdict: str) -> PatternFill:
    if verdict == "OK":
        return green_fill()
    if verdict == "WARN":
        return amber_fill()
    if verdict == "KO":
        return red_fill()
    return PatternFill()


def verdict_color(verdict: str) -> str:
    if verdict == "OK":
        return GREEN_OK
    if verdict == "WARN":
        return AMBER_WARN
    if verdict == "KO":
        return RED_NO
    return "0E0824"


# ────────────────────────────────────────────────────────────────────
# Données : 4 salons réels 2026
# ────────────────────────────────────────────────────────────────────
EVENTS = [
    {
        "id": "FAR-2026",
        "name": "Farnborough International Airshow 2026",
        "dates": "20-24 juillet 2026",
        "lieu": "Farnborough Airport, UK",
        "organisateur": "ADS Group",
        "publics": "B2B civil + défense — décideurs, CEO, ministres, presse spécialisée",
        "fuseau": "Europe/London (BST UTC+1)",
        "languages": "EN principal, FR / DE / IT secondaires",
        "tone": "Civil dominant, défense présent. Tonalité innovation + souveraineté EU.",
    },
    {
        "id": "ILA-2026",
        "name": "ILA Berlin 2026",
        "dates": "10-14 juin 2026",
        "lieu": "Berlin Brandenburg Airport (BER), Allemagne",
        "organisateur": "BDLI (Bundesverband der Deutschen Luft- und Raumfahrtindustrie)",
        "publics": "Civil + défense + spatial — focus EU, Bundeswehr, ESA",
        "fuseau": "Europe/Berlin (CEST UTC+2)",
        "languages": "DE principal, EN secondaire",
        "tone": "Souveraineté EU + transition énergétique. Tonalité plus industrielle.",
    },
    {
        "id": "EUR-2026",
        "name": "Eurosatory 2026",
        "dates": "15-19 juin 2026",
        "lieu": "Paris-Nord Villepinte, France",
        "organisateur": "COGES (Commissariat général du salon)",
        "publics": "Défense terrestre + drones + sécurité — délégations 100+ pays",
        "fuseau": "Europe/Paris (CEST UTC+2)",
        "languages": "FR principal, EN secondaire",
        "tone": "Strict défense — charte ASD impérative, contexte géopolitique post-Ukraine.",
    },
    {
        "id": "MEB-2026",
        "name": "MEBAA Show 2026",
        "dates": "8-10 décembre 2026",
        "lieu": "Dubai World Central (DWC), Émirats Arabes Unis",
        "organisateur": "MEBAA",
        "publics": "Aviation d'affaires + business jets — clientèle MEA + Asie",
        "fuseau": "Asia/Dubai (GST UTC+4)",
        "languages": "EN principal, AR secondaire",
        "tone": "Aviation d'affaires haut de gamme. Attention dual-use + sanctions Iran/Russie via UAE.",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : 10 règles encodage pack événementiel
# ────────────────────────────────────────────────────────────────────
RULES = [
    {
        "id": "E01",
        "label": "ASD-RESPONSIBLE",
        "axis": "Tonalité",
        "rule": "Charte ASD Europe « responsible defence comms » : pas de glorification, pas de visuel choquant, contexte stratégique sourcé.",
        "base": "ASD Europe Charter (mars 2025).",
    },
    {
        "id": "E02",
        "label": "AI-DISCLOSE",
        "axis": "AI Act",
        "rule": "Tout contenu généré par IA (post, brief, talking points) requiert mention IA — applicable 2 août 2026.",
        "base": "EU AI Act 2024/1689 art. 50.",
    },
    {
        "id": "E03",
        "label": "EXPORT-SAFE",
        "axis": "Export Control",
        "rule": "Pas d'information ITAR/EAR/dual-use sensible diffusée au-delà du périmètre de salon contrôlé.",
        "base": "ITAR 22 CFR 120-130, EAR 15 CFR 730-774, EU Reg. 2021/821.",
    },
    {
        "id": "E04",
        "label": "NO-CLASSIFIED",
        "axis": "Sécurité",
        "rule": "Aucune information classifiée DR / SD / TS dans contenu événementiel public (brief presse, social, talking points).",
        "base": "SGDSN — Code défense L.2311-1.",
    },
    {
        "id": "E05",
        "label": "SOURCE-FRESHNESS",
        "axis": "Veille",
        "rule": "Chiffres et benchmarks ≤ 24 mois (sauf normes pérennes).",
        "base": "Bonne pratique communication B2B.",
    },
    {
        "id": "E06",
        "label": "NUM-VALIDATED",
        "axis": "Précision",
        "rule": "Tout chiffre technique sourcé. Pas de chiffre rond marketing sans référence.",
        "base": "Code consommation L.121-1 + ASD Charter.",
    },
    {
        "id": "E07",
        "label": "TIMEZONE",
        "axis": "Logistique",
        "rule": "Embargos presse + horaires conf en heure locale du salon, mention explicite (BST/CEST/GST).",
        "base": "Bonne pratique relations presse internationales.",
    },
    {
        "id": "E08",
        "label": "LANGUE-LOCALE",
        "axis": "Multi-langue",
        "rule": "Contenu disponible dans la langue principale du salon (EN Farnborough/MEBAA, DE ILA, FR Eurosatory) + EN systématique.",
        "base": "Bonne pratique relations presse multi-pays.",
    },
    {
        "id": "E09",
        "label": "EMBARGO-OK",
        "axis": "Presse",
        "rule": "Embargo presse explicite (date + heure + fuseau) sur tout brief presse pré-salon.",
        "base": "Bonne pratique relations presse — éviter pré-fuites.",
    },
    {
        "id": "E10",
        "label": "GREENWASH",
        "axis": "ESG",
        "rule": "Pas de claims SAF/H2/électrique vagues sans LCA — pour audit complet renvoyer vers AeroSustainabilityComms.",
        "base": "EU Green Claims Directive 2024 + ReFuelEU Aviation.",
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Pack événementiel par salon (extrait)
# ────────────────────────────────────────────────────────────────────
PACKS = [
    {
        "event_id": "FAR-2026",
        "type": "Brief presse pré-salon",
        "title": "Brief presse — propulsion hybride régionale Farnborough",
        "extract": "EMBARGO 20/07/2026 06:00 BST. Notre démonstrateur de propulsion hybride pour aviation régionale sera présenté en vol le 21 juillet. Réduction CO2 de 70% démontrée sur le corridor London-Edinburgh (260 km). Premier client identifié : compagnie régionale UK (annonce conjointe lors du salon). Source autonomie : tests 2024 conditions ISA.",
        "ai": "OUI — rédaction par LLM",
        "verdicts": {"E01": "OK", "E02": "KO", "E03": "OK", "E04": "OK", "E05": "OK", "E06": "WARN", "E07": "OK", "E08": "WARN", "E09": "OK", "E10": "WARN"},
    },
    {
        "event_id": "ILA-2026",
        "type": "Post LinkedIn jour J",
        "title": "Post live ILA 2026 — partenariat industriel défense EU",
        "extract": "Heute auf der ILA 2026 — Wir kündigen unsere Partnerschaft mit dem deutschen Mittelstand für das Programm « European Sky Shield » an. Unser Système d'arme sol-air offre Reichweite 50 km + Engagement 12 cibles parallèles. EDIP funded 35%. Démonstration au Hall 4, Stand C-201. #ILA2026 #EuropeanDefence",
        "ai": "OUI — traduction et rédaction IA",
        "verdicts": {"E01": "WARN", "E02": "KO", "E03": "WARN", "E04": "OK", "E05": "OK", "E06": "WARN", "E07": "OK", "E08": "OK", "E09": "OK", "E10": "OK"},
    },
    {
        "event_id": "EUR-2026",
        "type": "Talking points VIP",
        "title": "Talking points dîner VIP Eurosatory — délégations gouvernementales",
        "extract": "Talking points pour rencontre déléguée Europe + 2 pays MENA. Points clés : (1) système anti-drone éprouvé sur le terrain ukrainien, (2) capacité de production 200 unités/an sur ligne LPM 2024-2030, (3) co-financement EDF (Eu Defence Fund) phase 2, (4) ouverts aux discussions de transferts de technologie pour pays partenaires. Mention sensibilité : éviter sujets RGE individuels.",
        "ai": "PARTIEL — synthèse benchmarks IA",
        "verdicts": {"E01": "KO", "E02": "WARN", "E03": "WARN", "E04": "OK", "E05": "OK", "E06": "WARN", "E07": "OK", "E08": "WARN", "E09": "OK", "E10": "OK"},
    },
    {
        "event_id": "MEB-2026",
        "type": "Brochure VIP + post Instagram",
        "title": "Pack MEBAA — brochure print VIP + Instagram story dual-language",
        "extract": "Brochure print + carrousel Instagram (EN/AR). « Discover our new business jet — range 6 500 nm, cabin 16 ft. Speed Mach 0.85. Sustainability built-in: SAF compatible up to 50% blend. Available for Q1 2027 delivery. Price on request. » Story Instagram avec démos vidéo ralenties + IA-generated cabin renders. Distribution VIP : 80 invités dont 12 high-net-worth de pays Moyen-Orient.",
        "ai": "OUI — renders cabine IA + traduction AR",
        "verdicts": {"E01": "OK", "E02": "KO", "E03": "OK", "E04": "OK", "E05": "OK", "E06": "WARN", "E07": "WARN", "E08": "OK", "E09": "OK", "E10": "WARN"},
    },
]

# ────────────────────────────────────────────────────────────────────
# Données : Redlines
# ────────────────────────────────────────────────────────────────────
REDLINES = [
    # FAR-2026
    {"case": "FAR-2026", "rule": "E02", "before": "[Aucune mention IA dans brief presse]", "after": "Footer brief : « Communiqué rédigé avec assistance IA — données validées par bureau d'études. UE 2024/1689 art. 50. »", "why": "AI Act art. 50 disclosure obligatoire applicable 02/08/2026"},
    {"case": "FAR-2026", "rule": "E06", "before": "Réduction CO2 de 70% démontrée sur le corridor London-Edinburgh", "after": "Réduction CO2 estimée à 70% (méthodologie ICAO CORSIA v2.3, 2024) sur le corridor London-Edinburgh (260 km), mode hybride. Hors phases production batterie.", "why": "Chiffre marketing sans périmètre LCA = trompeur"},
    {"case": "FAR-2026", "rule": "E08", "before": "[Brief uniquement EN]", "after": "Préparer une version FR du brief presse + traduction allemande pour audience continentale UE.", "why": "Farnborough = audience EU significative — multilingue recommandé"},
    {"case": "FAR-2026", "rule": "E10", "before": "(implicite mention transition énergétique)", "after": "Pour communication SAF/hybride/H2, renvoyer audit complet à AeroSustainabilityComms — éviter claims green non sourcés.", "why": "Greenwashing aviation = risque DGCCRF / EASA Decision 2024/015"},
    # ILA-2026
    {"case": "ILA-2026", "rule": "E01", "before": "Reichweite 50 km + Engagement 12 cibles parallèles", "after": "Capacité de portée 50 km avec engagement multi-cibles selon paramètres opérationnels du programme « European Sky Shield ».", "why": "Charte ASD : éviter chiffres tactiques précis dans communication grand public"},
    {"case": "ILA-2026", "rule": "E02", "before": "[Pas de mention IA dans le post LinkedIn]", "after": "Hashtag obligatoire en fin de post : #AIGenerated #EUAIAct ou note discrète : « Post avec assistance IA »", "why": "AI Act art. 50 — applicable aussi aux posts social"},
    {"case": "ILA-2026", "rule": "E03", "before": "Système d'arme sol-air offre Reichweite 50 km", "after": "[FLAG DefenseCommsGuard] Système d'arme sol-air = matériel de guerre LMG cat. 2 — diffusion publique des spécifications à valider compliance officer export.", "why": "Diffusion publique des specs d'un matériel de guerre = potentiellement non conforme L.2335-1"},
    {"case": "ILA-2026", "rule": "E06", "before": "EDIP funded 35%", "after": "EDIP funded 35% (Reg. UE 2025/588, conformément aux engagements EU Defence Industrial Strategy 2024)", "why": "Sourcer la référence du règlement EDIP"},
    # EUR-2026
    {"case": "EUR-2026", "rule": "E01", "before": "système anti-drone éprouvé sur le terrain ukrainien", "after": "système anti-drone éprouvé en conditions opérationnelles complexes (sans citation conflit en cours)", "why": "Charte ASD : éviter référence à conflit en cours pour des talking points VIP — risque géopolitique"},
    {"case": "EUR-2026", "rule": "E02", "before": "[Pas de mention IA dans talking points]", "after": "Mention en pied : « Synthèse benchmarks assistée IA — validation manuelle par dirigeant intervenant ».", "why": "AI Act applicable aux talking points même non publiés"},
    {"case": "EUR-2026", "rule": "E03", "before": "ouverts aux discussions de transferts de technologie pour pays partenaires", "after": "[FLAG DefenseCommsGuard] — Mention transferts de technologie = opération sous AT/IDP DGA. Reformuler : « Coopérations encadrées par les autorités françaises (DGA) selon les accords G2G applicables ».", "why": "Mention « transferts de technologie » sans encadrement compliance = risque export-control"},
    {"case": "EUR-2026", "rule": "E08", "before": "[Talking points en français uniquement]", "after": "Préparer version EN systématique + version arabe si délégations MENA confirmées.", "why": "Eurosatory accueille 100+ délégations — multilangue recommandé pour talking points VIP"},
    # MEB-2026
    {"case": "MEB-2026", "rule": "E02", "before": "[Pas de mention IA pour les renders cabine et la traduction AR]", "after": "Sticker discret sur les renders : « AI-generated visualization » — UE 2024/1689 art. 50 + AR translation watermark.", "why": "AI Act art. 50 art. visuels IA = disclosure obligatoire (deepfake/synthetic media)"},
    {"case": "MEB-2026", "rule": "E07", "before": "Story Instagram (sans mention horaire local)", "after": "Programmer story selon fuseau Asia/Dubai (GST UTC+4) — éviter publication 2-3am heure locale.", "why": "Multi-fuseau : programmation sociale doit respecter heure locale du salon"},
    {"case": "MEB-2026", "rule": "E10", "before": "Sustainability built-in: SAF compatible up to 50% blend", "after": "SAF compatibility up to 50% blend (HEFA certified ISCC EU, ReFuelEU Aviation Reg. 2023/2405). Lifecycle CO2eq study available on request.", "why": "Greenwashing potentiel : préciser type SAF + certification + LCA"},
    {"case": "MEB-2026", "rule": "E06", "before": "range 6 500 nm... Speed Mach 0.85", "after": "Range 6 500 nm (long-range cruise, ISA, max payload — réf. flight test FT-2024-Q3) — Speed Mach 0.85 (max operating cruise)", "why": "Chiffres techniques marketing sans contexte opérationnel = potentiellement trompeur"},
]

# ────────────────────────────────────────────────────────────────────
# Données : AI Act Disclosure modèles
# ────────────────────────────────────────────────────────────────────
AI_DISCLOSURE = [
    {"format": "Brief presse pré-salon", "fr": "Communiqué rédigé avec assistance IA — données techniques validées par bureau d'études. (UE 2024/1689 art. 50.)", "en": "Press release drafted with AI assistance — technical data validated by engineering office. (EU 2024/1689 art. 50.)"},
    {"format": "Post social (LinkedIn / X)", "fr": "Post composé avec assistance IA #AIGen", "en": "Post composed with AI assistance #AIGen #EU2024-1689"},
    {"format": "Story Instagram avec renders", "fr": "Visualisation générée par IA — réalité approximative — UE 2024/1689", "en": "AI-generated visualization — approximate render — EU 2024/1689"},
    {"format": "Talking points VIP (interne)", "fr": "Talking points avec synthèse IA — validation manuelle par intervenant. UE 2024/1689 art. 50.", "en": "Talking points with AI synthesis — manual validation by speaker. EU 2024/1689 art. 50."},
    {"format": "Brochure imprimée salon", "fr": "Document avec contenus partiellement générés par IA — UE 2024/1689 art. 50.", "en": "Document with content partially AI-generated — EU 2024/1689 art. 50."},
]

# ────────────────────────────────────────────────────────────────────
# Construction des onglets
# ────────────────────────────────────────────────────────────────────


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("0_README")
    set_widths(ws, {1: 26, 2: 92})
    write_title(ws, "AeroEventAI — Génération + audit packs événementiels salons aéro 2026", span=2)
    write_subtitle(ws, 2, "NEURAL — Branche Aéronautique / Marketing — Date de réf. : 25/04/2026", span=2)

    rows = [
        ("Mission", "Générer et auditer des packs événementiels (briefs presse, posts social, talking points VIP, brochures) pour les salons aéro/défense majeurs de Q2-Q4 2026, avec respect tonalité ASD + AI Act + multi-fuseau + multi-langue."),
        ("Calendrier salons", "4 salons réels Q2-Q4 2026 : ILA Berlin (10-14 juin), Eurosatory Paris (15-19 juin), Farnborough UK (20-24 juillet), MEBAA Dubai (8-10 décembre). Paris Air Show prochain en 2027 (biennal)."),
        ("Référentiel", "10 règles encodées (E01-E10) : ASD responsible comms, AI Act disclosure, export safe, no classified, source freshness, num validated, timezone, langue locale, embargo presse, anti-greenwashing."),
        ("Échelle verdict", "OK = conforme | WARN = ajustement recommandé | KO = bloquant, redline obligatoire avant diffusion."),
        ("Décision", "0 KO ET score ≥ 80% → DIFFUSION OK. Sinon : RELECTURE ou BLOQUÉ."),
        ("Inputs cas démo", "4 packs synthétiques inspirés du marché 2026 : brief presse FAR propulsion hybride, post LinkedIn ILA partenariat défense EU, talking points VIP Eurosatory, brochure VIP MEBAA business jets."),
        ("Outputs", "(1) Verdict matrice 10 × 4, (2) score conformité auto, (3) décision auto, (4) ~16 redlines concrètes, (5) 5 modèles disclosure IA FR + EN."),
        ("Cohérence cross-agents", "Sur ITAR/EAR → renvoi DefenseCommsGuard. Sur greenwashing → renvoi AeroSustainabilityComms. Cet agent se concentre sur tonalité + format + multi-langue + multi-fuseau."),
        ("Statut", "Démo recruteur — packs synthétiques. Aucun client réel ni programme classifié cité."),
        ("Disclaimer", "Outil indicatif. Ne remplace pas le directeur communication, le compliance officer export, ni le juriste défense. Toute diffusion finale doit être validée humainement."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 75 + 1))
        row += 1


def build_calendrier(wb: Workbook) -> None:
    ws = wb.create_sheet("1_Calendrier_Salons")
    set_widths(ws, {1: 12, 2: 38, 3: 22, 4: 30, 5: 24, 6: 50, 7: 22, 8: 22, 9: 50})
    write_title(ws, "Calendrier salons aéro/défense 2026 — Q2 à Q4", span=9)
    write_subtitle(ws, 2, "4 salons majeurs avec organisateur, fuseau, langues, tonalité requise.", span=9)

    headers = ["ID", "Salon", "Dates 2026", "Lieu", "Organisateur", "Publics cibles", "Fuseau", "Langues", "Tonalité"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 9)

    row = 5
    for i, e in enumerate(EVENTS):
        ws.cell(row=row, column=1, value=e["id"])
        ws.cell(row=row, column=2, value=e["name"])
        ws.cell(row=row, column=3, value=e["dates"])
        ws.cell(row=row, column=4, value=e["lieu"])
        ws.cell(row=row, column=5, value=e["organisateur"])
        ws.cell(row=row, column=6, value=e["publics"])
        ws.cell(row=row, column=7, value=e["fuseau"])
        ws.cell(row=row, column=8, value=e["languages"])
        ws.cell(row=row, column=9, value=e["tone"])
        style_body_row(ws, row, 9, alt=(i % 2 == 1), height=80)
        row += 1


def build_referentiel(wb: Workbook) -> None:
    ws = wb.create_sheet("2_Referentiel")
    set_widths(ws, {1: 8, 2: 22, 3: 16, 4: 60, 5: 45})
    write_title(ws, "Référentiel — 10 règles audit pack événementiel", span=5)
    write_subtitle(ws, 2, "Tonalité ASD + AI Act + Export safe + Multi-fuseau/langue + Embargo presse + ESG.", span=5)

    headers = ["ID", "Label", "Axe", "Règle", "Base / réf."]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        ws.cell(row=row, column=3, value=rule["axis"])
        ws.cell(row=row, column=4, value=rule["rule"])
        ws.cell(row=row, column=5, value=rule["base"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=70)
        row += 1


def build_packs(wb: Workbook) -> None:
    ws = wb.create_sheet("3_Packs_Evenementiels")
    set_widths(ws, {1: 12, 2: 22, 3: 38, 4: 18, 5: 80})
    write_title(ws, "Packs événementiels — 4 cas démo synthétiques", span=5)
    write_subtitle(ws, 2, "Un cas par salon. Anonymisés mais inspirés du marché 2026.", span=5)

    headers = ["Salon", "Type pack", "Titre", "IA utilisée", "Extrait"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, p in enumerate(PACKS):
        ws.cell(row=row, column=1, value=p["event_id"])
        ws.cell(row=row, column=2, value=p["type"])
        ws.cell(row=row, column=3, value=p["title"])
        ws.cell(row=row, column=4, value=p["ai"])
        ws.cell(row=row, column=5, value=p["extract"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=140)
        row += 1


def build_verdict(wb: Workbook) -> None:
    ws = wb.create_sheet("4_Verdict_Par_Pack")
    n_packs = len(PACKS)
    n_rules = len(RULES)
    cols_total = 2 + n_packs

    widths = {1: 8, 2: 22}
    for i in range(n_packs):
        widths[3 + i] = 14
    set_widths(ws, widths)

    write_title(ws, f"Matrice verdicts — {n_rules} règles × {n_packs} packs salons", span=cols_total)
    write_subtitle(ws, 2, "OK = conforme | WARN = ajustement recommandé | KO = bloquant. Score + décision en bas.", span=cols_total)

    ws.cell(row=4, column=1, value="ID")
    ws.cell(row=4, column=2, value="Règle")
    for i, p in enumerate(PACKS):
        ws.cell(row=4, column=3 + i, value=p["event_id"])
    style_header_row(ws, 4, cols_total)

    row = 5
    for r_idx, rule in enumerate(RULES):
        ws.cell(row=row, column=1, value=rule["id"])
        ws.cell(row=row, column=2, value=rule["label"])
        for c_idx, p in enumerate(PACKS):
            v = p["verdicts"].get(rule["id"], "")
            cell = ws.cell(row=row, column=3 + c_idx, value=v)
            cell.font = body_font(bold=True, color=verdict_color(v))
            cell.fill = verdict_fill(v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).font = body_font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).font = body_font()
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    write_section_header(ws, row, "Synthèse par pack", span=cols_total)
    row += 1

    last_verdict_row = 4 + n_rules

    for label, count_val in [("Nb OK", "OK"), ("Nb WARN", "WARN"), ("Nb KO", "KO")]:
        ws.cell(row=row, column=2, value=label)
        for c_idx in range(n_packs):
            col_letter = get_column_letter(3 + c_idx)
            ws.cell(row=row, column=3 + c_idx, value=f'=COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"{count_val}")')
        style_body_row(ws, row, cols_total, height=22)
        ws.cell(row=row, column=2).font = body_font(bold=True)
        row += 1

    ws.cell(row=row, column=2, value="Score conformité")
    for c_idx in range(n_packs):
        col_letter = get_column_letter(3 + c_idx)
        formula = f'=(COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN"))/{n_rules}'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.number_format = "0%"
    style_body_row(ws, row, cols_total, height=24)
    ws.cell(row=row, column=2).font = body_font(bold=True)
    row += 1

    ws.cell(row=row, column=2, value="Décision")
    for c_idx in range(n_packs):
        col_letter = get_column_letter(3 + c_idx)
        ko_count = f'COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"KO")'
        score = f'(COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"OK")+0.5*COUNTIF({col_letter}5:{col_letter}{last_verdict_row},"WARN"))/{n_rules}'
        formula = f'=IF({ko_count}>0,"BLOQUÉ",IF({score}>=0.8,"DIFFUSION OK","RELECTURE"))'
        cell = ws.cell(row=row, column=3 + c_idx, value=formula)
        cell.font = body_font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    style_body_row(ws, row, cols_total, height=26)
    ws.cell(row=row, column=2).font = body_font(bold=True)


def build_redlines(wb: Workbook) -> None:
    ws = wb.create_sheet("5_Redlines")
    set_widths(ws, {1: 12, 2: 8, 3: 50, 4: 65, 5: 45})
    write_title(ws, "Redlines — Suggestions de correction concrètes", span=5)
    write_subtitle(ws, 2, "Pour chaque KO/WARN, redline applicable avec base citée.", span=5)

    headers = ["Salon", "Règle", "Avant (extrait)", "Après (suggestion)", "Pourquoi"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 5)

    row = 5
    for i, rl in enumerate(REDLINES):
        ws.cell(row=row, column=1, value=rl["case"])
        ws.cell(row=row, column=2, value=rl["rule"])
        ws.cell(row=row, column=3, value=rl["before"])
        ws.cell(row=row, column=4, value=rl["after"])
        ws.cell(row=row, column=5, value=rl["why"])
        style_body_row(ws, row, 5, alt=(i % 2 == 1), height=75)
        row += 1


def build_ai_disclosure(wb: Workbook) -> None:
    ws = wb.create_sheet("6_AI_Act_Disclosure")
    set_widths(ws, {1: 28, 2: 60, 3: 60})
    write_title(ws, "AI Act art. 50 — Modèles de mention IA pour packs salons", span=3)
    write_subtitle(ws, 2, "Applicable 2 août 2026. Tous formats événementiels concernés.", span=3)

    headers = ["Format", "Mention FR", "Mention EN"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 3)

    row = 5
    for i, d in enumerate(AI_DISCLOSURE):
        ws.cell(row=row, column=1, value=d["format"])
        ws.cell(row=row, column=2, value=d["fr"])
        ws.cell(row=row, column=3, value=d["en"])
        style_body_row(ws, row, 3, alt=(i % 2 == 1), height=80)
        row += 1


def build_limites(wb: Workbook) -> None:
    ws = wb.create_sheet("7_Limites")
    set_widths(ws, {1: 30, 2: 80})
    write_title(ws, "Truth layer — Ce que AeroEventAI NE remplace PAS", span=2)
    write_subtitle(ws, 2, "Disclaimer obligatoire. Sujet : pack événementiel salons sensibles.", span=2)

    rows = [
        ("Directeur communication", "L'agent ne remplace pas le directeur communication / DirCom. Toute diffusion presse, social, brochure doit être validée par le DirCom."),
        ("Compliance officer export", "Pour toute information ITAR/EAR/dual-use → validation par DefenseCommsGuard + compliance officer obligatoire avant diffusion publique au salon."),
        ("Juriste défense", "Toute mention programme défense / cooperation agreement / G2G doit être validée par juriste défense."),
        ("Bureau d'études", "Tout chiffre technique (range, vitesse, MTBF, autonomie) doit être confirmé par bureau d'études signataire."),
        ("DGA/SGDSN — Classification", "L'agent ne classifie pas les contenus. Toute mention sensible doit être pré-validée par officier de sécurité."),
        ("Anti-greenwashing", "Pour tout claim SAF/H2/électrique/CO2 → renvoyer audit complet à AeroSustainabilityComms avant diffusion."),
        ("ASD Charter", "Charte « responsible defence comms » est volontaire (non opposable juridiquement). L'agent applique ses principes, mais l'arbitrage final reste au DirCom."),
        ("Multi-fuseau / multi-langue", "L'agent suggère des bonnes pratiques. La traduction finale (DE, FR, AR, EN) doit être validée par traducteur professionnel — pas par IA seule pour les contenus sensibles."),
        ("Embargo presse", "L'agent applique les conventions standard (date + heure + fuseau). Toute relation presse spécifique (exclusivité, embargo strict) reste de la responsabilité du DirCom."),
        ("Date de fraîcheur", "Calendrier 2026 figé au 25/04/2026. À rafraîchir avant tout planning événementiel — annulations / reports possibles."),
        ("Limite v1", "Pas de génération automatique de visuels/photos. Pas de scheduling automatique sur LinkedIn/X. v2 prévue avec connecteur outils MarTech."),
    ]

    row = 4
    for label, value in rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = body_font(bold=True)
        c1.fill = alt_fill()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = THIN_BORDER
        c2.font = body_font()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = max(40, 14 * (len(value) // 70 + 1))
        row += 1


def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb)
    build_calendrier(wb)
    build_referentiel(wb)
    build_packs(wb)
    build_verdict(wb)
    build_redlines(wb)
    build_ai_disclosure(wb)
    build_limites(wb)

    out = "AeroEventAI_NEURAL.xlsx"
    wb.save(out)
    print(f"OK — {out} written ({len(EVENTS)} salons, {len(PACKS)} packs, {len(RULES)} règles, {len(REDLINES)} redlines)")


if __name__ == "__main__":
    build()
