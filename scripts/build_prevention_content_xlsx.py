"""
Build PreventionContent_NEURAL.xlsx — démo Excel pour l'agent NEURAL
'PreventionContent' (branche Assurances — Marketing).

7 onglets : README, Contenus Source (3 domaines), ClaimGuard Audit,
SensitiveDataGuard RGPD, Contenus Optimisés, RGPD Compliance, Score Global.

Domaines : Prévention routière, Prévention santé, Prévention habitation.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# Branding NEURAL
# ────────────────────────────────────────────────────────────────────
NEURAL_VIOLET = "7C3AED"
NEURAL_INK = "0E0824"
GRAY_HEADER = "1E1242"
GRAY_LIGHT = "EAE6F5"
GREEN_OK = "16A34A"
AMBER_WARN = "D97706"
RED_NO = "DC2626"

FONT_FAMILY = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def title_font(size: int = 18, color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=True, color=color)


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=True, color=color)


def body_font(size: int = 10, bold: bool = False, color: str = "0E0824") -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color)


def title_fill() -> PatternFill:
    return PatternFill("solid", start_color=NEURAL_VIOLET, end_color=NEURAL_VIOLET)


def header_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_HEADER, end_color=GRAY_HEADER)


def alt_fill() -> PatternFill:
    return PatternFill("solid", start_color=GRAY_LIGHT, end_color=GRAY_LIGHT)


def green_fill() -> PatternFill:
    return PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")


def amber_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")


def red_fill() -> PatternFill:
    return PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")


def orange_fill() -> PatternFill:
    return PatternFill("solid", start_color="FFEDD5", end_color="FFEDD5")


def domain_fill(domain: str) -> PatternFill:
    colors = {"route": "1D4ED8", "sante": "0F766E", "habitation": "7C2D12"}
    c = colors.get(domain, NEURAL_VIOLET)
    return PatternFill("solid", start_color=c, end_color=c)


def write_title(ws, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36


def write_subtitle(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = body_font(size=10, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 22


def write_section_header(ws, row: int, text: str, span: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_body_row(ws, row: int, ncols: int, alt: bool = False, height: int = 70) -> None:
    fill = alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row].height = height


def set_widths(ws, widths: dict[int, float]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def verdict_style(cell, verdict: str) -> None:
    if verdict == "OK":
        cell.fill = green_fill()
        cell.font = body_font(bold=True, color=GREEN_OK)
    elif verdict == "WARN":
        cell.fill = amber_fill()
        cell.font = body_font(bold=True, color=AMBER_WARN)
    elif verdict == "KO":
        cell.fill = red_fill()
        cell.font = body_font(bold=True, color=RED_NO)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


# ────────────────────────────────────────────────────────────────────
# 3 contenus prévention sources (non optimisés)
# ────────────────────────────────────────────────────────────────────
CONTENUS_SOURCE = [
    {
        "id": "PREV-01",
        "domaine": "Prévention routière",
        "domain_key": "route",
        "format": "Guide numérique + boîtier télématique",
        "cible": "Assurés auto < 30 ans avec formule télématique YoungDriver",
        "titre": "10 règles pour réduire vos risques d'accident — et votre prime",
        "contenu": (
            "Votre boîtier YoungDriver enregistre votre style de conduite 24h/24 : vitesse, freinage brusque, "
            "accélérations, usage du téléphone au volant, horaires de nuit. Ces données sont analysées chaque "
            "semaine pour calculer votre score conducteur.\n\n"
            "Règle 1 : Ne conduisez jamais avec 0,5 g/L d'alcool ou plus — tout sinistre survenu sous influence "
            "sera rejeté et vous expose à des poursuites pénales.\n\n"
            "Règle 4 : Évitez les créneaux 22h-6h — votre prime augmente automatiquement si votre score nuit "
            "dépasse 20% de vos kilomètres totaux.\n\n"
            "Règle 7 : En cas d'accident, déclarez-le sous 5 jours ouvrés maximum, faute de quoi la garantie "
            "est susceptible d'être refusée.\n\n"
            "Règle 10 : Si votre score chute en dessous de 40/100 pendant 3 mois consécutifs, votre contrat "
            "sera réévalué à l'échéance."
        ),
        "source_donnees": "Boîtier télématique GPS + accéléromètre + gyroscope, transmission temps réel",
        "destinataire": "Email hebdomadaire + push notification appli mobile",
    },
    {
        "id": "PREV-02",
        "domaine": "Prévention santé",
        "domain_key": "sante",
        "format": "Newsletter mensuelle + questionnaire santé en ligne",
        "cible": "Adhérents mutuelle Senior Confort (60-70 ans)",
        "titre": "Votre bilan santé annuel : les examens à ne pas négliger",
        "contenu": (
            "Pour optimiser vos remboursements et maintenir votre couverture Senior Confort, complétez votre "
            "questionnaire de santé annuel avant le 31 mars. Ce questionnaire est obligatoire pour le maintien "
            "de votre formule dentaire Premium.\n\n"
            "Renseignez vos antécédents médicaux, traitements en cours, hospitalisations des 24 derniers mois, "
            "consommation de tabac, d'alcool et score IMC.\n\n"
            "Important : toute fausse déclaration peut entraîner la nullité du contrat (art. L.113-8 Code "
            "des assurances). Vos données de santé sont conservées 10 ans.\n\n"
            "Rappel : si vous n'avez pas déclaré une maladie chronique lors de la souscription, les soins liés "
            "à cette pathologie pourraient ne pas être remboursés.\n\n"
            "Dépistages recommandés cette année : coloscopie (si > 60 ans), mammographie, glycémie à jeun, "
            "bilan ophtalmologique complet."
        ),
        "source_donnees": "Questionnaire médical en ligne, données DSN employeur (si contrat collectif), historique soins",
        "destinataire": "Email mensuel + espace adhérent",
    },
    {
        "id": "PREV-03",
        "domaine": "Prévention habitation",
        "domain_key": "habitation",
        "format": "Guide PDF + proposition visite diagnostic",
        "cible": "Assurés MRH propriétaires > 5 ans de contrat",
        "titre": "Protégez votre logement : les bons réflexes avant l'été",
        "contenu": (
            "Votre contrat MRH Premium inclut une visite de prévention gratuite. Notre expert évalue les "
            "risques de votre logement et transmet un rapport d'évaluation à votre dossier.\n\n"
            "Avant la visite, transmettez-nous : plan du logement, factures d'entretien chaudière des 3 "
            "dernières années, photos des entrées, références de votre système d'alarme et de vos serrures "
            "certifiées A2P.\n\n"
            "Attention : en cas de sinistre vol ou incendie, si votre système de sécurité n'est pas conforme "
            "aux normes A2P mentionnées lors de la souscription, l'indemnisation pourra être réduite "
            "proportionnellement.\n\n"
            "Bon réflexe 3 : installez un détecteur de fumée NF dans chaque pièce — obligatoire depuis 2015 "
            "(DAAF). Son absence peut être invoquée en cas de sinistre incendie.\n\n"
            "Bon réflexe 6 : photographiez votre mobilier et objets de valeur et déposez les fichiers dans "
            "votre coffre-fort numérique NEURAL Vault — accessible à notre équipe sinistre en cas de besoin."
        ),
        "source_donnees": "Rapport visite expert, photos logement, données système alarme, historique sinistres",
        "destinataire": "Courrier + email + espace client",
    },
]

# ────────────────────────────────────────────────────────────────────
# ClaimGuard — détection patterns claim avoidance
# ────────────────────────────────────────────────────────────────────
CLAIMGUARD_FINDINGS = [
    # PREV-01 route
    {
        "prev_id": "PREV-01",
        "finding_id": "CG-01",
        "domaine": "Route",
        "pattern": "Menace de rejet a priori",
        "extrait": "tout sinistre survenu sous influence sera rejeté",
        "probleme": "Formulation préventive de rejet sans procédure contradictoire. Suggère l'impossibilité de contester. Contraire à L.113-9 (nullité ≠ refus automatique) et au droit à la réclamation ACPR.",
        "verdict": "KO",
        "correction": "Remplacer par : 'La conduite sous influence constitue une cause d'exclusion prévue au contrat. Tout sinistre fait l'objet d'une instruction individualisée.'",
        "ref_legale": "Code des assurances L.113-9, DDA art. 20, ACPR Reco 2020-R-02",
    },
    {
        "prev_id": "PREV-01",
        "finding_id": "CG-02",
        "domaine": "Route",
        "pattern": "Délai déclaration court + sanction implicite",
        "extrait": "déclarez-le sous 5 jours ouvrés maximum, faute de quoi la garantie est susceptible d'être refusée",
        "probleme": "Le délai légal est 5 jours ouvrés (L.113-2) mais la formulation 'susceptible d'être refusée' est intimidante et inexacte : le dépassement de délai n'entraîne refus que si préjudice prouvé pour l'assureur.",
        "verdict": "WARN",
        "correction": "Remplacer par : 'Déclarez votre sinistre dans les 5 jours ouvrés (délai légal). En cas de dépassement justifié, contactez votre conseiller — le délai peut être apprécié selon les circonstances.'",
        "ref_legale": "Code des assurances L.113-2 al. 4, jurisprudence Cass. 1re civ. 2019",
    },
    {
        "prev_id": "PREV-01",
        "finding_id": "CG-03",
        "domaine": "Route",
        "pattern": "Tarification punitive présentée comme règle de prévention",
        "extrait": "votre prime augmente automatiquement si votre score nuit dépasse 20%",
        "probleme": "Le contenu mélange prévention et tarification conditionnelle. Présenter une hausse tarifaire automatique dans un guide de prévention est assimilable à une menace commerciale déguisée en conseil de sécurité.",
        "verdict": "WARN",
        "correction": "Séparer clairement les rubriques 'Conseils de sécurité' et 'Fonctionnement de votre prime télématique'. Le contenu de prévention ne doit pas inclure de menaces tarifaires.",
        "ref_legale": "DDA art. 20 (information claire, non trompeuse), RGPD art. 22 (décision automatisée)",
    },
    # PREV-02 santé
    {
        "prev_id": "PREV-02",
        "finding_id": "CG-04",
        "domaine": "Santé",
        "pattern": "Questionnaire présenté comme condition de maintien de garantie",
        "extrait": "Ce questionnaire est obligatoire pour le maintien de votre formule dentaire Premium",
        "probleme": "Lie le maintien de la garantie à un questionnaire de santé post-souscription. Illégal : l'assureur ne peut modifier les garanties en cours de contrat sans accord de l'assuré (L.112-3). Risque de pratique de sélection médicale déguisée.",
        "verdict": "KO",
        "correction": "Supprimer le lien questionnaire/maintien garantie. Reformuler : 'Pour personnaliser vos conseils de prévention, vous pouvez compléter ce questionnaire de santé — son remplissage est facultatif et n'affecte pas vos garanties.'",
        "ref_legale": "Code des assurances L.112-3, L.113-15, CNIL délibération 2022-118 (données santé assurance)",
    },
    {
        "prev_id": "PREV-02",
        "finding_id": "CG-05",
        "domaine": "Santé",
        "pattern": "Fausse déclaration — formulé comme menace non nuancée",
        "extrait": "toute fausse déclaration peut entraîner la nullité du contrat (art. L.113-8)",
        "probleme": "L'article L.113-8 s'applique aux fausses déclarations INTENTIONNELLES. Citer cet article dans un guide de prévention sans cette nuance peut décourager un assuré de se souvenir d'une pathologie ancienne, par peur de nullité.",
        "verdict": "WARN",
        "correction": "Ajouter la nuance : 'L'art. L.113-8 sanctionne les fausses déclarations intentionnelles. En cas d'oubli ou d'imprécision non intentionnel, contactez votre conseiller pour une régularisation sans pénalité.'",
        "ref_legale": "Code des assurances L.113-8 (mauvaise foi requise), L.113-9 (omission non intentionnelle)",
    },
    {
        "prev_id": "PREV-02",
        "finding_id": "CG-06",
        "domaine": "Santé",
        "pattern": "Exclusion implicite pour pathologie non déclarée",
        "extrait": "les soins liés à cette pathologie pourraient ne pas être remboursés",
        "probleme": "Formulé comme avertissement préventif, ce passage suggère une exclusion anticipée que l'assureur se réserve de décider unilatéralement en cas de sinistre. Crée un effet dissuasif sur la déclaration de sinistre.",
        "verdict": "KO",
        "correction": "Supprimer. Remplacer par une procédure positive : 'Si vous pensez avoir omis une information à la souscription, contactez-nous pour une révision de votre contrat — cela ne génère pas de suspension de garantie.'",
        "ref_legale": "Code des assurances L.112-3, DDA art. 20 §3, ACPR Reco 2020-R-02 (relation client)",
    },
    # PREV-03 habitation
    {
        "prev_id": "PREV-03",
        "finding_id": "CG-07",
        "domaine": "Habitation",
        "pattern": "Réduction d'indemnisation implicite présentée comme conseil",
        "extrait": "l'indemnisation pourra être réduite proportionnellement",
        "probleme": "Présenter une réduction d'indemnisation dans un guide de prévention sans préciser la procédure contradictoire équivaut à une menace déguisée. La règle proportionnelle (L.113-9) ne s'applique qu'en cas de sous-assurance prouvée, pas d'absence de certification A2P.",
        "verdict": "KO",
        "correction": "Remplacer par : 'Pour bénéficier pleinement de vos garanties vol, nous vous recommandons des serrures certifiées A2P. En cas de sinistre, l'instruction examinera les circonstances — contactez votre conseiller si vous avez des doutes sur votre niveau de sécurité.'",
        "ref_legale": "Code des assurances L.113-9 (règle proportionnelle), L.124-3 (obligation d'information sinistre)",
    },
    {
        "prev_id": "PREV-03",
        "finding_id": "CG-08",
        "domaine": "Habitation",
        "pattern": "DAAF — obligation légale présentée comme levier d'exclusion",
        "extrait": "Son absence peut être invoquée en cas de sinistre incendie",
        "probleme": "La loi DAAF (2010) impose l'installation mais ne prévoit pas d'exclusion contractuelle automatique liée à l'absence. Formuler 'peut être invoquée' crée une menace d'exclusion non fondée en droit.",
        "verdict": "WARN",
        "correction": "Remplacer par : 'Le détecteur de fumée DAAF est obligatoire (loi du 9 mars 2010). Son installation protège votre famille — elle n'est pas une condition contractuelle d'indemnisation mais une obligation civique.'",
        "ref_legale": "Loi n°2010-238 du 9 mars 2010 (DAAF), Code des assurances — absence de clause d'exclusion légale DAAF",
    },
    {
        "prev_id": "PREV-03",
        "finding_id": "CG-09",
        "domaine": "Habitation",
        "pattern": "Accès assureur au coffre-fort numérique assuré",
        "extrait": "déposez les fichiers dans votre coffre-fort numérique NEURAL Vault — accessible à notre équipe sinistre",
        "probleme": "L'accès d'un tiers (assureur) aux données stockées par l'assuré constitue un traitement de données personnelles soumis au consentement RGPD. Formuler cela comme un conseil de prévention sans cadre de consentement est problématique.",
        "verdict": "KO",
        "correction": "Reformuler avec base légale explicite : 'Stockez vos photos dans votre espace client sécurisé. En cas de sinistre, vous pourrez partager ces éléments avec votre gestionnaire sinistre — le partage est à votre initiative et soumis à votre consentement.'",
        "ref_legale": "RGPD art. 6 §1a (consentement), art. 7 (conditions consentement), art. 9 (données photographiques si biométriques)",
    },
]

# ────────────────────────────────────────────────────────────────────
# SensitiveDataGuard — audit RGPD données sensibles
# ────────────────────────────────────────────────────────────────────
SENSITIVE_DATA_FINDINGS = [
    {
        "prev_id": "PREV-01",
        "finding_id": "SD-01",
        "domaine": "Route",
        "categorie_donnee": "Données comportementales (télématique)",
        "description": "Vitesse, freinage, accélérations, géolocalisation GPS, horaires conduite — collecte continue 24h/24",
        "article_rgpd": "Art. 6 §1b (contrat) — discutable, art. 22 (décision automatisée)",
        "base_legale_actuelle": "Consentement implicite via souscription formule télématique",
        "probleme": "Le consentement doit être explicite, spécifique et révocable (art. 7). La souscription groupée consentement+contrat ne suffit pas. La géolocalisation continue peut relever de données de déplacement (sensibles selon EDPB).",
        "verdict": "WARN",
        "action_requise": "Recueil consentement séparé et granulaire. Possibilité de révoquer sans résiliation du contrat. Information sur durée de conservation et droit à la portabilité (art. 20).",
        "duree_conservation": "Non précisée dans le contenu source",
        "verdict_duree": "KO",
        "droit_acces": "Non mentionné",
        "verdict_droits": "KO",
    },
    {
        "prev_id": "PREV-01",
        "finding_id": "SD-02",
        "domaine": "Route",
        "categorie_donnee": "Comportement substance — alcool/drogues",
        "description": "Inférence comportementale basée sur les données de conduite (horaires, type de route, freinage d'urgence)",
        "article_rgpd": "Art. 9 §1 — données révélant la santé ou les habitudes de consommation peuvent être catégorisées données sensibles",
        "base_legale_actuelle": "Non précisée — inclus dans le scoring général",
        "probleme": "L'inférence d'alcoolémie ou de conduite sous influence à partir de données télématiques peut constituer un traitement de données relatives à la santé (art. 9). Requiert consentement explicite ou intérêt vital.",
        "verdict": "KO",
        "action_requise": "Ne pas inférer de comportements liés à la santé à partir de données télémétriques sans consentement explicite art. 9. Anonymiser les inférences dans le scoring si impossible à isoler.",
        "duree_conservation": "Non précisée",
        "verdict_duree": "KO",
        "droit_acces": "Non mentionné",
        "verdict_droits": "KO",
    },
    {
        "prev_id": "PREV-02",
        "finding_id": "SD-03",
        "domaine": "Santé",
        "categorie_donnee": "Données de santé — questionnaire médical",
        "description": "Antécédents médicaux, traitements, hospitalisations, IMC, tabac, alcool — collecte via formulaire en ligne",
        "article_rgpd": "Art. 9 §1 — données de santé = catégorie spéciale — interdiction de principe",
        "base_legale_actuelle": "Consentement implicite non formalisé",
        "probleme": "Les données de santé (art. 9) requièrent un consentement EXPLICITE, SPÉCIFIQUE et DOCUMENTÉ. Un questionnaire intégré à une newsletter sans consentement séparé est illégal. Conservation 10 ans non justifiée.",
        "verdict": "KO",
        "action_requise": "Case à cocher distincte pour le consentement données santé. Séparation des finalités (prévention vs tarification). Durée conservation justifiée (recommandation CNIL : 2 ans pour données prévention santé). Droit d'effacement art. 17.",
        "duree_conservation": "10 ans (mentionné dans le contenu source)",
        "verdict_duree": "KO",
        "droit_acces": "Non mentionné",
        "verdict_droits": "KO",
    },
    {
        "prev_id": "PREV-02",
        "finding_id": "SD-04",
        "domaine": "Santé",
        "categorie_donnee": "Données addictions (tabac, alcool, IMC)",
        "description": "Consommation de tabac, d'alcool et IMC recueillis dans le questionnaire annuel",
        "article_rgpd": "Art. 9 — révèle l'état de santé et potentiellement des données relatives aux addictions",
        "base_legale_actuelle": "Consentement non formalisé — inclus dans questionnaire général",
        "probleme": "Tabac et alcool dans un contexte d'assurance santé sont des données pouvant révéler des pathologies (dépendances). Traitement séparé requis avec base légale explicite. Risque de profilage discriminatoire si utilisé pour tarification.",
        "verdict": "WARN",
        "action_requise": "Isoler ces données dans un module de prévention clairement distinct de la tarification. Informer l'assuré que ces données ne seront pas utilisées pour modifier ses garanties ou sa prime.",
        "duree_conservation": "10 ans (mélangé avec données médicales générales)",
        "verdict_duree": "WARN",
        "droit_acces": "Non mentionné",
        "verdict_droits": "KO",
    },
    {
        "prev_id": "PREV-03",
        "finding_id": "SD-05",
        "domaine": "Habitation",
        "categorie_donnee": "Données domicile — plan, photos, système alarme",
        "description": "Plan du logement, photos entrées et pièces, références alarme et serrures — transmis dans le cadre de la visite diagnostic",
        "article_rgpd": "Art. 6 §1b (contrat) — acceptable si nécessaire — mais photos domicile peuvent inclure données personnelles tiers",
        "base_legale_actuelle": "Intérêt légitime ou exécution contrat",
        "probleme": "Les photos de domicile peuvent capturer des membres de la famille, visiteurs, documents personnels. Base légale 'intérêt légitime' insuffisante si d'autres personnes sont visibles. Transmission à 'équipe sinistre' sans limitation de destinataires.",
        "verdict": "WARN",
        "action_requise": "Informer sur les destinataires précis des photos. Demander de ne pas photographier des personnes. Définir une durée de conservation spécifique pour les photos (recommandation : durée du contrat + 2 ans).",
        "duree_conservation": "Non précisée",
        "verdict_duree": "KO",
        "droit_acces": "Non mentionné",
        "verdict_droits": "KO",
    },
    {
        "prev_id": "PREV-03",
        "finding_id": "SD-06",
        "domaine": "Habitation",
        "categorie_donnee": "Accès tiers au coffre-fort numérique",
        "description": "Fichiers stockés par l'assuré dans NEURAL Vault — accès accordé à l'équipe sinistre selon le contenu source",
        "article_rgpd": "Art. 6 §1a (consentement), art. 28 (sous-traitant si NEURAL Vault ≠ assureur), art. 32 (sécurité)",
        "base_legale_actuelle": "Non précisée — accès présenté comme allant de soi",
        "probleme": "L'accès d'un tiers aux fichiers d'un coffre-fort numérique est un traitement de données personnelles. Consentement explicite requis pour chaque accès. Si NEURAL Vault est un sous-traitant, DPA (Data Processing Agreement) obligatoire.",
        "verdict": "KO",
        "action_requise": "Recueillir consentement explicite avant tout accès. Mettre en place un DPA si Vault ≠ assureur. Journal d'accès obligatoire (art. 30). Droit de retrait d'accès sans impact sur les garanties.",
        "duree_conservation": "Non précisée",
        "verdict_duree": "KO",
        "droit_acces": "Non mentionné",
        "verdict_droits": "KO",
    },
]

# ────────────────────────────────────────────────────────────────────
# Contenus optimisés post-audit
# ────────────────────────────────────────────────────────────────────
CONTENUS_OPTIMISES = [
    {
        "prev_id": "PREV-01",
        "domaine": "Prévention routière",
        "domain_key": "route",
        "titre_opt": "10 règles pour conduire plus sereinement — et mieux protéger votre prime",
        "contenu_opt": (
            "Votre boîtier YoungDriver analyse votre style de conduite pour vous aider à progresser. "
            "Vous avez accepté cette collecte lors de la souscription — vous pouvez en modifier les paramètres "
            "à tout moment depuis votre espace client (rubrique Mes données).\n\n"
            "Règle 1 — Zéro alcool au volant : la conduite sous influence est une cause d'exclusion prévue "
            "au contrat. Tout sinistre fait l'objet d'une instruction individualisée avec droit à la "
            "contestation.\n\n"
            "Règle 4 — Conduite de nuit : les créneaux 22h-6h sont plus accidentogènes. Votre score conducteur "
            "intègre ce paramètre — consultez l'explication détaillée de votre score dans l'appli.\n\n"
            "Règle 7 — En cas d'accident : déclarez votre sinistre dans les 5 jours ouvrés (délai légal). "
            "En cas de circonstances exceptionnelles, contactez votre conseiller — le délai est apprécié "
            "au cas par cas.\n\n"
            "Règle 10 — Suivi de votre score : si votre score évolue, nous vous contacterons pour en discuter "
            "ensemble à l'échéance de votre contrat."
        ),
        "corrections_appliquees": "CG-01 (rejet a priori → instruction individualisée), CG-02 (délai → apprécié cas par cas), CG-03 (prime séparée du conseil prévention)",
        "rgpd_corrections": "SD-01 (consentement révocable mentionné), SD-02 (inférence alcool supprimée)",
        "score_claimguard_avant": 35,
        "score_claimguard_apres": 88,
        "score_rgpd_avant": 25,
        "score_rgpd_apres": 72,
    },
    {
        "prev_id": "PREV-02",
        "domaine": "Prévention santé",
        "domain_key": "sante",
        "titre_opt": "Votre bilan santé annuel : les dépistages qui font la différence",
        "contenu_opt": (
            "Pour personaliser vos conseils de prévention, vous pouvez compléter un questionnaire de santé "
            "— son remplissage est entièrement facultatif et n'affecte en aucun cas vos remboursements ni "
            "vos garanties. (RGPD art. 9 — données de santé : votre consentement explicite vous sera demandé "
            "avant d'accéder au questionnaire.)\n\n"
            "Dépistages recommandés cette année :\n"
            "• Coloscopie si > 60 ans (Programme national de dépistage)\n"
            "• Mammographie bisannuelle (dépistage organisé)\n"
            "• Glycémie à jeun (risque diabète type 2)\n"
            "• Bilan ophtalmologique complet (glaucome, DMLA)\n\n"
            "Ces recommandations sont issues des guidelines HAS 2024 et sont indépendantes de votre contrat "
            "mutuelle. Elles ne constituent pas un diagnostic médical.\n\n"
            "Vos données de prévention sont conservées 2 ans et ne sont pas utilisées pour modifier "
            "vos garanties ou votre prime. Vous disposez d'un droit d'accès, de rectification et "
            "d'effacement (art. 17 RGPD) — exercez-le depuis votre espace adhérent."
        ),
        "corrections_appliquees": "CG-04 (questionnaire optionnel), CG-05 (nuance fausse déclaration), CG-06 (exclusion implicite supprimée)",
        "rgpd_corrections": "SD-03 (consentement explicite requis), SD-04 (addictions isolées de tarification), durée conservation réduite à 2 ans",
        "score_claimguard_avant": 28,
        "score_claimguard_apres": 91,
        "score_rgpd_avant": 15,
        "score_rgpd_apres": 78,
    },
    {
        "prev_id": "PREV-03",
        "domaine": "Prévention habitation",
        "domain_key": "habitation",
        "titre_opt": "Protégez votre logement : les bons réflexes avant l'été",
        "contenu_opt": (
            "Votre contrat MRH Premium inclut une visite de prévention gratuite. Cette visite est "
            "à votre initiative — le rapport produit vous appartient et vous seul décidez de le partager "
            "avec votre gestionnaire sinistre.\n\n"
            "Avant la visite, vous pouvez nous transmettre : plan du logement et factures d'entretien "
            "chaudière. Évitez de photographier des personnes — les photos transmises sont conservées "
            "pendant la durée de votre contrat + 2 ans, puis supprimées.\n\n"
            "Bon réflexe sécurité : serrures certifiées A2P — recommandées pour votre protection. "
            "Elles renforcent votre niveau de sécurité ; elles ne constituent pas une condition "
            "contractuelle d'indemnisation.\n\n"
            "Bon réflexe 3 — Détecteur de fumée DAAF : obligatoire depuis 2015 (loi du 9 mars 2010). "
            "Son installation protège votre famille — elle n'est pas une clause d'exclusion de votre contrat.\n\n"
            "Bon réflexe 6 — Inventaire mobilier : photographiez vos biens de valeur et conservez-les "
            "dans votre espace client sécurisé. En cas de sinistre, vous pourrez les partager avec "
            "votre gestionnaire — le partage est toujours à votre initiative."
        ),
        "corrections_appliquees": "CG-07 (réduction implicite → conseil positif), CG-08 (DAAF non exclusion), CG-09 (accès vault → initiative assuré)",
        "rgpd_corrections": "SD-05 (photos tiers mentionnés), SD-06 (accès vault conditionné consentement), durée conservation précisée",
        "score_claimguard_avant": 32,
        "score_claimguard_apres": 89,
        "score_rgpd_avant": 22,
        "score_rgpd_apres": 75,
    },
]

# ────────────────────────────────────────────────────────────────────
# RGPD Compliance — tableau traitements
# ────────────────────────────────────────────────────────────────────
RGPD_TRAITEMENTS = [
    {
        "traitement_id": "T-01",
        "prev_id": "PREV-01",
        "domaine": "Route",
        "nature_donnee": "Données télématiques (GPS, accéléromètre, gyroscope)",
        "categorie": "Non-sensible (comportementale)",
        "finalite": "Scoring conducteur, modulation prime, prévention accidents",
        "base_legale": "Art. 6 §1b (exécution contrat) + art. 6 §1a (consentement granulaire recommandé)",
        "duree_conservation": "Durée contrat + 5 ans (prescription quinquennale)",
        "destinataires": "Service tarification, service sinistres, sous-traitant analyse télématique",
        "transfert_hors_ue": "Possible si serveurs hors UE — clauses contractuelles types requises (art. 46)",
        "decision_automatisee": "OUI — score conducteur calculé automatiquement (art. 22 applicable)",
        "droit_opposition": "Droit de résiliation formule télématique sans pénalité",
        "verdict_conformite": "WARN",
        "action_prioritaire": "Consentement granulaire séparé + information art. 22 + DPA sous-traitant",
    },
    {
        "traitement_id": "T-02",
        "prev_id": "PREV-02",
        "domaine": "Santé",
        "nature_donnee": "Données de santé (antécédents, traitements, IMC, tabac, alcool)",
        "categorie": "SENSIBLE — Art. 9 RGPD",
        "finalite": "Prévention santé, personnalisation remboursements (risque : sélection médicale)",
        "base_legale": "Art. 9 §2a (consentement EXPLICITE) — base actuelle insuffisante",
        "duree_conservation": "10 ans déclarés — à réduire à 2 ans pour finalité prévention",
        "destinataires": "Service prévention, service souscription — séparation étanche requise",
        "transfert_hors_ue": "Non précisé",
        "decision_automatisee": "NON déclaré — à vérifier si scoring santé automatique en back-office",
        "droit_opposition": "Non mentionné — droit d'effacement art. 17 applicable",
        "verdict_conformite": "KO",
        "action_prioritaire": "Consentement explicite séparé + réduction durée conservation + séparation prévention/tarification + DPIA obligatoire (art. 35)",
    },
    {
        "traitement_id": "T-03",
        "prev_id": "PREV-02",
        "domaine": "Santé",
        "nature_donnee": "Données addictions (tabac, alcool quantifié)",
        "categorie": "SENSIBLE — données révélant l'état de santé (EDPB guidelines 03/2020)",
        "finalite": "Prévention addictions — risque dérive tarification",
        "base_legale": "Art. 9 §2a (consentement EXPLICITE) — absent actuellement",
        "duree_conservation": "Mélangée avec données santé générales — à isoler",
        "destinataires": "Service prévention uniquement — cloisonnement strict requis",
        "transfert_hors_ue": "Non précisé",
        "decision_automatisee": "Risque si intégré dans scoring de renouvellement",
        "droit_opposition": "Droit rectification (art. 16) + effacement (art. 17)",
        "verdict_conformite": "KO",
        "action_prioritaire": "Isoler dans module prévention distinct. Engagement contractuel de non-utilisation tarifaire. DPIA spécifique addictions.",
    },
    {
        "traitement_id": "T-04",
        "prev_id": "PREV-03",
        "domaine": "Habitation",
        "nature_donnee": "Photos domicile + plan logement + système alarme",
        "categorie": "Non-sensible — mais risque si personnes visibles sur photos",
        "finalite": "Évaluation risque prévention, constitution dossier sinistre",
        "base_legale": "Art. 6 §1b (contrat) — acceptable si photos strictement logement",
        "duree_conservation": "Durée contrat + 2 ans recommandés",
        "destinataires": "Expert visite, gestionnaire sinistre — liste limitative requise",
        "transfert_hors_ue": "Non précisé — à contrôler si stockage cloud externe",
        "decision_automatisee": "NON",
        "droit_opposition": "Droit d'accès (art. 15) + effacement à demande (art. 17)",
        "verdict_conformite": "WARN",
        "action_prioritaire": "Limiter destinataires par liste. Interdire photos avec personnes. Durée conservation précisée dans politique de confidentialité.",
    },
    {
        "traitement_id": "T-05",
        "prev_id": "PREV-03",
        "domaine": "Habitation",
        "nature_donnee": "Accès coffre-fort numérique assuré (NEURAL Vault)",
        "categorie": "Non-sensible — mais traitement de données personnelles par tiers",
        "finalite": "Facilitation déclaration sinistre",
        "base_legale": "Art. 6 §1a (consentement) — absent dans contenu source",
        "duree_conservation": "Non définie",
        "destinataires": "Équipe sinistre + NEURAL Vault (sous-traitant potentiel)",
        "transfert_hors_ue": "Dépend architecture Vault — à auditer",
        "decision_automatisee": "NON",
        "droit_opposition": "Droit retrait accès sans impact garanties — à formaliser",
        "verdict_conformite": "KO",
        "action_prioritaire": "Consentement explicite accès Vault. DPA si NEURAL Vault = sous-traitant distinct (art. 28). Journal d'accès (art. 30 registre traitements).",
    },
]


# ────────────────────────────────────────────────────────────────────
# Onglet 0 — README
# ────────────────────────────────────────────────────────────────────
def build_readme(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — Audit & génération contenu prévention assurance conforme RGPD")
    write_subtitle(ws, 2, "3 domaines (Route, Santé, Habitation) — ClaimGuard + SensitiveDataGuard + RGPD art. 9")

    ws.row_dimensions[3].height = 10

    write_section_header(ws, 4, "MISSION DE L'AGENT")
    mission_rows = [
        ("Problème", "Les contenus de prévention assurance mélangent souvent conseil de sécurité, menaces d'exclusion déguisées, et collecte de données sensibles sans cadre RGPD adéquat. Résultat : risque de pratique commerciale trompeuse (DGCCRF) + violations RGPD (CNIL) + contentieux assuré."),
        ("Solution NEURAL", "L'agent PreventionContent analyse chaque contenu sur 2 axes : ClaimGuard (détection clauses décourageant la déclaration de sinistre) et SensitiveDataGuard (audit RGPD données sensibles). Il génère une version optimisée conforme."),
        ("Domaines couverts", "1. Prévention routière (télématique YoungDriver) | 2. Prévention santé (questionnaire annuel senior) | 3. Prévention habitation (visite diagnostic MRH)"),
        ("Modules", "ClaimGuard — 9 findings (3 KO, 5 WARN, 1 OK) | SensitiveDataGuard — 6 traitements audités | RGPD Compliance — 5 traitements avec base légale + durée + droits"),
    ]
    for i, (k, v) in enumerate(mission_rows):
        row = 5 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font()
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 50

    ws.row_dimensions[9].height = 10

    write_section_header(ws, 10, "RÉFÉRENTIEL RÉGLEMENTAIRE")
    refs = [
        ("ClaimGuard", "Code des assurances L.112-3, L.113-8, L.113-9, L.124-3", "Prévention des clauses décourageant la déclaration de sinistre ou suggérant des exclusions non contractuelles"),
        ("SensitiveDataGuard", "RGPD art. 9 + EDPB Guidelines 03/2020 + CNIL délibération 2022-118", "Données de santé, comportementales (addictions, santé inférée), biométriques — consentement EXPLICITE requis"),
        ("Décision automatisée", "RGPD art. 22 + considérant 71", "Tout scoring automatisé influençant prime ou garantie = obligation information + droit opposition"),
        ("Données télématiques", "RGPD art. 6 + EDPB Opinion 05/2019 (véhicules connectés)", "Géolocalisation continue = données personnelles sensibles selon EDPB — consentement granulaire requis"),
        ("DPIA obligatoire", "RGPD art. 35 + liste CNIL des traitements à risque élevé", "Données santé + scoring + décision automatisée = DPIA obligatoire avant mise en œuvre"),
        ("Pratiques commerciales", "Code de la consommation L.121-1 (pratique trompeuse)", "Présenter une exclusion non contractuelle comme certaine = pratique commerciale trompeuse"),
    ]
    reg_headers = ["Module", "Base légale", "Description"]
    style_header_row(ws, 11, 3)
    for col, h in enumerate(reg_headers, 1):
        ws.cell(row=11, column=col, value=h)
    for i, (mod, base, desc) in enumerate(refs):
        row = 12 + i
        for col, val in enumerate([mod, base, desc], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = alt_fill()
        ws.row_dimensions[row].height = 45

    ws.row_dimensions[18].height = 10

    write_section_header(ws, 19, "MODE D'EMPLOI DES ONGLETS")
    tabs = [
        ("0_README", "Ce guide — mission, réglementation, mode d'emploi"),
        ("1_Contenus_Source", "3 contenus prévention bruts, non optimisés — avec indicateurs de risque initial"),
        ("2_ClaimGuard_Audit", "9 findings claim avoidance : verdict KO/WARN/OK + correction par l'agent"),
        ("3_SensitiveDataGuard", "6 traitements données sensibles audités : base légale, durée, droits, verdict RGPD"),
        ("4_Contenus_Optimises", "Versions corrigées après audit ClaimGuard + SensitiveDataGuard, avec scores avant/après"),
        ("5_RGPD_Compliance", "Tableau registre traitements (art. 30) : nature, finalité, base légale, destinataires, durée"),
        ("6_Score_Global", "Synthèse : score ClaimGuard + score RGPD + décision par contenu"),
    ]
    for i, (tab, desc) in enumerate(tabs):
        row = 20 + i
        ws.cell(row=row, column=1, value=tab).font = body_font(bold=True, color=NEURAL_VIOLET)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row=row, column=2, value=desc).font = body_font()
        ws.row_dimensions[row].height = 22

    set_widths(ws, {1: 22, 2: 40, 3: 40, 4: 20, 5: 20, 6: 20, 7: 20, 8: 20})


# ────────────────────────────────────────────────────────────────────
# Onglet 1 — Contenus Source
# ────────────────────────────────────────────────────────────────────
def build_contenus_source(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — 1_Contenus_Source : Données d'entrée brutes (avant audit)")
    write_subtitle(ws, 2, "3 contenus prévention tels que produits sans audit ClaimGuard ni SensitiveDataGuard — risques initiaux élevés")

    headers = ["ID", "Domaine", "Format", "Cible", "Titre", "Contenu brut (input)", "Source données", "Canal diffusion"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, c in enumerate(CONTENUS_SOURCE):
        row = 5 + i
        vals = [c["id"], c["domaine"], c["format"], c["cible"],
                c["titre"], c["contenu"], c["source_donnees"], c["destinataire"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=110)
        dom_cell = ws.cell(row=row, column=1)
        dom_cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        dom_cell.fill = domain_fill(c["domain_key"])
        dom_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_widths(ws, {1: 10, 2: 18, 3: 22, 4: 28, 5: 30, 6: 52, 7: 30, 8: 22})


# ────────────────────────────────────────────────────────────────────
# Onglet 2 — ClaimGuard Audit
# ────────────────────────────────────────────────────────────────────
def build_claimguard(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — 2_ClaimGuard_Audit : Détection clauses claim avoidance")
    write_subtitle(ws, 2, "9 findings sur 3 contenus — KO = clause potentiellement illégale, WARN = à reformuler, OK = conforme")

    headers = ["Finding ID", "Contenu", "Domaine", "Pattern détecté", "Extrait problématique", "Analyse de risque", "Verdict", "Correction suggérée", "Référence légale"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, f in enumerate(CLAIMGUARD_FINDINGS):
        row = 5 + i
        vals = [f["finding_id"], f["prev_id"], f["domaine"],
                f["pattern"], f["extrait"], f["probleme"],
                f["verdict"], f["correction"], f["ref_legale"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=90)
        verdict_style(ws.cell(row=row, column=7), f["verdict"])

    ws.row_dimensions[14].height = 10

    # Summary stats
    write_section_header(ws, 15, "SYNTHÈSE CLAIMGUARD — TOTAUX PAR VERDICT")
    ko_count = sum(1 for f in CLAIMGUARD_FINDINGS if f["verdict"] == "KO")
    warn_count = sum(1 for f in CLAIMGUARD_FINDINGS if f["verdict"] == "WARN")
    ok_count = sum(1 for f in CLAIMGUARD_FINDINGS if f["verdict"] == "OK")
    total = len(CLAIMGUARD_FINDINGS)

    sum_headers = ["Total findings", "KO (clause illégale)", "WARN (à reformuler)", "OK (conforme)", "Taux conformité"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws.cell(row=16, column=col, value=h)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[16].height = 28

    sum_vals = [total, ko_count, warn_count, ok_count,
                f"=((D17+0.5*C17)/B17)"]
    for col, val in enumerate(sum_vals, 1):
        cell = ws.cell(row=17, column=col, value=val)
        cell.font = body_font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.cell(row=17, column=2).fill = red_fill()
    ws.cell(row=17, column=3).fill = amber_fill()
    ws.cell(row=17, column=4).fill = green_fill()
    ws.cell(row=17, column=5).font = body_font(bold=True, color="0000FF")
    ws.cell(row=17, column=5).number_format = "0%"
    ws.row_dimensions[17].height = 24

    set_widths(ws, {1: 12, 2: 10, 3: 14, 4: 22, 5: 38, 6: 42, 7: 10, 8: 42, 9: 32})


# ────────────────────────────────────────────────────────────────────
# Onglet 3 — SensitiveDataGuard
# ────────────────────────────────────────────────────────────────────
def build_sensitivedataguard(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — 3_SensitiveDataGuard : Audit RGPD données sensibles")
    write_subtitle(ws, 2, "6 traitements de données identifiés dans les 3 contenus — focus art. 9 (données sensibles) + art. 22 (décision automatisée)")

    headers = ["Finding ID", "Contenu", "Domaine", "Catégorie données", "Description / contenu collecté", "Article applicable",
               "Base légale actuelle", "Problème identifié", "Verdict", "Action requise",
               "Durée conservation", "Verdict durée", "Droits exercés", "Verdict droits"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, f in enumerate(SENSITIVE_DATA_FINDINGS):
        row = 5 + i
        vals = [
            f["finding_id"], f["prev_id"], f["domaine"],
            f["categorie_donnee"], f["description"], f["article_rgpd"],
            f["base_legale_actuelle"], f["probleme"], f["verdict"],
            f["action_requise"], f["duree_conservation"], f["verdict_duree"],
            f["droit_acces"], f["verdict_droits"],
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=90)
        verdict_style(ws.cell(row=row, column=9), f["verdict"])
        verdict_style(ws.cell(row=row, column=12), f["verdict_duree"])
        verdict_style(ws.cell(row=row, column=14), f["verdict_droits"])

    # Art. 9 bloc
    ws.row_dimensions[11].height = 10
    write_section_header(ws, 12, "RGPD ART. 9 — DONNÉES SENSIBLES : RAPPEL DES OBLIGATIONS", span=14)
    art9_rows = [
        ("Définition", "Données révélant l'origine raciale/ethnique, opinions politiques, convictions religieuses, appartenance syndicale, données génétiques, biométriques, de santé, vie/orientation sexuelle.", "RGPD art. 9 §1"),
        ("Extension EDPB", "Données comportementales révélant l'état de santé (alcool, tabac, IMC, conduite sous influence inférée) — qualifiées comme données de santé par l'EDPB (Guidelines 03/2020).", "EDPB Guidelines 03/2020"),
        ("Base légale requise", "Art. 9 §2a — Consentement EXPLICITE, SPÉCIFIQUE, RÉVOCABLE. Pas de consentement implicite ou groupé avec acceptation CGV/contrat.", "Art. 9 §2 RGPD"),
        ("DPIA obligatoire", "Tout traitement à grande échelle de données de santé requiert une DPIA avant mise en œuvre (art. 35 + liste CNIL 2018).", "Art. 35 RGPD + CNIL délibération 2018-327"),
        ("Durée conservation", "Principe de minimisation (art. 5 §1e) : données santé prévention = 2 ans recommandés. Données tarification = durée contrat + 5 ans (prescription).", "Art. 5 §1e RGPD + CNIL recommandations assurance"),
    ]
    art9_headers = ["Point", "Règle", "Référence"]
    for col, h in enumerate(art9_headers, 1):
        cell = ws.cell(row=13, column=col, value=h)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[13].height = 28
    for i, (pt, rule, ref) in enumerate(art9_rows):
        row = 14 + i
        for col, val in enumerate([pt, rule, ref], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = alt_fill()
        ws.row_dimensions[row].height = 45

    set_widths(ws, {1: 12, 2: 10, 3: 12, 4: 24, 5: 22, 6: 26,
                    7: 26, 8: 38, 9: 10, 10: 38, 11: 22, 12: 12, 13: 20, 14: 12})


# ────────────────────────────────────────────────────────────────────
# Onglet 4 — Contenus Optimisés
# ────────────────────────────────────────────────────────────────────
def build_contenus_optimises(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — 4_Contenus_Optimisés : Versions post-audit ClaimGuard + RGPD")
    write_subtitle(ws, 2, "Contenus reformulés — clauses claim avoidance supprimées, mentions RGPD intégrées, scores avant/après")

    headers = ["ID", "Domaine", "Titre optimisé", "Contenu optimisé (output agent)", "Corrections ClaimGuard", "Corrections RGPD"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, c in enumerate(CONTENUS_OPTIMISES):
        row = 5 + i
        vals = [c["prev_id"], c["domaine"], c["titre_opt"],
                c["contenu_opt"], c["corrections_appliquees"], c["rgpd_corrections"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=130)
        dom_cell = ws.cell(row=row, column=1)
        dom_cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        dom_cell.fill = domain_fill(c["domain_key"])
        dom_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[8].height = 10

    # Score avant/après
    write_section_header(ws, 9, "SCORES AVANT / APRÈS OPTIMISATION — Calculés par l'agent")
    score_headers = ["ID", "Domaine", "Score ClaimGuard Avant", "Score ClaimGuard Après", "Δ ClaimGuard",
                     "Score RGPD Avant", "Score RGPD Après", "Δ RGPD"]
    style_header_row(ws, 10, len(score_headers))
    for col, h in enumerate(score_headers, 1):
        ws.cell(row=10, column=col, value=h)

    for i, c in enumerate(CONTENUS_OPTIMISES):
        row = 11 + i
        cg_av = c["score_claimguard_avant"]
        cg_ap = c["score_claimguard_apres"]
        rgpd_av = c["score_rgpd_avant"]
        rgpd_ap = c["score_rgpd_apres"]
        vals = [c["prev_id"], c["domaine"], cg_av, cg_ap,
                f"=D{row}-C{row}", rgpd_av, rgpd_ap, f"=G{row}-F{row}"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 28

        # Color ClaimGuard scores
        for col, score in [(3, cg_av), (4, cg_ap)]:
            cell = ws.cell(row=row, column=col)
            if score >= 80:
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)
            elif score >= 50:
                cell.fill = amber_fill()
                cell.font = body_font(bold=True, color=AMBER_WARN)
            else:
                cell.fill = red_fill()
                cell.font = body_font(bold=True, color=RED_NO)

        # Color RGPD scores
        for col, score in [(6, rgpd_av), (7, rgpd_ap)]:
            cell = ws.cell(row=row, column=col)
            if score >= 70:
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)
            elif score >= 50:
                cell.fill = amber_fill()
                cell.font = body_font(bold=True, color=AMBER_WARN)
            else:
                cell.fill = red_fill()
                cell.font = body_font(bold=True, color=RED_NO)

        # Delta cells blue (formulas)
        for col in [5, 8]:
            ws.cell(row=row, column=col).font = body_font(bold=True, color="0000FF")

    set_widths(ws, {1: 10, 2: 18, 3: 30, 4: 52, 5: 36, 6: 36})


# ────────────────────────────────────────────────────────────────────
# Onglet 5 — RGPD Compliance (registre traitements art. 30)
# ────────────────────────────────────────────────────────────────────
def build_rgpd_compliance(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — 5_RGPD_Compliance : Registre traitements (art. 30)")
    write_subtitle(ws, 2, "5 traitements — nature, finalité, base légale, destinataires, durée, droits exercés, verdict conformité")

    headers = ["T-ID", "Contenu", "Domaine", "Nature données", "Catégorie", "Finalité",
               "Base légale", "Durée conservation", "Destinataires", "Transfert hors UE",
               "Décision auto. (art. 22)", "Droit opposition", "Verdict conformité", "Action prioritaire"]
    style_header_row(ws, 4, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    for i, t in enumerate(RGPD_TRAITEMENTS):
        row = 5 + i
        vals = [
            t["traitement_id"], t["prev_id"], t["domaine"],
            t["nature_donnee"], t["categorie"], t["finalite"],
            t["base_legale"], t["duree_conservation"], t["destinataires"],
            t["transfert_hors_ue"], t["decision_automatisee"], t["droit_opposition"],
            t["verdict_conformite"], t["action_prioritaire"],
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        style_body_row(ws, row, len(headers), alt=(i % 2 == 1), height=80)
        verdict_style(ws.cell(row=row, column=13), t["verdict_conformite"])

        # Highlight SENSIBLE category
        cat_cell = ws.cell(row=row, column=5)
        if "SENSIBLE" in str(cat_cell.value or ""):
            cat_cell.fill = red_fill()
            cat_cell.font = body_font(bold=True, color=RED_NO)

    ws.row_dimensions[10].height = 10

    write_section_header(ws, 11, "OBLIGATIONS DPIA (ART. 35 RGPD) — TRAITEMENTS À RISQUE ÉLEVÉ", span=14)
    dpia_info = [
        ("DPIA obligatoire si", "Traitement à grande échelle de données de santé (art. 9) + scoring automatisé + croisement de plusieurs sources de données personnelles"),
        ("Contenus concernés", "PREV-01 (télématique + inférence santé) + PREV-02 (questionnaire santé + scoring) = DPIA obligatoire avant mise en production"),
        ("Contenu PREV-03", "PREV-03 (photos domicile + Vault) : DPIA recommandée si déploiement > 10 000 assurés — à vérifier selon périmètre"),
        ("Délai CNIL", "Notification CNIL si violation données personnelles sous 72h (art. 33). Notification assuré si risque élevé pour ses droits (art. 34)."),
        ("Registre traitements", "Ce tableau constitue un extrait du registre des activités de traitement requis par art. 30 RGPD. À compléter avec DPO avant mise en production."),
    ]
    for i, (k, v) in enumerate(dpia_info):
        row = 12 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font(color="475569")
        cell.alignment = Alignment(wrap_text=True)
        if i % 2 == 1:
            for col in range(1, 15):
                ws.cell(row=row, column=col).fill = alt_fill()
        ws.row_dimensions[row].height = 38

    set_widths(ws, {1: 8, 2: 10, 3: 12, 4: 24, 5: 22, 6: 28,
                    7: 30, 8: 22, 9: 26, 10: 18, 11: 18, 12: 26, 13: 12, 14: 36})


# ────────────────────────────────────────────────────────────────────
# Onglet 6 — Score Global
# ────────────────────────────────────────────────────────────────────
def build_score_global(ws) -> None:
    write_title(ws, "NEURAL · PreventionContent — 6_Score_Global : Synthèse conformité")
    write_subtitle(ws, 2, "Score ClaimGuard + Score RGPD + Décision par contenu — avant et après optimisation par l'agent")

    write_section_header(ws, 4, "MATRICE SCORES PAR CONTENU — 3 DOMAINES")
    headers = ["ID", "Domaine", "KO ClaimGuard", "WARN ClaimGuard", "Score CG Avant", "Score CG Après",
               "Traitements RGPD KO", "Score RGPD Avant", "Score RGPD Après", "Décision finale"]
    style_header_row(ws, 5, len(headers))
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)

    domain_stats = {
        "PREV-01": {"ko_cg": 1, "warn_cg": 2, "ko_rgpd": 2},
        "PREV-02": {"ko_cg": 2, "warn_cg": 1, "ko_rgpd": 2},
        "PREV-03": {"ko_cg": 2, "warn_cg": 1, "ko_rgpd": 2},
    }

    for i, c in enumerate(CONTENUS_OPTIMISES):
        row = 6 + i
        stats = domain_stats[c["prev_id"]]
        cg_av = c["score_claimguard_avant"]
        cg_ap = c["score_claimguard_apres"]
        rgpd_av = c["score_rgpd_avant"]
        rgpd_ap = c["score_rgpd_apres"]
        decision = "OPTIMISÉ — prêt après validation DPO" if cg_ap >= 85 and rgpd_ap >= 70 else "RÉVISION COMPLÉMENTAIRE"
        vals = [c["prev_id"], c["domaine"],
                stats["ko_cg"], stats["warn_cg"],
                cg_av, cg_ap,
                stats["ko_rgpd"], rgpd_av, rgpd_ap,
                decision]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 32

        # Domain badge
        dom_cell = ws.cell(row=row, column=1)
        dom_cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        dom_cell.fill = domain_fill(c["domain_key"])

        # KO/WARN cells
        ko_cell = ws.cell(row=row, column=3)
        ko_cell.fill = red_fill()
        ko_cell.font = body_font(bold=True, color=RED_NO)
        warn_cell = ws.cell(row=row, column=4)
        warn_cell.fill = amber_fill()
        warn_cell.font = body_font(bold=True, color=AMBER_WARN)

        # CG scores
        for col, score in [(5, cg_av), (6, cg_ap)]:
            cell = ws.cell(row=row, column=col)
            if score >= 80:
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)
            elif score >= 50:
                cell.fill = amber_fill()
                cell.font = body_font(bold=True, color=AMBER_WARN)
            else:
                cell.fill = red_fill()
                cell.font = body_font(bold=True, color=RED_NO)

        # RGPD scores
        for col, score in [(8, rgpd_av), (9, rgpd_ap)]:
            cell = ws.cell(row=row, column=col)
            if score >= 70:
                cell.fill = green_fill()
                cell.font = body_font(bold=True, color=GREEN_OK)
            elif score >= 50:
                cell.fill = amber_fill()
                cell.font = body_font(bold=True, color=AMBER_WARN)
            else:
                cell.fill = red_fill()
                cell.font = body_font(bold=True, color=RED_NO)

        # RGPD KO
        ws.cell(row=row, column=7).fill = red_fill()
        ws.cell(row=row, column=7).font = body_font(bold=True, color=RED_NO)

        # Decision
        dec_cell = ws.cell(row=row, column=10)
        if "prêt" in decision:
            dec_cell.fill = green_fill()
            dec_cell.font = body_font(bold=True, color=GREEN_OK)
        else:
            dec_cell.fill = amber_fill()
            dec_cell.font = body_font(bold=True, color=AMBER_WARN)

    ws.row_dimensions[9].height = 10

    # Aggregated formulas
    write_section_header(ws, 10, "STATISTIQUES GLOBALES — FORMULES")
    agg = [
        ("Nb KO ClaimGuard total", "=SUM(C6:C8)"),
        ("Nb WARN ClaimGuard total", "=SUM(D6:D8)"),
        ("Score CG moyen avant", "=AVERAGE(E6:E8)"),
        ("Score CG moyen après", "=AVERAGE(F6:F8)"),
        ("Gain CG moyen", "=B14-B13"),
        ("Nb traitements RGPD KO", "=SUM(G6:G8)"),
        ("Score RGPD moyen avant", "=AVERAGE(H6:H8)"),
        ("Score RGPD moyen après", "=AVERAGE(I6:I8)"),
        ("Gain RGPD moyen", "=B18-B17"),
    ]
    for i, (label, formula) in enumerate(agg):
        row = 11 + i
        ws.cell(row=row, column=1, value=label).font = body_font(bold=True)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = body_font(bold=True, color="0000FF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    ws.row_dimensions[20].height = 10

    # Verdict final & recommandation
    write_section_header(ws, 21, "VERDICT FINAL & RECOMMANDATIONS")
    recs = [
        ("Validation DPO obligatoire", "Avant toute diffusion des contenus optimisés, validation par le DPO ou le délégué à la protection des données (art. 37-39 RGPD). Les contenus santé (PREV-02) requièrent une DPIA complète."),
        ("DPIA PREV-01 + PREV-02", "Lancer les DPIA pour les traitements télématiques et données santé avant mise en production. Délai estimé : 4-6 semaines avec DPO + équipe technique."),
        ("Séparation stricte prévention/tarification", "Garantir contractuellement et techniquement que les données collectées via les contenus de prévention ne sont jamais utilisées pour modifier les garanties ou les primes."),
        ("Révision annuelle", "Les contenus de prévention doivent être re-audités annuellement par l'agent PreventionContent, notamment après toute évolution réglementaire (EDPB, CNIL, jurisprudence ACPR)."),
        ("Limite de l'agent", "L'agent PreventionContent détecte les risques et propose des reformulations. Il ne remplace pas un DPO, un juriste assurance ou un audit CNIL. Sa décision est indicative."),
    ]
    for i, (k, v) in enumerate(recs):
        row = 22 + i
        ws.cell(row=row, column=1, value=k).font = body_font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=2, value=v)
        cell.font = body_font(color="475569")
        cell.alignment = Alignment(wrap_text=True)
        if i % 2 == 1:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = alt_fill()
        ws.row_dimensions[row].height = 40

    set_widths(ws, {1: 18, 2: 20, 3: 16, 4: 16, 5: 16, 6: 16, 7: 18, 8: 16, 9: 16, 10: 28})


# ────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────
def main() -> None:
    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = "0_README"
    build_readme(ws_readme)

    ws_source = wb.create_sheet("1_Contenus_Source")
    build_contenus_source(ws_source)

    ws_cg = wb.create_sheet("2_ClaimGuard_Audit")
    build_claimguard(ws_cg)

    ws_sd = wb.create_sheet("3_SensitiveDataGuard")
    build_sensitivedataguard(ws_sd)

    ws_opt = wb.create_sheet("4_Contenus_Optimises")
    build_contenus_optimises(ws_opt)

    ws_rgpd = wb.create_sheet("5_RGPD_Compliance")
    build_rgpd_compliance(ws_rgpd)

    ws_score = wb.create_sheet("6_Score_Global")
    build_score_global(ws_score)

    out = "PreventionContent_NEURAL.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
