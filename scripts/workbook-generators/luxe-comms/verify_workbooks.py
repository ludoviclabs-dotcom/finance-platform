"""Verification post-correctifs : dates, formules, tables, DVs, CF."""
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

OUT_DIR = Path(r"C:\Users\Ludo\Desktop\IA projet entreprises\NEURAL - LUXE - Communication")

FILES = [
    "NEURAL_LUXE_COMMS_FOUNDATIONS.xlsx",
    "NEURAL_LUXE_COMMS_MASTER.xlsx",
    "NEURAL_AG001_MaisonVoiceGuard.xlsx",
    "NEURAL_AG002_LuxePressAgent.xlsx",
    "NEURAL_AG003_LuxeEventComms.xlsx",
    "NEURAL_AG004_HeritageComms.xlsx",
    "NEURAL_AG005_GreenClaimChecker.xlsx",
]


def main():
    ok = True

    print("=" * 74)
    print("  AUDIT POST-CORRECTIFS")
    print("=" * 74)

    for fname in FILES:
        p = OUT_DIR / fname
        if not p.exists():
            print(f"  [MISSING] {fname}")
            ok = False
            continue
        try:
            wb = load_workbook(p, data_only=False)
            sheets = wb.sheetnames
            tables = sum(len(wb[s].tables) for s in sheets)
            dvs = sum(len(wb[s].data_validations.dataValidation) for s in sheets)
            cfs = sum(len(wb[s].conditional_formatting._cf_rules) for s in sheets)
            size_kb = p.stat().st_size / 1024
            print(f"\n  {fname}  ({size_kb:.1f} KB)")
            print(f"       {len(sheets)} onglets | {tables} tables | {dvs} DV | {cfs} CF ranges")
            wb.close()
        except Exception as e:
            print(f"  [ERROR] {fname} -> {e}")
            ok = False

    print("\n" + "=" * 74)
    print("  SMOKE TESTS CIBLES")
    print("=" * 74)

    # Test 1 : FOUNDATIONS/5_HERITAGE — SRC-005 doit avoir REVIEW_DATE en datetime + formule STATUT
    wb = load_workbook(OUT_DIR / "NEURAL_LUXE_COMMS_FOUNDATIONS.xlsx")
    ws = wb["5_HERITAGE_SOURCEBOOK"]
    src005_date = ws.cell(row=9, column=8).value
    src005_fmt = ws.cell(row=9, column=8).number_format
    statut_formula = ws.cell(row=9, column=11).value
    print(f"\n  Test 1 — SRC-005 REVIEW_DATE 2025-01-10")
    print(f"    value: {src005_date}  (type={type(src005_date).__name__})")
    print(f"    format: {src005_fmt}")
    print(f"    STATUT formula: {statut_formula}")
    t1 = isinstance(src005_date, (date, datetime)) and "IF" in str(statut_formula)
    print(f"    -> {'OK' if t1 else 'FAIL'}")
    wb.close()

    # Test 2 : MASTER 9_GLOBAL_KPIS — UNIQUE present
    wb = load_workbook(OUT_DIR / "NEURAL_LUXE_COMMS_MASTER.xlsx")
    ws = wb["9_GLOBAL_KPIS"]
    runs_total = ws.cell(row=5, column=3).value
    print(f"\n  Test 2 — MASTER Runs total formula")
    print(f"    formula: {runs_total}")
    t2 = "UNIQUE" in str(runs_total)
    print(f"    -> {'OK' if t2 else 'FAIL'}")
    wb.close()

    # Test 3 : AG-001 KPI hardfail count — COUNTIFS (pas SUMIFS)
    wb = load_workbook(OUT_DIR / "NEURAL_AG001_MaisonVoiceGuard.xlsx")
    ws = wb["8_KPI_DASHBOARD"]
    # Chercher la ligne hardfail
    hf_formula = None
    for r in range(5, 15):
        lbl = ws.cell(row=r, column=2).value
        if lbl and "hard fail" in lbl.lower():
            hf_formula = ws.cell(row=r, column=3).value
            break
    print(f"\n  Test 3 — AG-001 Docs avec hardfail")
    print(f"    formula: {hf_formula}")
    t3 = hf_formula and "COUNTIFS" in str(hf_formula) and "SUMIFS" not in str(hf_formula)
    print(f"    -> {'OK' if t3 else 'FAIL'}")
    wb.close()

    # Test 4 : AG-001 scoring — XLOOKUP sur tblParamsAG001
    wb = load_workbook(OUT_DIR / "NEURAL_AG001_MaisonVoiceGuard.xlsx")
    ws = wb["5_SCORING_ENGINE"]
    thresh = ws.cell(row=5, column=14).value
    sla_target = ws.cell(row=5, column=18).value
    sla_met = ws.cell(row=5, column=19).value
    print(f"\n  Test 4 — AG-001 THRESHOLD + SLA branche CRISIS")
    print(f"    THRESHOLD formula: {str(thresh)[:80]}...")
    print(f"    SLA_TARGET_H formula: {str(sla_target)[:80]}...")
    print(f"    SLA_MET formula: {str(sla_met)[:80]}...")
    t4 = ("tblParamsAG001" in str(thresh) and "CRISIS_MODE_ON" in str(sla_target))
    print(f"    -> {'OK' if t4 else 'FAIL'}")
    wb.close()

    # Test 5 : tblBrandRules collision resolue
    wb1 = load_workbook(OUT_DIR / "NEURAL_LUXE_COMMS_FOUNDATIONS.xlsx")
    wb2 = load_workbook(OUT_DIR / "NEURAL_AG001_MaisonVoiceGuard.xlsx")
    f_tables = [t for s in wb1.sheetnames for t in wb1[s].tables]
    a_tables = [t for s in wb2.sheetnames for t in wb2[s].tables]
    print(f"\n  Test 5 — Collision tblBrandRules resolue")
    print(f"    FOUNDATIONS : {'tblBrandRulesFound' in f_tables}")
    print(f"    AG-001      : {'tblBrandRulesAG001' in a_tables}")
    t5 = "tblBrandRulesFound" in f_tables and "tblBrandRulesAG001" in a_tables
    print(f"    -> {'OK' if t5 else 'FAIL'}")
    wb1.close(); wb2.close()

    # Test 6 : AG-002 PRESS_PICKUP existe
    wb = load_workbook(OUT_DIR / "NEURAL_AG002_LuxePressAgent.xlsx")
    t6 = "11_PRESS_PICKUP" in wb.sheetnames
    print(f"\n  Test 6 — AG-002/11_PRESS_PICKUP cree (R-09)")
    print(f"    -> {'OK' if t6 else 'FAIL'}")
    wb.close()

    # Test 7 : LEARNINGS dans les 5 agents
    learnings_ok = {}
    for agent_file, sheet_name in [
        ("NEURAL_AG001_MaisonVoiceGuard.xlsx", "10_LEARNINGS"),
        ("NEURAL_AG002_LuxePressAgent.xlsx",   "12_LEARNINGS"),
        ("NEURAL_AG003_LuxeEventComms.xlsx",   "10_LEARNINGS"),
        ("NEURAL_AG004_HeritageComms.xlsx",    "9_LEARNINGS"),
        ("NEURAL_AG005_GreenClaimChecker.xlsx","11_LEARNINGS"),
    ]:
        wb = load_workbook(OUT_DIR / agent_file)
        learnings_ok[agent_file] = sheet_name in wb.sheetnames
        wb.close()
    t7 = all(learnings_ok.values())
    print(f"\n  Test 7 — LEARNINGS pose sur les 5 agents")
    for f, v in learnings_ok.items():
        print(f"    {f:50s} {'OK' if v else 'FAIL'}")
    print(f"    -> {'OK' if t7 else 'FAIL'}")

    # Test 8 : AG-001 TestSet executable
    wb = load_workbook(OUT_DIR / "NEURAL_AG001_MaisonVoiceGuard.xlsx")
    ws = wb["9_TESTSET"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, 11)]
    print(f"\n  Test 8 — AG-001 TestSet avec OBSERVED/PASS")
    print(f"    headers: {headers}")
    t8 = "PASS_FAIL" in headers and "OBSERVED_DECISION" in headers
    pass_formula = ws.cell(row=5, column=9).value
    print(f"    PASS_FAIL formula row 5: {pass_formula}")
    print(f"    -> {'OK' if t8 else 'FAIL'}")
    wb.close()

    # Test 9 : DOC_ID unifie (DOC-NNNN) — verifier AG-001 input queue
    wb = load_workbook(OUT_DIR / "NEURAL_AG001_MaisonVoiceGuard.xlsx")
    ws = wb["2_INPUT_QUEUE"]
    first_doc_id = ws.cell(row=5, column=1).value
    print(f"\n  Test 9 — DOC_ID unifie (DOC-NNNN)")
    print(f"    AG-001 input queue row 5 DOC_ID: {first_doc_id}")
    t9 = first_doc_id and first_doc_id.startswith("DOC-")
    print(f"    -> {'OK' if t9 else 'FAIL'}")
    wb.close()

    print("\n" + "=" * 74)
    results = [t1, t2, t3, t4, t5, t6, t7, t8, t9]
    score = sum(results)
    print(f"  RESULTAT : {score} / {len(results)} tests OK")
    print("=" * 74)
    return score == len(results)


if __name__ == "__main__":
    main()
